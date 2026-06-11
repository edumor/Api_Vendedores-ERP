"""
API Vendedores Microbell — FastAPI
Puerto: 8000
"""
import os
import math
import smtplib
import mimetypes
from io import BytesIO
from typing import Optional
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from fastapi import FastAPI, HTTPException, Query, File, UploadFile, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, StreamingResponse
from pydantic import BaseModel
import firebirdsql
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, HRFlowable
from reportlab.pdfbase import pdfmetrics

HOST     = '127.0.0.1'
PORT     = 3050
DATABASE      = 'c:/flexxus/DB/DB-Prueba.gdb'          # BD oficial (Prueba)
DATABASE_EST  = 'c:/flexxus/DB/DB-EST-Prueba.gdb'     # BD estadísticas (Prueba)
DATABASE_MLT  = 'c:/flexxus/DB/DB-MLT-Prueba.gdb'     # BD SW / Línea 2 (Prueba)
DB_USER  = 'SYSDBA'
DB_PASS  = '3122414422'

# ── SMTP Microbell ─────────────────────────────────────────────────────────────
SMTP_HOST     = 'mail.microbellsa.com.ar'
SMTP_PORT     = 465                          # 465 SSL (confirmado activo en mail.microbellsa.com.ar)
SMTP_USER     = 'noresponder2@microbellsa.com.ar'
SMTP_PASS     = 'Micro3880*'
SMTP_FROM     = 'noresponder2@microbellsa.com.ar'
SMTP_TO_PAGOS = 'pagos@microbellsa.com.ar'
# ──────────────────────────────────────────────────────────────────────────────

def conn(charset='WIN1252', db=None):
    return firebirdsql.connect(host=HOST, port=PORT, database=db or DATABASE,
                               user=DB_USER, password=DB_PASS, charset=charset)

# ── Debug global: captura errores y conteos de _query_db ─────────────────────
_QV_LAST_ERRORS: dict = {}
_QV_LAST_COUNTS: dict = {}

app = FastAPI(title="API Vendedores Microbell")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

FRONTEND_PATH = os.path.join(os.path.dirname(__file__), "frontend.html")
LOGO_PATH     = os.path.join(os.path.dirname(__file__), "microbellSA-color.png")
FAVICON_PATH  = os.path.join(os.path.dirname(__file__), "favicon.ico")

_BANCOS = [
    ("Santander Rio",  "0720131420000001149872", "131",    "11498/7"),
    ("Provincia",      "0140004501400404115211", "4004",   "041152/1"),
    ("HSBC",           "1500607500060732055732", "607",    "607-3-205573"),
    ("Galicia",        "0070154520000005006724", "154",    "5006-7-154/2"),
]
_MP_EMAIL = "marketing@microbellsa.com.ar"
_MP_CVU   = "0000003100004756934965"

def _fmt(v):
    """Formato moneda argentina: $ 1.234,56"""
    try:
        v = float(v)
    except (TypeError, ValueError):
        return "$ 0,00"
    neg = v < 0
    v = abs(v)
    ent = int(v)
    dec = round((v - ent) * 100)
    s = f"{ent:,}".replace(",", ".")
    return f"{'- ' if neg else ''}$ {s},{dec:02d}"

def _d(v):
    """Convierte datetime o string a dd/mm/yyyy"""
    if v is None:
        return ""
    s = str(v)
    try:
        from datetime import datetime
        if "T" in s:
            s = s.split("T")[0]
        elif " " in s:
            s = s.split(" ")[0]
        parts = s.split("-")
        if len(parts) == 3:
            return f"{parts[2]}/{parts[1]}/{parts[0]}"
    except Exception:
        pass
    return s

@app.get("/", response_class=HTMLResponse)
def index():
    with open(FRONTEND_PATH, encoding="utf-8") as f:
        return f.read()

@app.get("/logo")
def get_logo():
    from fastapi.responses import FileResponse as FR
    if os.path.exists(LOGO_PATH):
        return FR(LOGO_PATH, media_type="image/png")
    return HTMLResponse("", status_code=404)

@app.get("/favicon.ico")
def get_favicon():
    from fastapi.responses import FileResponse as FR
    if os.path.exists(FAVICON_PATH):
        return FR(FAVICON_PATH, media_type="image/x-icon",
                  headers={"Cache-Control": "public, max-age=86400"})
    return HTMLResponse("", status_code=404)

# ─── Auth ──────────────────────────────────────────────────────────────────────
PERFILES_PERMITIDOS = {'VENDEDORES', 'ADV', 'ADVJUAN', 'GERENTES', 'GTES FE'}

class LoginBody(BaseModel):
    usuario: str
    password: str

@app.post("/auth/login")
def login(body: LoginBody):
    c = conn()
    cur = c.cursor()
    cur.execute(
        'SELECT CODIGOUSUARIO, RAZONSOCIAL, PASSWORD1, ESVENDEDOR, CODIGOPERFIL '
        'FROM "USUARIOS" WHERE TRIM(CODIGOUSUARIO) = ? AND ACTIVO = ?',
        (body.usuario.strip().upper(), '1')
    )
    row = cur.fetchone()
    c.close()
    if not row:
        raise HTTPException(401, "Usuario o contraseña incorrectos")
    pwd_db = (row[2] or '').strip()
    if pwd_db != body.password.strip():
        raise HTTPException(401, "Usuario o contraseña incorrectos")
    perfil = (row[4] or '').strip()
    if perfil not in PERFILES_PERMITIDOS:
        raise HTTPException(403, "Sin acceso: perfil no autorizado")
    return {
        "codigousuario": (row[0] or '').strip(),
        "razonsocial":   (row[1] or '').strip(),
        "esvendedor":    str(row[3] or '').strip() == "1",
        "perfil":        perfil
    }

class CambiarPasswordBody(BaseModel):
    usuario: str
    password_actual: str
    nueva_password: str

@app.post("/auth/cambiar-password")
def cambiar_password(body: CambiarPasswordBody):
    c = conn()
    cur = c.cursor()
    cur.execute(
        'SELECT PASSWORD1 FROM "USUARIOS" WHERE TRIM(CODIGOUSUARIO) = ? AND ACTIVO = ?',
        (body.usuario.strip().upper(), '1')
    )
    row = cur.fetchone()
    if not row:
        c.close()
        raise HTTPException(404, "Usuario no encontrado")
    if (row[0] or '').strip() != body.password_actual.strip():
        c.close()
        raise HTTPException(401, "Contraseña actual incorrecta")
    cur.execute(
        'UPDATE "USUARIOS" SET PASSWORD1 = ? WHERE TRIM(CODIGOUSUARIO) = ?',
        (body.nueva_password.strip(), body.usuario.strip().upper())
    )
    c.commit()
    c.close()
    return {"ok": True}

# ─── Stock ─────────────────────────────────────────────────────────────────────
@app.get("/stock")
def get_stock(
    buscar: Optional[str] = None,
    gruposuperrubro: Optional[str] = None,
    superrubro: Optional[str] = None,
    rubro: Optional[str] = None,
    marca: Optional[str] = None,
    limit: int = Query(100, le=300),
    offset: int = 0
):
    c = conn()
    cur = c.cursor()
    wheres = ["a.ACTIVO = '1'"]
    params = []

    if buscar:
        wheres.append("(UPPER(a.DESCRIPCION) CONTAINING UPPER(?) OR a.CODIGOPARTICULAR CONTAINING ?)")
        params += [buscar, buscar]
    if rubro:
        wheres.append("a.CODIGORUBRO = ?")
        params.append(rubro)
    if superrubro:
        wheres.append("r.CODIGOSUPERRUBRO = ?")
        params.append(superrubro)
    if gruposuperrubro:
        wheres.append("sr.CODIGOGRUPOSUPERRUBRO = ?")
        params.append(gruposuperrubro)
    if marca:
        wheres.append("a.CODIGOMARCA = ?")
        params.append(marca)

    where_sql = " AND ".join(wheres)

    # Cotización USD → ARS actual
    try:
        cur.execute('SELECT CAMBIO FROM "MONEDAS" WHERE CODIGOMONEDA = ?', ('DOLARES',))
        row_m = cur.fetchone()
        cambio_usd = float(row_m[0]) if row_m else 1.0
    except Exception:
        cambio_usd = 1.0

    # Usar FMA_STOCK que ya calcula STOCKREAL y STOCKREMANENTE para depósitos 001+003
    sql = f"""
        SELECT FIRST {limit} SKIP {offset}
            s.ID_ARTICULO, a.CODIGOPARTICULAR, a.DESCRIPCION, a.CODIGOMARCA,
            a.PRECIOLISTA1, a.PRECIOLISTA2, a.PRECIOLISTA3,
            a.ALICUOTAIVA, a.CODIGOUNIDADMEDIDA,
            s.STOCKREAL, s.STOCKREMANENTE,
            r.DESCRIPCION, sr.DESCRIPCION, g.DESCRIPCION,
            a.CODIGOMONEDA
        FROM "FMA_STOCK"(NULL, NULL, '001,003', 1, 1) s
        JOIN "ARTICULOS" a ON a.CODIGOARTICULO = s.ID_ARTICULO
        LEFT JOIN "RUBROS" r ON r.CODIGORUBRO = a.CODIGORUBRO
        LEFT JOIN "SUPERRUBROS" sr ON sr.CODIGOSUPERRUBRO = r.CODIGOSUPERRUBRO
        LEFT JOIN "GRUPOSUPERRUBROS" g ON g.CODIGOGRUPOSUPERRUBRO = sr.CODIGOGRUPOSUPERRUBRO
        WHERE {where_sql}
          AND s.STOCKREMANENTE > 0
        ORDER BY a.DESCRIPCION
    """
    try:
        cur.execute(sql, params)
        rows = cur.fetchall()

        # Remanente por depósito individual para los artículos retornados
        ids = [r[0] for r in rows]
        rem_001_map = {}
        rem_003_map = {}
        if ids:
            cur.execute("SELECT ID_ARTICULO, STOCKREMANENTE FROM \"FMA_STOCK\"(NULL, NULL, '001', 1, 1)")
            rem_001_map = {r[0]: float(r[1] or 0) for r in cur.fetchall()}
            cur.execute("SELECT ID_ARTICULO, STOCKREMANENTE FROM \"FMA_STOCK\"(NULL, NULL, '003', 1, 1)")
            rem_003_map = {r[0]: float(r[1] or 0) for r in cur.fetchall()}

        c.close()
        resultado = []
        for r in rows:
            moneda = (r[14] or '').strip().upper()
            factor = cambio_usd if moneda == 'DOLARES' else 1.0
            def conv(v): return math.ceil(float(v) * factor * 100) / 100 if v else 0
            rem_001 = rem_001_map.get(r[0], 0)
            rem_003 = rem_003_map.get(r[0], 0)
            resultado.append({
                "codigo":           r[0],
                "codigoparticular": r[1] or r[0],
                "descripcion":      r[2],  "marca":   r[3],
                "precio1":          conv(r[4]),
                "precio2":          conv(r[5]),
                "precio3":          conv(r[6]),
                "iva":              r[7],  "unidad":  r[8],
                "stock":            float(r[9]) if r[9] else 0,
                "remanente":        float(r[10]) if r[10] else 0,
                "remanente_001":    rem_001,
                "remanente_003":    rem_003,
                "rubro":            r[11], "superrubro": r[12], "gruposuperrubro": r[13],
                "moneda":           r[14],
            })
        return resultado
    except Exception as e:
        c.close()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/buscar-articulos")
def buscar_articulos(q: str = Query("", min_length=2), db: str = Query("oficial")):
    # ARTICULOS siempre en BD principal, independiente de sw/oficial
    c = conn('WIN1252', db=DATABASE)
    cur = c.cursor()
    try:
        try:
            cur.execute('SELECT CAMBIO FROM "MONEDAS" WHERE CODIGOMONEDA = ?', ('DOLARES',))
            rm = cur.fetchone()
            cambio_usd = float(rm[0]) if rm else 1.0
        except Exception:
            cambio_usd = 1.0
        cur.execute(
            "SELECT FIRST 15 CODIGOARTICULO, CODIGOPARTICULAR, DESCRIPCION, PRECIOLISTA1, ALICUOTAIVA, CODIGOMONEDA "
            "FROM \"ARTICULOS\" WHERE UPPER(DESCRIPCION) CONTAINING UPPER(?) "
            "ORDER BY DESCRIPCION",
            (q,)
        )
        rows = cur.fetchall()
    finally:
        c.close()
    resultado = []
    for r in rows:
        moneda = (r[5] or '').strip().upper()
        factor = cambio_usd if moneda == 'DOLARES' else 1.0
        precio = math.ceil(float(r[3]) * factor * 100) / 100 if r[3] else 0
        resultado.append({"codigo": r[0], "codigoparticular": r[1] or r[0],
                          "descripcion": r[2], "precio": precio,
                          "iva": float(r[4]) if r[4] else 21})
    return resultado

# ─── Helpers stock export ──────────────────────────────────────────────────────
def _fetch_stock_data(buscar=None, gruposuperrubro=None, superrubro=None, rubro=None):
    """Devuelve todos los artículos con remanente > 0 según filtros, sin límite."""
    # ── 1. Cotización USD
    cambio_usd = 1.0
    try:
        cx = conn()
        cu = cx.cursor()
        cu.execute('SELECT CAMBIO FROM "MONEDAS" WHERE CODIGOMONEDA = ?', ('DOLARES',))
        rm = cu.fetchone()
        cambio_usd = float(rm[0]) if rm else 1.0
        cx.close()
    except Exception:
        pass

    # ── 2. Query principal (001+003 combinados)
    wheres = ["a.ACTIVO = '1'"]
    params = []
    if buscar:
        wheres.append("(UPPER(a.DESCRIPCION) CONTAINING UPPER(?) OR a.CODIGOPARTICULAR CONTAINING ?)")
        params += [buscar, buscar]
    if rubro:
        wheres.append("a.CODIGORUBRO = ?")
        params.append(rubro)
    if superrubro:
        wheres.append("r.CODIGOSUPERRUBRO = ?")
        params.append(superrubro)
    if gruposuperrubro:
        wheres.append("sr.CODIGOGRUPOSUPERRUBRO = ?")
        params.append(gruposuperrubro)
    where_sql = " AND ".join(wheres)

    sql = f"""
        SELECT s.ID_ARTICULO, a.CODIGOPARTICULAR, a.DESCRIPCION,
               a.PRECIOLISTA1, a.ALICUOTAIVA, a.CODIGOUNIDADMEDIDA,
               s.STOCKREAL, s.STOCKREMANENTE, a.CODIGOMONEDA,
               r.DESCRIPCION, sr.DESCRIPCION, g.DESCRIPCION
        FROM "FMA_STOCK"(NULL, NULL, '001,003', 1, 1) s
        JOIN "ARTICULOS" a ON a.CODIGOARTICULO = s.ID_ARTICULO
        LEFT JOIN "RUBROS" r ON r.CODIGORUBRO = a.CODIGORUBRO
        LEFT JOIN "SUPERRUBROS" sr ON sr.CODIGOSUPERRUBRO = r.CODIGOSUPERRUBRO
        LEFT JOIN "GRUPOSUPERRUBROS" g ON g.CODIGOGRUPOSUPERRUBRO = sr.CODIGOGRUPOSUPERRUBRO
        WHERE {where_sql} AND s.STOCKREMANENTE > 0
        ORDER BY a.DESCRIPCION
    """
    c1 = conn()
    cur1 = c1.cursor()
    cur1.execute(sql, params)
    rows = cur1.fetchall()
    c1.close()

    # ── 3. Remanente por depósito (paralelo para mayor velocidad)
    from concurrent.futures import ThreadPoolExecutor

    def _fetch_rem(dep):
        try:
            cx = conn()
            cu = cx.cursor()
            cu.execute(f'SELECT ID_ARTICULO, STOCKREMANENTE FROM "FMA_STOCK"(NULL, NULL, \'{dep}\', 1, 1)')
            result = {row[0]: float(row[1] or 0) for row in cu.fetchall()}
            cx.close()
            return result
        except Exception:
            return {}

    with ThreadPoolExecutor(max_workers=2) as ex:
        f001 = ex.submit(_fetch_rem, '001')
        f003 = ex.submit(_fetch_rem, '003')
        rem_001_map = f001.result()
        rem_003_map = f003.result()

    # ── 4. Armar resultado
    result = []
    for r in rows:
        moneda = (r[8] or '').strip().upper()
        factor = cambio_usd if moneda == 'DOLARES' else 1.0
        precio = math.ceil(float(r[3]) * factor * 100) / 100 if r[3] else 0
        result.append({
            "codigo":           r[1] or r[0],
            "descripcion":      (r[2] or '').strip(),
            "stock":            float(r[6] or 0),
            "rem_001":          rem_001_map.get(r[0], 0),
            "rem_003":          rem_003_map.get(r[0], 0),
            "rem_total":        float(r[7] or 0),
            "precio":           precio,
            "iva":              float(r[4]) if r[4] else 21,
            "rubro":            (r[9] or '').strip(),
            "superrubro":       (r[10] or '').strip(),
            "gruposuperrubro":  (r[11] or '').strip(),
        })
    return result


@app.get("/stock/exportar-excel")
def exportar_stock_excel(
    buscar: Optional[str] = None,
    gruposuperrubro: Optional[str] = None,
    superrubro: Optional[str] = None,
    rubro: Optional[str] = None,
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        rows = _fetch_stock_data(buscar, gruposuperrubro, superrubro, rubro)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock Remanente"

        hdr_fill = PatternFill("solid", fgColor="1A56DB")
        hdr_font = Font(bold=True, color="FFFFFFFF", size=10)
        alt_fill = PatternFill("solid", fgColor="FFEFF6FF")
        center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
        right_al = Alignment(horizontal="right",  vertical="center")
        left_al  = Alignment(horizontal="left",   vertical="center")

        headers    = ["Gr. Super Rubro", "Super Rubro", "Rubro",
                      "Código", "Descripción",
                      "Rem. VAC-LOG (001)", "Rem. Pacheco (003)", "Rem. Total",
                      "Precio s/IVA", "IVA %"]
        col_widths = [22, 22, 18, 14, 52, 18, 18, 14, 18, 8]

        ws.row_dimensions[1].height = 24
        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center
            ws.column_dimensions[get_column_letter(ci)].width = w

        for ri, row in enumerate(rows, 2):
            vals = [row["gruposuperrubro"], row["superrubro"], row["rubro"],
                    row["codigo"], row["descripcion"],
                    int(row["rem_001"]), int(row["rem_003"]), int(row["rem_total"]),
                    row["precio"], int(row["iva"])]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.alignment = left_al if ci <= 5 else right_al
                if ri % 2 == 0:
                    cell.fill = alt_fill

        ws.freeze_panes = "A2"
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=stock_remanente.xlsx"})
    except Exception as e:
        import traceback
        raise HTTPException(status_code=500, detail=traceback.format_exc())


@app.get("/stock/exportar-pdf")
def exportar_stock_pdf(
    buscar: Optional[str] = None,
    gruposuperrubro: Optional[str] = None,
    superrubro: Optional[str] = None,
    rubro: Optional[str] = None,
):
    from datetime import datetime
    from reportlab.lib.pagesizes import landscape
    rows = _fetch_stock_data(buscar, gruposuperrubro, superrubro, rubro)

    # Datos empresa
    razon_soc = cuit_emp = dir_emp = tel_emp = email_emp = ''
    try:
        cp = conn()
        ccp = cp.cursor()
        ccp.execute('SELECT RAZONSOCIAL, CUIT, DIRECCION, TELEFONO, EMAIL FROM "PARAMETROS" WHERE CODIGOPARAMETRO = 1')
        rp = ccp.fetchone()
        cp.close()
        if rp:
            razon_soc, cuit_emp, dir_emp, tel_emp, email_emp = [(v or '').strip() for v in rp]
    except Exception:
        pass

    buf = BytesIO()
    PAGE_W, PAGE_H = landscape(A4)
    mg = 14 * mm

    s_title = ParagraphStyle('t', fontSize=13, leading=16, fontName='Helvetica-Bold',
                              textColor=colors.HexColor('#1a56db'))
    s_sub   = ParagraphStyle('s', fontSize=8,  leading=10, fontName='Helvetica',
                              textColor=colors.HexColor('#6b7280'))
    s_ft    = ParagraphStyle('f', fontSize=6.5, leading=9, fontName='Helvetica',
                              alignment=TA_CENTER, textColor=colors.HexColor('#6b7280'))
    s_hdr   = ParagraphStyle('h', fontSize=7.5, leading=9, fontName='Helvetica-Bold',
                              alignment=TA_CENTER, textColor=colors.white)
    s_cell  = ParagraphStyle('c', fontSize=7.5, leading=9, fontName='Helvetica')
    s_cell_r= ParagraphStyle('cr', fontSize=7.5, leading=9, fontName='Helvetica', alignment=TA_RIGHT)

    usable_w = PAGE_W - 2 * mg
    footer_txt = f'{razon_soc}  ·  CUIT {cuit_emp}  ·  {dir_emp}  ·  Tel {tel_emp}  ·  {email_emp}'

    def _on_page(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor('#e5e7eb'))
        canvas.setLineWidth(0.5)
        canvas.line(mg, 10*mm, PAGE_W - mg, 10*mm)
        canvas.setFont('Helvetica', 6.5)
        canvas.setFillColor(colors.HexColor('#6b7280'))
        canvas.drawCentredString(PAGE_W / 2, 7*mm, footer_txt)
        canvas.drawRightString(PAGE_W - mg, 7*mm, f"Pág. {doc.page}")
        canvas.restoreState()

    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=mg, rightMargin=mg,
                            topMargin=mg, bottomMargin=18*mm)

    story = []

    # Logo + título
    logo_row = []
    if os.path.exists(LOGO_PATH):
        logo_row.append(Image(LOGO_PATH, width=38*mm, height=12*mm))
    else:
        logo_row.append(Paragraph('', s_sub))

    filtro_txt = " · ".join(filter(None, [
        f"Buscar: {buscar}" if buscar else None,
        f"GSR: {gruposuperrubro}" if gruposuperrubro else None,
        f"SR: {superrubro}" if superrubro else None,
        f"Rubro: {rubro}" if rubro else None,
        "Sin filtro" if not any([buscar, gruposuperrubro, superrubro, rubro]) else None,
    ]))
    titulo_cell = [
        Paragraph("Stock Remanente", s_title),
        Paragraph(f"{filtro_txt}  —  {datetime.now().strftime('%d/%m/%Y %H:%M')}  —  {len(rows)} artículos", s_sub),
    ]
    t_hdr = Table([[logo_row[0], titulo_cell]], colWidths=[42*mm, usable_w - 42*mm])
    t_hdr.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN',  (1,0), (1,0),  'RIGHT'),
    ]))
    story.append(t_hdr)
    story.append(Spacer(1, 4*mm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#1a56db')))
    story.append(Spacer(1, 3*mm))

    # Tabla de datos
    HDR_BG = colors.HexColor('#1a56db')
    ALT_BG = colors.HexColor('#eff6ff')
    COL_001 = colors.HexColor('#2563eb')
    COL_003 = colors.HexColor('#7c3aed')
    COL_TOT = colors.HexColor('#059669')

    # Anchos de columna — 10 cols llenando usable_w (landscape A4 ≈ 269mm)
    cw = [28*mm, 28*mm, 24*mm, 16*mm, 90*mm, 18*mm, 18*mm, 16*mm, 22*mm, 10*mm]
    # suma: 28+28+24+16+90+18+18+16+22+10 = 270mm ≈ usable_w

    data = [[
        Paragraph('Gr. Super\nRubro',   s_hdr),
        Paragraph('Super\nRubro',       s_hdr),
        Paragraph('Rubro',              s_hdr),
        Paragraph('Código',             s_hdr),
        Paragraph('Descripción',        s_hdr),
        Paragraph('Rem.\nVAC-LOG',      s_hdr),
        Paragraph('Rem.\nPacheco',      s_hdr),
        Paragraph('Rem.\nTotal',        s_hdr),
        Paragraph('Precio s/IVA',       s_hdr),
        Paragraph('IVA',                s_hdr),
    ]]

    for row in rows:
        data.append([
            Paragraph(row['gruposuperrubro'] or '—', s_cell),
            Paragraph(row['superrubro'] or '—',      s_cell),
            Paragraph(row['rubro'] or '—',           s_cell),
            Paragraph(str(row['codigo']),            s_cell),
            Paragraph(row['descripcion'],            s_cell),
            Paragraph(f"{int(row['rem_001']):,}".replace(',', '.'),   s_cell_r),
            Paragraph(f"{int(row['rem_003']):,}".replace(',', '.'),   s_cell_r),
            Paragraph(f"{int(row['rem_total']):,}".replace(',', '.'), s_cell_r),
            Paragraph('$'+f"{row['precio']:,.2f}".replace(',','X').replace('.',',').replace('X','.'), s_cell_r),
            Paragraph(f"{int(row['iva'])}%",         s_cell_r),
        ])

    tbl = Table(data, colWidths=cw, repeatRows=1)
    style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), HDR_BG),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, ALT_BG]),
        ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor('#e5e7eb')),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        # Color columnas remanente (índices 5,6,7 sin Stock Total)
        ('TEXTCOLOR', (5,1), (5,-1), COL_001),
        ('TEXTCOLOR', (6,1), (6,-1), COL_003),
        ('TEXTCOLOR', (7,1), (7,-1), COL_TOT),
        ('FONTNAME',  (5,1), (7,-1), 'Helvetica-Bold'),
    ])
    tbl.setStyle(style)
    story.append(tbl)

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": "inline; filename=stock_remanente.pdf"})


@app.get("/stock/{codigo}")
def get_articulo(codigo: str):
    c = conn()
    cur = c.cursor()
    # Cotización USD actual
    try:
        cur.execute('SELECT CAMBIO FROM "MONEDAS" WHERE CODIGOMONEDA = ?', ('DOLARES',))
        rm = cur.fetchone()
        cambio_usd = float(rm[0]) if rm else 1.0
    except Exception:
        cambio_usd = 1.0
    cur.execute(
        'SELECT a.CODIGOARTICULO, a.DESCRIPCION, a.DESCRIPCIONADICIONAL, a.CODIGOMARCA, '
        'a.PRECIOLISTA1, a.PRECIOLISTA2, a.PRECIOLISTA3, a.ALICUOTAIVA, a.CODIGOUNIDADMEDIDA, '
        'a.CODIGOMONEDA, a.CODIGOPARTICULAR '
        'FROM "ARTICULOS" a WHERE a.CODIGOPARTICULAR = ? OR a.CODIGOARTICULO = ?', (codigo, codigo)
    )
    row = cur.fetchone()
    if not row:
        c.close()
        raise HTTPException(404, "Artículo no encontrado")
    moneda = (row[9] or '').strip().upper()
    factor = cambio_usd if moneda == 'DOLARES' else 1.0
    def conv(v): return math.ceil(float(v) * factor * 100) / 100 if v else 0
    codigoarticulo = row[0]
    # Remanente por depósito usando FMA_STOCK
    rem = {}
    for dep in ['001', '003']:
        try:
            cur.execute(f"""
                SELECT FIRST 1 STOCKREMANENTE
                FROM "FMA_STOCK"(?, NULL, '{dep}', 1, 1)
            """, (codigoarticulo,))
            r = cur.fetchone()
            rem[dep] = float(r[0]) if r and r[0] is not None else 0.0
        except Exception:
            rem[dep] = 0.0
    c.close()
    return {
        "codigo": codigoarticulo, "codigoparticular": (row[10] or row[0] or '').strip(),
        "descripcion": row[1], "descripcion_adicional": row[2],
        "marca": row[3], "precio1": conv(row[4]), "precio2": conv(row[5]), "precio3": conv(row[6]),
        "iva": row[7], "unidad": row[8], "moneda": row[9],
        "remanente_001": rem['001'], "remanente_003": rem['003']
    }

# ─── Clientes (solo del vendedor) ─────────────────────────────────────────────
@app.get("/clientes")
def get_clientes(
    vendedor: Optional[str] = None,
    buscar: Optional[str] = None,
    limit: int = Query(100, le=300),
    offset: int = 0
):
    params = []
    where_vendedor = ""
    if vendedor:
        where_vendedor = "AND CODIGOVENDEDOR = ?"
        params = [vendedor.upper()]

    where_buscar = ""
    if buscar:
        where_buscar = "AND (UPPER(RAZONSOCIAL) CONTAINING UPPER(?) OR CODIGOCLIENTE CONTAINING ?)"
        params += [buscar, buscar]

    try:
        c1 = conn()
        cur1 = c1.cursor()
        cur1.execute(f"""
            SELECT FIRST {limit} SKIP {offset}
                CODIGOCLIENTE, RAZONSOCIAL, NOMBREFANTASIA, CUIT,
                TELEFONO, TELEFONOCELULAR, EMAIL, DIRECCION, LOCALIDAD,
                CODIGOPARTICULAR, REPARTOPROPIO, CONDICIONIVA,
                CODIGOMULTIPLAZO, MULTIPLAZOFIJO, COMENTARIOS,
                LIMITECREDITO, LIMITECREDITODOC
            FROM "CLIENTES"
            WHERE ACTIVO = '1' {where_vendedor}
                {where_buscar}
            ORDER BY RAZONSOCIAL
        """, params)
        rows = cur1.fetchall()
        c1.close()
        return [{
            "codigo": r[0], "razonsocial": r[1], "fantasia": r[2],
            "cuit":   r[3], "telefono":   r[4], "celular":  r[5],
            "email":  r[6], "direccion":  r[7], "localidad": r[8],
            "codigoparticular": (r[9] or "").strip(),
            "tipoiva": (r[11] or "").strip(),
            "discrimina_iva": (r[11] or "").strip().upper() == 'RI',
            "reparto_propio": str(r[10] or '0').strip() == '1',
            "codigomultiplazo": str(r[12] or '0').strip(),
            "multiplazofijo": int(r[13] or 0),
            "comentarios": (r[14] or '').strip(),
            "limitecredito": float(r[15] or 0),
            "limitecreditodoc": float(r[16] or 0),
        } for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error clientes: {e}")

@app.get("/clientes/{codigo}")
def get_cliente(codigo: str):
    c = conn()
    cur = c.cursor()
    cur.execute(
        'SELECT CODIGOCLIENTE, RAZONSOCIAL, NOMBREFANTASIA, CUIT, '
        'TELEFONO, TELEFONOCELULAR, EMAIL, DIRECCION, LOCALIDAD, COMENTARIOS '
        'FROM "CLIENTES" WHERE CODIGOCLIENTE = ? OR CODIGOPARTICULAR = ?', (codigo, codigo)
    )
    row = cur.fetchone()
    c.close()
    if not row:
        raise HTTPException(404, "Cliente no encontrado")
    return {
        "codigo": row[0], "razonsocial": row[1], "fantasia": row[2],
        "cuit": row[3], "telefono": row[4], "celular": row[5],
        "email": row[6], "direccion": row[7], "localidad": row[8],
        "comentarios": row[9]
    }

@app.get("/debug/cliente/{codigo}")
def debug_cliente(codigo: str):
    """Muestra todas las columnas de CLIENTES para un cliente dado (busca en todas las DBs)."""
    _DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    for db_path in [DATABASE, _DB_PROD]:
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            cur.execute(
                'SELECT FIRST 1 * FROM "CLIENTES" WHERE CODIGOCLIENTE = ? '
                'OR TRIM(CODIGOCLIENTE) = ? OR RAZONSOCIAL CONTAINING ?',
                (codigo, codigo.strip(), codigo)
            )
            row = cur.fetchone()
            if not row:
                c.close(); continue
            cols = [d[0] for d in cur.description]
            data = {}
            for k, v in zip(cols, row):
                if v is not None:
                    data[k] = str(v).strip() if isinstance(v, str) else v
            c.close()
            keywords = ['MULTI','PLAZO','COMENT','LIMITE','CREDITO','FIJO','INHAB','DESHAB']
            filtrado = {k: v for k, v in data.items() if any(kw in k.upper() for kw in keywords)}
            return {"db": db_path, "filtrado": filtrado, "todos_los_campos": list(cols)}
        except Exception as e:
            continue
    return {"error": "cliente no encontrado en ninguna DB"}

@app.get("/debug/sucursales/{codigo}")
def debug_sucursales(codigo: str):
    """Diagnóstico: muestra el paso a paso de resolución de sucursales para un cliente."""
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    resultado = {"codigo_recibido": codigo, "pasos": []}

    # Paso 1: DB-Prueba
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        cur.execute('SELECT CODIGOCLIENTE, CODIGOPARTICULAR, DIRECCION, RAZONSOCIAL '
                    'FROM "CLIENTES" WHERE CODIGOCLIENTE = ? OR CODIGOPARTICULAR = ?',
                    (codigo, codigo))
        r = cur.fetchone()
        c.close()
        resultado["pasos"].append({"db": "DB-Prueba", "encontrado": r is not None,
            "fila": [str(x) for x in r] if r else None})
        codigoparticular = (r[1] or '').strip() if r else codigo
    except Exception as e:
        resultado["pasos"].append({"db": "DB-Prueba", "error": str(e)})
        codigoparticular = codigo

    resultado["codigoparticular_resuelto"] = codigoparticular

    # Paso 2: DB-Microbell - resolver CODIGOCLIENTE
    try:
        c = conn('WIN1252', db=DB_PROD)
        cur = c.cursor()
        cur.execute('SELECT CODIGOCLIENTE, CODIGOPARTICULAR, DIRECCION, RAZONSOCIAL '
                    'FROM "CLIENTES" WHERE CODIGOPARTICULAR = ? OR CODIGOCLIENTE = ?',
                    (codigoparticular, codigoparticular))
        r = cur.fetchone()
        resultado["pasos"].append({"db": "DB-Microbell CLIENTES", "encontrado": r is not None,
            "fila": [str(x) for x in r] if r else None})
        cod_mb = str(r[0]).strip() if r else codigoparticular
        resultado["codigocliente_microbell"] = cod_mb

        # Paso 3: transporte a nivel cliente en CLIENTES
        cur.execute('SELECT CODIGOTRANSPORTE, TRANSPORTEFIJO, REPARTOPROPIO '
                    'FROM "CLIENTES" WHERE CODIGOCLIENTE = ?', (cod_mb,))
        r_cli2 = cur.fetchone()
        cols_cli2 = [d[0] for d in cur.description] if cur.description else []
        resultado["clientes_columnas_transporte"] = [c for c in cols_cli2 if 'TRANS' in c.upper() or 'REPARTO' in c.upper()]
        transp_cli = str(r_cli2[0]).strip() if r_cli2 and r_cli2[0] else None
        transp_fijo = str(r_cli2[1] or '0').strip() == '1' if r_cli2 else False
        if r_cli2:
            resultado["clientes_transporte_valores"] = {
                "CODIGOTRANSPORTE": transp_cli, "TRANSPORTEFIJO": transp_fijo
            }

        # Paso 4: SUCURSALESXCLIENTES
        cur.execute('SELECT CODIGOSUCURSAL, NOMBRE, DIRECCION, CODIGOTRANSPORTE '
                    'FROM "SUCURSALESXCLIENTES" WHERE CODIGOCLIENTE = ? ORDER BY CODIGOSUCURSAL',
                    (cod_mb,))
        rows = cur.fetchall()
        resultado["sucursales_count"] = len(rows)
        resultado["sucursales"] = [{"cod": str(r[0]), "nombre": str(r[1]), "dir": str(r[2]),
                                    "transp_suc": str(r[3]) if r[3] else None,
                                    "transp_efectivo": str(r[3]).strip() if r[3] else transp_cli,
                                    "fijo": transp_fijo} for r in rows]
        c.close()
    except Exception as e:
        resultado["pasos"].append({"db": "DB-Microbell", "error": str(e)})

    return resultado


@app.get("/debug/despachos/{tipo}/{numero}")
def debug_despachos(tipo: str, numero: str):
    """Busca despachos para un comprobante en todas las tablas posibles."""
    result = {}
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    for db_label, db_path in [('prueba', DATABASE), ('prod', DB_PROD)]:
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            # 1. Columnas de CUERPOCOMPROBANTES
            try:
                cur.execute("SELECT TRIM(f.RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS f WHERE f.RDB$RELATION_NAME='CUERPOCOMPROBANTES' ORDER BY f.RDB$FIELD_POSITION")
                result[f'{db_label}_cuerpo_cols'] = [r[0] for r in cur.fetchall()]
            except Exception as e:
                result[f'{db_label}_cuerpo_cols_err'] = str(e)
            # 2. Primera fila del comprobante (todos los campos)
            try:
                cur.execute('SELECT FIRST 1 * FROM "CUERPOCOMPROBANTES" WHERE TIPOCOMPROBANTE=? AND NUMEROCOMPROBANTE=?', (tipo, numero))
                row = cur.fetchone()
                if row:
                    cols = [d[0] for d in cur.description]
                    result[f'{db_label}_cuerpo_row'] = dict(zip(cols, [str(v) for v in row]))
            except Exception as e:
                result[f'{db_label}_cuerpo_row_err'] = str(e)
            # 3. Buscar tablas con "DESP" en el nombre
            try:
                cur.execute("SELECT DISTINCT TRIM(r.RDB$RELATION_NAME) FROM RDB$RELATIONS r WHERE r.RDB$RELATION_NAME LIKE '%DESP%' ORDER BY 1")
                result[f'{db_label}_tablas_desp'] = [r[0] for r in cur.fetchall()]
            except Exception as e:
                result[f'{db_label}_tablas_desp_err'] = str(e)
            # 4. Intentar cada tabla DESP encontrada
            for tbl in result.get(f'{db_label}_tablas_desp', []):
                try:
                    cur.execute(f'SELECT FIRST 3 * FROM "{tbl}"')
                    cols2 = [d[0] for d in cur.description]
                    rows2 = cur.fetchall()
                    result[f'{db_label}_{tbl}_cols'] = cols2
                    result[f'{db_label}_{tbl}_sample'] = [dict(zip(cols2,[str(v) for v in r])) for r in rows2]
                    # Buscar si tiene NUMEROCOMPROBANTE o TIPOCOMPROBANTE
                    if any(c in cols2 for c in ['NUMEROCOMPROBANTE','NROCOMPROBANTE']):
                        num_col = 'NUMEROCOMPROBANTE' if 'NUMEROCOMPROBANTE' in cols2 else 'NROCOMPROBANTE'
                        cur.execute(f'SELECT FIRST 5 * FROM "{tbl}" WHERE {num_col}=?', (numero,))
                        rows3 = cur.fetchall()
                        result[f'{db_label}_{tbl}_match'] = [dict(zip(cols2,[str(v) for v in r])) for r in rows3]
                except Exception as e:
                    result[f'{db_label}_{tbl}_err'] = str(e)
            c.close()
        except Exception as e:
            result[f'{db_label}_conn_err'] = str(e)
    return result


@app.get("/debug/cae/{numero}")
def debug_cae(numero: str):
    """Busca el CAE para un comprobante en tablas conocidas de Flexxus."""
    result = {}
    # NUMEROCOMPROBANTE almacenado como float sin formato
    num_float = numero  # el caller pasa el número tal como viene del comp_cols
    try:
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()
        # 1. Buscar en CAMPOSDINAMICOSCOMPROBANTES
        try:
            cur.execute(
                'SELECT CODIGOCAMPODINAMICO, VALOR FROM "CAMPOSDINAMICOSCOMPROBANTES" '
                'WHERE NUMEROCOMPROBANTE = ?', (num_float,)
            )
            result['CAMPOSDINAMICOS'] = [{"cod": r[0], "valor": str(r[1])} for r in cur.fetchall()]
        except Exception as e:
            result['CAMPOSDINAMICOS'] = str(e)
        # 2. Buscar en COMPROBANTESCAE si existe
        try:
            cur.execute(
                'SELECT * FROM "COMPROBANTESCAE" WHERE NUMEROCOMPROBANTE = ?', (num_float,)
            )
            row = cur.fetchone()
            if row:
                cols = [d[0] for d in cur.description]
                result['COMPROBANTESCAE'] = dict(zip(cols, [str(v) for v in row]))
            else:
                result['COMPROBANTESCAE'] = 'no encontrado'
        except Exception as e:
            result['COMPROBANTESCAE'] = str(e)
        # 3. Buscar tablas que contengan "CAE" en campos o nombre
        try:
            cur.execute(
                "SELECT DISTINCT f.RDB$RELATION_NAME "
                "FROM RDB$RELATION_FIELDS f "
                "WHERE f.RDB$FIELD_NAME CONTAINING 'CAE' "
                "AND f.RDB$SYSTEM_FLAG = 0"
            )
            result['tablas_con_CAE'] = [r[0].strip() for r in cur.fetchall()]
        except Exception as e:
            result['tablas_con_CAE'] = str(e)
        # 4. Consultar CAEAFIP
        try:
            cur.execute('SELECT FIRST 1 * FROM "CAEAFIP" WHERE NUMEROCOMPROBANTE = ?', (num_float,))
            row = cur.fetchone()
            if row:
                cols = [d[0] for d in cur.description]
                result['CAEAFIP'] = dict(zip(cols, [str(v) if v is not None else None for v in row]))
            else:
                # Mostrar estructura + últimos registros para ver formato
                cur.execute('SELECT FIRST 3 * FROM "CAEAFIP" ORDER BY 1 DESC')
                rows = cur.fetchall()
                cols = [d[0] for d in cur.description]
                result['CAEAFIP'] = {
                    'sin_dato_para_numero': num_float,
                    'muestra': [dict(zip(cols, [str(v) if v is not None else None for v in r])) for r in rows]
                }
        except Exception as e:
            result['CAEAFIP'] = str(e)
        c.close()
    except Exception as e:
        result['error'] = str(e)
    return result

@app.get("/debug/comp_cols/{numero}")
def debug_comp_cols(numero: str, db: str = Query("oficial")):
    """Muestra TODOS los campos de CABEZACOMPROBANTES para un comprobante dado."""
    db_path = DATABASE_MLT if db == 'sw' else DATABASE
    try:
        c = conn('WIN1252', db=db_path)
        cur = c.cursor()
        # Intentar búsqueda exacta primero, luego por LIKE
        cur.execute(
            'SELECT FIRST 1 * FROM "CABEZACOMPROBANTES" '
            'WHERE NUMEROCOMPROBANTE = ? OR NUMEROCOMPROBANTE LIKE ?',
            (numero, f'%{numero}%')
        )
        row = cur.fetchone()
        if not row:
            # Devolver los últimos 5 comprobantes para ver el formato real
            cur.execute(
                'SELECT FIRST 5 TIPOCOMPROBANTE, NUMEROCOMPROBANTE, FECHACOMPROBANTE '
                'FROM "CABEZACOMPROBANTES" ORDER BY FECHACOMPROBANTE DESC'
            )
            sample = cur.fetchall()
            c.close()
            return {"error": f"No encontrado", "formato_muestra": [
                {"tipo": r[0], "numero": r[1], "fecha": str(r[2])} for r in sample
            ]}
        cols = [d[0] for d in cur.description]
        c.close()
        return dict(zip(cols, [str(v) if v is not None else None for v in row]))
    except Exception as e:
        return {"error": str(e)}

# ─── Cuenta Corriente ──────────────────────────────────────────────────────────
def _resolve_codigos_en_db(db_path, codigoparticular):
    """Dado un CODIGOPARTICULAR, devuelve los CODIGOCLIENTEs que lo usan en esa BD via CUERPOCOMPROBANTES."""
    if not codigoparticular or not str(codigoparticular).strip():
        return []
    try:
        c = conn('LATIN1', db=db_path)
        cur = c.cursor()
        cur.execute(
            'SELECT DISTINCT CODIGOCLIENTE FROM "CUERPOCOMPROBANTES" WHERE CODIGOPARTICULAR = ?',
            (str(codigoparticular).strip(),)
        )
        rows = cur.fetchall()
        c.close()
        return [str(r[0]).strip() for r in rows if r[0] and str(r[0]).strip()]
    except Exception:
        return []

def _query_cta_part(db_path, codigos_base, codigoparticular, limit, offset):
    """Como _query_cta pero resuelve el CODIGOCLIENTE correcto para esta BD vía CODIGOPARTICULAR."""
    codigos = set(c for c in codigos_base if c)
    codigos.update(_resolve_codigos_en_db(db_path, codigoparticular))
    if not codigos:
        return []
    return _query_cta(db_path, list(codigos), limit, offset)

def _get_cambios(db_path):
    """Retorna dict {CODIGOMONEDA: CAMBIO} desde tabla MONEDAS."""
    try:
        c = conn('WIN1252', db=db_path)
        cur = c.cursor()
        cur.execute('SELECT CODIGOMONEDA, CAMBIO FROM "MONEDAS"')
        result = {str(r[0]).strip(): float(r[1] or 1) for r in cur.fetchall()}
        c.close()
        return result
    except Exception:
        return {}

def _query_cta(db_path, codigos, limit, offset):
    """Consulta CABEZACOMPROBANTES sin CAST/JOIN en SQL. Conversión de moneda en Python."""
    try:
        c = conn('WIN1252', db=db_path)
        cur = c.cursor()
        ph = ', '.join(['?'] * len(codigos))
        cur.execute(f"""
            SELECT FIRST {limit} SKIP {offset}
                TIPOCOMPROBANTE, NUMEROCOMPROBANTE, FECHACOMPROBANTE,
                TOTAL, IVA1, IVA2, PAGADO, COTIZACION, CODIGOMONEDA,
                FECHAVENCIMIENTO, CLASECOMPROBANTE
            FROM "CABEZACOMPROBANTES"
            WHERE CODIGOCLIENTE IN ({ph})
              AND CUENTACORRIENTE = '1'
              AND ANULADA = '0'
              AND TIPOCOMPROBANTE NOT IN ('RE', 'RI', 'INA')
            ORDER BY FECHAVENCIMIENTO ASC, FECHACOMPROBANTE ASC
        """, tuple(codigos))
        raw = cur.fetchall()
        c.close()
    except Exception:
        return []

    cambios = _get_cambios(db_path)
    result = []
    for r in raw:
        try:
            tipo   = r[0]; num   = r[1]; fecha = r[2]
            total  = float(r[3] or 0)
            iva1   = float(r[4] or 0)
            iva2   = float(r[5] or 0)
            pagado = float(r[6] or 0)
            cotiz  = float(r[7] or 1) or 1.0
            moneda = str(r[8] or '').strip()
            fvto   = r[9];  clase = r[10]
            neto   = total + iva1 + iva2
            debe   = neto - pagado
            cambio = cambios.get(moneda, 1.0) or 1.0
            deuda  = debe * cambio / cotiz
            if abs(deuda) >= 0.01:
                result.append((tipo, num, fecha, neto, pagado, deuda, fvto, clase))
        except Exception:
            continue
    return result

@app.get("/clientes/{codigo}/credito")
def get_credito_cliente(codigo: str):
    """Retorna límites de crédito y saldo deudor actual del cliente."""
    DB_PROD     = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    # Buscar primero en DATABASE (DB-Prueba = BD operativa de vendedores),
    # luego en DB_PROD como fallback. Los pedidos y clientes se operan en DATABASE.
    lim_cred = lim_doc = 0.0
    cod_real = codigo
    part_real = None
    _db_fuente = None
    _db_errores = {}
    for db_path in [DATABASE, DB_PROD]:
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            cur.execute(
                'SELECT LIMITECREDITO, LIMITECREDITODOC, CODIGOCLIENTE, '
                'MULTIPLAZOFIJO, CODIGOMULTIPLAZO, CODIGOPARTICULAR '
                'FROM "CLIENTES" WHERE CODIGOCLIENTE = ? OR CODIGOPARTICULAR = ?',
                (codigo, codigo)
            )
            row = cur.fetchone()
            c.close()
            if row:
                lim_cred       = float(row[0] or 0)
                lim_doc        = float(row[1] or 0)
                cod_real       = str(row[2]).strip()
                multiplazofijo = int(row[3] or 0)
                codigomulti    = str(row[4] or '0').strip()
                part_real      = str(row[5] or '').strip() or None
                _db_fuente     = db_path
                break
        except Exception as _e:
            _db_errores[db_path] = str(_e)
    # Saldo deudor: solo CODIGOCLIENTE (nunca mezclar con CODIGOPARTICULAR en CABEZACOMPROBANTES)
    saldo_deudor = 0.0
    seen_sd = set()
    for db_path in [DB_PROD, DB_MLT_PROD]:
        try:
            rows = _query_cta(db_path, [cod_real], 500, 0)
            for r in rows:
                key = (r[0], r[1])
                if key not in seen_sd:
                    seen_sd.add(key)
                    saldo_deudor += float(r[5] or 0)
        except Exception:
            pass
    saldo_deudor = round(saldo_deudor, 2)
    # Pedidos "A preparar": OPERACION='1' (Flexxus setea FECHATERMINADA al confirmar,
    # no al entregar — por eso filtramos por OPERACION, no por FECHATERMINADA IS NULL).
    # TOTAL en Flexxus = monto bruto con IVA incluido.
    pedidos_pendientes = 0.0
    for db_path in [DATABASE, DB_PROD]:
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            # Excluye pedidos que ya tienen remito (CANTIDADREMITIDA > 0 en alguna línea)
            # aunque Flexxus no haya cambiado el OPERACION manualmente.
            cur.execute(
                'SELECT COALESCE(SUM(cb.TOTAL), 0) FROM "CABEZAPEDIDOS" cb '
                'WHERE cb.TIPOCOMPROBANTE = ? AND cb.CODIGOCLIENTE = ? '
                'AND cb.ANULADA = ? AND cb.OPERACION = ? '
                'AND NOT EXISTS ('
                '  SELECT 1 FROM "CUERPOPEDIDOS" cue '
                '  WHERE cue.TIPOCOMPROBANTE = ? '
                '    AND cue.NUMEROCOMPROBANTE = cb.NUMEROCOMPROBANTE '
                '    AND cue.CANTIDADREMITIDA > 0'
                ')',
                ('NP', cod_real, '0', '1', 'NP')
            )
            row = cur.fetchone()
            c.close()
            if row:
                pedidos_pendientes += float(row[0] or 0)
                break  # BD operativa encontrada, no sumar la otra
        except Exception:
            pass
    pedidos_pendientes = round(pedidos_pendientes, 2)
    disponible_total = round(max(0, lim_cred + lim_doc - saldo_deudor - pedidos_pendientes), 2)
    return {
        "limitecredito":      lim_cred,
        "limitecreditodoc":   lim_doc,
        "saldo_deudor":       saldo_deudor,
        "pedidos_pendientes": pedidos_pendientes,
        "disponible_cred":    round(max(0, lim_cred - saldo_deudor), 2),
        "disponible_doc":     round(max(0, lim_doc  - saldo_deudor), 2),
        "disponible_total":   disponible_total,
        "multiplazofijo":     multiplazofijo,
        "codigomultiplazo":   codigomulti,
        "_db_fuente":         _db_fuente,
        "_db_errores":        _db_errores,
    }


@app.get("/clientes/{codigo}/cuenta_corriente")
def cuenta_corriente(codigo: str, limit: int = Query(200, le=500), offset: int = 0):
    # Lookup CODIGOPARTICULAR usando WIN1252 (charset de CLIENTES)
    c_cli = conn('WIN1252')
    cur_cli = c_cli.cursor()
    # Buscar por CODIGOCLIENTE o CODIGOPARTICULAR (el frontend puede mandar cualquiera)
    cur_cli.execute(
        'SELECT CODIGOCLIENTE, CODIGOPARTICULAR FROM "CLIENTES" '
        'WHERE CODIGOCLIENTE = ? OR CODIGOPARTICULAR = ?',
        (codigo, codigo)
    )
    cli = cur_cli.fetchone()
    c_cli.close()
    # Solo usar CODIGOCLIENTE para queries en CABEZACOMPROBANTES (CODIGOPARTICULAR puede ser más largo)
    codigocliente = None
    if cli and cli[0] and cli[0].strip():
        codigocliente = cli[0].strip()
    if not codigocliente:
        codigocliente = codigo
    codigos = [codigocliente]

    DB_PROD      = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD  = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    rows_prod     = _query_cta(DB_PROD,     codigos, limit, offset)
    rows_mlt_prod = _query_cta(DB_MLT_PROD, codigos, limit, offset)

    # Combinar y deduplicar por (tipo, numero)
    seen = set()
    combined = []
    for r in rows_prod + rows_mlt_prod:
        key = (r[0], r[1])  # tipo + numero
        if key not in seen:
            seen.add(key)
            combined.append(r)

    # Ordenar por fecha vencimiento
    combined.sort(key=lambda r: (r[6] or r[2], r[2]))

    return [{
        "tipo":      r[0], "numero":    r[1], "fecha":   r[2],
        "total":     float(r[3]) if r[3] else 0,
        "pagado":    float(r[4]) if r[4] else 0,
        "deuda":     float(r[5]) if r[5] else 0,
        "fecha_vto": r[6], "clase":     r[7],
    } for r in combined]


@app.get("/que-vendi/clientes")
def que_vendi_clientes(vendedor: str, buscar: Optional[str] = None):
    """Clientes del vendedor para autocomplete en ¿Qué Vendí?"""
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    params = [vendedor.upper()]
    where_buscar = ""
    if buscar:
        where_buscar = "AND (UPPER(RAZONSOCIAL) CONTAINING UPPER(?) OR CODIGOCLIENTE CONTAINING ? OR CODIGOPARTICULAR CONTAINING ?)"
        params += [buscar, buscar, buscar]
    try:
        c = conn('WIN1252', DB_PROD)
        cur = c.cursor()
        cur.execute(f"""
            SELECT FIRST 30 CODIGOCLIENTE, RAZONSOCIAL, CODIGOPARTICULAR
            FROM "CLIENTES"
            WHERE ACTIVO = '1' AND UPPER(CODIGOVENDEDOR) = ?
            {where_buscar}
            ORDER BY RAZONSOCIAL
        """, params)
        rows = cur.fetchall()
        c.close()
        return [{"codigo": (r[0] or '').strip(),
                 "razonsocial": (r[1] or '').strip(),
                 "codigoparticular": (r[2] or '').strip()} for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/que-vendi")
def que_vendi(
    vendedor: str,
    cliente: str,
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    limit: int = Query(500, le=2000),
    offset: int = 0
):
    """Artículos facturados a un cliente. Facturas en BDs Prueba; lookup de código en todas las BDs."""
    from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeout
    DB_PROD     = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    # Mismas 4 BDs que Cuenta Corriente — prueba + producción, L1 + SW
    DBS_FACT   = [DATABASE, DATABASE_MLT, DB_PROD, DB_MLT_PROD]
    # Lookup solo en las BDs que tienen tabla CLIENTES
    DBS_LOOKUP = [DATABASE, DB_PROD]

    # ── Lookup de códigos del cliente ────────────────────────────────────────
    def _lookup_codigos(db_path):
        try:
            c = conn('WIN1252', db_path)
            cur = c.cursor()
            cur.execute(
                'SELECT CODIGOCLIENTE, CODIGOPARTICULAR FROM "CLIENTES" '
                'WHERE CODIGOCLIENTE = ? OR CODIGOPARTICULAR = ?',
                (cliente, cliente)
            )
            row = cur.fetchone()
            c.close()
            if row:
                # str() porque CODIGOCLIENTE puede ser INTEGER en Firebird
                return [str(v).strip() for v in row if v is not None and str(v).strip()]
        except Exception:
            pass
        return []

    codigos = set([str(cliente).strip()])
    with ThreadPoolExecutor(max_workers=3) as ex:
        futs = {ex.submit(_lookup_codigos, db): db for db in DBS_LOOKUP}
        for f in futs:
            try:
                for cod in f.result(timeout=8):
                    codigos.add(cod)
            except FutureTimeout:
                pass
    codigos = list(codigos)

    # WHERE dinámico — filtra por CODIGOCLIENTE usando todos los códigos resueltos
    ph = ','.join('?' * len(codigos))
    TIPOS_FACTURA = ("('FA','FB','FE','FCA','FCB','FCE',"
                     "'FCCA','FCCB','FCCE',"
                     "'NCA','NCB','NCCA','NCCB')")
    TIPOS_NC = {'NCA', 'NCB', 'NCCA', 'NCCB'}   # valores se negarán
    params_extra = []
    where_parts = [f"cb.CODIGOCLIENTE IN ({ph})",
                   "cb.ANULADA = '0'",
                   f"cb.TIPOCOMPROBANTE IN {TIPOS_FACTURA}"]
    if desde:
        where_parts.append("cb.FECHACOMPROBANTE >= ?")
        params_extra.append(desde)
    if hasta:
        where_parts.append("cb.FECHACOMPROBANTE <= ?")
        params_extra.append(hasta)
    params_base = list(codigos) + params_extra
    where_sql = " AND ".join(where_parts)

    sql = f"""
        SELECT FIRST {limit} SKIP {offset}
            COALESCE(NULLIF(TRIM(cu.CODIGOPARTICULAR),''),
                     TRIM(cu.CODIGOARTICULO))             AS COD_ART,
            COALESCE(NULLIF(TRIM(cu.DESCRIPCION),''), '') AS DESCR,
            cb.FECHACOMPROBANTE,
            cb.TIPOCOMPROBANTE,
            cb.NUMEROCOMPROBANTE,
            CAST(cu.CANTIDAD       AS DOUBLE PRECISION)  AS CANT,
            CAST(cu.PRECIOUNITARIO AS DOUBLE PRECISION)  AS PU,
            CAST(cu.PRECIOTOTAL    AS DOUBLE PRECISION)  AS SUBTOTAL,
            COALESCE(CAST(cu.PORCENTAJEIVA AS DOUBLE PRECISION), 21) AS IVA_PCT
        FROM "CUERPOCOMPROBANTES" cu
        JOIN "CABEZACOMPROBANTES" cb
             ON cb.TIPOCOMPROBANTE   = cu.TIPOCOMPROBANTE
            AND cb.NUMEROCOMPROBANTE = cu.NUMEROCOMPROBANTE
        WHERE {where_sql}
        ORDER BY cb.FECHACOMPROBANTE DESC, cb.NUMEROCOMPROBANTE DESC
    """

    # ── Consulta paralela a las 4 BDs, timeout 25s por BD ────────────────────
    _db_errors = {}
    def _query_db(db_path):
        try:
            c = conn('LATIN1', db=db_path)
            cur = c.cursor()
            cur.execute(sql, params_base)
            rows = cur.fetchall()
            c.close()
            _QV_LAST_COUNTS[db_path] = len(rows)
            return rows
        except Exception as _e:
            _db_errors[db_path] = str(_e)
            _QV_LAST_ERRORS[db_path] = str(_e)
            _QV_LAST_COUNTS[db_path] = f"ERROR: {_e}"
            return []

    all_rows = []
    seen = set()
    with ThreadPoolExecutor(max_workers=4) as ex:
        futs = {ex.submit(_query_db, db): db for db in DBS_FACT}
        for f in futs:
            try:
                for r in f.result(timeout=25):
                    key = (str(r[3]).strip(), str(r[4]).strip(), str(r[0]).strip())
                    if key not in seen:
                        seen.add(key)
                        all_rows.append(r)
            except FutureTimeout:
                pass

    def _fmt(v):
        if hasattr(v, 'strftime'):
            return v.strftime('%Y-%m-%d')
        return str(v)[:10] if v else ''

    all_rows.sort(key=lambda r: (_fmt(r[2]), str(r[4] or '')), reverse=True)

    result = []
    for r in all_rows:
        try:
            tipo    = (r[3] or '').strip().upper()
            signo   = -1 if tipo in TIPOS_NC else 1
            cant    = signo * float(r[5] or 0)
            pu      = float(r[6] or 0)          # precio unitario siempre positivo
            subtot  = signo * float(r[7] or 0)
            iva_pct = float(r[8] or 21)
            total   = round(subtot * (1 + iva_pct / 100), 2)
            result.append({
                "cod_articulo":    (r[0] or '').strip(),
                "descripcion":     (r[1] or '').strip(),
                "fecha":           _fmt(r[2]),
                "tipo":            tipo,
                "numero":          str(r[4] or '').strip().replace('.0',''),
                "cantidad":        cant,
                "precio_unitario": pu,
                "importe":         subtot,
                "iva_pct":         iva_pct,
                "total":           total,
            })
        except Exception:
            pass   # fila con datos inválidos: se omite
    return result


@app.get("/que-vendi/pdf")
def que_vendi_pdf(vendedor: str, cliente: str,
                  desde: Optional[str] = None, hasta: Optional[str] = None,
                  razon: Optional[str] = None):
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from io import BytesIO
    from datetime import datetime

    rows = que_vendi(vendedor=vendedor, cliente=cliente, desde=desde, hasta=hasta, limit=2000)

    buf = BytesIO()
    # Landscape A4: usable width = 297 - 24 = 273 mm
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12*mm, rightMargin=12*mm,
                            topMargin=10*mm, bottomMargin=10*mm)

    azul = colors.HexColor('#1e429f')
    AR   = lambda v: f"{float(v or 0):,.2f}".replace(',','X').replace('.',',').replace('X','.')
    sDesc= ParagraphStyle('desc', fontSize=7, fontName='Helvetica',    leading=8)
    sHdr = ParagraphStyle('hdr',  fontSize=7, fontName='Helvetica-Bold',
                          textColor=colors.white, leading=8, alignment=1)  # 1=CENTER
    sSub = ParagraphStyle('sub',  fontSize=8, fontName='Helvetica', leading=10)
    sTit = ParagraphStyle('tit',  fontSize=12, fontName='Helvetica-Bold', leading=14)

    # ── Encabezado: logo izquierda + título derecha ─────────────────────────
    per = f"{desde or ''} a {hasta or ''}"
    emi = datetime.now().strftime('%d/%m/%Y %H:%M')
    logo_cell = ''
    if os.path.exists(LOGO_PATH):
        logo_cell = Image(LOGO_PATH, width=38*mm, height=12*mm,
                          kind='proportional')
    hdr_table = Table(
        [[logo_cell,
          [Paragraph('Microbell S.A. — ¿Qué Vendí?', sTit),
           Paragraph(f'Cliente: {razon or cliente}', sSub),
           Paragraph(f'Período: {per}   |   Emisión: {emi}', sSub)]]],
        colWidths=[42*mm, 231*mm]
    )
    hdr_table.setStyle(TableStyle([
        ('VALIGN',  (0,0),(-1,-1), 'MIDDLE'),
        ('ALIGN',   (1,0),(1,0),   'LEFT'),
        ('LEFTPADDING',  (0,0),(-1,-1), 0),
        ('RIGHTPADDING', (0,0),(-1,-1), 4),
        ('TOPPADDING',   (0,0),(-1,-1), 0),
        ('BOTTOMPADDING',(0,0),(-1,-1), 4),
    ]))

    # ── Columnas: suma = 273 mm ──────────────────────────────────────────────
    # CodArt Descripcion Fecha  Tipo  Nro   Cant  PUnit  Importe IVA  Total
    cw = [16*mm, 100*mm, 18*mm, 12*mm, 24*mm, 13*mm, 22*mm, 22*mm, 10*mm, 22*mm]
    # Encabezados como Paragraph para que también hagan wrap si es necesario
    hdrs = [Paragraph(h, sHdr) for h in
            ['Cód.Art.','Descripción','Fecha','Tipo','Nro. Comp.','Cant.','P.Unit.','Importe','IVA%','Total']]
    data = [hdrs]
    tot_imp = tot_tot = 0.0
    for r in rows:
        nro = str(r['numero']).replace('.0','').zfill(10)
        tot_imp += float(r['importe'] or 0)
        tot_tot += float(r['total']   or 0)
        fecha = r['fecha'][8:10]+'/'+r['fecha'][5:7]+'/'+r['fecha'][:4] if r['fecha'] else ''
        data.append([
            r['cod_articulo'],
            Paragraph(r['descripcion'] or '', sDesc),   # ← wrap automático
            fecha, r['tipo'], nro,
            str(int(float(r['cantidad'] or 0))), f"${AR(r['precio_unitario'])}",
            f"${AR(r['importe'])}", AR(r['iva_pct']), f"${AR(r['total'])}",
        ])
    data.append(['','','','','','',
                 Paragraph('TOTALES', ParagraphStyle('tb', fontSize=7, fontName='Helvetica-Bold')),
                 f"${AR(tot_imp)}", '', f"${AR(tot_tot)}"])

    t = Table(data, colWidths=cw, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',   (0,0), (-1,0),  azul),
        ('ROWBACKGROUNDS',(0,1),(-1,-2), [colors.white, colors.HexColor('#f3f4f6')]),
        ('BACKGROUND',   (0,-1),(-1,-1), colors.HexColor('#e0e7ff')),
        ('FONTNAME',     (0,1), (-1,-1), 'Helvetica'),
        ('FONTNAME',     (0,-1),(-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE',     (0,0), (-1,-1), 7),
        ('ALIGN',        (0,0), (-1,-1),  'LEFT'),
        ('ALIGN',        (0,0), (-1,0),  'CENTER'),   # toda la fila de headers
        ('ALIGN',        (0,1), (0,-1),  'CENTER'),   # Cód.Art. (datos)
        ('ALIGN',        (2,0), (2,-1),  'CENTER'),   # Fecha
        ('ALIGN',        (3,0), (3,-1),  'CENTER'),   # Tipo
        ('ALIGN',        (4,0), (4,-1),  'CENTER'),   # Nro
        ('ALIGN',        (5,0), (5,-1),  'CENTER'),   # Cant: centrado
        ('ALIGN',        (6,0), (-1,-1), 'RIGHT'),    # P.Unit → Total
        ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
        ('GRID',         (0,0), (-1,-1), 0.3, colors.HexColor('#d1d5db')),
        ('TOPPADDING',   (0,0), (-1,-1), 2),
        ('BOTTOMPADDING',(0,0), (-1,-1), 2),
        ('LEFTPADDING',  (0,0), (-1,-1), 3),
        ('RIGHTPADDING', (0,0), (-1,-1), 3),
    ]))

    story = [hdr_table, t]
    doc.build(story)
    buf.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(buf, media_type='application/pdf',
        headers={'Content-Disposition': f'inline; filename="que_vendi_{cliente}.pdf"'})


@app.get("/que-vendi/excel")
def que_vendi_excel(vendedor: str, cliente: str,
                    desde: Optional[str] = None, hasta: Optional[str] = None,
                    razon: Optional[str] = None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    rows = que_vendi(vendedor=vendedor, cliente=cliente, desde=desde, hasta=hasta, limit=2000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Que Vendi'

    azul  = PatternFill('solid', fgColor='1e429f')
    gris  = PatternFill('solid', fgColor='f3f4f6')
    azulc = PatternFill('solid', fgColor='e0e7ff')
    bF    = Font(bold=True, color='FFFFFF', size=9)
    bN    = Font(bold=True, size=9)
    nN    = Font(size=9)
    cen   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    der   = Alignment(horizontal='right',  vertical='center')
    izq   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin  = Side(style='thin', color='d1d5db')
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

    hdrs  = ['Cód.Art.','Descripción','Fecha','Tipo','Nro. Comp.','Cant.','P.Unit.','Importe','IVA%','Total']
    widths= [12,        45,           12,     8,     14,          10,     16,       16,       8,     16]
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = azul; cell.font = bF; cell.alignment = cen; cell.border = brd
        ws.column_dimensions[get_column_letter(ci)].width = w

    tot_imp = tot_tot = 0.0
    for ri, r in enumerate(rows, 2):
        nro = str(r['numero']).replace('.0','').zfill(10)
        fg  = None if ri % 2 == 0 else gris
        fecha = r['fecha'][8:10]+'/'+r['fecha'][5:7]+'/'+r['fecha'][:4] if r['fecha'] else ''
        vals = [r['cod_articulo'], r['descripcion'], fecha, r['tipo'], nro,
                float(r['cantidad'] or 0), float(r['precio_unitario'] or 0),
                float(r['importe'] or 0), float(r['iva_pct'] or 0), float(r['total'] or 0)]
        tot_imp += float(r['importe'] or 0); tot_tot += float(r['total'] or 0)
        aligns = [cen, izq, cen, cen, cen, cen, der, der, der, der]
        for ci, (v, al) in enumerate(zip(vals, aligns), 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = nN; cell.alignment = al; cell.border = brd
            if fg: cell.fill = fg
            if ci == 6:              cell.number_format = '#,##0'           # Cant: entero
            elif ci in (7, 8, 10): cell.number_format = '"$"#,##0.00'    # P.Unit/Importe/Total: $
            elif ci >= 7:          cell.number_format = '#,##0.00'        # IVA%: decimal

    tr = len(rows) + 2
    for ci, v in enumerate(['']*6 + ['TOTALES', tot_imp, '', tot_tot], 1):
        cell = ws.cell(row=tr, column=ci, value=v)
        cell.fill = azulc; cell.font = bN; cell.border = brd
        cell.alignment = der if ci >= 6 else cen
        if ci in (8, 10): cell.number_format = '"$"#,##0.00'

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="que_vendi_{cliente}.xlsx"'})


@app.get("/resumen-deudas")
def resumen_deudas(vendedor: Optional[str] = None):
    """Suma de deuda pendiente por cliente, ordenado por deuda desc."""
    DB_PROD     = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'

    # Obtener lista de clientes desde DB-Microbell (donde están los códigos reales de Flexxus)
    c_cli = conn('WIN1252', DB_PROD)
    cur_cli = c_cli.cursor()
    if vendedor:
        cur_cli.execute(
            'SELECT CODIGOCLIENTE, RAZONSOCIAL, CODIGOPARTICULAR FROM "CLIENTES" '
            'WHERE ACTIVO=? AND UPPER(CODIGOVENDEDOR)=? ORDER BY RAZONSOCIAL',
            ('1', vendedor.upper())
        )
    else:
        cur_cli.execute(
            'SELECT CODIGOCLIENTE, RAZONSOCIAL, CODIGOPARTICULAR FROM "CLIENTES" '
            "WHERE ACTIVO='1' ORDER BY RAZONSOCIAL"
        )
    clientes_rows = cur_cli.fetchall()
    c_cli.close()

    deudas = {}
    for cod, razon, part in clientes_rows:
        cod = (cod or '').strip(); razon = (razon or '').strip(); part = (part or '').strip()
        if not cod: continue
        # Solo CODIGOCLIENTE — mismo valor en DB_PROD y DB_MLT_PROD para el mismo cliente
        total_deuda = 0.0
        seen_cta = set()

        for db_path in [DB_PROD, DB_MLT_PROD]:
            try:
                rows = _query_cta(db_path, [cod], 500, 0)
                for r in rows:
                    key = (r[0], r[1])
                    if key not in seen_cta:
                        seen_cta.add(key)
                        total_deuda += float(r[5] or 0)
            except Exception:
                pass

        if total_deuda > 0:
            deudas[cod] = {'codigo': part or cod, 'razonsocial': razon, 'deuda': round(total_deuda, 2)}

    result = sorted(deudas.values(), key=lambda x: x['razonsocial'])
    return result

@app.get("/resumen-deudas/pdf")
def resumen_deudas_pdf(vendedor: Optional[str] = None):
    """PDF con resumen de deudas por cliente — logo + detalle de comprobantes."""
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, KeepTogether
    from reportlab.platypus.flowables import Flowable
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
    from reportlab.platypus import Image
    from datetime import date

    DB_PROD     = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'

    CELESTE = colors.HexColor('#4A90D9')
    GRIS    = colors.HexColor('#f3f4f6')
    ROJO    = colors.HexColor('#dc2626')
    AZUL_CLI= colors.HexColor('#1e3a5f')

    sN  = ParagraphStyle('rdN',  fontSize=7, leading=9,  fontName='Helvetica')
    sNc = ParagraphStyle('rdNc', fontSize=7, leading=9,  fontName='Helvetica',      alignment=TA_CENTER)
    sNr = ParagraphStyle('rdNr', fontSize=7, leading=9,  fontName='Helvetica',      alignment=TA_RIGHT)
    sNb = ParagraphStyle('rdNb', fontSize=7, leading=9,  fontName='Helvetica-Bold', alignment=TA_CENTER)
    sNbr= ParagraphStyle('rdNbr',fontSize=7, leading=9,  fontName='Helvetica-Bold', alignment=TA_CENTER)

    def _fmt_num(v):
        """Convierte 4400014955.0 → '4400014955' (10 dígitos, sin .0)"""
        try:
            return str(int(float(v))).zfill(10)
        except Exception:
            return str(v or '')

    def _fmt_date(v):
        """Convierte datetime o 'yyyy-mm-dd' → 'dd/mm/yyyy'"""
        if not v:
            return '—'
        try:
            from datetime import date as _date, datetime as _dt
            if hasattr(v, 'strftime'):
                return v.strftime('%d/%m/%Y')
            s = str(v)[:10]
            d = _dt.strptime(s, '%Y-%m-%d')
            return d.strftime('%d/%m/%Y')
        except Exception:
            return str(v)[:10]
    sCli= ParagraphStyle('rdCli',fontSize=9, leading=11, fontName='Helvetica-Bold', textColor=AZUL_CLI)
    sT  = ParagraphStyle('rdT',  fontSize=12,leading=15, fontName='Helvetica-Bold')
    sSub= ParagraphStyle('rdSub',fontSize=8, leading=10, fontName='Helvetica', textColor=colors.HexColor('#6b7280'))

    hoy = date.today().strftime('%d/%m/%Y')

    # ── Obtener clientes con deuda y sus comprobantes ─────────────────────────
    # Leer CLIENTES desde DB-Microbell donde están los códigos reales de Flexxus
    DB_PROD     = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    c_cli = conn('WIN1252', DB_PROD)
    cur_cli = c_cli.cursor()
    if vendedor:
        cur_cli.execute(
            'SELECT CODIGOCLIENTE, RAZONSOCIAL, CODIGOPARTICULAR, CUIT FROM "CLIENTES" '
            "WHERE ACTIVO='1' AND UPPER(CODIGOVENDEDOR)=? ORDER BY RAZONSOCIAL",
            (vendedor.upper(),)
        )
    else:
        cur_cli.execute(
            'SELECT CODIGOCLIENTE, RAZONSOCIAL, CODIGOPARTICULAR, CUIT FROM "CLIENTES" '
            "WHERE ACTIVO='1' ORDER BY RAZONSOCIAL"
        )
    clientes_rows = cur_cli.fetchall()
    c_cli.close()

    clientes_data = []
    gran_total_deuda = 0.0
    for cod, razon, part, cuit in clientes_rows:
        cod   = (cod   or '').strip()
        razon = (razon or '').strip()
        part  = (part  or '').strip()
        cuit  = (cuit  or '').strip()
        if not cod: continue
        comprobantes = []
        seen = set()

        for db_path in [DB_PROD, DB_MLT_PROD]:
            try:
                rows = _query_cta(db_path, [cod], 500, 0)
                for r in rows:
                    key = (r[0], r[1])
                    if key not in seen:
                        seen.add(key)
                        comprobantes.append(r)
            except Exception:
                pass

        comprobantes.sort(key=lambda r: (r[6] or r[2], r[2]))
        total_deuda = sum(float(r[5] or 0) for r in comprobantes)
        if total_deuda > 0:
            gran_total_deuda += total_deuda
            clientes_data.append({
                'codigo': part or cod, 'razonsocial': razon, 'cuit': cuit,
                'comprobantes': comprobantes, 'total_deuda': total_deuda
            })

    clientes_data.sort(key=lambda x: x['razonsocial'])

    # ── Armar PDF ─────────────────────────────────────────────────────────────
    buf = BytesIO()
    PAGE_W, PAGE_H = A4
    mg = 14 * mm
    usable_w = PAGE_W - 2 * mg

    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=mg, rightMargin=mg,
                            topMargin=12*mm, bottomMargin=12*mm)

    def _on_page(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 6.5)
        canvas.setFillColor(colors.HexColor('#9ca3af'))
        canvas.drawCentredString(PAGE_W/2, 8*mm,
            f'microbell S.A.  ·  CUIT 30-70839018-2  ·  Resumen de Deudas al {hoy}  ·  Pág. {doc.page}')
        canvas.restoreState()

    story = []

    # ── Header con logo ───────────────────────────────────────────────────────
    logo_cell = Image(LOGO_PATH, width=38*mm, height=13*mm) if os.path.exists(LOGO_PATH) \
                else Paragraph('<b>microbell S.A.</b>', sT)
    titulo_txt = f'Resumen de Deudas Pendientes'
    subtitulo  = f'Fecha: {hoy}'
    if vendedor:
        subtitulo += f'   |   Vendedor: {vendedor}'
    hdr_tbl = Table([[
        logo_cell,
        [Paragraph(titulo_txt, sT), Paragraph(subtitulo, sSub)]
    ]], colWidths=[45*mm, usable_w - 45*mm])
    hdr_tbl.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('ALIGN',(1,0),(1,0),'RIGHT'),
    ]))
    story.append(hdr_tbl)
    story.append(HRFlowable(width=usable_w, thickness=1.5, color=CELESTE, spaceAfter=4))

    # ── Resumen global ────────────────────────────────────────────────────────
    story.append(Paragraph(
        f'<b>Total general deuda: {_pesos(gran_total_deuda)}</b>  &nbsp;·&nbsp;  '
        f'{len(clientes_data)} clientes con saldo pendiente',
        ParagraphStyle('rdResumen', fontSize=9, fontName='Helvetica-Bold',
                       textColor=ROJO, leading=12)
    ))
    story.append(Spacer(1, 5*mm))

    # ── Detalle por cliente ───────────────────────────────────────────────────
    cw_tipo = 14*mm; cw_num = 28*mm; cw_fcomp = 20*mm; cw_fvto = 20*mm
    cw_total = 30*mm; cw_pago = 30*mm; cw_deuda = 30*mm
    cw_cli = usable_w  # encabezado cliente ocupa todo el ancho

    for cli in clientes_data:
        bloque = []

        # Encabezado cliente
        cli_hdr = Table([[
            Paragraph(f'{cli["razonsocial"]}', sCli),
            Paragraph(f'Cód: {cli["codigo"]}  |  CUIT: {cli["cuit"] or "—"}', sN),
            Paragraph(f'Deuda: <b>{_pesos(cli["total_deuda"])}</b>',
                      ParagraphStyle('rdDeu', fontSize=8, fontName='Helvetica-Bold',
                                     textColor=ROJO, alignment=TA_RIGHT)),
        ]], colWidths=[usable_w*0.45, usable_w*0.3, usable_w*0.25])
        cli_hdr.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,-1), colors.HexColor('#e8f0fe')),
            ('TOPPADDING',(0,0),(-1,-1),3), ('BOTTOMPADDING',(0,0),(-1,-1),3),
            ('LEFTPADDING',(0,0),(-1,-1),5), ('RIGHTPADDING',(0,0),(-1,-1),5),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ]))
        bloque.append(cli_hdr)

        # Tabla de comprobantes
        comp_rows = [[
            Paragraph('<b>Tipo</b>',    sNb),
            Paragraph('<b>Número</b>',  sNb),
            Paragraph('<b>F.Comp.</b>', sNb),
            Paragraph('<b>F.Vto.</b>',  sNb),
            Paragraph('<b>Total</b>',   sNbr),
            Paragraph('<b>Pagado</b>',  sNbr),
            Paragraph('<b>Saldo</b>',   sNbr),
        ]]
        for r in cli['comprobantes']:
            deuda = float(r[5] or 0)
            if deuda <= 0:
                continue
            comp_rows.append([
                Paragraph(str(r[0] or ''),          sNc),
                Paragraph(_fmt_num(r[1]),            sNc),
                Paragraph(_fmt_date(r[2]),           sNc),
                Paragraph(_fmt_date(r[6]),           sNc),
                Paragraph(_pesos(float(r[3] or 0)), sNr),
                Paragraph(_pesos(float(r[4] or 0)), sNr),
                Paragraph(_pesos(deuda),             sNr),
            ])

        comp_tbl = Table(comp_rows,
                         colWidths=[cw_tipo, cw_num, cw_fcomp, cw_fvto, cw_total, cw_pago, cw_deuda],
                         repeatRows=1)
        comp_tbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0), CELESTE), ('TEXTCOLOR',(0,0),(-1,0), colors.white),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, GRIS]),
            ('BOX',(0,0),(-1,-1),0.4,colors.grey),
            ('INNERGRID',(0,0),(-1,-1),0.2,colors.lightgrey),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),2), ('BOTTOMPADDING',(0,0),(-1,-1),2),
            ('LEFTPADDING',(0,0),(-1,-1),3),
        ]))
        bloque.append(comp_tbl)
        bloque.append(Spacer(1, 4*mm))

        story.append(KeepTogether(bloque))

    # ── Pie resumen ───────────────────────────────────────────────────────────
    story.append(HRFlowable(width=usable_w, thickness=1, color=CELESTE, spaceAfter=3))
    story.append(Paragraph(
        f'<b>TOTAL GENERAL DEUDA: {_pesos(gran_total_deuda)}</b>',
        ParagraphStyle('rdTot', fontSize=11, fontName='Helvetica-Bold',
                       textColor=ROJO, alignment=TA_RIGHT)
    ))

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    buf.seek(0)
    return StreamingResponse(buf, media_type='application/pdf',
        headers={'Content-Disposition': f'inline; filename="resumen_deudas_{hoy}.pdf"'})

@app.get("/clientes/{codigo}/cuenta_corriente/pdf")
def cuenta_corriente_pdf(codigo: str, limit: int = Query(500, le=2000), offset: int = 0):
    from datetime import datetime

    # ── 1. Datos cliente
    c_cli = conn('WIN1252')
    cur_cli = c_cli.cursor()
    cur_cli.execute(
        'SELECT CODIGOCLIENTE, CODIGOPARTICULAR, RAZONSOCIAL, CUIT, TELEFONO, TELEFONOCELULAR, DIRECCION, LOCALIDAD '
        'FROM "CLIENTES" WHERE CODIGOCLIENTE = ? OR CODIGOPARTICULAR = ?',
        (codigo, codigo)
    )
    cli = cur_cli.fetchone()
    c_cli.close()
    if not cli:
        raise HTTPException(404, "Cliente no encontrado")
    cli_cod_visible = (cli[1] or cli[0] or '').strip()
    cli_razon = (cli[2] or '').strip()
    cli_cuit  = (cli[3] or '').strip()
    cli_tel   = (cli[4] or cli[5] or '').strip()
    cli_dir   = (cli[6] or '').strip()
    cli_loc   = (cli[7] or '').strip()

    # ── 2. Movimientos (misma logica que /cuenta_corriente)
    codigos = set()
    if cli[0] and cli[0].strip(): codigos.add(cli[0].strip())
    if cli[1] and cli[1].strip(): codigos.add(cli[1].strip())
    if not codigos: codigos.add(codigo)
    codigos = list(codigos)

    DB_PROD     = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    rows_prod = _query_cta(DB_PROD,      codigos, limit, offset)
    rows_mp   = _query_cta(DB_MLT_PROD,  codigos, limit, offset)

    seen, combined = set(), []
    for r in rows_prod + rows_mp:
        key = (r[0], r[1])
        if key not in seen:
            seen.add(key)
            combined.append(r)
    combined.sort(key=lambda r: (r[6] or r[2], r[2]))

    # ── 3. Datos empresa
    razon_soc = cuit_emp = dir_emp = tel_emp = email_emp = ''
    try:
        cp = conn(); ccp = cp.cursor()
        ccp.execute('SELECT RAZONSOCIAL, CUIT, DIRECCION, TELEFONO, EMAIL FROM "PARAMETROS" WHERE CODIGOPARAMETRO = 1')
        rp = ccp.fetchone(); cp.close()
        if rp:
            razon_soc, cuit_emp, dir_emp, tel_emp, email_emp = [(v or '').strip() for v in rp]
    except Exception:
        pass

    # ── 4. PDF
    buf = BytesIO()
    PAGE_W, PAGE_H = A4
    mg = 14 * mm

    s_title = ParagraphStyle('t', fontSize=13, leading=16, fontName='Helvetica-Bold',
                              textColor=colors.HexColor('#1a56db'))
    s_sub   = ParagraphStyle('s', fontSize=8,  leading=10, fontName='Helvetica',
                              textColor=colors.HexColor('#6b7280'))
    s_label = ParagraphStyle('l', fontSize=8,  leading=10, fontName='Helvetica-Bold',
                              textColor=colors.HexColor('#374151'))
    s_val   = ParagraphStyle('v', fontSize=8,  leading=10, fontName='Helvetica')
    s_hdr   = ParagraphStyle('h', fontSize=7.5, leading=9, fontName='Helvetica-Bold',
                              alignment=TA_CENTER, textColor=colors.white)
    s_cell  = ParagraphStyle('c', fontSize=7.5, leading=9, fontName='Helvetica')
    s_cell_r= ParagraphStyle('cr', fontSize=7.5, leading=9, fontName='Helvetica', alignment=TA_RIGHT)
    s_total = ParagraphStyle('tt', fontSize=10, leading=12, fontName='Helvetica-Bold',
                              alignment=TA_RIGHT, textColor=colors.HexColor('#dc2626'))

    usable_w = PAGE_W - 2 * mg
    footer_txt = f'{razon_soc}  ·  CUIT {cuit_emp}  ·  {dir_emp}  ·  Tel {tel_emp}  ·  {email_emp}'

    def _on_page(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor('#e5e7eb'))
        canvas.setLineWidth(0.5)
        canvas.line(mg, 10*mm, PAGE_W - mg, 10*mm)
        canvas.setFont('Helvetica', 6.5)
        canvas.setFillColor(colors.HexColor('#6b7280'))
        canvas.drawCentredString(PAGE_W / 2, 7*mm, footer_txt)
        canvas.drawRightString(PAGE_W - mg, 7*mm, f"Pág. {doc.page}")
        canvas.restoreState()

    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=mg, rightMargin=mg,
                            topMargin=mg, bottomMargin=18*mm)
    story = []

    # Header logo + titulo
    logo_cell = Image(LOGO_PATH, width=38*mm, height=12*mm) if os.path.exists(LOGO_PATH) else Paragraph('', s_sub)
    titulo_cell = [
        Paragraph("Resumen de Cuenta Corriente", s_title),
        Paragraph(datetime.now().strftime('%d/%m/%Y %H:%M'), s_sub),
    ]
    t_hdr = Table([[logo_cell, titulo_cell]], colWidths=[42*mm, usable_w - 42*mm])
    t_hdr.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN',  (1,0), (1,0),  'RIGHT'),
    ]))
    story.append(t_hdr)
    story.append(Spacer(1, 4*mm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#1a56db')))
    story.append(Spacer(1, 3*mm))

    # Datos cliente
    cli_data = [
        [Paragraph('Cliente:',   s_label), Paragraph(cli_razon, s_val),
         Paragraph('Código:',    s_label), Paragraph(cli_cod_visible, s_val)],
        [Paragraph('CUIT:',      s_label), Paragraph(cli_cuit or '—', s_val),
         Paragraph('Teléfono:',  s_label), Paragraph(cli_tel or '—', s_val)],
        [Paragraph('Dirección:', s_label), Paragraph(cli_dir or '—', s_val),
         Paragraph('Localidad:', s_label), Paragraph(cli_loc or '—', s_val)],
    ]
    t_cli = Table(cli_data, colWidths=[22*mm, (usable_w/2 - 22*mm), 22*mm, (usable_w/2 - 22*mm)])
    t_cli.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOX',    (0,0), (-1,-1), 0.5, colors.HexColor('#e5e7eb')),
        ('INNERGRID', (0,0), (-1,-1), 0.3, colors.HexColor('#f3f4f6')),
        ('TOPPADDING',    (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING',   (0,0), (-1,-1), 6),
        ('RIGHTPADDING',  (0,0), (-1,-1), 6),
    ]))
    story.append(t_cli)
    story.append(Spacer(1, 4*mm))

    # Tabla movimientos
    HDR_BG = colors.HexColor('#1a56db')
    ALT_BG = colors.HexColor('#eff6ff')
    cw = [16*mm, 32*mm, 24*mm, 24*mm, 28*mm, 28*mm, 30*mm]

    data = [[
        Paragraph('Tipo',     s_hdr),
        Paragraph('Número',   s_hdr),
        Paragraph('F. Comp.', s_hdr),
        Paragraph('F. Vto.',  s_hdr),
        Paragraph('Total',    s_hdr),
        Paragraph('Pagado',   s_hdr),
        Paragraph('Deuda',    s_hdr),
    ]]

    def _fmt_money(n):
        return '$' + f"{float(n or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    def _fmt_fecha(f):
        if not f: return '—'
        s = str(f)
        return s[:10] if len(s) >= 10 else s

    total_total = total_pagado = total_deuda = 0.0
    hoy = datetime.now().date()

    for r in combined:
        tot = float(r[3] or 0); pag = float(r[4] or 0); deu = float(r[5] or 0)
        total_total += tot; total_pagado += pag; total_deuda += deu
        vto_str = _fmt_fecha(r[6])
        # Marcar vencido en rojo
        try:
            vto_d = r[6].date() if hasattr(r[6], 'date') else None
            vencido = bool(vto_d and vto_d < hoy and deu > 0.01)
        except Exception:
            vencido = False
        s_vto = ParagraphStyle('vt', parent=s_cell, textColor=colors.HexColor('#dc2626')) if vencido else s_cell

        data.append([
            Paragraph(str(r[0] or ''),    s_cell),
            Paragraph(str(r[1] or ''),    s_cell),
            Paragraph(_fmt_fecha(r[2]),   s_cell),
            Paragraph(vto_str + (' ⚠' if vencido else ''), s_vto),
            Paragraph(_fmt_money(tot),    s_cell_r),
            Paragraph(_fmt_money(pag),    s_cell_r),
            Paragraph(_fmt_money(deu),
                      ParagraphStyle('dr', parent=s_cell_r,
                                     textColor=colors.HexColor('#dc2626'),
                                     fontName='Helvetica-Bold')),
        ])

    tbl = Table(data, colWidths=cw, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND',     (0,0), (-1,0), HDR_BG),
        ('TEXTCOLOR',      (0,0), (-1,0), colors.white),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, ALT_BG]),
        ('GRID',           (0,0), (-1,-1), 0.4, colors.HexColor('#e5e7eb')),
        ('VALIGN',         (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',     (0,0), (-1,-1), 3),
        ('BOTTOMPADDING',  (0,0), (-1,-1), 3),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 4*mm))

    # Totales
    tot_data = [[
        Paragraph(f'{len(combined)} comprobantes', s_sub),
        Paragraph(f'Total: {_fmt_money(total_total)}    Pagado: {_fmt_money(total_pagado)}', s_sub),
        Paragraph(f'Deuda: {_fmt_money(total_deuda)}', s_total),
    ]]
    t_tot = Table(tot_data, colWidths=[40*mm, usable_w - 40*mm - 60*mm, 60*mm])
    t_tot.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN',  (2,0), (2,0),  'RIGHT'),
    ]))
    story.append(t_tot)

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    buf.seek(0)
    fname = f"CtaCte_{cli_cod_visible}.pdf"
    return StreamingResponse(buf, media_type='application/pdf',
        headers={"Content-Disposition": f"inline; filename={fname}"})


# ─── Multiplazos / Transportes / Sucursales ───────────────────────────────────
@app.get("/multiplazos")
def get_multiplazos():
    c = conn()
    cur = c.cursor()
    cur.execute('SELECT CODIGOMULTIPLAZO, DESCRIPCION FROM "MULTIPLAZOS" WHERE ACTIVO = ? ORDER BY DESCRIPCION', ('1',))
    rows = cur.fetchall()
    c.close()
    return [{"codigo": r[0], "descripcion": (r[1] or "").strip()} for r in rows]

@app.get("/transportes")
def get_transportes():
    c = conn()
    cur = c.cursor()
    cur.execute('SELECT FIRST 300 CODIGOTRANSPORTE, DESCRIPCION FROM "TRANSPORTES" ORDER BY DESCRIPCION')
    rows = cur.fetchall()
    c.close()
    return [{"codigo": r[0], "descripcion": (r[1] or "").strip()} for r in rows]

@app.get("/clientes/{codigo}/sucursales")
def get_sucursales_cliente(codigo: str):
    """
    Devuelve domicilios de entrega (SUCURSALESXCLIENTES) desde DB-Microbell.gdb.
    Resolución de código:
      1. Busca CODIGOPARTICULAR en DB-Prueba (fuente del codigo recibido)
      2. Usa CODIGOPARTICULAR para encontrar CODIGOCLIENTE en DB-Microbell
      3. Consulta SUCURSALESXCLIENTES con ese CODIGOCLIENTE
    """
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    sucursales = []
    direccion_principal = ''
    transp_cli  = None
    transp_fijo = False

    # ── Paso 1: resolver CODIGOPARTICULAR desde DB-Prueba ───────────────────
    codigoparticular = codigo   # fallback: usar el código tal cual
    try:
        c_pru = conn('WIN1252')   # DB-Prueba.gdb
        cur_pru = c_pru.cursor()
        cur_pru.execute(
            'SELECT CODIGOPARTICULAR, DIRECCION, LOCALIDAD FROM "CLIENTES" '
            'WHERE CODIGOCLIENTE = ? OR CODIGOPARTICULAR = ?',
            (codigo, codigo)
        )
        row_pru = cur_pru.fetchone()
        c_pru.close()
        if row_pru:
            cp = (row_pru[0] or '').strip()
            if cp:
                codigoparticular = cp
            # Dirección de fallback desde DB-Prueba
            partes_dir = [p for p in [(row_pru[1] or '').strip(),
                                      (row_pru[2] or '').strip()] if p]
            if partes_dir:
                direccion_principal = ' — '.join(partes_dir)
    except Exception:
        pass

    # ── Paso 2 y 3: buscar CODIGOCLIENTE en Microbell y traer sucursales ────
    c_prod = None
    try:
        c_prod = conn('WIN1252', db=DB_PROD)
        cur_prod = c_prod.cursor()

        # Resolver CODIGOCLIENTE interno en DB-Microbell via CODIGOPARTICULAR
        cur_prod.execute(
            'SELECT CODIGOCLIENTE, DIRECCION, LOCALIDAD FROM "CLIENTES" '
            'WHERE CODIGOPARTICULAR = ? OR CODIGOCLIENTE = ?',
            (codigoparticular, codigoparticular)
        )
        row_mb = cur_prod.fetchone()
        if row_mb:
            cod_mb = str(row_mb[0] or '').strip()
            partes_mb = [p for p in [str(row_mb[1] or '').strip(),
                                     str(row_mb[2] or '').strip()] if p]
            if partes_mb:
                direccion_principal = ' — '.join(partes_mb)
        else:
            cod_mb = codigoparticular

        # Transporte a nivel cliente (TRANSPORTEFIJO vive en CLIENTES, no en SUCURSALESXCLIENTES)
        cur_prod.execute(
            'SELECT CODIGOTRANSPORTE, TRANSPORTEFIJO, REPARTOPROPIO '
            'FROM "CLIENTES" WHERE CODIGOCLIENTE = ?', (cod_mb,)
        )
        row_transp = cur_prod.fetchone()
        transp_cli  = str(row_transp[0]).strip() if row_transp and row_transp[0] else None
        transp_fijo = str(row_transp[1] or '0').strip() == '1' if row_transp else False

        # Domicilios de entrega (sin TRANSPORTEFIJO — no existe en esta tabla)
        cur_prod.execute(
            'SELECT s.CODIGOSUCURSAL, s.NOMBRE, s.DIRECCION, s.CODIGOTRANSPORTE, '
            'p.NOMBRE, l.NOMBRE, s.TELEFONO, s.OBSERVACIONES '
            'FROM "SUCURSALESXCLIENTES" s '
            'LEFT JOIN "PROVINCIAS" p ON p.CODIGOPROVINCIA = s.CODIGOPROVINCIA '
            'LEFT JOIN "LOCALIDADES" l '
            '       ON l.CODIGOPROVINCIA = s.CODIGOPROVINCIA '
            '      AND l.CODIGOLOCALIDAD = s.CODIGOLOCALIDAD '
            'WHERE s.CODIGOCLIENTE = ? '
            'ORDER BY s.CODIGOSUCURSAL',
            (cod_mb,)
        )
        for r in cur_prod.fetchall():
            transp_suc = str(r[3]).strip() if r[3] is not None and r[3] != '' else None
            sucursales.append({
                "codigo":        str(r[0] or "").strip(),
                "nombre":        str(r[1] or "").strip(),
                "direccion":     str(r[2] or "").strip(),
                "transporte":    transp_suc or transp_cli,
                "provincia":     str(r[4] or "").strip(),
                "localidad":     str(r[5] or "").strip(),
                "telefono":      str(r[6] or "").strip(),
                "observaciones": str(r[7] or "").strip(),
                "transporteFijo": transp_fijo,
            })
    except Exception as _e_suc:
        sucursales = []
        _suc_error = str(_e_suc)
    else:
        _suc_error = None
    finally:
        if c_prod:
            try: c_prod.close()
            except Exception: pass

    resp = {
        "sucursales": sucursales,
        "direccion_principal": direccion_principal,
        "transporte_codigo": transp_cli,
        "transporte_fijo": transp_fijo,
    }
    if _suc_error:
        resp["_error"] = _suc_error   # solo para diagnóstico; se puede quitar luego
    return resp

# ─── Pedidos (solo del vendedor) ───────────────────────────────────────────────
@app.get("/pedidos")
def get_pedidos(
    vendedor: str,
    cliente: Optional[str] = None,
    limit: int = Query(50, le=200),
    offset: int = 0,
    db: str = Query("oficial")
):
    # SW y L1 usan el mismo DATABASE (CABEZAPEDIDOS)
    # Para SW se filtra además por PRIORIDAD = '2'
    c = conn('WIN1252', db=DATABASE)
    cur = c.cursor()
    params = [vendedor.upper()]
    where_cli = ""
    if cliente:
        where_cli = "AND CODIGOCLIENTE = ?"
        params.append(cliente)
    where_prioridad = "AND PRIORIDAD = '2'" if db == 'sw' else "AND (PRIORIDAD IS NULL OR PRIORIDAD = '1')"

    cur.execute(f"""
        SELECT FIRST {limit} SKIP {offset}
            NUMEROCOMPROBANTE, CODIGOCLIENTE, RAZONSOCIAL,
            FECHACOMPROBANTE, TOTAL, ANULADA, COMENTARIOS, FECHAENTREGA,
            OPERACION, FECHATERMINADA, CODIGOUSUARIO2
        FROM "CABEZAPEDIDOS"
        WHERE TIPOCOMPROBANTE = 'NP' AND CODIGOUSUARIO = ? AND ANULADA = '0'
            {where_cli} {where_prioridad}
        ORDER BY FECHACOMPROBANTE DESC, NUMEROCOMPROBANTE DESC
    """, params)
    rows = cur.fetchall()
    c.close()
    return [{
        "numero": r[0], "cod_cliente": r[1], "razonsocial": r[2],
        "fecha":  r[3], "total": r[4], "anulada": r[5],
        "comentarios": r[6], "fecha_entrega": r[7],
        "operacion": str(r[8]).strip() if r[8] is not None else "1",
        "fecha_terminada": str(r[9]) if r[9] else None,
        "responsable": str(r[10]).strip() if r[10] else None
    } for r in rows]

@app.post("/pedidos/{numero}/terminar")
def terminar_pedido(numero: str, codigousuario: str = Query(...)):
    """Marca el pedido como terminado: FECHATERMINADA = ahora, CODIGOUSUARIO2 = responsable."""
    from datetime import datetime
    ahora = datetime.now()
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        cur.execute("""
            UPDATE "CABEZAPEDIDOS"
            SET FECHATERMINADA = ?, CODIGOUSUARIO2 = ?, FECHAMODIFICACION = ?
            WHERE TIPOCOMPROBANTE = 'NP' AND NUMEROCOMPROBANTE = ?
        """, (ahora, codigousuario.upper(), ahora, numero))
        c.commit()
        c.close()
        return {"ok": True, "numero": numero, "fechaterminada": str(ahora), "responsable": codigousuario.upper()}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/pedidos/{numero}/detalle")
def get_pedido_detalle(numero: str, db: str = Query("oficial")):
    # SW y L1 usan el mismo DATABASE (CABEZAPEDIDOS/CUERPOPEDIDOS)
    c = conn('WIN1252', db=DATABASE)
    cur = c.cursor()
    cur.execute(
        "SELECT cp.LINEA, "
        "COALESCE(NULLIF(TRIM(cp.CODIGOPARTICULAR),''), NULLIF(TRIM(a.CODIGOPARTICULAR),''), TRIM(cp.CODIGOARTICULO)), "
        "cp.DESCRIPCION, cp.CANTIDAD, "
        "cp.DESCUENTO, cp.PRECIOUNITARIO, cp.PRECIOTOTAL, cp.PORCENTAJEIVA "
        "FROM \"CUERPOPEDIDOS\" cp "
        "LEFT JOIN \"ARTICULOS\" a ON a.CODIGOARTICULO = cp.CODIGOARTICULO "
        "WHERE cp.TIPOCOMPROBANTE = 'NP' AND cp.NUMEROCOMPROBANTE = ? ORDER BY cp.LINEA",
        (numero,)
    )
    rows = cur.fetchall()
    c.close()
    return [{
        "linea": r[0], "codigo": r[1], "descripcion": r[2],
        "cantidad": r[3], "descuento": r[4],
        "precio_unitario": r[5], "precio_total": r[6], "iva": r[7]
    } for r in rows]

# ─── PDF Nota de Pedido ────────────────────────────────────────────────────────
@app.get("/pedidos/{numero}/pdf")
def pedido_pdf(numero: str, db: str = Query("oficial")):
    db_path = DATABASE  # SW y L1 usan el mismo DATABASE (CABEZAPEDIDOS)

    # ── 1. Datos empresa ───────────────────────────────────────────────────────
    razon_soc = 'MICROBELL S.A.'
    dir_emp   = 'AV. MONROE 5088 PISO 2'
    tel_emp   = '+54 11 3988-0024'
    email_emp = 'info@microbellsa.com.ar'
    web_emp   = 'www.microbellsa.com'
    cuit_emp  = '30-70839018-2'
    try:
        c_e = conn('WIN1252', db=db_path)
        cur_e = c_e.cursor()
        cur_e.execute(
            'SELECT RAZONSOCIAL, DIRECCION, TELEFONO, EMAIL, DIRECCIONWEB, CUIT '
            'FROM "SUCURSALES" WHERE CODIGOSUCURSAL = ?', ('PRINCIPAL',)
        )
        emp = cur_e.fetchone()
        c_e.close()
        if emp:
            razon_soc = (emp[0] or razon_soc).strip()
            dir_emp   = (emp[1] or dir_emp).strip()
            tel_emp   = (emp[2] or tel_emp).strip()
            email_emp = (emp[3] or email_emp).strip()
            web_emp   = (emp[4] or web_emp).strip()
            cuit_emp  = (emp[5] or cuit_emp).strip()
    except Exception:
        pass

    # ── 2. Cabeza pedido ───────────────────────────────────────────────────────
    # Tanto SW como L1 usan CABEZAPEDIDOS en DATABASE
    cab_table  = '"CABEZAPEDIDOS"'
    item_table = '"CUERPOPEDIDOS"'
    extra_where = " AND TIPOCOMPROBANTE = 'NP'"
    try:
        c_h = conn('WIN1252', db=db_path)
        cur_h = c_h.cursor()
        cur_h.execute(
            f'SELECT CODIGOCLIENTE, RAZONSOCIAL, FECHACOMPROBANTE, FECHAENTREGA, '
            f'TOTAL, IVA1, COMENTARIOS, CODIGOUSUARIO, CODIGOMULTIPLAZO, '
            f'CODIGOTRANSPORTE, DIRECCION, TIPOIVA, TELEFONO '
            f'FROM {cab_table} WHERE NUMEROCOMPROBANTE = ?{extra_where}',
            (numero,)
        )
        cab = cur_h.fetchone()
        c_h.close()
    except Exception as ex:
        raise HTTPException(500, f"Cabeza: {ex}")
    if not cab:
        raise HTTPException(404, f"Pedido {numero} no encontrado")

    cod_cli, rs_cli, fec_comp, fec_entrega, total_cab, iva1_cab, \
    comentarios, cod_usu, cod_multi, cod_transp, dir_cli, tipo_iva, tel_cli = cab

    subtotal_cab = float(total_cab or 0)
    iva1_val     = float(iva1_cab or 0)
    total_final  = subtotal_cab + iva1_val

    # ── 3. Items ───────────────────────────────────────────────────────────────
    try:
        c_it = conn('WIN1252', db=db_path)
        cur_it = c_it.cursor()
        if db == 'sw':
            # DATABASE_MLT no tiene ARTICULOS — usar solo CODIGOPARTICULAR del cuerpo
            cur_it.execute(
                f'SELECT COALESCE(NULLIF(TRIM(CODIGOPARTICULAR),\'\'), TRIM(CODIGOARTICULO)), '
                f'DESCRIPCION, CANTIDAD, PRECIOUNITARIO, '
                f'DESCUENTO, PRECIOTOTAL, PORCENTAJEIVA '
                f'FROM {item_table} '
                f'WHERE NUMEROCOMPROBANTE = ?{extra_where} ORDER BY LINEA',
                (numero,)
            )
        else:
            cur_it.execute(
                f'SELECT COALESCE(NULLIF(TRIM(it.CODIGOPARTICULAR),\'\'), NULLIF(TRIM(a.CODIGOPARTICULAR),\'\'), TRIM(it.CODIGOARTICULO)), '
                f'it.DESCRIPCION, it.CANTIDAD, it.PRECIOUNITARIO, '
                f'it.DESCUENTO, it.PRECIOTOTAL, it.PORCENTAJEIVA '
                f'FROM {item_table} it '
                f'LEFT JOIN "ARTICULOS" a ON a.CODIGOARTICULO = it.CODIGOARTICULO '
                f'WHERE it.NUMEROCOMPROBANTE = ?{extra_where.replace("AND TIPOCOMPROBANTE", "AND it.TIPOCOMPROBANTE")} ORDER BY it.LINEA',
                (numero,)
            )
        items = cur_it.fetchall()
        c_it.close()
    except Exception as ex:
        raise HTTPException(500, f"Items: {ex}")

    # ── 4. Datos cliente ───────────────────────────────────────────────────────
    cuit_cli = ''
    tel_pdf  = str(tel_cli or '').strip().lstrip('-').strip().split()[0] if str(tel_cli or '').strip().lstrip('-').strip() else ''
    vendedor_nombre = cod_usu or ''
    try:
        c_cl = conn('WIN1252', db=db_path)
        cur_cl = c_cl.cursor()
        cur_cl.execute(
            'SELECT CUIT, TELEFONO, TELEFONOCELULAR FROM "CLIENTES" WHERE CODIGOCLIENTE = ?',
            (cod_cli,)
        )
        r_cl = cur_cl.fetchone()
        if r_cl:
            cuit_cli = (r_cl[0] or '').strip()
            if not tel_pdf:
                def _tp(v):
                    s = str(v or '').strip().lstrip('-').strip()
                    try: s = str(int(float(s))) if s and ('E' in s.upper() or '.' in s) else s
                    except: pass
                    return s
                tel_pdf = _tp(r_cl[1]) or _tp(r_cl[2])
        cur_cl.execute(
            'SELECT NOMBRE, APELLIDO FROM "USUARIOS" WHERE CODIGOUSUARIO = ?',
            (cod_usu,)
        )
        r_usu = cur_cl.fetchone()
        if r_usu:
            vendedor_nombre = f"{(r_usu[0] or '').strip()} {(r_usu[1] or '').strip()}".strip() or cod_usu
        c_cl.close()
    except Exception:
        pass

    # ── 5. Transporte descripción ──────────────────────────────────────────────
    transporte_desc = 'A CONVENIR'
    try:
        if cod_transp and str(cod_transp).strip() not in ('', '0'):
            # Buscar transporte en DATABASE principal (MLT no tiene la tabla completa)
            for _tr_db in ([db_path, DATABASE] if db == 'sw' else [db_path]):
                try:
                    c_tr = conn('WIN1252', db=_tr_db)
                    cur_tr = c_tr.cursor()
                    cur_tr.execute('SELECT DESCRIPCION FROM "TRANSPORTES" WHERE CODIGOTRANSPORTE = ?', (cod_transp,))
                    r_tr = cur_tr.fetchone()
                    c_tr.close()
                    if r_tr and (r_tr[0] or '').strip():
                        transporte_desc = r_tr[0].strip()
                        break
                except Exception:
                    pass
    except Exception:
        pass

    # ── 6. Condición de venta ──────────────────────────────────────────────────
    cond_venta = ''
    try:
        if cod_multi:
            c_mp = conn('WIN1252', db=db_path)
            cur_mp = c_mp.cursor()
            cur_mp.execute('SELECT DESCRIPCION FROM "MULTIPLAZOS" WHERE CODIGOMULTIPLAZO = ?', (cod_multi,))
            r_mp = cur_mp.fetchone()
            if r_mp: cond_venta = (r_mp[0] or '').strip()
            c_mp.close()
    except Exception:
        pass

    # ── 7. Depósito ────────────────────────────────────────────────────────────
    deposito_desc = ''
    try:
        c_dep = conn('WIN1252', db=db_path)
        cur_dep = c_dep.cursor()
        cur_dep.execute(
            f'SELECT FIRST 1 CODIGODEPOSITO FROM {item_table} WHERE NUMEROCOMPROBANTE = ?{extra_where}',
            (numero,)
        )
        r_dep = cur_dep.fetchone()
        if r_dep:
            cod_dep = str(r_dep[0] or '').strip()
            cur_dep.execute('SELECT NOMBRE FROM "DEPOSITOS" WHERE CODIGODEPOSITO = ?', (cod_dep,))
            r_dn = cur_dep.fetchone()
            deposito_desc = (r_dn[0] or cod_dep).strip() if r_dn else cod_dep
        c_dep.close()
    except Exception:
        pass

    # ── 8. Construcción PDF ────────────────────────────────────────────────────
    buf  = BytesIO()
    PAGE_W, PAGE_H = A4
    mg   = 14 * mm
    BOTTOM_BLOCK = 48 * mm   # espacio reservado al pie para totales + footer

    s_norm  = ParagraphStyle('pn',  fontSize=8,  leading=11, fontName='Helvetica')
    s_sm    = ParagraphStyle('psm', fontSize=7,  leading=9,  fontName='Helvetica')
    s_bold  = ParagraphStyle('pb',  fontSize=8,  leading=11, fontName='Helvetica-Bold')
    s_h2    = ParagraphStyle('ph2', fontSize=10, leading=13, fontName='Helvetica-Bold')
    s_c     = ParagraphStyle('pc',  fontSize=7.5,leading=10, fontName='Helvetica',      alignment=TA_CENTER)
    s_c_b   = ParagraphStyle('pcb', fontSize=7.5,leading=10, fontName='Helvetica-Bold', alignment=TA_CENTER)
    s_r     = ParagraphStyle('pr',  fontSize=8,  leading=11, fontName='Helvetica',      alignment=TA_RIGHT)
    s_r_b   = ParagraphStyle('prb', fontSize=9,  leading=12, fontName='Helvetica-Bold', alignment=TA_RIGHT)
    s_label = ParagraphStyle('pl',  fontSize=8,  leading=11, fontName='Helvetica-Bold', textColor=colors.HexColor('#374151'))
    s_val   = ParagraphStyle('pv',  fontSize=8,  leading=11, fontName='Helvetica',      textColor=colors.HexColor('#111827'))

    usable_w = PAGE_W - 2 * mg
    num_fmt  = f"0001-{int(numero):08d}"
    footer_txt = f'{razon_soc}  ·  CUIT {cuit_emp}  ·  {dir_emp}  ·  Tel {tel_emp}  ·  {email_emp}'
    s_ft = ParagraphStyle('ft', fontSize=6.5, leading=9, fontName='Helvetica',
                           alignment=TA_CENTER, textColor=colors.HexColor('#6b7280'))

    # ── Canvas callbacks: totales + firma al pie ───────────────────────────────
    def _draw_bottom(canvas, doc):
        canvas.saveState()
        x0    = mg
        right = PAGE_W - mg
        tot_y = BOTTOM_BLOCK - 8*mm        # Y desde la base de la hoja

        # Línea separadora
        canvas.setStrokeColor(colors.HexColor('#e5e7eb'))
        canvas.setLineWidth(0.5)
        canvas.line(x0, tot_y + 28*mm, right, tot_y + 28*mm)

        # Observaciones (si existen)
        obs_y = tot_y + 30*mm
        obs = str(comentarios or '').strip()
        if obs:
            canvas.setFont('Helvetica-Bold', 8)
            canvas.setFillColor(colors.HexColor('#374151'))
            canvas.drawString(x0, obs_y, 'Observaciones:')
            canvas.setFont('Helvetica', 8)
            canvas.drawString(x0 + 28*mm, obs_y, obs[:120])

        # Bloque totales (derecha)
        tw = 50 * mm
        tx = right - tw
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#374151'))
        canvas.drawRightString(tx - 2, tot_y + 20*mm, 'Subtotal:')
        canvas.setFont('Helvetica', 8)
        canvas.drawRightString(right,  tot_y + 20*mm, _fmt(subtotal_cab))
        canvas.drawRightString(tx - 2, tot_y + 12*mm, 'IVA 21%:')
        canvas.drawRightString(right,  tot_y + 12*mm, _fmt(iva1_val))
        # Línea sobre total
        canvas.setLineWidth(0.8)
        canvas.setStrokeColor(colors.black)
        canvas.line(tx - 5*mm, tot_y + 10*mm, right, tot_y + 10*mm)
        canvas.setFont('Helvetica-Bold', 10)
        canvas.setFillColor(colors.black)
        canvas.drawRightString(tx - 2, tot_y + 3*mm, 'TOTAL:')
        canvas.drawRightString(right,  tot_y + 3*mm, _fmt(total_final))

        # Líneas de firma (izquierda)
        sig_y = tot_y + 5*mm
        canvas.setLineWidth(0.4)
        canvas.setStrokeColor(colors.HexColor('#9ca3af'))
        canvas.line(x0, sig_y, x0 + 60*mm, sig_y)
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(colors.HexColor('#6b7280'))
        canvas.drawString(x0, sig_y - 4*mm, 'Firma y aclaración')

        # Footer empresa
        canvas.setFont('Helvetica', 6.5)
        canvas.setFillColor(colors.HexColor('#9ca3af'))
        canvas.drawCentredString(PAGE_W / 2, 8*mm, footer_txt)
        canvas.restoreState()

    doc_obj = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=mg, rightMargin=mg,
        topMargin=5*mm, bottomMargin=BOTTOM_BLOCK
    )

    story = []

    # ── Header ────────────────────────────────────────────────────────────────
    logo_w = 38 * mm
    if os.path.exists(LOGO_PATH):
        logo_cell = [Image(LOGO_PATH, width=logo_w, height=logo_w * 0.32)]
    else:
        logo_cell = [Paragraph(f'<b>{razon_soc}</b>', s_h2)]

    doc_w  = 52 * mm
    emp_w  = usable_w - logo_w - doc_w - 2*mm

    doc_cell = [
        Paragraph('<b>NOTA DE PEDIDO</b>',
                  ParagraphStyle('dt', fontSize=13, fontName='Helvetica-Bold', alignment=TA_CENTER)),
        Paragraph(f'<b>N°  {num_fmt}</b>',
                  ParagraphStyle('dn', fontSize=11, fontName='Helvetica-Bold', alignment=TA_CENTER)),
        Spacer(1, 3*mm),
        Paragraph(f'Fecha: <b>{_d(fec_comp)}</b>',
                  ParagraphStyle('df', fontSize=8.5, fontName='Helvetica', alignment=TA_CENTER)),
        Paragraph(f'Fecha Entrega: <b>{_d(fec_entrega)}</b>',
                  ParagraphStyle('de', fontSize=8.5, fontName='Helvetica', alignment=TA_CENTER)),
    ]
    emp_cell = [
        Paragraph(f'<b>{razon_soc}</b>', s_bold),
        Paragraph(dir_emp,    s_norm),
        Paragraph(f'Tel: {tel_emp}',  s_norm),
        Paragraph(f'CUIT: {cuit_emp}', s_norm),
        Paragraph(email_emp,  s_norm),
        Paragraph(web_emp,    s_norm),
    ]

    hdr_tbl = Table([[logo_cell, doc_cell, emp_cell]], colWidths=[logo_w, doc_w, emp_w])
    hdr_tbl.setStyle(TableStyle([
        ('VALIGN',       (0,0),(-1,-1), 'MIDDLE'),
        ('ALIGN',        (1,0),(1,0),   'CENTER'),
        ('ALIGN',        (2,0),(2,0),   'LEFT'),
        ('LEFTPADDING',  (0,0),(-1,-1), 4),
        ('RIGHTPADDING', (0,0),(-1,-1), 4),
        ('TOPPADDING',   (0,0),(-1,-1), 0),
        ('BOTTOMPADDING',(0,0),(-1,-1), 0),
    ]))
    story.append(hdr_tbl)
    story.append(HRFlowable(width='100%', thickness=1.5, color=colors.HexColor('#1a56db'),
                            spaceBefore=4*mm, spaceAfter=4*mm))

    # ── Datos cliente ──────────────────────────────────────────────────────────
    pad = TableStyle([
        ('TOPPADDING',   (0,0),(-1,-1), 3),
        ('BOTTOMPADDING',(0,0),(-1,-1), 3),
        ('LEFTPADDING',  (0,0),(-1,-1), 4),
        ('RIGHTPADDING', (0,0),(-1,-1), 4),
    ])
    lbl_w = 24 * mm
    half  = (usable_w - 5*mm) / 2
    cli_nombre = f"{cod_cli} - {rs_cli}".strip(' -') if cod_cli and rs_cli else (rs_cli or cod_cli or '')
    cli_l = Table([
        [Paragraph('Cliente:',   s_label), Paragraph(cli_nombre,    s_val)],
        [Paragraph('CUIT:',      s_label), Paragraph(cuit_cli or '', s_val)],
        [Paragraph('Dirección:', s_label), Paragraph(dir_cli or '',  s_val)],
        [Paragraph('Teléfono:',  s_label), Paragraph(tel_pdf or '',  s_val)],
    ], colWidths=[lbl_w, half - lbl_w])
    cli_r = Table([
        [Paragraph('Cond. Venta:', s_label), Paragraph(cond_venta or '',      s_val)],
        [Paragraph('Transporte:',  s_label), Paragraph(transporte_desc or '', s_val)],
        [Paragraph('Depósito:',    s_label), Paragraph(deposito_desc or '',   s_val)],
        [Paragraph('Vendedor:',    s_label), Paragraph(vendedor_nombre or '', s_val)],
    ], colWidths=[26*mm, half - 26*mm])
    for t in (cli_l, cli_r):
        t.setStyle(pad)

    cli_outer = Table([[cli_l, Spacer(5*mm,1), cli_r]], colWidths=[half, 5*mm, half])
    cli_outer.setStyle(TableStyle([
        ('VALIGN', (0,0),(-1,-1), 'TOP'),
        ('BOX',    (0,0),(0,0),   0.5, colors.HexColor('#d1d5db')),
        ('BOX',    (2,0),(2,0),   0.5, colors.HexColor('#d1d5db')),
        ('BACKGROUND',(0,0),(0,0), colors.HexColor('#f9fafb')),
        ('BACKGROUND',(2,0),(2,0), colors.HexColor('#f9fafb')),
    ]))
    story.append(cli_outer)
    story.append(Spacer(1, 5*mm))

    # ── Tabla artículos ────────────────────────────────────────────────────────
    # col_w suma = usable_w exacto
    col_w = [18*mm, 78*mm, 16*mm, 28*mm, 14*mm, 28*mm]   # = 182mm
    hdr_row = [
        Paragraph('Código',      s_c_b),
        Paragraph('Descripción', s_c_b),
        Paragraph('Cantidad',    s_c_b),
        Paragraph('P. Unitario', s_c_b),
        Paragraph('Dto %',       s_c_b),
        Paragraph('P. Total',    s_c_b),
    ]
    s_ri = ParagraphStyle('ri', fontSize=8, leading=11, fontName='Helvetica', alignment=TA_RIGHT)
    items_data = [hdr_row]
    for it in items:
        cod_art, desc_art, cant, pu, bonif, ptotal, piva = it
        cod_str = str(cod_art or '').strip()
        if cod_str.endswith('.0'):
            cod_str = cod_str[:-2]
        try:
            cant_str = str(int(float(cant))) if float(cant) == int(float(cant)) else str(cant)
        except Exception:
            cant_str = str(cant)
        items_data.append([
            Paragraph(cod_str,                    s_c),
            Paragraph(str(desc_art or '').strip(), s_norm),
            Paragraph(cant_str,                   s_c),
            Paragraph(_fmt(pu),                   s_ri),
            Paragraph(f"{float(bonif or 0):.1f}%", s_c),
            Paragraph(_fmt(ptotal),               s_ri),
        ])

    items_tbl = Table(items_data, colWidths=col_w, repeatRows=1)
    items_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),(-1,0),  colors.HexColor('#1a56db')),
        ('TEXTCOLOR',     (0,0),(-1,0),  colors.white),
        ('FONTNAME',      (0,0),(-1,0),  'Helvetica-Bold'),
        ('ROWBACKGROUNDS',(0,1),(-1,-1), [colors.white, colors.HexColor('#eff6ff')]),
        ('GRID',          (0,0),(-1,-1), 0.3, colors.HexColor('#e5e7eb')),
        ('TOPPADDING',    (0,0),(-1,-1), 4),
        ('BOTTOMPADDING', (0,0),(-1,-1), 4),
        ('LEFTPADDING',   (0,0),(-1,-1), 4),
        ('RIGHTPADDING',  (0,0),(-1,-1), 4),
        ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
    ]))
    story.append(items_tbl)

    doc_obj.build(story, onFirstPage=_draw_bottom, onLaterPages=_draw_bottom)
    buf.seek(0)
    fname = f"Nota_Pedido_{numero}.pdf"
    return StreamingResponse(buf, media_type='application/pdf',
                             headers={'Content-Disposition': f'inline; filename="{fname}"'})


class ItemDoc(BaseModel):
    codigoarticulo: str
    codigoparticular: str = ""
    descripcion: str
    cantidad: float
    preciounitario: float
    descuento: float = 0.0
    porcentajeiva: float = 21.0
    linea_presupuesto: Optional[int] = None   # línea en CUERPOPRESUPUESTOS si viene de presupuesto

class NuevoPedido(BaseModel):
    codigocliente: str
    razonsocial: str
    codigousuario: str
    comentarios: str = ""
    codigomultiplazo: Optional[str] = None
    codigotransporte: str = "0"
    codigodeposito: str = "001"
    domicilio_entrega: Optional[str] = None
    numero_presupuesto: Optional[str] = None  # si el pedido absorbe un presupuesto aprobado
    items: list[ItemDoc]

@app.post("/pedidos")
def crear_pedido(body: NuevoPedido, db: str = Query("oficial")):
    db_path = DATABASE_MLT if db == 'sw' else DATABASE
    # Obtener datos del cliente — busca en oficial si SW no tiene CLIENTES
    direccion, tipoiva, telefono, atencion_p, cuit_cli = '-', 'CF', '-', '', ''
    cli_transp_p, cli_reparto_p = '', ''
    def _tel_p(v):
        if v is None: return ''
        s = str(v).strip().lstrip('-').strip()
        try:
            s = str(int(float(s))) if s and ('E' in s.upper() or '.' in s) else s
        except Exception:
            pass
        return s
    for cli_db in ([db_path, DATABASE] if db == 'sw' else [db_path]):
        try:
            c_cli = conn('WIN1252', db=cli_db)
            cur_cli = c_cli.cursor()
            cur_cli.execute(
                'SELECT DIRECCION, TIPOIVA, TELEFONO, TELEFONOCELULAR, '
                'NOMBRE, APELLIDO, CODIGOTRANSPORTE, REPARTOPROPIO, CUIT '
                'FROM "CLIENTES" WHERE CODIGOCLIENTE = ?',
                (body.codigocliente,)
            )
            cli = cur_cli.fetchone()
            c_cli.close()
            if cli:
                direccion      = (cli[0] or '').strip() or '-'
                tipoiva        = (cli[1] or '').strip() or 'CF'
                tel_r          = _tel_p(cli[2])
                tel_c          = _tel_p(cli[3])
                telefono       = tel_r or tel_c or '-'
                nombre_c       = (cli[4] or '').strip()
                apellido_c     = (cli[5] or '').strip()
                atencion_p     = f"{nombre_c} {apellido_c}".strip()
                cli_transp_p   = str(cli[6] or '').strip()
                cli_reparto_p  = str(cli[7] or '').strip()
                cuit_cli       = str(cli[8] or '').strip()
                break
        except Exception:
            continue

    if body.domicilio_entrega:
        direccion = body.domicilio_entrega

    # Si cliente tiene reparto propio y no se eligió transporte, buscar código de REPARTO PROPIO
    if cli_reparto_p == '1' and (not body.codigotransporte or body.codigotransporte == '0'):
        try:
            c_tr = conn('WIN1252', db=DATABASE)
            cur_tr = c_tr.cursor()
            cur_tr.execute(
                "SELECT FIRST 1 CODIGOTRANSPORTE FROM \"TRANSPORTES\" "
                "WHERE UPPER(DESCRIPCION) CONTAINING 'REPARTO PROPIO' AND ACTIVO <> '0'"
            )
            row_tr = cur_tr.fetchone()
            c_tr.close()
            if row_tr:
                cli_transp_p = str(row_tr[0]).strip()
        except Exception:
            pass

    from datetime import datetime, timedelta
    now = datetime.now()
    fecha = now.strftime('%Y-%m-%d %H:%M:%S')
    fecha_entrega = (now + timedelta(days=2)).strftime('%Y-%m-%d %H:%M:%S')
    subtotal = sum(it.cantidad * it.preciounitario * (1 - it.descuento / 100) for it in body.items)
    iva1 = sum(it.cantidad * it.preciounitario * (1 - it.descuento / 100) * (it.porcentajeiva / 100) for it in body.items)
    total = subtotal

    # SW y L1 usan la misma BD (DATABASE) y la misma tabla (CABEZAPEDIDOS)
    # La diferencia: SW → PRIORIDAD='2', L1 → PRIORIDAD='1'
    c = conn('WIN1252', db=DATABASE)
    cur = c.cursor()
    try:
        if True:  # bloque unificado SW + L1
            # BD Oficial: usa CABEZAPEDIDOS / CUERPOPEDIDOS
            # Numeración desde PARAMETROS de DATABASE (BD principal) — mismo contador que SW
            # Usamos la misma conexión principal para lockear PARAMETROS atómicamente
            # WITH LOCK garantiza que Flexxus ERP no pueda leer el mismo número en paralelo
            cur.execute(
                "SELECT VALOR FROM \"PARAMETROS\" WHERE TRIM(TIPODOCUMENTO) = 'NP' WITH LOCK"
            )
            row_param_np = cur.fetchone()
            nuevo_num_int_o = int(float(row_param_np[0])) if row_param_np else 0
            # Garantía adicional: nunca usar un número ya existente en CABEZAPEDIDOS
            cur.execute(
                'SELECT MAX(CAST(NUMEROCOMPROBANTE AS INTEGER)) FROM "CABEZAPEDIDOS"'
                ' WHERE TIPOCOMPROBANTE = ?', ('NP',)
            )
            max_ped = int(cur.fetchone()[0] or 0)
            if nuevo_num_int_o <= max_ped:
                nuevo_num_int_o = max_ped + 1
            nuevo_num = str(nuevo_num_int_o)
            # Actualizar el contador YA (dentro del mismo lock) antes de cualquier INSERT
            cur.execute(
                "UPDATE \"PARAMETROS\" SET VALOR = ? WHERE TRIM(TIPODOCUMENTO) = 'NP'",
                (nuevo_num_int_o + 1,)
            )
            c_np = None  # ya no necesitamos conexión separada para numeración

            # Transporte: usar el del cliente como fallback
            transp_final = (
                body.codigotransporte if body.codigotransporte and body.codigotransporte != '0'
                else (cli_transp_p if cli_transp_p and cli_transp_p != '0' else '0')
            )

            deposito = body.codigodeposito or '001'

            cur.execute("""
                INSERT INTO "CABEZAPEDIDOS"
                (TIPOCOMPROBANTE, NUMEROCOMPROBANTE, CODIGOCLIENTE, RAZONSOCIAL,
                 FECHACOMPROBANTE, TOTAL, ANULADA, CODIGOUSUARIO, COMENTARIOS,
                 PORCIVA1, IVA1, PORCIVA2, IVA2, PAGADO, OPERACION, ENTREGAR,
                 DIRECCION, TIPOIVA, TELEFONO, FECHAENTREGA, PRIORIDAD,
                 VENTECNICOS, COMENTARIOSST, FORMAPAGO, CODIGOTRANSPORTE,
                 CODIGOMULTIPLAZO,
                 COEFICIENTEIVA, CODIGOMONEDA, COTIZACION,
                 NUMEROTRANSACCION, COTIZACIONFIJA, PORCENTAJEFLETE, MONTOFLETE,
                 LISTAPRECIO, CLASECOMPROBANTE, CODIGOACOPIO,
                 VALIDACTACTE, NUMEROAUTORIZACIONENTREGA, CUIT,
                 DESCUENTOPORCENTAJE, DESCUENTOMONTO, DESCUENTODESCRIPCION,
                 CODIGOUSUARIO2, FECHATERMINADA, FECHAMODIFICACION)
                VALUES ('NP', ?, ?, ?, ?, ?, '0', ?, ?,
                        0.0, ?, 0.0, 0.0, 0.0, '1', '1',
                        ?, ?, ?, ?, ?, '0', ' ', '', ?,
                        ?,
                        1.0, 'PESOS', 1.0, '0', '0', 0.0, 0.0,
                        '1', '0', '0',
                        1, '0', ?,
                        0.0, 0.0, '0,000000 %',
                        ' ', NULL, ?)
            """, (nuevo_num, body.codigocliente, body.razonsocial, fecha,
                  total, body.codigousuario.upper(), body.comentarios,
                  iva1, direccion, tipoiva, telefono, fecha_entrega,
                  '2' if db == 'sw' else '1',
                  transp_final,
                  int(body.codigomultiplazo) if body.codigomultiplazo else 0,
                  cuit_cli, fecha))

            for i, it in enumerate(body.items, 1):
                subtotal_item = it.cantidad * it.preciounitario * (1 - it.descuento / 100)
                cur.execute("""
                    INSERT INTO "CUERPOPEDIDOS"
                    (TIPOCOMPROBANTE, NUMEROCOMPROBANTE, LINEA, CODIGOARTICULO,
                     DESCRIPCION, CANTIDAD, DESCUENTO, PRECIOUNITARIO, PRECIOTOTAL,
                     PORCENTAJEIVA, CANTIDADREMITIDA, ESCONJUNTO,
                     GARANTIA, LOTE, FECHAMODIFICACION, NUMEROTRANSACCION,
                     CODIGODEPOSITO, CANTIDADPREPARADA, CANTIDADCOMPRA,
                     CANTIDADPRODUCCION, CANTIDADENVIADA, CANTIDADCANCELADA,
                     COEFICIENTECONVERSION, ESEXENTO, ESPRECIOPACTADO,
                     CODIGOPARTICULAR)
                    VALUES ('NP', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '0',
                            0, '000', ?, '0', ?, ?, 0,
                            0, 0, 0,
                            0.0, '0', '0',
                            ?)
                """, (nuevo_num, i, it.codigoarticulo, it.descripcion,
                      it.cantidad, it.descuento, it.preciounitario, subtotal_item,
                      it.porcentajeiva, 0, fecha, deposito, 0,
                      it.codigoparticular or it.codigoarticulo))

            # El UPDATE de PARAMETROS ya se hizo arriba dentro del mismo lock/transacción

        c.commit()

        # Si el pedido absorbe un presupuesto aprobado, actualizar CANTIDADREMITIDA
        if body.numero_presupuesto:
            try:
                c_pr = conn('LATIN1', db=DATABASE)
                cur_pr = c_pr.cursor()
                for it in body.items:
                    if it.linea_presupuesto is not None:
                        cur_pr.execute(
                            'UPDATE "CUERPOPRESUPUESTOS" '
                            'SET CANTIDADREMITIDA = COALESCE(CANTIDADREMITIDA, 0) + ? '
                            'WHERE NUMEROCOMPROBANTE = ? AND LINEA = ?',
                            (it.cantidad, body.numero_presupuesto, it.linea_presupuesto)
                        )
                c_pr.commit()
                c_pr.close()
            except Exception:
                pass  # No rollback del pedido por fallo en presupuesto

        return {"ok": True, "numero": nuevo_num, "total": round(total + iva1, 2)}
    except Exception as e:
        c.rollback()
        raise HTTPException(500, str(e))
    finally:
        c.close()

@app.get("/debug/cab_full/{numero}")
def debug_cab_full(numero: str):
    """Muestra TODOS los campos de CABEZAPEDIDOS para un NP dado."""
    try:
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()
        cur.execute(
            'SELECT * FROM "CABEZAPEDIDOS" WHERE TIPOCOMPROBANTE = ? AND NUMEROCOMPROBANTE = ?',
            ('NP', numero)
        )
        row = cur.fetchone()
        if not row:
            c.close()
            return {"error": f"NP {numero} no encontrado"}
        cols = [d[0] for d in cur.description]
        c.close()
        return dict(zip(cols, [str(v) if v is not None else None for v in row]))
    except Exception as e:
        return {"error": str(e)}

@app.get("/pedidos/proximo")
def get_proximo_pedido(db: str = Query("oficial")):
    # El contador NP siempre está en DATABASE (BD principal), sin importar si es SW u oficial
    try:
        c = conn('LATIN1', db=DATABASE)
        cur = c.cursor()
        cur.execute("SELECT VALOR FROM \"PARAMETROS\" WHERE TRIM(TIPODOCUMENTO) = 'NP'")
        row = cur.fetchone()
        c.close()
        if not row:
            raise HTTPException(500, "No se encontró parámetro de numeración NP")
        return {"proximo": int(float(row[0]))}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/presupuestos/proximo")
def get_proximo_presupuesto(db: str = Query("oficial")):
    db_path = DATABASE_MLT if db == 'sw' else DATABASE
    try:
        c = conn('LATIN1', db=db_path)
        cur = c.cursor()
        cur.execute("SELECT VALOR FROM \"PARAMETROS\" WHERE TIPODOCUMENTO = 'PR'")
        row = cur.fetchone()
        c.close()
        if not row:
            raise HTTPException(500, "No se encontró parámetro de numeración PR")
        return {"proximo": int(float(row[0]))}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))

# ─── Presupuestos (solo del vendedor) ─────────────────────────────────────────
@app.get("/presupuestos")
def get_presupuestos(
    vendedor: str,
    cliente: Optional[str] = None,
    limit: int = Query(50, le=200),
    offset: int = 0
):
    c = conn()
    cur = c.cursor()
    params = [vendedor.upper()]
    where_cli = ""
    if cliente:
        where_cli = "AND CODIGOCLIENTE = ?"
        params.append(cliente)

    cur.execute(f"""
        SELECT FIRST {limit} SKIP {offset}
            NUMEROCOMPROBANTE, CODIGOCLIENTE, RAZONSOCIAL,
            FECHACOMPROBANTE, FECHAVENCIMIENTO, TOTAL, ANULADA, COMENTARIOS
        FROM "CABEZAPRESUPUESTOS"
        WHERE CODIGOUSUARIO = ? AND ANULADA = '0'
            {where_cli}
        ORDER BY FECHACOMPROBANTE DESC, NUMEROCOMPROBANTE DESC
    """, params)
    rows = cur.fetchall()
    c.close()
    return [{
        "numero": r[0], "cod_cliente": r[1], "razonsocial": r[2],
        "fecha": r[3], "fecha_vto": r[4], "total": r[5],
        "anulada": r[6], "comentarios": r[7]
    } for r in rows]

@app.get("/presupuestos/pendientes/{codigocliente}")
def get_presupuestos_pendientes(codigocliente: str, db: str = Query("oficial")):
    """Presupuestos aprobados con items aún no completados para el cliente."""
    # CABEZAPRESUPUESTOS siempre en BD principal
    c = conn('LATIN1', db=DATABASE)
    cur = c.cursor()
    # Cabezas: aprobadas (FECHAAPROBADO IS NOT NULL), no anuladas, con items pendientes
    cur.execute("""
        SELECT DISTINCT cp.NUMEROCOMPROBANTE, cp.FECHACOMPROBANTE, cp.TOTAL, cp.COMENTARIOS,
               cp.CODIGOMULTIPLAZO, cp.CODIGOTRANSPORTE, cp.DIRECCION
        FROM "CABEZAPRESUPUESTOS" cp
        JOIN "CUERPOPRESUPUESTOS" cu ON cu.NUMEROCOMPROBANTE = cp.NUMEROCOMPROBANTE
        WHERE cp.CODIGOCLIENTE = ?
          AND cp.ANULADA = '0'
          AND cp.FECHAAPROBADO IS NOT NULL
          AND cp.FECHAAPROBADO > CAST('1900-01-02 00:00:00' AS TIMESTAMP)
          AND (cu.CANTIDAD - COALESCE(cu.CANTIDADREMITIDA, 0)) > 0
        ORDER BY cp.NUMEROCOMPROBANTE DESC
    """, (codigocliente,))
    cabezas = cur.fetchall()
    result = []
    for cab in cabezas:
        numero = cab[0]
        cur2 = c.cursor()
        cur2.execute("""
            SELECT LINEA, CODIGOARTICULO, DESCRIPCION,
                   CANTIDAD, COALESCE(CANTIDADREMITIDA, 0),
                   BONIFICACION, PRECIOUNITARIO, PORCENTAJEIVA,
                   CODIGOPARTICULAR
            FROM "CUERPOPRESUPUESTOS"
            WHERE NUMEROCOMPROBANTE = ?
              AND (CANTIDAD - COALESCE(CANTIDADREMITIDA, 0)) > 0
            ORDER BY LINEA
        """, (numero,))
        items = []
        for r in cur2.fetchall():
            pendiente = float(r[3]) - float(r[4])
            cod_particular = (r[8] or '').strip() or (r[1] or '').strip()
            items.append({
                "linea": r[0], "codigo": r[1], "codigoparticular": cod_particular,
                "descripcion": r[2],
                "cantidad_total": float(r[3]),
                "cantidad_remitida": float(r[4]),
                "cantidad_pendiente": pendiente,
                "descuento": float(r[5]),
                "precio_unitario": float(r[6]),
                "iva": float(r[7])
            })
        result.append({
            "numero": numero,
            "fecha": str(cab[1]),
            "total": float(cab[2]) if cab[2] else 0,
            "comentarios": cab[3] or "",
            "codigomultiplazo": str(cab[4] or '0'),
            "codigotransporte": str(cab[5] or '0'),
            "direccion": cab[6] or "",
            "items": items
        })
    c.close()
    return result

@app.get("/presupuestos/{numero}/detalle")
def get_presupuesto_detalle(numero: str):
    c = conn()
    cur = c.cursor()
    cur.execute(
        "SELECT cp.LINEA, "
        "COALESCE(NULLIF(TRIM(cp.CODIGOPARTICULAR),''), NULLIF(TRIM(a.CODIGOPARTICULAR),''), TRIM(cp.CODIGOARTICULO)), "
        "cp.DESCRIPCION, cp.CANTIDAD, "
        "cp.BONIFICACION, cp.PRECIOUNITARIO, cp.PRECIOTOTAL, cp.PORCENTAJEIVA "
        "FROM \"CUERPOPRESUPUESTOS\" cp "
        "LEFT JOIN \"ARTICULOS\" a ON a.CODIGOARTICULO = cp.CODIGOARTICULO "
        "WHERE cp.NUMEROCOMPROBANTE = ? ORDER BY cp.LINEA",
        (numero,)
    )
    rows = cur.fetchall()
    c.close()
    return [{
        "linea": r[0], "codigo": r[1], "descripcion": r[2],
        "cantidad": r[3], "descuento": r[4],
        "precio_unitario": r[5], "precio_total": r[6], "iva": r[7]
    } for r in rows]

class NuevoPresupuesto(BaseModel):
    codigocliente: str
    razonsocial: str
    codigousuario: str
    comentarios: str = ""
    codigomultiplazo: Optional[str] = None
    codigotransporte: str = "0"
    domicilio_entrega: Optional[str] = None
    items: list[ItemDoc]

@app.post("/presupuestos")
def crear_presupuesto(body: NuevoPresupuesto, db: str = Query("oficial")):
    db_path = DATABASE_MLT if db == 'sw' else DATABASE
    # Obtener datos del cliente (WIN1252)
    try:
        c_cli = conn('WIN1252', db=db_path)
        cur_cli = c_cli.cursor()
        cur_cli.execute(
            'SELECT DIRECCION, TIPOIVA, TELEFONO, TELEFONOCELULAR, NOMBRE, APELLIDO, '
            'CODIGOTRANSPORTE, REPARTOPROPIO '
            'FROM "CLIENTES" WHERE CODIGOCLIENTE = ?',
            (body.codigocliente,)
        )
        cli = cur_cli.fetchone()
        c_cli.close()
        if cli:
            direccion     = (cli[0] or '').strip() or ''
            tipoiva       = (cli[1] or '').strip() or 'CF'
            def _tel(v):
                if v is None: return ''
                s = str(v).strip().lstrip('-').strip()
                # Si Firebird devuelve numérico (ej: 4.51E+14), convertir a entero
                try:
                    s = str(int(float(s))) if s and ('E' in s.upper() or '.' in s) else s
                except Exception:
                    pass
                return s
            tel_raw       = _tel(cli[2])
            tel_cel       = _tel(cli[3])
            telefono      = tel_raw or tel_cel
            nombre_c      = (cli[4] or '').strip()
            apellido_c    = (cli[5] or '').strip()
            atencion      = f"{nombre_c} {apellido_c}".strip()
            cli_transp    = str(cli[6] or '').strip()
            cli_reparto   = str(cli[7] or '').strip()
        else:
            direccion, tipoiva, telefono, atencion = '', 'CF', '', ''
            cli_transp, cli_reparto = '', ''
    except Exception:
        direccion, tipoiva, telefono, atencion = '', 'CF', '', ''
        cli_transp, cli_reparto = '', ''

    if body.domicilio_entrega:
        direccion = body.domicilio_entrega

    from datetime import datetime, timedelta
    now = datetime.now()
    fecha = now.strftime('%Y-%m-%d %H:%M:%S')
    fecha_vto = (now + timedelta(days=7)).strftime('%Y-%m-%d %H:%M:%S')
    subtotal = sum(it.cantidad * it.preciounitario * (1 - it.descuento / 100) for it in body.items)
    iva1 = sum(it.cantidad * it.preciounitario * (1 - it.descuento / 100) * (it.porcentajeiva / 100) for it in body.items)
    total = subtotal

    c = conn('LATIN1', db=db_path)
    cur = c.cursor()
    try:
        cur.execute("SELECT VALOR FROM \"PARAMETROS\" WHERE TIPODOCUMENTO = 'PR'")
        row_param = cur.fetchone()
        if not row_param:
            raise Exception("No se encontró parámetro de numeración PR en PARAMETROS")
        nuevo_num = str(int(float(row_param[0])))

        cur.execute("""
                INSERT INTO "CABEZAPRESUPUESTOS"
                (TIPOCOMPROBANTE, NUMEROCOMPROBANTE, CODIGOCLIENTE, RAZONSOCIAL,
                 FECHACOMPROBANTE, FECHAVENCIMIENTO, TOTAL, ANULADA,
                 CODIGOUSUARIO, COMENTARIOS,
                 PORCIVA1, IVA1, PORCIVA2, IVA2,
                 DIRECCION, TIPOIVA, TELEFONO,
                 CODIGOUSUARIO2, COEFICIENTEIVA, FECHAMODIFICACION,
                 CODIGORESPONSABLE, CODIGOMONEDA, COTIZACION,
                 ATENCION, NUMEROTRANSACCION, FORMAPAGO, CODIGOOPERACION,
                 LISTAPRECIO, COTIZACIONFIJA, PORCENTAJEFLETE, MONTOFLETE,
                 CLASECOMPROBANTE, CODIGOACOPIO,
                 CODIGOMULTIPLAZO, CODIGOTRANSPORTE,
                 FECHAAPROBADO, CODIGOUSUARIOAPROBACION)
                VALUES ('PR', ?, ?, ?, ?, ?, ?, '0', ?, ?,
                        21.0, ?, 0.0, 0.0,
                        ?, ?, ?,
                        ?, 1.0, ?,
                        ?, 'PESOS', 1.0,
                        ?, '0', '', '1',
                        '1', '0', 0.0, 0.0,
                        '0', '0',
                        ?, ?,
                        '1900-01-01 00:00:00', '')
            """, (nuevo_num, body.codigocliente, body.razonsocial, fecha,
                  fecha_vto, total, body.codigousuario.upper(), body.comentarios,
                  iva1,
                  direccion, tipoiva, telefono,
                  body.codigousuario.upper(), fecha,
                  body.codigousuario.upper(), atencion,
                  int(body.codigomultiplazo) if body.codigomultiplazo else 0,
                  body.codigotransporte if body.codigotransporte and body.codigotransporte != '0'
                  else (cli_transp if cli_transp and cli_transp != '0' else '0')))

        for i, it in enumerate(body.items, 1):
            subtotal_item = it.cantidad * it.preciounitario * (1 - it.descuento / 100)
            cur.execute("""
                INSERT INTO "CUERPOPRESUPUESTOS"
                (TIPOCOMPROBANTE, NUMEROCOMPROBANTE, LINEA, CODIGOARTICULO,
                 DESCRIPCION, CANTIDAD, BONIFICACION, PRECIOUNITARIO, PRECIOTOTAL,
                 PORCENTAJEIVA, CANTIDADREMITIDA, ESCONJUNTO,
                 GARANTIA, FECHAMODIFICACION, NUMEROTRANSACCION,
                 CODIGOPARTICULAR,
                 LOTE, INTERES, COEFICIENTECONVERSION,
                 ITEMGANADO, ESEXENTO, ESALTERNATIVO, ESPRECIOPACTADO)
                VALUES ('PR', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '0',
                        0, ?, '0',
                        ?,
                        '000', 0.0, 0.0,
                        '1', '0', '0', '0')
            """, (nuevo_num, i, it.codigoarticulo, it.descripcion,
                  it.cantidad, it.descuento, it.preciounitario, subtotal_item,
                  it.porcentajeiva, 0, fecha,
                  it.codigoparticular or it.codigoarticulo))

        cur.execute(
            "UPDATE \"PARAMETROS\" SET VALOR = ? WHERE TIPODOCUMENTO = 'PR'",
            (int(float(row_param[0])) + 1,)
        )
        c.commit()
        c.close()
        return {"ok": True, "numero": nuevo_num, "total": round(total + iva1, 2)}

    except Exception as e:
        c.rollback()
        c.close()
        raise HTTPException(500, str(e))

# ─── PDF Presupuesto ───────────────────────────────────────────────────────────
@app.get("/presupuestos/{numero}/pdf")
def presupuesto_pdf(numero: str, db: str = Query("oficial")):
    from datetime import datetime
    db_path = DATABASE_MLT if db == 'sw' else DATABASE

    # ── 1. Datos empresa ───────────────────────────────────────────────────────
    razon_soc = 'MICROBELL S.A.'
    dir_emp   = 'AV. MONROE 5088 PISO 2'
    tel_emp   = '+54 11 3988-0024'
    email_emp = 'info@microbellsa.com.ar'
    web_emp   = 'www.microbellsa.com'
    cuit_emp  = '30-70839018-2'
    iibb_emp  = 'CM 901-068199-0'
    fi_emp    = '2021-05-01'
    try:
        c_e = conn('WIN1252', db=db_path)
        cur_e = c_e.cursor()
        cur_e.execute(
            'SELECT RAZONSOCIAL, DIRECCION, TELEFONO, EMAIL, DIRECCIONWEB, '
            'CUIT, INGRESOSBRUTOS, FE_FECHAINC '
            'FROM "SUCURSALES" WHERE CODIGOSUCURSAL = ?', ('PRINCIPAL',)
        )
        emp = cur_e.fetchone()
        c_e.close()
        if emp:
            razon_soc = (emp[0] or razon_soc).strip()
            dir_emp   = (emp[1] or dir_emp).strip()
            tel_emp   = (emp[2] or tel_emp).strip()
            email_emp = (emp[3] or email_emp).strip()
            web_emp   = (emp[4] or web_emp).strip()
            cuit_emp  = (emp[5] or cuit_emp).strip()
            iibb_emp  = (emp[6] or iibb_emp).strip()
            fi_emp    = emp[7] if emp[7] else fi_emp
    except Exception:
        pass  # usa los valores hardcodeados

    # ── 2. Cabeza presupuesto ──────────────────────────────────────────────────
    try:
        c_h = conn('WIN1252', db=db_path)
        cur_h = c_h.cursor()
        cur_h.execute(
            'SELECT CODIGOCLIENTE, RAZONSOCIAL, FECHACOMPROBANTE, FECHAVENCIMIENTO, '
            'TOTAL, IVA1, IVA2, COMENTARIOS, CODIGOUSUARIO, CODIGOMULTIPLAZO, '
            'CODIGOTRANSPORTE, DIRECCION, TIPOIVA, TELEFONO '
            'FROM "CABEZAPRESUPUESTOS" WHERE NUMEROCOMPROBANTE = ?', (numero,)
        )
        cab = cur_h.fetchone()
        c_h.close()
    except Exception as ex:
        raise HTTPException(500, f"Cabeza: {ex}")

    if not cab:
        raise HTTPException(404, f"Presupuesto {numero} no encontrado")

    cod_cli, rs_cli, fec_comp, fec_vto, total_cab, iva1_cab, iva2_cab, \
    comentarios, cod_usu, cod_multi, cod_transp, dir_cli, tipo_iva, tel_cli = cab

    subtotal_cab = float(total_cab or 0)
    iva1_val     = float(iva1_cab or 0)
    iva2_val     = float(iva2_cab or 0)
    total_final  = subtotal_cab + iva1_val + iva2_val

    # ── 3. Items ───────────────────────────────────────────────────────────────
    try:
        c_it = conn('WIN1252', db=db_path)
        cur_it = c_it.cursor()
        cur_it.execute(
            'SELECT COALESCE(NULLIF(TRIM(it.CODIGOPARTICULAR),\'\'), NULLIF(TRIM(a.CODIGOPARTICULAR),\'\'), TRIM(it.CODIGOARTICULO)), '
            'it.DESCRIPCION, it.CANTIDAD, it.PRECIOUNITARIO, '
            'it.BONIFICACION, it.PRECIOTOTAL, it.PORCENTAJEIVA '
            'FROM "CUERPOPRESUPUESTOS" it '
            'LEFT JOIN "ARTICULOS" a ON a.CODIGOARTICULO = it.CODIGOARTICULO '
            'WHERE it.NUMEROCOMPROBANTE = ? ORDER BY it.LINEA',
            (numero,)
        )
        items = cur_it.fetchall()
        c_it.close()
    except Exception as ex:
        raise HTTPException(500, f"Items: {ex}")

    # ── 4. Datos cliente (CUIT, teléfono, vendedor) ───────────────────────────
    cuit_cli = ''
    vendedor_nombre = cod_usu or ''
    # tel_cli viene de CABEZAPRESUPUESTOS; si está vacío lo buscamos en CLIENTES
    tel_pdf = str(tel_cli or '').strip().lstrip('-').strip()
    try:
        c_cl = conn('WIN1252', db=db_path)
        cur_cl = c_cl.cursor()
        cur_cl.execute(
            'SELECT CUIT, TELEFONO, TELEFONOCELULAR FROM "CLIENTES" WHERE CODIGOCLIENTE = ?',
            (cod_cli,)
        )
        r_cl = cur_cl.fetchone()
        if r_cl:
            cuit_cli = (r_cl[0] or '').strip()
            if not tel_pdf:
                tel_pdf = (r_cl[1] or r_cl[2] or '').strip()
        cur_cl.execute(
            'SELECT RAZONSOCIAL FROM "USUARIOS" WHERE CODIGOUSUARIO = ?', (cod_usu,)
        )
        r_vend = cur_cl.fetchone()
        if r_vend:
            vendedor_nombre = (r_vend[0] or '').strip()
        c_cl.close()
    except Exception:
        pass

    # ── 5. Transporte ──────────────────────────────────────────────────────────
    transporte_desc = str(cod_transp or '').strip()
    if transporte_desc and transporte_desc != '0':
        for _tr_db in ([db_path, DATABASE] if db == 'sw' else [db_path]):
            try:
                c_tr = conn('WIN1252', db=_tr_db)
                cur_tr = c_tr.cursor()
                cur_tr.execute('SELECT DESCRIPCION FROM "TRANSPORTES" WHERE CODIGOTRANSPORTE = ?', (cod_transp,))
                r_tr = cur_tr.fetchone()
                c_tr.close()
                if r_tr and (r_tr[0] or '').strip():
                    transporte_desc = r_tr[0].strip()
                    break
            except Exception:
                pass
        else:
            transporte_desc = 'A CONVENIR'
    else:
        transporte_desc = 'A CONVENIR'

    # ── 6. Condición de pago ───────────────────────────────────────────────────
    cond_pago = str(cod_multi or '').strip()
    if cond_pago and cond_pago not in ('0', ''):
        try:
            c_mp = conn('WIN1252', db=db_path)
            cur_mp = c_mp.cursor()
            cur_mp.execute('SELECT DESCRIPCION FROM "MULTIPLAZOS" WHERE CODIGOMULTIPLAZO = ?', (cod_multi,))
            r_mp = cur_mp.fetchone()
            if r_mp:
                cond_pago = (r_mp[0] or '').strip()
            c_mp.close()
        except Exception:
            pass
    else:
        cond_pago = ''

    # ── 7. Construir PDF ───────────────────────────────────────────────────────
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=10*mm, rightMargin=10*mm,
        topMargin=4*mm, bottomMargin=12*mm
    )

    styles = getSampleStyleSheet()
    W = A4[0] - 20*mm  # ancho útil

    s_norm  = ParagraphStyle('norm',  fontSize=7,  leading=9,  fontName='Helvetica')
    s_bold  = ParagraphStyle('bold',  fontSize=7,  leading=9,  fontName='Helvetica-Bold')
    s_title = ParagraphStyle('title', fontSize=11, leading=13, fontName='Helvetica-Bold', alignment=TA_CENTER)
    s_small = ParagraphStyle('small', fontSize=6,  leading=7,  fontName='Helvetica')
    s_c     = ParagraphStyle('c',     fontSize=7,  leading=9,  fontName='Helvetica', alignment=TA_CENTER)
    s_r     = ParagraphStyle('r',     fontSize=7,  leading=9,  fontName='Helvetica', alignment=TA_RIGHT)
    s_rb    = ParagraphStyle('rb',    fontSize=7,  leading=9,  fontName='Helvetica-Bold', alignment=TA_RIGHT)

    GRIS   = colors.HexColor('#D9D9D9')
    OSCURO = colors.HexColor('#1F3864')
    AZUL   = colors.HexColor('#2E74B5')

    story = []

    # ── ENCABEZADO ─────────────────────────────────────────────────────────────
    # Logo
    logo_cell = ''
    if os.path.exists(LOGO_PATH):
        try:
            _logo_img = Image(LOGO_PATH, width=36*mm, height=36*0.32*mm)
            _logo_img.hAlign = 'CENTER'
            logo_cell = _logo_img
        except Exception:
            logo_cell = Paragraph(razon_soc, s_bold)
    else:
        logo_cell = Paragraph(razon_soc, s_bold)

    emp_info = (
        f"<b>{razon_soc}</b><br/>"
        f"{dir_emp}<br/>"
        f"Tel: {tel_emp}<br/>"
        f"{email_emp}<br/>"
        f"{web_emp}"
    )

    # Caja tipo comprobante "P" (Presupuesto)
    tipo_box = Table(
        [[Paragraph('<b>P</b>', ParagraphStyle('pb', fontSize=22, fontName='Helvetica-Bold', alignment=TA_CENTER))]],
        colWidths=[15*mm], rowHeights=[15*mm]
    )
    tipo_box.setStyle(TableStyle([
        ('BOX',       (0,0),(0,0), 1.5, colors.black),
        ('VALIGN',    (0,0),(0,0), 'MIDDLE'),
        ('ALIGN',     (0,0),(0,0), 'CENTER'),
    ]))
    cod_info = Paragraph('COD.908', ParagraphStyle('cod', fontSize=8, fontName='Helvetica', alignment=TA_CENTER))

    nro_fmt = numero.zfill(8)
    doc_info = (
        f"<b>PRESUPUESTO</b><br/>"
        f"Nº 0001-{nro_fmt}<br/>"
        f"Fecha: {_d(fec_comp)}<br/>"
        f"CUIT: {cuit_emp}<br/>"
        f"IIBB: {iibb_emp}<br/>"
        f"F.Inicio: {_d(fi_emp)}<br/>"
        f"<font size='5'><b>DOCUMENTO NO VÁLIDO COMO FACTURA</b></font>"
    )

    header_data = [[
        logo_cell,
        Paragraph(emp_info, s_norm),
        Table([[tipo_box],[cod_info]], colWidths=[20*mm]),
        Paragraph(doc_info, s_norm),
    ]]
    header_table = Table(header_data, colWidths=[45*mm, 70*mm, 22*mm, None])
    header_table.setStyle(TableStyle([
        ('VALIGN',    (0,0),(-1,-1), 'TOP'),
        ('ALIGN',     (0,0),(0,0),   'CENTER'),
        ('LEFTPADDING',  (0,0),(-1,-1), 2),
        ('RIGHTPADDING', (0,0),(-1,-1), 2),
        ('TOPPADDING',   (0,0),(-1,-1), 0),
        ('BOTTOMPADDING',(0,0),(-1,-1), 0),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 3*mm))

    # ── DATOS CLIENTE ──────────────────────────────────────────────────────────
    cli_left = (
        f"<b>Cliente:</b> {cod_cli} - {(rs_cli or '').strip()}<br/>"
        f"<b>Dirección:</b> {(dir_cli or '').strip()}<br/>"
        f"<b>Tel:</b> {tel_pdf}<br/>"
        f"<b>CUIT:</b> {cuit_cli}"
    )
    cli_right = (
        f"<b>Cond. IVA:</b> {(tipo_iva or '').strip()}<br/>"
        f"<b>Transporte:</b> {transporte_desc}<br/>"
        f"<b>Asistente:</b> {cod_usu or ''}<br/>"
        f"<b>Vendedor:</b> {vendedor_nombre}"
    )
    cli_table = Table(
        [[Paragraph(cli_left, s_norm), Paragraph(cli_right, s_norm)]],
        colWidths=[W*0.55, W*0.45]
    )
    cli_table.setStyle(TableStyle([
        ('BACKGROUND',   (0,0),(-1,-1), GRIS),
        ('BOX',          (0,0),(-1,-1), 0.5, colors.grey),
        ('LEFTPADDING',  (0,0),(-1,-1), 4),
        ('RIGHTPADDING', (0,0),(-1,-1), 4),
        ('TOPPADDING',   (0,0),(-1,-1), 3),
        ('BOTTOMPADDING',(0,0),(-1,-1), 3),
        ('VALIGN',       (0,0),(-1,-1), 'TOP'),
    ]))
    story.append(cli_table)
    story.append(Spacer(1, 2*mm))

    # ── TABLA ITEMS ────────────────────────────────────────────────────────────
    col_w = [22*mm, None, 18*mm, 28*mm, 18*mm, 28*mm]
    # cabecera
    hdr = ['Código', 'Descripción', 'Cantidad', 'Precio Unit.', 'Bonif.%', 'Subtotal']
    items_data = [[Paragraph(f'<b>{h}</b>', s_c) for h in hdr]]

    for it in items:
        cod_art, desc_art, cant, pu, bonif, ptotal, piva = it
        # Limpiar código: quitar espacios y el .0 si viene como float
        cod_str = str(cod_art or '').strip()
        if cod_str.endswith('.0'):
            cod_str = cod_str[:-2]
        try:
            cant_str = str(int(float(cant))) if float(cant) == int(float(cant)) else str(cant)
        except Exception:
            cant_str = str(cant or '')
        items_data.append([
            Paragraph(cod_str, s_c),
            Paragraph(str(desc_art or ''), s_norm),
            Paragraph(cant_str, s_c),
            Paragraph(_fmt(pu), s_r),
            Paragraph(f"{float(bonif or 0):.1f}%", s_c),
            Paragraph(_fmt(ptotal), s_r),
        ])

    # fila totales
    items_data.append(['', '', '', '', Paragraph('<b>SUBTOTAL</b>', s_rb), Paragraph(_fmt(subtotal_cab), s_rb)])
    if iva1_val:
        items_data.append(['', '', '', '', Paragraph('<b>IVA 21%</b>', s_rb), Paragraph(_fmt(iva1_val), s_rb)])
    if iva2_val:
        items_data.append(['', '', '', '', Paragraph('<b>IVA 10.5%</b>', s_rb), Paragraph(_fmt(iva2_val), s_rb)])
    items_data.append(['', '', '', '', Paragraph('<b>TOTAL</b>', s_rb), Paragraph(_fmt(total_final), s_rb)])

    items_table = Table(items_data, colWidths=col_w, repeatRows=1)
    n_items = len(items) + 1  # +1 header
    ts = TableStyle([
        ('BACKGROUND',    (0,0),(-1,0),          AZUL),
        ('TEXTCOLOR',     (0,0),(-1,0),          colors.white),
        ('FONTNAME',      (0,0),(-1,0),          'Helvetica-Bold'),
        ('FONTSIZE',      (0,0),(-1,-1),         7),
        ('ROWBACKGROUNDS',(0,1),(-1,n_items-1),  [colors.white, colors.HexColor('#EEF3F8')]),
        ('GRID',          (0,0),(-1,n_items-1),  0.3, colors.grey),
        ('LINEABOVE',     (0,n_items),(-1,n_items), 0.8, colors.black),
        ('TOPPADDING',    (0,0),(-1,-1),         2),
        ('BOTTOMPADDING', (0,0),(-1,-1),         2),
        ('LEFTPADDING',   (0,0),(-1,-1),         3),
        ('RIGHTPADDING',  (0,0),(-1,-1),         3),
        ('VALIGN',        (0,0),(-1,-1),         'MIDDLE'),
    ])
    # línea separadora antes del TOTAL
    ts.add('LINEABOVE', (4,-1), (-1,-1), 1.2, colors.black)
    items_table.setStyle(ts)
    story.append(items_table)
    story.append(Spacer(1, 3*mm))

    # ── CONDICIONES DE VENTA ───────────────────────────────────────────────────
    cond_data = [
        [Paragraph('<b>Condiciones de Venta</b>', s_bold), ''],
        [Paragraph(f'<b>Cond. Pago:</b> {cond_pago}', s_norm),
         Paragraph(f'<b>Tipo Cambio:</b> $ 1,00', s_norm)],
        [Paragraph(f'<b>Domicilio Entrega:</b> {(dir_cli or "").strip()}', s_norm),
         Paragraph(f'<b>Fecha Venc.:</b> {_d(fec_vto)}', s_norm)],
        [Paragraph(f'<b>Observaciones:</b> {(comentarios or "").strip()}', s_norm), ''],
    ]
    cond_table = Table(cond_data, colWidths=[W*0.6, W*0.4])
    cond_table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),(-1,0), GRIS),
        ('SPAN',          (0,0),(-1,0)),
        ('SPAN',          (0,3),(-1,3)),
        ('BOX',           (0,0),(-1,-1), 0.5, colors.grey),
        ('INNERGRID',     (0,0),(-1,-1), 0.2, colors.lightgrey),
        ('LEFTPADDING',   (0,0),(-1,-1), 4),
        ('RIGHTPADDING',  (0,0),(-1,-1), 4),
        ('TOPPADDING',    (0,0),(-1,-1), 2),
        ('BOTTOMPADDING', (0,0),(-1,-1), 2),
        ('FONTSIZE',      (0,0),(-1,-1), 7),
    ]))
    story.append(cond_table)
    story.append(Spacer(1, 3*mm))

    # ── CUENTAS BANCARIAS (en flujo, centrada, angosta) ───────────────────────
    s_bh  = ParagraphStyle('bh',  fontSize=6, fontName='Helvetica-Bold', textColor=colors.white, alignment=TA_CENTER, leading=8)
    s_bno = ParagraphStyle('bno', fontSize=6, fontName='Helvetica', leading=8)
    s_bc  = ParagraphStyle('bc',  fontSize=6, fontName='Helvetica', leading=8, alignment=TA_CENTER)
    s_bct = ParagraphStyle('bct', fontSize=7, fontName='Helvetica-Bold', textColor=colors.white, alignment=TA_CENTER, leading=9)

    # Fila título "CUENTAS BANCARIAS" que abarca todas las columnas
    # Fila cabecera columnas
    bank_data = [
        [Paragraph('<b>CUENTAS BANCARIAS</b>', s_bct), '', '', ''],
        [Paragraph('<b>BANCO</b>',   s_bh),
         Paragraph('<b>CBU</b>',     s_bh),
         Paragraph('<b>SUCURSAL</b>',s_bh),
         Paragraph('<b>CTA. CTE.</b>',s_bh)],
    ]
    for nombre, cbu, suc, cta in _BANCOS:
        bank_data.append([
            Paragraph(nombre, s_bno), Paragraph(cbu, s_bno),
            Paragraph(suc, s_bc),     Paragraph(cta, s_bno),
        ])
    bank_data.append([
        Paragraph('<b>Mercado Pago</b>', ParagraphStyle('bmp', fontSize=6, fontName='Helvetica-Bold', leading=8)),
        Paragraph(f'{_MP_EMAIL}', s_bno),
        Paragraph('CVU', s_bc),
        Paragraph(_MP_CVU, s_bno),
    ])

    # Ancho total ~118mm, columnas: banco 26 + CBU 54 + suc 15 + cta 23
    BW = [26*mm, 54*mm, 15*mm, 23*mm]
    bank_table = Table(bank_data, colWidths=BW)
    bank_table.setStyle(TableStyle([
        # Título
        ('BACKGROUND',    (0,0),(-1,0),  OSCURO),
        ('SPAN',          (0,0),(-1,0)),
        ('ALIGN',         (0,0),(-1,0),  'CENTER'),
        ('TOPPADDING',    (0,0),(-1,0),  3),
        ('BOTTOMPADDING', (0,0),(-1,0),  3),
        # Cabecera columnas
        ('BACKGROUND',    (0,1),(-1,1),  OSCURO),
        # Filas datos
        ('ROWBACKGROUNDS',(0,2),(-1,-1), [colors.white, colors.HexColor('#EEF3F8')]),
        ('BOX',           (0,0),(-1,-1), 0.5, colors.grey),
        ('INNERGRID',     (0,1),(-1,-1), 0.2, colors.lightgrey),
        ('LEFTPADDING',   (0,0),(-1,-1), 3),
        ('RIGHTPADDING',  (0,0),(-1,-1), 3),
        ('TOPPADDING',    (0,1),(-1,-1), 2),
        ('BOTTOMPADDING', (0,1),(-1,-1), 2),
        ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
    ]))
    bank_table.hAlign = 'CENTER'

    story.append(Spacer(1, 4*mm))
    story.append(bank_table)

    doc.build(story)
    buf.seek(0)
    fname = f"Presupuesto_{nro_fmt}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{fname}"'}
    )

# ─── Detalle de comprobante ────────────────────────────────────────────────────
@app.get("/comprobantes/{tipo}/{numero}/detalle")
def detalle_comprobante(tipo: str, numero: str):
    sql = """
        SELECT CODIGOARTICULO, DESCRIPCION, CANTIDAD,
               PRECIOUNITARIO, PORCENTAJEIVA, DESCUENTO, PRECIOTOTAL
        FROM "CUERPOCOMPROBANTES"
        WHERE TIPOCOMPROBANTE = ? AND NUMEROCOMPROBANTE = ?
        ORDER BY LINEA
    """
    DB_PROD     = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    for db_path in [DATABASE, DATABASE_MLT, DB_PROD, DB_MLT_PROD]:
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            cur.execute(sql, (tipo, numero))
            rows = cur.fetchall()
            c.close()
            if rows:
                return [{
                    "codigo":      r[0], "descripcion": r[1],
                    "cantidad":    float(r[2]) if r[2] else 0,
                    "precio_unit": float(r[3]) if r[3] else 0,
                    "iva":         float(r[4]) if r[4] else 0,
                    "descuento":   float(r[5]) if r[5] else 0,
                    "total":       float(r[6]) if r[6] else 0,
                } for r in rows]
        except Exception:
            continue
    return []

# ─── PDF Comprobante (FA/FB/FCA/FE/NCA/NCB/NDA/NDB y Proforma SW) ────────────

_TIPO_NOMBRE = {
    'FA':'FACTURA A','FB':'FACTURA B','FCA':'FACTURA DE CREDITO A',
    'FCB':'FACTURA DE CREDITO B','FE':'FACTURA E',
    'NCA':'NOTA DE CREDITO A','NCB':'NOTA DE CREDITO B',
    'NCCA':'NOTA DE CREDITO DE CREDITO A','NCE':'NOTA DE CREDITO E',
    'NDA':'NOTA DE DEBITO A','NDB':'NOTA DE DEBITO B','NDE':'NOTA DE DEBITO E',
}
_TIPO_LETRA = {
    'FA':'A','FCA':'A','NCA':'A','NDA':'A','NCCA':'A',
    'FB':'B','FCB':'B','NCB':'B','NDB':'B',
    'FE':'E','NCE':'E','NDE':'E',
}
_BANCOS = [
    ('Santander Rio','0720131420000001149872','131','11498/7'),
    ('Prov. Bs. As.','0140004501400404115211','4004','041152/1'),
    ('HSBC','1500607500060732055732','607','607-3-205573'),
    ('Galicia','0070154520000005006724','154','5006-7-154/2'),
]
_CVU = 'E-mail: marketing@microbellsa.com.ar\nCVU: 000000310000004756934965'

def _fmt_num(numero: str):
    """4400014918.0 → ('0044','00014918','0044-00014918')"""
    try:
        n = int(float(numero))
        pv  = n // 100_000_000
        seq = n %  100_000_000
        return f'{pv:04d}', f'{seq:08d}', f'{pv:04d}-{seq:08d}'
    except Exception:
        return '0001', numero, numero

def _pesos(v): return f'$ {float(v or 0):,.2f}'.replace(',','X').replace('.',',').replace('X','.')

@app.get("/comprobantes/{tipo}/{numero}/pdf")
def comprobante_pdf_route(tipo: str, numero: str):
    from reportlab.platypus import KeepTogether
    from reportlab.platypus.flowables import Flowable

    class BottomSpacer(Flowable):
        """Empuja el contenido siguiente al pie de página."""
        def __init__(self, footer_h):
            Flowable.__init__(self)
            self._fh = footer_h
        def wrap(self, availWidth, availHeight):
            return availWidth, max(0, availHeight - self._fh)
        def draw(self):
            pass
    from reportlab.lib.utils import ImageReader
    DB_PROD     = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'

    # ── 1. Buscar cabeza ───────────────────────────────────────────────────────
    cab = None; items = []; found_db = None; is_mlt = False
    for db_path, mlt in [(DATABASE,False),(DB_PROD,False),(DATABASE_MLT,True),(DB_MLT_PROD,True)]:
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            cur.execute('SELECT * FROM "CABEZACOMPROBANTES" WHERE TIPOCOMPROBANTE=? AND NUMEROCOMPROBANTE=?',(tipo,numero))
            row = cur.fetchone()
            if row:
                cols = [d[0] for d in cur.description]
                cab = {k: (v.strip() if isinstance(v,str) else v) for k,v in zip(cols,row)}
                cur.execute(
                    'SELECT LINEA,CODIGOARTICULO,DESCRIPCION,CANTIDAD,'
                    'PRECIOUNITARIO,DESCUENTO,PORCENTAJEIVA,PRECIOTOTAL '
                    'FROM "CUERPOCOMPROBANTES" '
                    'WHERE TIPOCOMPROBANTE=? AND NUMEROCOMPROBANTE=? ORDER BY LINEA',
                    (tipo, numero)
                )
                items = cur.fetchall()
                found_db = db_path; is_mlt = mlt
                c.close(); break
            c.close()
        except Exception:
            pass
    if not cab:
        raise HTTPException(404, f"{tipo} {numero} no encontrado")

    # ── 2. CAE (solo Línea 1) ──────────────────────────────────────────────────
    cae = None; vto_cae = None
    if not is_mlt:
        for db_path in [found_db, DATABASE, DB_PROD]:
            try:
                c = conn('WIN1252', db=db_path)
                cur = c.cursor()
                cur.execute('SELECT CAE,VENCIMIENTOCAE FROM "CAEAFIP" WHERE TIPOCOMPROBANTE=? AND NUMEROCOMPROBANTE=?',(tipo,numero))
                r = cur.fetchone(); c.close()
                if r and r[0]:
                    cae = str(r[0]).strip()
                    vto_cae = r[1]; break
            except Exception:
                pass

    # ── 2b. Despachos (solo Línea 1) ──────────────────────────────────────────
    # Busca despachos por los códigos de artículo que componen la factura
    # usando STOCKXDESPACHO (CODIGOARTICULO → DESPACHO → ADUANA)
    despachos_lista = []  # lista de "nro - aduana" únicos
    if not is_mlt and items:
        codigos_articulos = list({str(it[1] or '').strip() for it in items if it[1]})
        if codigos_articulos:
            for db_path in [found_db, DATABASE, DB_PROD]:
                try:
                    c_d = conn('WIN1252', db=db_path)
                    cur_d = c_d.cursor()
                    placeholders = ','.join(['?' for _ in codigos_articulos])
                    # MAX(DESPACHO) por artículo = despacho más reciente de importación
                    cur_d.execute(
                        f'SELECT MAX(s.DESPACHO), MAX(s.ADUANA) '
                        f'FROM "STOCKXDESPACHO" s '
                        f'WHERE s.CODIGOARTICULO IN ({placeholders}) '
                        f"AND TRIM(s.DESPACHO) <> '' "
                        f'GROUP BY s.CODIGOARTICULO',
                        codigos_articulos
                    )
                    for r in cur_d.fetchall():
                        nrd    = str(r[0] or '').strip()
                        aduana = str(r[1] or '').strip()
                        # descartar despachos vacíos o con solo ceros
                        if not nrd or not nrd.replace('0','').strip():
                            continue
                        if not aduana or aduana == '-':
                            aduana = 'Aduana de Buenos Aires'
                        entry = f'{nrd} - {aduana}'
                        if entry not in despachos_lista:
                            despachos_lista.append(entry)
                    despachos_lista.sort()
                    c_d.close()
                    if despachos_lista:
                        break
                except Exception:
                    pass

    # ── 3. Datos cliente extra ─────────────────────────────────────────────────
    cli_ingbrutos = ''; cli_localidad = ''; cli_provincia = ''; cli_cp = ''
    try:
        c_cl = conn('WIN1252', db=DATABASE)
        cur_cl = c_cl.cursor()
        cur_cl.execute(
            'SELECT INGRESOSBRUTOS, LOCALIDAD, PROVINCIA, CP '
            'FROM "CLIENTES" WHERE CODIGOCLIENTE=?', (cab.get('CODIGOCLIENTE',''),)
        )
        r_cl = cur_cl.fetchone(); c_cl.close()
        if r_cl:
            cli_ingbrutos = str(r_cl[0] or '').strip()
            cli_localidad = str(r_cl[1] or '').strip()
            cli_provincia = str(r_cl[2] or '').strip()
            cli_cp        = str(r_cl[3] or '').strip()
    except Exception:
        pass

    # ── 4. Vendedor / asistente ────────────────────────────────────────────────
    cod_vend = str(cab.get('CODIGOUSUARIO','') or '').strip()
    cod_asist= str(cab.get('CODIGOUSUARIO2','') or '').strip()

    # ── 5. Condición de venta ──────────────────────────────────────────────────
    cond_venta = ''
    try:
        c_mp = conn('WIN1252', db=DATABASE)
        cur_mp = c_mp.cursor()
        cur_mp.execute('SELECT DESCRIPCION FROM "MULTIPLAZOS" WHERE CODIGOMULTIPLAZO=?',(cab.get('CODIGOMULTIPLAZO',''),))
        r_mp = cur_mp.fetchone(); c_mp.close()
        if r_mp: cond_venta = str(r_mp[0] or '').strip()
    except Exception:
        pass

    pv_str, seq_str, num_display = _fmt_num(numero)

    # ── 6. Generar PDF ─────────────────────────────────────────────────────────
    buf = BytesIO()
    W, H = A4
    m = 14*mm
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=m, rightMargin=m,
                            topMargin=10*mm, bottomMargin=10*mm)

    AZUL    = colors.HexColor('#1a3a5c')
    CELESTE = colors.HexColor('#4A90D9')   # azul claro para tablas bancarias
    GRIS    = colors.HexColor('#f4f6f9')
    NEGRO   = colors.black
    ROJO    = colors.HexColor('#cc0000')

    base   = getSampleStyleSheet()
    sN     = ParagraphStyle('sN',  fontSize=7,  leading=9,  fontName='Helvetica')
    sNb    = ParagraphStyle('sNb', fontSize=7,  leading=9,  fontName='Helvetica-Bold')
    sNc    = ParagraphStyle('sNc', fontSize=7,  leading=9,  fontName='Helvetica', alignment=TA_CENTER)
    sNr    = ParagraphStyle('sNr', fontSize=7,  leading=9,  fontName='Helvetica', alignment=TA_RIGHT)
    sT     = ParagraphStyle('sT',  fontSize=9,  leading=11, fontName='Helvetica-Bold')
    sTc    = ParagraphStyle('sTc', fontSize=9,  leading=11, fontName='Helvetica-Bold', alignment=TA_CENTER)
    sCAE   = ParagraphStyle('sCAE',fontSize=8,  leading=10, fontName='Helvetica-Bold')
    sCAEv  = ParagraphStyle('sCAEv',fontSize=8, leading=10, fontName='Helvetica')

    cw = W - 2*m  # ancho útil

    story = []

    if is_mlt:
        # ══════════════════════════════════════════════════════════════════════
        # PROFORMA REMITO (SW / DATABASE_MLT)
        # ══════════════════════════════════════════════════════════════════════
        fecha_str = ''
        if cab.get('FECHACOMPROBANTE'):
            try: fecha_str = cab['FECHACOMPROBANTE'].strftime('%d/%m/%Y')
            except Exception: fecha_str = str(cab['FECHACOMPROBANTE'])[:10]

        # Cabeza: logo | badge A | "Proforma Remito / RN-N° / FECHA"
        logo_img = Image(LOGO_PATH, width=40*mm, height=14*mm) if os.path.exists(LOGO_PATH) else Paragraph('microbell S.A.', sT)

        badge_tbl = Table([[Paragraph('<b>A</b>',ParagraphStyle('ba',fontSize=28,fontName='Helvetica-Bold',alignment=TA_CENTER))]],
                          colWidths=[18*mm], rowHeights=[20*mm])
        badge_tbl.setStyle(TableStyle([('BOX',(0,0),(-1,-1),1.5,NEGRO),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(0,0),(-1,-1),'CENTER')]))

        sRsw  = ParagraphStyle('sRsw',  fontSize=7, leading=9, fontName='Helvetica',      alignment=TA_RIGHT)
        sRbsw = ParagraphStyle('sRbsw', fontSize=7, leading=9, fontName='Helvetica-Bold', alignment=TA_RIGHT)
        right_hdr = [
            Paragraph('<b>Proforma Remito</b>',             ParagraphStyle('ph', fontSize=11,fontName='Helvetica-Bold',alignment=TA_RIGHT)),
            Paragraph('<b>DOCUMENTO NO VALIDO COMO FACTURA</b>', ParagraphStyle('phd',fontSize=7,fontName='Helvetica-Bold',textColor=ROJO,alignment=TA_RIGHT)),
            Paragraph(f'RN - N°: {num_display}',            ParagraphStyle('phn',fontSize=10,fontName='Helvetica-Bold',alignment=TA_RIGHT)),
            Paragraph(f'FECHA: {fecha_str}',                sRbsw),
            Paragraph(f'CUIT: 30-70839018-2',               sRsw),
            Paragraph(f'INGRESOS BRUTOS: CM 901-068199-0',  sRsw),
            Paragraph(f'Inicio de Actividades: 05/09/2005', sRsw),
        ]

        emp_left_sw = [
            logo_img,
            Paragraph('Dirección: AV. MONROE 5088 PISO 2', sN),
            Paragraph('C.A.B.A. CAPITAL FEDERAL C1431CAP', sN),
            Paragraph('Télefono: +54 11 3988-0024', sN),
            Paragraph('Email: info@microbellsa.com.ar  www.microbellsa.com', sN),
        ]

        hdr_tbl = Table([[emp_left_sw, badge_tbl, right_hdr]],
                        colWidths=[cw*0.35, 22*mm, cw*0.50])
        hdr_tbl.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP'),('LEFTPADDING',(0,0),(0,-1),4),('RIGHTPADDING',(0,0),(-1,-1),0)]))
        story.append(hdr_tbl)
        story.append(HRFlowable(width=cw, thickness=1.5, color=AZUL, spaceAfter=4))

        # Bloque cliente
        cod_cli  = str(cab.get('CODIGOCLIENTE','') or '').strip()
        rs_cli   = str(cab.get('RAZONSOCIAL','') or '').strip()
        dir_cli  = str(cab.get('DIRECCION','') or '').strip()
        cuit_cli = str(cab.get('CUIT','') or '').strip()
        tel_cli  = str(cab.get('TELEFONO','') or '').strip()

        cli_rows = [
            [Paragraph(f'<b>Cliente:</b> {cod_cli}  {rs_cli}', sNb), Paragraph(f'<b>CUIT:</b> {cuit_cli}', sN)],
            [Paragraph(f'<b>Dirección:</b> {dir_cli}', sN), Paragraph(f'<b>Cod.Vend.:</b> {cod_vend}', sN)],
            [Paragraph(f'<b>Localidad:</b> {cli_localidad}   <b>Provincia:</b> {cli_provincia}', sN), Paragraph(f'<b>Cod.Asist.:</b> {cod_asist}', sN)],
            [Paragraph(f'<b>Cond. de Venta:</b> {cond_venta}', sN), Spacer(1,1)],
        ]
        cli_tbl = Table(cli_rows, colWidths=[cw*0.62, cw*0.38])
        cli_tbl.setStyle(TableStyle([
            ('BOX',(0,0),(-1,-1),0.5,colors.grey),
            ('INNERGRID',(0,0),(-1,-1),0.3,colors.lightgrey),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
            ('LEFTPADDING',(0,0),(-1,-1),6),
        ]))
        story.append(cli_tbl)
        story.append(Spacer(1,4))

        # Items SW
        hdrs_sw = ['CANT','CODIGO','ARTÍCULO','DTO','PRECIO UNITARIO','TOTAL']
        cws_sw  = [14*mm, 18*mm, cw-14*mm-18*mm-18*mm-32*mm-28*mm, 18*mm, 32*mm, 28*mm]
        it_rows = [[Paragraph(h, sNb) for h in hdrs_sw]]
        total_sw = 0.0
        for it in items:
            cant  = float(it[3] or 0)
            pu    = float(it[4] or 0)
            dto   = float(it[5] or 0)
            total_item = float(it[7] or 0)
            total_sw += total_item
            obs_it = ''
            it_rows.append([
                Paragraph(str(int(cant)) if cant == int(cant) else str(cant), sNc),
                Paragraph(str(it[1] or ''), sN),
                Paragraph(str(it[2] or ''), sN),
                Paragraph(f'{dto:.2f} %', sNr),
                Paragraph(_pesos(pu), sNr),
                Paragraph(_pesos(total_item), sNr),
            ])
        it_tbl = Table(it_rows, colWidths=cws_sw, repeatRows=1)
        it_tbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),CELESTE),('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,GRIS]),
            ('BOX',(0,0),(-1,-1),0.5,colors.grey),
            ('INNERGRID',(0,0),(-1,-1),0.3,colors.lightgrey),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
            ('LEFTPADDING',(0,0),(-1,-1),4),
        ]))
        story.append(it_tbl)
        story.append(Spacer(1,6))

        # Observaciones + cuentas bancarias + total
        obs_sw = str(cab.get('COMENTARIOS','') or '').strip()
        if obs_sw:
            story.append(Paragraph(f'<b>Observaciones:</b> {obs_sw}', sN))
            story.append(Spacer(1,4))

        story.append(BottomSpacer(65*mm))   # empuja footer al pie

        # Tabla bancaria compacta (125mm) centrada
        bw = [26*mm, 46*mm, 18*mm, 26*mm]   # total ~116mm
        bk_data_sw = [
            [Paragraph('<b>CUENTAS BANCARIAS</b>', ParagraphStyle('bct',fontSize=7,fontName='Helvetica-Bold',textColor=colors.white,alignment=TA_CENTER,leading=9)),'','',''],
            [Paragraph('<b>BANCO</b>',sNb),Paragraph('<b>CBU</b>',sNb),Paragraph('<b>SUCURSAL</b>',sNb),Paragraph('<b>CTA. CTE.</b>',sNb)],
        ]
        for b,cbu,suc,cta in _BANCOS:
            bk_data_sw.append([Paragraph(b,sN),Paragraph(cbu,sN),Paragraph(suc,sNc),Paragraph(cta,sN)])
        bk_data_sw.append([Paragraph(_CVU,sN),'','',''])
        bk_tbl_sw = Table(bk_data_sw, colWidths=bw)
        bk_tbl_sw.setStyle(TableStyle([
            ('SPAN',(0,0),(-1,0)),
            ('BACKGROUND',(0,0),(-1,0),CELESTE),('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('BACKGROUND',(0,1),(-1,1),colors.HexColor('#d0e4f7')),
            ('ROWBACKGROUNDS',(0,2),(-1,-2),[colors.white,GRIS]),
            ('SPAN',(0,-1),(-1,-1)),
            ('BOX',(0,0),(-1,-1),0.5,colors.grey),('INNERGRID',(0,0),(-1,-1),0.3,colors.lightgrey),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2),
            ('LEFTPADDING',(0,0),(-1,-1),4),
        ]))

        bk_total_sw = sum(bw)
        pad_sw = (cw - bk_total_sw) / 2   # centrado

        cotiz = float(cab.get('COTIZACION',1) or 1)
        tot_str = _pesos(total_sw)
        sTotal = ParagraphStyle('sTot', fontSize=12, fontName='Helvetica-Bold', alignment=TA_RIGHT)

        footer_sw = Table([
            [Spacer(1,1), bk_tbl_sw, Spacer(1,1)],
            ['', Paragraph(f'Tipo de Cambio: $ {cotiz:.0f}', sN), ''],
            ['', Paragraph(f'<b>TOTAL:  {tot_str}</b>', sTotal), ''],
        ], colWidths=[pad_sw, bk_total_sw, pad_sw])
        footer_sw.setStyle(TableStyle([
            ('VALIGN',(0,0),(-1,-1),'TOP'),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ]))
        story.append(KeepTogether([footer_sw]))

    else:
        # ══════════════════════════════════════════════════════════════════════
        # COMPROBANTE LÍNEA 1 (FA, NCA, NDA, FCA, FE, etc.)
        # ══════════════════════════════════════════════════════════════════════
        tipo_nombre = _TIPO_NOMBRE.get(tipo, tipo)
        letra       = _TIPO_LETRA.get(tipo, 'A')

        fecha_str = ''
        if cab.get('FECHACOMPROBANTE'):
            try: fecha_str = cab['FECHACOMPROBANTE'].strftime('%d/%m/%Y')
            except Exception: fecha_str = str(cab['FECHACOMPROBANTE'])[:10]

        # Logo
        logo_img = Image(LOGO_PATH, width=38*mm, height=13*mm) if os.path.exists(LOGO_PATH) else Paragraph('<b>microbell S.A.</b>', sT)

        emp_left = [
            logo_img,
            Paragraph('Dirección: AV. MONROE 5088 PISO 2', sN),
            Paragraph('C.A.B.A. CAPITAL FEDERAL C1431CAP', sN),
            Paragraph('Télefono: +54 11 3988-0024', sN),
            Paragraph('Email: info@microbellsa.com.ar  www.microbellsa.com', sN),
        ]

        badge_cell = Table([[Paragraph(f'<b>{letra}</b>', ParagraphStyle('ltr',fontSize=30,fontName='Helvetica-Bold',alignment=TA_CENTER))]],
                            colWidths=[20*mm], rowHeights=[22*mm])
        badge_cell.setStyle(TableStyle([('BOX',(0,0),(-1,-1),2,NEGRO),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(0,0),(-1,-1),'CENTER')]))

        cod_pv_label = Paragraph(f'COD. N° {pv_str}', sNc)

        badge_full = Table([[badge_cell],[cod_pv_label]], colWidths=[22*mm])
        badge_full.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(0,0),(-1,-1),'CENTER')]))

        sR  = ParagraphStyle('sR',  fontSize=7, leading=9, fontName='Helvetica',      alignment=TA_RIGHT)
        sRb = ParagraphStyle('sRb', fontSize=7, leading=9, fontName='Helvetica-Bold', alignment=TA_RIGHT)
        emp_right = [
            Paragraph(f'<b>{tipo_nombre}</b>', ParagraphStyle('tn',fontSize=9,fontName='Helvetica-Bold',alignment=TA_RIGHT)),
            Paragraph(f'N° {num_display}',     ParagraphStyle('nn',fontSize=9,fontName='Helvetica-Bold',alignment=TA_RIGHT)),
            Paragraph(f'Fecha emisión: {fecha_str}',        sR),
            Paragraph(f'CUIT: 30-70839018-2',               sR),
            Paragraph(f'Ing. Brutos: CM 901-068199-0',       sR),
            Paragraph(f'Inic. Activ.: 05/09/2005',           sR),
            Paragraph(f'RESPONSABLE INSCRIPTO',              sRb),
        ]

        hdr_tbl = Table([[emp_left, badge_full, emp_right]],
                        colWidths=[cw*0.38, 26*mm, cw*0.50])
        hdr_tbl.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP'),('LEFTPADDING',(0,0),(0,-1),4),('RIGHTPADDING',(0,0),(-1,-1),0)]))
        story.append(hdr_tbl)
        story.append(HRFlowable(width=cw, thickness=2, color=AZUL, spaceAfter=4))

        # Bloque cliente
        cod_cli  = str(cab.get('CODIGOCLIENTE','') or '').strip()
        rs_cli   = str(cab.get('RAZONSOCIAL','') or '').strip()
        dir_cli  = str(cab.get('DIRECCION','') or '').strip()
        cuit_cli = str(cab.get('CUIT','') or '').strip()
        tipoiva  = str(cab.get('TIPOIVA','') or '').strip()
        tel_cli  = str(cab.get('TELEFONO','') or '').strip()

        localidad_full = cli_localidad
        if cli_provincia: localidad_full += f'  Provincia: {cli_provincia}'
        if cli_cp:        localidad_full += f'  C.P.: {cli_cp}'

        cli_data = [
            [Paragraph(f'<b>Cliente:</b>   {cod_cli}  {rs_cli}', sN),
             Paragraph(f'<b>C.U.I.T.:</b>  {cuit_cli}', sN)],
            [Paragraph(f'<b>Dirección:</b>  {dir_cli}', sN),
             Paragraph(f'<b>Ing. Brutos:</b>  {cli_ingbrutos}', sN)],
            [Paragraph(f'<b>Localidad:</b>  {localidad_full}', sN),
             Paragraph(f'<b>Cond. Vta.:</b>  {cond_venta}', sN)],
            [Paragraph(f'<b>I.V.A.:</b>  {tipoiva}   <b>C. P.:</b>  {cli_cp}', sN),
             Paragraph(f'<b>Cod. Vend.:</b>  {cod_vend}   <b>Cod. Asist.:</b>  {cod_asist}', sN)],
        ]
        cli_tbl = Table(cli_data, colWidths=[cw*0.55, cw*0.45])
        cli_tbl.setStyle(TableStyle([
            ('BOX',(0,0),(-1,-1),0.5,colors.grey),
            ('INNERGRID',(0,0),(-1,-1),0.3,colors.lightgrey),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
            ('LEFTPADDING',(0,0),(-1,-1),6),
        ]))
        story.append(cli_tbl)
        story.append(Spacer(1,5))

        # Items
        hdrs_l1 = ['CODIGO','DESCRIPCIÓN','CANTIDAD','P. UNITARIO','DESCUENTO','IVA','TOTAL']
        cws_l1  = [18*mm, cw-18*mm-22*mm-26*mm-22*mm-14*mm-26*mm, 22*mm, 26*mm, 22*mm, 14*mm, 26*mm]
        it_rows = [[Paragraph(h, sNb) for h in hdrs_l1]]
        subtotal = 0.0; iva21_tot = 0.0; iva105_tot = 0.0
        total_neto = 0.0
        for it in items:
            cant  = float(it[3] or 0)
            pu    = float(it[4] or 0)
            dto   = float(it[5] or 0)
            piva  = float(it[6] or 0)
            ptot  = float(it[7] or 0)
            neto_item = pu * cant * (1 - dto/100)
            total_neto += neto_item
            iva_item = neto_item * piva / 100
            if piva >= 20: iva21_tot  += iva_item
            elif piva > 0: iva105_tot += iva_item
            subtotal += neto_item
            it_rows.append([
                Paragraph(str(it[1] or ''), sNc),
                Paragraph(str(it[2] or ''), sN),
                Paragraph(f'{cant:g}', sNc),
                Paragraph(_pesos(pu), sNr),
                Paragraph(f'{dto:.2f}%', sNc),
                Paragraph(f'{piva:.2f}%', sNr),
                Paragraph(_pesos(ptot), sNr),
            ])

        it_tbl = Table(it_rows, colWidths=cws_l1, repeatRows=1)
        it_tbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),CELESTE),('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,GRIS]),
            ('BOX',(0,0),(-1,-1),0.5,colors.grey),
            ('INNERGRID',(0,0),(-1,-1),0.3,colors.lightgrey),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
            ('LEFTPADDING',(0,0),(-1,-1),4),
        ]))
        story.append(it_tbl)

        # Sección Despachos (antes del footer) — en línea separados por |
        if despachos_lista:
            story.append(Spacer(1,4))
            desp_txt = ' &nbsp;|&nbsp; '.join(despachos_lista)
            story.append(Paragraph(f'<b>Despachos:</b> {desp_txt}',
                                   ParagraphStyle('sDesp', fontSize=7, leading=9,
                                                  fontName='Helvetica', wordWrap='CJK')))

        story.append(BottomSpacer(95*mm))   # empuja footer al pie

        # Footer: cuentas bancarias centradas + totales IVA derecha
        obs   = str(cab.get('COMENTARIOS','') or '').strip()
        cotiz = float(cab.get('COTIZACION',1) or 1)
        total_cab = float(cab.get('TOTAL',0) or 0)
        iva1_cab  = float(cab.get('IVA1',0) or 0)
        iva2_cab  = float(cab.get('IVA2',0) or 0)
        gran_total = total_cab + iva1_cab + iva2_cab
        dto_total  = float(cab.get('DESCUENTOMONTO',0) or 0)

        bw_l1 = [26*mm, 46*mm, 18*mm, 26*mm]   # 116mm total — compacto, centrado
        bk_data = [[Paragraph('<b>CUENTAS BANCARIAS</b>', sTc), '', '', '']]
        bk_data.append([Paragraph('<b>BANCO</b>',sNb),Paragraph('<b>CBU</b>',sNb),Paragraph('<b>SUC.</b>',sNb),Paragraph('<b>CTA. CTE.</b>',sNb)])
        for b,cbu,suc,cta in _BANCOS:
            bk_data.append([Paragraph(b,sN),Paragraph(cbu,sN),Paragraph(suc,sNc),Paragraph(cta,sN)])
        bk_data.append([Paragraph(_CVU, sN), '', '', ''])
        bk_tbl = Table(bk_data, colWidths=bw_l1)
        bk_tbl.setStyle(TableStyle([
            ('SPAN',(0,0),(-1,0)), ('BACKGROUND',(0,0),(-1,0),CELESTE), ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('BACKGROUND',(0,1),(-1,1),colors.HexColor('#d0e4f7')),
            ('BOX',(0,0),(-1,-1),0.5,colors.grey), ('INNERGRID',(0,0),(-1,-1),0.3,colors.lightgrey),
            ('SPAN',(0,-1),(-1,-1)),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'), ('TOPPADDING',(0,0),(-1,-1),2),
            ('BOTTOMPADDING',(0,0),(-1,-1),2), ('LEFTPADDING',(0,0),(-1,-1),3),
        ]))

        sTotL1 = ParagraphStyle('sTotL1', fontSize=12, fontName='Helvetica-Bold', alignment=TA_RIGHT)
        tot_rows = [
            [Paragraph('Neto gravado:', sN), Paragraph(_pesos(total_neto), sNr)],
            [Paragraph('Descuento %:', sN),  Paragraph(_pesos(dto_total), sNr)],
            [Paragraph('Subtotal:', sN),      Paragraph(_pesos(total_cab), sNr)],
            [Paragraph('Perc:', sN),          Paragraph('$ 0,00', sNr)],
            [Paragraph('I.V.A. 21%:', sN),    Paragraph(_pesos(iva1_cab), sNr)],
            [Paragraph('I.V.A. 10,5%:', sN),  Paragraph(_pesos(iva2_cab), sNr)],
            [Paragraph('<b>TOTAL:</b>', sTotL1), Paragraph(f'<b>{_pesos(gran_total)}</b>', sTotL1)],
        ]
        tot_tbl = Table(tot_rows, colWidths=[32*mm, 40*mm])
        tot_tbl.setStyle(TableStyle([
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'), ('TOPPADDING',(0,0),(-1,-1),2), ('BOTTOMPADDING',(0,0),(-1,-1),2),
            ('LINEABOVE',(0,-1),(-1,-1),1.5,CELESTE),
        ]))

        bk_total_l1 = sum(bw_l1)         # 116mm
        tot_total_l1 = 72*mm             # 32+40
        pad_l1 = (cw - bk_total_l1 - tot_total_l1) / 2  # centra banco dejando totales a la derecha

        footer_tbl = Table(
            [[Spacer(1,1), bk_tbl, Spacer(1,4), tot_tbl]],
            colWidths=[pad_l1, bk_total_l1, 4*mm, tot_total_l1]
        )
        footer_tbl.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'BOTTOM')]))
        footer_items = [footer_tbl]
        if obs:
            footer_items.append(Spacer(1,3))
            footer_items.append(Paragraph(f'Observaciones: {obs}', sN))
        footer_items.append(Paragraph(f'Tipo de Cambio: $ {cotiz:.2f}', sN))
        footer_items.append(Spacer(1,8))
        footer_items.append(HRFlowable(width=cw, thickness=0.5, color=colors.grey, spaceAfter=6))
        if cae:
            vto_cae_str = ''
            if vto_cae:
                try: vto_cae_str = vto_cae.strftime('%d/%m/%Y')
                except Exception: vto_cae_str = str(vto_cae)[:10]

            # Construir URL QR AFIP
            import json as _json, base64 as _b64
            _TIPO_AFIP_MAP = {
                'FA':1,'NDA':2,'NCA':3,'FB':6,'NDB':7,'NCB':8,
                'FCA':201,'FCB':206,'FE':19,'NCCA':203,'NDE':12,'NCE':13,
                'FE_AFIP':19,
            }
            gran_total_cae = float(cab.get('TOTAL',0) or 0) + float(cab.get('IVA1',0) or 0) + float(cab.get('IVA2',0) or 0)
            cuit_rec_raw   = str(cab.get('CUIT','0') or '0')
            cuit_rec_num   = int(''.join(c for c in cuit_rec_raw if c.isdigit()) or '0')
            pv_int         = int(pv_str)
            seq_int        = int(seq_str)
            fecha_afip     = ''
            if cab.get('FECHACOMPROBANTE'):
                try: fecha_afip = cab['FECHACOMPROBANTE'].strftime('%Y-%m-%d')
                except Exception: fecha_afip = str(cab['FECHACOMPROBANTE'])[:10]
            qr_data = {
                "ver":1, "fecha":fecha_afip,
                "cuit":30708390182, "ptoVta":pv_int,
                "tipoCmp":_TIPO_AFIP_MAP.get(tipo,1), "nroCmp":seq_int,
                "importe":round(gran_total_cae,2), "moneda":"PES", "ctz":1.0,
                "tipoDocRec":80, "nroDocRec":cuit_rec_num,
                "tipoCodAut":"E", "codAut":int(cae)
            }
            # Base64 estándar (sin padding) — mismo formato que ARCA emite
            import urllib.parse as _uparse
            _json_bytes = _json.dumps(qr_data, separators=(',',':')).encode()
            qr_b64 = _b64.b64encode(_json_bytes).decode().rstrip('=')
            # URL-encode el parámetro para que + y / no rompan la URL
            qr_afip_url = ('https://servicioscf.afip.gob.ar/publico/comprobantes/cae.aspx?p='
                           + _uparse.quote(qr_b64, safe=''))
            # URL sitio web empresa (derecho)
            qr_web_url = 'https://www.microbellsa.com'

            # Generar imágenes QR
            qr_img_l = qr_img_r = None
            try:
                import qrcode as _qrc
                # QR izquierdo: AFIP CAE — ERROR_CORRECT_L para menor densidad y mejor escaneo
                _qr_l = _qrc.QRCode(error_correction=_qrc.constants.ERROR_CORRECT_L, box_size=6, border=2)
                _qr_l.add_data(qr_afip_url); _qr_l.make(fit=True)
                _pil_l = _qr_l.make_image(fill_color='black', back_color='white')
                _buf_l = BytesIO(); _pil_l.save(_buf_l, format='PNG'); _buf_l.seek(0)
                qr_img_l = Image(_buf_l, width=32*mm, height=32*mm)
                # QR derecho: sitio web empresa — URL corta, baja densidad
                _qr_r = _qrc.QRCode(error_correction=_qrc.constants.ERROR_CORRECT_L, box_size=6, border=2)
                _qr_r.add_data(qr_web_url); _qr_r.make(fit=True)
                _pil_r = _qr_r.make_image(fill_color='black', back_color='white')
                _buf_r = BytesIO(); _pil_r.save(_buf_r, format='PNG'); _buf_r.seek(0)
                qr_img_r = Image(_buf_r, width=32*mm, height=32*mm)
            except Exception:
                qr_img_l = Paragraph('[ QR AFIP ]', sN)
                qr_img_r = Paragraph('[ QR Web ]', sN)

            sCAEc = ParagraphStyle('sCAEc', fontSize=8, leading=11, fontName='Helvetica')
            sCAEb = ParagraphStyle('sCAEb', fontSize=8, leading=11, fontName='Helvetica-Bold')
            cae_center = [
                Paragraph(f'Factura Electrónica / CAE:', sCAEc),
                Paragraph(f'<b>{cae}</b>', sCAEb),
                Spacer(1,4),
                Paragraph(f'Fecha Vencimiento CAE:', sCAEc),
                Paragraph(f'<b>{vto_cae_str}</b>', sCAEb),
            ]
            qr_w = 34*mm
            mid_w = cw - 2*qr_w
            cae_tbl = Table([[qr_img_l, cae_center, qr_img_r]],
                            colWidths=[qr_w, mid_w, qr_w])
            cae_tbl.setStyle(TableStyle([
                ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                ('ALIGN',(0,0),(0,-1),'LEFT'),
                ('ALIGN',(2,0),(2,-1),'RIGHT'),
                ('ALIGN',(1,0),(1,-1),'LEFT'),
                ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2),
            ]))
            footer_items.append(cae_tbl)
        else:
            footer_items.append(Paragraph('Factura Electrónica / CAE: (pendiente)', sN))
        story.append(KeepTogether(footer_items))

    doc.build(story)
    buf.seek(0)
    pv_s, seq_s, _ = _fmt_num(numero)
    fname = f'{tipo}_{pv_s}-{seq_s}.pdf'
    return StreamingResponse(buf, media_type='application/pdf',
                             headers={'Content-Disposition': f'inline; filename="{fname}"'})

# ─── Debug despachos comprobante ──────────────────────────────────────────────
@app.get("/debug/despachos2/{tipo}/{numero}")
def debug_despachos2(tipo: str, numero: str):
    """Muestra contenido de DETALLEDESPACHOVENTAS para el comprobante dado,
       y los artículos del comprobante cruzados con STOCKXDESPACHO."""
    result = {}
    _DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    for db_path in [DATABASE, _DB_PROD]:
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()

            # 1. Artículos de la factura
            cur.execute(
                'SELECT LINEA,CODIGOARTICULO FROM "CUERPOCOMPROBANTES" '
                'WHERE TIPOCOMPROBANTE=? AND NUMEROCOMPROBANTE=? ORDER BY LINEA',
                (tipo, numero)
            )
            arts = [{'linea': r[0], 'articulo': str(r[1] or '').strip()} for r in cur.fetchall()]
            result['articulos'] = arts

            # 2. DETALLEDESPACHOVENTAS — raw con exactamente ese tipo/numero
            cur.execute(
                'SELECT FIRST 50 TIPOCOMPROBANTE, NUMEROCOMPROBANTE, LINEA, DESPACHO, CANTIDAD '
                'FROM "DETALLEDESPACHOVENTAS" '
                'WHERE TIPOCOMPROBANTE=? AND NUMEROCOMPROBANTE=? ORDER BY LINEA, DESPACHO',
                (tipo, numero)
            )
            rows_ddv = [{'tipo': str(r[0]).strip(), 'numero': str(r[1]).strip(),
                         'linea': r[2], 'despacho': str(r[3] or '').strip(), 'cantidad': r[4]}
                        for r in cur.fetchall()]
            result['DETALLEDESPACHOVENTAS_exacto'] = rows_ddv

            # 3. DETALLEDESPACHOVENTAS — busca con LIKE para ver si numero está en otro formato
            cur.execute(
                'SELECT FIRST 20 TIPOCOMPROBANTE, NUMEROCOMPROBANTE, LINEA, DESPACHO '
                'FROM "DETALLEDESPACHOVENTAS" '
                "WHERE TIPOCOMPROBANTE=? AND TRIM(DESPACHO)<>'' "
                "AND DESPACHO<>'000000000000000' ORDER BY NUMEROCOMPROBANTE DESC",
                (tipo,)
            )
            rows_sample = [{'tipo': str(r[0]).strip(), 'numero': str(r[1]).strip(),
                            'linea': r[2], 'despacho': str(r[3] or '').strip()}
                           for r in cur.fetchall()]
            result['DETALLEDESPACHOVENTAS_sample_mismo_tipo'] = rows_sample

            # 4. STOCKXDESPACHO por artículos de la factura
            codigos = [a['articulo'] for a in arts if a['articulo']]
            if codigos:
                ph = ','.join(['?' for _ in codigos])
                cur.execute(
                    f'SELECT CODIGOARTICULO, DESPACHO, ADUANA '
                    f'FROM "STOCKXDESPACHO" WHERE CODIGOARTICULO IN ({ph}) '
                    f"AND TRIM(DESPACHO)<>'' ORDER BY CODIGOARTICULO, DESPACHO",
                    codigos
                )
                result['STOCKXDESPACHO_por_articulo'] = [
                    {'articulo': str(r[0]).strip(), 'despacho': str(r[1] or '').strip(),
                     'aduana': str(r[2] or '').strip()}
                    for r in cur.fetchall()
                ]
            c.close()
            result['db_usado'] = db_path
            break
        except Exception as e:
            result[f'error_{db_path}'] = str(e)
    return result

# ─── Rubros / jerarquía ────────────────────────────────────────────────────────
@app.get("/debug/tablas_rubro")
def debug_tablas_rubro():
    c = conn()
    cur = c.cursor()
    # Columnas de cada tabla de jerarquía
    result = {}
    for tabla in ['GRUPOSUPERRUBROS','SUPERRUBROS','RUBROS']:
        cur.execute(f"SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME='{tabla}' ORDER BY RDB$FIELD_POSITION")
        result[f'cols_{tabla}'] = [r[0] for r in cur.fetchall()]
        try:
            cur.execute(f'SELECT FIRST 2 * FROM "{tabla}"')
            result[f'muestra_{tabla}'] = [list(r) for r in cur.fetchall()]
        except Exception as e:
            result[f'muestra_{tabla}_err'] = str(e)
    c.close()
    return result

@app.get("/gruposuperrubros")
def get_gruposuperrubros():
    c = conn()
    cur = c.cursor()
    # Solo los que tienen artículos con remanente > 0
    cur.execute("""
        SELECT DISTINCT g.CODIGOGRUPOSUPERRUBRO, g.DESCRIPCION
        FROM "GRUPOSUPERRUBROS" g
        WHERE EXISTS (
            SELECT 1 FROM "ARTICULOS" a
            JOIN "RUBROS" r ON r.CODIGORUBRO = a.CODIGORUBRO
            JOIN "SUPERRUBROS" sr ON sr.CODIGOSUPERRUBRO = r.CODIGOSUPERRUBRO
            WHERE sr.CODIGOGRUPOSUPERRUBRO = g.CODIGOGRUPOSUPERRUBRO
              AND a.ACTIVO = '1'
        )
        ORDER BY g.DESCRIPCION
    """)
    rows = cur.fetchall()
    c.close()
    return [{"codigo": r[0], "descripcion": r[1]} for r in rows]

@app.get("/superrubros")
def get_superrubros(grupo: Optional[str] = None):
    c = conn()
    cur = c.cursor()
    if grupo:
        cur.execute("""
            SELECT DISTINCT sr.CODIGOSUPERRUBRO, sr.DESCRIPCION
            FROM "SUPERRUBROS" sr
            WHERE sr.CODIGOGRUPOSUPERRUBRO = ?
              AND EXISTS (
                SELECT 1 FROM "ARTICULOS" a
                JOIN "RUBROS" r ON r.CODIGORUBRO = a.CODIGORUBRO
                WHERE r.CODIGOSUPERRUBRO = sr.CODIGOSUPERRUBRO AND a.ACTIVO = '1'
              )
            ORDER BY sr.DESCRIPCION
        """, (grupo,))
    else:
        cur.execute('SELECT CODIGOSUPERRUBRO, DESCRIPCION, CODIGOGRUPOSUPERRUBRO FROM "SUPERRUBROS" ORDER BY DESCRIPCION')
    rows = cur.fetchall()
    c.close()
    return [{"codigo": r[0], "descripcion": r[1], "grupo": r[2] if len(r)>2 else None} for r in rows]

@app.get("/rubros")
def get_rubros(superrubro: Optional[str] = None, grupo: Optional[str] = None):
    c = conn()
    cur = c.cursor()
    params = []
    filtro_sr = ""
    if superrubro:
        filtro_sr = "AND r.CODIGOSUPERRUBRO = ?"
        params.append(superrubro)
    elif grupo:
        filtro_sr = "AND sr.CODIGOGRUPOSUPERRUBRO = ?"
        params.append(grupo)
    cur.execute(f"""
        SELECT DISTINCT r.CODIGORUBRO, r.DESCRIPCION, r.CODIGOSUPERRUBRO
        FROM "RUBROS" r
        JOIN "SUPERRUBROS" sr ON sr.CODIGOSUPERRUBRO = r.CODIGOSUPERRUBRO
        WHERE EXISTS (
            SELECT 1 FROM "ARTICULOS" a
            WHERE a.CODIGORUBRO = r.CODIGORUBRO AND a.ACTIVO = '1'
        )
        {filtro_sr}
        ORDER BY r.DESCRIPCION
    """, params)
    rows = cur.fetchall()
    c.close()
    return [{"codigo": r[0], "descripcion": r[1], "superrubro": r[2]} for r in rows]

@app.get("/marcas")
def get_marcas():
    c = conn()
    cur = c.cursor()
    cur.execute("""
        SELECT DISTINCT m.CODIGOMARCA, m.DESCRIPCION
        FROM "MARCAS" m
        WHERE EXISTS (
            SELECT 1 FROM "ARTICULOS" a
            WHERE a.CODIGOMARCA = m.CODIGOMARCA AND a.ACTIVO = '1'
        )
        ORDER BY m.DESCRIPCION
    """)
    rows = cur.fetchall()
    c.close()
    return [{"codigo": r[0], "descripcion": r[1]} for r in rows]

# ─── Debug ────────────────────────────────────────────────────────────────────
@app.get("/debug/stock")
def debug_stock():
    try:
        c = conn()
        cur = c.cursor()
        resultado = {}

        # Columnas de STOCK
        cur.execute("SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME='STOCK' ORDER BY RDB$FIELD_POSITION")
        resultado["columnas_stock"] = [r[0] for r in cur.fetchall()]

        # Primeras 5 filas de STOCK (ver qué campos tiene)
        cur.execute('SELECT FIRST 5 * FROM "STOCK"')
        cols = [d[0] for d in cur.description]
        resultado["stock_muestra_cols"] = cols
        resultado["stock_muestra_rows"] = [list(r) for r in cur.fetchall()]

        # Valores distintos del campo que identifica depósito (CODIGODEPOSITO o CODIGOSUCURSAL)
        for campo in ["CODIGODEPOSITO", "CODIGOSUCURSAL", "DEPOSITO"]:
            try:
                cur.execute(f'SELECT DISTINCT "{campo}" FROM "STOCK"')
                resultado[f"distintos_{campo}"] = [r[0] for r in cur.fetchall()]
            except Exception:
                pass

        c.close()
        return resultado
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/debug/presupuesto/{numero}")
def debug_presupuesto(numero: str):
    """Muestra los campos de aprobación y remisión tal como están en la DB."""
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        cur.execute(
            'SELECT NUMEROCOMPROBANTE, FECHAAPROBADO, CODIGOUSUARIOAPROBACION, '
            'CODIGOOPERACION, ANULADA, CLASECOMPROBANTE, CODIGOMULTIPLAZO, '
            'COTIZACIONFIJA, LISTAPRECIO, NUMEROTRANSACCION, CODIGORESPONSABLE, '
            'FECHAVENCIMIENTO, CODIGOUSUARIO, CODIGOUSUARIO2 '
            'FROM "CABEZAPRESUPUESTOS" WHERE NUMEROCOMPROBANTE = ?', (numero,)
        )
        cab = cur.fetchone()
        cur.execute(
            'SELECT LINEA, CANTIDAD, CANTIDADREMITIDA '
            'FROM "CUERPOPRESUPUESTOS" WHERE NUMEROCOMPROBANTE = ? ORDER BY LINEA',
            (numero,)
        )
        items = cur.fetchall()
        # Revisar triggers sobre CABEZAPRESUPUESTOS (nombre + tipo + source)
        cur.execute(
            "SELECT TRIM(RDB$TRIGGER_NAME), RDB$TRIGGER_TYPE, RDB$TRIGGER_SOURCE "
            "FROM RDB$TRIGGERS "
            "WHERE RDB$RELATION_NAME = 'CABEZAPRESUPUESTOS' AND RDB$SYSTEM_FLAG = 0"
        )
        triggers = [{"name": r[0], "type": r[1], "source": str(r[2])[:800] if r[2] else None}
                    for r in cur.fetchall()]
        # Default del campo CODIGOUSUARIOAPROBACION
        cur.execute("""
            SELECT rf.RDB$DEFAULT_SOURCE
            FROM RDB$RELATION_FIELDS rf
            WHERE rf.RDB$RELATION_NAME = 'CABEZAPRESUPUESTOS'
              AND TRIM(rf.RDB$FIELD_NAME) = 'CODIGOUSUARIOAPROBACION'
        """)
        col_def = cur.fetchone()
        c.close()
        return {
            "cabeza": {
                "numero":                   cab[0],
                "fechaaprobado":            str(cab[1]),
                "codigousuarioaprobacion":  repr(cab[2]),
                "codigooperacion":          cab[3],
                "anulada":                  cab[4],
                "clasecomprobante":         cab[5],
                "codigomultiplazo":         cab[6],
                "cotizacionfija":           cab[7],
                "listaprecio":              cab[8],
                "numerotransaccion":        cab[9],
                "codigoresponsable":        repr(cab[10]),
                "fechavencimiento":         str(cab[11]),
                "codigousuario":            cab[12],
                "codigousuario2":           cab[13],
            } if cab else None,
            "items": [{"linea": r[0], "cantidad": r[1], "cantidadremitida": r[2]} for r in items],
            "triggers": triggers,
            "col_default_aprobacion": str(col_def[0]) if col_def and col_def[0] else None,
        }
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/debug/depositos")
def debug_depositos():
    try:
        c = conn()
        cur = c.cursor()
        resultado = {}

        for tabla in ["DEPOSITOS", "SUBDEPOSITOS", "WEB_STOCK", "STOCKPORUSUARIO", "STOCKPORUSUARIODETALLE"]:
            try:
                cur.execute(f"SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME='{tabla}' ORDER BY RDB$FIELD_POSITION")
                resultado[f"cols_{tabla}"] = [r[0] for r in cur.fetchall()]
            except Exception as ex:
                resultado[f"cols_{tabla}_error"] = str(ex)

        # Muestra los depósitos existentes
        try:
            cur.execute('SELECT FIRST 20 * FROM "DEPOSITOS"')
            cols = [d[0] for d in cur.description]
            resultado["depositos_cols"] = cols
            resultado["depositos_rows"] = [list(r) for r in cur.fetchall()]
        except Exception as ex:
            resultado["depositos_error"] = str(ex)

        # Muestra sample de WEB_STOCK
        try:
            cur.execute('SELECT FIRST 3 * FROM "WEB_STOCK"')
            cols = [d[0] for d in cur.description]
            resultado["web_stock_cols"] = cols
            resultado["web_stock_rows"] = [list(r) for r in cur.fetchall()]
        except Exception as ex:
            resultado["web_stock_error"] = str(ex)

        c.close()
        return resultado
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/debug/articulos_monedas")
def debug_articulos_monedas():
    try:
        c = conn()
        cur = c.cursor()
        resultado = {}
        # Campos de ARTICULOS que tengan CODIGO o PARTICULAR
        cur.execute("SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME='ARTICULOS' AND (RDB$FIELD_NAME CONTAINING 'PARTICULAR' OR RDB$FIELD_NAME CONTAINING 'PRECIO') ORDER BY RDB$FIELD_POSITION")
        resultado["articulos_precio_particular"] = [r[0] for r in cur.fetchall()]
        # Muestra artículo 01315 - valores raw de precio
        cur.execute(
            "SELECT CODIGOARTICULO, CODIGOPARTICULAR, CODIGOMONEDA, "
            "PRECIOLISTA1, PRECIOLISTA2, PRECIOLISTA3 "
            "FROM \"ARTICULOS\" WHERE CODIGOARTICULO = '01315' OR CODIGOPARTICULAR = '01315'"
        )
        row = cur.fetchone()
        if row:
            resultado["art_01315"] = {
                "CODIGOARTICULO": row[0], "CODIGOPARTICULAR": row[1],
                "CODIGOMONEDA": row[2],
                "PRECIOLISTA1_raw": str(row[3]), "PRECIOLISTA1_float": float(row[3]) if row[3] else None,
                "PRECIOLISTA2_raw": str(row[4]), "PRECIOLISTA3_raw": str(row[5]),
            }
        # CAMBIO raw de MONEDAS para DOLARES
        cur.execute('SELECT CODIGOMONEDA, CAMBIO, CAST(CAMBIO AS VARCHAR(30)) FROM "MONEDAS" WHERE CODIGOMONEDA = \'DOLARES\'')
        rm = cur.fetchone()
        if rm:
            resultado["dolares_cambio"] = {"raw": str(rm[1]), "str": rm[2], "float": float(rm[1]) if rm[1] else None}
            if row and row[3] and rm[1]:
                resultado["calculo"] = {
                    "preciolista1_float": float(row[3]),
                    "cambio_float": float(rm[1]),
                    "resultado_sin_round": float(row[3]) * float(rm[1]),
                    "resultado_round2": round(float(row[3]) * float(rm[1]), 2),
                    "resultado_round4": round(float(row[3]) * float(rm[1]), 4),
                }
        # Estructura y datos de MONEDAS
        cur.execute("SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME='MONEDAS' ORDER BY RDB$FIELD_POSITION")
        resultado["monedas_cols"] = [r[0] for r in cur.fetchall()]
        # Columnas de CLIENTES (para identificar campo IVA)
        cur.execute("SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME='CLIENTES' ORDER BY RDB$FIELD_POSITION")
        all_cols = [r[0] for r in cur.fetchall()]
        resultado["clientes_cols"] = all_cols
        # Busca columnas relacionadas a IVA/condición fiscal
        resultado["clientes_cols_iva"] = [c for c in all_cols if any(k in c for k in ['IVA','CONDIC','CATEG','FISCAL','RESPON'])]
        # Buscar generadores/secuencias de Firebird (para numeración PR)
        cur.execute("SELECT TRIM(RDB$GENERATOR_NAME), RDB$GENERATOR_ID FROM RDB$GENERATORS WHERE RDB$SYSTEM_FLAG=0 ORDER BY RDB$GENERATOR_NAME")
        resultado["generators"] = [{"name": r[0], "id": r[1]} for r in cur.fetchall()]
        # Tablas que podrían tener el contador (buscar tablas con columnas NUMERO+TIPO)
        cur.execute("""
            SELECT DISTINCT TRIM(a.RDB$RELATION_NAME)
            FROM RDB$RELATION_FIELDS a
            JOIN RDB$RELATION_FIELDS b ON a.RDB$RELATION_NAME=b.RDB$RELATION_NAME
            WHERE (a.RDB$FIELD_NAME CONTAINING 'NUMERO' OR a.RDB$FIELD_NAME = 'NUMERO')
              AND (b.RDB$FIELD_NAME CONTAINING 'TIPO' OR b.RDB$FIELD_NAME CONTAINING 'COMPROBANTE' OR b.RDB$FIELD_NAME CONTAINING 'DOCUMENTO')
              AND a.RDB$RELATION_NAME NOT STARTING WITH 'RDB$'
              AND a.RDB$RELATION_NAME NOT CONTAINING 'CABEZA'
              AND a.RDB$RELATION_NAME NOT CONTAINING 'CUERPO'
            ORDER BY 1
        """)
        resultado["tablas_contador_candidatas"] = [r[0] for r in cur.fetchall()]

        # Buscar triggers sobre CABEZAPRESUPUESTOS
        try:
            cur.execute("""
                SELECT TRIM(t.RDB$TRIGGER_NAME), t.RDB$TRIGGER_TYPE,
                       CAST(t.RDB$TRIGGER_SOURCE AS VARCHAR(500))
                FROM RDB$TRIGGERS t
                WHERE t.RDB$RELATION_NAME = 'CABEZAPRESUPUESTOS'
                  AND t.RDB$SYSTEM_FLAG = 0
            """)
            resultado["triggers_cabezapresupuestos"] = [
                {"name": r[0], "type": r[1], "src_snippet": str(r[2])[:300] if r[2] else ""}
                for r in cur.fetchall()
            ]
        except Exception as ex:
            resultado["triggers_err"] = str(ex)

        # Valor actual del generador AUXILIAR (podría ser el contador PR)
        try:
            cur.execute("SELECT GEN_ID(AUXILIAR, 0) FROM RDB$DATABASE")
            resultado["gen_auxiliar_value"] = cur.fetchone()[0]
        except Exception as ex:
            resultado["gen_auxiliar_err"] = str(ex)

        c.close()
        return resultado
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/setup/fechaaprobado_nullable")
def setup_fechaaprobado_nullable():
    """Ejecutar una sola vez: quita NOT NULL de FECHAAPROBADO en CABEZAPRESUPUESTOS."""
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        # Quitar NOT NULL y DEFAULT de FECHAAPROBADO
        cur.execute("""
            UPDATE RDB$RELATION_FIELDS
            SET RDB$NULL_FLAG = NULL,
                RDB$DEFAULT_VALUE = NULL,
                RDB$DEFAULT_SOURCE = NULL
            WHERE RDB$RELATION_NAME = 'CABEZAPRESUPUESTOS'
              AND RDB$FIELD_NAME = 'FECHAAPROBADO'
        """)
        c.commit()
        # Verificar
        cur.execute("""
            SELECT RDB$NULL_FLAG, RDB$DEFAULT_SOURCE
            FROM RDB$RELATION_FIELDS
            WHERE RDB$RELATION_NAME = 'CABEZAPRESUPUESTOS'
              AND RDB$FIELD_NAME = 'FECHAAPROBADO'
        """)
        row = cur.fetchone()
        c.close()
        return {"ok": True,
                "RDB$NULL_FLAG": row[0] if row else '?',
                "RDB$DEFAULT_SOURCE": row[1] if row else '?',
                "msg": "Ambos deben ser None/null para que FECHAAPROBADO quede en NULL al insertar"}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/debug/procedimientos")
def debug_procedimientos():
    try:
        c = conn()
        cur = c.cursor()
        resultado = {}
        # Procedimientos almacenados
        cur.execute("SELECT TRIM(RDB$PROCEDURE_NAME) FROM RDB$PROCEDURES WHERE RDB$SYSTEM_FLAG=0 ORDER BY RDB$PROCEDURE_NAME")
        procs = [r[0] for r in cur.fetchall()]
        resultado["procedimientos"] = [p for p in procs if any(x in p for x in ['STOCK','DEPO','REMANEN','SALDO'])]
        resultado["total_procedimientos"] = len(procs)
        # Probar si STOCK tiene datos recientes por depósito via CUERPOCOMPROBANTES
        # Artículo 02785 depósito 003
        cur.execute("""
            SELECT cc.CODIGODEPOSITO,
                   SUM(CASE WHEN cab.TIPOCOMPROBANTE IN ('FA','FE','FCA','FCE','RE','NC','NCA','SIV','FDI')
                             THEN cc.CANTIDAD ELSE 0 END) AS CANT_COMP,
                   COUNT(*) AS FILAS
            FROM "CUERPOCOMPROBANTES" cc
            JOIN "CABEZACOMPROBANTES" cab ON cab.TIPOCOMPROBANTE=cc.TIPOCOMPROBANTE AND cab.NUMEROCOMPROBANTE=cc.NUMEROCOMPROBANTE
            WHERE cc.CODIGOARTICULO='02785' AND cc.CODIGODEPOSITO IN ('001','003')
            GROUP BY cc.CODIGODEPOSITO
        """)
        cols = [d[0] for d in cur.description]
        resultado["movimientos_02785"] = [dict(zip(cols,list(r))) for r in cur.fetchall()]
        # Ver tipos de comprobante distintos en CUERPOCOMPROBANTES
        cur.execute("SELECT DISTINCT TIPOCOMPROBANTE FROM \"CUERPOCOMPROBANTES\" ORDER BY TIPOCOMPROBANTE")
        resultado["tipos_comprobante_cuerpo"] = [r[0] for r in cur.fetchall()]
        c.close()
        return resultado
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/debug/procs_stock")
def debug_procs_stock():
    try:
        c = conn()
        cur = c.cursor()
        resultado = {}
        procs_interesantes = [
            'FMA_CALCULA_STOCKREMANENTE','FMA_CALCULASTOCKREMANENTE',
            'FMA_CALCULA_STOCKREAL','FMA_CALCULASTOCKREAL',
            'FMA_DETALLESTOCK','FMA_STOCK','FMA_DEPOSITOS'
        ]
        for proc in procs_interesantes:
            try:
                cur.execute("""
                    SELECT TRIM(p.RDB$PARAMETER_NAME), p.RDB$PARAMETER_TYPE,
                           TRIM(f.RDB$FIELD_TYPE), p.RDB$PARAMETER_NUMBER
                    FROM RDB$PROCEDURE_PARAMETERS p
                    JOIN RDB$FIELDS f ON f.RDB$FIELD_NAME = p.RDB$FIELD_SOURCE
                    WHERE p.RDB$PROCEDURE_NAME = ?
                    ORDER BY p.RDB$PARAMETER_TYPE, p.RDB$PARAMETER_NUMBER
                """, (proc,))
                params = [{"nombre": r[0], "tipo": "INPUT" if r[1]==0 else "OUTPUT", "campo": r[2]} for r in cur.fetchall()]
                if params:
                    resultado[proc] = params
            except Exception as ex:
                resultado[proc + "_error"] = str(ex)
        # Intentar llamar FMA_DETALLESTOCK con artículo conocido
        for call in [
            ("FMA_DETALLESTOCK", "EXECUTE PROCEDURE \"FMA_DETALLESTOCK\" '02785'"),
            ("FMA_STOCK", "EXECUTE PROCEDURE \"FMA_STOCK\" '02785'"),
        ]:
            try:
                cur.execute(call[1])
                cols = [d[0] for d in cur.description]
                rows = cur.fetchall()
                resultado[call[0]+"_result"] = [dict(zip(cols,list(r))) for r in rows[:5]]
            except Exception as ex:
                resultado[call[0]+"_call_error"] = str(ex)
        c.close()
        return resultado
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/debug/stock_condiciones")
def debug_stock_condiciones():
    """Verifica OPERACIONES.COMPROMETESTOCK y CASILLEROS para entender por qué no descuenta stock"""
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        res = {}
        # 1. Ver COMPROMETESTOCK para cada OPERACION usada en NPs
        cur.execute('SELECT CODIGOOPERACION, DESCRIPCION, COMPROMETESTOCK FROM "OPERACIONES" ORDER BY CODIGOOPERACION')
        res['operaciones'] = [{'codigo': str(r[0]), 'descripcion': str(r[1]).strip(), 'comprometestock': r[2]} for r in cur.fetchall()]
        # 2. Ver si existen registros en CASILLEROS para artículo 00590 (el de la antena)
        cur.execute('SELECT CODIGOARTICULO, LOTE, CODIGODEPOSITO FROM "CASILLEROS" WHERE CODIGOARTICULO = ? ORDER BY LOTE', ('00590',))
        rows = cur.fetchall()
        res['casilleros_00590'] = [{'art': r[0], 'lote': str(r[1]), 'deposito': str(r[2])} for r in rows]
        # 3. Ver qué LOTE usamos en CUERPOPEDIDOS para ese artículo
        cur.execute("SELECT FIRST 3 LOTE, CODIGODEPOSITO, CANTIDAD, CANTIDADREMITIDA FROM \"CUERPOPEDIDOS\" WHERE CODIGOARTICULO = '00590' ORDER BY NUMEROCOMPROBANTE DESC")
        res['cuerpopedidos_00590'] = [{'lote': str(r[0]), 'deposito': str(r[1]), 'cantidad': r[2], 'remitida': r[3]} for r in cur.fetchall()]
        # 4. Ver cuántos casilleros existen en total (para saber si la tabla está poblada)
        cur.execute('SELECT COUNT(*) FROM "CASILLEROS"')
        res['total_casilleros'] = cur.fetchone()[0]
        c.close()
        return res
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/debug/stock_profundo")
def debug_stock_profundo():
    """Fuente de FMA_CALCULA_STOCKPEDIDO + artículo real en CUERPOPEDIDOS + CASILLEROS sample"""
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        res = {}
        # 1. Fuente de FMA_CALCULA_STOCKPEDIDO (sin el filtro EXISTS)
        for sp in ['FMA_CALCULA_STOCKPEDIDO', 'FMA_STOCK']:
            cur.execute('SELECT RDB$PROCEDURE_SOURCE FROM RDB$PROCEDURES WHERE TRIM(RDB$PROCEDURE_NAME)=?', (sp,))
            row = cur.fetchone()
            res[f'src_{sp}'] = str(row[0])[:2000] if row and row[0] else 'NO ENCONTRADO'
        # 2. Buscar el artículo 00590 en ARTICULOS para ver su CODIGOARTICULO interno
        cur.execute("SELECT CODIGOARTICULO, CODIGOPARTICULAR, DESCRIPCION FROM \"ARTICULOS\" WHERE CODIGOPARTICULAR='00590' OR CODIGOARTICULO='00590'")
        rows = cur.fetchall()
        res['articulos_00590'] = [{'interno': r[0], 'particular': r[1], 'desc': str(r[2])[:40]} for r in rows]
        # 3. CUERPOPEDIDOS de los últimos 5 pedidos NP (cualquier artículo)
        cur.execute("SELECT FIRST 5 NUMEROCOMPROBANTE, CODIGOARTICULO, LOTE, CANTIDAD, CODIGODEPOSITO FROM \"CUERPOPEDIDOS\" WHERE TIPOCOMPROBANTE='NP' ORDER BY NUMEROCOMPROBANTE DESC")
        res['cuerpopedidos_ultimos'] = [{'num': str(r[0]), 'art': r[1], 'lote': str(r[2]), 'cant': r[3], 'dep': str(r[4])} for r in cur.fetchall()]
        # 4. Sample de CASILLEROS para ver qué lote usan
        cur.execute("SELECT FIRST 5 CODIGOARTICULO, LOTE, CODIGODEPOSITO FROM \"CASILLEROS\"")
        res['casilleros_sample'] = [{'art': r[0], 'lote': str(r[1]), 'dep': str(r[2])} for r in cur.fetchall()]
        c.close()
        return res
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/debug/sp_remanente")
def debug_sp_remanente():
    """Lee el código fuente de FMA_CALCULASTOCKREMANENTE para ver qué tablas/condiciones usa"""
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        res = {}
        for sp in ['FMA_CALCULASTOCKREMANENTE', 'FMA_CALCULA_STOCKREMANENTE']:
            try:
                cur.execute("""
                    SELECT TRIM(RDB$PROCEDURE_NAME), RDB$PROCEDURE_SOURCE
                    FROM RDB$PROCEDURES
                    WHERE TRIM(RDB$PROCEDURE_NAME) = ?
                """, (sp,))
                row = cur.fetchone()
                res[sp] = str(row[1]) if row and row[1] else 'NO ENCONTRADO'
            except Exception as e2:
                res[sp] = str(e2)
        c.close()
        return res
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/debug/fma_stock")
def debug_fma_stock():
    try:
        c = conn()
        cur = c.cursor()
        resultado = {}
        # Test FMA_STOCK con filtro de depósitos 001 y 003
        for depositos in ["001,003", "001|003", "001;003", "001 003"]:
            try:
                cur.execute(f"""
                    SELECT FIRST 3 ID_ARTICULO, CODIGO_PRODUCTO, STOCKREAL, STOCKREMANENTE
                    FROM "FMA_STOCK"(NULL, NULL, '{depositos}', 1, 1)
                    WHERE STOCKREAL > 0
                """)
                cols = [d[0] for d in cur.description]
                rows = [dict(zip(cols, list(r))) for r in cur.fetchall()]
                resultado[f"FMA_STOCK_depositos_{depositos}"] = rows
                break
            except Exception as ex:
                resultado[f"FMA_STOCK_depositos_{depositos}_error"] = str(ex)
        # Test FMA_CALCULASTOCKREMANENTE para artículo 02785
        for depo in ['001', '003']:
            try:
                cur.execute('EXECUTE PROCEDURE "FMA_CALCULASTOCKREMANENTE" ?, ?, ?', ('02785', '000', depo))
                row = cur.fetchone()
                resultado[f"remanente_02785_dep{depo}"] = float(row[0]) if row else None
            except Exception as ex:
                resultado[f"remanente_02785_dep{depo}_error"] = str(ex)
        # Test FMA_CALCULA_STOCKREAL para artículo 02785
        for depo in ['001', '003']:
            try:
                cur.execute('EXECUTE PROCEDURE "FMA_CALCULA_STOCKREAL" ?, ?, ?', ('02785', '000', depo))
                row = cur.fetchone()
                resultado[f"stockreal_02785_dep{depo}"] = float(row[0]) if row else None
            except Exception as ex:
                resultado[f"stockreal_02785_dep{depo}_error"] = str(ex)
        c.close()
        return resultado
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/debug/cta/{codigo}")
def debug_cta(codigo: str):
    resultado = {}
    # 1. Lookup en CLIENTES
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        cur.execute('SELECT CODIGOCLIENTE, CODIGOPARTICULAR, RAZONSOCIAL FROM "CLIENTES" WHERE CODIGOCLIENTE = ?', (codigo,))
        row = cur.fetchone()
        c.close()
        resultado['clientes'] = {"codigocliente": row[0], "codigoparticular": row[1], "razonsocial": row[2]} if row else None
    except Exception as e:
        resultado['clientes_error'] = str(e)

    DATABASE_MLT = 'c:/flexxus/db/DB-MLT-Microbell.gdb'
    # 2. BD principal - todos los comprobantes del cliente (sin filtro de saldo)
    for db_key, db_path in [('db_main', DATABASE), ('db_est', DATABASE_EST), ('db_mlt', DATABASE_MLT)]:
        try:
            c = conn('LATIN1', db=db_path)
            cur = c.cursor()
            # Buscar por codigo y por codigoparticular
            cp = resultado.get('clientes', {}) or {}
            codigos = list({codigo, cp.get('codigoparticular','')})
            codigos = [x for x in codigos if x]
            placeholders = ','.join(['?' for _ in codigos])
            cur.execute(f"""
                SELECT TIPOCOMPROBANTE, NUMEROCOMPROBANTE, FECHACOMPROBANTE,
                       CODIGOCLIENTE, TOTAL, IVA1, IVA2, PAGADO, CUENTACORRIENTE, ANULADA
                FROM "CABEZACOMPROBANTES"
                WHERE CODIGOCLIENTE IN ({placeholders})
                  AND TIPOCOMPROBANTE IN ('FA','FB','FE','FCA','FCB','DI','SIV')
                ORDER BY FECHACOMPROBANTE DESC
            """, tuple(codigos))
            rows = cur.fetchall()
            c.close()
            resultado[db_key] = [{"tipo":r[0],"numero":r[1],"fecha":str(r[2]),"cod_cli":r[3],
                                   "total":float(r[4] or 0),"iva1":float(r[5] or 0),"iva2":float(r[6] or 0),
                                   "pagado":float(r[7] or 0),"ctacte":r[8],"anulada":r[9]} for r in rows]
        except Exception as e:
            # Si falla, listar tablas disponibles en esa BD
            resultado[db_key+'_error'] = str(e)
            try:
                c2 = conn('LATIN1', db=db_path)
                cur2 = c2.cursor()
                cur2.execute("SELECT TRIM(RDB$RELATION_NAME) FROM RDB$RELATIONS WHERE RDB$SYSTEM_FLAG = 0 ORDER BY RDB$RELATION_NAME")
                tablas = [r[0] for r in cur2.fetchall()]
                c2.close()
                resultado[db_key+'_tablas'] = tablas
            except Exception as e2:
                resultado[db_key+'_tablas_error'] = str(e2)
    return resultado

@app.get("/debug/esquema_docs")
def debug_esquema_docs():
    """Inspecciona columnas, NOT NULL, defaults y generators de pedidos/presupuestos."""
    try:
        c = conn('LATIN1')
        cur = c.cursor()
        resultado = {}

        # Columnas + NOT NULL + DEFAULT de cada tabla
        for tabla in ['CABEZAPEDIDOS', 'CUERPOPEDIDOS', 'CABEZAPRESUPUESTOS', 'CUERPOPRESUPUESTOS']:
            try:
                cur.execute(f"""
                    SELECT TRIM(rf.RDB$FIELD_NAME),
                           rf.RDB$NULL_FLAG,
                           TRIM(rf.RDB$DEFAULT_SOURCE)
                    FROM RDB$RELATION_FIELDS rf
                    WHERE rf.RDB$RELATION_NAME = '{tabla}'
                    ORDER BY rf.RDB$FIELD_POSITION
                """)
                resultado[f'cols_{tabla}'] = [
                    {"campo": r[0], "not_null": r[1]==1, "default": r[2]}
                    for r in cur.fetchall()
                ]
            except Exception as ex:
                resultado[f'cols_{tabla}_error'] = str(ex)

        # Muestra fila real de CABEZAPEDIDOS (para ver qué trae Flexxus)
        for tabla in ['CABEZAPEDIDOS', 'CABEZAPRESUPUESTOS']:
            try:
                cur.execute(f'SELECT FIRST 1 * FROM "{tabla}" ORDER BY NUMEROCOMPROBANTE DESC')
                cols = [d[0] for d in cur.description]
                row = cur.fetchone()
                resultado[f'muestra_{tabla}'] = dict(zip(cols, [str(v) if v is not None else None for v in row])) if row else None
            except Exception as ex:
                resultado[f'muestra_{tabla}_error'] = str(ex)

        # Generators relacionados con pedidos/presupuestos
        try:
            cur.execute("""
                SELECT TRIM(RDB$GENERATOR_NAME), RDB$GENERATOR_ID
                FROM RDB$GENERATORS
                WHERE RDB$SYSTEM_FLAG = 0
                ORDER BY RDB$GENERATOR_NAME
            """)
            todos = [{"nombre": r[0], "id": r[1]} for r in cur.fetchall()]
            resultado['generators'] = [g for g in todos if any(x in g['nombre'].upper() for x in ['PED','PRE','COMP','DOC','NUMERO','NP'])]
            resultado['generators_todos'] = todos
        except Exception as ex:
            resultado['generators_error'] = str(ex)

        # Triggers en CABEZAPEDIDOS (puede haber lógica de numeración)
        try:
            cur.execute("""
                SELECT TRIM(RDB$TRIGGER_NAME), RDB$TRIGGER_TYPE, TRIM(RDB$TRIGGER_SOURCE)
                FROM RDB$TRIGGERS
                WHERE RDB$RELATION_NAME IN ('CABEZAPEDIDOS','CABEZAPRESUPUESTOS')
                  AND RDB$SYSTEM_FLAG = 0
                ORDER BY RDB$TRIGGER_NAME
            """)
            resultado['triggers'] = [{"nombre": r[0], "tipo": r[1], "fuente": r[2]} for r in cur.fetchall()]
        except Exception as ex:
            resultado['triggers_error'] = str(ex)

        c.close()
        return resultado
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/debug/esquema_pedidos")
def debug_esquema_pedidos():
    """Columnas CABEZAPEDIDOS + valores OPERACION + muestra filas NP recientes"""
    res = {}
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        # Columnas de CABEZAPEDIDOS
        cur.execute("""
            SELECT TRIM(rf.RDB$FIELD_NAME), rf.RDB$FIELD_POSITION,
                   f.RDB$NULL_FLAG, f.RDB$DEFAULT_VALUE
            FROM RDB$RELATION_FIELDS rf
            JOIN RDB$FIELDS f ON rf.RDB$FIELD_SOURCE = f.RDB$FIELD_NAME
            WHERE rf.RDB$RELATION_NAME = 'CABEZAPEDIDOS'
            ORDER BY rf.RDB$FIELD_POSITION
        """)
        res['cols_cabezapedidos'] = [r[0] for r in cur.fetchall()]
        # Valores distintos de OPERACION en NPs
        cur.execute("SELECT DISTINCT OPERACION, COUNT(*) FROM \"CABEZAPEDIDOS\" WHERE TIPOCOMPROBANTE='NP' GROUP BY OPERACION ORDER BY OPERACION")
        res['operacion_values'] = [{'operacion': str(r[0]), 'count': r[1]} for r in cur.fetchall()]
        # Últimos 3 NPs con campos clave
        try:
            cur.execute("""
                SELECT FIRST 3 NUMEROCOMPROBANTE, OPERACION, FECHATERMINADA,
                       CODIGOUSUARIO, CODIGOUSUARIO2, CODIGOTECNICO, CLASECOMPROBANTE
                FROM "CABEZAPEDIDOS" WHERE TIPOCOMPROBANTE='NP' AND OPERACION IN ('2','4')
                ORDER BY NUMEROCOMPROBANTE DESC
            """)
            res['sample_np_terminados'] = [
                {'num': str(r[0]), 'operacion': str(r[1]), 'fechaterminada': str(r[2]),
                 'usuario': str(r[3]), 'usuario2': str(r[4]), 'tecnico': str(r[5]), 'clase': str(r[6])}
                for r in cur.fetchall()
            ]
        except Exception as e2:
            res['sample_np_extended_error'] = str(e2)
        c.close()
    except Exception as e:
        res['error'] = str(e)
    return res

@app.get("/debug/comparar_np/{num_a}/{num_b}")
def debug_comparar_np(num_a: str, num_b: str):
    """Compara campo a campo dos NPs en CABEZAPEDIDOS."""
    res = {}
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        cur.execute("SELECT * FROM \"CABEZAPEDIDOS\" WHERE NUMEROCOMPROBANTE IN (?,?) AND TIPOCOMPROBANTE='NP'", (num_a, num_b))
        cols = [d[0] for d in cur.description]
        rows = {str(r[cols.index('NUMEROCOMPROBANTE')]): dict(zip(cols, [str(v) if v is not None else None for v in r])) for r in cur.fetchall()}
        res['cabeza'] = rows

        # Diferencias entre los dos
        if num_a in rows and num_b in rows:
            diffs = {}
            for col in cols:
                va = rows[num_a].get(col)
                vb = rows[num_b].get(col)
                if va != vb:
                    diffs[col] = {num_a: va, num_b: vb}
            res['diferencias'] = diffs

        # Triggers sobre CABEZAPEDIDOS
        cur.execute("""
            SELECT TRIM(t.RDB$TRIGGER_NAME), t.RDB$TRIGGER_TYPE, t.RDB$TRIGGER_SOURCE
            FROM RDB$TRIGGERS t
            WHERE t.RDB$RELATION_NAME = 'CABEZAPEDIDOS' AND t.RDB$SYSTEM_FLAG = 0
            ORDER BY t.RDB$TRIGGER_SEQUENCE
        """)
        res['triggers'] = [{'nombre': r[0], 'tipo': r[1], 'source': (r[2] or '')[:300]} for r in cur.fetchall()]
        c.close()
    except Exception as e:
        res['error'] = str(e)
    return res

@app.get("/debug/operacion_np")
def debug_operacion_np():
    """Muestra OPERACION de los últimos 10 NPs y todos los valores distintos existentes."""
    res = {}
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        # Últimos 10 NPs: número, operacion, fechacomprobante
        cur.execute("""
            SELECT FIRST 10 NUMEROCOMPROBANTE, OPERACION, FECHACOMPROBANTE, CODIGOUSUARIO
            FROM "CABEZAPEDIDOS" WHERE TIPOCOMPROBANTE='NP'
            ORDER BY CAST(NUMEROCOMPROBANTE AS INTEGER) DESC
        """)
        res['ultimos_10'] = [{'num': str(r[0]), 'operacion': str(r[1]).strip(),
                               'fecha': str(r[2])[:10], 'usuario': str(r[3]).strip()}
                              for r in cur.fetchall()]
        # Todos los valores de OPERACION con conteo
        cur.execute("SELECT DISTINCT OPERACION, COUNT(*) FROM \"CABEZAPEDIDOS\" WHERE TIPOCOMPROBANTE='NP' GROUP BY OPERACION ORDER BY OPERACION")
        res['operacion_dist'] = [{'valor': repr(r[0]), 'count': r[1]} for r in cur.fetchall()]
        c.close()
    except Exception as e:
        res['error'] = str(e)
    return res

@app.get("/debug/comparar_pedidos")
def debug_comparar_pedidos():
    """Compara todos los campos de dos NPs: 100023473 (Flexxus) vs 100023547 (APP)"""
    res = {}
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        for num in ['100023473', '100023547']:
            cur.execute('SELECT * FROM "CABEZAPEDIDOS" WHERE TIPOCOMPROBANTE=\'NP\' AND NUMEROCOMPROBANTE=?', (num,))
            row = cur.fetchone()
            if row:
                cols = [d[0] for d in cur.description]
                res[num] = {cols[i]: str(row[i]) if row[i] is not None else None for i in range(len(cols))}
            else:
                res[num] = 'NO ENCONTRADO'
        c.close()
    except Exception as e:
        res['error'] = str(e)
    return res

@app.get("/debug/esquema_mlt")
def debug_esquema_mlt():
    """Columnas de CABEZACOMPROBANTES y CUERPOCOMPROBANTES en DB-MLT"""
    try:
        c = conn('LATIN1', db=DATABASE_MLT)
        cur = c.cursor()
        resultado = {}
        for tabla in ['CABEZACOMPROBANTES', 'CUERPOCOMPROBANTES']:
            cur.execute("""
                SELECT TRIM(f.RDB$FIELD_NAME),
                       TRIM(tp.RDB$TYPE_NAME),
                       f.RDB$NULL_FLAG,
                       f.RDB$DEFAULT_SOURCE
                FROM RDB$RELATION_FIELDS f
                JOIN RDB$FIELDS ff ON ff.RDB$FIELD_NAME = f.RDB$FIELD_SOURCE
                LEFT JOIN RDB$TYPES tp ON tp.RDB$TYPE = ff.RDB$FIELD_TYPE AND tp.RDB$FIELD_NAME = 'RDB$FIELD_TYPE'
                WHERE f.RDB$RELATION_NAME = ?
                ORDER BY f.RDB$FIELD_POSITION
            """, (tabla,))
            resultado[tabla] = [{"col": r[0], "tipo": r[1], "notnull": r[2], "default": r[3]} for r in cur.fetchall()]
        # Muestra de 1 fila para ver valores reales
        for tabla in ['CABEZACOMPROBANTES', 'CUERPOCOMPROBANTES']:
            try:
                cur.execute(f'SELECT FIRST 1 * FROM "{tabla}"')
                cols = [d[0] for d in cur.description]
                row = cur.fetchone()
                resultado[tabla+'_muestra'] = dict(zip(cols, [str(v) for v in row])) if row else {}
            except Exception as ex:
                resultado[tabla+'_muestra_err'] = str(ex)
        c.close()
        return resultado
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/mlt_cab/{numero}")
def debug_mlt_cab(numero: str):
    """Ver CABEZACOMPROBANTES de DATABASE_MLT para un número de comprobante."""
    try:
        c = conn('LATIN1', DATABASE_MLT)
        cur = c.cursor()
        cur.execute("SELECT FIRST 1 * FROM \"CABEZACOMPROBANTES\" "
                    "WHERE NUMEROCOMPROBANTE = ?", (numero,))
        row = cur.fetchone()
        cols = [d[0] for d in cur.description]
        c.close()
        if row:
            return dict(zip(cols, [str(v) for v in row]))
        return {"error": "no encontrado", "numero": numero}
    except Exception as e:
        return {"error": str(e)}


@app.get("/debug/mlt_tablas")
def debug_mlt_tablas():
    """Lista todas las tablas de DATABASE_MLT."""
    try:
        c = conn('LATIN1', DATABASE_MLT)
        cur = c.cursor()
        cur.execute("SELECT TRIM(RDB$RELATION_NAME) FROM RDB$RELATIONS "
                    "WHERE RDB$SYSTEM_FLAG = 0 ORDER BY RDB$RELATION_NAME")
        tablas = [r[0] for r in cur.fetchall()]
        c.close()
        return {"tablas": tablas}
    except Exception as e:
        return {"error": str(e)}


@app.get("/debug/quevendi_mlt/{cliente}")
def debug_quevendi_mlt(cliente: str):
    """Diagnostica por qué que_vendi no trae resultados de DATABASE_MLT para un cliente."""
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    result = {"cliente": cliente, "pasos": []}

    # 1. Lookup en DB-MLT-Microbell.gdb (prod MLT tiene CLIENTES, DATABASE_MLT no)
    cod_mlt = cliente
    try:
        c = conn('WIN1252', DB_MLT_PROD)
        cur = c.cursor()
        cur.execute('SELECT CODIGOCLIENTE, CODIGOPARTICULAR FROM "CLIENTES" '
                    'WHERE CODIGOCLIENTE = ? OR CODIGOPARTICULAR = ?', (cliente, cliente))
        row = cur.fetchone()
        c.close()
        result["pasos"].append({"paso": "lookup_clientes_mlt_prod", "encontrado": row is not None,
                                "fila": [str(x) for x in row] if row else None})
        if row:
            cod_mlt = str(row[0]).strip()
    except Exception as e:
        result["pasos"].append({"paso": "lookup_clientes_mlt_prod", "error": str(e)})

    result["codigocliente_mlt"] = cod_mlt

    # 2. Verificar columnas de CUERPOCOMPROBANTES en DATABASE_MLT
    try:
        c = conn('LATIN1', DATABASE_MLT)
        cur = c.cursor()
        cur.execute("SELECT FIRST 1 * FROM \"CUERPOCOMPROBANTES\"")
        cols = [d[0] for d in cur.description]
        result["cuerpocomprobantes_cols"] = cols
        result["tiene_codigoparticular"] = "CODIGOPARTICULAR" in cols

        # 3. Buscar comprobantes del cliente
        cur.execute(
            'SELECT FIRST 5 cb.TIPOCOMPROBANTE, cb.NUMEROCOMPROBANTE, cb.FECHACOMPROBANTE '
            'FROM "CABEZACOMPROBANTES" cb '
            "WHERE cb.CODIGOCLIENTE = ? AND cb.ANULADA = '0' "
            "AND cb.TIPOCOMPROBANTE IN ('FA','FB','FE','FCA','FCB','FCE','FCCA','FCCB','FCCE','NCA','NCB','NCCA','NCCB') "
            'ORDER BY cb.FECHACOMPROBANTE DESC',
            (cod_mlt,)
        )
        rows_cab = cur.fetchall()
        result["cabeza_count"] = len(rows_cab)
        result["cabeza_sample"] = [[str(x) for x in r] for r in rows_cab]

        # 4. Si tiene CODIGOPARTICULAR: probar el SQL completo
        if "CODIGOPARTICULAR" in cols and rows_cab:
            try:
                cur.execute(
                    'SELECT FIRST 3 '
                    'COALESCE(NULLIF(TRIM(cu.CODIGOPARTICULAR),\'\'), TRIM(cu.CODIGOARTICULO)) AS COD_ART, '
                    'cu.DESCRIPCION, cb.TIPOCOMPROBANTE, cb.NUMEROCOMPROBANTE, '
                    'CAST(cu.CANTIDAD AS DOUBLE PRECISION) '
                    'FROM "CUERPOCOMPROBANTES" cu '
                    'JOIN "CABEZACOMPROBANTES" cb '
                    '  ON cb.TIPOCOMPROBANTE = cu.TIPOCOMPROBANTE '
                    ' AND cb.NUMEROCOMPROBANTE = cu.NUMEROCOMPROBANTE '
                    "WHERE cb.CODIGOCLIENTE = ? AND cb.ANULADA = '0' "
                    "AND cb.TIPOCOMPROBANTE IN ('FA','FB','FE','FCA','FCB','FCE','FCCA','FCCB','FCCE','NCA','NCB','NCCA','NCCB')",
                    (cod_mlt,)
                )
                rows_cuerpo = cur.fetchall()
                result["cuerpo_count"] = len(rows_cuerpo)
                result["cuerpo_sample"] = [[str(x) for x in r] for r in rows_cuerpo]
            except Exception as e2:
                result["cuerpo_error"] = str(e2)
        elif "CODIGOPARTICULAR" not in cols:
            result["nota"] = "CODIGOPARTICULAR no existe en CUERPOCOMPROBANTES de DATABASE_MLT"
        c.close()
    except Exception as e:
        result["pasos"].append({"paso": "query_mlt", "error": str(e)})

    return result


@app.get("/debug/quevendi_prod/{cliente}")
def debug_quevendi_prod(cliente: str):
    """Busca FA 100001668 y 100001684 en las 4 BDs y diagnostica que_vendi para un cliente."""
    DB_PROD     = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    result = {"cliente": cliente}

    # 1. Resolver codigos
    codigos = set([cliente])
    for db_label, db_path in [("DATABASE", DATABASE), ("DB_PROD", DB_PROD)]:
        try:
            c = conn('WIN1252', db_path)
            cur = c.cursor()
            cur.execute('SELECT CODIGOCLIENTE, CODIGOPARTICULAR FROM "CLIENTES" '
                        'WHERE CODIGOCLIENTE = ? OR CODIGOPARTICULAR = ?', (cliente, cliente))
            row = cur.fetchone()
            c.close()
            if row:
                for v in row:
                    if v is not None and str(v).strip():
                        codigos.add(str(v).strip())
            result[f"lookup_{db_label}"] = [str(x) for x in row] if row else None
        except Exception as e:
            result[f"lookup_{db_label}_error"] = str(e)

    codigos = list(codigos)
    result["codigos"] = codigos
    ph = ','.join('?' * len(codigos))

    # 2. Buscar FA 100001668 y 100001684 en las 4 BDs
    for db_label, db_path in [("DATABASE", DATABASE), ("DATABASE_MLT", DATABASE_MLT),
                               ("DB_PROD", DB_PROD), ("DB_MLT_PROD", DB_MLT_PROD)]:
        try:
            c = conn('LATIN1', db=db_path)
            cur = c.cursor()
            # Buscar las facturas objetivo
            cur.execute(
                'SELECT TIPOCOMPROBANTE, NUMEROCOMPROBANTE, CODIGOCLIENTE, FECHACOMPROBANTE '
                'FROM "CABEZACOMPROBANTES" WHERE NUMEROCOMPROBANTE IN (100001668, 100001684)'
            )
            rows_t = cur.fetchall()
            result[f"facturas_target_{db_label}"] = [[str(x) for x in r] for r in rows_t]

            # Contar comprobantes del cliente
            cur.execute(
                f'SELECT COUNT(*) FROM "CABEZACOMPROBANTES" '
                f"WHERE CODIGOCLIENTE IN ({ph}) AND ANULADA = '0' "
                f"AND TIPOCOMPROBANTE IN ('FA','FB','FE','FCA','FCB','FCE','FCCA','FCCB','FCCE','NCA','NCB','NCCA','NCCB')",
                codigos
            )
            result[f"count_{db_label}"] = cur.fetchone()[0]

            # Contar en CUERPOCOMPROBANTES (para verificar el JOIN)
            cur.execute(
                f'SELECT COUNT(*) FROM "CUERPOCOMPROBANTES" cu '
                f'JOIN "CABEZACOMPROBANTES" cb '
                f'  ON cb.TIPOCOMPROBANTE = cu.TIPOCOMPROBANTE '
                f' AND cb.NUMEROCOMPROBANTE = cu.NUMEROCOMPROBANTE '
                f"WHERE cb.CODIGOCLIENTE IN ({ph}) AND cb.ANULADA = '0' "
                f"AND cb.TIPOCOMPROBANTE IN ('FA','FB','FE','FCA','FCB','FCE','FCCA','FCCB','FCCE','NCA','NCB','NCCA','NCCB')",
                codigos
            )
            result[f"count_con_join_{db_label}"] = cur.fetchone()[0]
            c.close()
        except Exception as e:
            result[f"error_{db_label}"] = str(e)

    return result


@app.get("/debug/qv_errors")
def debug_qv_errors():
    """Muestra errores y conteo de filas por BD de la última llamada a que_vendi."""
    return {"errors": _QV_LAST_ERRORS, "counts": _QV_LAST_COUNTS}


@app.get("/debug/flexxus_deudas_schema")
def debug_flexxus_deudas_schema():
    """Busca tablas relevantes (saldo/deuda/vendedor/cuenta) y columnas de CABEZACOMPROBANTES en prod."""
    DB_PROD     = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    result = {}
    keywords = ['DEUDA', 'SALDO', 'VENDEDOR', 'CUENTA', 'COBRANZA', 'VENC']
    for label, db_path in [('DB_PROD', DB_PROD), ('DB_MLT_PROD', DB_MLT_PROD)]:
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            # Todas las tablas
            cur.execute("SELECT TRIM(RDB$RELATION_NAME) FROM RDB$RELATIONS "
                        "WHERE RDB$SYSTEM_FLAG = 0 ORDER BY RDB$RELATION_NAME")
            todas = [r[0] for r in cur.fetchall()]
            relevantes = [t for t in todas if any(k in t.upper() for k in keywords)]
            result[f'{label}_tablas_relevantes'] = relevantes
            # Columnas de CABEZACOMPROBANTES
            cur.execute("SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS "
                        "WHERE TRIM(RDB$RELATION_NAME) = 'CABEZACOMPROBANTES' "
                        "ORDER BY RDB$FIELD_POSITION")
            cols_cab = [r[0] for r in cur.fetchall()]
            result[f'{label}_cabeza_cols'] = cols_cab
            # Columnas de CUERPOCOMPROBANTES (solo las 20 primeras)
            cur.execute("SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS "
                        "WHERE TRIM(RDB$RELATION_NAME) = 'CUERPOCOMPROBANTES' "
                        "ORDER BY RDB$FIELD_POSITION")
            cols_cue = [r[0] for r in cur.fetchall()]
            result[f'{label}_cuerpo_cols'] = cols_cue
            c.close()
        except Exception as e:
            result[f'{label}_error'] = str(e)
    return result


@app.get("/debug/cliente_vendedor/{nombre}")
def debug_cliente_vendedor(nombre: str):
    """Muestra CODIGOVENDEDOR de un cliente en DB_PROD y sus comprobantes pendientes."""
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    result = {}
    try:
        c = conn('WIN1252', DB_PROD)
        cur = c.cursor()
        cur.execute(
            "SELECT CODIGOCLIENTE, CODIGOPARTICULAR, RAZONSOCIAL, CODIGOVENDEDOR, ACTIVO "
            'FROM "CLIENTES" WHERE UPPER(RAZONSOCIAL) CONTAINING UPPER(?)',
            (nombre,)
        )
        rows = cur.fetchall()
        result['clientes'] = [{'cod': r[0], 'part': r[1], 'razon': r[2], 'vendedor': r[3], 'activo': r[4]} for r in rows]
        # Para cada cliente encontrado, contar sus comprobantes pendientes
        for cli in result['clientes']:
            cod = (cli['cod'] or '').strip()
            if not cod: continue
            cur.execute(
                "SELECT COUNT(*) FROM \"CABEZACOMPROBANTES\" "
                "WHERE CODIGOCLIENTE=? AND CUENTACORRIENTE='1' AND ANULADA='0' "
                "AND TIPOCOMPROBANTE NOT IN ('RE','RI','INA')",
                (cod,)
            )
            cli['comprobantes_cta'] = cur.fetchone()[0]
        c.close()
    except Exception as e:
        result['error'] = str(e)
    return result


@app.get("/debug/cliente_en_mlt/{nombre}")
def debug_cliente_en_mlt(nombre: str):
    """Diagnostica la presencia de un cliente en DB_MLT_PROD buscando por nombre parcial en CABEZACOMPROBANTES."""
    DB_PROD     = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    result = {}

    # 1. Buscar en DB_PROD CLIENTES
    try:
        c = conn('WIN1252', DB_PROD)
        cur = c.cursor()
        cur.execute(
            "SELECT CODIGOCLIENTE, CODIGOPARTICULAR, RAZONSOCIAL, CODIGOVENDEDOR "
            'FROM "CLIENTES" WHERE UPPER(RAZONSOCIAL) CONTAINING UPPER(?) AND ACTIVO=\'1\'',
            (nombre,)
        )
        rows = cur.fetchall()
        c.close()
        result['DB_PROD_clientes'] = [
            {'cod': r[0], 'part': r[1], 'razon': r[2], 'vendedor': r[3]} for r in rows
        ]
    except Exception as e:
        result['DB_PROD_error'] = str(e)

    # 2. Buscar en DB_PROD CABEZACOMPROBANTES por RAZONSOCIAL (para verificar si la FA está ahí)
    try:
        c = conn('WIN1252', DB_PROD)
        cur = c.cursor()
        cur.execute(
            "SELECT FIRST 10 CODIGOCLIENTE, TIPOCOMPROBANTE, NUMEROCOMPROBANTE, RAZONSOCIAL, FECHACOMPROBANTE, CUENTACORRIENTE, ANULADA "
            'FROM "CABEZACOMPROBANTES" WHERE UPPER(RAZONSOCIAL) CONTAINING UPPER(?) '
            "AND CUENTACORRIENTE='1' AND ANULADA='0' ORDER BY FECHACOMPROBANTE DESC",
            (nombre,)
        )
        rows = cur.fetchall()
        c.close()
        result['DB_PROD_cabeza_por_razon'] = [
            {'cod': r[0], 'tipo': r[1], 'num': r[2], 'razon': r[3], 'fecha': str(r[4]), 'cta': r[5], 'anul': r[6]} for r in rows
        ]
    except Exception as e:
        result['DB_PROD_cabeza_error'] = str(e)

    # 3. Buscar en DB_MLT_PROD CABEZACOMPROBANTES por RAZONSOCIAL
    try:
        c = conn('WIN1252', DB_MLT_PROD)
        cur = c.cursor()
        cur.execute(
            "SELECT FIRST 5 CODIGOCLIENTE, TIPOCOMPROBANTE, NUMEROCOMPROBANTE, RAZONSOCIAL, FECHACOMPROBANTE "
            'FROM "CABEZACOMPROBANTES" WHERE UPPER(RAZONSOCIAL) CONTAINING UPPER(?)',
            (nombre,)
        )
        rows = cur.fetchall()
        c.close()
        result['DB_MLT_PROD_cabeza_por_razon'] = [
            {'cod': r[0], 'tipo': r[1], 'num': r[2], 'razon': r[3], 'fecha': str(r[4])} for r in rows
        ]
    except Exception as e:
        result['DB_MLT_PROD_cabeza_error'] = str(e)

    # 3. Si encontramos clientes en DB_PROD, buscar su CODIGOPARTICULAR en CUERPOCOMPROBANTES de DB_MLT_PROD
    for cli in result.get('DB_PROD_clientes', []):
        part = cli.get('part', '')
        cod  = cli.get('cod', '')
        key  = f"lookup_part_{part or cod}"
        if part:
            try:
                c = conn('LATIN1', DB_MLT_PROD)
                cur = c.cursor()
                cur.execute(
                    'SELECT FIRST 5 DISTINCT CODIGOCLIENTE FROM "CUERPOCOMPROBANTES" WHERE CODIGOPARTICULAR = ?',
                    (part,)
                )
                rows = cur.fetchall()
                c.close()
                result[key] = {'codigoscliente_en_mlt': [str(r[0]) for r in rows]}
            except Exception as e:
                result[key] = {'error': str(e)}

    return result


@app.get("/debug/vista_deuda_directo/{codigocliente}")
def debug_vista_deuda_directo(codigocliente: str):
    """Consulta directa de VISTADEUDACLIENTES + CABEZACOMPROBANTES para un CODIGOCLIENTE."""
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    result = {}
    # 1. VISTADEUDACLIENTES directo
    for charset in ['LATIN1', 'WIN1252']:
        try:
            c = conn(charset, DB_PROD)
            cur = c.cursor()
            cur.execute(
                'SELECT TIPOCOMPROBANTE, NUMEROCOMPROBANTE, PAGADO, NETO, DEBE, TOTALACTUALIZADO, CODIGOMONEDA, COTIZACION '
                'FROM "VISTADEUDACLIENTES" WHERE CODIGOCLIENTE = ?',
                (codigocliente,)
            )
            rows = cur.fetchall()
            c.close()
            result[f'vista_{charset}'] = [
                {'tipo': r[0], 'num': r[1], 'pagado': r[2], 'neto': r[3],
                 'debe': r[4], 'total_act': r[5], 'moneda': r[6], 'cotiz': r[7]}
                for r in rows
            ]
            result[f'vista_{charset}_suma'] = sum(float(r[5] or 0) for r in rows)
        except Exception as e:
            result[f'vista_{charset}_error'] = str(e)
    # 2. CABEZACOMPROBANTES directo (sin MONEDAS) para ver registros raw
    try:
        c = conn('WIN1252', DB_PROD)
        cur = c.cursor()
        cur.execute(
            "SELECT TIPOCOMPROBANTE, NUMEROCOMPROBANTE, TOTAL, IVA1, IVA2, PAGADO, CODIGOMONEDA, COTIZACION, ANULADA, CUENTACORRIENTE "
            'FROM "CABEZACOMPROBANTES" WHERE CODIGOCLIENTE = ? '
            "AND CUENTACORRIENTE='1' AND ANULADA='0' "
            "AND TIPOCOMPROBANTE NOT IN ('RE','RI','INA') "
            "AND ABS(CAST(TOTAL AS DOUBLE PRECISION)+CAST(IVA1 AS DOUBLE PRECISION)+CAST(IVA2 AS DOUBLE PRECISION)-CAST(PAGADO AS DOUBLE PRECISION)) >= 0.01 "
            "ORDER BY FECHACOMPROBANTE DESC",
            (codigocliente,)
        )
        rows = cur.fetchall()
        c.close()
        result['cabeza_raw'] = [
            {'tipo': r[0], 'num': r[1], 'total': r[2], 'iva1': r[3], 'iva2': r[4],
             'pagado': r[5], 'moneda': r[6], 'cotiz': r[7], 'anul': r[8], 'cta': r[9]}
            for r in rows
        ]
        result['cabeza_raw_suma_simple'] = sum(
            float((r[2] or 0)) + float((r[3] or 0)) + float((r[4] or 0)) - float((r[5] or 0))
            for r in rows
        )
    except Exception as e:
        result['cabeza_raw_error'] = str(e)
    return result


@app.get("/debug/vista_deuda_clientes")
def debug_vista_deuda_clientes():
    """Inspecciona VISTADEUDACLIENTES: columnas, definición SQL y muestra de datos."""
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    result = {}
    try:
        c = conn('WIN1252', db=DB_PROD)
        cur = c.cursor()
        # Columnas de la vista
        cur.execute("SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS "
                    "WHERE TRIM(RDB$RELATION_NAME) = 'VISTADEUDACLIENTES' "
                    "ORDER BY RDB$FIELD_POSITION")
        cols = [r[0] for r in cur.fetchall()]
        result['columnas'] = cols
        # Definición SQL de la vista
        cur.execute("SELECT RDB$VIEW_SOURCE FROM RDB$RELATIONS "
                    "WHERE TRIM(RDB$RELATION_NAME) = 'VISTADEUDACLIENTES'")
        row = cur.fetchone()
        result['sql_vista'] = str(row[0]) if row else None
        # Muestra de 3 filas
        try:
            cur.execute('SELECT FIRST 3 * FROM "VISTADEUDACLIENTES"')
            sample_cols = [d[0] for d in cur.description]
            sample_rows = cur.fetchall()
            result['muestra'] = [dict(zip(sample_cols, [str(v) if v is not None else None for v in r])) for r in sample_rows]
        except Exception as e2:
            result['muestra_error'] = str(e2)
        # Total deuda en la vista (para comparar con Flexxus)
        try:
            deuda_col = next((col for col in cols if 'DEUDA' in col.upper() or 'SALDO' in col.upper()), None)
            if deuda_col:
                cur.execute(f'SELECT COUNT(*), SUM(CAST("{deuda_col}" AS DOUBLE PRECISION)) FROM "VISTADEUDACLIENTES"')
                r = cur.fetchone()
                result[f'total_filas'] = r[0]
                result[f'suma_{deuda_col}'] = float(r[1] or 0)
        except Exception as e3:
            result['suma_error'] = str(e3)
        c.close()
    except Exception as e:
        result['error'] = str(e)
    return result


@app.get("/debug/cta_vendedor/{vendedor}")
def debug_cta_vendedor(vendedor: str):
    """Diagnóstico: suma deuda directo desde CABEZACOMPROBANTES con CODIGOVENDEDOR si existe."""
    DB_PROD     = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    result = {}
    for label, db_path in [('DB_PROD', DB_PROD), ('DB_MLT_PROD', DB_MLT_PROD)]:
        try:
            c = conn('LATIN1', db=db_path)
            cur = c.cursor()
            # Verificar si CABEZACOMPROBANTES tiene CODIGOVENDEDOR
            cur.execute("SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS "
                        "WHERE TRIM(RDB$RELATION_NAME) = 'CABEZACOMPROBANTES' "
                        "AND TRIM(RDB$FIELD_NAME) CONTAINING 'VENDEDOR'")
            cols_vend = [r[0] for r in cur.fetchall()]
            result[f'{label}_cols_vendedor_en_cabeza'] = cols_vend
            if cols_vend:
                col = cols_vend[0]
                cur.execute(f"""
                    SELECT COUNT(*),
                           SUM(CAST(TOTAL AS DOUBLE PRECISION) + CAST(IVA1 AS DOUBLE PRECISION) + CAST(IVA2 AS DOUBLE PRECISION)),
                           SUM(CAST(TOTAL AS DOUBLE PRECISION) + CAST(IVA1 AS DOUBLE PRECISION) + CAST(IVA2 AS DOUBLE PRECISION) - CAST(PAGADO AS DOUBLE PRECISION))
                    FROM "CABEZACOMPROBANTES"
                    WHERE UPPER({col}) = UPPER(?)
                      AND CUENTACORRIENTE = '1'
                      AND ANULADA = '0'
                      AND TIPOCOMPROBANTE IN ('FA','FB','FE','FCA','FCB','DI','SIV','NCA','NCB','NDA','NDB','NCAE','NDAE')
                      AND ABS(CAST(TOTAL AS DOUBLE PRECISION) + CAST(IVA1 AS DOUBLE PRECISION) + CAST(IVA2 AS DOUBLE PRECISION) - CAST(PAGADO AS DOUBLE PRECISION)) > 0.01
                """, (vendedor.upper(),))
                row = cur.fetchone()
                result[f'{label}_directo_count'] = row[0]
                result[f'{label}_directo_total_bruto'] = float(row[1] or 0)
                result[f'{label}_directo_deuda'] = float(row[2] or 0)
            c.close()
        except Exception as e:
            result[f'{label}_error'] = str(e)
    return result


@app.get("/debug/camping_query/{cod}")
def debug_camping_query(cod: str):
    """Debug exacto de _query_cta para un cliente específico (ej: 348 = CAMPING LA PLATA)."""
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    out = {"cod": cod, "db_prod": {}, "db_mlt": {}}
    for db_key, db_path in [("db_prod", DB_PROD), ("db_mlt", DB_MLT_PROD)]:
        res = {"charset_test": {}, "fetchone_loop": [], "errors": []}
        # Test 1: conectar y contar sin filtros extra
        for charset in ['WIN1252', 'LATIN1']:
            try:
                c = conn(charset, db=db_path)
                cur = c.cursor()
                cur.execute('SELECT COUNT(*) FROM "CABEZACOMPROBANTES" WHERE CODIGOCLIENTE = ?', (cod,))
                row = cur.fetchone()
                c.close()
                res["charset_test"][charset] = {"count_total": row[0] if row else None}
            except Exception as e:
                res["charset_test"][charset] = {"error": str(e)}

        # Test 2: SIN FILTROS — ver todos los registros para detectar debe>0 excluidos
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            cur.execute("""
                SELECT FIRST 500 SKIP 0
                    TIPOCOMPROBANTE, NUMEROCOMPROBANTE, FECHACOMPROBANTE,
                    TOTAL, IVA1, IVA2, PAGADO, COTIZACION, CODIGOMONEDA,
                    FECHAVENCIMIENTO, CLASECOMPROBANTE, CUENTACORRIENTE, ANULADA
                FROM "CABEZACOMPROBANTES"
                WHERE CODIGOCLIENTE = ?
                ORDER BY FECHACOMPROBANTE ASC
            """, (cod,))
            all_rows = []
            while True:
                try:
                    r = cur.fetchone()
                    if r is None: break
                    total=float(r[3] or 0); iva1=float(r[4] or 0); iva2=float(r[5] or 0)
                    pagado=float(r[6] or 0); neto=total+iva1+iva2; debe=neto-pagado
                    if abs(debe) >= 0.01:
                        all_rows.append({"tipo":str(r[0]),"num":str(r[1]),"fecha":str(r[2]),
                            "neto":round(neto,2),"pagado":round(pagado,2),"debe":round(debe,2),
                            "cotiz":float(r[7] or 1),"moneda":str(r[8] or '').strip(),
                            "ctacte":str(r[11]),"anulada":str(r[12])})
                except Exception as e:
                    all_rows.append({"row_error": str(e)})
                    break
            res["sin_filtros_con_saldo"] = all_rows
            c.close()
        except Exception as exec_e:
            res["sin_filtros_error"] = str(exec_e)

        # Test 3: fetchone loop con WIN1252 — misma query que _query_cta
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            cur.execute("""
                SELECT FIRST 500 SKIP 0
                    TIPOCOMPROBANTE, NUMEROCOMPROBANTE, FECHACOMPROBANTE,
                    TOTAL, IVA1, IVA2, PAGADO, COTIZACION, CODIGOMONEDA,
                    FECHAVENCIMIENTO, CLASECOMPROBANTE
                FROM "CABEZACOMPROBANTES"
                WHERE CODIGOCLIENTE = ?
                  AND CUENTACORRIENTE = '1'
                  AND ANULADA = '0'
                  AND TIPOCOMPROBANTE NOT IN ('RE', 'RI', 'INA')
                ORDER BY FECHAVENCIMIENTO ASC, FECHACOMPROBANTE ASC
            """, (cod,))
            n = 0
            while True:
                try:
                    r = cur.fetchone()
                    if r is None:
                        break
                    n += 1
                    # intento conversión
                    try:
                        total = float(r[3] or 0); iva1 = float(r[4] or 0); iva2 = float(r[5] or 0)
                        pagado = float(r[6] or 0); cotiz = float(r[7] or 1) or 1.0
                        neto = total + iva1 + iva2; debe = neto - pagado
                        res["fetchone_loop"].append({
                            "n": n, "tipo": str(r[0]), "num": str(r[1]),
                            "fecha": str(r[2]), "neto": neto, "debe": debe,
                            "cotiz": cotiz, "moneda": str(r[8] or '').strip()
                        })
                    except Exception as conv_e:
                        res["fetchone_loop"].append({"n": n, "conv_error": str(conv_e), "raw": [str(x) for x in r]})
                except Exception as fetch_e:
                    res["errors"].append({"at_row": n, "fetch_error": str(fetch_e)})
                    break
            res["total_fetched"] = n
            c.close()
        except Exception as exec_e:
            res["execute_error"] = str(exec_e)

        out[db_key] = res
    return out

@app.get("/debug/listar_dbs")
def listar_dbs():
    import glob
    archivos = glob.glob('c:/flexxus/**/*.gdb', recursive=True) + \
               glob.glob('c:/flexxus/**/*.FDB', recursive=True)
    return {"archivos": sorted(archivos)}

@app.get("/debug/generators")
def debug_generators():
    """Lista generators y su valor actual en ambas BDs"""
    resultado = {}
    for nombre, db_path in [('oficial', DATABASE), ('sw', DATABASE_MLT)]:
        try:
            c = conn('LATIN1', db=db_path)
            cur = c.cursor()
            cur.execute(
                "SELECT TRIM(RDB$GENERATOR_NAME), RDB$GENERATOR_ID "
                "FROM RDB$GENERATORS WHERE RDB$SYSTEM_FLAG = 0 "
                "ORDER BY RDB$GENERATOR_NAME"
            )
            gens = cur.fetchall()
            vals = {}
            for g in gens:
                try:
                    cur.execute(f'SELECT GEN_ID("{g[0]}", 0) FROM RDB$DATABASE')
                    vals[g[0]] = cur.fetchone()[0]
                except Exception:
                    vals[g[0]] = None
            resultado[nombre] = vals
            c.close()
        except Exception as e:
            resultado[nombre] = {"error": str(e)}
    return resultado

@app.get("/debug/tablas_mlt")
def debug_tablas_mlt():
    """Lista todas las tablas de usuario en DB-MLT-Microbell.gdb"""
    try:
        c = conn('LATIN1', db=DATABASE_MLT)
        cur = c.cursor()
        cur.execute(
            "SELECT TRIM(RDB$RELATION_NAME) FROM RDB$RELATIONS "
            "WHERE RDB$SYSTEM_FLAG = 0 AND RDB$VIEW_BLR IS NULL "
            "ORDER BY RDB$RELATION_NAME"
        )
        tablas = [r[0] for r in cur.fetchall()]
        c.close()
        return {"tablas": tablas}
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/esquema_clientes")
def debug_esquema_clientes():
    """Muestra todas las columnas de la tabla CLIENTES en DB-Prueba.gdb"""
    try:
        c = conn()
        cur = c.cursor()
        cur.execute(
            "SELECT TRIM(RDB$FIELD_NAME), RDB$FIELD_POSITION "
            "FROM RDB$RELATION_FIELDS "
            "WHERE TRIM(RDB$RELATION_NAME) = 'CLIENTES' "
            "ORDER BY RDB$FIELD_POSITION"
        )
        cols = [{"pos": r[1], "nombre": r[0]} for r in cur.fetchall()]
        c.close()
        return {"columnas": cols}
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/muestra_clientes")
def debug_muestra_clientes(vendedor: str = "KRAFFT"):
    """Muestra 3 clientes del vendedor para ver valores de columnas IVA"""
    try:
        c = conn()
        cur = c.cursor()
        cur.execute(
            'SELECT FIRST 3 CODIGOCLIENTE, RAZONSOCIAL, CODIGOPARTICULAR '
            'FROM "CLIENTES" WHERE ACTIVO = ? AND CODIGOVENDEDOR = ? ORDER BY RAZONSOCIAL',
            ('1', vendedor.upper())
        )
        rows = cur.fetchall()
        c.close()
        return [{"codigo": r[0], "razonsocial": r[1], "codigoparticular": r[2]} for r in rows]
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/tablas_of")
def debug_tablas_of():
    """Lista todas las tablas de DB-Prueba.gdb"""
    try:
        c = conn('LATIN1', db=DATABASE)
        cur = c.cursor()
        cur.execute("SELECT RDB$RELATION_NAME FROM RDB$RELATIONS WHERE RDB$SYSTEM_FLAG=0 ORDER BY RDB$RELATION_NAME")
        rows = cur.fetchall()
        c.close()
        return [r[0].strip() for r in rows]
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/contadores_of")
def debug_contadores_of():
    """Muestra todos los contadores de la BD oficial (DB-Prueba.gdb)"""
    try:
        c = conn('LATIN1', db=DATABASE)
        cur = c.cursor()
        cur.execute('SELECT CODIGOCONTADOR, DESCRIPCION, VALOR FROM "CONTADORES" ORDER BY CODIGOCONTADOR')
        rows = cur.fetchall()
        c.close()
        return [{"codigo": r[0], "descripcion": (r[1] or "").strip(), "valor": r[2]} for r in rows]
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/contadores_sw")
def debug_contadores_sw():
    """Muestra todos los contadores de la BD SW (DB-MLT-Prueba.gdb)"""
    try:
        c = conn('LATIN1', db=DATABASE_MLT)
        cur = c.cursor()
        cur.execute('SELECT CODIGOCONTADOR, DESCRIPCION, VALOR FROM "CONTADORES" ORDER BY CODIGOCONTADOR')
        rows = cur.fetchall()
        c.close()
        return [{"codigo": r[0], "descripcion": (r[1] or "").strip(), "valor": r[2]} for r in rows]
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/info_bd")
def debug_info_bd():
    """Muestra las rutas de BD configuradas y cuenta de registros clave"""
    result = {
        "DATABASE": DATABASE,
        "DATABASE_MLT": DATABASE_MLT,
    }
    try:
        c = conn()
        cur = c.cursor()
        cur.execute('SELECT COUNT(*) FROM "CABEZAPRESUPUESTOS"')
        result["presupuestos_total"] = cur.fetchone()[0]
        cur.execute('SELECT MAX(CAST(NUMEROCOMPROBANTE AS INTEGER)) FROM "CABEZAPRESUPUESTOS"')
        result["presupuesto_max"] = cur.fetchone()[0]
        cur.execute('SELECT COUNT(*) FROM "CABEZAPEDIDOS"')
        result["pedidos_total"] = cur.fetchone()[0]
        cur.execute('SELECT MAX(CAST(NUMEROCOMPROBANTE AS INTEGER)) FROM "CABEZAPEDIDOS"')
        result["pedido_max"] = cur.fetchone()[0]
        c.close()
    except Exception as e:
        result["error"] = str(e)
    return result

@app.get("/debug/config_prueba")
def debug_config_prueba():
    """Lee configuracion.ini del perfil Prueba"""
    import os
    result = {}
    paths = [
        'c:/flexxus/FlexxusERP/Prueba/bin/configuracion.ini',
        'c:/flexxus/FlexxusERP/BIN/FlexxusServer.ini',
    ]
    for p in paths:
        try:
            with open(p, 'r', encoding='latin1', errors='replace') as f:
                result[p] = f.read()
        except Exception as e:
            result[p] = f"Error: {e}"
    return result

@app.get("/debug/config_flexxus")
def debug_config_flexxus():
    """Busca archivos de configuración de Flexxus que definen la BD por empresa"""
    import glob, os
    result = {"archivos": [], "contenido": {}}
    patterns = [
        'c:/flexxus/**/*.ini', 'c:/flexxus/**/*.cfg',
        'c:/flexxus/**/*.config', 'c:/flexxus/**/Empresas*.xml',
        'c:/flexxus/**/empresas*.ini', 'c:/flexxus/**/conexion*.ini',
        'c:/flexxus/FlexxusERP/*.ini', 'c:/flexxus/FlexxusERP/*.cfg',
        'c:/flexxus/FlexxusERP/*.xml',
    ]
    for p in patterns:
        for f in glob.glob(p, recursive=True):
            result["archivos"].append(f)
            try:
                with open(f, 'r', encoding='latin1', errors='replace') as fp:
                    content = fp.read(3000)
                if 'DB-' in content or 'Prueba' in content or 'Microbell' in content or 'gdb' in content.lower():
                    result["contenido"][f] = content
            except Exception as e:
                result["contenido"][f] = f"Error: {e}"
    return result

@app.get("/debug/comparar_bases")
def debug_comparar_bases():
    """Compara presupuesto MAX en DB-Prueba vs DB-Microbell"""
    result = {}
    for nombre, path in [
        ("DB-Prueba", "c:/flexxus/DB/DB-Prueba.gdb"),
        ("DB-Microbell", "c:/flexxus/DB/DB-Microbell.gdb"),
    ]:
        try:
            c = firebirdsql.connect(host=HOST, port=PORT, database=path,
                                    user=DB_USER, password=DB_PASS, charset='LATIN1')
            cur = c.cursor()
            cur.execute('SELECT MAX(CAST(NUMEROCOMPROBANTE AS INTEGER)), COUNT(*) FROM "CABEZAPRESUPUESTOS"')
            row = cur.fetchone()
            cur.execute('SELECT MAX(CAST(NUMEROCOMPROBANTE AS INTEGER)), COUNT(*) FROM "CABEZAPEDIDOS"')
            row2 = cur.fetchone()
            c.close()
            result[nombre] = {
                "presupuesto_max": row[0], "presupuestos_total": row[1],
                "pedido_max": row2[0], "pedidos_total": row2[1]
            }
        except Exception as e:
            result[nombre] = {"error": str(e)}
    return result

@app.get("/debug/listar_gdbs")
def debug_listar_gdbs():
    """Lista todos los archivos .gdb y .fdb en c:/flexxus/"""
    import glob, os
    result = []
    for pattern in ['c:/flexxus/**/*.gdb', 'c:/flexxus/**/*.fdb',
                    'c:/Flexxus/**/*.gdb', 'c:/Flexxus/**/*.fdb']:
        for f in glob.glob(pattern, recursive=True):
            try:
                size = os.path.getsize(f)
                mtime = os.path.getmtime(f)
                from datetime import datetime
                result.append({
                    "path": f,
                    "size_mb": round(size / 1024 / 1024, 1),
                    "modificado": datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M')
                })
            except Exception:
                result.append({"path": f})
    return sorted(result, key=lambda x: x.get("modificado",""), reverse=True)

@app.get("/debug/esquema_cuerpos")
def debug_esquema_cuerpos():
    """Columnas reales de CUERPOPRESUPUESTOS y CUERPOPEDIDOS en DB-Prueba.gdb"""
    result = {}
    c = conn()
    cur = c.cursor()
    for tabla in ['CUERPOPRESUPUESTOS', 'CUERPOPEDIDOS']:
        cur.execute(
            "SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS "
            "WHERE TRIM(RDB$RELATION_NAME) = ? ORDER BY RDB$FIELD_POSITION",
            (tabla,))
        result[tabla] = [r[0] for r in cur.fetchall()]
    c.close()
    return result

@app.get("/debug/esquema_cabezas")
def debug_esquema_cabezas():
    """Columnas de CABEZAPRESUPUESTOS y CABEZAPEDIDOS"""
    result = {}
    c = conn()
    cur = c.cursor()
    for tabla in ['CABEZAPRESUPUESTOS', 'CABEZAPEDIDOS']:
        cur.execute(
            "SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS "
            "WHERE TRIM(RDB$RELATION_NAME) = ? ORDER BY RDB$FIELD_POSITION",
            (tabla,))
        result[tabla] = [r[0] for r in cur.fetchall()]
    c.close()
    return result

@app.get("/debug/muestra_presupuesto/{numero}")
def debug_muestra_presupuesto(numero: str):
    """Muestra cabeza + cuerpo de un presupuesto específico"""
    try:
        c = conn()
        cur = c.cursor()
        cur.execute('SELECT * FROM "CABEZAPRESUPUESTOS" WHERE NUMEROCOMPROBANTE = ?', (numero,))
        cols_cab = [d[0] for d in cur.description]
        row = cur.fetchone()
        cabeza = dict(zip(cols_cab, [str(v) if v is not None else None for v in row])) if row else None
        cur.execute('SELECT * FROM "CUERPOPRESUPUESTOS" WHERE NUMEROCOMPROBANTE = ? ORDER BY LINEA', (numero,))
        cols_cue = [d[0] for d in cur.description]
        cuerpos = [dict(zip(cols_cue, [str(v) if v is not None else None for v in r])) for r in cur.fetchall()]
        c.close()
        return {"cabeza": cabeza, "cuerpos": cuerpos}
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/codigos_presupuesto/{numero}")
def debug_codigos_presupuesto(numero: str):
    """Compara CODIGOARTICULO vs CODIGOPARTICULAR en CUERPOPRESUPUESTOS."""
    try:
        c = conn()
        cur = c.cursor()
        cur.execute(
            "SELECT LINEA, CODIGOARTICULO, CODIGOPARTICULAR, DESCRIPCION "
            "FROM \"CUERPOPRESUPUESTOS\" WHERE NUMEROCOMPROBANTE = ? ORDER BY LINEA",
            (numero,)
        )
        rows = cur.fetchall()
        c.close()
        return [{"linea": r[0], "codigoarticulo": r[1], "codigoparticular": r[2], "descripcion": r[3]} for r in rows]
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/ultimos_presupuestos")
def debug_ultimos_presupuestos():
    """Últimos 5 presupuestos en DB-Prueba.gdb"""
    try:
        c = conn()
        cur = c.cursor()
        cur.execute(
            'SELECT FIRST 5 TIPOCOMPROBANTE, NUMEROCOMPROBANTE, CODIGOCLIENTE, '
            'RAZONSOCIAL, FECHACOMPROBANTE, TOTAL, CODIGOUSUARIO '
            'FROM "CABEZAPRESUPUESTOS" ORDER BY FECHAMODIFICACION DESC')
        rows = cur.fetchall()
        c.close()
        return [{"tipo": r[0], "numero": r[1], "cliente": r[2],
                 "razon": r[3], "fecha": str(r[4]), "total": r[5], "usuario": r[6]}
                for r in rows]
    except Exception as e:
        return {"error": str(e)}


@app.get("/debug/comparar/{num_bueno}/{num_malo}")
def debug_comparar(num_bueno: str, num_malo: str):
    """
    Compara TODAS las columnas de dos presupuestos.
    Uso: /debug/comparar/7849/7857
    Resalta las columnas que difieren.
    """
    try:
        c = conn()
        cur = c.cursor()

        def leer(numero):
            cur.execute('SELECT * FROM "CABEZAPRESUPUESTOS" WHERE NUMEROCOMPROBANTE = ?', (numero,))
            cols = [d[0] for d in cur.description]
            row = cur.fetchone()
            if not row:
                return None, cols
            return dict(zip(cols, [str(v) if v is not None else '__NULL__' for v in row])), cols

        datos_bueno, cols = leer(num_bueno)
        datos_malo, _    = leer(num_malo)

        # Defaults del dominio para cada columna
        cur.execute("""
            SELECT TRIM(rf.RDB$FIELD_NAME), rf.RDB$NULL_FLAG, rf.RDB$DEFAULT_SOURCE
            FROM RDB$RELATION_FIELDS rf
            WHERE rf.RDB$RELATION_NAME = 'CABEZAPRESUPUESTOS'
            ORDER BY rf.RDB$FIELD_POSITION
        """)
        defaults = {r[0].strip(): {"not_null": r[1]==1, "default": str(r[2]).strip() if r[2] else None}
                    for r in cur.fetchall()}

        # Todos los triggers (INSERT + UPDATE)
        cur.execute("""
            SELECT TRIM(RDB$TRIGGER_NAME), RDB$TRIGGER_TYPE, CAST(RDB$TRIGGER_SOURCE AS VARCHAR(4000))
            FROM RDB$TRIGGERS
            WHERE RDB$RELATION_NAME = 'CABEZAPRESUPUESTOS' AND RDB$SYSTEM_FLAG = 0
        """)
        triggers = [{"name": r[0], "type": r[1],
                     "type_desc": {1:"BEFORE INSERT",2:"AFTER INSERT",3:"BEFORE UPDATE",4:"AFTER UPDATE",
                                   5:"BEFORE DELETE",6:"AFTER DELETE"}.get(r[1],"?"),
                     "source": r[2]} for r in cur.fetchall()]

        c.close()

        if not datos_bueno:
            return {"error": f"No encontrado: {num_bueno}"}
        if not datos_malo:
            return {"error": f"No encontrado: {num_malo}"}

        diferencias = {}
        iguales = {}
        for col in cols:
            v1 = datos_bueno.get(col)
            v2 = datos_malo.get(col)
            if v1 != v2:
                diferencias[col] = {
                    f"bueno_{num_bueno}": v1,
                    f"malo_{num_malo}": v2,
                    "default_info": defaults.get(col.strip())
                }
            else:
                iguales[col] = v1

        return {
            "diferencias": diferencias,
            "total_diferencias": len(diferencias),
            "iguales": iguales,
            "triggers": triggers,
            "defaults_columnas": defaults,
        }
    except Exception as e:
        return {"error": str(e)}


@app.post("/setup/restaurar_defaults_aprobacion")
def restaurar_defaults_aprobacion():
    """
    Restaura el DEFAULT original de FECHAAPROBADO ('1900-01-01') y CODIGOUSUARIOAPROBACION ('')
    que fue eliminado por setup/fechaaprobado_nullable.
    Ejecutar UNA SOLA VEZ.
    """
    try:
        import struct
        c = conn('WIN1252')
        cur = c.cursor()

        # Restaurar DEFAULT '1900-01-01 00:00:00' en FECHAAPROBADO
        # En Firebird, el DEFAULT_SOURCE es texto SQL, y DEFAULT_VALUE es BLR binario.
        # Actualizamos solo DEFAULT_SOURCE (texto); Firebird lo recompila al arrancar.
        cur.execute("""
            UPDATE RDB$RELATION_FIELDS
            SET RDB$DEFAULT_SOURCE = 'DEFAULT ''1900-01-01 00:00:00'''
            WHERE RDB$RELATION_NAME = 'CABEZAPRESUPUESTOS'
              AND TRIM(RDB$FIELD_NAME) = 'FECHAAPROBADO'
        """)
        # Restaurar DEFAULT '' en CODIGOUSUARIOAPROBACION
        cur.execute("""
            UPDATE RDB$RELATION_FIELDS
            SET RDB$DEFAULT_SOURCE = 'DEFAULT '''''
            WHERE RDB$RELATION_NAME = 'CABEZAPRESUPUESTOS'
              AND TRIM(RDB$FIELD_NAME) = 'CODIGOUSUARIOAPROBACION'
        """)
        c.commit()

        # Verificar
        cur.execute("""
            SELECT TRIM(RDB$FIELD_NAME), RDB$NULL_FLAG, RDB$DEFAULT_SOURCE
            FROM RDB$RELATION_FIELDS
            WHERE RDB$RELATION_NAME = 'CABEZAPRESUPUESTOS'
              AND TRIM(RDB$FIELD_NAME) IN ('FECHAAPROBADO', 'CODIGOUSUARIOAPROBACION')
        """)
        rows = cur.fetchall()
        c.close()
        return {
            "ok": True,
            "columnas": [{"campo": r[0], "not_null": r[1], "default": str(r[2]) if r[2] else None} for r in rows],
            "instruccion": "Reiniciar Firebird para que tome los nuevos defaults"
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/debug/parametros_np")
def debug_parametros_np():
    """Verifica el valor de NP en PARAMETROS y hace prueba de TRIM."""
    res = {}
    try:
        c = conn('LATIN1')
        cur = c.cursor()
        # Buscar con TRIM
        cur.execute("SELECT TIPODOCUMENTO, VALOR FROM \"PARAMETROS\" WHERE TRIM(TIPODOCUMENTO) = 'NP'")
        row = cur.fetchone()
        res['con_trim'] = {'tipodocumento': repr(row[0]), 'valor': row[1]} if row else None
        # Buscar sin TRIM para ver el valor crudo
        cur.execute("SELECT TIPODOCUMENTO, VALOR FROM \"PARAMETROS\" WHERE TIPODOCUMENTO LIKE '%NP%'")
        rows = cur.fetchall()
        res['like_NP'] = [{'tipodocumento': repr(r[0]), 'valor': r[1]} for r in rows]
        # Ver todas las filas con comprobante tipo 2 chars
        cur.execute("SELECT TIPODOCUMENTO, VALOR FROM \"PARAMETROS\" ORDER BY TIPODOCUMENTO")
        all_rows = cur.fetchall()
        res['primeros_20'] = [{'tipodocumento': repr(r[0]), 'valor': r[1]} for r in all_rows[:20]]
        c.close()
    except Exception as e:
        res['error'] = str(e)
    return res

@app.get("/debug/ultimos_pedidos")
def debug_ultimos_pedidos():
    """Últimos 5 pedidos en DB-Prueba.gdb"""
    try:
        c = conn()
        cur = c.cursor()
        cur.execute(
            'SELECT FIRST 5 TIPOCOMPROBANTE, NUMEROCOMPROBANTE, CODIGOCLIENTE, '
            'RAZONSOCIAL, FECHACOMPROBANTE, TOTAL, CODIGOUSUARIO '
            'FROM "CABEZAPEDIDOS" ORDER BY FECHAMODIFICACION DESC')
        rows = cur.fetchall()
        c.close()
        return [{"tipo": r[0], "numero": r[1], "cliente": r[2],
                 "razon": r[3], "fecha": str(r[4]), "total": r[5], "usuario": r[6]}
                for r in rows]
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/sw_fix_cantremitida")
def debug_sw_fix_cantremitida():
    """Corrige CANTIDADREMITIDA=0 en CUERPOCOMPROBANTES para todos los NPs SW creados por la API."""
    try:
        c = conn('LATIN1', db=DATABASE_MLT)
        cur = c.cursor()
        # Buscar NPs que tengan CANTIDADREMITIDA = CANTIDAD (los incorrectos)
        cur.execute("""
            SELECT cc.NUMEROCOMPROBANTE, cc.LINEA, cc.CANTIDAD, cc.CANTIDADREMITIDA
            FROM "CUERPOCOMPROBANTES" cc
            JOIN "CABEZACOMPROBANTES" cab ON cab.NUMEROCOMPROBANTE = cc.NUMEROCOMPROBANTE
              AND cab.TIPOCOMPROBANTE = 'NP'
            WHERE cc.TIPOCOMPROBANTE = 'NP'
              AND cc.CANTIDADREMITIDA > 0
              AND cc.CANTIDADREMITIDA = cc.CANTIDAD
        """)
        rows = cur.fetchall()
        afectados = [{'num': str(r[0]), 'linea': str(r[1]), 'cantidad': str(r[2]), 'remitida': str(r[3])} for r in rows]

        if not rows:
            c.close()
            return {"mensaje": "No hay registros con CANTIDADREMITIDA incorrecta", "afectados": []}

        # Corregir
        cur.execute("""
            UPDATE "CUERPOCOMPROBANTES"
            SET CANTIDADREMITIDA = 0
            WHERE TIPOCOMPROBANTE = 'NP'
              AND CANTIDADREMITIDA > 0
              AND CANTIDADREMITIDA = CANTIDAD
        """)
        c.commit()
        c.close()
        return {"mensaje": f"Corregidos {len(rows)} ítems", "afectados": afectados}
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/sw_np_flexxus")
def debug_sw_np_flexxus():
    """Busca NPs en DATABASE_MLT creados por Flexxus (no por la API) para ver qué tabla usan."""
    res = {}
    try:
        c = conn('LATIN1', db=DATABASE_MLT)
        cur = c.cursor()

        # NPs en CABEZACOMPROBANTES anteriores a mayo 2026 (creados por Flexxus, no por la API)
        cur.execute("""
            SELECT FIRST 5 NUMEROCOMPROBANTE, TIPOCOMPROBANTE, CODIGOCLIENTE,
                   RAZONSOCIAL, FECHACOMPROBANTE, CODIGOUSUARIO, FECHAMODIFICACION
            FROM "CABEZACOMPROBANTES"
            WHERE TIPOCOMPROBANTE = 'NP'
              AND FECHACOMPROBANTE < '2026-05-01 00:00:00'
            ORDER BY NUMEROCOMPROBANTE DESC
        """)
        rows = cur.fetchall()
        res['np_anteriores_en_cabezacomprobantes'] = [
            {'num': str(r[0]), 'tipo': r[1], 'cli': r[2], 'razon': r[3],
             'fecha': str(r[4]), 'usuario': r[5], 'fechamod': str(r[6]) if r[6] else None}
            for r in rows
        ] if rows else 'NINGUNO — Flexxus SW nunca creó NPs en esta tabla'

        # Contar todos los tipos de comprobante en CABEZACOMPROBANTES
        cur.execute("""
            SELECT TRIM(TIPOCOMPROBANTE), COUNT(*)
            FROM "CABEZACOMPROBANTES"
            GROUP BY TIPOCOMPROBANTE
            ORDER BY COUNT(*) DESC
        """)
        res['tipos_en_cabezacomprobantes'] = {r[0]: r[1] for r in cur.fetchall()}

        # Ver todas las tablas de la BD MLT para entender su estructura
        cur.execute("""
            SELECT TRIM(RDB$RELATION_NAME)
            FROM RDB$RELATIONS
            WHERE RDB$SYSTEM_FLAG = 0
              AND RDB$VIEW_SOURCE IS NULL
            ORDER BY RDB$RELATION_NAME
        """)
        res['todas_las_tablas'] = [r[0] for r in cur.fetchall()]

        c.close()
    except Exception as e:
        res['error'] = str(e)
    return res

@app.get("/debug/sw_tablas_np")
def debug_sw_tablas_np():
    """Verifica si DATABASE_MLT tiene CABEZAPEDIDOS y muestra sus últimos NPs."""
    res = {}
    try:
        c = conn('LATIN1', db=DATABASE_MLT)
        cur = c.cursor()
        # Verificar qué tablas NP-relacionadas existen en MLT
        cur.execute("""
            SELECT TRIM(RDB$RELATION_NAME)
            FROM RDB$RELATIONS
            WHERE RDB$SYSTEM_FLAG = 0
              AND (RDB$RELATION_NAME LIKE '%PEDIDO%' OR RDB$RELATION_NAME LIKE '%COMPROBANTE%')
            ORDER BY RDB$RELATION_NAME
        """)
        res['tablas'] = [r[0] for r in cur.fetchall()]

        # Si existe CABEZAPEDIDOS en MLT, mostrar los últimos NPs
        if 'CABEZAPEDIDOS' in res['tablas']:
            try:
                cur.execute("""
                    SELECT FIRST 5 NUMEROCOMPROBANTE, TIPOCOMPROBANTE, CODIGOCLIENTE,
                           RAZONSOCIAL, FECHACOMPROBANTE, OPERACION, ANULADA, FECHAMODIFICACION
                    FROM "CABEZAPEDIDOS"
                    WHERE TIPOCOMPROBANTE = 'NP'
                    ORDER BY NUMEROCOMPROBANTE DESC
                """)
                cols = ['numero','tipo','cliente','razon','fecha','operacion','anulada','fechamod']
                res['cabezapedidos_ultimos'] = [dict(zip(cols, [str(v) if v else None for v in r])) for r in cur.fetchall()]
            except Exception as ex:
                res['cabezapedidos_error'] = str(ex)

        c.close()
    except Exception as e:
        res['error'] = str(e)
    return res

@app.get("/debug/np_l1_full/{numero}")
def debug_np_l1_full(numero: str):
    """Muestra NP en DATABASE (Línea 1) CABEZAPEDIDOS con tablas relacionadas."""
    res = {}
    try:
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()
        cur.execute("SELECT * FROM \"CABEZAPEDIDOS\" WHERE NUMEROCOMPROBANTE=? AND TIPOCOMPROBANTE='NP'", (numero,))
        row = cur.fetchone()
        if row:
            cols = [d[0] for d in cur.description]
            res['cabeza'] = {cols[i]: str(row[i]) if row[i] is not None else None for i in range(len(cols))}
        else:
            res['cabeza'] = 'NO ENCONTRADO en CABEZAPEDIDOS'

        # Campos dinámicos en DATABASE
        try:
            cur.execute("SELECT * FROM \"CAMPOSDINAMICOSCOMPROBANTES\" WHERE NUMEROCOMPROBANTE=?", (numero,))
            rows = cur.fetchall()
            if rows:
                cols2 = [d[0] for d in cur.description]
                res['campos_dinamicos'] = [{cols2[i]: str(r[i]) if r[i] is not None else None for i in range(len(cols2))} for r in rows]
            else:
                res['campos_dinamicos'] = []
        except Exception as ex:
            res['campos_dinamicos_err'] = str(ex)

        # Diferencias con un NP de Línea 1 puro (100023558)
        ref = '100023558'
        cur.execute("SELECT * FROM \"CABEZAPEDIDOS\" WHERE NUMEROCOMPROBANTE=? AND TIPOCOMPROBANTE='NP'", (ref,))
        row2 = cur.fetchone()
        if row2 and row:
            cols3 = [d[0] for d in cur.description]
            d1 = {cols3[i]: str(row[i]) if row[i] is not None else None for i in range(len(cols3))}
            d2 = {cols3[i]: str(row2[i]) if row2[i] is not None else None for i in range(len(cols3))}
            res['diferencias_vs_L1_100023558'] = {k: {'sw': d1[k], 'l1': d2[k]} for k in cols3 if d1.get(k) != d2.get(k) and k not in ('NUMEROCOMPROBANTE','CODIGOCLIENTE','RAZONSOCIAL','FECHACOMPROBANTE','TOTAL','IVA1','COMENTARIOS','DIRECCION','TELEFONO','FECHAENTREGA','CODIGOMULTIPLAZO','CODIGOTRANSPORTE','CUIT','CODIGOUSUARIO','FECHAMODIFICACION')}
        c.close()
    except Exception as e:
        res['error'] = str(e)
    return res

@app.get("/debug/sw_np_full/{numero}")
def debug_sw_np_full(numero: str):
    """Muestra TODOS los registros relacionados con un NP SW en todas las tablas relevantes."""
    res = {}
    try:
        c = conn('LATIN1', db=DATABASE_MLT)
        cur = c.cursor()

        # Cabeza
        cur.execute("SELECT * FROM \"CABEZACOMPROBANTES\" WHERE NUMEROCOMPROBANTE=? AND TIPOCOMPROBANTE='NP'", (numero,))
        row = cur.fetchone()
        if row:
            cols = [d[0] for d in cur.description]
            res['cabeza'] = {cols[i]: str(row[i]) if row[i] is not None else None for i in range(len(cols))}
        else:
            res['cabeza'] = 'NO ENCONTRADO'

        # Tablas relacionadas
        tablas_rel = [
            'CAMBIOSESTADOSCOMPROBANTES',
            'CAMPOSDINAMICOSCOMPROBANTES',
            'COMPROBANTESASOCIADOS',
            'COMPROBANTESASIENTOS',
        ]
        for tabla in tablas_rel:
            try:
                cur.execute(f'SELECT * FROM "{tabla}" WHERE NUMEROCOMPROBANTE=?', (numero,))
                rows = cur.fetchall()
                if rows:
                    cols2 = [d[0] for d in cur.description]
                    res[tabla] = [{cols2[i]: str(r[i]) if r[i] is not None else None for i in range(len(cols2))} for r in rows]
                else:
                    res[tabla] = []
            except Exception as ex:
                res[tabla+'_err'] = str(ex)

        c.close()
    except Exception as e:
        res['error'] = str(e)
    return res

@app.get("/debug/sw_np/{numero}")
def debug_sw_np(numero: str):
    """Muestra todos los campos de CABEZACOMPROBANTES y CUERPOCOMPROBANTES para un NP SW."""
    res = {}
    try:
        c = conn('LATIN1', db=DATABASE_MLT)
        cur = c.cursor()
        # Cabeza
        cur.execute("SELECT * FROM \"CABEZACOMPROBANTES\" WHERE NUMEROCOMPROBANTE = ? AND TIPOCOMPROBANTE = 'NP'", (numero,))
        row = cur.fetchone()
        if row:
            cols = [d[0] for d in cur.description]
            res['cabeza'] = {cols[i]: str(row[i]) if row[i] is not None else None for i in range(len(cols))}
        else:
            res['cabeza'] = f'NO ENCONTRADO con numero={numero}'
        # Buscar por aproximación si no encontró
        cur.execute("SELECT FIRST 5 NUMEROCOMPROBANTE, TIPOCOMPROBANTE, CODIGOCLIENTE, FECHACOMPROBANTE, FECHAMODIFICACION FROM \"CABEZACOMPROBANTES\" WHERE TIPOCOMPROBANTE='NP' ORDER BY NUMEROCOMPROBANTE DESC")
        res['ultimos_sw'] = [{'num': str(r[0]), 'tipo': r[1], 'cli': r[2], 'fecha': str(r[3]), 'fechamod': str(r[4]) if r[4] else None} for r in cur.fetchall()]
        # Cuerpo
        cur.execute("SELECT * FROM \"CUERPOCOMPROBANTES\" WHERE NUMEROCOMPROBANTE = ? AND TIPOCOMPROBANTE = 'NP'", (numero,))
        rows = cur.fetchall()
        if rows:
            cols2 = [d[0] for d in cur.description]
            res['items'] = [{cols2[i]: str(rows[j][i]) if rows[j][i] is not None else None for i in range(len(cols2))} for j in range(len(rows))]
        else:
            res['items'] = 'SIN ITEMS'
        c.close()
    except Exception as e:
        res['error'] = str(e)
    return res


@app.post("/clientes/{codigo}/informar-pago")
async def informar_pago(
    codigo: str,
    nombre: str = Form(...),
    vendedor: str = Form(""),
    comentario: str = Form(""),
    comprobante: Optional[UploadFile] = File(None),
):
    """Envía un mail a pagos@ con comentario del vendedor y comprobante adjunto."""
    # Leer adjunto
    adjunto_bytes = None
    adjunto_nombre = None
    adjunto_mime = None
    if comprobante and comprobante.filename:
        adjunto_bytes = await comprobante.read()
        adjunto_nombre = comprobante.filename
        adjunto_mime, _ = mimetypes.guess_type(adjunto_nombre)
        if not adjunto_mime:
            adjunto_mime = 'application/octet-stream'

    # Armar mail
    msg = MIMEMultipart()
    msg['From']    = SMTP_FROM
    msg['To']      = SMTP_TO_PAGOS
    tiene_adjunto = "con comprobante adjunto" if adjunto_bytes else "sin comprobante"
    msg['Subject'] = f"Aviso de Pago — {nombre} (Cód. {codigo}) — {tiene_adjunto}"

    cuerpo = (
        f"Se informa un pago realizado por el cliente.\n"
        f"El importe puede ser parcial o total respecto a las facturas pendientes.\n\n"
        f"Cliente : {nombre}\n"
        f"Código  : {codigo}\n"
        f"Vendedor: {vendedor or '—'}\n\n"
        f"{'Detalle del pago informado por el vendedor:' if comentario else 'El vendedor no ingresó comentario adicional.'}\n"
        + (f"{'—'*50}\n{comentario}\n{'—'*50}\n\n" if comentario else "\n")
        + f"{'✔ Se adjunta comprobante de pago.' if adjunto_bytes else '✘ No se adjuntó comprobante.'}\n"
    )
    msg.attach(MIMEText(cuerpo, 'plain', 'utf-8'))

    if adjunto_bytes:
        main_type, sub_type = adjunto_mime.split('/', 1)
        part = MIMEBase(main_type, sub_type)
        part.set_payload(adjunto_bytes)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment', filename=adjunto_nombre)
        msg.attach(part)

    # Enviar via SSL (puerto 465)
    try:
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=15) as srv:
            srv.ehlo()
            if SMTP_PASS:
                srv.login(SMTP_USER, SMTP_PASS)
            srv.sendmail(SMTP_FROM, [SMTP_TO_PAGOS], msg.as_bytes())
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al enviar email: {e}")

    return {"ok": True}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
