"""
API Vendedores Microbell — FastAPI
Puerto: 8000
"""
import os
import math
import uuid
import smtplib
import mimetypes
import sqlite3
import json
import shutil
import urllib.parse
import urllib.request
import urllib.error
from io import BytesIO
from typing import Optional, List
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from fastapi import FastAPI, HTTPException, Query, File, UploadFile, Form, Depends, Request, BackgroundTasks, Body
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse, RedirectResponse, FileResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
import uvicorn
import threading
import asyncio
import time
import queue
import firebirdsql
from dotenv import load_dotenv
from jose import JWTError, jwt
from datetime import datetime, timedelta
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, HRFlowable
from reportlab.pdfbase import pdfmetrics

# ── Cargar variables de entorno desde .env ────────────────────────────────────
load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))

HOST          = os.getenv('FB_HOST', '127.0.0.1')
PORT          = int(os.getenv('FB_PORT', 3050))
DATABASE      = os.getenv('DB_L1',  'c:/flexxus/DB/DB-Prueba.gdb')
DATABASE_EST  = os.getenv('DB_EST', 'c:/flexxus/DB/DB-EST-Prueba.gdb')
DATABASE_MLT  = os.getenv('DB_MLT', 'c:/flexxus/DB/DB-MLT-Prueba.gdb')
DB_USER       = os.getenv('FB_USER', 'SYSDBA')
DB_PASS       = os.getenv('FB_PASS', '')

# ── SMTP Microbell ─────────────────────────────────────────────────────────────
SMTP_HOST     = os.getenv('SMTP_HOST', '')
SMTP_PORT     = int(os.getenv('SMTP_PORT', 587))
SMTP_USER     = os.getenv('SMTP_USER', '')
SMTP_PASS     = os.getenv('SMTP_PASS', '')
SMTP_FROM     = os.getenv('SMTP_FROM', '')

SMTP_TO_PAGOS = os.getenv('SMTP_TO_PAGOS', '')

# ── OneSignal Push Notifications ─────────────────────────────────────────────
ONESIGNAL_APP_ID  = os.getenv('ONESIGNAL_APP_ID', '')
ONESIGNAL_API_KEY = os.getenv('ONESIGNAL_API_KEY', '')

# ── WhatsApp Business (Meta Cloud API) ────────────────────────────────────────
WA_PHONE_NUMBER_ID = os.getenv('WHATSAPP_PHONE_ID') or os.getenv('WA_PHONE_NUMBER_ID', '')
WA_ACCESS_TOKEN    = os.getenv('WHATSAPP_TOKEN') or os.getenv('WA_ACCESS_TOKEN', '')
WA_WABA_ID         = os.getenv('WHATSAPP_WABA_ID') or os.getenv('WA_WABA_ID', '')        # para crear plantillas
WA_TEMPLATE_CAT    = os.getenv('WA_TEMPLATE_CAT', 'microbell_catalogo')   # nombre plantilla catálogo
WA_TEMPLATE_SLIDE  = os.getenv('WA_TEMPLATE_SLIDE', 'microbell_catalogo') # nombre plantilla slide (puede ser la misma)
WA_TEMPLATE_COBRANZAS = os.getenv('WA_TEMPLATE_COBRANZAS', 'microbell_cobranzas_v1')  # plantilla aviso a Cobranzas
WA_COBRANZAS_CEL   = os.getenv('WA_COBRANZAS_CEL', '5491168561985')  # celular Área de Cobranzas
WA_TEMPLATE_REACTIVACION = os.getenv('WA_TEMPLATE_REACTIVACION', 'microbell_reactivacion_v1')  # plantilla reactivación de clientes

# ── Catálogos ──────────────────────────────────────────────────────────────────
_BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
CATALOGOS_DIR = os.path.join(_BASE_DIR, os.getenv('CATALOGOS_DIR', 'catalogos'))
os.makedirs(CATALOGOS_DIR, exist_ok=True)
# ──────────────────────────────────────────────────────────────────────────────

# ── Pool de conexiones Firebird ────────────────────────────────────────────────
# Antes conn() abría una conexión TCP nueva en CADA llamada y el resto del código la
# cierra con c.close() al terminar (225 usos en este archivo). Abrir/cerrar la conexión
# de red contra Firebird es lo que generaba los ~2seg de demora por artículo agregado,
# no el cálculo del stock (eso ya estaba cacheado). La solución: conn() ahora devuelve
# un wrapper que se comporta EXACTAMENTE igual (mismo .cursor()/.commit()/.rollback()/
# .close()) pero .close() no cierra el socket — hace rollback() de cualquier
# transacción sin confirmar (mismo efecto que tenía cerrar la conexión antes) y la
# devuelve a un pool para reusarla en el próximo request. Ningún otro lugar del código
# necesita cambiar. DB_POOL_ENABLED=0 en .env desactiva el pool y vuelve al
# comportamiento anterior (conexión nueva por request) sin tocar código, por si hiciera
# falta revertir rápido en producción.
_DB_POOL_ENABLED = os.getenv('DB_POOL_ENABLED', '1') == '1'
_DB_POOL_SIZE     = int(os.getenv('DB_POOL_SIZE', 8))   # máx. conexiones ociosas guardadas por (db, charset)
_db_pools: dict         = {}   # (db_path, charset) -> queue.Queue de conexiones firebirdsql "crudas"
_db_pool_created: dict  = {}   # (db_path, charset) -> cuántas conexiones vivas hay creadas (para logs/diagnóstico)
_db_pools_lock = threading.Lock()

def _db_pool_get_queue(key):
    with _db_pools_lock:
        pool = _db_pools.get(key)
        if pool is None:
            pool = queue.Queue(maxsize=_DB_POOL_SIZE)
            _db_pools[key] = pool
            _db_pool_created[key] = 0
        return pool

def _db_pool_dec(key):
    with _db_pools_lock:
        _db_pool_created[key] = max(0, _db_pool_created.get(key, 1) - 1)

class _PooledConnection:
    """Wrapper transparente sobre una conexión firebirdsql real. Todo lo que no sea
    close() se delega tal cual (cursor, commit, rollback, etc.) — el resto del código
    no nota la diferencia."""
    def __init__(self, raw, key):
        self._raw = raw
        self._key = key
        self._broken = False

    def cursor(self, *a, **kw):
        try:
            return self._raw.cursor(*a, **kw)
        except Exception:
            self._broken = True
            raise

    def commit(self, *a, **kw):
        try:
            return self._raw.commit(*a, **kw)
        except Exception:
            self._broken = True
            raise

    def rollback(self, *a, **kw):
        try:
            return self._raw.rollback(*a, **kw)
        except Exception:
            self._broken = True
            raise

    def close(self):
        if not _DB_POOL_ENABLED:
            try: self._raw.close()
            except Exception: pass
            return
        raw, key = self._raw, self._key
        if self._broken:
            try: raw.close()
            except Exception: pass
            _db_pool_dec(key)
            return
        try:
            raw.rollback()  # limpia transacción abierta sin confirmar, igual que antes al cerrar
        except Exception:
            try: raw.close()
            except Exception: pass
            _db_pool_dec(key)
            return
        pool = _db_pool_get_queue(key)
        try:
            pool.put_nowait(raw)
        except queue.Full:
            try: raw.close()
            except Exception: pass
            _db_pool_dec(key)

    def __getattr__(self, name):
        return getattr(self._raw, name)

def conn(charset='WIN1252', db=None):
    db_path = db or DATABASE
    if not _DB_POOL_ENABLED:
        return firebirdsql.connect(host=HOST, port=PORT, database=db_path,
                                   user=DB_USER, password=DB_PASS, charset=charset)
    key = (db_path, charset)
    pool = _db_pool_get_queue(key)
    # 1) Intentar reusar una conexión ociosa del pool, verificando que siga viva
    while True:
        try:
            raw = pool.get_nowait()
        except queue.Empty:
            break
        try:
            _hc = raw.cursor()
            _hc.execute('SELECT 1 FROM RDB$DATABASE')
            _hc.fetchone()
            return _PooledConnection(raw, key)
        except Exception:
            try: raw.close()
            except Exception: pass
            _db_pool_dec(key)
            continue  # probar la siguiente ociosa, si hay más
    # 2) No había ninguna reusable: crear una nueva
    raw = firebirdsql.connect(host=HOST, port=PORT, database=db_path,
                              user=DB_USER, password=DB_PASS, charset=charset)
    with _db_pools_lock:
        _db_pool_created[key] = _db_pool_created.get(key, 0) + 1
    return _PooledConnection(raw, key)

# ── Cache FMA_STOCK ───────────────────────────────────────────────────────────
# TTL configurable via .env: FMA_CACHE_TTL=90 (segundos). 0 = sin caché.
_FMA_CACHE_TTL   = int(os.getenv('FMA_CACHE_TTL', 90))   # default 90 segundos (1.5 min)
_FMA_ALL_DEPS    = ['001', '002', '003', '005', '013', '016']
_fma_cache: dict = {}
_fma_cache_lock  = threading.Lock()

def _fma_stock_bulk(dep: str, charset='WIN1252') -> dict:
    """Devuelve {ID_ARTICULO: STOCKREMANENTE} para un depósito con caché TTL."""
    if _FMA_CACHE_TTL > 0:
        with _fma_cache_lock:
            entry = _fma_cache.get(dep)
        if entry and (time.time() - entry[0]) < _FMA_CACHE_TTL:
            return entry[1]
    try:
        c2 = conn(charset)
        cur2 = c2.cursor()
        cur2.execute(f'SELECT ID_ARTICULO, STOCKREMANENTE FROM "FMA_STOCK"(NULL, NULL, \'{dep}\', 1, 1)')
        data = {row[0]: float(row[1] or 0) for row in cur2.fetchall()}
        c2.close()
    except Exception:
        data = {}
    if _FMA_CACHE_TTL > 0:
        with _fma_cache_lock:
            _fma_cache[dep] = (time.time(), data)
    return data

def _fma_stock_parallel(deps: list, charset='WIN1252') -> dict:
    """Ejecuta _fma_stock_bulk para varios depósitos en paralelo."""
    from concurrent.futures import ThreadPoolExecutor
    result = {}
    with ThreadPoolExecutor(max_workers=max(len(deps), 1)) as ex:
        futures = {ex.submit(_fma_stock_bulk, dep, charset): dep for dep in deps}
        for fut, dep in futures.items():
            try:
                result[dep] = fut.result()
            except Exception:
                result[dep] = {}
    return result

def _fma_cache_invalidate(deps: list = None):
    """Invalida el caché de los depósitos indicados (o todos si deps=None)."""
    with _fma_cache_lock:
        if deps is None:
            _fma_cache.clear()
        else:
            for d in deps:
                _fma_cache.pop(d, None)

def _prewarm_fma_cache():
    """Precalienta el caché FMA_STOCK en background al arrancar el servidor y lo
    mantiene tibio con refresco recurrente (antes de que expire el TTL) — así ninguna
    búsqueda de stock paga el costo de una consulta en frío a Firebird, ni siquiera
    tras minutos de inactividad."""
    if _FMA_CACHE_TTL <= 0:
        return
    time.sleep(4)   # espera que el servidor termine de arrancar
    _intervalo = max(10, _FMA_CACHE_TTL - 15)
    while True:
        try:
            _fma_stock_parallel(_deps_activos())   # todos los depósitos activos, en paralelo
        except Exception:
            pass        # si falla el pre-calentamiento, no tumbar el servidor
        time.sleep(_intervalo)

threading.Thread(target=_prewarm_fma_cache, daemon=True).start()

# ── Cache Catálogo Artículos ──────────────────────────────────────────────────
# TTL configurable via .env: CATALOG_CACHE_TTL=1800 (segundos). 0 = sin caché.
_CATALOG_CACHE_TTL  = int(os.getenv('CATALOG_CACHE_TTL', 1800))  # default 30 min
_catalog_cache: dict = {}       # {CODIGOARTICULO: {...campos...}}
_catalog_cache_ts: float = 0.0
_catalog_cache_lock = threading.Lock()
_catalog_cambio_usd: float = 1.0

def _s(v):
    """Convierte cualquier valor Firebird a str limpio."""
    if v is None: return ''
    return str(v).strip()

def _sin_tildes(s):
    """Quita acentos/diacríticos (NFKD) — usado para normalizar texto de búsqueda
    (tolera tildes que cambian con reconocimiento de voz) y, en el catálogo en
    memoria, para precomputar campos de búsqueda equivalentes a la colación
    insensible a acentos que ya usa Firebird en CONTAINING/UPPER."""
    import unicodedata as _ud
    return ''.join(c for c in _ud.normalize('NFKD', s or '') if not _ud.combining(c))

def _redondear_precio(v, factor=1.0):
    """Política de redondeo de importes unitarios: entero más cercano (sin decimales),
    tras aplicar la conversión de moneda si corresponde. Única fuente de verdad —
    usada en listados/búsqueda de stock, informes (Excel/PDF) y carga de precio
    unitario en pedidos/presupuestos."""
    if not v: return 0
    return round(float(v) * factor)

# ── Tabla de Comisiones de Vendedores ────────────────────────────────────────
# Cargada desde comisiones.json (generado desde la tabla Paradox).
# Clave: (vendedor_upper, gruposuperrubro_upper, superrubro_upper, rubro_upper)
# Valor: porcentaje de comisión (float)
_COMISIONES: dict = {}

def _load_comisiones():
    """Carga tabla comisiones_vendedores desde admin.db en memoria."""
    global _COMISIONES
    import sqlite3 as _sq3
    db_path = os.path.join(os.path.dirname(__file__), 'admin.db')
    if not os.path.exists(db_path):
        return
    try:
        _db = _sq3.connect(db_path)
        rows = _db.execute(
            'SELECT vendedor, gruposuperrubro, superrubro, rubro, porcentaje '
            'FROM comisiones_vendedores'
        ).fetchall()
        _db.close()
        _COMISIONES = {(r[0], r[1], r[2], r[3]): r[4] for r in rows}
    except Exception:
        pass

_load_comisiones()

def _load_catalog(charset='WIN1252') -> tuple:
    """Carga todos los artículos activos con jerarquía completa desde Firebird."""
    c = conn(charset)
    cur = c.cursor()
    try:
        cur.execute('SELECT CAMBIO FROM "MONEDAS" WHERE CODIGOMONEDA=?', ('DOLARES',))
        row_m = cur.fetchone()
        cambio_usd = float(row_m[0]) if row_m else 1.0
    except Exception:
        cambio_usd = 1.0
    cur.execute("""
        SELECT
            a.CODIGOARTICULO, a.CODIGOPARTICULAR, a.DESCRIPCION, a.CODIGOMARCA,
            a.PRECIOLISTA1, a.PRECIOLISTA2, a.PRECIOLISTA3, a.PRECIOLISTA5,
            a.ALICUOTAIVA, a.COEFICIENTE, a.CODIGOUNIDADMEDIDA, a.CODIGOMONEDA,
            a.CODIGORUBRO,
            r.DESCRIPCION, r.CODIGOSUPERRUBRO,
            sr.DESCRIPCION, sr.CODIGOGRUPOSUPERRUBRO,
            g.DESCRIPCION,
            a.COEFICIENTESEGUNRUBRO, r.COEFICIENTE,
            a.DTOMAXIMO1, a.APLICABLEABONIFICACION, a.PERMITESTOCKNEGATIVO
        FROM "ARTICULOS" a
        LEFT JOIN "RUBROS" r ON r.CODIGORUBRO = a.CODIGORUBRO
        LEFT JOIN "SUPERRUBROS" sr ON sr.CODIGOSUPERRUBRO = r.CODIGOSUPERRUBRO
        LEFT JOIN "GRUPOSUPERRUBROS" g ON g.CODIGOGRUPOSUPERRUBRO = sr.CODIGOGRUPOSUPERRUBRO
        WHERE a.ACTIVO = '1'
          AND (g.DESCRIPCION IS NULL OR UPPER(g.DESCRIPCION) NOT IN ('TERCEROS','SERVICIOS'))
    """)
    catalog = {}
    for row in cur.fetchall():
        art_id = row[0]
        # IVA real (validado contra Flexxus): si COEFICIENTESEGUNRUBRO=1, el
        # coeficiente aplicado es el del RUBRO (r.COEFICIENTE); si =0 (Manual),
        # es el propio a.COEFICIENTE. ALICUOTAIVA NO se usa: siempre viene 0.
        _coef_manual = float(row[9] or 0)
        _coef_rubro = float(row[19] or 0)
        _usa_rubro = _s(row[18]).strip() == '1'
        _coef_final = _coef_rubro if _usa_rubro else _coef_manual
        _cod_particular = _s(row[1]) or _s(art_id)
        _descripcion    = _s(row[2])
        catalog[art_id] = {
            'codigo':                 art_id,
            'codigoparticular':       _cod_particular,
            'descripcion':            _descripcion,
            'codigomarca':            _s(row[3]),
            'precio1':                float(row[4] or 0),
            'precio2':                float(row[5] or 0),
            'precio3':                float(row[6] or 0),
            'precio5':                float(row[7] or 0),
            'alicuotaiva':            row[8],
            'coeficiente':            _coef_manual,
            'coeficiente_segun_rubro': _usa_rubro,
            'rubro_coeficiente':      _coef_rubro,
            'iva':                    round(_coef_final * 21, 2),
            'unidad':                 _s(row[10]),
            'codigomoneda':           _s(row[11]).upper(),
            'codigo_rubro':           _s(row[12]),
            'rubro':                  _s(row[13]),
            'codigo_superrubro':      _s(row[14]),
            'superrubro':             _s(row[15]),
            'codigo_gruposuperrubro': _s(row[16]),
            'gruposuperrubro':        _s(row[17]),
            'dtomaximo1_raw':             row[20],
            'aplicableabonificacion_raw': row[21],
            'permitestocknegativo_raw':   row[22],
            # Campos precomputados para búsqueda de texto en memoria (/buscar-articulos)
            # — evita pagar un CONTAINING en vivo contra Firebird en cada letra escrita.
            # Sin acentos + mayúsculas, igual que la colación insensible a acentos que
            # ya usa Firebird acá.
            '_cn': _sin_tildes(_cod_particular).upper(),
            '_dn': _sin_tildes(_descripcion).upper(),
        }
    c.close()
    return catalog, cambio_usd

def _get_catalog() -> tuple:
    """Devuelve (catalog_dict, cambio_usd) desde caché o recarga si TTL expiró."""
    global _catalog_cache, _catalog_cache_ts, _catalog_cambio_usd
    now = time.time()
    if _CATALOG_CACHE_TTL > 0 and _catalog_cache and (now - _catalog_cache_ts) < _CATALOG_CACHE_TTL:
        return _catalog_cache, _catalog_cambio_usd
    with _catalog_cache_lock:
        now = time.time()
        if _CATALOG_CACHE_TTL > 0 and _catalog_cache and (now - _catalog_cache_ts) < _CATALOG_CACHE_TTL:
            return _catalog_cache, _catalog_cambio_usd
        try:
            cat, usd = _load_catalog()
            _catalog_cache = cat
            _catalog_cache_ts = time.time()
            _catalog_cambio_usd = usd
        except Exception as _e:
            if not _catalog_cache:
                raise  # primera carga fallida: propagar para que el endpoint devuelva 500
            # recarga fallida pero hay caché viejo: lo usamos
    return _catalog_cache, _catalog_cambio_usd

def _catalog_invalidate():
    """Fuerza recarga del catálogo en la próxima consulta."""
    global _catalog_cache_ts
    with _catalog_cache_lock:
        _catalog_cache_ts = 0.0

# ── Cache Combos Filtro (GSR/SR/Rubro/Marca/Depósitos) ───────────────────────
# Evita ir a Firebird cada vez que se abre la sección Stock (admin.html) o el
# formulario de pedido/presupuesto (frontend.html) solo para llenar combos.
# Mismo TTL que el catálogo (30 min por defecto) — estas listas cambian poco.
_filtros_cache: dict = {}
_filtros_cache_ts: float = 0.0
_filtros_cache_lock = threading.Lock()

def _load_filtros_combos() -> dict:
    c = conn('WIN1252'); cur = c.cursor()
    cur.execute("""
        SELECT DISTINCT g.CODIGOGRUPOSUPERRUBRO, g.DESCRIPCION
        FROM "GRUPOSUPERRUBROS" g
        JOIN "SUPERRUBROS" sr ON sr.CODIGOGRUPOSUPERRUBRO = g.CODIGOGRUPOSUPERRUBRO
        JOIN "RUBROS" r ON r.CODIGOSUPERRUBRO = sr.CODIGOSUPERRUBRO
        JOIN "ARTICULOS" a ON a.CODIGORUBRO = r.CODIGORUBRO
        WHERE a.ACTIVO = '1'
          AND UPPER(g.DESCRIPCION) NOT IN ('TERCEROS','SERVICIOS')
        ORDER BY g.DESCRIPCION
    """)
    gsr = [{"codigo": str(r[0] or '').strip(), "descripcion": str(r[1] or '').strip()} for r in cur.fetchall()]

    cur.execute('SELECT CODIGOSUPERRUBRO, DESCRIPCION FROM "SUPERRUBROS" ORDER BY DESCRIPCION')
    sr = [{"codigo": str(r[0] or '').strip(), "descripcion": str(r[1] or '').strip()} for r in cur.fetchall()]

    cur.execute('SELECT CODIGORUBRO, DESCRIPCION FROM "RUBROS" ORDER BY DESCRIPCION')
    rubro = [{"codigo": str(r[0] or '').strip(), "descripcion": str(r[1] or '').strip()} for r in cur.fetchall()]

    cur.execute("""
        SELECT DISTINCT m.CODIGOMARCA, m.DESCRIPCION
        FROM "MARCAS" m
        JOIN "ARTICULOS" a ON a.CODIGOMARCA = m.CODIGOMARCA
        WHERE a.ACTIVO = '1'
        ORDER BY m.DESCRIPCION
    """)
    marca = [{"codigo": str(r[0] or '').strip(), "descripcion": str(r[1] or '').strip()} for r in cur.fetchall()]

    cur.execute('SELECT CODIGODEPOSITO, DESCRIPCION FROM "DEPOSITOS" WHERE ACTIVO=1 ORDER BY CODIGODEPOSITO')
    depositos = [{"codigo": str(r[0] or '').strip(), "nombre": str(r[1] or '').strip()}
                 for r in cur.fetchall() if str(r[0] or '').strip()]
    c.close()
    return {'gsr': gsr, 'sr': sr, 'rubro': rubro, 'marca': marca, 'depositos': depositos}

def _get_filtros_combos() -> dict:
    """Devuelve los combos de filtro desde caché o recarga si el TTL expiró."""
    global _filtros_cache, _filtros_cache_ts
    now = time.time()
    if _filtros_cache and (now - _filtros_cache_ts) < _CATALOG_CACHE_TTL:
        return _filtros_cache
    with _filtros_cache_lock:
        now = time.time()
        if _filtros_cache and (now - _filtros_cache_ts) < _CATALOG_CACHE_TTL:
            return _filtros_cache
        try:
            data = _load_filtros_combos()
            _filtros_cache = data
            _filtros_cache_ts = time.time()
        except Exception:
            if not _filtros_cache:
                raise
    return _filtros_cache

def _filtros_invalidate():
    """Fuerza recarga de los combos de filtro en la próxima consulta."""
    global _filtros_cache_ts
    with _filtros_cache_lock:
        _filtros_cache_ts = 0.0

# _deps_activos: lista de códigos de depósito ACTIVOS (3 dígitos), leída dinámicamente
# de la tabla DEPOSITOS de Firebird vía _get_filtros_combos (ya cacheada, TTL igual al
# catálogo). Reemplaza las listas fijas de 6 depósitos que /stock, /stock/batch y
# /stock/{codigo} tenían hardcodeadas (001,002,003,005,013,016) — cualquier depósito
# agregado después (ej. 017 SARANDI) quedaba afuera para siempre, así que un vendedor
# con ese depósito habilitado veía remanente 0 aunque el stock real fuera >0 (bug
# reportado 2026-07-31: Krafft Ariel con depósito SARANDI, remanente real 51, el
# frontend mostraba 0 porque el JSON de /stock nunca traía la clave "remanente_017").
# Con esto, si mañana se crea un depósito 018, aparece solo con activarlo en Flexxus.
def _deps_activos():
    try:
        deps = _get_filtros_combos().get('depositos') or []
        result = sorted({str(d['codigo']).strip().zfill(3) for d in deps if d.get('codigo')})
        if result:
            return result
    except Exception:
        pass
    return list(_FMA_ALL_DEPS)  # fallback si la consulta a DEPOSITOS falla

def _depositos_arma_pedidos_map():
    """Devuelve {codigo: bool} — si un depósito no tiene fila explícita en
    depositos_config (SQLite admin.db), se asume arma_pedidos=True (opt-out: nada
    cambia hasta que Eduardo desmarque explícitamente un depósito administrativo/
    logístico, ej. SCRAP, MARKET PLACE, RMA, DESTRUCCION TOTAL). Usado por la
    sugerencia de transferencia automática en el popup de stock del Pedido — NO
    afecta ninguna otra vista de remanente (Presupuesto, Stock por Depósito, etc.)."""
    result = {}
    try:
        db = _admin_db()
        for row in db.execute("SELECT codigo, arma_pedidos FROM depositos_config"):
            result[str(row['codigo']).strip().zfill(3)] = bool(row['arma_pedidos'])
        db.close()
    except Exception:
        pass
    return result

# ── Cache Combos Filtro PÚBLICOS (frontend.html — /gruposuperrubros, /superrubros,
# /rubros, /marcas) ───────────────────────────────────────────────────────────
# Se piden los 4 juntos al hacer login del vendedor (cargarGruposSuperRubro), antes
# de tocar la sección Stock. Mismo TTL que el catálogo (30 min).
_pub_filtros_cache: dict = {}
_pub_filtros_cache_ts: float = 0.0
_pub_filtros_cache_lock = threading.Lock()

def _load_pub_filtros() -> dict:
    c = conn(); cur = c.cursor()
    cur.execute("""
        SELECT DISTINCT g.CODIGOGRUPOSUPERRUBRO, g.DESCRIPCION
        FROM "GRUPOSUPERRUBROS" g
        WHERE UPPER(g.DESCRIPCION) NOT IN ('TERCEROS','SERVICIOS')
          AND EXISTS (
            SELECT 1 FROM "ARTICULOS" a
            JOIN "RUBROS" r ON r.CODIGORUBRO = a.CODIGORUBRO
            JOIN "SUPERRUBROS" sr ON sr.CODIGOSUPERRUBRO = r.CODIGOSUPERRUBRO
            WHERE sr.CODIGOGRUPOSUPERRUBRO = g.CODIGOGRUPOSUPERRUBRO
              AND a.ACTIVO = '1'
        )
        ORDER BY g.DESCRIPCION
    """)
    gsr = [{"codigo": r[0], "descripcion": r[1]} for r in cur.fetchall()]

    cur.execute('SELECT CODIGOSUPERRUBRO, DESCRIPCION, CODIGOGRUPOSUPERRUBRO FROM "SUPERRUBROS" ORDER BY DESCRIPCION')
    sr = [{"codigo": r[0], "descripcion": r[1], "grupo": r[2] if len(r) > 2 else None} for r in cur.fetchall()]

    cur.execute("""
        SELECT DISTINCT r.CODIGORUBRO, r.DESCRIPCION, r.CODIGOSUPERRUBRO
        FROM "RUBROS" r
        JOIN "SUPERRUBROS" sr ON sr.CODIGOSUPERRUBRO = r.CODIGOSUPERRUBRO
        WHERE EXISTS (
            SELECT 1 FROM "ARTICULOS" a
            WHERE a.CODIGORUBRO = r.CODIGORUBRO AND a.ACTIVO = '1'
        )
        ORDER BY r.DESCRIPCION
    """)
    rubro = [{"codigo": r[0], "descripcion": r[1], "superrubro": r[2]} for r in cur.fetchall()]

    cur.execute("""
        SELECT DISTINCT m.CODIGOMARCA, m.DESCRIPCION
        FROM "MARCAS" m
        WHERE EXISTS (
            SELECT 1 FROM "ARTICULOS" a
            WHERE a.CODIGOMARCA = m.CODIGOMARCA AND a.ACTIVO = '1'
        )
        ORDER BY m.DESCRIPCION
    """)
    marca = [{"codigo": r[0], "descripcion": r[1]} for r in cur.fetchall()]
    c.close()
    return {'gsr': gsr, 'sr': sr, 'rubro': rubro, 'marca': marca}

def _get_pub_filtros() -> dict:
    global _pub_filtros_cache, _pub_filtros_cache_ts
    now = time.time()
    if _pub_filtros_cache and (now - _pub_filtros_cache_ts) < _CATALOG_CACHE_TTL:
        return _pub_filtros_cache
    with _pub_filtros_cache_lock:
        now = time.time()
        if _pub_filtros_cache and (now - _pub_filtros_cache_ts) < _CATALOG_CACHE_TTL:
            return _pub_filtros_cache
        try:
            data = _load_pub_filtros()
            _pub_filtros_cache = data
            _pub_filtros_cache_ts = time.time()
        except Exception:
            if not _pub_filtros_cache:
                raise
    return _pub_filtros_cache

def _pub_filtros_invalidate():
    global _pub_filtros_cache_ts
    with _pub_filtros_cache_lock:
        _pub_filtros_cache_ts = 0.0

def _search_stock_cache(
    buscar=None, gruposuperrubro=None, superrubro=None, rubro=None, marca=None,
    dep_lista=None, limit=100, offset=0, cambio_usd_override=None
):
    """Búsqueda de stock en memoria combinando catálogo + FMA cache.
    Retorna (pagina, total, cambio_usd)."""
    if dep_lista is None:
        dep_lista = ['001', '003']

    catalog, cambio_usd = _get_catalog()
    if cambio_usd_override is not None:
        cambio_usd = cambio_usd_override

    # Asegurar caché FMA actualizado para los depósitos pedidos
    _fma_stock_parallel(dep_lista)
    with _fma_cache_lock:
        dep_caches = {dep: (_fma_cache.get(dep) or (0, {}))[1] for dep in dep_lista}

    buscar_norm = None
    if buscar:
        buscar = _sanitizar_buscar(buscar)
        buscar_norm = buscar.upper()

    resultados = []
    for art_id, art in catalog.items():
        # Filtro texto
        if buscar_norm:
            desc_up = art['descripcion'].upper()
            cod_up  = art['codigoparticular'].upper()
            if buscar_norm not in desc_up and buscar_norm not in cod_up:
                continue
        # Filtros jerarquía
        if rubro           and art['codigo_rubro']            != rubro:           continue
        if superrubro      and art['codigo_superrubro']       != superrubro:      continue
        if gruposuperrubro and art['codigo_gruposuperrubro']  != gruposuperrubro: continue
        if marca           and art['codigomarca']             != marca:           continue

        # Stock remanente en los depósitos solicitados
        rem_dep = {dep: dep_caches[dep].get(art_id, 0) for dep in dep_lista}
        rem_total = sum(rem_dep.values())
        if rem_total <= 0:
            continue

        resultados.append((art, rem_dep, rem_total))

    # Orden igual al SQL original
    resultados.sort(key=lambda x: x[0]['codigoparticular'])
    total = len(resultados)
    return resultados[offset:offset + limit], total, cambio_usd

def _prewarm_catalog():
    """Precalienta el catálogo en background al arrancar y lo mantiene tibio con
    refresco recurrente antes de que expire el TTL."""
    if _CATALOG_CACHE_TTL <= 0:
        return
    time.sleep(8)  # después del FMA prewarm
    _intervalo = max(30, _CATALOG_CACHE_TTL - 60)
    while True:
        try:
            _get_catalog()
        except Exception:
            pass
        time.sleep(_intervalo)

threading.Thread(target=_prewarm_catalog, daemon=True).start()

# ── Debug global: captura errores y conteos de _query_db ─────────────────────
_QV_LAST_ERRORS: dict = {}
_QV_LAST_COUNTS: dict = {}

app = FastAPI(title="API Vendedores Microbell")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
app.mount("/static", StaticFiles(directory="static"), name="static")
os.makedirs("catalogo_imagenes", exist_ok=True)  # evita crash si aún no se copió el contenido sincronizado
app.mount("/catalogo-imagenes", StaticFiles(directory="catalogo_imagenes"), name="catalogo_imagenes")

from meli import router as meli_router
app.include_router(meli_router)

from import_budget import router as import_budget_router
app.include_router(import_budget_router)

from import_research import router as import_research_router
app.include_router(import_research_router)

# ── JWT ───────────────────────────────────────────────────────────────────────
JWT_SECRET = os.getenv('JWT_SECRET_KEY', 'dev-secret-CAMBIAR')
JWT_ALGO   = os.getenv('JWT_ALGORITHM', 'HS256')
JWT_HOURS  = int(os.getenv('JWT_EXPIRE_HOURS', 10))
_bearer    = HTTPBearer(auto_error=False)

def _create_token(data: dict) -> str:
    payload = data.copy()
    payload['exp'] = datetime.utcnow() + timedelta(hours=JWT_HOURS)
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)

def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(_bearer)):
    if not credentials:
        raise HTTPException(401, "No autenticado", headers={"WWW-Authenticate": "Bearer"})
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGO])
        if not payload.get('sub'):
            raise HTTPException(401, "Token inválido")
        return payload
    except JWTError:
        raise HTTPException(401, "Token inválido o expirado", headers={"WWW-Authenticate": "Bearer"})

class _LoginReq(BaseModel):
    usuario: str
    password: str

class _CambiarPassReq(BaseModel):
    usuario: str
    password_actual: str
    nueva_password: str

@app.post("/auth/login")
def auth_login(req: _LoginReq):
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        cur.execute(
            'SELECT CODIGOUSUARIO, RAZONSOCIAL, CODIGOPERFIL, ESVENDEDOR, ACTIVO, '
            'BONIFICACIONMAXIMA, PORCENTAJEINCREMENTOPRECIO, PORCENTAJEDECREMENTOPRECIO '
            'FROM "USUARIOS" WHERE UPPER(CODIGOUSUARIO)=? AND UPPER(PASSWORD1)=?',
            (req.usuario.upper(), req.password.upper())
        )
        row = cur.fetchone()
        c.close()
    except Exception as e:
        raise HTTPException(500, f"Error DB: {e}")
    if not row or str(row[4] or '').strip() != '1':
        raise HTTPException(401, "Usuario o contraseña incorrectos")
    cod    = str(row[0] or '').strip()
    razon  = str(row[1] or '').strip()
    perfil = str(row[2] or '').strip()
    esvend = str(row[3] or '0').strip()
    bonif_max   = float(row[5]) if row[5] is not None else 0.0
    pct_inc     = float(row[6]) if row[6] is not None else 0.0
    pct_dec     = float(row[7]) if row[7] is not None else 0.0
    _PERFILES_OK = {'VENDEDORES', 'ADV', 'ADVJUAN', 'GERENTES', 'GTES FE'}
    if perfil not in _PERFILES_OK:
        raise HTTPException(403, "Sin acceso: perfil no autorizado")
    token = _create_token({'sub': cod, 'nombre': razon, 'perfil': perfil, 'esvendedor': esvend,
                           'bonificacion_maxima': bonif_max, 'pct_incremento': pct_inc, 'pct_decremento': pct_dec})
    return {"codigousuario": cod, "razonsocial": razon, "perfil": perfil,
            "esvendedor": esvend, "token": token,
            "bonificacion_maxima": bonif_max, "pct_incremento": pct_inc, "pct_decremento": pct_dec}

@app.post("/auth/cambiar-password")
def auth_cambiar_password(req: _CambiarPassReq, user=Depends(get_current_user)):
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        cur.execute(
            'SELECT CODIGOUSUARIO FROM "USUARIOS" '
            'WHERE UPPER(CODIGOUSUARIO)=? AND UPPER(PASSWORD1)=? AND ACTIVO=?',
            (req.usuario.upper(), req.password_actual.upper(), '1')
        )
        if not cur.fetchone():
            c.close()
            raise HTTPException(401, "Contraseña actual incorrecta")
        cur.execute(
            'UPDATE "USUARIOS" SET PASSWORD1=? WHERE UPPER(CODIGOUSUARIO)=?',
            (req.nueva_password.upper(), req.usuario.upper())
        )
        c.commit()
        c.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error DB: {e}")
    return {"ok": True}
# ─────────────────────────────────────────────────────────────────────────────

FRONTEND_PATH = os.path.join(os.path.dirname(__file__), "frontend.html")
ADMIN_PATH    = os.path.join(os.path.dirname(__file__), "admin.html")
LOGO_PATH     = os.path.join(os.path.dirname(__file__), "microbellSA-color.png")
FAVICON_PATH  = os.path.join(os.path.dirname(__file__), "favicon.ico")
ADMIN_DB_PATH = os.path.join(os.path.dirname(__file__), "admin.db")

# ─── Perfiles autorizados para el Control Panel ───────────────────────────────
_ADMIN_PERFILES = {'ADV', 'DISENO', 'GERENTES', 'GTES FE', 'ADVJUAN'}

# ─── SQLite: inicialización ───────────────────────────────────────────────────
def _admin_db():
    c = sqlite3.connect(ADMIN_DB_PATH)
    c.row_factory = sqlite3.Row
    return c

def _init_admin_db():
    c = _admin_db()
    cur = c.cursor()
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS vendor_profiles (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo TEXT UNIQUE NOT NULL,
        nombre TEXT NOT NULL,
        activo INTEGER DEFAULT 1,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS vendor_profile_assignments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigousuario TEXT NOT NULL,
        profile_id INTEGER NOT NULL,
        FOREIGN KEY (profile_id) REFERENCES vendor_profiles(id) ON DELETE CASCADE,
        UNIQUE(codigousuario, profile_id)
    );
    CREATE TABLE IF NOT EXISTS feature_flags (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigousuario TEXT,
        feature TEXT NOT NULL,
        enabled INTEGER DEFAULT 1,
        UNIQUE(codigousuario, feature)
    );
    CREATE TABLE IF NOT EXISTS multiplazos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        dias TEXT NOT NULL,
        activo INTEGER DEFAULT 1
    );
    CREATE TABLE IF NOT EXISTS vendor_multiplazos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigousuario TEXT,
        multiplazo_id INTEGER NOT NULL,
        FOREIGN KEY (multiplazo_id) REFERENCES multiplazos(id) ON DELETE CASCADE,
        UNIQUE(codigousuario, multiplazo_id)
    );
    CREATE TABLE IF NOT EXISTS vendor_multiplazos_fb (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigousuario TEXT NOT NULL,
        codigo_multiplazo TEXT NOT NULL,
        UNIQUE(codigousuario, codigo_multiplazo)
    );
    CREATE TABLE IF NOT EXISTS catalogs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        descripcion TEXT,
        url TEXT NOT NULL,
        activo INTEGER DEFAULT 1,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS catalog_profiles (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        catalog_id INTEGER NOT NULL,
        profile_id INTEGER NOT NULL,
        FOREIGN KEY (catalog_id) REFERENCES catalogs(id) ON DELETE CASCADE,
        FOREIGN KEY (profile_id) REFERENCES vendor_profiles(id) ON DELETE CASCADE,
        UNIQUE(catalog_id, profile_id)
    );
    CREATE TABLE IF NOT EXISTS offers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        tipo TEXT NOT NULL,
        descripcion TEXT,
        fecha_desde TEXT,
        fecha_hasta TEXT,
        activo INTEGER DEFAULT 1,
        created_at TEXT DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS offer_product_details (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        offer_id INTEGER NOT NULL,
        codigo_producto TEXT,
        descuento_pct REAL DEFAULT 0,
        bonificacion_pct REAL DEFAULT 0,
        FOREIGN KEY (offer_id) REFERENCES offers(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS offer_financial_details (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        offer_id INTEGER NOT NULL,
        porcentaje REAL NOT NULL,
        orden INTEGER DEFAULT 0,
        FOREIGN KEY (offer_id) REFERENCES offers(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS offer_conditions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        offer_id INTEGER NOT NULL,
        condicion_comercial TEXT NOT NULL,
        FOREIGN KEY (offer_id) REFERENCES offers(id) ON DELETE CASCADE,
        UNIQUE(offer_id, condicion_comercial)
    );
    CREATE TABLE IF NOT EXISTS offer_vendors (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        offer_id INTEGER NOT NULL,
        codigousuario TEXT,
        FOREIGN KEY (offer_id) REFERENCES offers(id) ON DELETE CASCADE,
        UNIQUE(offer_id, codigousuario)
    );
    CREATE TABLE IF NOT EXISTS creditos_internos_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha TEXT DEFAULT (datetime('now','localtime')),
        codigousuario TEXT NOT NULL,
        codigocliente TEXT NOT NULL,
        cliente_razonsocial TEXT DEFAULT '',
        oferta_id INTEGER,
        oferta_nombre TEXT DEFAULT '',
        escalon_monto_minimo REAL,
        monto REAL NOT NULL,
        pedido_numero TEXT,
        pedido_db TEXT DEFAULT 'oficial',
        numero_ci REAL,
        codigo_asiento REAL,
        estado TEXT NOT NULL DEFAULT 'ok',
        error_detalle TEXT,
        numero_di REAL,
        di_fecha TEXT,
        di_motivo TEXT
    );
    CREATE TABLE IF NOT EXISTS depositos_config (
        codigo TEXT PRIMARY KEY,
        arma_pedidos INTEGER NOT NULL DEFAULT 1,
        fecha_modificacion TEXT DEFAULT (datetime('now','localtime')),
        modificado_por TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS stock_ajuste_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha TEXT DEFAULT (datetime('now','localtime')),
        usuario TEXT NOT NULL,
        deposito TEXT NOT NULL,
        filtro_desc TEXT,
        total_articulos INTEGER DEFAULT 0,
        con_pendientes INTEGER DEFAULT 0,
        estado TEXT DEFAULT 'ok',
        detalle TEXT
    );
    CREATE TABLE IF NOT EXISTS stock_ajuste_backup (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        log_id INTEGER NOT NULL,
        codigo_articulo TEXT NOT NULL,
        codigo_particular TEXT NOT NULL,
        descripcion TEXT,
        stock_anterior REAL,
        stock_nuevo REAL,
        diferencia REAL,
        remanente_anterior REAL,
        pedidos_pendientes INTEGER DEFAULT 0,
        FOREIGN KEY (log_id) REFERENCES stock_ajuste_log(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS admin_audit_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha TEXT DEFAULT (datetime('now','localtime')),
        usuario TEXT NOT NULL,
        metodo TEXT NOT NULL,
        endpoint TEXT NOT NULL,
        ip TEXT
    );
    CREATE TABLE IF NOT EXISTS stock_reservas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tipo TEXT NOT NULL,
        codigo_articulo TEXT,
        codigo_particular TEXT,
        descripcion_articulo TEXT,
        tipo_grupo TEXT,
        valor_grupo TEXT,
        nombre_grupo TEXT,
        tipo_cantidad TEXT NOT NULL DEFAULT 'unidades',
        cantidad REAL NOT NULL DEFAULT 0,
        deposito TEXT DEFAULT '',
        cantidad_utilizada REAL DEFAULT 0,
        motivo TEXT NOT NULL DEFAULT '',
        fecha_hasta TEXT,
        creado_por TEXT NOT NULL DEFAULT '',
        creado_at TEXT DEFAULT (datetime('now','localtime')),
        activo INTEGER DEFAULT 1
    );
    -- Migración: agregar columnas si ya existía la tabla sin ellas

    """)
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS vendedores_contacto (
        codigo      TEXT PRIMARY KEY,
        nombre      TEXT NOT NULL,
        mail        TEXT DEFAULT '',
        celular     TEXT DEFAULT '',
        apikey_wa   TEXT DEFAULT '',
        activo      INTEGER DEFAULT 1
    );
    CREATE TABLE IF NOT EXISTS catalogos (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre      TEXT NOT NULL,
        descripcion TEXT DEFAULT '',
        filename    TEXT NOT NULL,
        token       TEXT NOT NULL UNIQUE,
        subido_por  TEXT NOT NULL,
        fecha       TEXT DEFAULT (datetime('now','localtime')),
        activo        INTEGER DEFAULT 1,
        email_enviado INTEGER DEFAULT 0,
        email_count   INTEGER DEFAULT 0,
        push_enviado  INTEGER DEFAULT 0,
        push_count    INTEGER DEFAULT 0,
        wa_enviado    INTEGER DEFAULT 0,
        wa_count      INTEGER DEFAULT 0,
        perfiles_texto TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS catalogo_vendedores (
        catalogo_id INTEGER NOT NULL,
        codigo      TEXT    NOT NULL,
        PRIMARY KEY (catalogo_id, codigo),
        FOREIGN KEY (catalogo_id) REFERENCES catalogos(id) ON DELETE CASCADE
    );
    """)
    # Tabla: perfiles de vendedor asociados a la oferta
    cur.execute("""
    CREATE TABLE IF NOT EXISTS offer_profiles (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        offer_id INTEGER NOT NULL,
        perfil_codigo TEXT NOT NULL,
        FOREIGN KEY (offer_id) REFERENCES offers(id) ON DELETE CASCADE,
        UNIQUE(offer_id, perfil_codigo)
    )""")
    # Tabla: filtros de categoría de artículos (gruposuperrubro / superrubro / rubro / marca)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS offer_category_filters (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        offer_id INTEGER NOT NULL,
        nivel TEXT NOT NULL,
        valor TEXT NOT NULL,
        FOREIGN KEY (offer_id) REFERENCES offers(id) ON DELETE CASCADE,
        UNIQUE(offer_id, nivel, valor)
    )""")
    # Tabla: alcance del DESCUENTO por escalón de facturación neta — más restrictivo
    # (opcional) que offer_category_filters. offer_category_filters decide qué artículos
    # SUMAN al neto para alcanzar un escalón; offer_discount_filters decide a cuáles de
    # esos artículos se les aplica efectivamente el % cuando se alcanza el escalón (ej.
    # Hasbro+Microelectronics suman al neto, pero el % solo se da a los Hasbro). Vacío =
    # aplica a todo el alcance general (offer_category_filters), igual que antes de este
    # campo existir.
    cur.execute("""
    CREATE TABLE IF NOT EXISTS offer_discount_filters (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        offer_id INTEGER NOT NULL,
        nivel TEXT NOT NULL,
        valor TEXT NOT NULL,
        FOREIGN KEY (offer_id) REFERENCES offers(id) ON DELETE CASCADE,
        UNIQUE(offer_id, nivel, valor)
    )""")
    cur.execute("""
    CREATE TABLE IF NOT EXISTS offer_combo_escalones (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        offer_id INTEGER NOT NULL,
        min_combos INTEGER NOT NULL,
        descuento_pct REAL NOT NULL,
        FOREIGN KEY (offer_id) REFERENCES offers(id) ON DELETE CASCADE
    )""")
    cur.execute("""
    CREATE TABLE IF NOT EXISTS offer_amount_escalones (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        offer_id INTEGER NOT NULL,
        monto_minimo REAL NOT NULL,
        descuento_pct REAL NOT NULL DEFAULT 0,
        condicion_comercial TEXT,
        FOREIGN KEY (offer_id) REFERENCES offers(id) ON DELETE CASCADE
    )""")
    try:
        cur.execute("ALTER TABLE offer_amount_escalones ADD COLUMN condicion_comercial TEXT")
    except Exception:
        pass
    # Alcance de oferta por CLIENTE específico (ej. descuento de reactivación armado para
    # un cliente puntual que dejó de comprar) — análogo a offer_vendors, pero por cliente.
    # Cuando una oferta tiene filas acá, el alcance por cliente prevalece sobre
    # offer_vendors/offer_profiles (cualquier corredor que atienda a ese cliente puede
    # aplicarla), y sigue respetando offer_category_filters si los tuviera.
    cur.execute("""
    CREATE TABLE IF NOT EXISTS offer_clients (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        offer_id INTEGER NOT NULL,
        codigocliente TEXT NOT NULL,
        razonsocial TEXT DEFAULT '',
        descuento_extra_pct REAL DEFAULT 0,
        vencimiento_extra TEXT DEFAULT '',
        monto_minimo_extra REAL DEFAULT 0,
        FOREIGN KEY (offer_id) REFERENCES offers(id) ON DELETE CASCADE,
        UNIQUE(offer_id, codigocliente)
    )""")
    # descuento_extra_pct / vencimiento_extra / monto_minimo_extra: "recupero de
    # cartera" — cuando un cliente puntual viene de un análisis de Reactivación con
    # descuento adicional cargado, se guarda ACÁ (dentro de la oferta de categoría
    # que ya le aplicaba, ej. Jugueteria/Outdoors/Tecnologia) en vez de crear una
    # oferta aparte. El % se inyecta en /ofertas SOLO para ese código de cliente, y
    # SOLO en el escalón de financial_escalones ya alcanzado — nunca destraba el
    # escalón por sí solo. monto_minimo_extra permite exigir un neto de facturación
    # propio para el recupero, independiente (y potencialmente mayor) al escalón de
    # la oferta base — si es menor o igual al escalón base, no cambia nada.
    for _col, _ddl in (('descuento_extra_pct', "ALTER TABLE offer_clients ADD COLUMN descuento_extra_pct REAL DEFAULT 0"),
                       ('vencimiento_extra',   "ALTER TABLE offer_clients ADD COLUMN vencimiento_extra TEXT DEFAULT ''"),
                       ('monto_minimo_extra',  "ALTER TABLE offer_clients ADD COLUMN monto_minimo_extra REAL DEFAULT 0"),
                       # Soporte para el panel independiente "Descuento por Cartera" (oferta
                       # tipo='cartera'): cada tramo por cliente puede otorgar UNA de dos
                       # recompensas excluyentes — % de descuento (descuento_extra_pct, como
                       # antes) o una condición comercial distinta (condicion_comercial_extra).
                       # tipo_cartera indica cuál de las dos aplica para esa fila.
                       ('tipo_cartera',           "ALTER TABLE offer_clients ADD COLUMN tipo_cartera TEXT DEFAULT 'descuento'"),
                       ('condicion_comercial_extra', "ALTER TABLE offer_clients ADD COLUMN condicion_comercial_extra TEXT DEFAULT ''")):
        try:
            cur.execute(_ddl)
        except Exception:
            pass
    # ── Reactivación de Clientes ────────────────────────────────────────────────
    # Un "análisis" define los parámetros de búsqueda de clientes importantes que
    # dejaron de comprar: Periodo A (referencia histórica para el ranking de
    # facturación, vacío=toda la historia), Periodo B (ventana deslizante de N días
    # para chequear inactividad, siempre relativa a "hoy" en cada corrida), filtros
    # de vendedor/categoría, descuento adicional opcional con su propio vencimiento,
    # y el día/hora en que corre automáticamente.
    cur.execute("""
    CREATE TABLE IF NOT EXISTS reactivacion_analisis (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        vendedor_codigo TEXT DEFAULT '',
        gruposuperrubro TEXT DEFAULT '',
        superrubro TEXT DEFAULT '',
        rubro TEXT DEFAULT '',
        periodo_a_desde TEXT DEFAULT '',
        periodo_a_hasta TEXT DEFAULT '',
        periodo_b_dias INTEGER NOT NULL DEFAULT 60,
        descuento_pct REAL DEFAULT 0,
        oferta_vencimiento TEXT DEFAULT '',
        descuento_monto_minimo REAL DEFAULT 0,
        dia_semana INTEGER NOT NULL DEFAULT 0,
        hora TEXT NOT NULL DEFAULT '09:00',
        activo INTEGER DEFAULT 1,
        ultima_corrida TEXT DEFAULT '',
        ultimo_error TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now','localtime'))
    )""")
    try:
        cur.execute("ALTER TABLE reactivacion_analisis ADD COLUMN ultimo_error TEXT DEFAULT ''")
    except Exception:
        pass
    try:
        cur.execute("ALTER TABLE reactivacion_analisis ADD COLUMN descuento_monto_minimo REAL DEFAULT 0")
    except Exception:
        pass
    try:
        # Recompensa alternativa al % cuando descuento_pct queda en 0: al llegar al
        # mínimo propio, en vez de (o además de) un %, se otorga esta condición
        # comercial distinta — igual mecanismo que "Descuento por Cartera" en Ofertas.
        cur.execute("ALTER TABLE reactivacion_analisis ADD COLUMN condicion_comercial_extra TEXT DEFAULT ''")
    except Exception:
        pass
    # Resultado de UNA corrida (se recalcula todo desde cero en cada corrida — puede
    # haber clientes que entran/salen de una corrida a otra). estado: pendiente (recién
    # detectado, aún no se generó/envió el PDF) -> notificado (ya se avisó al vendedor)
    # -> comprado (dejó de estar inactivo, se cierra positivo) | cerrado (venció la
    # oferta especial sin compra, se deja de notificar).
    cur.execute("""
    CREATE TABLE IF NOT EXISTS reactivacion_resultados (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        analisis_id INTEGER NOT NULL,
        fecha_corrida TEXT NOT NULL,
        codigocliente TEXT NOT NULL,
        codigoparticular TEXT DEFAULT '',
        razonsocial TEXT DEFAULT '',
        vendedor_codigo TEXT DEFAULT '',
        vendedor_nombre TEXT DEFAULT '',
        importe_total REAL NOT NULL DEFAULT 0,
        estado TEXT NOT NULL DEFAULT 'pendiente',
        fecha_ultima_notificacion TEXT DEFAULT '',
        offer_id INTEGER,
        FOREIGN KEY (analisis_id) REFERENCES reactivacion_analisis(id) ON DELETE CASCADE
    )""")
    # codigoparticular: agregado después de la primera versión — el código que ve el
    # vendedor/admin debe ser el CODIGOPARTICULAR de Flexxus, no el CODIGOCLIENTE
    # interno (que es solo la clave de join contra Firebird).
    try:
        cur.execute("ALTER TABLE reactivacion_resultados ADD COLUMN codigoparticular TEXT DEFAULT ''")
    except Exception:
        pass
    # Detalle línea a línea (factura + artículo) que compone el importe_total de un
    # resultado — solo para consulta/auditoría en el modal expandible del panel, nunca
    # se envía a los corredores.
    cur.execute("""
    CREATE TABLE IF NOT EXISTS reactivacion_resultado_detalle (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        resultado_id INTEGER NOT NULL,
        tipo_comprobante TEXT DEFAULT '',
        numero_comprobante TEXT DEFAULT '',
        fecha TEXT DEFAULT '',
        codigo_articulo TEXT DEFAULT '',
        descripcion_articulo TEXT DEFAULT '',
        rubro TEXT DEFAULT '',
        superrubro TEXT DEFAULT '',
        cantidad REAL DEFAULT 0,
        importe REAL DEFAULT 0,
        FOREIGN KEY (resultado_id) REFERENCES reactivacion_resultados(id) ON DELETE CASCADE
    )""")
    cur.execute("""
    CREATE TABLE IF NOT EXISTS app_config (
        clave TEXT PRIMARY KEY,
        valor TEXT,
        updated_at TEXT DEFAULT (datetime('now'))
    )""")
    # Migraciones no destructivas
    try: cur.execute("ALTER TABLE offers ADD COLUMN deposito TEXT DEFAULT ''")
    except Exception: pass
    try: cur.execute("ALTER TABLE offers ADD COLUMN cupo INTEGER DEFAULT 0")
    except Exception: pass
    try: cur.execute("ALTER TABLE offers ADD COLUMN usos INTEGER DEFAULT 0")
    except Exception: pass
    try: cur.execute("ALTER TABLE offers ADD COLUMN tipo_financiero TEXT DEFAULT 'descuento_total'")
    except Exception: pass
    try: cur.execute("ALTER TABLE offers ADD COLUMN monto_minimo REAL DEFAULT 0")
    except Exception: pass
    try: cur.execute("ALTER TABLE offers ADD COLUMN financial_escalones TEXT")
    except Exception: pass
    # Tope de bonificación en $ acumulado a lo largo de la vida de la oferta (0 = ilimitado,
    # mismo criterio que cupo/usos pero en pesos en vez de cantidad de veces). Al llegar al
    # tope, la oferta se auto-desactiva igual que cuando se agota el cupo de usos.
    try: cur.execute("ALTER TABLE offers ADD COLUMN tope_bonificacion_pesos REAL DEFAULT 0")
    except Exception: pass
    try: cur.execute("ALTER TABLE offers ADD COLUMN bonificado_acumulado_pesos REAL DEFAULT 0")
    except Exception: pass
    # Marca si esta oferta de bonificación puede SUMARSE con otras bonificaciones vigentes
    # que matcheen el mismo artículo (p.ej. una campaña puntual como "Día del Niño" que debe
    # sumarse a la bonificación de Jugueteria ya existente). Por defecto 0 (no acumulable):
    # cuando varias bonificaciones no-acumulables matchean el mismo artículo, se toma solo la
    # mayor (comportamiento histórico) — evita que dos bonificaciones de categoría no
    # relacionadas (p.ej. Jugueteria + Outdoors) se sumen entre sí sin que eso sea intencional.
    try: cur.execute("ALTER TABLE offers ADD COLUMN acumulable INTEGER DEFAULT 0")
    except Exception: pass
    try: cur.execute("ALTER TABLE offer_product_details ADD COLUMN descripcion TEXT DEFAULT ''")
    except Exception: pass
    try: cur.execute("ALTER TABLE offer_product_details ADD COLUMN cantidad REAL DEFAULT 1")
    except Exception: pass
    try: cur.execute("ALTER TABLE stock_ajuste_log ADD COLUMN deposito_nombre TEXT")
    except Exception: pass
    try: cur.execute("ALTER TABLE admin_audit_log ADD COLUMN accion TEXT")
    except Exception: pass
    try: cur.execute("ALTER TABLE admin_audit_log ADD COLUMN detalle TEXT")
    except Exception: pass
    try: cur.execute("ALTER TABLE admin_audit_log ADD COLUMN seccion TEXT")
    except Exception: pass
    # Migraciones catalogos: estado de notificaciones y perfiles asociados
    try: cur.execute("ALTER TABLE catalogos ADD COLUMN email_enviado INTEGER DEFAULT 0")
    except Exception: pass
    try: cur.execute("ALTER TABLE catalogos ADD COLUMN wa_enviado INTEGER DEFAULT 0")
    except Exception: pass
    try: cur.execute("ALTER TABLE catalogos ADD COLUMN push_enviado INTEGER DEFAULT 0")
    except Exception: pass
    try: cur.execute("ALTER TABLE catalogos ADD COLUMN perfiles_texto TEXT DEFAULT ''")
    except Exception: pass
    try: cur.execute("ALTER TABLE catalogos ADD COLUMN email_count INTEGER DEFAULT 0")
    except Exception: pass
    try: cur.execute("ALTER TABLE catalogos ADD COLUMN wa_count INTEGER DEFAULT 0")
    except Exception: pass
    try: cur.execute("ALTER TABLE catalogos ADD COLUMN push_count INTEGER DEFAULT 0")
    except Exception: pass
    # Transferencia entre depósitos
    cur.execute("""
    CREATE TABLE IF NOT EXISTS transferencia_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha TEXT DEFAULT (datetime('now','localtime')),
        usuario TEXT NOT NULL,
        deposito_origen TEXT NOT NULL,
        deposito_origen_nombre TEXT,
        deposito_destino TEXT NOT NULL,
        deposito_destino_nombre TEXT,
        total_articulos INTEGER DEFAULT 0,
        estado TEXT DEFAULT 'ok',
        detalle TEXT,
        revertida INTEGER DEFAULT 0
    )""")
    cur.execute("""
    CREATE TABLE IF NOT EXISTS transferencia_detalle (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        log_id INTEGER NOT NULL,
        codigo_articulo TEXT NOT NULL,
        codigo_particular TEXT NOT NULL,
        descripcion TEXT,
        cantidad REAL NOT NULL,
        stock_origen_anterior REAL,
        stock_destino_anterior REAL,
        FOREIGN KEY (log_id) REFERENCES transferencia_log(id) ON DELETE CASCADE
    )""")
    cur.execute("""
    CREATE TABLE IF NOT EXISTS articulo_imagenes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo_particular TEXT NOT NULL,
        ruta_imagen TEXT NOT NULL,
        carpeta_origen TEXT,
        orden INTEGER DEFAULT 0,
        fecha_creacion TEXT DEFAULT (datetime('now','localtime')),
        UNIQUE(codigo_particular, ruta_imagen)
    )""")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_articulo_imagenes_cod ON articulo_imagenes(codigo_particular)")
    # Seed: depósitos exclusivos para ECOMMERCE (002 y 013)
    for _dep_seed in ('deposito_exclusivo_002', 'deposito_exclusivo_013'):
        try:
            cur.execute(
                "INSERT OR IGNORE INTO feature_flags (codigousuario, feature, enabled) VALUES (?,?,?)",
                ('ECOMMERCE', _dep_seed, 1)
            )
        except Exception: pass
    # Seed: ECOMMERCE usa Lista 5 de precios
    try:
        cur.execute(
            "INSERT OR IGNORE INTO feature_flags (codigousuario, feature, enabled) VALUES (?,?,?)",
            ('ECOMMERCE', 'usa_lista5', 1)
        )
    except Exception: pass
    # Seed: ECOMMERCE puede crear pedidos en L1 (DATABASE) — forzar habilitado (UPSERT)
    try:
        cur.execute(
            """INSERT INTO feature_flags (codigousuario, feature, enabled) VALUES (?,?,?)
               ON CONFLICT(codigousuario, feature) DO UPDATE SET enabled=1""",
            ('ECOMMERCE', 'pedidos', 1)
        )
    except Exception: pass
    # MIGRACION: eliminar TODOS los flags pedidos=false (global e individuales)
    # El flag pedidos=false incorrectamente fuerza a L1 vendors a usar la BD SW.
    # Solo debe existir pedidos=false para usuarios exclusivamente SW-only.
    # El admin debe volver a configurar pedidos=false solo para usuarios SW puros.
    try:
        cur.execute("DELETE FROM feature_flags WHERE feature='pedidos' AND enabled=0")
    except Exception: pass

    # Limpiar filas duplicadas en feature_flags con codigousuario IS NULL
    # (SQLite no detecta duplicados con NULL en UNIQUE, pueden haberse acumulado)
    try:
        cur.execute("""
            DELETE FROM feature_flags
            WHERE codigousuario IS NULL
              AND rowid NOT IN (
                SELECT MAX(rowid) FROM feature_flags
                WHERE codigousuario IS NULL
                GROUP BY feature
              )
        """)
    except Exception: pass

    c.commit()
    c.close()

_init_admin_db()

# ─── Config persistente (key-value en admin.db) ──────────────────────────────
def _config_get(clave: str):
    db = _admin_db()
    row = db.execute("SELECT valor FROM app_config WHERE clave=?", (clave,)).fetchone()
    db.close()
    return row['valor'] if row else None

def _config_set(clave: str, valor: str):
    db = _admin_db()
    db.execute(
        "INSERT INTO app_config (clave, valor, updated_at) VALUES (?,?,datetime('now')) "
        "ON CONFLICT(clave) DO UPDATE SET valor=excluded.valor, updated_at=excluded.updated_at",
        (clave, valor)
    )
    db.commit()
    db.close()

# ─── Helper auditoría ────────────────────────────────────────────────────────
def _audit(usuario: str, accion: str, detalle: str = '', ip: str = '', seccion: str = ''):
    try:
        db = _admin_db()
        db.execute(
            "INSERT INTO admin_audit_log (usuario, metodo, endpoint, ip, accion, detalle, seccion) VALUES (?,?,?,?,?,?,?)",
            (usuario, '', '', ip, accion, detalle, seccion)
        )
        db.commit(); db.close()
    except Exception:
        pass

# ─── Admin JWT ────────────────────────────────────────────────────────────────
JWT_ADMIN_HOURS = 8

def _create_admin_token(data: dict) -> str:
    payload = data.copy()
    payload['role'] = 'admin'
    payload['exp'] = datetime.utcnow() + timedelta(hours=JWT_ADMIN_HOURS)
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)

def get_admin_user(credentials: HTTPAuthorizationCredentials = Depends(_bearer)):
    if not credentials:
        raise HTTPException(401, "No autenticado")
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGO])
        if payload.get('role') != 'admin':
            raise HTTPException(403, "Acceso denegado")
        return payload
    except JWTError:
        raise HTTPException(401, "Token inválido o expirado")

def get_admin_download_auth(
    request: Request,
    access_token: Optional[str] = Query(None)
):
    """Dependency para descargas: acepta Bearer header O ?access_token= (WebView)."""
    token = access_token
    if not token:
        auth = request.headers.get('Authorization', '')
        if auth.startswith('Bearer '):
            token = auth[7:]
    if not token:
        raise HTTPException(401, "No autenticado")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
        if payload.get('role') != 'admin':
            raise HTTPException(403, "Acceso denegado")
        return payload
    except JWTError:
        raise HTTPException(401, "Token inválido o expirado")

# ─── Admin: login ─────────────────────────────────────────────────────────────
@app.post("/admin/login")
def admin_login(req: _LoginReq, request: Request):
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        cur.execute(
            'SELECT CODIGOUSUARIO, RAZONSOCIAL, CODIGOPERFIL, ACTIVO '
            'FROM "USUARIOS" WHERE UPPER(CODIGOUSUARIO)=? AND UPPER(PASSWORD1)=?',
            (req.usuario.upper(), req.password.upper())
        )
        row = cur.fetchone()
        c.close()
    except Exception as e:
        raise HTTPException(500, f"Error DB: {e}")
    if not row or str(row[3] or '').strip() != '1':
        raise HTTPException(401, "Usuario o contraseña incorrectos")
    perfil = str(row[2] or '').strip().upper()
    if perfil not in _ADMIN_PERFILES:
        raise HTTPException(403, f"Perfil '{perfil}' no tiene acceso al panel de control")
    cod   = str(row[0] or '').strip()
    razon = str(row[1] or '').strip()
    token = _create_admin_token({'sub': cod, 'nombre': razon, 'perfil': perfil})
    _ip = request.client.host if request.client else ''
    _audit(cod, 'Inicio de sesión', f'Perfil: {perfil}', _ip)
    return {"token": token, "usuario": cod, "nombre": razon, "perfil": perfil}

# ─── Admin: token de impersonación (crea JWT de vendedor para usar en iframe) ─
@app.get("/admin/impersonate-token")
def admin_impersonate_token(codigousuario: str, _u=Depends(get_admin_user)):
    """Genera un JWT temporal (1 hora) como si fuera el vendedor indicado.
    Permite al admin cargar el frontend con el contexto de ese vendedor."""
    cod = codigousuario.strip().upper()
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        cur.execute(
            'SELECT CODIGOUSUARIO, RAZONSOCIAL, CODIGOPERFIL, ESVENDEDOR, ACTIVO '
            'FROM "USUARIOS" WHERE UPPER(CODIGOUSUARIO)=?', (cod,)
        )
        row = cur.fetchone()
        # Permisos del admin actuante (no del vendedor impersonado)
        admin_cod = _u.get('sub', '')
        cur.execute(
            'SELECT BONIFICACIONMAXIMA, PORCENTAJEINCREMENTOPRECIO, PORCENTAJEDECREMENTOPRECIO, CODIGOPERFIL '
            'FROM "USUARIOS" WHERE UPPER(CODIGOUSUARIO)=?', (admin_cod.upper(),)
        )
        row_admin = cur.fetchone()
        c.close()
    except Exception as e:
        raise HTTPException(500, str(e))
    if not row:
        raise HTTPException(404, f"Vendedor '{cod}' no encontrado")
    razon  = str(row[1] or '').strip()
    perfil = str(row[2] or '').strip()
    esvend = str(row[3] or '').strip()
    bonif_max = float(row_admin[0]) if row_admin and row_admin[0] is not None else 0.0
    pct_inc   = float(row_admin[1]) if row_admin and row_admin[1] is not None else 0.0
    pct_dec   = float(row_admin[2]) if row_admin and row_admin[2] is not None else 0.0
    admin_perfil = str(row_admin[3] or '').strip() if row_admin else ''
    payload = {'sub': cod, 'nombre': razon, 'perfil': perfil,
               'esvendedor': esvend, 'impersonated_by': admin_cod,
               'admin_perfil': admin_perfil,
               'bonificacion_maxima': bonif_max, 'pct_incremento': pct_inc, 'pct_decremento': pct_dec,
               'exp': datetime.utcnow() + timedelta(hours=1)}
    token = jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)
    return {"token": token, "codigousuario": cod, "nombre": razon}

# ─── Admin: buscar clientes con vendedor asignado (para Pedido/Ppto) ──────────
@app.get("/admin/pedido/buscar-cliente")
def admin_pedido_buscar_cliente(q: str = Query('', min_length=2), _u=Depends(get_admin_user)):
    """Busca clientes activos en Firebird para autocompletar en sec-pedido admin.
    Devuelve codigovendedor para rellenar el selector de vendedor automáticamente."""
    try:
        c = conn()
        cur = c.cursor()
        cur.execute("""
            SELECT FIRST 30
                CODIGOCLIENTE, RAZONSOCIAL, TRIM(CODIGOPARTICULAR),
                UPPER(TRIM(CODIGOVENDEDOR))
            FROM "CLIENTES"
            WHERE ACTIVO = '1'
              AND (UPPER(RAZONSOCIAL) CONTAINING UPPER(?) OR CODIGOCLIENTE CONTAINING ?)
            ORDER BY RAZONSOCIAL
        """, [q, q])
        rows = cur.fetchall()
        c.close()
        return [
            {
                "codigocliente":    str(r[0] or '').strip(),
                "razonsocial":      str(r[1] or '').strip(),
                "codigoparticular": str(r[2] or '').strip(),
                "codigovendedor":   str(r[3] or '').strip(),
            }
            for r in rows
        ]
    except Exception as e:
        raise HTTPException(500, detail=str(e))

# ─── Admin: sirve admin.html ──────────────────────────────────────────────────
@app.get("/admin", response_class=HTMLResponse)
@app.get("/ctrl", response_class=HTMLResponse)
def admin_panel():
    if os.path.exists(ADMIN_PATH):
        with open(ADMIN_PATH, encoding="utf-8") as f:
            return f.read()
    return HTMLResponse("<h1>admin.html no encontrado</h1>", status_code=404)

# ─── Admin: vendedores (lista desde Firebird) ─────────────────────────────────
@app.get("/admin/vendedores")
def admin_get_vendedores(_u=Depends(get_admin_user)):
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        cur.execute(
            "SELECT CODIGOUSUARIO, RAZONSOCIAL, CODIGOPERFIL FROM \"USUARIOS\" "
            "WHERE ACTIVO='1' AND UPPER(TRIM(CODIGOPERFIL))='VENDEDORES' ORDER BY RAZONSOCIAL"
        )
        rows = cur.fetchall()
        c.close()
    except Exception as e:
        raise HTTPException(500, str(e))
    return [{"codigo": str(r[0] or '').strip(), "nombre": str(r[1] or '').strip(),
             "perfil_flexxus": str(r[2] or '').strip()} for r in rows]

@app.get("/vendedores-lista")
def get_vendedores_lista(u=Depends(get_current_user)):
    """Lista de vendedores accesible con token de impersonación admin."""
    if not u.get('admin_perfil'):
        raise HTTPException(403, "Requiere perfil admin")
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        cur.execute(
            "SELECT CODIGOUSUARIO, RAZONSOCIAL FROM \"USUARIOS\" "
            "WHERE ACTIVO='1' AND UPPER(TRIM(CODIGOPERFIL))='VENDEDORES' ORDER BY RAZONSOCIAL"
        )
        rows = cur.fetchall()
        c.close()
    except Exception as e:
        raise HTTPException(500, str(e))
    return [{"codigo": str(r[0] or '').strip(), "nombre": str(r[1] or '').strip()} for r in rows]

# ─── Admin: perfiles de vendedor ABM ─────────────────────────────────────────
@app.get("/admin/perfiles")
def admin_get_perfiles(_u=Depends(get_admin_user)):
    c = _admin_db(); rows = c.execute("SELECT * FROM vendor_profiles ORDER BY nombre").fetchall(); c.close()
    return [dict(r) for r in rows]

@app.post("/admin/perfiles")
def admin_create_perfil(data: dict, _u=Depends(get_admin_user)):
    codigo = (data.get('codigo') or '').strip().upper()
    nombre = (data.get('nombre') or '').strip()
    if not codigo or not nombre:
        raise HTTPException(400, "codigo y nombre requeridos")
    try:
        c = _admin_db()
        c.execute("INSERT INTO vendor_profiles (codigo, nombre) VALUES (?,?)", (codigo, nombre))
        c.commit(); id_ = c.execute("SELECT last_insert_rowid()").fetchone()[0]; c.close()
    except sqlite3.IntegrityError:
        raise HTTPException(409, f"Perfil '{codigo}' ya existe")
    return {"id": id_, "codigo": codigo, "nombre": nombre, "activo": 1}

@app.put("/admin/perfiles/{id}")
def admin_update_perfil(id: int, data: dict, _u=Depends(get_admin_user)):
    c = _admin_db()
    c.execute("UPDATE vendor_profiles SET codigo=?, nombre=?, activo=? WHERE id=?",
              (data.get('codigo','').strip().upper(), data.get('nombre','').strip(), data.get('activo',1), id))
    c.commit(); c.close()
    return {"ok": True}

@app.delete("/admin/perfiles/{id}")
def admin_delete_perfil(id: int, _u=Depends(get_admin_user)):
    c = _admin_db(); c.execute("DELETE FROM vendor_profiles WHERE id=?", (id,)); c.commit(); c.close()
    return {"ok": True}

# ─── Admin: asignación perfiles ↔ vendedores ──────────────────────────────────
@app.get("/admin/asignaciones")
def admin_get_asignaciones(_u=Depends(get_admin_user)):
    c = _admin_db()
    rows = c.execute("""
        SELECT a.codigousuario, p.id, p.codigo, p.nombre
        FROM vendor_profile_assignments a
        JOIN vendor_profiles p ON p.id=a.profile_id
        ORDER BY a.codigousuario, p.nombre
    """).fetchall(); c.close()
    return [dict(r) for r in rows]

@app.post("/admin/asignaciones")
def admin_set_asignacion(data: dict, _u=Depends(get_admin_user)):
    cod = (data.get('codigousuario') or '').strip().upper()
    pid = data.get('profile_id')
    if not cod or not pid:
        raise HTTPException(400, "codigousuario y profile_id requeridos")
    c = _admin_db()
    try:
        c.execute("INSERT OR IGNORE INTO vendor_profile_assignments (codigousuario, profile_id) VALUES (?,?)", (cod, pid))
        c.commit()
    finally:
        c.close()
    return {"ok": True}

@app.delete("/admin/asignaciones")
def admin_del_asignacion(codigousuario: str, profile_id: int, _u=Depends(get_admin_user)):
    c = _admin_db()
    c.execute("DELETE FROM vendor_profile_assignments WHERE codigousuario=? AND profile_id=?",
              (codigousuario.upper(), profile_id))
    c.commit(); c.close()
    return {"ok": True}

# ─── Admin: feature flags ─────────────────────────────────────────────────────
@app.get("/admin/flags")
def admin_get_flags(_u=Depends(get_admin_user)):
    c = _admin_db()
    rows = c.execute("SELECT * FROM feature_flags ORDER BY codigousuario, feature").fetchall(); c.close()
    return [dict(r) for r in rows]

@app.post("/admin/flags")
def admin_set_flag(data: dict, _u=Depends(get_admin_user)):
    cod     = (data.get('codigousuario') or None)
    if cod: cod = cod.strip().upper() or None
    feature = (data.get('feature') or '').strip()
    enabled = int(data.get('enabled', 1))
    if not feature:
        raise HTTPException(400, "feature requerido")
    c = _admin_db()
    if cod is None:
        # SQLite no detecta duplicados con NULL en UNIQUE, usar DELETE+INSERT explícito
        c.execute("DELETE FROM feature_flags WHERE codigousuario IS NULL AND feature=?", (feature,))
        c.execute("INSERT INTO feature_flags (codigousuario, feature, enabled) VALUES (NULL,?,?)", (feature, enabled))
    else:
        c.execute("""INSERT INTO feature_flags (codigousuario, feature, enabled) VALUES (?,?,?)
                     ON CONFLICT(codigousuario, feature) DO UPDATE SET enabled=excluded.enabled""",
                  (cod, feature, enabled))
    c.commit(); c.close()
    return {"ok": True}

# Endpoint público para que el frontend consulte sus flags
@app.post("/ofertas/{id}/usar")
def registrar_uso_oferta(id: int, body: dict = Body(default={})):
    """Registra un uso de la promo. Si usos >= cupo (y cupo > 0), o si el acumulado en $
    bonificado llega al tope_bonificacion_pesos (y tope > 0), la desactiva.
    body opcional: {"monto_bonificado": <pesos bonificados en este uso>} — usado por las
    ofertas tipo bonificación/financiero, que no tienen un "uso" discreto por unidad como
    las de tipo producto sino un monto acumulado por pedido/presupuesto."""
    monto_bonificado = float((body or {}).get('monto_bonificado', 0) or 0)
    c = _admin_db()
    row = c.execute("SELECT cupo, usos, activo, tope_bonificacion_pesos, bonificado_acumulado_pesos FROM offers WHERE id=?", (id,)).fetchone()
    if not row:
        c.close(); raise HTTPException(404, "Oferta no encontrada")
    cupo = row['cupo'] or 0
    usos = (row['usos'] or 0) + 1
    activo = row['activo']
    tope = row['tope_bonificacion_pesos'] or 0
    acumulado = (row['bonificado_acumulado_pesos'] or 0) + monto_bonificado
    if cupo > 0 and usos >= cupo:
        activo = 0
    if tope > 0 and acumulado >= tope:
        activo = 0
    c.execute("UPDATE offers SET usos=?, activo=?, bonificado_acumulado_pesos=? WHERE id=?", (usos, activo, acumulado, id))
    c.commit(); c.close()
    agotada_cupo = cupo > 0 and usos >= cupo
    agotada_tope = tope > 0 and acumulado >= tope
    return {"usos": usos, "cupo": cupo, "activo": activo,
            "tope_bonificacion_pesos": tope, "bonificado_acumulado_pesos": acumulado,
            "agotada": agotada_cupo or agotada_tope,
            "agotada_cupo": agotada_cupo, "agotada_tope": agotada_tope}

@app.get("/vendor-perfiles")
def get_vendor_perfiles(vendedor: str):
    """Devuelve los codigos de perfil (TECNOLOGIA, JUGUETERIA, etc.) asignados a un vendedor."""
    c = _admin_db()
    rows = c.execute("""
        SELECT vp.codigo FROM vendor_profiles vp
        JOIN vendor_profile_assignments vpa ON vpa.profile_id = vp.id
        WHERE UPPER(vpa.codigousuario) = ?
    """, (vendedor.upper(),)).fetchall()
    c.close()
    return [r[0] for r in rows]

@app.get("/flags")
def get_flags_for_vendor(vendedor: Optional[str] = None):
    c = _admin_db()
    vend_upper = vendedor.upper() if vendedor else ''

    # 1. Global flags (codigousuario IS NULL)
    global_rows = c.execute(
        "SELECT feature, enabled FROM feature_flags WHERE codigousuario IS NULL"
    ).fetchall()

    # 2. Profile flags del vendedor (codigousuario = 'PERFIL:XXX')
    perfil_rows = []
    if vend_upper:
        perfil_codigos = c.execute("""
            SELECT 'PERFIL:' || vp.codigo
            FROM vendor_profiles vp
            JOIN vendor_profile_assignments vpa ON vpa.profile_id = vp.id
            WHERE UPPER(vpa.codigousuario) = ?
        """, (vend_upper,)).fetchall()
        if perfil_codigos:
            placeholders = ','.join('?' * len(perfil_codigos))
            keys = [r[0] for r in perfil_codigos]
            perfil_rows = c.execute(
                f"SELECT feature, enabled FROM feature_flags WHERE codigousuario IN ({placeholders})",
                keys
            ).fetchall()

    # 3. Individual flags del vendedor
    ind_rows = c.execute(
        "SELECT feature, enabled FROM feature_flags WHERE codigousuario=?",
        (vend_upper,)
    ).fetchall() if vend_upper else []

    c.close()

    # Aplicar precedencia: global → perfil → individual
    result = {}
    for r in global_rows:
        result[r['feature']] = bool(r['enabled'])
    for r in perfil_rows:
        result[r['feature']] = bool(r['enabled'])
    for r in ind_rows:
        result[r['feature']] = bool(r['enabled'])

    # Derivar deposito_exclusivo desde flags deposito_exclusivo_XXX
    # Si hay ALGÚN flag configurado (aunque todos estén deshabilitados), siempre incluir la clave
    # para que el frontend distinga "sin configurar" (clave ausente) de "configurado a vacío" (clave = "")
    dep_all = [k for k in result if k.startswith('deposito_exclusivo_')]
    dep_enabled = sorted([k.replace('deposito_exclusivo_', '') for k in dep_all if result[k]])
    if dep_all:
        result['deposito_exclusivo'] = ','.join(dep_enabled)  # puede ser '' si todos deshabilitados
    return result

@app.post("/admin/flags/bulk")
def admin_set_flags_bulk(data: dict, _u=Depends(get_admin_user)):
    feature  = (data.get('feature') or '').strip()
    enabled  = 1 if data.get('enabled') else 0
    usuarios = data.get('codigousuarios') or []
    if not feature or not usuarios:
        raise HTTPException(400, "feature y codigousuarios requeridos")
    c = _admin_db()
    for cod in usuarios:
        cod = str(cod).strip().upper()
        if cod:
            c.execute("""INSERT INTO feature_flags (codigousuario, feature, enabled) VALUES (?,?,?)
                         ON CONFLICT(codigousuario, feature) DO UPDATE SET enabled=excluded.enabled""",
                      (cod, feature, enabled))
    c.commit(); c.close()
    return {"ok": True, "actualizados": len(usuarios)}

@app.post("/admin/flags/reset-individuales")
def admin_reset_flags_individuales(data: dict, _u=Depends(get_admin_user)):
    """Borra todos los overrides individuales de una feature, dejando solo el global."""
    feature = (data.get('feature') or '').strip()
    if not feature:
        raise HTTPException(400, "feature requerida")
    c = _admin_db()
    c.execute("DELETE FROM feature_flags WHERE feature=? AND codigousuario IS NOT NULL AND TRIM(codigousuario)<>''", (feature,))
    c.commit(); c.close()
    return {"ok": True}

@app.delete("/admin/flags/depositos-vendedor/{codigousuario}")
def admin_delete_depositos_vendedor(codigousuario: str, _u=Depends(get_admin_user)):
    """Elimina todas las filas deposito_exclusivo_* para un vendedor o perfil, dejando que herede global/perfil."""
    cod = codigousuario.strip()
    if not cod:
        raise HTTPException(400, "codigousuario requerido")
    c = _admin_db()
    c.execute("DELETE FROM feature_flags WHERE codigousuario=? AND feature LIKE 'deposito_exclusivo_%'", (cod,))
    c.commit(); c.close()
    return {"ok": True}

@app.delete("/admin/flags/depositos-todos-individuales")
def admin_delete_depositos_todos_individuales(_u=Depends(get_admin_user)):
    """Elimina TODOS los overrides individuales y de perfil de deposito_exclusivo_*, dejando solo el global."""
    c = _admin_db()
    # Borrar individual y de perfiles (PERFIL:XXX), mantener solo codigousuario IS NULL (global)
    c.execute("DELETE FROM feature_flags WHERE feature LIKE 'deposito_exclusivo_%' AND codigousuario IS NOT NULL")
    c.commit(); c.close()
    return {"ok": True}

# ─── Admin: multiplazos ───────────────────────────────────────────────────────
@app.get("/admin/multiplazos")
def admin_get_multiplazos(_u=Depends(get_admin_user)):
    c = _admin_db(); rows = c.execute("SELECT * FROM multiplazos ORDER BY nombre").fetchall(); c.close()
    return [dict(r) for r in rows]

@app.post("/admin/multiplazos")
def admin_create_multiplazo(data: dict, _u=Depends(get_admin_user)):
    nombre = (data.get('nombre') or '').strip()
    dias   = (data.get('dias') or '').strip()
    if not nombre or not dias:
        raise HTTPException(400, "nombre y dias requeridos")
    c = _admin_db()
    c.execute("INSERT INTO multiplazos (nombre, dias) VALUES (?,?)", (nombre, dias))
    c.commit(); id_ = c.execute("SELECT last_insert_rowid()").fetchone()[0]; c.close()
    return {"id": id_, "nombre": nombre, "dias": dias, "activo": 1}

@app.put("/admin/multiplazos/{id}")
def admin_update_multiplazo(id: int, data: dict, _u=Depends(get_admin_user)):
    c = _admin_db()
    c.execute("UPDATE multiplazos SET nombre=?, dias=?, activo=? WHERE id=?",
              (data.get('nombre','').strip(), data.get('dias','').strip(), data.get('activo',1), id))
    c.commit(); c.close()
    return {"ok": True}

@app.delete("/admin/multiplazos/{id}")
def admin_delete_multiplazo(id: int, _u=Depends(get_admin_user)):
    c = _admin_db(); c.execute("DELETE FROM multiplazos WHERE id=?", (id,)); c.commit(); c.close()
    return {"ok": True}

@app.get("/admin/vendor-multiplazos")
def admin_get_vendor_multiplazos(_u=Depends(get_admin_user)):
    c = _admin_db()
    rows = c.execute("""
        SELECT vm.codigousuario, m.id, m.nombre, m.dias
        FROM vendor_multiplazos vm JOIN multiplazos m ON m.id=vm.multiplazo_id
        ORDER BY vm.codigousuario, m.nombre
    """).fetchall(); c.close()
    return [dict(r) for r in rows]

@app.post("/admin/vendor-multiplazos")
def admin_set_vendor_multiplazo(data: dict, _u=Depends(get_admin_user)):
    cod = (data.get('codigousuario') or None)
    if cod: cod = cod.strip().upper() or None
    mid = data.get('multiplazo_id')
    c = _admin_db()
    c.execute("INSERT OR IGNORE INTO vendor_multiplazos (codigousuario, multiplazo_id) VALUES (?,?)", (cod, mid))
    c.commit(); c.close()
    return {"ok": True}

@app.delete("/admin/vendor-multiplazos")
def admin_del_vendor_multiplazo(codigousuario: str, multiplazo_id: int, _u=Depends(get_admin_user)):
    cod = codigousuario.upper() if codigousuario != '__global__' else None
    c = _admin_db()
    c.execute("DELETE FROM vendor_multiplazos WHERE codigousuario IS ? AND multiplazo_id=?", (cod, multiplazo_id))
    c.commit(); c.close()
    return {"ok": True}

# ─── Multiplazos desde Firebird ───────────────────────────────────────────────
@app.get("/admin/multiplazos-fb")
def admin_get_multiplazos_fb(_u=Depends(get_admin_user)):
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        try:
            cur.execute('SELECT CODIGOMULTIPLAZO, DESCRIPCION, FACTURAPEDIDOSCTACTE FROM "MULTIPLAZOS" WHERE ACTIVO=? ORDER BY DESCRIPCION', ('1',))
            rows = cur.fetchall()
            result = [{"codigo": str(r[0] or '').strip(), "descripcion": str(r[1] or '').strip(), "cuentacorriente": bool(r[2])} for r in rows]
        except Exception:
            cur2 = c.cursor()
            cur2.execute('SELECT CODIGOMULTIPLAZO, DESCRIPCION FROM "MULTIPLAZOS" WHERE ACTIVO=? ORDER BY DESCRIPCION', ('1',))
            rows = cur2.fetchall()
            result = [{"codigo": str(r[0] or '').strip(), "descripcion": str(r[1] or '').strip(), "cuentacorriente": False} for r in rows]
        c.close()
    except Exception as e:
        raise HTTPException(500, str(e))
    return result

@app.get("/admin/vendor-multiplazos-fb")
def admin_get_vendor_multiplazos_fb(_u=Depends(get_admin_user)):
    db = _admin_db()
    rows = db.execute("SELECT codigousuario, codigo_multiplazo FROM vendor_multiplazos_fb ORDER BY codigousuario, codigo_multiplazo").fetchall()
    db.close()
    return [dict(r) for r in rows]

@app.post("/admin/vendor-multiplazos-fb")
def admin_set_vendor_multiplazo_fb(data: dict, _u=Depends(get_admin_user)):
    cod  = (data.get('codigousuario') or '').strip().upper()
    mps  = data.get('codigos_multiplazo') or []  # lista de códigos FB
    if not cod or not mps:
        raise HTTPException(400, "codigousuario y codigos_multiplazo requeridos")
    db = _admin_db()
    for mp in mps:
        mp = str(mp).strip()
        if mp:
            db.execute("INSERT OR IGNORE INTO vendor_multiplazos_fb (codigousuario, codigo_multiplazo) VALUES (?,?)", (cod, mp))
    db.commit(); db.close()
    return {"ok": True}

@app.delete("/admin/vendor-multiplazos-fb")
def admin_del_vendor_multiplazo_fb(codigousuario: str, codigo_multiplazo: str, _u=Depends(get_admin_user)):
    db = _admin_db()
    db.execute("DELETE FROM vendor_multiplazos_fb WHERE codigousuario=? AND codigo_multiplazo=?",
               (codigousuario.upper(), codigo_multiplazo.strip()))
    db.commit(); db.close()
    return {"ok": True}

@app.delete("/admin/vendor-multiplazos-fb/bulk")
def admin_del_vendor_multiplazo_fb_bulk(codigousuario: str, _u=Depends(get_admin_user)):
    """Elimina TODOS los multiplazos asignados a un vendedor"""
    db = _admin_db()
    db.execute("DELETE FROM vendor_multiplazos_fb WHERE codigousuario=?", (codigousuario.upper(),))
    db.commit(); db.close()
    return {"ok": True}

@app.post("/admin/vendor-multiplazos-fb/bulk-perfil")
def admin_set_multiplazos_fb_perfil(data: dict, _u=Depends(get_admin_user)):
    """Asigna una lista de multiplazos a todos los vendedores de un perfil (vendor_profile_assignments)"""
    profile_id       = data.get('profile_id')
    codigos_mp       = data.get('codigos_multiplazo') or []
    reemplazar       = data.get('reemplazar', False)
    if not profile_id or not codigos_mp:
        raise HTTPException(400, "profile_id y codigos_multiplazo requeridos")
    db = _admin_db()
    # Buscar vendedores con ese perfil
    vendors = db.execute("SELECT DISTINCT codigousuario FROM vendor_profile_assignments WHERE profile_id=?", (profile_id,)).fetchall()
    if not vendors:
        db.close(); raise HTTPException(404, "Sin vendedores en ese perfil")
    for v in vendors:
        cod = v['codigousuario']
        if reemplazar:
            db.execute("DELETE FROM vendor_multiplazos_fb WHERE codigousuario=?", (cod,))
        for mp in codigos_mp:
            mp = str(mp).strip()
            if mp:
                db.execute("INSERT OR IGNORE INTO vendor_multiplazos_fb (codigousuario, codigo_multiplazo) VALUES (?,?)", (cod, mp))
    db.commit(); db.close()
    return {"ok": True, "vendedores": len(vendors)}

# Endpoint público para frontend — devuelve multiplazos de Firebird filtrados por asignación del vendedor
@app.get("/multiplazos")
def get_multiplazos_for_vendor(vendedor: Optional[str] = None):
    try:
        # Obtener todos los activos de Firebird
        fb = conn('WIN1252')
        cur = fb.cursor()
        try:
            cur.execute('SELECT CODIGOMULTIPLAZO, DESCRIPCION, FACTURAPEDIDOSCTACTE FROM "MULTIPLAZOS" WHERE ACTIVO=? ORDER BY DESCRIPCION', ('1',))
            todos = [{"codigo": ('' if r[0] is None else str(r[0]).strip()), "descripcion": str(r[1] or '').strip(), "cuentacorriente": bool(r[2])} for r in cur.fetchall()]
        except Exception:
            cur2 = fb.cursor()
            cur2.execute('SELECT CODIGOMULTIPLAZO, DESCRIPCION FROM "MULTIPLAZOS" WHERE ACTIVO=? ORDER BY DESCRIPCION', ('1',))
            todos = [{"codigo": ('' if r[0] is None else str(r[0]).strip()), "descripcion": str(r[1] or '').strip(), "cuentacorriente": False} for r in cur2.fetchall()]
        fb.close()
    except Exception:
        todos = []
    if not vendedor:
        return todos
    # Filtrar por asignaciones del vendedor en admin.db
    db = _admin_db()
    asignados = [r[0] for r in db.execute(
        "SELECT codigo_multiplazo FROM vendor_multiplazos_fb WHERE codigousuario=?",
        (vendedor.upper(),)
    ).fetchall()]
    db.close()
    if not asignados:
        return todos  # sin asignaciones: devuelve todos
    return [m for m in todos if m['codigo'] in asignados]

# ─── Admin: catálogos ─────────────────────────────────────────────────────────
@app.get("/admin/catalogs-legacy")
def admin_get_catalogos(_u=Depends(get_admin_user)):
    """Endpoint legacy — tabla 'catalogs' (vieja). Usar /admin/catalogos para la nueva."""
    c = _admin_db()
    cats = [dict(r) for r in c.execute("SELECT * FROM catalogs ORDER BY nombre").fetchall()]
    for cat in cats:
        profs = c.execute("""
            SELECT p.id, p.codigo, p.nombre FROM catalog_profiles cp
            JOIN vendor_profiles p ON p.id=cp.profile_id WHERE cp.catalog_id=?
        """, (cat['id'],)).fetchall()
        cat['profiles'] = [dict(p) for p in profs]
    c.close()
    return cats

@app.post("/admin/catalogos")
def admin_create_catalogo(data: dict, _u=Depends(get_admin_user)):
    nombre = (data.get('nombre') or '').strip()
    url    = (data.get('url') or '').strip()
    if not nombre or not url:
        raise HTTPException(400, "nombre y url requeridos")
    c = _admin_db()
    c.execute("INSERT INTO catalogs (nombre, descripcion, url) VALUES (?,?,?)",
              (nombre, data.get('descripcion','').strip(), url))
    c.commit(); id_ = c.execute("SELECT last_insert_rowid()").fetchone()[0]
    for pid in (data.get('profile_ids') or []):
        try: c.execute("INSERT OR IGNORE INTO catalog_profiles (catalog_id, profile_id) VALUES (?,?)", (id_, pid))
        except: pass
    c.commit(); c.close()
    return {"id": id_, "ok": True}

@app.put("/admin/catalogos/{id}")
def admin_update_catalogo(id: int, data: dict, _u=Depends(get_admin_user)):
    c = _admin_db()
    c.execute("UPDATE catalogs SET nombre=?, descripcion=?, url=?, activo=? WHERE id=?",
              (data.get('nombre','').strip(), data.get('descripcion','').strip(),
               data.get('url','').strip(), data.get('activo',1), id))
    if 'profile_ids' in data:
        c.execute("DELETE FROM catalog_profiles WHERE catalog_id=?", (id,))
        for pid in (data['profile_ids'] or []):
            try: c.execute("INSERT OR IGNORE INTO catalog_profiles (catalog_id, profile_id) VALUES (?,?)", (id, pid))
            except: pass
    c.commit(); c.close()
    return {"ok": True}

@app.delete("/admin/catalogs-legacy/{id}")
def admin_delete_catalogo_legacy(id: int, _u=Depends(get_admin_user)):
    """Legacy: borraba de tabla 'catalogs' vieja. Reemplazado por /admin/catalogos/{cat_id}."""
    c = _admin_db(); c.execute("DELETE FROM catalogs WHERE id=?", (id,)); c.commit(); c.close()
    return {"ok": True}

# Endpoint público para frontend
@app.get("/catalogos")
def get_catalogos_for_vendor(vendedor: Optional[str] = None):
    c = _admin_db()
    if vendedor:
        rows = c.execute("""
            SELECT DISTINCT ca.id, ca.nombre, ca.descripcion, ca.url
            FROM catalogs ca
            JOIN catalog_profiles cp ON cp.catalog_id=ca.id
            JOIN vendor_profile_assignments vpa ON vpa.profile_id=cp.profile_id
            WHERE ca.activo=1 AND vpa.codigousuario=?
        """, (vendedor.upper(),)).fetchall()
    else:
        rows = c.execute("SELECT id, nombre, descripcion, url FROM catalogs WHERE activo=1").fetchall()
    c.close()
    return [dict(r) for r in rows]

# ─── Admin: invalidación manual de cachés ────────────────────────────────────
@app.post("/admin/cache/refresh")
def admin_cache_refresh(target: str = Query("all", description="'catalog', 'stock' o 'all'"), _u=Depends(get_admin_user)):
    """Fuerza recarga del caché de catálogo y/o stock sin reiniciar el servidor."""
    if target in ("catalog", "all"):
        _catalog_invalidate()
        _get_catalog()   # recarga sincrónica
        _filtros_invalidate()
        _get_filtros_combos()   # recarga sincrónica
        _pub_filtros_invalidate()
        _get_pub_filtros()      # recarga sincrónica
    if target in ("stock", "all"):
        _fma_cache_invalidate()
        _fma_stock_parallel(_deps_activos())  # recarga sincrónica
    return {"ok": True, "refreshed": target}

@app.get("/admin/cache/status")
def admin_cache_status(_u=Depends(get_admin_user)):
    """Devuelve estado de los cachés (edad en segundos)."""
    now = time.time()
    catalog_age = int(now - _catalog_cache_ts) if _catalog_cache_ts else None
    fma_ages = {}
    with _fma_cache_lock:
        for dep, entry in _fma_cache.items():
            fma_ages[dep] = int(now - entry[0])
    return {
        "catalog": {
            "articulos": len(_catalog_cache),
            "edad_seg":  catalog_age,
            "ttl_seg":   _CATALOG_CACHE_TTL,
        },
        "fma_stock": {
            "depositos": fma_ages,
            "ttl_seg":   _FMA_CACHE_TTL,
        },
        "filtros_combos": {
            "edad_seg": int(now - _filtros_cache_ts) if _filtros_cache_ts else None,
            "ttl_seg":  _CATALOG_CACHE_TTL,
        }
    }

# ─── Admin: depósitos disponibles ────────────────────────────────────────────
@app.get("/admin/depositos")
def admin_get_depositos(_u=Depends(get_admin_user)):
    try:
        result = _get_filtros_combos()['depositos']
        if result:
            return result
    except Exception:
        pass
    return [
        {"codigo": "001", "nombre": "DEPOSITO VAC-LOG"},
        {"codigo": "002", "nombre": "DEPOSITO MARKET PLACE"},
        {"codigo": "003", "nombre": "DEPOSITO PACHECO"},
        {"codigo": "005", "nombre": "DEPOSITO OUTLET"},
        {"codigo": "016", "nombre": "DEPOSITO EXPO"},
    ]

@app.get("/depositos")
def get_depositos_publico():
    """Endpoint público: lista de depósitos activos (para frontend de vendedores).
    Incluye arma_pedidos (2026-07-31, ver depositos_config) — metadata inofensiva
    para exponer a todos los vendedores, la sugerencia de transferencia automática
    que la usa se sigue mostrando solo a los perfiles habilitados (front + back)."""
    armaMap = _depositos_arma_pedidos_map()
    try:
        result = [{"codigo": d["codigo"].zfill(3), "nombre": d["nombre"],
                    "arma_pedidos": armaMap.get(d["codigo"].zfill(3), True)}
                  for d in _get_filtros_combos()['depositos']]
        if result:
            return result
    except Exception:
        pass
    return [
        {"codigo": "001", "nombre": "DEPOSITO VAC-LOG", "arma_pedidos": armaMap.get('001', True)},
        {"codigo": "002", "nombre": "DEPOSITO MARKET PLACE", "arma_pedidos": armaMap.get('002', True)},
        {"codigo": "003", "nombre": "DEPOSITO PACHECO", "arma_pedidos": armaMap.get('003', True)},
        {"codigo": "005", "nombre": "DEPOSITO OUTLET", "arma_pedidos": armaMap.get('005', True)},
        {"codigo": "013", "nombre": "DEPOSITO FULL ML", "arma_pedidos": armaMap.get('013', True)},
        {"codigo": "016", "nombre": "DEPOSITO EXPO", "arma_pedidos": armaMap.get('016', True)},
    ]

@app.get("/admin/depositos-config")
def get_depositos_config(_u=Depends(get_admin_user)):
    """Lista TODOS los depósitos activos de Firebird con su flag arma_pedidos actual
    (default True si no hay fila explícita en depositos_config) — para el checklist
    de admin.html donde Eduardo marca cuáles arman pedidos de verdad."""
    armaMap = _depositos_arma_pedidos_map()
    try:
        deps = _get_filtros_combos()['depositos']
    except Exception:
        deps = []
    return [{"codigo": d["codigo"].zfill(3), "nombre": d["nombre"],
              "arma_pedidos": armaMap.get(d["codigo"].zfill(3), True)} for d in deps]

@app.post("/admin/depositos-config")
def set_depositos_config(body: dict = Body(...), u=Depends(get_admin_user)):
    """Guarda en bloque el flag arma_pedidos de cada depósito. body: {"depositos":
    [{"codigo":"006","arma_pedidos":false}, ...]}. Solo afecta la sugerencia de
    transferencia automática en el popup de stock del Pedido (2026-07-31) — ninguna
    otra vista de remanente lee depositos_config."""
    items = (body or {}).get('depositos') or []
    db = _admin_db()
    usuario = (u or {}).get('sub') or (u or {}).get('nombre') or ''
    for it in items:
        cod = str(it.get('codigo', '')).strip().zfill(3)
        if not cod:
            continue
        arma = 1 if it.get('arma_pedidos') else 0
        db.execute(
            "INSERT INTO depositos_config (codigo, arma_pedidos, fecha_modificacion, modificado_por) "
            "VALUES (?,?,datetime('now','localtime'),?) "
            "ON CONFLICT(codigo) DO UPDATE SET arma_pedidos=excluded.arma_pedidos, "
            "fecha_modificacion=excluded.fecha_modificacion, modificado_por=excluded.modificado_por",
            (cod, arma, usuario or '')
        )
    db.commit(); db.close()
    return {"ok": True, "actualizados": len(items)}

# ─── Admin: Ajuste de Stock ───────────────────────────────────────────────────
_PERFILES_GERENTES       = {'GERENTES', 'GTES FE'}
_PERFILES_TRANSFERENCIA  = {'GERENTES', 'GTES FE', 'ADV', 'ADVJUAN'}

def get_gerente_user(credentials: HTTPAuthorizationCredentials = Depends(_bearer)):
    if not credentials:
        raise HTTPException(401, "No autenticado")
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGO])
        if payload.get('role') != 'admin':
            raise HTTPException(403, "Acceso denegado")
        perfil = str(payload.get('perfil') or '').strip().upper()
        if perfil not in _PERFILES_GERENTES:
            raise HTTPException(403, f"Perfil '{perfil}' no tiene acceso a Ajuste de Stock")
        return payload
    except JWTError:
        raise HTTPException(401, "Token inválido o expirado")

@app.post("/admin/ajuste-stock/preview")
def ajuste_stock_preview(body: dict, deposito: str, _u=Depends(get_gerente_user)):
    """CSV define los SKUs. Depósito obligatorio. Sin filtros adicionales."""
    try:
        articulos_csv = body.get('articulos') or []
        if not articulos_csv:
            raise HTTPException(400, "El CSV está vacío")
        if not deposito:
            raise HTTPException(400, "Seleccioná un depósito")
        csv_map = {str(a['codigo']).strip(): float(a.get('cantidad', 0))
                   for a in articulos_csv if a.get('codigo')}
        cod_list = list(csv_map.keys())
        placeholders = ','.join(['?' for _ in cod_list])
        c = conn('WIN1252')
        cur = c.cursor()
        cur.execute(f'''
            SELECT CODIGOARTICULO, CODIGOPARTICULAR, DESCRIPCION
            FROM "ARTICULOS"
            WHERE CODIGOPARTICULAR IN ({placeholders})
            ORDER BY CODIGOPARTICULAR
        ''', cod_list)
        arts = cur.fetchall()
        # Remanente del depósito seleccionado
        stock_map = {}
        try:
            cur2 = c.cursor()
            cur2.execute(f'SELECT ID_ARTICULO, STOCKREMANENTE FROM "FMA_STOCK"(NULL, NULL, \'{deposito}\', 1, 1)')
            stock_map = {str(r[0] or '').strip(): float(r[1] or 0) for r in cur2.fetchall()}
        except Exception:
            pass
        # Stock total Flexxus (STOCK table, CODIGOSUCURSAL=PRINCIPAL)
        stock_total_map = {}
        try:
            cur_st = c.cursor()
            cur_st.execute("SELECT CODIGOARTICULO, STOCKACTUAL FROM \"STOCK\" WHERE CODIGOSUCURSAL='PRINCIPAL' AND LOTE='000'")
            stock_total_map = {str(r[0] or '').strip(): float(r[1] or 0) for r in cur_st.fetchall()}
        except Exception:
            pass
        # Remanente total Flexxus = suma FMA_STOCK de todos los depósitos activos
        _DEPS_ACTIVOS = ['001','002','003','005','013','016']
        rem_total_map = {}
        try:
            cur_rt = c.cursor()
            cur_rt.execute("SELECT ID_ARTICULO, STOCKREMANENTE FROM \"FMA_STOCK\"(NULL, NULL, NULL, 1, 1)")
            for r in cur_rt.fetchall():
                k = str(r[0] or '').strip()
                rem_total_map[k] = rem_total_map.get(k, 0.0) + float(r[1] or 0)
        except Exception:
            # fallback: sumar por cada depósito activo
            try:
                for dep in _DEPS_ACTIVOS:
                    cur_rt2 = c.cursor()
                    cur_rt2.execute(f"SELECT ID_ARTICULO, STOCKREMANENTE FROM \"FMA_STOCK\"(NULL, NULL, '{dep}', 1, 1)")
                    for r in cur_rt2.fetchall():
                        k = str(r[0] or '').strip()
                        rem_total_map[k] = rem_total_map.get(k, 0.0) + float(r[1] or 0)
            except Exception:
                pass
        pendientes_map = {}
        try:
            cur3 = c.cursor()
            cur3.execute('''
                SELECT cc.CODIGOARTICULO, SUM(cc.CANTIDAD - cc.CANTIDADPREPARADA)
                FROM "CUERPOCOMPROBANTES" cc
                JOIN "CABEZACOMPROBANTES" cab ON cab.NUMEROCOMPROBANTE = cc.NUMEROCOMPROBANTE
                WHERE cab.TIPOCOMPROBANTE = 'PE'
                  AND cc.CANTIDADPREPARADA < cc.CANTIDAD
                  AND cc.CODIGODEPOSITO = ?
                GROUP BY cc.CODIGOARTICULO
            ''', (deposito,))
            pendientes_map = {str(r[0] or '').strip(): int(r[1] or 0) for r in cur3.fetchall()}
        except Exception:
            pass
        c.close()
        encontrados = {str(a[1] or '').strip() for a in arts}
        no_encontrados = [x for x in cod_list if x not in encontrados]
        resultado = []
        for a in arts:
            cod_art  = str(a[0] or '').strip()
            cod_part = str(a[1] or '').strip()
            stock_ant = stock_map.get(cod_art, 0.0)
            stock_nvo = csv_map.get(cod_part, stock_ant)
            diff = round(stock_nvo - stock_ant, 4)
            rem_ant = stock_map.get(cod_art, 0.0)
            rem_nvo = round(rem_ant + diff, 4)
            resultado.append({
                "codigo_articulo": cod_art, "codigo_particular": cod_part,
                "descripcion": str(a[2] or '').strip(),
                "stock_anterior": stock_ant, "stock_nuevo": stock_nvo,
                "diferencia": diff,
                "remanente_anterior": rem_ant,
                "remanente_nuevo": rem_nvo,
                "stock_total_flexxus": stock_total_map.get(cod_art, 0.0),
                "remanente_total_flexxus": round(rem_total_map.get(cod_art, 0.0), 4),
                "pedidos_pendientes": pendientes_map.get(cod_art, 0)
            })
        return {"deposito": deposito, "articulos": resultado, "no_encontrados": no_encontrados}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/admin/ajuste-stock/jerarquia")
def ajuste_stock_jerarquia(_u=Depends(get_gerente_user)):
    """Devuelve SuperRubro -> Rubros usando WEB_STOCK (sin JOIN a tablas faltantes)."""
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        # WEB_STOCK ya tiene CODIGOSUPERRUBRO y CODIGORUBRO directamente
        cur.execute("""
            SELECT DISTINCT
                TRIM(CODIGOSUPERRUBRO),
                TRIM(CODIGORUBRO)
            FROM "WEB_STOCK"
            WHERE CODIGOSUPERRUBRO IS NOT NULL
              AND CODIGORUBRO IS NOT NULL
            ORDER BY 1, 2
        """)
        rows = cur.fetchall()
        c.close()
        from collections import defaultdict
        tree = defaultdict(list)
        for sr_val, rub in rows:
            sr_val = (sr_val or '').strip()
            rub = (rub or '').strip()
            if sr_val and rub:
                tree[sr_val].append(rub)
        # Sin tabla GRUPOSUPERRUBROS accesible, usamos SR como nivel principal
        result = [
            {"gsr": sr_val, "superrubros": [{"sr": sr_val, "rubros": sorted(rubros)}]}
            for sr_val, rubros in sorted(tree.items())
        ]
        return result
    except Exception:
        return []

@app.post("/admin/ajuste-stock/procesar")
def ajuste_stock_procesar(data: dict, _u=Depends(get_gerente_user)):
    deposito  = (data.get('deposito') or '').strip()
    articulos = data.get('articulos') or []
    filtro    = data.get('filtro') or {}
    usuario   = str(_u.get('sub') or '').strip().upper()
    if not deposito or not articulos:
        raise HTTPException(400, "deposito y articulos requeridos")
    filtro_parts = []
    if filtro.get('gsr'):   filtro_parts.append(f"GSR:{filtro['gsr']}")
    if filtro.get('sr'):    filtro_parts.append(f"SR:{filtro['sr']}")
    if filtro.get('rubro'): filtro_parts.append(f"RUB:{filtro['rubro']}")
    filtro_desc = ', '.join(filtro_parts) or 'Todos'
    con_pendientes = sum(1 for a in articulos if a.get('pedidos_pendientes', 0) > 0)
    estado = 'ok'; detalle_errors = []
    # Resolver nombre del depósito
    dep_nombre = deposito
    try:
        _c = conn('WIN1252'); _cur = _c.cursor()
        _cur.execute('SELECT DESCRIPCION FROM "DEPOSITOS" WHERE CODIGODEPOSITO=?', (deposito,))
        _r = _cur.fetchone()
        if _r: dep_nombre = str(_r[0] or '').strip()
        _c.close()
    except Exception: pass
    # 1. Backup SQLite
    db = _admin_db()
    try:
        cur_db = db.cursor()
        cur_db.execute(
            "INSERT INTO stock_ajuste_log (usuario, deposito, deposito_nombre, filtro_desc, total_articulos, con_pendientes, estado) VALUES (?,?,?,?,?,?,?)",
            (usuario, deposito, dep_nombre, filtro_desc, len(articulos), con_pendientes, 'procesando'))
        log_id = cur_db.lastrowid
        for a in articulos:
            cur_db.execute(
                "INSERT INTO stock_ajuste_backup (log_id,codigo_articulo,codigo_particular,descripcion,stock_anterior,stock_nuevo,diferencia,remanente_anterior,pedidos_pendientes) VALUES (?,?,?,?,?,?,?,?,?)",
                (log_id, a.get('codigo_articulo',''), a.get('codigo_particular',''),
                 a.get('descripcion',''), a.get('stock_anterior',0), a.get('stock_nuevo',0),
                 a.get('diferencia',0), a.get('remanente',0), a.get('pedidos_pendientes',0)))
        db.commit()
    except Exception as e:
        db.close(); raise HTTPException(500, f"Error backup: {e}")
    # 2. Escribir Firebird
    procesados = 0
    try:
        fb = conn('WIN1252'); fb_cur = fb.cursor()
        fb_cur.execute('SELECT MAX(NUMEROMOVIMIENTO) FROM "CORRECCIONESSTOCKMANUALES"')
        next_num = int(fb_cur.fetchone()[0] or 0) + 1
        from datetime import date as _date, datetime as _dt, time as _time
        hoy_date = _date.today()          # objeto date para campos DATE de Firebird
        hoy_str  = hoy_date.isoformat()   # string solo para observaciones
        for a in articulos:
            cod_art = a.get('codigo_articulo','').strip()
            diff    = float(a.get('diferencia', 0))
            stock_nvo = float(a.get('stock_nuevo', 0))
            if diff == 0: continue
            ingreso = max(diff, 0); egreso = max(-diff, 0)
            try:
                ahora = _dt.now()
                hora_time = ahora.time().replace(microsecond=0)
                # STOCK en CORRECCIONESSTOCKMANUALES es VARCHAR (auditoría)
                stock_desc = f'AJUSTE API {hoy_str}'
                fb_cur.execute("""
                    INSERT INTO "CORRECCIONESSTOCKMANUALES"
                    (NUMEROMOVIMIENTO,FECHA,CODIGOUSUARIO,INGRESO,EGRESO,
                     CODIGOARTICULO,LOTE,STOCK,OBSERVACIONES,CODIGODEPOSITO,
                     NUMEROTRANSACCION,COSTOUNITARIO,CODIGOMOTIVOAJUSTE,
                     FECHAMODIFICACION,HORA)
                    VALUES (?,?,?,?,?,?,?,?,?,?,0,0.0,1,?,?)
                """, (next_num, hoy_date, usuario, ingreso, egreso,
                      cod_art, '000', stock_desc,
                      f'Ajuste inventario {hoy_str} dep {deposito}', deposito,
                      ahora, ahora))
                next_num += 1
                # 1. CASILLEROS: stock real por depósito (FMA_STOCK lee de aquí)
                fb_cur.execute(
                    "UPDATE \"CASILLEROS\" SET STOCKACTUAL=? WHERE CODIGOARTICULO=? AND CODIGODEPOSITO=? AND LOTE='000'",
                    (stock_nvo, cod_art, deposito))
                if fb_cur.rowcount == 0:
                    fb_cur.execute(
                        "UPDATE \"CASILLEROS\" SET STOCKACTUAL=? WHERE CODIGOARTICULO=? AND CODIGODEPOSITO=?",
                        (stock_nvo, cod_art, deposito))
                # 2. STOCK global: recalcular como SUM(CASILLEROS) para mantener consistencia
                fb_cur.execute(
                    "UPDATE \"STOCK\" SET STOCKACTUAL=(SELECT SUM(STOCKACTUAL) FROM \"CASILLEROS\" WHERE CODIGOARTICULO=?),FECHAMODIFICACION=? WHERE CODIGOARTICULO=?",
                    (cod_art, ahora, cod_art))
                procesados += 1
            except Exception as e_art:
                detalle_errors.append(f"{cod_art}: {e_art}")
        fb.commit(); fb.close()
        estado = 'ok' if not detalle_errors else 'parcial'
    except Exception as e:
        estado = 'error'; detalle_errors.append(str(e))
    db.execute("UPDATE stock_ajuste_log SET estado=?,detalle=?,total_articulos=? WHERE id=?",
               (estado, '; '.join(detalle_errors)[:500] or None, procesados, log_id))
    db.commit(); db.close()
    if estado == 'error':
        raise HTTPException(500, detalle_errors[0] if detalle_errors else 'Error')
    return {"ok": True, "procesados": procesados, "con_pendientes": con_pendientes,
            "estado": estado, "errores": detalle_errors}

@app.post("/admin/ajuste-stock/revertir/{log_id}")
def ajuste_stock_revertir(log_id: int, _u=Depends(get_gerente_user)):
    """Revierte un ajuste restaurando CASILLEROS y STOCK desde el backup de SQLite."""
    db = _admin_db()
    log = db.execute(
        "SELECT id, deposito, estado FROM stock_ajuste_log WHERE id=?", (log_id,)
    ).fetchone()
    if not log:
        db.close(); raise HTTPException(404, f"Log {log_id} no encontrado")
    deposito = log["deposito"]
    backups = db.execute(
        "SELECT codigo_articulo, stock_anterior FROM stock_ajuste_backup WHERE log_id=?", (log_id,)
    ).fetchall()
    if not backups:
        db.close(); raise HTTPException(404, "Sin backup para este ajuste")
    db.close()

    from datetime import datetime as _dt
    ahora_rev = _dt.now()
    revertidos = []; errores = []
    try:
        fb = conn('WIN1252'); cur = fb.cursor()
        for b in backups:
            cod_art = str(b["codigo_articulo"]).strip()
            stock_ant = float(b["stock_anterior"])
            try:
                cur.execute(
                    "UPDATE \"CASILLEROS\" SET STOCKACTUAL=? WHERE CODIGOARTICULO=? AND CODIGODEPOSITO=? AND LOTE='000'",
                    (stock_ant, cod_art, deposito))
                if cur.rowcount == 0:
                    cur.execute(
                        "UPDATE \"CASILLEROS\" SET STOCKACTUAL=? WHERE CODIGOARTICULO=? AND CODIGODEPOSITO=?",
                        (stock_ant, cod_art, deposito))
                cur.execute(
                    "UPDATE \"STOCK\" SET STOCKACTUAL=(SELECT SUM(STOCKACTUAL) FROM \"CASILLEROS\" WHERE CODIGOARTICULO=?),FECHAMODIFICACION=? WHERE CODIGOARTICULO=?",
                    (cod_art, ahora_rev, cod_art))
                revertidos.append(cod_art)
            except Exception as e:
                errores.append(f"{cod_art}: {e}")
        fb.commit(); fb.close()
    except Exception as e:
        raise HTTPException(500, str(e))
    return {"ok": True, "log_id": log_id, "deposito": deposito,
            "revertidos": len(revertidos), "errores": errores}

@app.get("/admin/ajuste-stock/historial")
def ajuste_stock_historial(_u=Depends(get_gerente_user)):
    db = _admin_db()
    rows = db.execute(
        "SELECT id,fecha,usuario,deposito,deposito_nombre,filtro_desc,total_articulos,con_pendientes,estado,detalle "
        "FROM stock_ajuste_log ORDER BY fecha DESC LIMIT 50"
    ).fetchall()
    result = [dict(r) for r in rows]
    db.close()
    # Resolver nombres faltantes desde Firebird y persistirlos
    codigos_sin_nombre = list({r['deposito'] for r in result if not r.get('deposito_nombre')})
    if codigos_sin_nombre:
        try:
            placeholders = ','.join('?' * len(codigos_sin_nombre))
            _c = conn('WIN1252'); _cur = _c.cursor()
            _cur.execute(f'SELECT CODIGODEPOSITO, DESCRIPCION FROM "DEPOSITOS" WHERE CODIGODEPOSITO IN ({placeholders})', codigos_sin_nombre)
            mapa = {str(r[0]).strip(): str(r[1] or '').strip() for r in _cur.fetchall()}
            _c.close()
            # Actualizar SQLite para que próximas consultas ya traigan el nombre
            _db2 = _admin_db()
            for cod, nombre in mapa.items():
                if nombre:
                    _db2.execute("UPDATE stock_ajuste_log SET deposito_nombre=? WHERE deposito=? AND (deposito_nombre IS NULL OR deposito_nombre='')", (nombre, cod))
            _db2.commit(); _db2.close()
            # Aplicar al resultado en memoria
            for r in result:
                if not r.get('deposito_nombre'):
                    r['deposito_nombre'] = mapa.get(str(r['deposito']).strip(), r['deposito'])
        except Exception:
            pass
    return result

@app.delete("/admin/ajuste-stock/historial")
def ajuste_stock_limpiar_historial(_u=Depends(get_gerente_user)):
    """Borra todo el historial de ajustes de stock (y sus backups en cascada)."""
    db = _admin_db()
    db.execute("DELETE FROM stock_ajuste_log")
    db.commit()
    deleted = db.execute("SELECT changes()").fetchone()[0]
    db.close()
    return {"deleted": deleted}

# ═══════════════════════════════════════════════════════════════════════════════
# ─── TRANSFERENCIA ENTRE DEPÓSITOS ───────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════

def get_transferencia_user(credentials: HTTPAuthorizationCredentials = Depends(_bearer)):
    if not credentials:
        raise HTTPException(401, "No autenticado")
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGO])
        if payload.get('role') != 'admin':
            raise HTTPException(403, "Acceso denegado")
        perfil = str(payload.get('perfil') or '').strip().upper()
        if perfil not in _PERFILES_TRANSFERENCIA:
            raise HTTPException(403, f"Perfil '{perfil}' no tiene acceso a Transferencia")
        return payload
    except JWTError:
        raise HTTPException(401, "Token inválido o expirado")

@app.get("/admin/transferencia/buscar-articulo")
def transferencia_buscar_articulo(q: str = Query(""), _u=Depends(get_transferencia_user)):
    """Autocompletar artículo por SKU (codigoparticular) o descripción."""
    q = q.strip()
    if len(q) < 2:
        return []
    try:
        c = conn('WIN1252'); cur = c.cursor()
        cur.execute("""
            SELECT FIRST 15 CODIGOARTICULO, CODIGOPARTICULAR, DESCRIPCION
            FROM "ARTICULOS"
            WHERE (UPPER(CODIGOPARTICULAR) STARTING WITH UPPER(?))
               OR (UPPER(DESCRIPCION) CONTAINING UPPER(?))
            ORDER BY CODIGOPARTICULAR
        """, (q.upper(), q.upper()))
        rows = cur.fetchall(); c.close()
        return [{"codigo_articulo": str(r[0] or '').strip(),
                 "codigo_particular": str(r[1] or '').strip(),
                 "descripcion": str(r[2] or '').strip()} for r in rows]
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/admin/transferencia/preview")
def transferencia_preview(body: dict, _u=Depends(get_transferencia_user)):
    dep_origen  = str(body.get('deposito_origen')  or '').strip()
    dep_destino = str(body.get('deposito_destino') or '').strip()
    articulos   = body.get('articulos') or []  # [{sku, cantidad}]
    if not dep_origen or not dep_destino:
        raise HTTPException(400, "Seleccioná origen y destino")
    if dep_origen == dep_destino:
        raise HTTPException(400, "Origen y destino deben ser distintos")
    if not articulos:
        raise HTTPException(400, "Sin artículos")

    skus     = [str(a['sku']).strip() for a in articulos if a.get('sku')]
    qty_map  = {str(a['sku']).strip(): float(a.get('cantidad', 0)) for a in articulos if a.get('sku')}
    if not skus:
        raise HTTPException(400, "Sin SKUs válidos")
    placeholders = ','.join(['?' for _ in skus])

    try:
        c = conn('WIN1252'); cur = c.cursor()
        cur.execute(f'SELECT CODIGOARTICULO, CODIGOPARTICULAR, DESCRIPCION FROM "ARTICULOS" WHERE CODIGOPARTICULAR IN ({placeholders})', skus)
        arts = {str(r[1] or '').strip(): {'codigo_articulo': str(r[0] or '').strip(), 'descripcion': str(r[2] or '').strip()} for r in cur.fetchall()}

        cod_arts = [v['codigo_articulo'] for v in arts.values() if v['codigo_articulo']]
        stock_origen = {}; stock_destino = {}
        if cod_arts:
            ph2 = ','.join(['?' for _ in cod_arts])
            cur.execute(f"SELECT CODIGOARTICULO, STOCKACTUAL FROM \"CASILLEROS\" WHERE CODIGOARTICULO IN ({ph2}) AND CODIGODEPOSITO=? AND LOTE='000'", cod_arts + [dep_origen])
            stock_origen = {str(r[0] or '').strip(): float(r[1] or 0) for r in cur.fetchall()}
            cur.execute(f"SELECT CODIGOARTICULO, STOCKACTUAL FROM \"CASILLEROS\" WHERE CODIGOARTICULO IN ({ph2}) AND CODIGODEPOSITO=? AND LOTE='000'", cod_arts + [dep_destino])
            stock_destino = {str(r[0] or '').strip(): float(r[1] or 0) for r in cur.fetchall()}

        rem_origen = {}; rem_destino = {}
        try:
            cur.execute(f"SELECT ID_ARTICULO, STOCKREMANENTE FROM \"FMA_STOCK\"(NULL, NULL, '{dep_origen}', 1, 1)")
            rem_origen = {str(r[0] or '').strip(): float(r[1] or 0) for r in cur.fetchall()}
        except Exception: pass
        try:
            cur.execute(f"SELECT ID_ARTICULO, STOCKREMANENTE FROM \"FMA_STOCK\"(NULL, NULL, '{dep_destino}', 1, 1)")
            rem_destino = {str(r[0] or '').strip(): float(r[1] or 0) for r in cur.fetchall()}
        except Exception: pass
        c.close()

        no_encontrados = [s for s in skus if s not in arts]
        resultado = []
        for sku in skus:
            if sku not in arts: continue
            art = arts[sku]; cod = art['codigo_articulo']
            cant = qty_map.get(sku, 0)
            st_orig = stock_origen.get(cod, 0); st_dest = stock_destino.get(cod, 0)
            rem_orig = rem_origen.get(cod, 0);  rem_dest = rem_destino.get(cod, 0)
            resultado.append({
                'codigo_articulo': cod, 'codigo_particular': sku,
                'descripcion': art['descripcion'], 'cantidad': cant,
                'stock_origen': st_orig,        'stock_origen_post': st_orig - cant,
                'remanente_origen': rem_orig,   'remanente_origen_post': rem_orig - cant,
                'stock_destino': st_dest,       'stock_destino_post': st_dest + cant,
                'remanente_destino': rem_dest,  'remanente_destino_post': rem_dest + cant,
                'alerta': st_orig < cant
            })
        return {'articulos': resultado, 'no_encontrados': no_encontrados}
    except HTTPException: raise
    except Exception as e: raise HTTPException(500, str(e))

def _procesar_transferencia_stock(dep_origen, dep_destino, dep_origen_nombre, dep_destino_nombre,
                                   articulos, usuario, fecha_operacion=None):
    """Lógica real (Firebird) de transferencia entre depósitos — mueve CASILLEROS.STOCKACTUAL,
    recalcula STOCK y deja auditoría en CORRECCIONESSTOCKMANUALES. Extraída de
    /admin/transferencia/procesar para poder reutilizarla también desde el endpoint de
    transferencia automática disparado desde frontend.html (sesión de vendedor)."""
    dep_origen         = str(dep_origen or '').strip()
    dep_destino        = str(dep_destino or '').strip()
    dep_origen_nombre  = str(dep_origen_nombre or dep_origen).strip()
    dep_destino_nombre = str(dep_destino_nombre or dep_destino).strip()

    if not dep_origen or not dep_destino:
        raise HTTPException(400, "Seleccioná origen y destino")
    if dep_origen == dep_destino:
        raise HTTPException(400, "Origen y destino deben ser distintos")
    if not articulos:
        raise HTTPException(400, "Sin artículos")

    from datetime import date as _date, datetime as _dt
    # Fecha de operación: la que manda el frontend (para registrar en Flexxus)
    fecha_op_str = str(fecha_operacion or '').strip()
    try:
        hoy_date = _date.fromisoformat(fecha_op_str) if fecha_op_str else _date.today()
    except ValueError:
        hoy_date = _date.today()
    ahora = _dt.now()

    db = _admin_db(); cur_db = db.cursor()
    cur_db.execute(
        "INSERT INTO transferencia_log (usuario, deposito_origen, deposito_origen_nombre, deposito_destino, deposito_destino_nombre, estado) VALUES (?,?,?,?,?,'pending')",
        (usuario, dep_origen, dep_origen_nombre, dep_destino, dep_destino_nombre)
    )
    log_id = cur_db.lastrowid; db.commit()

    procesados = 0; errores = []

    try:
        fb = conn('WIN1252'); fb_cur = fb.cursor()
        fb_cur.execute('SELECT MAX(NUMEROMOVIMIENTO) FROM "CORRECCIONESSTOCKMANUALES"')
        next_num = int(fb_cur.fetchone()[0] or 0) + 1

        for a in articulos:
            cod_art  = str(a.get('codigo_articulo') or '').strip()
            sku      = str(a.get('codigo_particular') or '').strip()
            desc     = str(a.get('descripcion') or '').strip()
            cantidad = float(a.get('cantidad', 0))
            if not cod_art or cantidad <= 0: continue
            try:
                # Leer stock actual de origen
                fb_cur.execute("SELECT STOCKACTUAL FROM \"CASILLEROS\" WHERE CODIGOARTICULO=? AND CODIGODEPOSITO=? AND LOTE='000'", (cod_art, dep_origen))
                row = fb_cur.fetchone(); st_orig = float(row[0] if row else 0)
                # Leer stock actual de destino
                fb_cur.execute("SELECT STOCKACTUAL FROM \"CASILLEROS\" WHERE CODIGOARTICULO=? AND CODIGODEPOSITO=? AND LOTE='000'", (cod_art, dep_destino))
                row = fb_cur.fetchone(); st_dest = float(row[0] if row else 0)

                # Guardar backup en SQLite
                cur_db.execute(
                    "INSERT INTO transferencia_detalle (log_id, codigo_articulo, codigo_particular, descripcion, cantidad, stock_origen_anterior, stock_destino_anterior) VALUES (?,?,?,?,?,?,?)",
                    (log_id, cod_art, sku, desc, cantidad, st_orig, st_dest)
                )

                # Actualizar CASILLEROS origen (baja)
                fb_cur.execute("UPDATE \"CASILLEROS\" SET STOCKACTUAL=? WHERE CODIGOARTICULO=? AND CODIGODEPOSITO=? AND LOTE='000'", (max(0.0, st_orig - cantidad), cod_art, dep_origen))
                if fb_cur.rowcount == 0:
                    fb_cur.execute("UPDATE \"CASILLEROS\" SET STOCKACTUAL=? WHERE CODIGOARTICULO=? AND CODIGODEPOSITO=?", (max(0.0, st_orig - cantidad), cod_art, dep_origen))

                # Actualizar CASILLEROS destino (suba)
                fb_cur.execute("UPDATE \"CASILLEROS\" SET STOCKACTUAL=? WHERE CODIGOARTICULO=? AND CODIGODEPOSITO=? AND LOTE='000'", (st_dest + cantidad, cod_art, dep_destino))
                if fb_cur.rowcount == 0:
                    fb_cur.execute("UPDATE \"CASILLEROS\" SET STOCKACTUAL=? WHERE CODIGOARTICULO=? AND CODIGODEPOSITO=?", (st_dest + cantidad, cod_art, dep_destino))

                # Recalcular STOCK global
                fb_cur.execute("UPDATE \"STOCK\" SET STOCKACTUAL=(SELECT SUM(STOCKACTUAL) FROM \"CASILLEROS\" WHERE CODIGOARTICULO=?),FECHAMODIFICACION=? WHERE CODIGOARTICULO=?", (cod_art, ahora, cod_art))

                stock_desc = f'TRANSF API {hoy_date.isoformat()}'
                # Auditoría — EGRESO de origen
                fb_cur.execute("""
                    INSERT INTO "CORRECCIONESSTOCKMANUALES"
                    (NUMEROMOVIMIENTO,FECHA,CODIGOUSUARIO,INGRESO,EGRESO,
                     CODIGOARTICULO,LOTE,STOCK,OBSERVACIONES,CODIGODEPOSITO,
                     NUMEROTRANSACCION,COSTOUNITARIO,CODIGOMOTIVOAJUSTE,FECHAMODIFICACION,HORA)
                    VALUES (?,?,?,0,?,?,?,?,?,?,0,0.0,1,?,?)
                """, (next_num, hoy_date, usuario, cantidad, cod_art, '000', stock_desc,
                      f'Transferencia a dep {dep_destino} | TRF-{log_id}', dep_origen, ahora, ahora))
                next_num += 1
                # Auditoría — INGRESO a destino
                fb_cur.execute("""
                    INSERT INTO "CORRECCIONESSTOCKMANUALES"
                    (NUMEROMOVIMIENTO,FECHA,CODIGOUSUARIO,INGRESO,EGRESO,
                     CODIGOARTICULO,LOTE,STOCK,OBSERVACIONES,CODIGODEPOSITO,
                     NUMEROTRANSACCION,COSTOUNITARIO,CODIGOMOTIVOAJUSTE,FECHAMODIFICACION,HORA)
                    VALUES (?,?,?,?,0,?,?,?,?,?,0,0.0,1,?,?)
                """, (next_num, hoy_date, usuario, cantidad, cod_art, '000', stock_desc,
                      f'Transferencia desde dep {dep_origen} | TRF-{log_id}', dep_destino, ahora, ahora))
                next_num += 1
                procesados += 1
            except Exception as e_art:
                errores.append(f"{sku}: {e_art}")
        fb.commit(); fb.close()
        db.commit()
    except Exception as e:
        errores.append(str(e))

    estado = 'ok' if not errores else ('parcial' if procesados > 0 else 'error')
    db.execute("UPDATE transferencia_log SET estado=?,detalle=?,total_articulos=? WHERE id=?",
               (estado, '; '.join(errores)[:500] or None, procesados, log_id))
    db.commit(); db.close()
    if estado == 'error':
        raise HTTPException(500, errores[0] if errores else 'Error')
    return {"ok": True, "log_id": log_id, "procesados": procesados, "estado": estado, "errores": errores}

@app.post("/admin/transferencia/procesar")
def transferencia_procesar(body: dict, _u=Depends(get_transferencia_user)):
    return _procesar_transferencia_stock(
        body.get('deposito_origen'), body.get('deposito_destino'),
        body.get('deposito_origen_nombre'), body.get('deposito_destino_nombre'),
        body.get('articulos') or [], _u.get('sub', '?'),
        fecha_operacion=body.get('fecha_operacion')
    )

# ─── Transferencia automática disparada desde frontend.html (sesión de vendedor) ──
# Cuando un perfil privilegiado (ADV/ADVJUAN/GERENTES/GTES FE) carga un pedido y el
# artículo no tiene remanente en el depósito seleccionado pero SÍ en un depósito que
# no arma pedidos (ver depositos_config / arma_pedidos), el frontend ofrece transferir
# automáticamente ese faltante. El token de frontend.html (impersonación) nunca trae
# role='admin', por eso get_transferencia_user (que lo exige) no sirve acá — este
# dependency acepta también sesiones de vendedor, validando el perfil igual de estricto.
def get_transferencia_actor(credentials: HTTPAuthorizationCredentials = Depends(_bearer)):
    if not credentials:
        raise HTTPException(401, "No autenticado")
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGO])
    except JWTError:
        raise HTTPException(401, "Token inválido o expirado")
    if payload.get('role') == 'admin':
        perfil = str(payload.get('perfil') or '').strip().upper()
    else:
        # Sesión de vendedor/impersonación: el perfil relevante es el del admin que
        # está actuando (admin_perfil), no el del vendedor impersonado (perfil).
        perfil = str(payload.get('admin_perfil') or payload.get('perfil') or '').strip().upper()
    if perfil not in _PERFILES_TRANSFERENCIA:
        raise HTTPException(403, f"Perfil '{perfil}' no tiene acceso a Transferencia")
    return payload

class TransferenciaAutomaticaBody(BaseModel):
    deposito_origen: str
    deposito_destino: str
    deposito_origen_nombre: str = ''
    deposito_destino_nombre: str = ''
    sku: str
    cantidad: float

@app.post("/pedidos/transferencia-automatica")
def pedidos_transferencia_automatica(body: TransferenciaAutomaticaBody, _u=Depends(get_transferencia_actor)):
    dep_origen  = body.deposito_origen.strip()
    dep_destino = body.deposito_destino.strip()
    sku         = body.sku.strip()
    cantidad    = float(body.cantidad)
    if not dep_origen or not dep_destino:
        raise HTTPException(400, "Depósito origen/destino requerido")
    if dep_origen == dep_destino:
        raise HTTPException(400, "Origen y destino deben ser distintos")
    if not sku or cantidad <= 0:
        raise HTTPException(400, "SKU/cantidad inválidos")

    try:
        c = conn('WIN1252'); cur = c.cursor()
        cur.execute('SELECT CODIGOARTICULO, DESCRIPCION FROM "ARTICULOS" WHERE CODIGOPARTICULAR=?', (sku,))
        row = cur.fetchone(); c.close()
    except Exception as e:
        raise HTTPException(500, f"Error consultando artículo: {e}")
    if not row:
        raise HTTPException(404, f"Artículo {sku} no encontrado")
    cod_art     = str(row[0] or '').strip()
    descripcion = str(row[1] or '').strip()

    usuario = _u.get('sub', '?')
    resultado = _procesar_transferencia_stock(
        dep_origen, dep_destino,
        body.deposito_origen_nombre or dep_origen, body.deposito_destino_nombre or dep_destino,
        [{'codigo_articulo': cod_art, 'codigo_particular': sku, 'descripcion': descripcion, 'cantidad': cantidad}],
        usuario
    )
    return resultado

@app.post("/admin/transferencia/revertir/{log_id}")
def transferencia_revertir(log_id: int, _u=Depends(get_transferencia_user)):
    db = _admin_db()
    log = db.execute("SELECT * FROM transferencia_log WHERE id=?", (log_id,)).fetchone()
    if not log:
        db.close(); raise HTTPException(404, f"Log {log_id} no encontrado")
    if log['revertida']:
        db.close(); raise HTTPException(400, "Esta transferencia ya fue revertida")
    detalles = db.execute("SELECT * FROM transferencia_detalle WHERE log_id=?", (log_id,)).fetchall()
    if not detalles:
        db.close(); raise HTTPException(404, "Sin detalle para esta transferencia")
    dep_origen  = log['deposito_origen']
    dep_destino = log['deposito_destino']
    db.close()

    from datetime import datetime as _dt
    ahora = _dt.now()
    revertidos = []; errores = []
    try:
        fb = conn('WIN1252'); fb_cur = fb.cursor()
        for d in detalles:
            cod_art    = d['codigo_articulo']
            st_orig_ant = float(d['stock_origen_anterior'] or 0)
            st_dest_ant = float(d['stock_destino_anterior'] or 0)
            try:
                fb_cur.execute("UPDATE \"CASILLEROS\" SET STOCKACTUAL=? WHERE CODIGOARTICULO=? AND CODIGODEPOSITO=? AND LOTE='000'", (st_orig_ant, cod_art, dep_origen))
                fb_cur.execute("UPDATE \"CASILLEROS\" SET STOCKACTUAL=? WHERE CODIGOARTICULO=? AND CODIGODEPOSITO=? AND LOTE='000'", (st_dest_ant, cod_art, dep_destino))
                fb_cur.execute("UPDATE \"STOCK\" SET STOCKACTUAL=(SELECT SUM(STOCKACTUAL) FROM \"CASILLEROS\" WHERE CODIGOARTICULO=?),FECHAMODIFICACION=? WHERE CODIGOARTICULO=?", (cod_art, ahora, cod_art))
                revertidos.append(cod_art)
            except Exception as e:
                errores.append(f"{cod_art}: {e}")
        fb.commit(); fb.close()
    except Exception as e:
        raise HTTPException(500, str(e))

    db2 = _admin_db()
    det = f"Revertida {ahora.strftime('%Y-%m-%d %H:%M')}" + (f" | Errores: {'; '.join(errores)}" if errores else '')
    db2.execute("UPDATE transferencia_log SET revertida=1, detalle=? WHERE id=?", (det, log_id))
    db2.commit(); db2.close()
    return {"ok": True, "log_id": log_id, "revertidos": len(revertidos), "errores": errores}

@app.get("/admin/transferencia/historial")
def transferencia_historial(_u=Depends(get_transferencia_user)):
    db = _admin_db()
    rows = db.execute(
        "SELECT id,fecha,usuario,deposito_origen,deposito_origen_nombre,deposito_destino,deposito_destino_nombre,"
        "total_articulos,estado,detalle,revertida FROM transferencia_log ORDER BY fecha DESC LIMIT 50"
    ).fetchall()
    result = []
    for r in rows:
        row = dict(r)
        detalles = db.execute(
            "SELECT codigo_particular,descripcion,cantidad,stock_origen_anterior,stock_destino_anterior FROM transferencia_detalle WHERE log_id=?",
            (r['id'],)
        ).fetchall()
        row['detalle_items'] = [dict(d) for d in detalles]
        result.append(row)
    db.close()
    return result


def _transferencia_get(log_id: int):
    db = _admin_db()
    log = db.execute(
        "SELECT id,fecha,usuario,deposito_origen,deposito_origen_nombre,deposito_destino,deposito_destino_nombre,"
        "total_articulos,estado,detalle,revertida FROM transferencia_log WHERE id=?", (log_id,)
    ).fetchone()
    if not log:
        db.close()
        raise HTTPException(404, f"Transferencia {log_id} no encontrada")
    detalles = db.execute(
        "SELECT codigo_articulo,codigo_particular,descripcion,cantidad,stock_origen_anterior,stock_destino_anterior "
        "FROM transferencia_detalle WHERE log_id=? ORDER BY id", (log_id,)
    ).fetchall()
    db.close()
    return dict(log), [dict(d) for d in detalles]


@app.get("/admin/transferencia/{log_id}/excel")
def transferencia_exportar_excel(log_id: int, _u=Depends(get_transferencia_user)):
    log, detalles = _transferencia_get(log_id)
    try:
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transferencia"

        hdr_fill   = PatternFill("solid", fgColor="1A56DB")
        hdr_font   = Font(bold=True, color="FFFFFFFF", size=10)
        title_font = Font(bold=True, size=13, color="111827")
        sub_font   = Font(size=9.5, color="374151")
        err_font   = Font(size=8.5, italic=True, color="DC2626")
        right_al   = Alignment(horizontal="right", vertical="center")
        center_al  = Alignment(horizontal="center", vertical="center")
        left_al    = Alignment(horizontal="left", vertical="center")
        NUM_COLS   = 5

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
        c = ws.cell(1, 1, f"MICROBELL S.A. — Comprobante de Transferencia entre Depósitos  ·  TRF-{log['id']}")
        c.font = title_font; c.alignment = left_al
        ws.row_dimensions[1].height = 22

        origen  = log.get('deposito_origen_nombre') or log.get('deposito_origen')
        destino = log.get('deposito_destino_nombre') or log.get('deposito_destino')
        estado  = 'Revertida' if log.get('revertida') else log.get('estado')
        resumen = (f"Fecha: {log.get('fecha')}  ·  Usuario: {log.get('usuario')}  ·  "
                   f"Origen: {origen}  →  Destino: {destino}  ·  "
                   f"Artículos: {log.get('total_articulos')}  ·  Estado: {estado}")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NUM_COLS)
        c2 = ws.cell(2, 1, resumen)
        c2.font = sub_font; c2.alignment = left_al
        ws.row_dimensions[2].height = 18

        next_row = 4
        if log.get('detalle'):
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=NUM_COLS)
            c3 = ws.cell(3, 1, str(log['detalle'])[:300])
            c3.font = err_font; c3.alignment = left_al
            ws.row_dimensions[3].height = 16
            next_row = 5
        ws.row_dimensions[next_row - 1].height = 6

        HEADERS = ["SKU", "Descripción", "Cantidad", "Stock Origen (antes)", "Stock Destino (antes)"]
        for ci, h in enumerate(HEADERS, 1):
            cell = ws.cell(next_row, ci, h)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center_al
        ws.row_dimensions[next_row].height = 20

        for ri, d in enumerate(detalles, next_row + 1):
            ws.cell(ri, 1, d.get('codigo_particular', ''))
            wc2 = ws.cell(ri, 2, d.get('descripcion', '')); wc2.alignment = left_al
            wc3 = ws.cell(ri, 3, d.get('cantidad', 0));     wc3.alignment = right_al
            wc4 = ws.cell(ri, 4, d.get('stock_origen_anterior', 0));  wc4.alignment = right_al
            wc5 = ws.cell(ri, 5, d.get('stock_destino_anterior', 0)); wc5.alignment = right_al

        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 44
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        fname = f"transferencia_TRF-{log_id}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/admin/transferencia/{log_id}/pdf")
def transferencia_exportar_pdf(log_id: int, _u=Depends(get_transferencia_user)):
    log, detalles = _transferencia_get(log_id)
    try:
        import io

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                                leftMargin=15*mm, rightMargin=15*mm,
                                topMargin=14*mm, bottomMargin=14*mm)

        st_title = ParagraphStyle('trf_title', fontName='Helvetica-Bold', fontSize=14,
                                   textColor=colors.HexColor('#111827'), spaceAfter=4)
        st_sub   = ParagraphStyle('trf_sub', fontName='Helvetica', fontSize=9.5,
                                   textColor=colors.HexColor('#374151'), spaceAfter=2, leading=13)
        st_err   = ParagraphStyle('trf_err', fontName='Helvetica-Oblique', fontSize=8.5,
                                   textColor=colors.HexColor('#DC2626'), spaceAfter=8)
        st_desc  = ParagraphStyle('trf_desc', fontName='Helvetica', fontSize=8.5,
                                   textColor=colors.HexColor('#111827'), leading=10)

        origen  = log.get('deposito_origen_nombre') or log.get('deposito_origen')
        destino = log.get('deposito_destino_nombre') or log.get('deposito_destino')
        estado  = 'Revertida' if log.get('revertida') else log.get('estado')

        story = []
        story.append(Paragraph(
            f"MICROBELL S.A. — Comprobante de Transferencia entre Depósitos &nbsp;·&nbsp; TRF-{log['id']}",
            st_title
        ))
        story.append(Paragraph(
            f"Fecha: {log.get('fecha')} &nbsp;·&nbsp; Usuario: {log.get('usuario')} &nbsp;·&nbsp; "
            f"Origen: <b>{origen}</b> &nbsp;→&nbsp; Destino: <b>{destino}</b> &nbsp;·&nbsp; "
            f"Artículos: {log.get('total_articulos')} &nbsp;·&nbsp; Estado: {estado}", st_sub
        ))
        if log.get('detalle'):
            story.append(Paragraph(str(log['detalle'])[:300], st_err))
        story.append(Spacer(1, 8))

        BLUE  = colors.HexColor('#1A56DB')
        ALTBG = colors.HexColor('#EFF6FF')

        tbl_data = [["SKU", "Descripción", "Cantidad", "Stock Origen\n(antes)", "Stock Destino\n(antes)"]]
        for d in detalles:
            tbl_data.append([
                d.get('codigo_particular', ''),
                Paragraph(d.get('descripcion', '') or '', st_desc),
                str(d.get('cantidad', 0)),
                str(d.get('stock_origen_anterior', 0)),
                str(d.get('stock_destino_anterior', 0)),
            ])

        tbl = Table(tbl_data, colWidths=[28*mm, 82*mm, 22*mm, 30*mm, 30*mm], repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,0), BLUE),
            ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
            ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,0), (-1,0), 8.5),
            ('ALIGN',         (0,0), (-1,0), 'CENTER'),
            ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
            ('FONTNAME',      (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE',      (0,1), (-1,-1), 8.5),
            ('ROWBACKGROUND', (0,1), (-1,-1), [colors.white, ALTBG]),
            ('GRID',          (0,0), (-1,-1), 0.3, colors.HexColor('#D1D5DB')),
            ('ALIGN',         (2,1), (4,-1), 'RIGHT'),
        ]))
        story.append(tbl)
        doc.build(story)
        buf.seek(0)
        fname = f"transferencia_TRF-{log_id}.pdf"
        return StreamingResponse(buf, media_type="application/pdf",
                                  headers={"Content-Disposition": f'attachment; filename="{fname}"'})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/admin/audit-event")
def admin_audit_event(data: dict, request: Request, _u=Depends(get_admin_user)):
    """El frontend registra acciones con contexto semántico (navegación, creación, etc.)"""
    accion  = str(data.get('accion')  or '').strip()[:120]
    detalle = str(data.get('detalle') or '').strip()[:300]
    seccion = str(data.get('seccion') or '').strip()[:60]
    if not accion:
        raise HTTPException(400, "accion requerida")
    ip = request.client.host if request.client else ''
    _audit(_u.get('sub', '?'), accion, detalle, ip, seccion)
    return {"ok": True}

@app.get("/admin/audit-log")
def admin_audit_log(
    desde:  Optional[str] = Query(None),
    hasta:  Optional[str] = Query(None),
    offset: int = Query(0, ge=0),
    limit:  int = Query(30, ge=1, le=200),
    _u=Depends(get_gerente_user)
):
    db = _admin_db()
    conds, params = [], []
    if desde:
        conds.append("fecha >= ?"); params.append(desde + " 00:00:00")
    if hasta:
        conds.append("fecha <= ?"); params.append(hasta + " 23:59:59")
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    total = db.execute(f"SELECT COUNT(*) FROM admin_audit_log {where}", params).fetchone()[0]
    rows  = db.execute(
        f"SELECT id,fecha,usuario,seccion,accion,detalle,ip,metodo,endpoint FROM admin_audit_log {where} ORDER BY fecha DESC LIMIT ? OFFSET ?",
        params + [limit, offset]
    ).fetchall()
    db.close()
    return {"total": total, "offset": offset, "limit": limit, "rows": [dict(r) for r in rows]}

# ═══════════════════════════════════════════════════════════════════════════════
# ─── CATÁLOGOS (archivos PDF/Excel) ──────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════

# ── Helpers de notificación ───────────────────────────────────────────────────
def _send_email_catalogo(destinatarios: list[str], nombre_catalogo: str, url: str, descripcion: str = ''):
    """Envía email de nuevo catálogo a lista de mails."""
    if not destinatarios or not SMTP_HOST:
        return
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f'Nuevo catálogo publicado: {nombre_catalogo}'
        msg['From']    = SMTP_FROM
        msg['To']      = ', '.join(destinatarios)
        html = f"""
        <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto">
          <div style="background:#1a56db;padding:20px;text-align:center">
            <h2 style="color:#fff;margin:0">Microbell S.A.</h2>
          </div>
          <div style="padding:24px;background:#f9fafb;border:1px solid #e5e7eb">
            <h3 style="color:#1a1a2e">📚 Nuevo catálogo disponible</h3>
            <p style="color:#374151">Se publicó el catálogo <strong>{nombre_catalogo}</strong>.</p>
            <p style="text-align:center;margin:28px 0">
              <a href="{url}" style="background:#1a56db;color:#fff;padding:12px 28px;border-radius:8px;text-decoration:none;font-weight:700">
                Ver catálogo
              </a>
            </p>
            <p style="color:#9ca3af;font-size:.82rem">Si el botón no funciona, copiá este link: {url}</p>
          </div>
        </div>"""
        msg.attach(MIMEText(html, 'html', 'utf-8'))
        if SMTP_PORT == 465:
            with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT) as s:
                s.login(SMTP_USER, SMTP_PASS)
                s.sendmail(SMTP_FROM, destinatarios, msg.as_bytes())
        else:
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
                s.ehlo(); s.starttls(); s.ehlo()
                s.login(SMTP_USER, SMTP_PASS)
                s.sendmail(SMTP_FROM, destinatarios, msg.as_bytes())
    except Exception as e:
        import traceback
        print(f"[EMAIL] Error: {e}\n{traceback.format_exc()}")
        raise  # propagar para que _notificar_catalogo_bg lo capture

def _send_push_catalogo(nombre_catalogo: str, url: str, descripcion: str = '') -> int:
    """Envía push notification via OneSignal a todos los suscriptores.
    Retorna 1 si se envió OK, 0 si no configurado o error.
    """
    if not ONESIGNAL_APP_ID or not ONESIGNAL_API_KEY:
        return 0
    try:
        mensaje = descripcion if descripcion else f"Nuevo catálogo disponible: {nombre_catalogo}"
        payload = json.dumps({
            "app_id": ONESIGNAL_APP_ID,
            "included_segments": ["Total Subscriptions"],
            "headings": {"es": nombre_catalogo, "en": nombre_catalogo},
            "contents": {"es": mensaje, "en": mensaje},
            "url": url,
        }).encode()
        req = urllib.request.Request(
            "https://onesignal.com/api/v1/notifications",
            data=payload,
            headers={
                'Authorization': f'Key {ONESIGNAL_API_KEY}',
                'Content-Type': 'application/json'
            },
            method='POST'
        )
        resp = urllib.request.urlopen(req, timeout=15)
        body = resp.read().decode('utf-8', errors='replace')
        print(f"[PUSH] OK: {body}")
        return 1
    except urllib.error.HTTPError as e:
        print(f"[PUSH] HTTPError {e.code}: {e.read().decode('utf-8', errors='replace')}")
    except Exception as e:
        print(f"[PUSH] Error: {e}")
    return 0

def _normalizar_celular_ar(cel) -> str:
    """Formato requerido por WhatsApp Cloud API para moviles argentinos: 54 9 <area><numero>.
    Si falta el 9 despues del codigo de pais (54), lo inserta. Sin esto, Meta devuelve
    error 133010 'Account not registered' aunque el numero sea correcto."""
    import re
    cel = re.sub(r'\D', '', str(cel or ''))
    if cel.startswith('54') and not cel.startswith('549'):
        cel = '549' + cel[2:]
    return cel


def _send_whatsapp_catalogo(celulares: list, nombre_catalogo: str, url: str, descripcion: str = '', template_name: str = None) -> int:
    """Envía WA vía Meta Cloud API con template aprobada. Retorna cantidad enviada OK."""
    if not WA_PHONE_NUMBER_ID or not WA_ACCESS_TOKEN:
        print("[WA] WA_PHONE_NUMBER_ID o WA_ACCESS_TOKEN no configurados")
        return 0
    tpl = template_name or WA_TEMPLATE_CAT
    api_url = f"https://graph.facebook.com/v20.0/{WA_PHONE_NUMBER_ID}/messages"
    headers = {'Authorization': f'Bearer {WA_ACCESS_TOKEN}', 'Content-Type': 'application/json'}
    suffix = url.rsplit('/', 1)[-1]
    ok_count = 0
    for cel in celulares:
        cel = _normalizar_celular_ar(cel)
        if not cel:
            continue
        payload = json.dumps({
            "messaging_product": "whatsapp",
            "to": cel,
            "type": "template",
            "template": {
                "name": tpl,
                "language": {"code": "es_AR"},
                "components": [
                    {
                        "type": "body",
                        "parameters": [
                            {"type": "text", "text": nombre_catalogo}
                        ]
                    },
                    {
                        "type": "button",
                        "sub_type": "url",
                        "index": "0",
                        "parameters": [
                            {"type": "text", "text": suffix}
                        ]
                    }
                ]
            }
        }).encode()
        try:
            req_http = urllib.request.Request(api_url, data=payload, headers=headers, method='POST')
            with urllib.request.urlopen(req_http, timeout=15) as resp:
                body = resp.read().decode('utf-8', errors='replace')
                print(f"[WA] OK → {cel}: {body}")
                ok_count += 1
        except urllib.error.HTTPError as e:
            err_body = ''
            try: err_body = e.read().decode('utf-8', errors='replace')
            except Exception: pass
            print(f"[WA] HTTPError {e.code} → {cel}: {err_body}")
        except Exception as e:
            print(f"[WA] Error → {cel}: {e}")
    return ok_count


def _send_whatsapp_cobranzas(tipo_doc: str, numero: str, corredor: str, cliente_cod: str,
                              cliente_nombre: str, observacion: str) -> bool:
    """Notifica al Área de Cobranzas por WhatsApp que se cargó un pedido/presupuesto
    a un cliente con observaciones en el ABM, pendiente de su confirmación.
    Requiere la plantilla WA_TEMPLATE_COBRANZAS (categoría UTILITY) aprobada en Meta."""
    if not WA_PHONE_NUMBER_ID or not WA_ACCESS_TOKEN:
        print("[WA-COB] WA_PHONE_NUMBER_ID o WA_ACCESS_TOKEN no configurados")
        return False
    cel = _normalizar_celular_ar(WA_COBRANZAS_CEL)
    if not cel:
        print("[WA-COB] WA_COBRANZAS_CEL no configurado")
        return False
    api_url = f"https://graph.facebook.com/v20.0/{WA_PHONE_NUMBER_ID}/messages"
    headers = {'Authorization': f'Bearer {WA_ACCESS_TOKEN}', 'Content-Type': 'application/json'}
    tipo_lbl = 'Pedido' if tipo_doc == 'pedido' else 'Presupuesto'
    params = [tipo_lbl, str(numero), corredor or '-', f"{cliente_nombre} ({cliente_cod})", (observacion or '-')[:1000]]
    payload = json.dumps({
        "messaging_product": "whatsapp",
        "to": cel,
        "type": "template",
        "template": {
            "name": WA_TEMPLATE_COBRANZAS,
            "language": {"code": "es_AR"},
            "components": [
                {"type": "body", "parameters": [{"type": "text", "text": p} for p in params]}
            ]
        }
    }).encode()
    try:
        req_http = urllib.request.Request(api_url, data=payload, headers=headers, method='POST')
        with urllib.request.urlopen(req_http, timeout=15) as resp:
            body = resp.read().decode('utf-8', errors='replace')
            print(f"[WA-COB] OK → {cel}: {body}")
        return True
    except urllib.error.HTTPError as e:
        err_body = ''
        try: err_body = e.read().decode('utf-8', errors='replace')
        except Exception: pass
        print(f"[WA-COB] HTTPError {e.code} → {cel}: {err_body}")
    except Exception as e:
        print(f"[WA-COB] Error → {cel}: {e}")
    return False


def _notificar_catalogo_bg(catalogo_id: int, nombre: str, token: str, base_url: str):
    """Corre en background: lee destinatarios, envía email + push y registra resultado."""
    db = _admin_db()
    rows = db.execute("""
        SELECT vc.mail, vc.nombre, vc.celular
        FROM catalogo_vendedores cv
        JOIN vendedores_contacto vc ON vc.codigo = cv.codigo
        WHERE cv.catalogo_id = ? AND vc.activo = 1
    """, (catalogo_id,)).fetchall()
    cat_row = db.execute("SELECT descripcion FROM catalogos WHERE id=?", (catalogo_id,)).fetchone()
    descripcion = (cat_row['descripcion'] or '') if cat_row else ''
    url   = f"{base_url}/catalogo/{token}"
    mails    = [r['mail']    for r in rows if (r['mail']    or '').strip()]
    celulares = [r['celular'] for r in rows if (r['celular'] or '').strip()]
    # Enviar email
    email_ok = 0
    if mails:
        try:
            _send_email_catalogo(mails, nombre, url, descripcion)
            email_ok = len(mails)
        except Exception as e:
            print(f"[CAT] Email error: {e}")
    # Enviar push notification
    push_ok = _send_push_catalogo(nombre, url, descripcion)
    # Enviar WhatsApp
    wa_ok = _send_whatsapp_catalogo(celulares, nombre, url, descripcion)
    print(f"[CAT] Notificaciones: email={email_ok} push={push_ok} wa={wa_ok}")
    db.execute(
        "UPDATE catalogos SET email_enviado=?, email_count=?, push_enviado=?, push_count=?, wa_enviado=?, wa_count=? WHERE id=?",
        (1 if email_ok > 0 else 0, email_ok,
         push_ok, push_ok,
         1 if wa_ok > 0 else 0, wa_ok,
         catalogo_id)
    )
    db.commit()
    db.close()

# ── CRUD contactos de vendedores ──────────────────────────────────────────────
@app.get("/admin/vendedores-contacto")
def get_vendedores_contacto(_u=Depends(get_admin_user)):
    db = _admin_db()
    rows = db.execute("SELECT * FROM vendedores_contacto ORDER BY nombre").fetchall()
    db.close()
    return [dict(r) for r in rows]

@app.post("/admin/vendedores-contacto")
def upsert_vendedor_contacto(data: dict, _u=Depends(get_admin_user)):
    codigo  = str(data.get('codigo') or '').strip().upper()
    nombre  = str(data.get('nombre') or '').strip()
    if not codigo or not nombre:
        raise HTTPException(400, "codigo y nombre requeridos")
    db = _admin_db()
    db.execute("""INSERT INTO vendedores_contacto (codigo,nombre,mail,celular,activo)
                  VALUES (?,?,?,?,?)
                  ON CONFLICT(codigo) DO UPDATE SET
                    nombre=excluded.nombre, mail=excluded.mail,
                    celular=excluded.celular, activo=excluded.activo""",
               (codigo, nombre, data.get('mail',''), data.get('celular',''),
                1 if data.get('activo',1) else 0))
    db.commit(); db.close()
    return {"ok": True}

@app.delete("/admin/vendedores-contacto/{codigo}")
def delete_vendedor_contacto(codigo: str, _u=Depends(get_admin_user)):
    db = _admin_db()
    db.execute("DELETE FROM vendedores_contacto WHERE codigo=?", (codigo.upper(),))
    db.commit(); db.close()
    return {"ok": True}

# ── Upload catálogo ───────────────────────────────────────────────────────────
@app.post("/admin/catalogos/upload")
async def upload_catalogo(
    request: Request,
    background_tasks: BackgroundTasks,
    nombre:      str        = Form(...),
    descripcion: str        = Form(''),
    codigos:     str        = Form(''),      # JSON array de codigos individuales
    perfil_ids:  str        = Form(''),      # JSON array de profile_ids (resuelve codigos desde vendor_profile_assignments)
    archivo:     UploadFile = File(...),
    _u=Depends(get_admin_user)
):
    # Validar extensión
    ext = os.path.splitext(archivo.filename or '')[1].lower()
    if ext not in ('.pdf', '.xlsx', '.xls'):
        raise HTTPException(400, "Solo se permiten archivos PDF o Excel (.xlsx/.xls)")
    # Guardar archivo con nombre único
    token    = uuid.uuid4().hex
    filename = f"{token}{ext}"
    dest     = os.path.join(CATALOGOS_DIR, filename)
    with open(dest, 'wb') as f:
        shutil.copyfileobj(archivo.file, f)
    # Parsear vendedores individuales
    try:
        vend_codigos = set(str(c).strip().upper() for c in json.loads(codigos) if c) if codigos else set()
    except Exception:
        vend_codigos = set()
    # Resolver vendedores por perfil (intersectar con vendedores_contacto)
    try:
        pids = [int(x) for x in json.loads(perfil_ids) if str(x).isdigit()] if perfil_ids else []
    except Exception:
        pids = []
    if pids:
        db_tmp = _admin_db()
        ph = ','.join('?' * len(pids))
        rows_pa = db_tmp.execute(
            f"SELECT DISTINCT vpa.codigousuario FROM vendor_profile_assignments vpa "
            f"JOIN vendedores_contacto vc ON UPPER(vc.codigo)=UPPER(vpa.codigousuario) "
            f"WHERE vpa.profile_id IN ({ph}) AND vc.activo=1", pids
        ).fetchall()
        db_tmp.close()
        for r in rows_pa:
            vend_codigos.add(str(r['codigousuario']).strip().upper())
    vend_codigos = list(vend_codigos)
    # Insertar en DB
    db = _admin_db()
    cur = db.execute(
        "INSERT INTO catalogos (nombre, descripcion, filename, token, subido_por) VALUES (?,?,?,?,?)",
        (nombre.strip(), descripcion.strip(), filename, token, _u.get('sub','?'))
    )
    cat_id = cur.lastrowid
    for cod in vend_codigos:
        if cod:
            try:
                db.execute("INSERT OR IGNORE INTO catalogo_vendedores (catalogo_id,codigo) VALUES (?,?)", (cat_id, cod))
            except Exception:
                pass
    # Guardar texto de perfiles asociados para auditoría/display
    perfiles_texto = ''
    if pids:
        db_p = _admin_db()
        ph2 = ','.join('?' * len(pids))
        prows = db_p.execute(f"SELECT codigo FROM vendor_profiles WHERE id IN ({ph2})", pids).fetchall()
        db_p.close()
        perfiles_texto = ', '.join(r['codigo'] for r in prows) if prows else ''
    if perfiles_texto:
        db.execute("UPDATE catalogos SET perfiles_texto=? WHERE id=?", (perfiles_texto, cat_id))
    db.commit(); db.close()
    _audit(_u.get('sub','?'), 'Subió catálogo', f'{nombre} ({ext[1:].upper()}) → {len(vend_codigos)} vendedores', '', 'Catálogos')
    # Notificar en background
    base_url = str(request.base_url).rstrip('/')
    background_tasks.add_task(_notificar_catalogo_bg, cat_id, nombre.strip(), token, base_url)
    return {"ok": True, "id": cat_id, "token": token}

# ── Listar catálogos (admin, paginado) ────────────────────────────────────────
@app.get("/admin/catalogos")
def admin_list_catalogos(
    offset: int = Query(0, ge=0),
    limit:  int = Query(30, ge=1, le=100),
    _u=Depends(get_admin_user)
):
    db = _admin_db()
    total = db.execute("SELECT COUNT(*) FROM catalogos WHERE activo=1").fetchone()[0]
    try:
        rows = db.execute(
            "SELECT id,nombre,descripcion,filename,token,subido_por,fecha,"
            "COALESCE(email_enviado,0) AS email_enviado,"
            "COALESCE(email_count,0)   AS email_count,"
            "COALESCE(push_enviado,0)  AS push_enviado,"
            "COALESCE(push_count,0)    AS push_count,"
            "COALESCE(wa_enviado,0)    AS wa_enviado,"
            "COALESCE(wa_count,0)      AS wa_count,"
            "COALESCE(perfiles_texto,'') AS perfiles_texto "
            "FROM catalogos WHERE activo=1 ORDER BY fecha DESC LIMIT ? OFFSET ?",
            (limit, offset)
        ).fetchall()
    except Exception:
        # Fallback si las columnas nuevas aún no existen (servidor no reiniciado)
        rows = db.execute(
            "SELECT id,nombre,descripcion,filename,token,subido_por,fecha,"
            "0 AS email_enviado,0 AS email_count,0 AS push_enviado,0 AS push_count,'' AS perfiles_texto "
            "FROM catalogos WHERE activo=1 ORDER BY fecha DESC LIMIT ? OFFSET ?",
            (limit, offset)
        ).fetchall()
    result = []
    for r in rows:
        rd = dict(r)
        vends = db.execute("SELECT codigo FROM catalogo_vendedores WHERE catalogo_id=?", (r['id'],)).fetchall()
        rd['vendedores'] = [v['codigo'] for v in vends]
        result.append(rd)
    db.close()
    return {"total": total, "offset": offset, "limit": limit, "rows": result}

# ── Eliminar catálogo ─────────────────────────────────────────────────────────
@app.delete("/admin/catalogos/{cat_id}")
def delete_catalogo_file(cat_id: int, _u=Depends(get_admin_user)):
    db = _admin_db()
    row = db.execute("SELECT filename, nombre FROM catalogos WHERE id=?", (cat_id,)).fetchone()
    if not row:
        db.close(); raise HTTPException(404, "Catálogo no encontrado")
    # Borrar físico
    try:
        os.remove(os.path.join(CATALOGOS_DIR, row['filename']))
    except Exception:
        pass
    db.execute("DELETE FROM catalogos WHERE id=?", (cat_id,))
    db.commit(); db.close()
    _audit(_u.get('sub','?'), 'Eliminó catálogo', row['nombre'], '', 'Catálogos')
    return {"ok": True}

# ── Reenviar catálogo a destinatarios ────────────────────────────────────────
@app.post("/admin/catalogos/{cat_id}/reenviar")
async def reenviar_catalogo(cat_id: int, data: dict, background_tasks: BackgroundTasks,
                            request: Request, _u=Depends(get_admin_user)):
    db = _admin_db()
    cat = db.execute("SELECT id,nombre,token FROM catalogos WHERE id=? AND activo=1", (cat_id,)).fetchone()
    if not cat:
        db.close(); raise HTTPException(404, "Catálogo no encontrado")
    nombre = cat['nombre']
    token  = cat['token']
    # Resolver destinatarios igual que en upload
    perfil_ids = [int(x) for x in (data.get('perfil_ids') or []) if str(x).isdigit()]
    codigos    = list({str(c).strip().upper() for c in (data.get('codigos') or []) if c})
    if perfil_ids:
        ph = ','.join('?' * len(perfil_ids))
        rows_pa = db.execute(
            f"SELECT DISTINCT vpa.codigousuario FROM vendor_profile_assignments vpa "
            f"JOIN vendedores_contacto vc ON UPPER(vc.codigo)=UPPER(vpa.codigousuario) "
            f"WHERE vpa.profile_id IN ({ph}) AND vc.activo=1", perfil_ids
        ).fetchall()
        for r in rows_pa:
            codigos.append(str(r['codigousuario']).strip().upper())
    codigos = list(set(codigos))
    if not codigos:
        db.close()
        raise HTTPException(400, "No hay destinatarios con datos de contacto para los perfiles/vendedores seleccionados")
    # Dar acceso al catálogo en el portal del vendedor (visibilidad en /vendedor/catalogos)
    for cod in codigos:
        try:
            db.execute("INSERT OR IGNORE INTO catalogo_vendedores (catalogo_id,codigo) VALUES (?,?)", (cat_id, cod))
        except Exception:
            pass
    db.commit(); db.close()
    # Buscar contactos y enviar
    db2 = _admin_db()
    ph2 = ','.join('?' * len(codigos))
    contactos = [dict(r) for r in db2.execute(
        f"SELECT mail, celular, nombre FROM vendedores_contacto WHERE UPPER(codigo) IN ({ph2}) AND activo=1",
        [c.upper() for c in codigos]
    ).fetchall()]
    db2.close()
    base_url = str(request.base_url).rstrip('/')
    url = f"{base_url}/catalogo/{token}"
    mails     = [c['mail']    for c in contactos if (c['mail']    or '').strip()]
    celulares = [c['celular'] for c in contactos if (c['celular'] or '').strip()]
    background_tasks.add_task(_send_email_catalogo, mails, nombre, url)
    background_tasks.add_task(_send_push_catalogo, nombre, url)
    background_tasks.add_task(_send_whatsapp_catalogo, celulares, nombre, url)
    _audit(_u.get('sub','?'), 'Reenvió catálogo', f'{nombre} → {len(contactos)} destinatarios', '', 'Catálogos')
    return {"ok": True, "destinatarios": len(contactos), "emails": len(mails), "whatsapps": len(celulares)}

# ── Servir catálogo por token (público) ───────────────────────────────────────
@app.get("/catalogo/{token}")
def serve_catalogo(token: str):
    db = _admin_db()
    row = db.execute(
        "SELECT filename, nombre FROM catalogos WHERE token=? AND activo=1", (token,)
    ).fetchone()
    db.close()
    if not row:
        raise HTTPException(404, "Catálogo no encontrado")
    filepath = os.path.join(CATALOGOS_DIR, row['filename'])
    if not os.path.exists(filepath):
        raise HTTPException(404, "Archivo no disponible")
    ext  = os.path.splitext(row['filename'])[1].lower()
    mime = 'application/pdf' if ext == '.pdf' else 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    nombre_archivo = f"{row['nombre']}{ext}"
    return FileResponse(filepath, media_type=mime, headers={"Content-Disposition": f'inline; filename="{nombre_archivo}"'})

# ── Catálogos disponibles para un vendedor (frontend) ────────────────────────
@app.get("/vendedor/catalogos")
def vendedor_catalogos(current_user=Depends(get_current_user)):
    cod = current_user.get('sub', '').upper()
    db  = _admin_db()
    # Muestra: catálogos asignados al vendedor O catálogos sin restricción (sin entradas en catalogo_vendedores = "para todos")
    rows = db.execute("""
        SELECT DISTINCT c.nombre, c.token, c.descripcion, c.fecha
        FROM catalogos c
        WHERE c.activo = 1
          AND (
            EXISTS (SELECT 1 FROM catalogo_vendedores cv WHERE cv.catalogo_id = c.id AND UPPER(cv.codigo) = ?)
            OR NOT EXISTS (SELECT 1 FROM catalogo_vendedores cv2 WHERE cv2.catalogo_id = c.id)
          )
        ORDER BY c.fecha DESC LIMIT 20
    """, (cod,)).fetchall()
    db.close()
    return [dict(r) for r in rows]

@app.get("/admin/debug/catalogos")
def debug_catalogos():
    """Endpoint de diagnóstico: muestra estado real de la tabla catalogos."""
    db = _admin_db()
    cols = [r[1] for r in db.execute("PRAGMA table_info(catalogos)").fetchall()]
    cats = db.execute("SELECT id,nombre,activo,fecha,subido_por FROM catalogos ORDER BY fecha DESC LIMIT 10").fetchall()
    cv   = db.execute("SELECT catalogo_id, GROUP_CONCAT(codigo) as vends FROM catalogo_vendedores GROUP BY catalogo_id").fetchall()
    db.close()
    return {
        "columnas": cols,
        "total": len(cats),
        "catalogos": [dict(r) for r in cats],
        "asignaciones": [dict(r) for r in cv]
    }

# ─── Admin: catálogos de Firebird para selects ───────────────────────────────
@app.get("/admin/gruposuperrubros")
def admin_get_gsr(_u=Depends(get_admin_user)):
    return _get_filtros_combos()['gsr']

@app.get("/admin/superrubros")
def admin_get_sr(grupo: Optional[str] = None, _u=Depends(get_admin_user)):
    if not grupo:
        return _get_filtros_combos()['sr']
    # Filtrado en cascada (usuario ya seleccionó un GSR): consulta puntual en vivo
    c = conn('WIN1252'); cur = c.cursor()
    cur.execute("""
        SELECT DISTINCT sr.CODIGOSUPERRUBRO, sr.DESCRIPCION
        FROM "SUPERRUBROS" sr
        JOIN "RUBROS" r ON r.CODIGOSUPERRUBRO = sr.CODIGOSUPERRUBRO
        JOIN "ARTICULOS" a ON a.CODIGORUBRO = r.CODIGORUBRO
        WHERE sr.CODIGOGRUPOSUPERRUBRO = ? AND a.ACTIVO = '1'
        ORDER BY sr.DESCRIPCION
    """, (grupo,))
    rows = cur.fetchall(); c.close()
    return [{"codigo": str(r[0] or '').strip(), "descripcion": str(r[1] or '').strip()} for r in rows]

@app.get("/admin/rubros")
def admin_get_rubros_admin(superrubro: Optional[str] = None, _u=Depends(get_admin_user)):
    if not superrubro:
        return _get_filtros_combos()['rubro']
    # Filtrado en cascada (usuario ya seleccionó un SR): consulta puntual en vivo
    c = conn('WIN1252'); cur = c.cursor()
    cur.execute("""
        SELECT DISTINCT r.CODIGORUBRO, r.DESCRIPCION
        FROM "RUBROS" r
        JOIN "ARTICULOS" a ON a.CODIGORUBRO = r.CODIGORUBRO
        WHERE r.CODIGOSUPERRUBRO = ? AND a.ACTIVO = '1'
        ORDER BY r.DESCRIPCION
    """, (superrubro,))
    rows = cur.fetchall(); c.close()
    return [{"codigo": str(r[0] or '').strip(), "descripcion": str(r[1] or '').strip()} for r in rows]

@app.get("/admin/marcas")
def admin_get_marcas_list(_u=Depends(get_admin_user)):
    return _get_filtros_combos()['marca']

class _TestEmailReq(BaseModel):
    destinatario: str

class _TestWAReq(BaseModel):
    celular: str  # con código de país, ej: 5491112345678

@app.post("/admin/test-email")
def admin_test_email(req: _TestEmailReq, _u=Depends(get_admin_user)):
    """Envía un email de prueba y retorna diagnóstico detallado."""
    resultado = {
        "config": {
            "SMTP_HOST": SMTP_HOST or "(vacío)",
            "SMTP_PORT": SMTP_PORT,
            "SMTP_USER": SMTP_USER or "(vacío)",
            "SMTP_FROM": SMTP_FROM or "(vacío)",
            "destinatario": req.destinatario,
        },
        "ok": False,
        "error": None,
        "etapa": None,
    }
    if not SMTP_HOST:
        resultado["error"] = "SMTP_HOST no configurado en .env"
        return resultado
    if not SMTP_USER or not SMTP_PASS:
        resultado["error"] = "SMTP_USER o SMTP_PASS no configurados en .env"
        return resultado
    try:
        resultado["etapa"] = "construyendo_mensaje"
        msg = MIMEMultipart('alternative')
        msg['Subject'] = 'Test diagnóstico — API Microbell'
        msg['From']    = SMTP_FROM
        msg['To']      = req.destinatario
        msg.attach(MIMEText('<p>Email de prueba desde API Microbell. Si ves esto, SMTP funciona correctamente.</p>', 'html', 'utf-8'))
        resultado["etapa"] = "conectando_smtp"
        if SMTP_PORT == 465:
            ctx = smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=10)
        else:
            ctx = smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=10)
        with ctx as s:
            if SMTP_PORT != 465:
                resultado["etapa"] = "starttls"
                s.ehlo(); s.starttls(); s.ehlo()
            resultado["etapa"] = "login"
            s.login(SMTP_USER, SMTP_PASS)
            resultado["etapa"] = "enviando"
            s.sendmail(SMTP_FROM, [req.destinatario], msg.as_bytes())
        resultado["ok"] = True
        resultado["etapa"] = "enviado"
    except Exception as e:
        resultado["error"] = f"{type(e).__name__}: {e}"
    return resultado

@app.post("/admin/test-whatsapp")
def admin_test_whatsapp(req: _TestWAReq, _u=Depends(get_admin_user)):
    """Envía un WhatsApp de prueba via Meta Cloud API y retorna diagnóstico detallado."""
    cel = _normalizar_celular_ar(req.celular)
    resultado = {
        "config": {
            "WA_PHONE_NUMBER_ID": WA_PHONE_NUMBER_ID or "(vacío)",
            "WA_ACCESS_TOKEN": (WA_ACCESS_TOKEN[:8] + "***") if WA_ACCESS_TOKEN else "(vacío)",
            "celular_original": req.celular,
            "celular_enviado": cel,
        },
        "ok": False,
        "error": None,
        "response_status": None,
        "response_body": None,
    }
    if not WA_PHONE_NUMBER_ID or not WA_ACCESS_TOKEN:
        resultado["error"] = "WA_PHONE_NUMBER_ID o WA_ACCESS_TOKEN no configurados en .env"
        return resultado
    if not cel:
        resultado["error"] = "Número de celular vacío o inválido"
        return resultado
    try:
        api_url = f"https://graph.facebook.com/v20.0/{WA_PHONE_NUMBER_ID}/messages"
        payload = json.dumps({
            "messaging_product": "whatsapp",
            "to": cel,
            "type": "template",
            "template": {
                "name": WA_TEMPLATE_CAT,
                "language": {"code": "es_AR"},
                "components": [
                    {
                        "type": "body",
                        "parameters": [
                            {"type": "text", "text": "Test Catálogo"}
                        ]
                    },
                    {
                        "type": "button",
                        "sub_type": "url",
                        "index": "0",
                        "parameters": [
                            {"type": "text", "text": "test123"}
                        ]
                    }
                ]
            }
        }).encode()
        req_http = urllib.request.Request(
            api_url, data=payload,
            headers={'Authorization': f'Bearer {WA_ACCESS_TOKEN}', 'Content-Type': 'application/json'},
            method='POST'
        )
        with urllib.request.urlopen(req_http, timeout=15) as resp:
            resultado["response_status"] = resp.status
            resultado["response_body"]   = resp.read().decode('utf-8', errors='replace')
        resultado["ok"] = True
    except urllib.error.HTTPError as e:
        resultado["error"]           = f"HTTPError {e.code}: {e.reason}"
        resultado["response_status"] = e.code
        try:
            resultado["response_body"] = e.read().decode('utf-8', errors='replace')
        except Exception:
            pass
    except Exception as e:
        resultado["error"] = f"{type(e).__name__}: {e}"
    return resultado

@app.post("/admin/wa/crear-plantilla")
def admin_wa_crear_plantilla(_u=Depends(get_admin_user)):
    """Crea (o actualiza) la plantilla 'microbell_catalogo' en Meta Business.
    Requiere WA_WABA_ID y WA_ACCESS_TOKEN configurados en .env"""
    if not WA_WABA_ID or not WA_ACCESS_TOKEN:
        raise HTTPException(400, "WA_WABA_ID y WA_ACCESS_TOKEN deben estar configurados en .env")
    api_url = f"https://graph.facebook.com/v20.0/{WA_WABA_ID}/message_templates"
    payload = json.dumps({
        "name": WA_TEMPLATE_CAT,
        "language": "es_AR",
        "category": "MARKETING",
        "components": [
            {
                "type": "HEADER",
                "format": "TEXT",
                "text": "Nuevo catalogo Microbell"
            },
            {
                "type": "BODY",
                "text": "Se publico el catalogo {{1}}. Entra a la app Vendedores Microbell S.A. para verlo.",
                "example": {"body_text": [["Verano 2026"]]}
            },
            {
                "type": "FOOTER",
                "text": "Microbell S.A. - Sistema de Vendedores"
            },
            {
                "type": "BUTTONS",
                "buttons": [
                    {"type": "URL", "text": "Ver catalogo", "url": "https://vendedores.microbellsa.com.ar/catalogo/{{1}}", "example": ["abc123"]}
                ]
            }
        ]
    }).encode()
    try:
        req_http = urllib.request.Request(
            api_url, data=payload,
            headers={'Authorization': f'Bearer {WA_ACCESS_TOKEN}', 'Content-Type': 'application/json'},
            method='POST'
        )
        with urllib.request.urlopen(req_http, timeout=15) as resp:
            body = resp.read().decode('utf-8', errors='replace')
        return {"ok": True, "response": json.loads(body)}
    except urllib.error.HTTPError as e:
        err = ''
        try: err = e.read().decode('utf-8', errors='replace')
        except Exception: pass
        raise HTTPException(e.code, f"Meta API error: {err}")
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/admin/wa/crear-plantilla-cobranzas")
def admin_wa_crear_plantilla_cobranzas(_u=Depends(get_admin_user)):
    """Crea (o actualiza) la plantilla 'microbell_cobranzas_v1' en Meta Business.
    Se usa para avisar al Área de Cobranzas cuando un corredor carga un pedido/presupuesto
    a un cliente con observaciones en el ABM. Categoría UTILITY (notificación transaccional,
    no marketing) — requiere aprobación de Meta antes de poder enviarse.
    Requiere WA_WABA_ID y WA_ACCESS_TOKEN configurados en .env"""
    if not WA_WABA_ID or not WA_ACCESS_TOKEN:
        raise HTTPException(400, "WA_WABA_ID y WA_ACCESS_TOKEN deben estar configurados en .env")
    api_url = f"https://graph.facebook.com/v20.0/{WA_WABA_ID}/message_templates"
    payload = json.dumps({
        "name": WA_TEMPLATE_COBRANZAS,
        "language": "es_AR",
        "category": "UTILITY",
        "components": [
            {
                "type": "HEADER",
                "format": "TEXT",
                "text": "Aviso: cliente con observación"
            },
            {
                "type": "BODY",
                "text": "Se cargó un {{1}} Nro: {{2}}\nCorredor: {{3}}\nCliente: {{4}}\n\nObservación del cliente:\n{{5}}\n\nPendiente de confirmación por Cobranzas Microbell.",
                "example": {"body_text": [["Pedido", "100024114", "ALVAREZ FERNANDO", "SUCESION DE GOMEZ GREGORIO OSVALDO (9267)", "NOSIS 27-09 CHEQ SIN FONDO - no cumple con los pagos"]]}
            },
            {
                "type": "FOOTER",
                "text": "Microbell S.A. - Sistema de Vendedores"
            }
        ]
    }).encode()
    try:
        req_http = urllib.request.Request(
            api_url, data=payload,
            headers={'Authorization': f'Bearer {WA_ACCESS_TOKEN}', 'Content-Type': 'application/json'},
            method='POST'
        )
        with urllib.request.urlopen(req_http, timeout=15) as resp:
            body = resp.read().decode('utf-8', errors='replace')
        return {"ok": True, "response": json.loads(body)}
    except urllib.error.HTTPError as e:
        err = ''
        try: err = e.read().decode('utf-8', errors='replace')
        except Exception: pass
        raise HTTPException(e.code, f"Meta API error: {err}")
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/admin/wa/crear-plantilla-reactivacion")
def admin_wa_crear_plantilla_reactivacion(_u=Depends(get_admin_user)):
    """Crea (o actualiza) la plantilla 'microbell_reactivacion_v1' en Meta Business.
    Avisa al corredor que se detectaron clientes importantes sin compras recientes —
    el detalle completo va por mail (PDF adjunto), acá solo se lo alerta a revisar el
    correo. Categoría UTILITY — requiere aprobación de Meta antes de poder enviarse.
    Requiere WA_WABA_ID y WA_ACCESS_TOKEN configurados en .env"""
    if not WA_WABA_ID or not WA_ACCESS_TOKEN:
        raise HTTPException(400, "WA_WABA_ID y WA_ACCESS_TOKEN deben estar configurados en .env")
    api_url = f"https://graph.facebook.com/v20.0/{WA_WABA_ID}/message_templates"
    payload = json.dumps({
        "name": WA_TEMPLATE_REACTIVACION,
        "language": "es_AR",
        "category": "UTILITY",
        "components": [
            {
                "type": "HEADER",
                "format": "TEXT",
                "text": "Reactivación de clientes"
            },
            {
                "type": "BODY",
                "text": "Hola {{1}}, detectamos {{2}} cliente(s) tuyo(s) con alta facturación histórica sin compras recientes. Te enviamos el detalle por mail, con ofertas vigentes para ofrecerles.",
                "example": {"body_text": [["Juan", "3"]]}
            },
            {
                "type": "FOOTER",
                "text": "Microbell S.A. - Sistema de Vendedores"
            }
        ]
    }).encode()
    try:
        req_http = urllib.request.Request(
            api_url, data=payload,
            headers={'Authorization': f'Bearer {WA_ACCESS_TOKEN}', 'Content-Type': 'application/json'},
            method='POST'
        )
        with urllib.request.urlopen(req_http, timeout=15) as resp:
            body = resp.read().decode('utf-8', errors='replace')
        return {"ok": True, "response": json.loads(body)}
    except urllib.error.HTTPError as e:
        err = ''
        try: err = e.read().decode('utf-8', errors='replace')
        except Exception: pass
        raise HTTPException(e.code, f"Meta API error: {err}")
    except Exception as e:
        raise HTTPException(500, str(e))


def _send_whatsapp_reactivacion(celular: str, vendedor_nombre: str, cantidad: int) -> tuple:
    """Envía WA vía Meta Cloud API con la plantilla de reactivación. No lleva PDF adjunto
    (WhatsApp Cloud API no permite adjuntar documentos arbitrarios en templates de texto) —
    el detalle completo va por mail. Retorna (ok: bool, error: str) — error vacío si ok,
    con el detalle devuelto por Meta si falló (para poder diagnosticar sin consola del servidor)."""
    if not WA_PHONE_NUMBER_ID or not WA_ACCESS_TOKEN:
        return False, "WA_PHONE_NUMBER_ID/WA_ACCESS_TOKEN no configurados"
    cel = _normalizar_celular_ar(celular)
    if not cel:
        return False, f"celular inválido: {celular!r}"
    api_url = f"https://graph.facebook.com/v20.0/{WA_PHONE_NUMBER_ID}/messages"
    payload = json.dumps({
        "messaging_product": "whatsapp",
        "to": cel,
        "type": "template",
        "template": {
            "name": WA_TEMPLATE_REACTIVACION,
            "language": {"code": "es_AR"},
            "components": [{
                "type": "body",
                "parameters": [
                    {"type": "text", "text": vendedor_nombre or ''},
                    {"type": "text", "text": str(cantidad)}
                ]
            }]
        }
    }).encode()
    try:
        req_http = urllib.request.Request(
            api_url, data=payload,
            headers={'Authorization': f'Bearer {WA_ACCESS_TOKEN}', 'Content-Type': 'application/json'},
            method='POST'
        )
        with urllib.request.urlopen(req_http, timeout=15) as resp:
            resp.read()
        return True, ""
    except urllib.error.HTTPError as e:
        try: err_body = e.read().decode('utf-8', errors='replace')
        except Exception: err_body = str(e)
        print(f"[REACTIVACION WA] HTTPError {e.code} → {cel}: {err_body}")
        return False, f"HTTP {e.code}: {err_body[:300]}"
    except Exception as e:
        print(f"[REACTIVACION WA] error → {cel}: {e}")
        return False, str(e)


@app.post("/admin/cambiar-password")
def admin_cambiar_password(req: _CambiarPassReq, _u=Depends(get_admin_user)):
    try:
        c = conn('WIN1252'); cur = c.cursor()
        cur.execute(
            'SELECT CODIGOUSUARIO FROM "USUARIOS" '
            'WHERE UPPER(CODIGOUSUARIO)=? AND UPPER(PASSWORD1)=? AND ACTIVO=?',
            (req.usuario.upper(), req.password_actual.upper(), '1')
        )
        if not cur.fetchone():
            c.close(); raise HTTPException(401, "Contraseña actual incorrecta")
        cur.execute(
            'UPDATE "USUARIOS" SET PASSWORD1=? WHERE UPPER(CODIGOUSUARIO)=?',
            (req.nueva_password, req.usuario.upper())
        )
        c.commit(); c.close()
    except HTTPException: raise
    except Exception as e: raise HTTPException(500, str(e))
    return {"ok": True}

# ─── Admin: stock con múltiples depósitos ─────────────────────────────────────
# ─── Admin: Reservas de Stock ─────────────────────────────────────────────────
def _migrate_stock_reservas():
    """Agrega columnas nuevas si la tabla ya existía sin ellas."""
    c = _admin_db()
    cols = [row[1] for row in c.execute("PRAGMA table_info(stock_reservas)").fetchall()]
    if 'deposito' not in cols:
        c.execute("ALTER TABLE stock_reservas ADD COLUMN deposito TEXT DEFAULT ''")
    if 'cantidad_utilizada' not in cols:
        c.execute("ALTER TABLE stock_reservas ADD COLUMN cantidad_utilizada REAL DEFAULT 0")
    if 'es_preventa' not in cols:
        c.execute("ALTER TABLE stock_reservas ADD COLUMN es_preventa INTEGER DEFAULT 0")
    if 'codigo_multiplazo' not in cols:
        c.execute("ALTER TABLE stock_reservas ADD COLUMN codigo_multiplazo TEXT")
    c.commit()
    c.close()

try:
    _migrate_stock_reservas()
except Exception:
    pass

def _sanitizar_buscar(buscar: str) -> str:
    """Limpia el parámetro buscar que puede venir del autocomplete con formato
    'CODE — DESCRIPTION'. Extrae solo el primer token antes del separador y
    limita la longitud para evitar el error Firebird -303 (string truncation)."""
    if not buscar:
        return buscar
    for sep in [' — ', ' — ', ' - ']:   # em dash, guión largo, guión
        if sep in buscar:
            buscar = buscar.split(sep)[0].strip()
            break
    return buscar[:80]

def _purgar_reservas_vencidas():
    """Marca como inactivas las reservas cuya fecha_hasta ya pasó."""
    c = _admin_db()
    today = datetime.now().strftime('%Y-%m-%d')
    c.execute(
        "UPDATE stock_reservas SET activo=0 WHERE activo=1 AND fecha_hasta IS NOT NULL AND fecha_hasta < ?",
        (today,)
    )
    c.commit()
    c.close()

def _job_purgar_reservas():
    """Hilo de fondo: purga reservas vencidas al inicio y luego cada hora."""
    _purgar_reservas_vencidas()          # ejecución inmediata al arrancar
    while True:
        time.sleep(3600)                 # espera 1 hora
        try:
            _purgar_reservas_vencidas()
        except Exception:
            pass                         # nunca tumbar el hilo por error puntual

_t = threading.Thread(target=_job_purgar_reservas, daemon=True)
_t.start()

def _get_reservas_activas():
    """Purga vencidas y retorna reservas activas vigentes."""
    _purgar_reservas_vencidas()
    c = _admin_db()
    today = datetime.now().strftime('%Y-%m-%d')
    rows = c.execute(
        "SELECT * FROM stock_reservas WHERE activo=1 AND (fecha_hasta IS NULL OR fecha_hasta >= ?)",
        (today,)
    ).fetchall()
    c.close()
    return [dict(r) for r in rows]

def _apply_reservas(resultado, reservas, rem_key='remanente_total'):
    """Agrega 'reservado' y 'reservado_por_deposito' a cada item,
    y descuenta directamente los campos remanente_XXX por depósito."""
    for item in resultado:
        reservado = 0.0
        reservado_por_dep: dict = {}
        rem = float(item.get(rem_key) or 0)
        for rv in reservas:
            applies = False
            if rv['tipo'] == 'articulo':
                # Comparar por codigo_articulo (interno) Y por codigo_particular
                # para cubrir el caso donde uno u otro fue almacenado en la reserva
                rv_art  = str(rv.get('codigo_articulo')  or '').strip()
                rv_part = str(rv.get('codigo_particular') or '').strip()
                it_cod  = str(item.get('codigo')           or '').strip()
                it_part = str(item.get('codigoparticular') or '').strip()
                applies = bool(
                    (rv_art  and rv_art  == it_cod)  or
                    (rv_part and rv_part == it_part) or
                    (rv_art  and rv_art  == it_part) or
                    (rv_part and rv_part == it_cod)
                )
            elif rv['tipo'] == 'grupo':
                tg = rv.get('tipo_grupo', '')
                vg = (rv.get('valor_grupo') or '').strip().upper()
                if tg == 'gruposuperrubro':
                    applies = (item.get('codigo_gruposuperrubro') or '').strip().upper() == vg
                elif tg == 'superrubro':
                    applies = (item.get('codigo_superrubro') or '').strip().upper() == vg
                elif tg == 'rubro':
                    applies = (item.get('codigo_rubro') or '').strip().upper() == vg
                elif tg == 'marca':
                    applies = (item.get('marca') or '').strip().upper() == vg
            if applies:
                cant_total     = float(rv.get('cantidad') or 0)
                cant_utilizada = float(rv.get('cantidad_utilizada') or 0)
                cant_neta      = cant_total - cant_utilizada
                # Reserva agotada: ignorar
                if cant_neta <= 0:
                    continue
                es_pct = rv.get('tipo_cantidad') == 'porcentaje'
                # ── Columna "Res." y validaciones → saldo real (cant_neta)
                amount_display = (rem * cant_neta / 100.0) if es_pct else cant_neta
                reservado += amount_display
                # ── Resta del remanente visible → cantidad TOTAL reservada
                # Los utilizados están comprometidos en presupuestos pero Firebird
                # aún no los descontó: el remanente NO debe subir al consumir la reserva.
                amount_stock = (rem * cant_total / 100.0) if es_pct else cant_total
                dep = (rv.get('deposito') or '').strip()
                if dep:
                    reservado_por_dep[dep] = reservado_por_dep.get(dep, 0.0) + amount_stock
        item['reservado'] = round(reservado)
        item['reservado_por_deposito'] = {k: round(v) for k, v in reservado_por_dep.items()}
        # Descontar directamente en los campos remanente por depósito (mínimo 0).
        # field se arma dinámicamente (antes era un diccionario fijo de 6 depósitos
        # que dejaba sin descontar la reserva en cualquier depósito nuevo, ej. 017
        # SARANDI — mismo bug de fondo que en /stock, /stock/batch y /stock/{codigo}).
        for dep, amount in reservado_por_dep.items():
            field = f"remanente_{dep.strip().zfill(3)}"
            if field in item:
                item[field] = max(0.0, float(item[field] or 0) - amount)
    return resultado

class ReservaStock(BaseModel):
    tipo: str
    codigo_articulo: Optional[str] = None
    codigo_particular: Optional[str] = None
    descripcion_articulo: Optional[str] = None
    tipo_grupo: Optional[str] = None
    valor_grupo: Optional[str] = None
    nombre_grupo: Optional[str] = None
    tipo_cantidad: str = 'unidades'
    cantidad: float = 0
    deposito: str = ''
    motivo: str = ''
    fecha_hasta: Optional[str] = None
    activo: int = 1
    es_preventa: bool = False
    codigo_multiplazo: Optional[str] = None

class ConsumoReserva(BaseModel):
    cantidad: float
    pedido_id: str = ''

@app.get("/admin/reservas-stock")
def get_reservas_stock(_u=Depends(get_admin_user)):
    c = _admin_db()
    rows = c.execute("SELECT * FROM stock_reservas ORDER BY creado_at DESC").fetchall()
    c.close()
    return [dict(r) for r in rows]

@app.get("/reservas-activas")
def get_reservas_activas_frontend(_u=Depends(get_current_user)):
    """Devuelve reservas activas y vigentes. Accesible con token de vendedor o impersonación."""
    return _get_reservas_activas()

@app.post("/admin/reservas-stock")
def create_reserva_stock(body: ReservaStock, _u=Depends(get_admin_user)):
    # Validar stock disponible en depósito si es reserva por artículo (no aplica para preventa)
    if not body.es_preventa and body.tipo == 'articulo' and body.codigo_articulo and body.deposito and body.tipo_cantidad == 'unidades':
        try:
            fb = conn('WIN1252')
            cur = fb.cursor()
            cur.execute(f'SELECT STOCKREMANENTE FROM "FMA_STOCK"(NULL, NULL, \'{body.deposito}\', 1, 1) WHERE ID_ARTICULO=?',
                        (body.codigo_articulo,))
            row = cur.fetchone()
            fb.close()
            rem_dep = float(row[0] or 0) if row else 0.0
            # Sumar reservas activas existentes para ese artículo+depósito
            c2 = _admin_db()
            today = datetime.now().strftime('%Y-%m-%d')
            rows_act = c2.execute(
                """SELECT cantidad, cantidad_utilizada FROM stock_reservas
                   WHERE activo=1 AND tipo='articulo' AND codigo_articulo=? AND deposito=?
                   AND (fecha_hasta IS NULL OR fecha_hasta >= ?)""",
                (body.codigo_articulo, body.deposito, today)
            ).fetchall()
            c2.close()
            ya_reservado = sum(max(0, float(r[0] or 0) - float(r[1] or 0)) for r in rows_act)
            disponible = rem_dep - ya_reservado
            if body.cantidad > disponible:
                raise HTTPException(400, f"Stock insuficiente en depósito {body.deposito}: remanente={round(rem_dep)}, ya reservado={round(ya_reservado)}, disponible={round(disponible)}, solicitado={round(body.cantidad)}")
        except HTTPException:
            raise
        except Exception as e:
            pass  # Si Firebird no responde, no bloquear creación

    c = _admin_db()
    c.execute(
        """INSERT INTO stock_reservas (tipo, codigo_articulo, codigo_particular, descripcion_articulo,
           tipo_grupo, valor_grupo, nombre_grupo, tipo_cantidad, cantidad, deposito, cantidad_utilizada,
           motivo, fecha_hasta, creado_por, activo, es_preventa, codigo_multiplazo)
           VALUES (?,?,?,?,?,?,?,?,?,?,0,?,?,?,1,?,?)""",
        (body.tipo, body.codigo_articulo, body.codigo_particular, body.descripcion_articulo,
         body.tipo_grupo, body.valor_grupo, body.nombre_grupo,
         body.tipo_cantidad, body.cantidad, body.deposito,
         body.motivo, body.fecha_hasta, _u['sub'], int(body.es_preventa), body.codigo_multiplazo)
    )
    c.commit()
    new_id = c.execute("SELECT last_insert_rowid()").fetchone()[0]
    c.close()
    _audit(_u['sub'], 'Nueva reserva stock', f"tipo={body.tipo} dep={body.deposito} motivo={body.motivo}")
    return {"id": new_id}

@app.put("/admin/reservas-stock/{rid}")
def update_reserva_stock(rid: int, body: ReservaStock, _u=Depends(get_admin_user)):
    c = _admin_db()
    c.execute(
        """UPDATE stock_reservas SET tipo=?, codigo_articulo=?, codigo_particular=?, descripcion_articulo=?,
           tipo_grupo=?, valor_grupo=?, nombre_grupo=?, tipo_cantidad=?, cantidad=?, deposito=?,
           motivo=?, fecha_hasta=?, activo=?, es_preventa=?, codigo_multiplazo=?
           WHERE id=?""",
        (body.tipo, body.codigo_articulo, body.codigo_particular, body.descripcion_articulo,
         body.tipo_grupo, body.valor_grupo, body.nombre_grupo,
         body.tipo_cantidad, body.cantidad, body.deposito,
         body.motivo, body.fecha_hasta, body.activo, int(body.es_preventa), body.codigo_multiplazo, rid)
    )
    c.commit()
    c.close()
    _audit(_u['sub'], 'Actualizar reserva stock', f"id={rid}")
    return {"ok": True}

@app.patch("/admin/reservas-stock/{rid}")
def patch_reserva_stock(rid: int, body: dict, _u=Depends(get_admin_user)):
    """Actualización parcial: solo los campos enviados en el body."""
    allowed = {'cantidad','activo','motivo','deposito','fecha_hasta','tipo_cantidad',
               'codigo_articulo','codigo_particular','descripcion_articulo','es_preventa','codigo_multiplazo'}
    fields = {k: v for k, v in body.items() if k in allowed}
    if not fields:
        raise HTTPException(400, "No se enviaron campos válidos para actualizar")
    set_sql = ', '.join(f"{k}=?" for k in fields)
    vals = list(fields.values()) + [rid]
    c = _admin_db()
    c.execute(f"UPDATE stock_reservas SET {set_sql} WHERE id=?", vals)
    c.commit()
    c.close()
    _audit(_u['sub'], 'Patch reserva stock', f"id={rid} campos={list(fields)}")
    return {"ok": True}

@app.post("/admin/reservas-stock/{rid}/consumir")
def consumir_reserva(rid: int, body: ConsumoReserva, _u=Depends(get_current_user)):
    """Registra uso de cantidad de una reserva (desde pedido/presupuesto)."""
    c = _admin_db()
    row = c.execute("SELECT * FROM stock_reservas WHERE id=? AND activo=1", (rid,)).fetchone()
    if not row:
        c.close()
        raise HTTPException(404, "Reserva no encontrada o inactiva")
    r = dict(row)
    cant_neta = float(r['cantidad'] or 0) - float(r['cantidad_utilizada'] or 0)
    if body.cantidad > cant_neta + 0.001:
        c.close()
        raise HTTPException(400, f"Cantidad solicitada ({body.cantidad}) supera disponible en reserva ({round(cant_neta)})")
    nueva_utilizada = float(r['cantidad_utilizada'] or 0) + body.cantidad
    c.execute("UPDATE stock_reservas SET cantidad_utilizada=? WHERE id=?", (nueva_utilizada, rid))
    c.commit()
    c.close()
    _audit(_u['sub'], 'Consumo de reserva stock', f"reserva_id={rid} cantidad={body.cantidad} pedido={body.pedido_id}")
    return {"ok": True, "cantidad_utilizada": nueva_utilizada, "restante": float(r['cantidad'] or 0) - nueva_utilizada}

@app.get("/debug/reserva/{codigo}")
def debug_reserva(codigo: str):
    """Endpoint de diagnóstico — muestra reservas raw de admin.db para un código. Sin auth."""
    c = _admin_db()
    rows = c.execute(
        "SELECT id, codigo_articulo, codigo_particular, deposito, cantidad, cantidad_utilizada, es_preventa, codigo_multiplazo, activo, motivo, fecha_hasta FROM stock_reservas WHERE codigo_articulo=? OR codigo_particular=? ORDER BY id DESC",
        (codigo, codigo)
    ).fetchall()
    c.close()
    return [dict(r) for r in rows]

@app.get("/reservas-activas-articulo/{codigo}")
def get_reservas_activas_articulo(codigo: str, _u=Depends(get_current_user)):
    """Devuelve reservas activas con remanente disponible para un artículo (para usar en pedido)."""
    _purgar_reservas_vencidas()
    c = _admin_db()
    today = datetime.now().strftime('%Y-%m-%d')
    rows = c.execute(
        """SELECT * FROM stock_reservas
           WHERE activo=1 AND tipo='articulo'
           AND (codigo_articulo=? OR codigo_particular=?)
           AND (fecha_hasta IS NULL OR fecha_hasta >= ?)
           ORDER BY CASE WHEN fecha_hasta IS NULL THEN 1 ELSE 0 END, fecha_hasta ASC""",
        (codigo, codigo, today)
    ).fetchall()
    c.close()
    reservas = [dict(row) for row in rows]

    # Consultar remanente para reservas tipo articulo con es_preventa.
    # Reutiliza la caché FMA_STOCK por depósito (ya con TTL/pre-warm) en vez de
    # correr FMA_STOCK(NULL,NULL,NULL,1,1) sin caché en cada llamada.
    art_preventa = [r for r in reservas if r.get('es_preventa') and r.get('codigo_articulo')]
    rem_map = {}
    if art_preventa:
        try:
            art_set = {str(r['codigo_articulo']) for r in art_preventa}
            fma_por_dep = _fma_stock_parallel(_deps_activos())
            for dep_data in fma_por_dep.values():
                for art_id, stock in dep_data.items():
                    sid = str(art_id)
                    if sid in art_set:
                        rem_map[sid] = rem_map.get(sid, 0.0) + stock
        except Exception:
            pass

    result = []
    for r in reservas:
        restante = float(r['cantidad'] or 0) - float(r['cantidad_utilizada'] or 0)
        if restante > 0:
            r['restante'] = round(restante)
            if r.get('es_preventa') and r.get('codigo_articulo'):
                r['remanente_firebird'] = rem_map.get(str(r['codigo_articulo']))
            else:
                r['remanente_firebird'] = None
            result.append(r)
    return result

@app.delete("/admin/reservas-stock/{rid}")
def delete_reserva_stock(rid: int, _u=Depends(get_admin_user)):
    c = _admin_db()
    c.execute("DELETE FROM stock_reservas WHERE id=?", (rid,))
    c.commit()
    c.close()
    _audit(_u['sub'], 'Eliminar reserva stock', f"id={rid}")
    return {"ok": True}

@app.get("/admin/reservas-stock/con-stock")
def get_reservas_con_stock(_u=Depends(get_admin_user)):
    """Devuelve reservas enriquecidas con remanente actual de Firebird."""
    c = _admin_db()
    rows = c.execute("SELECT * FROM stock_reservas ORDER BY creado_at DESC").fetchall()
    c.close()
    reservas = [dict(r) for r in rows]

    # Para reservas por artículo, consultar remanente en Firebird
    art_codes = [r['codigo_articulo'] for r in reservas if r['tipo'] == 'articulo' and r.get('codigo_articulo')]
    rem_map = {}
    if art_codes:
        try:
            fb = conn('WIN1252')
            cur = fb.cursor()
            cur.execute("SELECT ID_ARTICULO, STOCKREMANENTE FROM \"FMA_STOCK\"(NULL, NULL, NULL, 1, 1)")
            for row in cur.fetchall():
                if str(row[0]) in [str(c) for c in art_codes]:
                    rem_map[str(row[0])] = float(row[1] or 0)
            fb.close()
        except Exception:
            pass

    for r in reservas:
        if r['tipo'] == 'articulo' and r.get('codigo_articulo'):
            rem = rem_map.get(str(r['codigo_articulo']), 0)
            r['remanente_firebird'] = rem
            if r['tipo_cantidad'] == 'porcentaje':
                r['reservado_unidades'] = round(rem * float(r['cantidad'] or 0) / 100)
            else:
                r['reservado_unidades'] = round(float(r['cantidad'] or 0))
            r['disponible'] = max(0, round(rem - r['reservado_unidades']))
        else:
            r['remanente_firebird'] = None
            r['reservado_unidades'] = None
            r['disponible'] = None
    return reservas

@app.get("/admin/reservas-stock/exportar-pdf")
def exportar_reservas_pdf(token: Optional[str] = None, request: Request = None):
    # Acepta token por header (Bearer) o por query param (?token=...) para window.open
    _u = None
    auth_header = request.headers.get('Authorization', '') if request else ''
    raw_token = token or (auth_header.replace('Bearer ', '') if auth_header.startswith('Bearer ') else None)
    if not raw_token:
        raise HTTPException(401, "No autenticado")
    try:
        _u = jwt.decode(raw_token, JWT_SECRET, algorithms=[JWT_ALGO])
        if _u.get('role') != 'admin':
            raise HTTPException(403, "Acceso denegado")
    except JWTError:
        raise HTTPException(401, "Token inválido o expirado")
    """PDF de todas las reservas de stock activas con impacto en remanente."""
    import io
    from reportlab.lib.pagesizes import landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT

    # Obtener datos enriquecidos
    resp = get_reservas_con_stock(_u=_u)
    reservas = resp if isinstance(resp, list) else []
    today = datetime.now().strftime('%Y-%m-%d')

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12*mm, rightMargin=12*mm,
                            topMargin=14*mm, bottomMargin=12*mm)

    styles = getSampleStyleSheet()
    s_title = ParagraphStyle('t', parent=styles['Heading1'], fontSize=13, textColor=colors.HexColor('#1a56db'), spaceAfter=4)
    s_sub   = ParagraphStyle('s', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#6b7280'), spaceAfter=6)
    s_hdr   = ParagraphStyle('h', parent=styles['Normal'], fontSize=7, textColor=colors.white, alignment=TA_CENTER, fontName='Helvetica-Bold', leading=9)
    s_cell  = ParagraphStyle('c', parent=styles['Normal'], fontSize=7, leading=9)
    s_cell_c= ParagraphStyle('cc', parent=styles['Normal'], fontSize=7, leading=9, alignment=TA_CENTER)
    s_cell_r= ParagraphStyle('cr', parent=styles['Normal'], fontSize=7, leading=9, alignment=TA_RIGHT)
    s_green = ParagraphStyle('g', parent=styles['Normal'], fontSize=7, leading=9, alignment=TA_RIGHT, textColor=colors.HexColor('#059669'))
    s_red   = ParagraphStyle('r', parent=styles['Normal'], fontSize=7, leading=9, alignment=TA_RIGHT, textColor=colors.HexColor('#dc2626'))
    s_orange= ParagraphStyle('o', parent=styles['Normal'], fontSize=7, leading=9, alignment=TA_RIGHT, textColor=colors.HexColor('#c2410c'))

    fmt_n = lambda v: f"{v:,.0f}".replace(',', '.') if v is not None else '—'
    col_names = ["Tipo", "Alcance / Artículo", "Cant. Reservada", "Rem. Firebird", "Disponible", "Motivo", "Vence", "Estado"]
    table_data = [[Paragraph(h, s_hdr) for h in col_names]]

    for r in reservas:
        activo = bool(r.get('activo')) and (not r.get('fecha_hasta') or r['fecha_hasta'] >= today)
        tipo_label = 'Artículo' if r['tipo'] == 'articulo' else {
            'gruposuperrubro': 'G.S.Rubro', 'superrubro': 'S.Rubro',
            'rubro': 'Rubro', 'marca': 'Marca'
        }.get(r.get('tipo_grupo', ''), r.get('tipo_grupo', 'Grupo'))

        if r['tipo'] == 'articulo':
            alcance = f"{r.get('codigo_particular','')} — {r.get('descripcion_articulo','')}"
        else:
            alcance = f"{r.get('nombre_grupo') or r.get('valor_grupo', '')}"

        cant_str = f"{r['cantidad']:.0f}%" if r.get('tipo_cantidad') == 'porcentaje' else f"{r['cantidad']:,.0f} u."
        rem_str = fmt_n(r.get('remanente_firebird'))
        disp = r.get('disponible')
        disp_style = s_green if (disp or 0) > 0 else s_red
        disp_str = fmt_n(disp)
        vence = r.get('fecha_hasta') or 'Sin venc.'
        estado = 'Activa' if activo else 'Vencida'

        table_data.append([
            Paragraph(tipo_label, s_cell_c),
            Paragraph(alcance[:60], s_cell),
            Paragraph(cant_str, s_orange),
            Paragraph(rem_str, s_cell_r),
            Paragraph(disp_str, disp_style),
            Paragraph((r.get('motivo') or '')[:40], s_cell),
            Paragraph(vence, s_cell_c),
            Paragraph(estado, s_cell_c),
        ])

    col_widths = [22*mm, 80*mm, 28*mm, 28*mm, 28*mm, 60*mm, 24*mm, 18*mm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a56db')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#fff7ed')]),
        ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#e5e7eb')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
    ]))

    fecha_gen = datetime.now().strftime('%d/%m/%Y %H:%M')
    story = [
        Paragraph("📌 Reservas de Stock — Microbell S.A.", s_title),
        Paragraph(f"Generado el {fecha_gen} por {_u.get('sub','')}", s_sub),
        Spacer(1, 4*mm),
        t
    ]
    doc.build(story)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": "inline; filename=reservas_stock.pdf"})

@app.get("/admin/stock")
def admin_get_stock(
    buscar: Optional[str] = None,
    depositos: Optional[str] = None,
    gruposuperrubro: Optional[str] = None,
    superrubro: Optional[str] = None,
    rubro: Optional[str] = None,
    marca: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
    _u=Depends(get_admin_user)
):
    dep_lista = [d.strip() for d in (depositos or '001,003').split(',') if d.strip()]
    # Usa la misma caché TTL (90s) que frontend.html — prioriza velocidad sobre exactitud al segundo
    try:
        pagina, total_count, cambio_usd = _search_stock_cache(
            buscar=buscar, gruposuperrubro=gruposuperrubro, superrubro=superrubro,
            rubro=rubro, marca=marca, dep_lista=dep_lista, limit=limit, offset=offset
        )
        resultado = []
        for art, rem_dep, rem_total in pagina:
            factor = cambio_usd if art['codigomoneda'] == 'DOLARES' else 1.0
            precio = _redondear_precio(art['precio1'], factor)
            item = {
                "codigo":               art['codigo'],
                "codigoparticular":     art['codigoparticular'],
                "descripcion":          art['descripcion'],
                "marca":                art['codigomarca'],
                "precio1":              precio,
                "iva":                  art['iva'],
                "unidad":               art['unidad'],
                "stock_total":          rem_total,
                "remanente_total":      rem_total,
                "rubro":                art['rubro'],
                "superrubro":           art['superrubro'],
                "gruposuperrubro":      art['gruposuperrubro'],
                "codigo_rubro":         art['codigo_rubro'],
                "codigo_superrubro":    art['codigo_superrubro'],
                "codigo_gruposuperrubro": art['codigo_gruposuperrubro'],
            }
            for dep in dep_lista:
                item[f"rem_{dep}"] = rem_dep.get(dep, 0)
            resultado.append(item)
        _apply_reservas(resultado, _get_reservas_activas())
        resp = JSONResponse(content=resultado)
        resp.headers['X-Total-Count'] = str(total_count)
        resp.headers['Access-Control-Expose-Headers'] = 'X-Total-Count'
        return resp
    except Exception as e:
        raise HTTPException(500, str(e))

# ─── Admin: exportar stock ───────────────────────────────────────────────────
def _dep_nombres(dep_lista: list) -> dict:
    """Devuelve {codigo: nombre} consultando Firebird para los depósitos dados."""
    if not dep_lista:
        return {}
    try:
        placeholders = ','.join('?' * len(dep_lista))
        _c = conn('WIN1252'); _cur = _c.cursor()
        _cur.execute(
            f'SELECT TRIM(CODIGODEPOSITO), TRIM(DESCRIPCION) FROM "DEPOSITOS" WHERE CODIGODEPOSITO IN ({placeholders})',
            dep_lista
        )
        mapa = {str(r[0]).strip(): str(r[1] or r[0]).strip() for r in _cur.fetchall()}
        _c.close()
        return mapa
    except Exception:
        return {d: d for d in dep_lista}  # fallback: usar el código

def _admin_stock_data(buscar=None, depositos=None, gruposuperrubro=None,
                      superrubro=None, rubro=None, marca=None):
    """Devuelve todos los artículos activos según filtros (sin paginación) para export.
    Usa catálogo + FMA por depósito individual (igual que Rotación) porque FMA_STOCK
    no acepta CSV de depósitos — retorna resultados incorrectos/vacíos con varios deps.
    """
    dep_lista = [d.strip() for d in (depositos or '001,002,003,005,013,016').split(',') if d.strip()]

    catalog, cambio_usd = _get_catalog()
    fma_data = _fma_stock_parallel(dep_lista)  # {dep: {art_id: stock}}

    buscar_norm = _sanitizar_buscar(buscar).upper() if buscar else None

    resultado = []
    for art_id, art in catalog.items():
        # Filtros de texto y jerarquía
        if buscar_norm:
            if buscar_norm not in art['descripcion'].upper() and buscar_norm not in art['codigoparticular'].upper():
                continue
        if rubro           and art.get('codigo_rubro')            != rubro:           continue
        if superrubro      and art.get('codigo_superrubro')       != superrubro:      continue
        if gruposuperrubro and art.get('codigo_gruposuperrubro')  != gruposuperrubro: continue
        if marca           and art.get('codigomarca')             != marca:           continue

        rem_dep = {dep: fma_data.get(dep, {}).get(art_id, 0) for dep in dep_lista}
        rem_total = sum(rem_dep.values())
        if rem_total <= 0:
            continue

        factor = cambio_usd if art.get('codigomoneda', '').upper() == 'DOLARES' else 1.0
        precio = _redondear_precio(art.get('precio1', 0), factor)

        item = {
            "codigo":        art.get('codigoparticular', ''),
            "descripcion":   art.get('descripcion', ''),
            "precio1":       precio,
            "iva":           art.get('iva', 0),
            "marca":         art.get('codigomarca', ''),
            "rubro":         art.get('rubro', ''),
            "superrubro":    art.get('superrubro', ''),
            "gruposuperrubro": art.get('gruposuperrubro', ''),
        }
        for dep in dep_lista:
            item[f"rem_{dep}"] = rem_dep[dep]
        item["rem_total"] = rem_total
        resultado.append(item)

    resultado.sort(key=lambda x: x['codigo'])
    return resultado, dep_lista


@app.get("/admin/stock/exportar-excel")
def admin_exportar_stock_excel(
    buscar: Optional[str] = None,
    depositos: Optional[str] = None,
    gruposuperrubro: Optional[str] = None,
    superrubro: Optional[str] = None,
    rubro: Optional[str] = None,
    marca: Optional[str] = None,
    _u=Depends(get_admin_download_auth)
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        rows, dep_lista = _admin_stock_data(buscar, depositos, gruposuperrubro, superrubro, rubro, marca)

        DEP_LABELS = _dep_nombres(dep_lista)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock Admin"

        hdr_fill = PatternFill("solid", fgColor="1A56DB")
        hdr_font = Font(bold=True, color="FFFFFFFF", size=10)
        alt_fill = PatternFill("solid", fgColor="FFEFF6FF")
        right_al = Alignment(horizontal="right", vertical="center")
        center   = Alignment(horizontal="center", vertical="center", wrap_text=True)

        headers = ["Gr.SR", "Super Rubro", "Rubro", "Código", "Descripción"] + [DEP_LABELS.get(d, d) for d in dep_lista] + ["R.Total", "P.Unit.", "IVA%"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center
        ws.row_dimensions[1].height = 28

        for ri, row in enumerate(rows, 2):
            ws.cell(ri, 1, row.get("gruposuperrubro", ""))
            ws.cell(ri, 2, row.get("superrubro", ""))
            ws.cell(ri, 3, row.get("rubro", ""))
            ws.cell(ri, 4, row["codigo"])
            ws.cell(ri, 5, row["descripcion"])
            for di, dep in enumerate(dep_lista, 6):
                c = ws.cell(ri, di, round(row[f"rem_{dep}"], 2))
                c.alignment = right_al
                v = row[f"rem_{dep}"]
                if v < 0: c.font = Font(color="DC2626")
                elif v > 0: c.font = Font(color="059669")
            col_rt = 6 + len(dep_lista)
            ws.cell(ri, col_rt, round(row["rem_total"], 2)).alignment = right_al
            ws.cell(ri, col_rt+1, row["precio1"]).alignment = right_al
            _c_iva = ws.cell(ri, col_rt+2, row["iva"])
            _c_iva.alignment = right_al
            _c_iva.number_format = '0.00'
            if ri % 2 == 0:
                for ci2 in range(1, len(headers)+1):
                    ws.cell(ri, ci2).fill = alt_fill

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 45
        for ci2 in range(6, len(headers)+1):
            ws.column_dimensions[get_column_letter(ci2)].width = 13

        import io
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        from fastapi.responses import StreamingResponse
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=stock_admin.xlsx"})
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/admin/stock/exportar-pdf")
def admin_exportar_stock_pdf(
    buscar: Optional[str] = None,
    depositos: Optional[str] = None,
    gruposuperrubro: Optional[str] = None,
    superrubro: Optional[str] = None,
    rubro: Optional[str] = None,
    marca: Optional[str] = None,
    _u=Depends(get_admin_download_auth)
):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_RIGHT, TA_CENTER

        rows, dep_lista = _admin_stock_data(buscar, depositos, gruposuperrubro, superrubro, rubro, marca)
        DEP_LABELS = _dep_nombres(dep_lista)

        import io
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=10*mm, rightMargin=10*mm,
                                topMargin=12*mm, bottomMargin=10*mm)

        styles = getSampleStyleSheet()
        s_title = ParagraphStyle('t', parent=styles['Heading1'], fontSize=12, textColor=colors.HexColor('#1a56db'))
        s_hdr   = ParagraphStyle('h', parent=styles['Normal'], fontSize=6, textColor=colors.white, alignment=TA_CENTER, fontName='Helvetica-Bold', leading=8)
        s_cell  = ParagraphStyle('c', parent=styles['Normal'], fontSize=6, leading=8)
        s_cell_r= ParagraphStyle('cr', parent=styles['Normal'], fontSize=6, alignment=TA_RIGHT, leading=8)
        s_neg   = ParagraphStyle('neg', parent=styles['Normal'], fontSize=6, alignment=TA_RIGHT, textColor=colors.HexColor('#dc2626'), leading=8)
        s_pos   = ParagraphStyle('pos', parent=styles['Normal'], fontSize=6, alignment=TA_RIGHT, textColor=colors.HexColor('#059669'), leading=8)

        dep_headers = [DEP_LABELS.get(d, d) for d in dep_lista]
        col_names = ["Gr.SR", "Super Rubro", "Rubro", "Código", "Descripción"] + dep_headers + ["R.Total", "P.Unit.", "IVA%"]

        table_data = [[Paragraph(h, s_hdr) for h in col_names]]
        fmt_n = lambda v: f"{v:,.0f}".replace(',','.') if v else '—'
        fmt_p = lambda v: f"${v:,.2f}".replace(',','#').replace('.',',').replace('#','.')

        for row in rows:
            r_row = [
                Paragraph(row.get("gruposuperrubro", ""), s_cell),
                Paragraph(row.get("superrubro", ""), s_cell),
                Paragraph(row.get("rubro", ""), s_cell),
                Paragraph(row["codigo"], s_cell),
                Paragraph(row["descripcion"], s_cell),
            ]
            for dep in dep_lista:
                v = row[f"rem_{dep}"]
                style = s_neg if v < 0 else s_pos if v > 0 else s_cell_r
                r_row.append(Paragraph(fmt_n(v) if v != 0 else '—', style))
            r_row.append(Paragraph(fmt_n(row["rem_total"]), s_cell_r))
            r_row.append(Paragraph(fmt_p(row["precio1"]), s_cell_r))
            r_row.append(Paragraph((f"{row['iva']:.2f}".replace('.', ',')+'%') if row['iva'] else '—', s_cell_r))
            table_data.append(r_row)

        n_dep = len(dep_lista)
        # Ancho disponible = A4 landscape - margenes
        page_w = landscape(A4)[0] - 20*mm
        # GSR/SR/Rub mas anchos, Desc acotada
        fixed_w = (22 + 28 + 20 + 14)*mm + n_dep*16*mm + (16 + 18 + 11)*mm
        desc_w = min(50*mm, max(30*mm, page_w - fixed_w))
        col_widths = [22*mm, 28*mm, 20*mm, 14*mm, desc_w] + [16*mm]*n_dep + [16*mm, 18*mm, 11*mm]
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a56db')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#eff6ff')]),
            ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#e5e7eb')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ]))

        story = [Paragraph("Stock por Depósito — Microbell", s_title), Spacer(1, 4*mm), t]
        doc.build(story)
        buf.seek(0)
        from fastapi.responses import StreamingResponse
        return StreamingResponse(buf, media_type="application/pdf",
                                 headers={"Content-Disposition": "inline; filename=stock_admin.pdf"})
    except Exception as e:
        raise HTTPException(500, str(e))


# ─── Admin: motor de ofertas ──────────────────────────────────────────────────
def _resolver_codigoparticular_clientes(offers: list):
    """Resuelve codigoparticular (el código que reconoce el equipo comercial) para cada
    cliente de offer_clients, SOLO para mostrar en el panel — codigocliente (el interno
    de Firebird) no se toca, porque es lo que frontend.html manda como ?cliente= al
    matchear ofertas."""
    todos_cod = sorted({cl['codigocliente'] for o in offers for cl in o.get('clients', []) if cl.get('codigocliente')})
    if not todos_cod:
        return
    try:
        fb = conn('WIN1252', db='c:/flexxus/DB/DB-Microbell.gdb')
        cur = fb.cursor()
        ph2 = ','.join('?' * len(todos_cod))
        cur.execute(f"""SELECT CODIGOCLIENTE, TRIM(CODIGOPARTICULAR) FROM "CLIENTES"
                        WHERE CODIGOCLIENTE IN ({ph2})""", todos_cod)
        part_map = {str(r[0]).strip(): (r[1] or '').strip() for r in cur.fetchall()}
        fb.close()
        for o in offers:
            for cl in o.get('clients', []):
                cl['codigoparticular'] = part_map.get(cl['codigocliente']) or cl['codigocliente']
    except Exception as e:
        print(f"[OFERTAS] error resolviendo codigoparticular para mostrar: {e}")


def _load_offer_relations(c, o):
    oid = o['id']
    o['product_details']    = [dict(r) for r in c.execute("SELECT * FROM offer_product_details WHERE offer_id=?", (oid,)).fetchall()]
    o['financial_details']  = [dict(r) for r in c.execute("SELECT * FROM offer_financial_details WHERE offer_id=? ORDER BY orden", (oid,)).fetchall()]
    o['conditions']         = [r[0] for r in c.execute("SELECT condicion_comercial FROM offer_conditions WHERE offer_id=?", (oid,)).fetchall()]
    o['vendors']            = [r[0] for r in c.execute("SELECT codigousuario FROM offer_vendors WHERE offer_id=?", (oid,)).fetchall()]
    o['profiles']           = [r[0] for r in c.execute("SELECT perfil_codigo FROM offer_profiles WHERE offer_id=?", (oid,)).fetchall()]
    o['category_filters']   = [dict(r) for r in c.execute("SELECT nivel, valor FROM offer_category_filters WHERE offer_id=?", (oid,)).fetchall()]
    o['clients']            = [dict(r) for r in c.execute("SELECT codigocliente, razonsocial, descuento_extra_pct, vencimiento_extra, monto_minimo_extra, tipo_cartera, condicion_comercial_extra FROM offer_clients WHERE offer_id=?", (oid,)).fetchall()]
    o['discount_filters']   = [dict(r) for r in c.execute("SELECT nivel, valor FROM offer_discount_filters WHERE offer_id=?", (oid,)).fetchall()]
    o['combo_escalones']    = [dict(r) for r in c.execute("SELECT min_combos, descuento_pct FROM offer_combo_escalones WHERE offer_id=? ORDER BY min_combos", (oid,)).fetchall()]
    o['amount_escalones']   = [dict(r) for r in c.execute("SELECT monto_minimo, descuento_pct, condicion_comercial FROM offer_amount_escalones WHERE offer_id=? ORDER BY monto_minimo", (oid,)).fetchall()]
    for k,d in [('deposito',''),('tipo_financiero','descuento_total'),('monto_minimo',0),('cupo',0),('usos',0),('tope_bonificacion_pesos',0),('bonificado_acumulado_pesos',0),('acumulable',0)]:
        if k not in o: o[k] = d
    o['financial_escalones'] = json.loads(o['financial_escalones']) if o.get('financial_escalones') else []
    _resolver_codigoparticular_clientes([o])
    return o

def _fmt_entero(v):
    """NUMEROCOMPROBANTE en Firebird viaja como NUMERIC/DOUBLE y el driver lo
    entrega como float (7938.0). Los números de comprobante son siempre
    enteros, así que se muestran sin el '.0'."""
    if v is None:
        return ''
    s = str(v).strip()
    try:
        f = float(s)
        return str(int(f)) if f == int(f) else s
    except (ValueError, TypeError):
        return s


def _fmt_cant(v):
    """Cantidades e IVA%: enteros sin '.0' cuando no tienen parte decimal."""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return v
    return int(f) if f == int(f) else round(f, 2)


def _admin_pendientes_query(c, tabla_cab, tabla_cuerpo, tipo_label, es_pedido,
                             fecha_desde, fecha_hasta, grupo, superrubro, rubro, marca, articulo, vendedor=None):
    """Devuelve líneas de Pedidos o Presupuestos con saldo pendiente de remitir
    (CANTIDAD - CANTIDADREMITIDA > 0), cruzadas con ARTICULOS/RUBROS para poder
    filtrar por categoría."""
    wheres = ["cb.ANULADA = '0'",
              '(cu.CANTIDAD - COALESCE(cu.CANTIDADREMITIDA, 0)) > 0',
              'cb.FECHACOMPROBANTE >= ?', 'cb.FECHACOMPROBANTE < ?']
    if es_pedido:
        wheres.insert(0, "cb.TIPOCOMPROBANTE = 'NP'")
    else:
        # Solo presupuestos aprobados: los que no se aprobaron nunca no están
        # "pendientes de cumplimentar", son borradores.
        wheres.append('cb.FECHAAPROBADO IS NOT NULL')
        wheres.append("cb.FECHAAPROBADO > CAST('1900-01-02 00:00:00' AS TIMESTAMP)")
    if vendedor:
        wheres.append('UPPER(cb.CODIGOUSUARIO) = ?')
    if grupo:
        wheres.append('g.CODIGOGRUPOSUPERRUBRO = ?')
    if superrubro:
        wheres.append('sr.CODIGOSUPERRUBRO = ?')
    if rubro:
        wheres.append('a.CODIGORUBRO = ?')
    if marca:
        wheres.append('a.CODIGOMARCA = ?')
    if articulo:
        wheres.append("(UPPER(a.CODIGOPARTICULAR) CONTAINING UPPER(?) OR UPPER(cu.DESCRIPCION) CONTAINING UPPER(?))")

    # params deben ir en el mismo orden que aparecen los "?" en el WHERE final
    params = [fecha_desde, fecha_hasta]
    if vendedor:   params.append(vendedor.strip().upper())
    if grupo:      params.append(grupo)
    if superrubro: params.append(superrubro)
    if rubro:      params.append(rubro)
    if marca:      params.append(marca)
    if articulo:   params += [articulo, articulo]

    join_tipo = " AND cu.TIPOCOMPROBANTE = cb.TIPOCOMPROBANTE" if es_pedido else ""
    sql = f"""
        SELECT cb.NUMEROCOMPROBANTE, cb.FECHACOMPROBANTE, cb.CODIGOCLIENTE, cb.RAZONSOCIAL,
               cu.LINEA, cu.CODIGOARTICULO,
               COALESCE(NULLIF(TRIM(cu.CODIGOPARTICULAR),''), NULLIF(TRIM(a.CODIGOPARTICULAR),''), TRIM(cu.CODIGOARTICULO)),
               cu.DESCRIPCION, cu.CANTIDAD, COALESCE(cu.CANTIDADREMITIDA,0),
               cu.PRECIOUNITARIO, cu.PORCENTAJEIVA,
               a.CODIGOMARCA, m.DESCRIPCION, a.CODIGORUBRO, r.DESCRIPCION,
               sr.CODIGOSUPERRUBRO, sr.DESCRIPCION, g.DESCRIPCION,
               cb.CODIGOUSUARIO, u.RAZONSOCIAL,
               a.COEFICIENTESEGUNRUBRO, a.COEFICIENTE, r.COEFICIENTE
        FROM "{tabla_cab}" cb
        JOIN "{tabla_cuerpo}" cu ON cu.NUMEROCOMPROBANTE = cb.NUMEROCOMPROBANTE{join_tipo}
        LEFT JOIN "ARTICULOS" a ON a.CODIGOARTICULO = cu.CODIGOARTICULO
        LEFT JOIN "MARCAS" m ON m.CODIGOMARCA = a.CODIGOMARCA
        LEFT JOIN "RUBROS" r ON r.CODIGORUBRO = a.CODIGORUBRO
        LEFT JOIN "SUPERRUBROS" sr ON sr.CODIGOSUPERRUBRO = r.CODIGOSUPERRUBRO
        LEFT JOIN "GRUPOSUPERRUBROS" g ON g.CODIGOGRUPOSUPERRUBRO = sr.CODIGOGRUPOSUPERRUBRO
        LEFT JOIN "USUARIOS" u ON u.CODIGOUSUARIO = cb.CODIGOUSUARIO
        WHERE {' AND '.join(wheres)}
        ORDER BY cb.FECHACOMPROBANTE DESC
    """
    cur = c.cursor()
    cur.execute(sql, params)
    rows = []
    for r in cur.fetchall():
        cant_total = float(r[8] or 0)
        cant_rem   = float(r[9] or 0)
        # IVA real (mismo criterio que el resto de la app): si COEFICIENTESEGUNRUBRO=1
        # usar el coeficiente del RUBRO, si=0 (Manual) usar el propio del artículo.
        # cu.PORCENTAJEIVA queda desactualizado en comprobantes viejos y no se usa.
        _coef_manual = float(r[22] or 0)
        _coef_rubro  = float(r[23] or 0)
        _usa_rubro   = str(r[21] or '').strip() == '1'
        _coef_final  = _coef_rubro if _usa_rubro else _coef_manual
        iva_pct      = round(_coef_final * 21, 2)
        rows.append({
            "tipo":              tipo_label,
            "numero":            _fmt_entero(r[0]),
            "fecha":             str(r[1])[:10] if r[1] else '',
            "codigocliente":     str(r[2] or '').strip(),
            "razonsocial":       (r[3] or '').strip(),
            "linea":             r[4],
            "codigo":            r[5],
            "codigoparticular":  (r[6] or r[5] or '').strip(),
            "descripcion":       (r[7] or '').strip(),
            "cantidad_total":    cant_total,
            "cantidad_remitida": cant_rem,
            "cantidad_pendiente": round(cant_total - cant_rem, 3),
            "precio_unitario":   float(r[10] or 0),
            "iva":               iva_pct,
            "codigomarca":       (r[12] or '').strip(),
            "marca":             (r[13] or '').strip(),
            "codigo_rubro":      (r[14] or '').strip(),
            "rubro":             (r[15] or '').strip(),
            "codigo_superrubro": (r[16] or '').strip(),
            "superrubro":        (r[17] or '').strip(),
            "gruposuperrubro":   (r[18] or '').strip(),
            "codigo_vendedor":   (r[19] or '').strip(),
            "vendedor":          (r[20] or r[19] or '').strip(),
        })
    return rows


@app.get("/admin/pendientes-cumplimentar")
def admin_pendientes_cumplimentar(
    fecha_desde: str = Query(...),
    fecha_hasta: str = Query(...),
    grupo: str = None,
    superrubro: str = None,
    rubro: str = None,
    marca: str = None,
    articulo: str = None,
    vendedor: str = None,
    _u=Depends(get_admin_user)
):
    """Artículos de Pedidos y Presupuestos (aprobados) con saldo pendiente de
    remitir, dentro del rango de FECHACOMPROBANTE indicado. Sin filtros de
    categoría/vendedor seleccionados, trae todos los pendientes del rango."""
    # FECHACOMPROBANTE < hasta+1día para incluir el día completo de "hasta"
    hasta_excl = (datetime.strptime(fecha_hasta, '%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%d')
    c = conn('WIN1252')
    try:
        rows_pedidos = _admin_pendientes_query(
            c, 'CABEZAPEDIDOS', 'CUERPOPEDIDOS', 'Pedido', True,
            fecha_desde, hasta_excl, grupo, superrubro, rubro, marca, articulo, vendedor
        )
        rows_presu = _admin_pendientes_query(
            c, 'CABEZAPRESUPUESTOS', 'CUERPOPRESUPUESTOS', 'Presupuesto', False,
            fecha_desde, hasta_excl, grupo, superrubro, rubro, marca, articulo, vendedor
        )
    finally:
        c.close()
    result = rows_pedidos + rows_presu
    # Más antiguo primero: lo más urgente de cumplimentar va en la página 1.
    result.sort(key=lambda x: x['fecha'])
    return result


_PENDIENTES_HEADERS = ["Tipo", "Número", "Fecha", "Vendedor", "Cliente", "Código", "Descripción",
                        "Cant. Total", "Cant. Remitida", "Cant. Pendiente",
                        "Precio Unit.", "IVA%", "Marca", "Rubro"]

def _pendientes_row_vals(r):
    return [r['tipo'], r['numero'], r['fecha'], r['vendedor'], r['razonsocial'], r['codigoparticular'], r['descripcion'],
            _fmt_cant(r['cantidad_total']), _fmt_cant(r['cantidad_remitida']), _fmt_cant(r['cantidad_pendiente']),
            round(r['precio_unitario'], 2), _fmt_cant(r['iva']), r['marca'], r['rubro']]


@app.get("/admin/pendientes-cumplimentar/exportar-excel")
def admin_pendientes_exportar_excel(
    fecha_desde: str = Query(...),
    fecha_hasta: str = Query(...),
    grupo: str = None,
    superrubro: str = None,
    rubro: str = None,
    marca: str = None,
    articulo: str = None,
    vendedor: str = None,
    _u=Depends(get_admin_download_auth)
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import io

        rows = admin_pendientes_cumplimentar(
            fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, grupo=grupo,
            superrubro=superrubro, rubro=rubro, marca=marca, articulo=articulo,
            vendedor=vendedor, _u=_u
        )
        if not rows:
            raise HTTPException(404, "Sin artículos pendientes para los filtros indicados")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pendientes"

        hdr_fill = PatternFill("solid", fgColor="1A56DB")
        hdr_font = Font(bold=True, color="FFFFFFFF", size=10)
        alt_fill = PatternFill("solid", fgColor="EFF6FF")
        title_font = Font(bold=True, size=13, color="111827")
        right_al = Alignment(horizontal="right", vertical="center")
        left_al  = Alignment(horizontal="left", vertical="center", wrap_text=True)
        center_al = Alignment(horizontal="center", vertical="center")
        NUM_COLS = len(_PENDIENTES_HEADERS)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
        fecha_gen = datetime.now().strftime('%d/%m/%Y %H:%M')
        c1 = ws.cell(1, 1, f"MICROBELL S.A. — Artículos Pendientes de Cumplimentar  ({fecha_desde} a {fecha_hasta})  ·  Generado: {fecha_gen}")
        c1.font = title_font; c1.alignment = left_al
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 6

        for ci, h in enumerate(_PENDIENTES_HEADERS, 1):
            cell = ws.cell(3, ci, h)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center_al
        ws.row_dimensions[3].height = 20

        for ri, r in enumerate(rows, 4):
            fill = alt_fill if ri % 2 == 0 else None
            vals = _pendientes_row_vals(r)
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(ri, ci, v)
                if fill: cell.fill = fill
                if ci in (8, 9, 10, 11, 12): cell.alignment = right_al
                elif ci in (4, 5, 7): cell.alignment = left_al

        widths = [11, 9, 11, 14, 26, 11, 36, 11, 13, 13, 12, 8, 16, 16]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64+i) if i <= 26 else 'A'].width = w

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        from fastapi.responses import StreamingResponse
        fname = f"pendientes_{fecha_desde}_a_{fecha_hasta}.xlsx"
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f"attachment; filename={fname}"})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/admin/pendientes-cumplimentar/exportar-pdf")
def admin_pendientes_exportar_pdf(
    fecha_desde: str = Query(...),
    fecha_hasta: str = Query(...),
    grupo: str = None,
    superrubro: str = None,
    rubro: str = None,
    marca: str = None,
    articulo: str = None,
    vendedor: str = None,
    _u=Depends(get_admin_download_auth)
):
    try:
        import io
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_LEFT
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        rows = admin_pendientes_cumplimentar(
            fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, grupo=grupo,
            superrubro=superrubro, rubro=rubro, marca=marca, articulo=articulo,
            vendedor=vendedor, _u=_u
        )
        if not rows:
            raise HTTPException(404, "Sin artículos pendientes para los filtros indicados")

        fecha_gen = datetime.now().strftime('%d/%m/%Y %H:%M')
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=6*mm, rightMargin=6*mm,
                                topMargin=12*mm, bottomMargin=12*mm)
        styles = getSampleStyleSheet()
        st_title = ParagraphStyle('pend_title', fontName='Helvetica-Bold', fontSize=13,
                                  textColor=colors.HexColor('#111827'), spaceAfter=6)
        st_desc  = ParagraphStyle('pend_desc', fontName='Helvetica', fontSize=7.5,
                                  textColor=colors.HexColor('#111827'), leading=9, wordWrap='LTR')
        st_cli   = ParagraphStyle('pend_cli', fontName='Helvetica', fontSize=7.5,
                                  textColor=colors.HexColor('#111827'), leading=9, wordWrap='LTR')
        st_txt   = ParagraphStyle('pend_txt', fontName='Helvetica', fontSize=7.5,
                                  textColor=colors.HexColor('#111827'), leading=9, wordWrap='LTR')
        st_head  = ParagraphStyle('pend_head', fontName='Helvetica-Bold', fontSize=7.5,
                                  textColor=colors.white, leading=9, wordWrap='LTR')

        story = [Paragraph(
            f"MICROBELL S.A. — Artículos Pendientes de Cumplimentar &nbsp;({fecha_desde} a {fecha_hasta}) &nbsp;·&nbsp; Generado: {fecha_gen} &nbsp;·&nbsp; {len(rows)} líneas",
            st_title
        )]

        # Encabezados abreviados (evita que el texto del header se salga de columnas angostas)
        pdf_headers = ["Tipo", "Número", "Fecha", "Vendedor", "Cliente", "Código", "Descripción",
                       "Cant.Total", "Cant.Remit.", "Cant.Pend.", "P.Unit.", "IVA%", "Marca", "Rubro"]
        data = [[Paragraph(h, st_head) for h in pdf_headers]]
        for r in rows:
            vals = _pendientes_row_vals(r)
            vals[0]  = Paragraph(str(vals[0] or ''), st_txt)   # Tipo con wrap (Presupuesto no entra en 1 línea corta)
            vals[1]  = Paragraph(str(vals[1] or ''), st_txt)   # Número con wrap
            vals[2]  = Paragraph(str(vals[2] or ''), st_txt)   # Fecha con wrap
            vals[3]  = Paragraph(str(vals[3] or ''), st_txt)   # Vendedor con wrap
            vals[4]  = Paragraph(str(vals[4] or ''), st_cli)   # Cliente con wrap (evita que invada la celda de al lado)
            vals[6]  = Paragraph(str(vals[6] or ''), st_desc)  # Descripción con wrap
            vals[12] = Paragraph(str(vals[12] or ''), st_txt)  # Marca con wrap
            vals[13] = Paragraph(str(vals[13] or ''), st_txt)  # Rubro con wrap
            data.append(vals)

        # Anchos ajustados para que Vendedor/Cliente/Descripción tengan lugar de sobra
        # y hagan wrap en vez de superponerse con la columna siguiente.
        col_widths = [16*mm, 14*mm, 16*mm, 30*mm, 39*mm, 11*mm, 66*mm, 11*mm, 12*mm, 12*mm, 13*mm, 8*mm, 17*mm, 17*mm]
        tbl = Table(data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1A56DB')),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,-1), 7.5),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#EFF6FF')]),
            ('GRID', (0,0), (-1,-1), 0.4, colors.lightgrey),
            ('ALIGN', (7,1), (11,-1), 'RIGHT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ]))
        story.append(tbl)
        doc.build(story)
        buf.seek(0)
        from fastapi.responses import StreamingResponse
        fname = f"pendientes_{fecha_desde}_a_{fecha_hasta}.pdf"
        return StreamingResponse(buf, media_type="application/pdf",
                                 headers={"Content-Disposition": f'inline; filename="{fname}"'})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/admin/rotacion-filtros")
def admin_rotacion_filtros(_u=Depends(get_admin_user)):
    """Devuelve GrupoSuperRubros, SuperRubros, Rubros y Marcas para los filtros de Rotación."""
    try:
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()
        cur.execute('SELECT TRIM(CODIGOGRUPOSUPERRUBRO), TRIM(DESCRIPCION) FROM "GRUPOSUPERRUBROS" ORDER BY DESCRIPCION')
        grupos = [{'codigo': r[0], 'descripcion': r[1] or r[0]} for r in cur.fetchall()
                  if r[0] and (r[1] or '').strip().upper() not in ('TERCEROS', 'SERVICIOS')]
        cur.execute('SELECT TRIM(CODIGOSUPERRUBRO), TRIM(DESCRIPCION), TRIM(CODIGOGRUPOSUPERRUBRO) FROM "SUPERRUBROS" ORDER BY DESCRIPCION')
        superrubros = [{'codigo': r[0], 'descripcion': r[1] or r[0], 'grupo': r[2] or ''} for r in cur.fetchall() if r[0]]
        cur.execute('SELECT TRIM(CODIGORUBRO), TRIM(DESCRIPCION), TRIM(CODIGOSUPERRUBRO) FROM "RUBROS" ORDER BY DESCRIPCION')
        rubros = [{'codigo': r[0], 'descripcion': r[1] or r[0], 'superrubro': r[2] or ''} for r in cur.fetchall() if r[0]]
        try:
            cur.execute('SELECT TRIM(CODIGOMARCA), TRIM(DESCRIPCION) FROM "MARCAS" ORDER BY DESCRIPCION')
            marcas = [{'codigo': r[0], 'descripcion': r[1] or r[0]} for r in cur.fetchall() if r[0]]
        except Exception:
            marcas = []
        c.close()
        return {'grupos': grupos, 'superrubros': superrubros, 'rubros': rubros, 'marcas': marcas}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/admin/debug-multiplazos-columns")
def admin_debug_multiplazos_columns(_u=Depends(get_admin_user)):
    """Debug: lista todas las columnas de la tabla MULTIPLAZOS en Firebird."""
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        cur.execute("""
            SELECT rf.RDB$FIELD_NAME, f.RDB$FIELD_TYPE, f.RDB$FIELD_LENGTH
            FROM RDB$RELATION_FIELDS rf
            JOIN RDB$FIELDS f ON f.RDB$FIELD_NAME = rf.RDB$FIELD_SOURCE
            WHERE rf.RDB$RELATION_NAME = 'MULTIPLAZOS'
            ORDER BY rf.RDB$FIELD_POSITION
        """)
        cols = [{"nombre": r[0].strip(), "tipo": r[1], "longitud": r[2]} for r in cur.fetchall()]
        # También devolver la fila del multiplazo 36 (PREVENTA) para ver valores
        cur2 = c.cursor()
        cur2.execute('SELECT * FROM "MULTIPLAZOS" WHERE CODIGOMULTIPLAZO=?', ('36',))
        row = cur2.fetchone()
        row_dict = {cols[i]['nombre']: row[i] for i in range(len(cols))} if row else {}
        c.close()
        return {"columnas": cols, "ejemplo_multiplazo_36": row_dict}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/admin/debug-clientes-codigo")
def admin_debug_clientes_codigo(codigos: str = Query(..., description="CODIGOCLIENTE separados por coma, ej: 12440,9748"),
                                 _u=Depends(get_admin_user)):
    """Debug temporal: para un CODIGOCLIENTE interno dado, muestra lado a lado
    CODIGOCLIENTE, CODIGOPARTICULAR y RAZONSOCIAL tal cual están en Firebird —
    para diagnosticar por qué Reactivación muestra un código que no coincide
    con lo esperado en el módulo de Ofertas."""
    lista = [c.strip() for c in codigos.split(',') if c.strip()]
    if not lista:
        return []
    try:
        c = conn('WIN1252', db='c:/flexxus/DB/DB-Microbell.gdb')
        cur = c.cursor()
        ph = ','.join('?' * len(lista))
        cur.execute(f"""
            SELECT CODIGOCLIENTE, TRIM(CODIGOPARTICULAR), TRIM(RAZONSOCIAL), ACTIVO
            FROM "CLIENTES" WHERE CODIGOCLIENTE IN ({ph})
        """, lista)
        rows = cur.fetchall()
        c.close()
        return [{"codigocliente": str(r[0]).strip(), "codigoparticular": (r[1] or '').strip(),
                  "razonsocial": (r[2] or '').strip(), "activo": str(r[3])} for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/admin/debug-tipos-comprobante")
def admin_debug_tipos_comprobante(_u=Depends(get_admin_user)):
    """Debug: lista todos los TIPOCOMPROBANTE distintos en CABEZACOMPROBANTES con cantidad de registros."""
    try:
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()
        cur.execute("""
            SELECT TIPOCOMPROBANTE, COUNT(*) AS cant, MAX(FECHACOMPROBANTE) AS ultima_fecha
            FROM "CABEZACOMPROBANTES"
            WHERE ANULADA = 0
            GROUP BY TIPOCOMPROBANTE
            ORDER BY cant DESC
        """)
        rows = cur.fetchall()
        c.close()
        return [{"tipo": (r[0] or '').strip(), "cantidad": int(r[1] or 0),
                 "ultima_fecha": str(r[2])[:10] if r[2] else None} for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/admin/analisis-rotacion")
def admin_analisis_rotacion(
    meses: int = 12,
    grupo: str = None,
    superrubro: str = None,
    rubro: str = None,
    marca: str = None,
    articulo: str = None,
    depositos: str = None,   # CSV de códigos, ej "001,003" — vacío/None = todos
    pct_operativo: float = 30,   # % gasto operativo sobre Costo
    pct_utilidad: float = 25,    # % utilidad sobre (Costo + Operativo)
    pct_meli: float = 80,        # % para Pub. ML sobre el Mayorista resultante
    pct_venta_max: float = 20,   # % máx. de (vendido en el período / stock) para entrar
    compra_hasta: str = None,    # YYYY-MM-DD — solo artículos cuya última compra sea ANTES de esta fecha (o sin compra registrada)
    _u=Depends(get_admin_user)
):
    """
    Artículos con stockremanente > 0 (stock físico, siempre L1) y BAJA ROTACIÓN
    en el período: (cantidad vendida neta / stock remanente actual) × 100 <= pct_venta_max.
    Solo cuentan como "venta" las facturas/remitos (FA, FB, FE) — NO pedidos (NP)
    ni presupuestos, que no son ventas confirmadas. Las notas de crédito
    (FCA, FCB, FCE, FCCA, FCCB, FCCE) restan, son devoluciones.
    Considera ventas tanto de Línea 1 como de SW (comparten el mismo stock físico).
    Calcula precio lista1 en ARS, precio ML sugerido, costo, margen%, bulto sugerido.
    """
    fecha_corte = (datetime.now() - timedelta(days=meses * 30.44))

    try:
        # 1. Catálogo completo + cambio USD — misma fuente que "Stock por Depósito"
        #    (clave = CODIGOARTICULO interno, NO CODIGOPARTICULAR).
        catalog, cambio_usd = _get_catalog()

        # 2. Stock remanente — misma función _fma_stock_parallel que usa
        #    "Stock por Depósito" (probada y consistente). Pasar un CSV de
        #    varios depósitos directo a FMA_STOCK en una sola llamada daba
        #    resultados incorrectos; acá se llama una vez por depósito.
        dep_lista = [d.strip() for d in depositos.split(',') if d.strip()] if depositos and depositos.strip() else _deps_activos()
        fma_data = _fma_stock_parallel(dep_lista)  # {dep: {art_id: stock}}
        stock_map = {}  # art_id (CODIGOARTICULO) -> stock total
        for art_id in catalog.keys():
            total = sum(fma_data.get(dep, {}).get(art_id, 0) for dep in dep_lista)
            if total > 0:
                stock_map[art_id] = total

        if not stock_map:
            return []

        # 3. Ventas reales (facturas/remitos) por artículo en el período —
        #    combinado de L1 y SW. NO se cuentan NP (pedido) ni presupuestos:
        #    no son ventas confirmadas. Facturas (FA/FB/FE) suman; notas de
        #    crédito (FCA/FCB/FCE/FCCA/FCCB/FCCE) restan (son devoluciones).
        #    También se guarda la fecha de la última venta (solo informativo).
        # FA/FB/FE: cantidad positiva; NCA/NCB/FC*: cantidad negativa (ya firmada en DB).
        # No se invierte signo: la suma directa da la cantidad neta correcta.
        TIPOS_VENTA_TODOS = ('FA', 'FB', 'FE', 'NCA', 'NCB', 'FCA', 'FCB', 'FCE', 'FCCA', 'FCCB', 'FCCE')
        vendido_map  = {}  # CODIGOARTICULO -> cantidad vendida neta en el período
        ultimo_venta = {}  # CODIGOARTICULO -> fecha de la venta más reciente (cualquier período)

        ph_tipos = ','.join(f"'{t}'" for t in TIPOS_VENTA_TODOS)

        # L1: facturas reales en CABEZACOMPROBANTES/CUERPOCOMPROBANTES.
        # CABEZAPEDIDOS/CUERPOPEDIDOS solo contiene pedidos NP, NO facturas FA.
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()
        try:
            cur.execute(f"""
                SELECT cu.CODIGOARTICULO, cb.FECHACOMPROBANTE, cu.CANTIDAD, cb.TIPOCOMPROBANTE
                FROM "CUERPOCOMPROBANTES" cu
                JOIN "CABEZACOMPROBANTES" cb
                  ON cb.TIPOCOMPROBANTE=cu.TIPOCOMPROBANTE
                 AND cb.NUMEROCOMPROBANTE=cu.NUMEROCOMPROBANTE
                WHERE cu.TIPOCOMPROBANTE IN ({ph_tipos}) AND cb.ANULADA=0
            """)
            for r in cur.fetchall():
                cod = (r[0] or '').strip()
                if not cod or not r[1]:
                    continue
                cant = float(r[2] or 0)
                if cant == 0:
                    continue
                dt = r[1] if isinstance(r[1], datetime) else datetime.strptime(str(r[1])[:10], '%Y-%m-%d')
                if cant > 0:
                    ultimo_venta[cod] = max(ultimo_venta.get(cod, datetime.min), dt)
                if dt >= fecha_corte:
                    vendido_map[cod] = vendido_map.get(cod, 0) + cant
        except Exception:
            pass
        c.close()

        # Ventas SW — conexión aparte (DATABASE_MLT, LATIN1). Si SW no está
        # disponible, no debe romper el análisis: se sigue solo con L1.
        try:
            c_sw = conn('LATIN1', db=DATABASE_MLT)
            cur_sw = c_sw.cursor()
            try:
                cur_sw.execute(f"""
                    SELECT cu.CODIGOARTICULO, cb.FECHACOMPROBANTE, cu.CANTIDAD, cb.TIPOCOMPROBANTE
                    FROM "CUERPOCOMPROBANTES" cu
                    JOIN "CABEZACOMPROBANTES" cb
                      ON cb.TIPOCOMPROBANTE=cu.TIPOCOMPROBANTE
                     AND cb.NUMEROCOMPROBANTE=cu.NUMEROCOMPROBANTE
                    WHERE cu.TIPOCOMPROBANTE IN ({ph_tipos}) AND cb.ANULADA=0
                """)
                for r in cur_sw.fetchall():
                    cod = (r[0] or '').strip()
                    if not cod or not r[1]:
                        continue
                    cant = float(r[2] or 0)
                    if cant == 0:
                        continue
                    dt = r[1] if isinstance(r[1], datetime) else datetime.strptime(str(r[1])[:10], '%Y-%m-%d')
                    if cant > 0:
                        ultimo_venta[cod] = max(ultimo_venta.get(cod, datetime.min), dt)
                    if dt >= fecha_corte:
                        vendido_map[cod] = vendido_map.get(cod, 0) + cant
            finally:
                c_sw.close()
        except Exception:
            pass

        # 4b. Última compra (RC = Remito Compra) por artículo — L1 únicamente.
        ultima_compra = {}  # CODIGOARTICULO -> fecha más reciente de RC
        c_rc = conn('WIN1252', db=DATABASE)
        cur_rc = c_rc.cursor()
        try:
            cur_rc.execute("""
                SELECT cu.CODIGOARTICULO, MAX(cb.FECHACOMPROBANTE)
                FROM "CUERPOCOMPROBANTES" cu
                JOIN "CABEZACOMPROBANTES" cb
                  ON cb.TIPOCOMPROBANTE = cu.TIPOCOMPROBANTE
                 AND cb.NUMEROCOMPROBANTE = cu.NUMEROCOMPROBANTE
                WHERE cu.TIPOCOMPROBANTE = 'RE' AND cb.ANULADA = 0
                GROUP BY cu.CODIGOARTICULO
            """)
            for row in cur_rc.fetchall():
                cod = (row[0] or '').strip()
                if cod and row[1]:
                    dt = row[1] if isinstance(row[1], datetime) else datetime.strptime(str(row[1])[:10], '%Y-%m-%d')
                    ultima_compra[cod] = dt
        except Exception:
            pass
        finally:
            c_rc.close()

        # 4. Costo (PRECIOCOMPRA) — confirmado que solo aparece correcto
        #    buscando por CODIGOPARTICULAR (no por CODIGOARTICULO). Se indexa
        #    por codigoparticular y se cruza con catalog[art_id]['codigoparticular'].
        costo_por_particular = {}
        particulares = list({catalog[aid]['codigoparticular'] for aid in stock_map if aid in catalog})
        c2 = conn('WIN1252', db=DATABASE)
        cur2 = c2.cursor()
        for i in range(0, len(particulares), 400):
            chunk = particulares[i:i + 400]
            ph = ','.join(['?' for _ in chunk])
            try:
                cur2.execute(
                    f'SELECT CODIGOPARTICULAR, PRECIOCOMPRA FROM "ARTICULOS" WHERE CODIGOPARTICULAR IN ({ph})',
                    chunk
                )
                for row in cur2.fetchall():
                    cp = (row[0] or '').strip()
                    if cp:
                        costo_por_particular[cp] = float(row[1] or 0)
            except Exception:
                pass
        c2.close()

        # 5. Filtrar por jerarquía/marca/artículo usando el catálogo en memoria
        #    (igual criterio que _search_stock_cache) y construir resultado.
        resultado = []
        for art_id, rem in stock_map.items():
            art = catalog.get(art_id)
            if not art:
                continue
            if rubro and art['codigo_rubro'] != rubro:
                continue
            if superrubro and art['codigo_superrubro'] != superrubro:
                continue
            if grupo and art['codigo_gruposuperrubro'] != grupo:
                continue
            if marca and art['codigomarca'] != marca:
                continue
            if articulo:
                au = articulo.upper()
                if au not in (art['codigoparticular'] or '').upper() and au not in (art['descripcion'] or '').upper():
                    continue

            cod = art['codigoparticular']
            art_id_str = str(art_id).strip()
            vendido = vendido_map.get(art_id_str, 0) or vendido_map.get(cod, 0)
            vendido = max(vendido, 0)  # más notas de crédito que ventas no es "rotación negativa", es 0
            pct_venta = round((vendido / rem * 100), 1) if rem > 0 else 0.0
            # Baja rotación: solo entran los que vendieron <= pct_venta_max
            # de su stock actual en el período (0% = no vendieron nada).
            if pct_venta > pct_venta_max:
                continue
            ultimo_mov = ultimo_venta.get(art_id_str) or ultimo_venta.get(cod)

            # Filtro por fecha de última compra:
            # - Si ultima_compra existe y es >= compra_hasta → EXCLUIR (compraron recientemente)
            # - Si ultima_compra es None → INCLUIR siempre (sin registro en Firebird, no se castiga)
            uc_pre = ultima_compra.get(art_id_str) or ultima_compra.get(cod)
            ultima_compra_str = uc_pre.strftime('%Y-%m-%d') if uc_pre else None
            if compra_hasta and ultima_compra_str and ultima_compra_str >= compra_hasta:
                continue

            precio1 = art['precio1']
            moneda  = art['codigomoneda']
            precio1_ars = precio1 * cambio_usd if moneda == 'DOLARES' else precio1

            # Costo en ARS y margen sobre lista
            costo = costo_por_particular.get(cod)
            costo_ars = None
            margen_lista = None
            if costo and costo > 0:
                costo_ars = costo * cambio_usd if moneda == 'DOLARES' else costo
                if precio1_ars > 0:
                    margen_lista = round((precio1_ars - costo_ars) / costo_ars * 100, 1)

            # Precio Oferta Mayorista (Lista1 ARS = techo, nunca se supera).
            # Mercadería sin reposición futura: el costo de reposición cargado
            # en Flexxus es la única referencia de costo disponible.
            # Cálculo EN CASCADA con los % ingresados como variables en el modal:
            #   1) Punto de equilibrio = Costo × (1 + pct_operativo/100)
            #   2) Precio Mayorista objetivo = equilibrio × (1 + pct_utilidad/100)
            #      (la utilidad se aplica sobre el resultado del paso 1, no
            #      sobre el costo original)
            #   3) Pub. ML = Mayorista × (1 + pct_meli/100)
            # margen_mayorista_pct = utilidad real lograda sobre el punto de
            # equilibrio: pct_utilidad = cumple objetivo, 0% = breakeven,
            # negativo = ni cubre el gasto operativo (tope de Lista1 de por medio).
            op_factor   = 1 + (pct_operativo / 100)
            util_factor = 1 + (pct_utilidad / 100)
            meli_factor = 1 + (pct_meli / 100)
            if costo_ars and costo_ars > 0:
                punto_equilibrio = costo_ars * op_factor
                precio_objetivo  = punto_equilibrio * util_factor
                precio_mayorista = round(min(precio1_ars, precio_objetivo), 2)
                margen_mayorista = round((precio_mayorista / punto_equilibrio - 1) * 100, 1)
                precio_bulto_unitario = round(min(precio1_ars, punto_equilibrio), 2)
            else:
                precio_mayorista  = round(precio1_ars, 2)
                margen_mayorista  = None
                precio_bulto_unitario = round(precio1_ars, 2)
                punto_equilibrio = None

            # Precio competitivo en MercadoLibre = Mayorista (oferta) × meli_factor.
            # Se basa en el precio de OFERTA, no en Lista1, para que la
            # publicación en ML refleje la liquidación real.
            precio_meli_pub = round(precio_mayorista * meli_factor, 2)
            puede_pub_meli = (precio_meli_pub > punto_equilibrio) if punto_equilibrio else True

            # Sugerencia de bulto según precio unitario en ARS (artículos de bajo valor)
            if precio1_ars < 1000:
                bulto = 100
            elif precio1_ars < 5000:
                bulto = 20
            elif precio1_ars < 15000:
                bulto = 10
            else:
                bulto = None

            # No sugerir bulto si el stock disponible no alcanza para armar al menos un pack
            if bulto and rem < bulto:
                bulto = None

            # Precios de bulto (total del pack)
            precio_bulto_meli      = round(precio_meli_pub        * bulto, 2) if bulto else None
            precio_bulto_mayorista = round(precio_bulto_unitario  * bulto, 2) if bulto else None

            ultimo_str = None
            if ultimo_mov and ultimo_mov != datetime.min:
                try:
                    ultimo_str = ultimo_mov.strftime('%Y-%m-%d')
                except Exception:
                    ultimo_str = str(ultimo_mov)[:10]

            # ultima_compra_str ya calculado arriba en el filtro
            resultado.append({
                'codigo':             cod,
                'art_id':             art_id_str,
                'codigoparticular':   art['codigoparticular'] or cod,
                'descripcion':        art['descripcion'],
                'stock':              round(rem),
                'precio1':            round(precio1, 2),
                'moneda':             moneda,
                'precio1_ars':        round(precio1_ars, 2),
                'precio_meli_pub':    precio_meli_pub,
                'precio_mayorista':   precio_mayorista,
                'margen_mayorista_pct': margen_mayorista,
                'costo':              round(costo_ars, 2) if costo_ars else None,
                'margen_lista_pct':   margen_lista,
                'puede_pub_meli':     puede_pub_meli,
                'ultimo_movimiento':  ultimo_str,
                'ultima_compra':      ultima_compra_str,
                'cantidad_vendida_periodo': round(vendido, 2),
                'pct_venta':          pct_venta,
                'bulto_sugerido':     bulto,
                'precio_bulto_meli':      precio_bulto_meli,
                'precio_bulto_mayorista': precio_bulto_mayorista,
                'alicuotaiva':        art['iva'],
                'cambio_usd':         cambio_usd,
                'pct_operativo':      pct_operativo,
                'pct_utilidad':       pct_utilidad,
                'pct_meli':           pct_meli,
                'pct_venta_max':      pct_venta_max,
            })

        # Ordenar por valor de stock desc
        resultado.sort(key=lambda x: x['stock'] * x['precio1_ars'], reverse=True)
        return resultado

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


_PACK_QTYS = [10, 20, 30, 40]


def _rotacion_precio_pack(r: dict, qty: int):
    """Replica el cálculo del frontend (admin.html): precio unitario de pack
    usa precio_bulto_mayorista/bulto_sugerido si hay bulto calculado, sino
    cae a precio_mayorista. Devuelve None si no hay precio unitario."""
    bulto = r.get('bulto_sugerido')
    pbm = r.get('precio_bulto_mayorista')
    precio_unit = (pbm / bulto) if (pbm and bulto) else (r.get('precio_mayorista') or 0)
    return round(precio_unit * qty) if precio_unit > 0 else None


@app.get("/admin/rotacion/exportar-excel")
def admin_rotacion_exportar_excel(
    meses: int = 12,
    grupo: str = None,
    superrubro: str = None,
    rubro: str = None,
    marca: str = None,
    articulo: str = None,
    depositos: str = None,
    pct_operativo: float = 30,
    pct_utilidad: float = 25,
    pct_meli: float = 80,
    pct_venta_max: float = 20,
    compra_hasta: str = None,
    _u=Depends(get_admin_download_auth)
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import io

        rows = admin_analisis_rotacion(
            meses=meses, grupo=grupo, superrubro=superrubro, rubro=rubro, marca=marca,
            articulo=articulo, depositos=depositos, pct_operativo=pct_operativo,
            pct_utilidad=pct_utilidad, pct_meli=pct_meli, pct_venta_max=pct_venta_max,
            compra_hasta=compra_hasta, _u=_u
        )
        if not rows:
            raise HTTPException(404, "Sin datos de rotación para los filtros indicados")

        cambio   = rows[0].get('cambio_usd', 1)
        pct_op   = rows[0].get('pct_operativo', pct_operativo)
        pct_ut   = rows[0].get('pct_utilidad', pct_utilidad)
        pct_ml   = rows[0].get('pct_meli', pct_meli)
        pct_vtx  = rows[0].get('pct_venta_max', pct_venta_max)

        sin_ventas   = sum(1 for r in rows if not (r.get('cantidad_vendida_periodo', 0) > 0))
        total_costo  = sum(r['stock'] * (r.get('costo') or 0) for r in rows)
        sin_costo    = sum(1 for r in rows if not r.get('costo'))
        con_bulto    = sum(1 for r in rows if r.get('bulto_sugerido'))
        con_perdida  = sum(1 for r in rows if r.get('margen_mayorista_pct') is not None and r['margen_mayorista_pct'] < 0)
        sin_compra   = sum(1 for r in rows if not r.get('ultima_compra'))

        ch_fmt = '/'.join(reversed(compra_hasta.split('-'))) if compra_hasta else None
        linea_resumen = (
            f"{len(rows)} artículos de baja rotación (ventas ≤ {pct_vtx}% del stock)"
            + (f" · Compras anteriores al {ch_fmt}" if ch_fmt else "")
            + f" · {sin_ventas} sin ninguna venta en el período"
            + f" · Capital inmovilizado (costo): ${round(total_costo):,} ARS"
            + f" · Cambio USD: ${cambio:,}"
            + (f" · {con_bulto} con venta por bulto sugerida" if con_bulto else "")
            + (f" · {con_perdida} no cubren ni el punto de equilibrio" if con_perdida else "")
            + (f" · {sin_costo} sin costo cargado" if sin_costo else "")
            + (f" · {sin_compra} sin remito de compra registrado" if sin_compra else "")
        )
        linea_leyenda = (
            f"Baja rotación = ventas NV en el período ÷ stock actual ≤ {pct_vtx}%"
            f"  |  Punto de equilibrio = Costo × (1+{pct_op}%)"
            f"  |  Mayorista = equilibrio × (1+{pct_ut}%), sin superar Lista1"
            f"  |  Pub. ML = Mayorista × (1+{pct_ml}%)"
            f"  |  Bulto = al punto de equilibrio (más agresivo)"
            f"  |  Margen May. = utilidad real ya cubierto el operativo"
        )
        fecha_gen = datetime.now().strftime('%d/%m/%Y %H:%M')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rotación"

        hdr_fill  = PatternFill("solid", fgColor="1A56DB")
        hdr_font  = Font(bold=True, color="FFFFFFFF", size=10)
        alt_fill  = PatternFill("solid", fgColor="EFF6FF")
        title_font = Font(bold=True, size=13, color="111827")
        sub_font   = Font(size=9, color="374151")
        leg_font   = Font(size=8, italic=True, color="6B7280")
        right_al  = Alignment(horizontal="right", vertical="center")
        center_al = Alignment(horizontal="center", vertical="center")
        left_al   = Alignment(horizontal="left", vertical="center", wrap_text=True)
        NUM_COLS  = 16

        # Fila 1: Título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
        c = ws.cell(1, 1, f"MICROBELL S.A. — Análisis de Rotación de Stock  ({meses} meses)  ·  Generado: {fecha_gen}")
        c.font = title_font; c.alignment = left_al
        ws.row_dimensions[1].height = 22

        # Fila 2: Resumen
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NUM_COLS)
        c2 = ws.cell(2, 1, linea_resumen)
        c2.font = sub_font; c2.alignment = left_al
        ws.row_dimensions[2].height = 18

        # Fila 3: Leyenda
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=NUM_COLS)
        c3 = ws.cell(3, 1, linea_leyenda)
        c3.font = leg_font; c3.alignment = left_al
        ws.row_dimensions[3].height = 16

        # Fila 4: vacía
        ws.row_dimensions[4].height = 6

        # Fila 5: Encabezados de tabla
        HEADERS = ["Código", "Descripción", "Stock", "Costo ARS", "Últ. Compra",
                   "Lista1 ARS", "Mayorista", "Pub. ML", "Margen May.%",
                   "Pack ×10", "Pack ×20", "Pack ×30", "Pack ×40",
                   "Vendido (u.)", "Rotación %", "Última venta"]
        for ci, h in enumerate(HEADERS, 1):
            cell = ws.cell(5, ci, h)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center_al
        ws.row_dimensions[5].height = 20

        # Datos
        for ri, r in enumerate(rows, 6):
            is_alt = ri % 2 == 0
            fill = alt_fill if is_alt else None

            def wc(col, val, al=None, fmt=None):
                c = ws.cell(ri, col, val)
                if fill: c.fill = fill
                if al: c.alignment = al
                if fmt: c.number_format = fmt
                return c

            wc(1,  r.get('codigoparticular', ''))
            wc(2,  r.get('descripcion', ''),        left_al)
            wc(3,  round(r.get('stock', 0)),         right_al)
            costo = r.get('costo')
            wc(4,  round(costo) if costo else '—',   right_al, '#,##0' if costo else None)
            wc(5,  r.get('ultima_compra') or '—',    center_al)
            wc(6,  round(r.get('precio1_ars', 0)),   right_al, '#,##0')
            wc(7,  round(r.get('precio_mayorista', 0) or 0), right_al, '#,##0')
            wc(8,  round(r.get('precio_meli_pub', 0) or 0),  right_al, '#,##0')
            mg = r.get('margen_mayorista_pct')
            c_mg = wc(9, (str(mg)+'%') if mg is not None else '—', center_al)
            if mg is not None:
                if mg < 0:        c_mg.font = Font(color="DC2626", bold=True)
                elif mg < pct_ut: c_mg.font = Font(color="D97706")
                else:             c_mg.font = Font(color="15803D", bold=True)
            for pi, qty in enumerate(_PACK_QTYS):
                pv = _rotacion_precio_pack(r, qty)
                wc(10 + pi, pv if pv else '—', right_al, '#,##0' if pv else None)
            wc(14, round(r.get('cantidad_vendida_periodo', 0) or 0), right_al)
            wc(15, str(r.get('pct_venta', 0) or 0)+'%', center_al)
            wc(16, r.get('ultimo_movimiento') or '—')

        # Anchos de columna
        ws.column_dimensions['A'].width = 11   # Código
        ws.column_dimensions['B'].width = 42   # Descripción
        ws.column_dimensions['C'].width = 9    # Stock
        ws.column_dimensions['D'].width = 13   # Costo ARS
        ws.column_dimensions['E'].width = 13   # Últ. Compra
        ws.column_dimensions['F'].width = 13   # Lista1 ARS
        ws.column_dimensions['G'].width = 13   # Mayorista
        ws.column_dimensions['H'].width = 13   # Pub. ML
        ws.column_dimensions['I'].width = 13   # Margen May.%
        ws.column_dimensions['J'].width = 11   # Pack ×10
        ws.column_dimensions['K'].width = 11   # Pack ×20
        ws.column_dimensions['L'].width = 11   # Pack ×30
        ws.column_dimensions['M'].width = 11   # Pack ×40
        ws.column_dimensions['N'].width = 13   # Vendido (u.)
        ws.column_dimensions['O'].width = 12   # Rotación %
        ws.column_dimensions['P'].width = 16   # Última venta

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        from fastapi.responses import StreamingResponse
        fname = f"rotacion_{meses}m_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f"attachment; filename={fname}"})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/admin/rotacion/exportar-pdf")
def admin_rotacion_exportar_pdf(
    meses: int = 12,
    grupo: str = None,
    superrubro: str = None,
    rubro: str = None,
    marca: str = None,
    articulo: str = None,
    depositos: str = None,
    pct_operativo: float = 30,
    pct_utilidad: float = 25,
    pct_meli: float = 80,
    pct_venta_max: float = 20,
    compra_hasta: str = None,
    _u=Depends(get_admin_download_auth)
):
    try:
        import io
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        rows = admin_analisis_rotacion(
            meses=meses, grupo=grupo, superrubro=superrubro, rubro=rubro, marca=marca,
            articulo=articulo, depositos=depositos, pct_operativo=pct_operativo,
            pct_utilidad=pct_utilidad, pct_meli=pct_meli, pct_venta_max=pct_venta_max,
            compra_hasta=compra_hasta, _u=_u
        )
        if not rows:
            raise HTTPException(404, "Sin datos de rotación para los filtros indicados")

        cambio   = rows[0].get('cambio_usd', 1)
        pct_op   = rows[0].get('pct_operativo', pct_operativo)
        pct_ut   = rows[0].get('pct_utilidad', pct_utilidad)
        pct_ml   = rows[0].get('pct_meli', pct_meli)
        pct_vtx  = rows[0].get('pct_venta_max', pct_venta_max)

        sin_ventas   = sum(1 for r in rows if not (r.get('cantidad_vendida_periodo', 0) > 0))
        total_costo  = sum(r['stock'] * (r.get('costo') or 0) for r in rows)
        sin_costo    = sum(1 for r in rows if not r.get('costo'))
        con_bulto    = sum(1 for r in rows if r.get('bulto_sugerido'))
        con_perdida  = sum(1 for r in rows if r.get('margen_mayorista_pct') is not None and r['margen_mayorista_pct'] < 0)
        sin_compra   = sum(1 for r in rows if not r.get('ultima_compra'))

        ch_fmt = '/'.join(reversed(compra_hasta.split('-'))) if compra_hasta else None
        linea_resumen = (
            f"<b>{len(rows)}</b> artículos de baja rotación (ventas ≤ {pct_vtx}% del stock)"
            + (f" &nbsp;·&nbsp; Compras anteriores al <b>{ch_fmt}</b>" if ch_fmt else "")
            + f" &nbsp;·&nbsp; {sin_ventas} sin ninguna venta en el período"
            + f" &nbsp;·&nbsp; Capital inmovilizado (costo): <b>${round(total_costo):,} ARS</b>"
            + f" &nbsp;·&nbsp; Cambio USD: ${cambio:,}"
            + (f" &nbsp;·&nbsp; <font color='#7c3aed'><b>{con_bulto} con venta por bulto sugerida</b></font>" if con_bulto else "")
            + (f" &nbsp;·&nbsp; <font color='#dc2626'><b>{con_perdida} no cubren el punto de equilibrio</b></font>" if con_perdida else "")
            + (f" &nbsp;·&nbsp; <font color='#d97706'>{sin_costo} sin costo cargado</font>" if sin_costo else "")
            + (f" &nbsp;·&nbsp; <font color='#6b7280'>{sin_compra} sin remito de compra en el sistema</font>" if sin_compra else "")
        )
        linea_leyenda = (
            f"Baja rotación = ventas NV ÷ stock actual ≤ {pct_vtx}%  |  "
            f"Punto de equilibrio = Costo × (1+{pct_op}%)  |  "
            f"Mayorista = equilibrio × (1+{pct_ut}%), sin superar Lista1  |  "
            f"Pub. ML = Mayorista × (1+{pct_ml}%)  |  "
            f"Bulto = al punto de equilibrio (más agresivo)  |  "
            f"Margen May. = utilidad real ya cubierto el operativo"
        )
        fecha_gen = datetime.now().strftime('%d/%m/%Y %H:%M')

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=10*mm, rightMargin=10*mm,
                                topMargin=12*mm, bottomMargin=12*mm)

        styles = getSampleStyleSheet()
        st_title = ParagraphStyle('rot_title', fontName='Helvetica-Bold', fontSize=13,
                                  textColor=colors.HexColor('#111827'), spaceAfter=4)
        st_sub   = ParagraphStyle('rot_sub',   fontName='Helvetica', fontSize=9,
                                  textColor=colors.HexColor('#374151'), spaceAfter=2)
        st_leg   = ParagraphStyle('rot_leg',   fontName='Helvetica-Oblique', fontSize=7.5,
                                  textColor=colors.HexColor('#6b7280'), spaceAfter=8)
        st_desc  = ParagraphStyle('rot_desc',  fontName='Helvetica', fontSize=7.5,
                                  textColor=colors.HexColor('#111827'), leading=9,
                                  wordWrap='LTR')

        story = []
        story.append(Paragraph(f"MICROBELL S.A. — Análisis de Rotación de Stock &nbsp;({meses} meses) &nbsp;·&nbsp; Generado: {fecha_gen}", st_title))
        story.append(Paragraph(linea_resumen, st_sub))
        story.append(Paragraph(linea_leyenda, st_leg))

        # Tabla
        BLUE    = colors.HexColor('#1A56DB')
        RED     = colors.HexColor('#DC2626')
        ORANGE  = colors.HexColor('#D97706')
        GREEN   = colors.HexColor('#15803D')
        PURPLE  = colors.HexColor('#7C3AED')
        ALTBG   = colors.HexColor('#EFF6FF')

        def fmt_ars(v): return f"${round(v):,}".replace(',', '.') if v else '—'
        def fmt_pct(v): return f"{v}%" if v is not None else '—'

        col_headers = ["Código", "Descripción", "Stock", "Costo ARS", "Últ.\nCompra",
                       "Lista1", "Mayorista", "Pub. ML", "Margen\nMay.%",
                       "Pack\n×10", "Pack\n×20", "Pack\n×30", "Pack\n×40",
                       "Vendido\n(u.)", "Rot.%", "Última\nventa"]
        col_widths  = [18*mm, 46*mm, 11*mm, 15*mm, 15*mm, 15*mm, 15*mm, 15*mm, 12*mm,
                       12*mm, 12*mm, 12*mm, 12*mm, 13*mm, 10*mm, 15*mm]

        tbl_data = [col_headers]
        row_colors = []  # (row_idx, color) for margen color
        for ri, r in enumerate(rows, 1):
            mg   = r.get('margen_mayorista_pct')
            if mg is not None:
                if mg < 0:        row_colors.append((ri, RED, 8))
                elif mg < pct_ut: row_colors.append((ri, ORANGE, 8))
                else:             row_colors.append((ri, GREEN, 8))
            tbl_data.append([
                r.get('codigoparticular', ''),
                Paragraph(r.get('descripcion', ''), st_desc),
                str(round(r.get('stock', 0))),
                fmt_ars(r.get('costo')),
                r.get('ultima_compra') or '—',
                fmt_ars(r.get('precio1_ars')),
                fmt_ars(r.get('precio_mayorista')),
                fmt_ars(r.get('precio_meli_pub')),
                fmt_pct(mg),
                fmt_ars(_rotacion_precio_pack(r, 10)),
                fmt_ars(_rotacion_precio_pack(r, 20)),
                fmt_ars(_rotacion_precio_pack(r, 30)),
                fmt_ars(_rotacion_precio_pack(r, 40)),
                str(round(r.get('cantidad_vendida_periodo', 0) or 0)),
                fmt_pct(r.get('pct_venta')),
                r.get('ultimo_movimiento') or '—',
            ])

        base_style = [
            ('BACKGROUND',  (0,0), (-1,0), BLUE),
            ('TEXTCOLOR',   (0,0), (-1,0), colors.white),
            ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',    (0,0), (-1,0), 8),
            ('ALIGN',       (0,0), (-1,0), 'CENTER'),
            ('VALIGN',      (0,0), (-1,-1), 'MIDDLE'),
            ('FONTNAME',    (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE',    (0,1), (-1,-1), 7.5),
            ('ROWBACKGROUND', (0,1), (-1,-1), [colors.white, ALTBG]),
            ('GRID',        (0,0), (-1,-1), 0.3, colors.HexColor('#D1D5DB')),
            ('ALIGN',       (2,1), (3,-1), 'RIGHT'),   # Stock, Costo ARS
            ('ALIGN',       (4,1), (4,-1), 'CENTER'),  # Últ. Compra
            ('ALIGN',       (5,1), (7,-1), 'RIGHT'),   # Lista1, Mayorista, Pub. ML
            ('ALIGN',       (8,1), (8,-1), 'CENTER'),  # Margen
            ('ALIGN',       (9,1), (12,-1), 'RIGHT'),  # Pack ×10..×40
            ('ALIGN',       (13,1), (13,-1), 'RIGHT'), # Vendido
            ('ALIGN',       (14,1), (15,-1), 'CENTER'),# Rot%, Última venta
        ]
        for (ri, clr, ci) in row_colors:
            base_style.append(('TEXTCOLOR', (ci, ri), (ci, ri), clr))
            base_style.append(('FONTNAME',  (ci, ri), (ci, ri), 'Helvetica-Bold'))

        tbl = Table(tbl_data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle(base_style))
        story.append(tbl)

        doc.build(story)
        buf.seek(0)
        from fastapi.responses import StreamingResponse
        fname = f"rotacion_{meses}m_{datetime.now().strftime('%Y%m%d')}.pdf"
        return StreamingResponse(buf, media_type="application/pdf",
                                 headers={"Content-Disposition": f"inline; filename={fname}"})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/admin/rotacion-detalle-ventas")
def admin_rotacion_detalle_ventas(art_id: str, meses: int = 12, _u=Depends(get_admin_user)):
    """Detalle de comprobantes de venta (FA/FB/FE) y notas de crédito
    (FCA/FCB/FCE/FCCA/FCCB/FCCE, con signo negativo) de un artículo puntual
    en el período, para el modal que se abre al hacer clic en la columna de
    ventas de Rotación.

    art_id = CODIGOARTICULO interno (NO CODIGOPARTICULAR) — el mismo valor
    usado para construir vendido_map en /admin/analisis-rotacion. Se busca
    directo por ese campo en CUERPOPEDIDOS (L1) y CUERPOCOMPROBANTES (SW),
    sin pasar por un lookup adicional en ARTICULOS por CODIGOPARTICULAR,
    porque ese lookup puede resolver a un CODIGOARTICULO distinto cuando
    hay códigos particulares duplicados — causaba que el modal apareciera
    vacío aunque el análisis principal sí contara ventas."""
    # FA/FB/FE: cantidad positiva; NCA/NCB/FC*: cantidad negativa (ya firmada en DB).
    TIPOS_VENTA_TODOS = ('FA', 'FB', 'FE', 'NCA', 'NCB', 'FCA', 'FCB', 'FCE', 'FCCA', 'FCCB', 'FCCE')
    ph_tipos = ','.join(f"'{t}'" for t in TIPOS_VENTA_TODOS)
    fecha_corte = (datetime.now() - timedelta(days=meses * 30.44))
    detalle = []

    def _calc(cant_raw, precio_u, dto, iva_pct):
        cant = float(cant_raw or 0)  # ya firmada (negativa para NC*)
        precio_u = float(precio_u or 0)
        dto = float(dto or 0)
        iva_pct = float(iva_pct) if iva_pct is not None else 21.0
        importe_neto = round(cant * precio_u * (1 - dto / 100), 2)
        iva_monto = round(importe_neto * iva_pct / 100, 2)
        importe_con_iva = round(importe_neto + iva_monto, 2)
        return cant, importe_neto, iva_pct, iva_monto, importe_con_iva

    # L1: CABEZACOMPROBANTES/CUERPOCOMPROBANTES (DATABASE, WIN1252).
    # CABEZAPEDIDOS solo contiene pedidos NP, no facturas FA.
    try:
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()
        cur.execute(f"""
            SELECT cb.FECHACOMPROBANTE, cb.TIPOCOMPROBANTE, cb.NUMEROCOMPROBANTE,
                   cu.CANTIDAD, cu.PRECIOUNITARIO, cu.DESCUENTO, cu.PORCENTAJEIVA,
                   cb.RAZONSOCIAL
            FROM "CUERPOCOMPROBANTES" cu
            JOIN "CABEZACOMPROBANTES" cb
              ON cb.TIPOCOMPROBANTE=cu.TIPOCOMPROBANTE
             AND cb.NUMEROCOMPROBANTE=cu.NUMEROCOMPROBANTE
            WHERE cu.CODIGOARTICULO=? AND cu.TIPOCOMPROBANTE IN ({ph_tipos}) AND cb.ANULADA=0
              AND cb.FECHACOMPROBANTE >= ?
            ORDER BY cb.FECHACOMPROBANTE DESC
        """, (art_id, fecha_corte))
        for r in cur.fetchall():
            tipo = (r[1] or '').strip().upper()
            cant, importe, iva_pct, iva_monto, importe_con_iva = _calc(r[3], r[4], r[5], r[6])
            detalle.append({
                'origen': 'L1',
                'fecha': r[0].strftime('%Y-%m-%d') if hasattr(r[0], 'strftime') else str(r[0])[:10],
                'tipo': tipo,
                'numero': (r[2] or '').strip() if isinstance(r[2], str) else r[2],
                'razonsocial': (r[7] or '').strip() if isinstance(r[7], str) else r[7],
                'cantidad': cant,
                'importe': importe,
                'iva_pct': iva_pct,
                'iva_monto': iva_monto,
                'importe_con_iva': importe_con_iva,
            })
        c.close()
    except Exception:
        pass

    try:
        c_sw = conn('LATIN1', db=DATABASE_MLT)
        cur_sw = c_sw.cursor()
        cur_sw.execute(f"""
            SELECT cb.FECHACOMPROBANTE, cb.TIPOCOMPROBANTE, cb.NUMEROCOMPROBANTE,
                   cu.CANTIDAD, cu.PRECIOUNITARIO, cu.DESCUENTO, cu.PORCENTAJEIVA,
                   cb.RAZONSOCIAL
            FROM "CUERPOCOMPROBANTES" cu
            JOIN "CABEZACOMPROBANTES" cb
              ON cb.TIPOCOMPROBANTE=cu.TIPOCOMPROBANTE
             AND cb.NUMEROCOMPROBANTE=cu.NUMEROCOMPROBANTE
            WHERE cu.CODIGOARTICULO=? AND cu.TIPOCOMPROBANTE IN ({ph_tipos}) AND cb.ANULADA=0
              AND cb.FECHACOMPROBANTE >= ?
            ORDER BY cb.FECHACOMPROBANTE DESC
        """, (art_id, fecha_corte))
        for r in cur_sw.fetchall():
            tipo = (r[1] or '').strip().upper()
            cant, importe, iva_pct, iva_monto, importe_con_iva = _calc(r[3], r[4], r[5], r[6])
            detalle.append({
                'origen': 'SW',
                'fecha': r[0].strftime('%Y-%m-%d') if hasattr(r[0], 'strftime') else str(r[0])[:10],
                'tipo': tipo,
                'numero': (r[2] or '').strip() if isinstance(r[2], str) else r[2],
                'razonsocial': (r[7] or '').strip() if isinstance(r[7], str) else r[7],
                'cantidad': cant,
                'importe': importe,
                'iva_pct': iva_pct,
                'iva_monto': iva_monto,
                'importe_con_iva': importe_con_iva,
            })
        c_sw.close()
    except Exception:
        pass

    detalle.sort(key=lambda d: d['fecha'], reverse=True)
    return detalle


@app.get("/debug/rotacion-fca-articulo")
def debug_rotacion_fca_articulo(art_id: str = None, codigoparticular: str = None, _u=Depends(get_admin_user)):
    """DEBUG TEMPORAL: trae TODAS las filas de CUERPOPEDIDOS/CABEZAPEDIDOS (L1)
    y CUERPOCOMPROBANTES/CABEZACOMPROBANTES (SW) para un CODIGOARTICULO dado,
    de CUALQUIER tipo de comprobante, fecha y estado ANULADA — sin ningún
    filtro — para detectar si un comprobante puntual (ej. una FCA que el
    usuario sabe que existe en Flexxus) está en la base, y por qué no pasa
    los filtros normales de /admin/rotacion-detalle-ventas (fecha, ANULADA,
    o el tipo de comprobante).

    Acepta art_id (CODIGOARTICULO, recomendado) o, si no se tiene a mano,
    codigoparticular (resuelve a CODIGOARTICULO vía ARTICULOS — solo para
    este debug, no usar este camino en lógica de producción)."""
    if not art_id and codigoparticular:
        try:
            c0 = conn('WIN1252', db=DATABASE)
            cur0 = c0.cursor()
            cur0.execute('SELECT CODIGOARTICULO FROM "ARTICULOS" WHERE CODIGOPARTICULAR=?', (codigoparticular,))
            row0 = cur0.fetchone()
            art_id = str(row0[0]).strip() if row0 else None
            c0.close()
        except Exception:
            pass
    if not art_id:
        raise HTTPException(status_code=400, detail="Debes pasar art_id o codigoparticular")
    out = {'art_id_usado': art_id, 'l1': [], 'sw': [], 'error_l1': None, 'error_sw': None}
    try:
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()
        cur.execute("""
            SELECT cb.TIPOCOMPROBANTE, cb.NUMEROCOMPROBANTE, cb.FECHACOMPROBANTE,
                   cb.ANULADA, cp.CANTIDAD, cp.PRECIOUNITARIO, cb.RAZONSOCIAL
            FROM "CUERPOPEDIDOS" cp
            JOIN "CABEZAPEDIDOS" cb
              ON cb.TIPOCOMPROBANTE=cp.TIPOCOMPROBANTE
             AND cb.NUMEROCOMPROBANTE=cp.NUMEROCOMPROBANTE
            WHERE cp.CODIGOARTICULO=?
            ORDER BY cb.FECHACOMPROBANTE DESC
        """, (art_id,))
        for r in cur.fetchall():
            out['l1'].append({
                'tipo': (r[0] or '').strip() if isinstance(r[0], str) else r[0],
                'numero': (r[1] or '').strip() if isinstance(r[1], str) else r[1],
                'fecha': r[2].strftime('%Y-%m-%d') if hasattr(r[2], 'strftime') else str(r[2]),
                'anulada': r[3],
                'cantidad': r[4],
                'precio_u': r[5],
                'razonsocial': (r[6] or '').strip() if isinstance(r[6], str) else r[6],
            })
        c.close()
    except Exception as e:
        out['error_l1'] = str(e)

    # Test de hipótesis: ¿existe también CABEZACOMPROBANTES/CUERPOCOMPROBANTES
    # DENTRO de la base de Línea1 (no la de SW), con las facturas FA reales que
    # no aparecen en CABEZAPEDIDOS (que parece guardar solo NP/pedidos)?
    out['l1_comprobantes'] = []
    out['error_l1_comprobantes'] = None
    try:
        c1c = conn('WIN1252', db=DATABASE)
        cur1c = c1c.cursor()
        cur1c.execute("""
            SELECT cb.TIPOCOMPROBANTE, cb.NUMEROCOMPROBANTE, cb.FECHACOMPROBANTE,
                   cb.ANULADA, cu.CANTIDAD, cu.PRECIOUNITARIO, cb.RAZONSOCIAL
            FROM "CUERPOCOMPROBANTES" cu
            JOIN "CABEZACOMPROBANTES" cb
              ON cb.TIPOCOMPROBANTE=cu.TIPOCOMPROBANTE
             AND cb.NUMEROCOMPROBANTE=cu.NUMEROCOMPROBANTE
            WHERE cu.CODIGOARTICULO=?
            ORDER BY cb.FECHACOMPROBANTE DESC
        """, (art_id,))
        for r in cur1c.fetchall():
            out['l1_comprobantes'].append({
                'tipo': (r[0] or '').strip() if isinstance(r[0], str) else r[0],
                'numero': (r[1] or '').strip() if isinstance(r[1], str) else r[1],
                'fecha': r[2].strftime('%Y-%m-%d') if hasattr(r[2], 'strftime') else str(r[2]),
                'anulada': r[3],
                'cantidad': r[4],
                'precio_u': r[5],
                'razonsocial': (r[6] or '').strip() if isinstance(r[6], str) else r[6],
            })
        c1c.close()
    except Exception as e:
        out['error_l1_comprobantes'] = str(e)

    try:
        c_sw = conn('LATIN1', db=DATABASE_MLT)
        cur_sw = c_sw.cursor()
        cur_sw.execute("""
            SELECT cb.TIPOCOMPROBANTE, cb.NUMEROCOMPROBANTE, cb.FECHACOMPROBANTE,
                   cb.ANULADA, cu.CANTIDAD, cu.PRECIOUNITARIO, cb.RAZONSOCIAL
            FROM "CUERPOCOMPROBANTES" cu
            JOIN "CABEZACOMPROBANTES" cb
              ON cb.TIPOCOMPROBANTE=cu.TIPOCOMPROBANTE
             AND cb.NUMEROCOMPROBANTE=cu.NUMEROCOMPROBANTE
            WHERE cu.CODIGOARTICULO=?
            ORDER BY cb.FECHACOMPROBANTE DESC
        """, (art_id,))
        for r in cur_sw.fetchall():
            out['sw'].append({
                'tipo': (r[0] or '').strip() if isinstance(r[0], str) else r[0],
                'numero': (r[1] or '').strip() if isinstance(r[1], str) else r[1],
                'fecha': r[2].strftime('%Y-%m-%d') if hasattr(r[2], 'strftime') else str(r[2]),
                'anulada': r[3],
                'cantidad': r[4],
                'precio_u': r[5],
                'razonsocial': (r[6] or '').strip() if isinstance(r[6], str) else r[6],
            })
        c_sw.close()
    except Exception as e:
        out['error_sw'] = str(e)

    return out


@app.get("/debug/fma-stock-articulo")
def debug_fma_stock_articulo(codigo: str = '03375', _u=Depends(get_admin_user)):
    """DEBUG TEMPORAL: llama FMA_STOCK depósito por depósito (igual que el
    resto del proyecto) y muestra el STOCKREMANENTE crudo de un artículo
    puntual en cada uno, para comparar contra lo que muestra Stock por
    Depósito y aislar dónde está la discrepancia. Una sola conexión
    reutilizada (la versión anterior abría 24 conexiones y daba timeout)."""
    deps = ['001', '002', '003', '005', '006', '007', '008', '010', '011', '013', '016', '017']
    try:
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()
        cur.execute('SELECT CODIGOARTICULO, CODIGOPARTICULAR FROM "ARTICULOS" WHERE CODIGOPARTICULAR=?', (codigo,))
        row0 = cur.fetchone()
        codigo_articulo_interno = str(row0[0]).strip() if row0 and row0[0] is not None else None

        por_codigoparticular = {}
        por_codigoarticulo = {}
        for dep in deps:
            try:
                cur.execute(
                    f'SELECT ID_ARTICULO, STOCKREMANENTE FROM "FMA_STOCK"(NULL, NULL, \'{dep}\', 1, 1) '
                    f'WHERE ID_ARTICULO=?', (codigo,)
                )
                row = cur.fetchone()
                por_codigoparticular[dep] = float(row[1]) if row else None
            except Exception as e_dep:
                por_codigoparticular[dep] = f"ERROR: {e_dep}"
            if codigo_articulo_interno and codigo_articulo_interno != codigo:
                try:
                    cur.execute(
                        f'SELECT ID_ARTICULO, STOCKREMANENTE FROM "FMA_STOCK"(NULL, NULL, \'{dep}\', 1, 1) '
                        f'WHERE ID_ARTICULO=?', (codigo_articulo_interno,)
                    )
                    row_i = cur.fetchone()
                    por_codigoarticulo[dep] = float(row_i[1]) if row_i else None
                except Exception as e_dep2:
                    por_codigoarticulo[dep] = f"ERROR: {e_dep2}"
        c.close()
        suma_part = sum(v for v in por_codigoparticular.values() if isinstance(v, (int, float)))
        suma_art  = sum(v for v in por_codigoarticulo.values() if isinstance(v, (int, float)))
        return {
            "codigo_buscado_codigoparticular": codigo,
            "codigoarticulo_interno_real": codigo_articulo_interno,
            "stock_por_deposito_usando_CODIGOPARTICULAR": por_codigoparticular,
            "suma_usando_CODIGOPARTICULAR": suma_part,
            "stock_por_deposito_usando_CODIGOARTICULO_interno": por_codigoarticulo,
            "suma_usando_CODIGOARTICULO_interno": suma_art,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/debug/articulos-columnas-costo")
def debug_articulos_columnas_costo(codigo: str = '03748', _u=Depends(get_admin_user)):
    """DEBUG TEMPORAL: lista todas las columnas de ARTICULOS y sus valores para
    un código puntual, para identificar cuál corresponde al Costo Reposición
    que se ve en Flexxus (la columna COSTO no lo trae)."""
    try:
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()
        cur.execute('SELECT * FROM "ARTICULOS" WHERE CODIGOARTICULO=?', (codigo,))
        row = cur.fetchone()
        if not row:
            c.close()
            return {"error": f"artículo {codigo} no encontrado"}
        cols = [d[0] for d in cur.description]
        c.close()
        # Solo columnas que contengan COSTO o PRECIO en el nombre, para no saturar
        relevantes = {cols[i]: row[i] for i in range(len(cols)) if 'COSTO' in cols[i].upper() or 'PRECIO' in cols[i].upper()}
        return {"codigo": codigo, "columnas_costo_precio": relevantes, "todas_las_columnas": cols}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/debug/tablas-costo")
def debug_tablas_costo(_u=Depends(get_admin_user)):
    """DEBUG TEMPORAL: lista tablas del esquema cuyo nombre sugiere que guardan
    el costo de reposición de artículos (no está en ARTICULOS directamente)."""
    try:
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()
        cur.execute(
            "SELECT TRIM(RDB$RELATION_NAME) FROM RDB$RELATIONS "
            "WHERE RDB$SYSTEM_FLAG = 0 AND ("
            "UPPER(RDB$RELATION_NAME) CONTAINING 'COSTO' OR "
            "UPPER(RDB$RELATION_NAME) CONTAINING 'REPOSICION' OR "
            "UPPER(RDB$RELATION_NAME) CONTAINING 'ARTICULOPRECIO' OR "
            "UPPER(RDB$RELATION_NAME) CONTAINING 'PRECIOCOSTO'"
            ") ORDER BY 1"
        )
        tablas = [r[0] for r in cur.fetchall()]
        c.close()
        return {"tablas_candidatas": tablas}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/debug/cambiosdecostos")
def debug_cambiosdecostos(codigo: str = '03748', _u=Depends(get_admin_user)):
    """DEBUG TEMPORAL: inspecciona CAMBIOSDECOSTOS, candidata a guardar el
    Costo Reposición actual de un artículo (no está en ARTICULOS)."""
    try:
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()
        cur.execute('SELECT FIRST 1 * FROM "CAMBIOSDECOSTOS"')
        row0 = cur.fetchone()
        cols = [d[0] for d in cur.description] if row0 else []
        # Buscar columna de artículo entre las disponibles
        col_art = next((cc for cc in cols if 'ARTICULO' in cc.upper()), None)
        registros = []
        if col_art:
            cur.execute(
                f'SELECT FIRST 5 * FROM "CAMBIOSDECOSTOS" WHERE {col_art}=? ORDER BY 1 DESC',
                (codigo,)
            )
            for r in cur.fetchall():
                registros.append({cols[i]: r[i] for i in range(len(cols))})
        c.close()
        return {"columnas": cols, "columna_articulo_detectada": col_art, "registros_para_codigo": registros}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/admin/articulos-buscar")
def admin_articulos_buscar(q: str = Query("", min_length=2), _u=Depends(get_admin_user)):
    """Autocomplete rápido de artículos para filtro Rotación (sin FMA_STOCK)."""
    try:
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()
        cur.execute(
            "SELECT FIRST 15 CODIGOARTICULO, CODIGOPARTICULAR, DESCRIPCION "
            "FROM \"ARTICULOS\" WHERE ACTIVO = '1' "
            "AND (UPPER(CODIGOPARTICULAR) CONTAINING UPPER(?) OR UPPER(DESCRIPCION) CONTAINING UPPER(?)) "
            "ORDER BY CASE WHEN UPPER(CODIGOPARTICULAR) STARTING WITH UPPER(?) THEN 0 ELSE 1 END, "
            "CODIGOPARTICULAR, DESCRIPCION",
            (q, q, q)
        )
        rows = cur.fetchall()
        c.close()
        return [{'codigo': r[0], 'codigoparticular': r[1] or r[0], 'descripcion': r[2] or ''} for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/admin/ofertas")
def admin_get_ofertas(_u=Depends(get_admin_user)):
    c = _admin_db()
    offers = [dict(r) for r in c.execute("SELECT * FROM offers ORDER BY created_at DESC").fetchall()]
    if not offers:
        c.close()
        return offers
    ids = [o['id'] for o in offers]
    idx = {o['id']: o for o in offers}
    # Inicializar relaciones y defaults
    for o in offers:
        o['product_details']  = []
        o['financial_details'] = []
        o['conditions']       = []
        o['vendors']          = []
        o['profiles']         = []
        o['category_filters'] = []
        o['discount_filters'] = []
        o['combo_escalones']   = []
        o['amount_escalones']  = []
        o['clients']           = []
        for k, d in [('deposito',''),('tipo_financiero','descuento_total'),('monto_minimo',0),('cupo',0),('usos',0),('tope_bonificacion_pesos',0),('bonificado_acumulado_pesos',0),('acumulable',0)]:
            if k not in o: o[k] = d
    ph = ','.join('?' * len(ids))
    for r in c.execute(f"SELECT * FROM offer_product_details WHERE offer_id IN ({ph})", ids):
        d = dict(r)
        if d['offer_id'] in idx: idx[d['offer_id']]['product_details'].append(d)
    for r in c.execute(f"SELECT * FROM offer_financial_details WHERE offer_id IN ({ph}) ORDER BY orden", ids):
        d = dict(r)
        if d['offer_id'] in idx: idx[d['offer_id']]['financial_details'].append(d)
    for r in c.execute(f"SELECT offer_id, condicion_comercial FROM offer_conditions WHERE offer_id IN ({ph})", ids):
        if r[0] in idx: idx[r[0]]['conditions'].append(r[1])
    for r in c.execute(f"SELECT offer_id, codigousuario FROM offer_vendors WHERE offer_id IN ({ph})", ids):
        if r[0] in idx: idx[r[0]]['vendors'].append(r[1])
    for r in c.execute(f"SELECT offer_id, perfil_codigo FROM offer_profiles WHERE offer_id IN ({ph})", ids):
        if r[0] in idx: idx[r[0]]['profiles'].append(r[1])
    for r in c.execute(f"SELECT offer_id, nivel, valor FROM offer_category_filters WHERE offer_id IN ({ph})", ids):
        if r[0] in idx: idx[r[0]]['category_filters'].append({'nivel': r[1], 'valor': r[2]})
    for r in c.execute(f"SELECT offer_id, nivel, valor FROM offer_discount_filters WHERE offer_id IN ({ph})", ids):
        if r[0] in idx: idx[r[0]]['discount_filters'].append({'nivel': r[1], 'valor': r[2]})
    for r in c.execute(f"SELECT offer_id, min_combos, descuento_pct FROM offer_combo_escalones WHERE offer_id IN ({ph}) ORDER BY min_combos", ids):
        if r[0] in idx: idx[r[0]]['combo_escalones'].append({'min_combos': r[1], 'descuento_pct': r[2]})
    for r in c.execute(f"SELECT offer_id, monto_minimo, descuento_pct, condicion_comercial FROM offer_amount_escalones WHERE offer_id IN ({ph}) ORDER BY monto_minimo", ids):
        if r[0] in idx: idx[r[0]]['amount_escalones'].append({'monto_minimo': r[1], 'descuento_pct': r[2], 'condicion_comercial': r[3]})
    for r in c.execute(f"SELECT offer_id, codigocliente, razonsocial, descuento_extra_pct, vencimiento_extra, monto_minimo_extra, tipo_cartera, condicion_comercial_extra FROM offer_clients WHERE offer_id IN ({ph})", ids):
        if r[0] in idx: idx[r[0]]['clients'].append({'codigocliente': r[1], 'razonsocial': r[2], 'descuento_extra_pct': r[3], 'vencimiento_extra': r[4], 'monto_minimo_extra': r[5], 'tipo_cartera': r[6], 'condicion_comercial_extra': r[7]})
    c.close()
    _resolver_codigoparticular_clientes(offers)
    return offers

def _save_offer_relations(c, id_, data):
    for e in (data.get('combo_escalones') or []):
        mc = int(e.get('min_combos') or 1)
        dp = float(e.get('descuento_pct') or 0)
        if mc >= 1 and dp > 0:
            c.execute("INSERT INTO offer_combo_escalones (offer_id, min_combos, descuento_pct) VALUES (?,?,?)", (id_, mc, dp))
    for e in (data.get('amount_escalones') or []):
        mm = float(e.get('monto_minimo') or 0)
        dp = float(e.get('descuento_pct') or 0)
        cc = str(e.get('condicion_comercial') or '').strip() or None
        if (dp > 0 or cc):  # mm=0 válido: condición se aplica sin mínimo de compra
            c.execute("INSERT INTO offer_amount_escalones (offer_id, monto_minimo, descuento_pct, condicion_comercial) VALUES (?,?,?,?)", (id_, mm, dp, cc))
    for d in (data.get('product_details') or []):
        c.execute("INSERT INTO offer_product_details (offer_id, codigo_producto, descripcion, cantidad, bonificacion_pct) VALUES (?,?,?,?,?)",
                  (id_, d.get('codigo_producto'), d.get('descripcion',''), d.get('cantidad',1), d.get('bonificacion_pct',0)))
    for i, d in enumerate(data.get('financial_details') or []):
        c.execute("INSERT INTO offer_financial_details (offer_id, porcentaje, orden) VALUES (?,?,?)", (id_, d.get('porcentaje',0), i))
    fe = data.get('financial_escalones') or []
    c.execute("UPDATE offers SET financial_escalones=? WHERE id=?", (json.dumps(fe) if fe else None, id_))
    for cond in (data.get('conditions') or []):
        try: c.execute("INSERT OR IGNORE INTO offer_conditions (offer_id, condicion_comercial) VALUES (?,?)", (id_, cond))
        except: pass
    for vend in (data.get('vendors') or []):
        v = vend.upper() if vend else None
        if v:
            try: c.execute("INSERT OR IGNORE INTO offer_vendors (offer_id, codigousuario) VALUES (?,?)", (id_, v))
            except: pass
    for perf in (data.get('profiles') or []):
        if perf:
            try: c.execute("INSERT OR IGNORE INTO offer_profiles (offer_id, perfil_codigo) VALUES (?,?)", (id_, perf.upper()))
            except: pass
    for f in (data.get('category_filters') or []):
        if f.get('nivel') and f.get('valor'):
            try: c.execute("INSERT OR IGNORE INTO offer_category_filters (offer_id, nivel, valor) VALUES (?,?,?)", (id_, f['nivel'], f['valor']))
            except: pass
    for f in (data.get('discount_filters') or []):
        if f.get('nivel') and f.get('valor'):
            try: c.execute("INSERT OR IGNORE INTO offer_discount_filters (offer_id, nivel, valor) VALUES (?,?,?)", (id_, f['nivel'], f['valor']))
            except: pass
    for cli in (data.get('clients') or []):
        cod = str(cli.get('codigocliente') or '').strip()
        if cod:
            try: c.execute("""INSERT OR IGNORE INTO offer_clients
                               (offer_id, codigocliente, razonsocial, descuento_extra_pct, vencimiento_extra, monto_minimo_extra, tipo_cartera, condicion_comercial_extra)
                               VALUES (?,?,?,?,?,?,?,?)""",
                           (id_, cod, str(cli.get('razonsocial') or '').strip(),
                            float(cli.get('descuento_extra_pct') or 0), str(cli.get('vencimiento_extra') or ''),
                            float(cli.get('monto_minimo_extra') or 0),
                            str(cli.get('tipo_cartera') or 'descuento'),
                            str(cli.get('condicion_comercial_extra') or '')))
            except: pass
    c.commit()

@app.get("/admin/ofertas/{id}")
def admin_get_oferta_by_id(id: int, _u=Depends(get_admin_user)):
    c = _admin_db()
    row = c.execute("SELECT * FROM offers WHERE id=?", (id,)).fetchone()
    if not row: c.close(); raise HTTPException(404, "Oferta no encontrada")
    o = dict(row)
    _load_offer_relations(c, o)
    c.close()
    return o

@app.post("/admin/ofertas")
def admin_create_oferta(data: dict, _u=Depends(get_admin_user)):
    nombre = (data.get('nombre') or '').strip()
    tipo   = (data.get('tipo') or '').strip()
    if not nombre or not tipo:
        raise HTTPException(400, "nombre y tipo requeridos")
    c = _admin_db()
    c.execute("INSERT INTO offers (nombre, tipo, descripcion, fecha_desde, fecha_hasta, deposito, tipo_financiero, monto_minimo, cupo, usos, tope_bonificacion_pesos, bonificado_acumulado_pesos, acumulable) VALUES (?,?,?,?,?,?,?,?,?,0,?,0,?)",
              (nombre, tipo, data.get('descripcion','').strip(), data.get('fecha_desde'), data.get('fecha_hasta'),
               data.get('deposito','').strip(), data.get('tipo_financiero','descuento_total'),
               data.get('monto_minimo',0), int(data.get('cupo',0) or 0), float(data.get('tope_bonificacion_pesos',0) or 0),
               1 if data.get('acumulable') else 0))
    c.commit(); id_ = c.execute("SELECT last_insert_rowid()").fetchone()[0]
    _save_offer_relations(c, id_, data)
    c.close()
    return {"id": id_, "ok": True}

@app.put("/admin/ofertas/{id}")
def admin_update_oferta(id: int, data: dict, _u=Depends(get_admin_user)):
    c = _admin_db()
    nuevo_cupo = int(data.get('cupo',0) or 0)
    nuevo_tope = float(data.get('tope_bonificacion_pesos',0) or 0)
    # Al ampliar el cupo o el tope de bonificación, reactivar si estaba inactiva por
    # agotamiento de cualquiera de los dos (mismo criterio para ambos).
    row = c.execute("SELECT usos, activo, bonificado_acumulado_pesos FROM offers WHERE id=?", (id,)).fetchone()
    usos_actual = row['usos'] if row else 0
    acumulado_actual = (row['bonificado_acumulado_pesos'] if row else 0) or 0
    activo_nuevo = data.get('activo',1)  # respetar lo que manda el formulario
    c.execute("""UPDATE offers SET nombre=?, tipo=?, descripcion=?, fecha_desde=?, fecha_hasta=?,
                 activo=?, deposito=?, tipo_financiero=?, monto_minimo=?, cupo=?, tope_bonificacion_pesos=?, acumulable=? WHERE id=?""",
              (data.get('nombre','').strip(), data.get('tipo','').strip(),
               data.get('descripcion','').strip(), data.get('fecha_desde'), data.get('fecha_hasta'),
               activo_nuevo, data.get('deposito','').strip(),
               data.get('tipo_financiero','descuento_total'), data.get('monto_minimo',0), nuevo_cupo, nuevo_tope,
               1 if data.get('acumulable') else 0, id))
    for tbl in ('offer_product_details','offer_financial_details','offer_conditions',
                'offer_vendors','offer_profiles','offer_category_filters','offer_discount_filters',
                'offer_combo_escalones','offer_amount_escalones','offer_clients'):
        c.execute(f"DELETE FROM {tbl} WHERE offer_id=?", (id,))
    _save_offer_relations(c, id, data)
    c.close()
    return {"ok": True}

@app.delete("/admin/ofertas/{id}")
def admin_delete_oferta(id: int, _u=Depends(get_admin_user)):
    c = _admin_db(); c.execute("DELETE FROM offers WHERE id=?", (id,)); c.commit(); c.close()
    return {"ok": True}

# Endpoint público para frontend
@app.get("/ofertas")
def get_ofertas_for_vendor(vendedor: Optional[str] = None, perfil: Optional[str] = None, cliente: Optional[str] = None):
    from datetime import date
    hoy = date.today().isoformat()
    c = _admin_db()
    vend_up = vendedor.upper() if vendedor else None
    perf_up = perfil.upper() if perfil else None
    cli_up  = cliente.strip() if cliente else None
    # Una oferta aplica si:
    #   - No tiene NINGUNA restricción de vendor/profile (offer_clients NO cuenta acá —
    #     ver nota abajo), O
    #   - El vendedor está en offer_vendors, O
    #   - El perfil del vendedor está en offer_profiles, O
    #   - El cliente de la operación está en offer_clients (alcance por cliente prevalece:
    #     cualquier corredor que atienda a ese cliente puede aplicarla, aunque la oferta
    #     tenga cargados otros vendedores/perfiles distintos al que consulta).
    #
    # IMPORTANTE: offer_clients es SIEMPRE complementario, nunca restrictivo. Antes,
    # una oferta sin ningún vendor/perfil cargado (o sea "para todos") dejaba de ser
    # "para todos" en cuanto Reactivación le agregaba filas de recupero de cartera a
    # offer_clients — porque la condición de "sin restricción" exigía ADEMÁS que
    # offer_clients estuviera vacía. Eso rompía la oferta base (ej. escalón Hasbro-
    # Microelectronics) para CUALQUIER cliente que no fuera uno de los pocos cargados
    # para el recupero, salvo que el vendedor matcheara offer_vendors/offer_profiles.
    # offer_clients solo debe SUMAR acceso (a esos clientes puntuales, para cualquier
    # corredor), nunca sacarle el alcance general a la oferta.
    base_cond = """
        SELECT DISTINCT o.* FROM offers o
        WHERE o.activo=1
          AND (o.fecha_desde IS NULL OR o.fecha_desde <= ?)
          AND (o.fecha_hasta IS NULL OR o.fecha_hasta >= ?)
          AND (
            (NOT EXISTS (SELECT 1 FROM offer_vendors ov WHERE ov.offer_id=o.id AND ov.codigousuario IS NOT NULL)
             AND NOT EXISTS (SELECT 1 FROM offer_profiles op WHERE op.offer_id=o.id))
    """
    params = [hoy, hoy]
    if vend_up:
        base_cond += " OR EXISTS (SELECT 1 FROM offer_vendors ov WHERE ov.offer_id=o.id AND ov.codigousuario=?)"
        params.append(vend_up)
    if perf_up:
        base_cond += " OR EXISTS (SELECT 1 FROM offer_profiles op WHERE op.offer_id=o.id AND op.perfil_codigo=?)"
        params.append(perf_up)
    if cli_up:
        base_cond += " OR EXISTS (SELECT 1 FROM offer_clients oc WHERE oc.offer_id=o.id AND oc.codigocliente=?)"
        params.append(cli_up)
    base_cond += ")"
    offers = c.execute(base_cond, params).fetchall()
    if not offers:
        c.close()
        return []

    # ── Recolectar condiciones de texto para resolver en UNA sola conexión Firebird ──
    # (Evita abrir N conexiones Firebird — una por cada condición de texto — que
    # causaba demoras de 5-15 s en cargarOfertas() del frontend.)
    ids = [o['id'] for o in offers]
    ph  = ','.join('?' * len(ids))
    all_conds_rows = c.execute(
        f"SELECT offer_id, condicion_comercial FROM offer_conditions WHERE offer_id IN ({ph})", ids
    ).fetchall()

    # Separar las que ya son código numérico de las que son texto libre
    text_conds = set()
    for row in all_conds_rows:
        val = str(row[1] or '').strip()
        if val and not val.isdigit():
            text_conds.add(val)

    # Una sola conexión Firebird para resolver TODOS los textos pendientes
    text_to_code = {}
    if text_conds:
        try:
            fb = conn('WIN1252')
            cur_fb = fb.cursor()
            for txt in text_conds:
                cur_fb.execute(
                    'SELECT FIRST 1 CODIGOMULTIPLAZO FROM "MULTIPLAZOS" '
                    'WHERE UPPER(TRIM(DESCRIPCION)) = UPPER(TRIM(?))', (txt,))
                row = cur_fb.fetchone()
                text_to_code[txt] = str(row[0]).strip() if row else txt
            fb.close()
        except Exception:
            pass  # si Firebird falla, se usa el texto tal cual

    # Agrupar condiciones por offer_id con resolución ya lista
    conds_by_offer = {}
    for row in all_conds_rows:
        oid = row[0]
        val = str(row[1] or '').strip()
        resolved_val = val if val.isdigit() else text_to_code.get(val, val)
        conds_by_offer.setdefault(oid, []).append(resolved_val)

    # ── Batch queries SQLite para el resto de tablas relacionadas ──────────────
    fin_by_offer  = {}
    for row in c.execute(f"SELECT offer_id, porcentaje, orden FROM offer_financial_details WHERE offer_id IN ({ph}) ORDER BY orden", ids):
        fin_by_offer.setdefault(row[0], []).append({'porcentaje': row[1], 'orden': row[2]})

    prod_by_offer = {}
    for row in c.execute(f"SELECT * FROM offer_product_details WHERE offer_id IN ({ph})", ids):
        prod_by_offer.setdefault(row['offer_id'], []).append(dict(row))

    cat_by_offer  = {}
    for row in c.execute(f"SELECT offer_id, nivel, valor FROM offer_category_filters WHERE offer_id IN ({ph})", ids):
        cat_by_offer.setdefault(row[0], []).append({'nivel': row[1], 'valor': row[2]})

    disc_by_offer = {}
    for row in c.execute(f"SELECT offer_id, nivel, valor FROM offer_discount_filters WHERE offer_id IN ({ph})", ids):
        disc_by_offer.setdefault(row[0], []).append({'nivel': row[1], 'valor': row[2]})

    esc_by_offer  = {}
    for row in c.execute(f"SELECT offer_id, min_combos, descuento_pct FROM offer_combo_escalones WHERE offer_id IN ({ph}) ORDER BY min_combos", ids):
        esc_by_offer.setdefault(row[0], []).append({'min_combos': row[1], 'descuento_pct': row[2]})
    amt_esc_by_offer = {}
    for row in c.execute(f"SELECT offer_id, monto_minimo, descuento_pct, condicion_comercial FROM offer_amount_escalones WHERE offer_id IN ({ph}) ORDER BY monto_minimo", ids):
        amt_esc_by_offer.setdefault(row[0], []).append({'monto_minimo': row[1], 'descuento_pct': row[2], 'condicion_comercial': row[3]})

    # "Recupero de cartera": % adicional de Reactivación, cargado por cliente
    # puntual dentro de offer_clients (ver _reactivacion_vincular_cliente_recupero).
    # Se inyecta MÁS ABAJO únicamente en el escalón ya alcanzado de esta oferta para
    # ESTE cliente puntual — nunca destraba el escalón por sí solo. monto_minimo_extra
    # permite exigir, además, un neto propio del recupero (independiente del escalón
    # base — si es menor o igual, no cambia nada).
    extra_by_offer = {}
    if cli_up:
        for row in c.execute(
            f"SELECT offer_id, descuento_extra_pct, vencimiento_extra, monto_minimo_extra, tipo_cartera, condicion_comercial_extra FROM offer_clients WHERE offer_id IN ({ph}) AND codigocliente=?",
            ids + [cli_up]):
            pct_extra  = row[1] or 0
            vto_extra  = (row[2] or '').strip()
            tipo_extra = (row[4] or 'descuento').strip()
            cond_extra = (row[5] or '').strip()
            vigente = (not vto_extra or vto_extra >= hoy)
            # La recompensa puede ser % de descuento, condición comercial distinta, o
            # ambas a la vez — alcanza con que cualquiera de las dos esté cargada,
            # sin importar la etiqueta tipo_extra (solo informativa).
            if vigente and (pct_extra > 0 or cond_extra):
                extra_by_offer[row[0]] = {'pct': pct_extra, 'monto_minimo': row[3] or 0,
                                           'tipo': tipo_extra, 'condicion': cond_extra}

    result = []
    for o in offers:
        od = dict(o)
        oid = od['id']
        od['financial_details'] = fin_by_offer.get(oid, [])
        od['product_details']   = prod_by_offer.get(oid, [])
        od['conditions']        = conds_by_offer.get(oid, [])
        od['category_filters']  = cat_by_offer.get(oid, [])
        od['discount_filters']  = disc_by_offer.get(oid, [])
        od['combo_escalones']   = esc_by_offer.get(oid, [])
        od['amount_escalones']  = amt_esc_by_offer.get(oid, [])
        for k, d in [('deposito',''),('tipo_financiero','descuento_total'),('monto_minimo',0),('cupo',0),('usos',0),('tope_bonificacion_pesos',0),('bonificado_acumulado_pesos',0),('acumulable',0)]:
            if k not in od: od[k] = d
        # Parsear financial_escalones de JSON string a lista
        fe_raw = od.get('financial_escalones')
        od['financial_escalones'] = json.loads(fe_raw) if fe_raw else []
        extra = extra_by_offer.get(oid)
        if od.get('tipo') == 'cartera':
            # Oferta del panel "Descuento por Cartera": no tiene alcance de
            # categoría/vendedor/perfil ni financial_escalones propios en la DB — el
            # único tramo que existe es el cargado para ESTE cliente puntual en
            # offer_clients. Sin cliente coincidente, la oferta queda inerte (sin
            # escalones, _checkDescuentoMonto del frontend la ignora).
            if extra:
                # % y condición no son excluyentes — se aplican los dos si están cargados.
                esc = {'monto_minimo': extra['monto_minimo'], 'porcentajes': [], 'condicion_comercial': None}
                if extra['pct'] > 0:
                    esc['porcentajes'] = [extra['pct']]
                    esc['_recupero_pct'] = extra['pct']
                if extra['condicion']:
                    esc['condicion_comercial'] = extra['condicion']
                    esc['_recupero_condicion'] = extra['condicion']
                od['financial_escalones'] = [esc]
            else:
                od['financial_escalones'] = []
        elif extra and od['financial_escalones']:
            # El recupero exige, como mínimo, el escalón de mayor monto_minimo que ya
            # tiene la oferta — el pedido tiene que alcanzarlo primero. Si además se
            # cargó un monto_minimo_extra propio (mayor al del escalón base), ese pasa
            # a ser el umbral real: el % extra recién se otorga a partir de ahí, pero
            # sumado al % base (no lo reemplaza) para no perder el beneficio ya ganado.
            top = max(od['financial_escalones'], key=lambda e: e.get('monto_minimo', 0))
            umbral = max(top.get('monto_minimo', 0), extra['monto_minimo'])
            # % y condición no son excluyentes — se aplican los dos si están cargados,
            # sumando el % al del escalón base (nunca lo reemplaza) y reemplazando la
            # condición del escalón alcanzado si se cargó una distinta. _recupero_condicion
            # (análogo a _recupero_pct) le avisa al frontend que la condición vino del
            # recupero de cartera, no del escalón base, para informarlo con su propio cartel.
            if umbral <= top.get('monto_minimo', 0):
                if extra['condicion']:
                    top['condicion_comercial'] = extra['condicion']
                    top['_recupero_condicion'] = extra['condicion']
                if extra['pct'] > 0:
                    top['porcentajes'] = list(top.get('porcentajes') or []) + [extra['pct']]
                    top['_recupero_pct'] = extra['pct']
            else:
                nuevo_escalon = dict(top)
                nuevo_escalon['monto_minimo'] = umbral
                if extra['condicion']:
                    nuevo_escalon['condicion_comercial'] = extra['condicion']
                    nuevo_escalon['_recupero_condicion'] = extra['condicion']
                if extra['pct'] > 0:
                    nuevo_escalon['porcentajes'] = list(top.get('porcentajes') or []) + [extra['pct']]
                    nuevo_escalon['_recupero_pct'] = extra['pct']
                od['financial_escalones'] = od['financial_escalones'] + [nuevo_escalon]
        result.append(od)
    c.close()
    return result

_BANCOS = [
    ("Santander Rio",  "0720131420000001149872", "131",    "11498/7"),
    ("Provincia",      "0140004501400404115211", "4004",   "041152/1"),
    ("HSBC",           "1500607500060732055732", "607",    "607-3-205573"),
    ("Galicia",        "0070154520000005006724", "154",    "5006-7-154/2"),
]
_MP_EMAIL = "marketing@microbellsa.com.ar"
_MP_CVU   = "0000003100004756934965"

def _fecha_ddmmyyyy(s):
    """Convierte una fecha ISO (YYYY-MM-DD, con o sin hora) a dd/mm/yyyy — formato
    español latinoamericano. Si no matchea el patrón, devuelve el string tal cual
    en vez de romper (puede venir ya en otro formato)."""
    s = (s or '').strip()
    if not s:
        return ''
    try:
        return datetime.strptime(s[:10], '%Y-%m-%d').strftime('%d/%m/%Y')
    except Exception:
        return s

def _fmt(v):
    """Formato moneda argentina: $ 1.234,56"""
    try:
        v = float(v)
    except (TypeError, ValueError):
        return "$ 0,00"
    neg = v < 0
    v = abs(v)
    ent = int(v)
    dec = round((v - ent) * 100)
    s = f"{ent:,}".replace(",", ".")
    return f"{'- ' if neg else ''}$ {s},{dec:02d}"

def _d(v):
    """Convierte datetime o string a dd/mm/yyyy"""
    if v is None:
        return ""
    s = str(v)
    try:
        from datetime import datetime
        if "T" in s:
            s = s.split("T")[0]
        elif " " in s:
            s = s.split(" ")[0]
        parts = s.split("-")
        if len(parts) == 3:
            return f"{parts[2]}/{parts[1]}/{parts[0]}"
    except Exception:
        pass
    return s

@app.get("/panel")
def panel_redirect():
    return RedirectResponse(url="/admin", status_code=302)

@app.get("/", response_class=HTMLResponse)
def index():
    with open(FRONTEND_PATH, encoding="utf-8") as f:
        return f.read()

@app.get("/logo")
def get_logo():
    from fastapi.responses import FileResponse as FR
    if os.path.exists(LOGO_PATH):
        return FR(LOGO_PATH, media_type="image/png")
    return HTMLResponse("", status_code=404)

@app.get("/manifest.json")
def get_manifest():
    from fastapi.responses import FileResponse as FR
    path = os.path.join(os.path.dirname(__file__), "manifest.json")
    if os.path.exists(path):
        return FR(path, media_type="application/manifest+json",
                  headers={"Cache-Control": "public, max-age=86400"})
    return HTMLResponse("", status_code=404)

@app.get("/sw.js")
def get_sw():
    from fastapi.responses import FileResponse as FR
    path = os.path.join(os.path.dirname(__file__), "sw.js")
    if os.path.exists(path):
        return FR(path, media_type="application/javascript",
                  headers={"Cache-Control": "no-cache, no-store, must-revalidate",
                           "Service-Worker-Allowed": "/"})
    return HTMLResponse("", status_code=404)

@app.get("/favicon.ico")
def get_favicon():
    from fastapi.responses import FileResponse as FR
    if os.path.exists(FAVICON_PATH):
        return FR(FAVICON_PATH, media_type="image/x-icon",
                  headers={"Cache-Control": "public, max-age=86400"})
    return HTMLResponse("", status_code=404)

@app.get("/icons/{filename}")
def get_icon(filename: str):
    from fastapi.responses import FileResponse as FR
    path = os.path.join(os.path.dirname(__file__), filename)
    if os.path.exists(path) and filename.lower().endswith('.png'):
        return FR(path, media_type="image/png",
                  headers={"Cache-Control": "public, max-age=86400"})
    return HTMLResponse("", status_code=404)


# ─── Stock ─────────────────────────────────────────────────────────────────────
@app.get("/stock/cache-ts")
def stock_cache_ts():
    """Devuelve la antigüedad en segundos de cada depósito en cache."""
    now = time.time()
    result = {}
    with _fma_cache_lock:
        for dep, (ts, _) in _fma_cache.items():
            result[dep] = round(now - ts)
    return {"cache": result, "ttl": _FMA_CACHE_TTL, "ts": now}

@app.get("/stock")
def get_stock(
    buscar: Optional[str] = None,
    gruposuperrubro: Optional[str] = None,
    superrubro: Optional[str] = None,
    rubro: Optional[str] = None,
    marca: Optional[str] = None,
    deposito: Optional[str] = None,
    limit: int = Query(100, le=300),
    offset: int = 0,
    _user=Depends(get_current_user)
):
    _dep_fma = deposito if deposito else '001,003'
    dep_lista = [d.strip() for d in _dep_fma.split(',') if d.strip()] or ['001', '003']
    try:
        pagina, _total, cambio_usd = _search_stock_cache(
            buscar=buscar, gruposuperrubro=gruposuperrubro, superrubro=superrubro,
            rubro=rubro, marca=marca, dep_lista=dep_lista, limit=limit, offset=offset
        )
        # Asegurar TODOS los depósitos activos en caché para los remanente_XXX del
        # frontend (antes solo se aseguraban 6 depósitos fijos — ver _deps_activos).
        _deps_all = _deps_activos()
        _fma_stock_parallel(_deps_all)
        with _fma_cache_lock:
            all_rem = {d: (_fma_cache.get(d) or (0, {}))[1] for d in _deps_all}

        resultado = []
        for art, rem_dep, rem_total in pagina:
            factor = cambio_usd if art['codigomoneda'] == 'DOLARES' else 1.0
            def conv(v): return _redondear_precio(v, factor)
            item = {
                "codigo":           art['codigo'],
                "codigoparticular": art['codigoparticular'],
                "descripcion":      art['descripcion'],
                "marca":            art['codigomarca'],
                "precio1":          conv(art['precio1']),
                "precio2":          conv(art['precio2']),
                "precio3":          conv(art['precio3']),
                "precio5":          conv(art['precio5']),
                "iva":              art['iva'],
                "unidad":           art['unidad'],
                "stock":            rem_total,
                "remanente":        rem_total,
                "rubro":            art['rubro'],
                "superrubro":       art['superrubro'],
                "gruposuperrubro":  art['gruposuperrubro'],
                "moneda":           art['codigomoneda'],
                "codigo_rubro":     art['codigo_rubro'],
                "codigo_superrubro": art['codigo_superrubro'],
                "codigo_gruposuperrubro": art['codigo_gruposuperrubro'],
            }
            # remanente_XXX para CADA depósito activo (antes solo 001/002/003/005/013/016
            # — cualquier depósito nuevo, ej. 017 SARANDI, quedaba sin su clave y el
            # frontend lo mostraba en 0 aunque tuviera stock real).
            for _d in _deps_all:
                item[f"remanente_{_d}"] = all_rem[_d].get(art['codigo'], 0)
            resultado.append(item)
        _apply_reservas(resultado, _get_reservas_activas(), rem_key='remanente')
        return resultado
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/buscar-articulos")
def buscar_articulos(q: str = Query("", min_length=2), db: str = Query("oficial"), deposito: Optional[str] = None, skip_stock: int = Query(0), solo_con_stock: int = Query(0)):
    from concurrent.futures import ThreadPoolExecutor
    import re as _re
    ALL_DEPS = ['001','002','003','005','013','016']
    deps_to_fetch = [d.strip() for d in deposito.split(',')] if deposito else ALL_DEPS

    # ── 1. Artículos — búsqueda de texto contra el catálogo en memoria (2026-07-31) ──
    # Antes esto era un CONTAINING en vivo contra Firebird en CADA letra escrita (sin
    # índice útil para substring → varios segundos por búsqueda, con la conexión yendo
    # y viniendo cada vez — ver reporte de Eduardo). El catálogo (_get_catalog) ya vive
    # en memoria con refresco automático cada CATALOG_CACHE_TTL (30 min por default,
    # con pre-calentamiento en background — ver _prewarm_catalog) para otras pantallas;
    # acá se reutiliza para que la búsqueda de texto sea instantánea, sin tocar
    # Firebird en absoluto. Misma semántica que el SQL anterior: detección de código
    # puro, AND de palabras sueltas con tolerancia a tildes/plural, FIRST 30, mismo
    # orden — _cn/_dn son los campos normalizados (sin tildes, mayúsculas) precomputados
    # en _load_catalog.
    catalog, cambio_usd = _get_catalog()
    _es_codigo = bool(_re.match(r'^[A-Za-z0-9\-]{2,15}$', q) and ' ' not in q)
    q_norm = _sin_tildes(q).upper()

    if _es_codigo:
        matched = [a for a in catalog.values() if q_norm in a['_cn'] or q_norm in a['_dn']]
        matched.sort(key=lambda a: (0 if a['_cn'].startswith(q_norm) else 1, a['codigoparticular'], a['descripcion']))
    else:
        # Búsqueda por palabras sueltas: si hay más de una palabra, exige que TODAS
        # aparezcan en DESCRIPCION (en cualquier orden), en vez de exigir la frase
        # exacta contigua. Tolera reconocimiento de voz que cambia el orden o los
        # conectores (ej. "pistola para dardos" vs "pistola de dardos").
        _words = [w for w in q.split() if w]
        if len(_words) > 1:
            # Normaliza tildes y, en palabras de 4+ letras terminadas en "s", también
            # prueba sin esa "s" final — tolera diferencias de singular/plural y
            # tildes que introduce el reconocimiento de voz (ej. "baterías" dicho por
            # voz vs "BATERIA" cargado en Flexxus).
            _word_norms = []
            for _w in _words:
                _wn = _sin_tildes(_w).upper()
                _wn_sin_s = _wn[:-1] if (len(_wn) >= 4 and _wn[-1] == 'S') else None
                _word_norms.append((_wn, _wn_sin_s))
            def _todas_las_palabras(a):
                return all((wn in a['_dn']) or (wn2 and wn2 in a['_dn']) for wn, wn2 in _word_norms)
            matched = [a for a in catalog.values() if _todas_las_palabras(a) or q_norm in a['_cn']]
        else:
            matched = [a for a in catalog.values() if q_norm in a['_dn'] or q_norm in a['_cn']]
        matched.sort(key=lambda a: (a['descripcion'], a['codigoparticular']))
    arts_matched = matched[:30]

    # 2. FMA_STOCK en paralelo - una conexion por deposito
    def _fetch_dep(dep):
        try:
            cx = conn('WIN1252', db=DATABASE)
            cu = cx.cursor()
            cu.execute(f"SELECT ID_ARTICULO, STOCKREMANENTE FROM \"FMA_STOCK\"(NULL, NULL, '{dep}', 1, 1)")
            result = {row[0]: float(row[1] or 0) for row in cu.fetchall()}
            cx.close()
            return dep, result
        except Exception:
            return dep, {}

    rem_maps = {d: {} for d in ALL_DEPS}
    if arts_matched and not skip_stock:
        with ThreadPoolExecutor(max_workers=len(deps_to_fetch)) as ex:
            for dep, data in ex.map(_fetch_dep, deps_to_fetch):
                rem_maps[dep] = data

    # Filtro "solo con stock" para el autocomplete del vendedor — reusa el caché FMA
    # compartido (TTL, mismo que usa /stock/batch) en vez de _fetch_dep de arriba, para
    # no reintroducir conexiones sin cachear por cada tecla escrita. Se queda afuera un
    # artículo solo si no tiene remanente en NINGUNO de los depósitos consultados
    # (deps_to_fetch = los asignados al vendedor, o todos si no tiene restricción) —
    # salvo que tenga PERMITESTOCKNEGATIVO activo, que siempre se muestra.
    rem_stock_filtro = None
    if arts_matched and solo_con_stock:
        rem_stock_filtro = _fma_stock_parallel(deps_to_fetch)

    resultado = []
    for art in arts_matched:
        cod = art['codigo']
        moneda = art['codigomoneda']
        factor = cambio_usd if moneda == 'DOLARES' else 1.0
        precio = _redondear_precio(art['precio1'], factor)
        dto_max = float(art['dtomaximo1_raw']) if art['dtomaximo1_raw'] is not None else None
        aplica_bonif = str(art['aplicableabonificacion_raw'] or '0').strip() == '1'
        if not aplica_bonif:
            dto_max = None
        precio5 = _redondear_precio(art['precio5'], factor)
        _coef = art['rubro_coeficiente'] if art['coeficiente_segun_rubro'] else art['coeficiente']
        permite_neg = str(art['permitestocknegativo_raw'] or '0').strip() == '1'
        if rem_stock_filtro is not None and not permite_neg:
            total_rem = sum(rem_stock_filtro.get(dep, {}).get(cod, 0.0) for dep in deps_to_fetch)
            if total_rem <= 0:
                continue
        item = {"codigo": cod, "codigoparticular": art['codigoparticular'],
                "descripcion": art['descripcion'], "precio": precio, "precio5": precio5,
                "iva": round(_coef * 21, 2),
                "dto_max": dto_max,
                "permite_stock_negativo": permite_neg}
        for dep in ALL_DEPS:
            item[f'rem{dep}'] = rem_maps.get(dep, {}).get(cod, 0)
        resultado.append(item)
    return resultado

# ─── Helpers stock export ──────────────────────────────────────────────────────
def _fetch_stock_data(buscar=None, gruposuperrubro=None, superrubro=None, rubro=None, marca=None):
    """Devuelve todos los artículos con remanente > 0 según filtros, sin límite."""
    # ── 1. Cotización USD
    cambio_usd = 1.0
    try:
        cx = conn()
        cu = cx.cursor()
        cu.execute('SELECT CAMBIO FROM "MONEDAS" WHERE CODIGOMONEDA = ?', ('DOLARES',))
        rm = cu.fetchone()
        cambio_usd = float(rm[0]) if rm else 1.0
        cx.close()
    except Exception:
        pass

    # ── 2. Query principal (001+003 combinados)
    wheres = ["a.ACTIVO = '1'", "(g.DESCRIPCION IS NULL OR UPPER(g.DESCRIPCION) NOT IN ('TERCEROS','SERVICIOS'))"]
    params = []
    if buscar:
        buscar = _sanitizar_buscar(buscar)
        wheres.append("(UPPER(a.DESCRIPCION) CONTAINING UPPER(?) OR a.CODIGOPARTICULAR CONTAINING ?)")
        params += [buscar, buscar]
    if rubro:
        wheres.append("a.CODIGORUBRO = ?")
        params.append(rubro)
    if superrubro:
        wheres.append("r.CODIGOSUPERRUBRO = ?")
        params.append(superrubro)
    if gruposuperrubro:
        wheres.append("sr.CODIGOGRUPOSUPERRUBRO = ?")
        params.append(gruposuperrubro)
    if marca:
        wheres.append("a.CODIGOMARCA = ?")
        params.append(marca)
    where_sql = " AND ".join(wheres)

    sql = f"""
        SELECT s.ID_ARTICULO, a.CODIGOPARTICULAR, a.DESCRIPCION,
               a.PRECIOLISTA1, a.COEFICIENTESEGUNRUBRO, a.CODIGOUNIDADMEDIDA,
               s.STOCKREAL, s.STOCKREMANENTE, a.CODIGOMONEDA,
               r.DESCRIPCION, sr.DESCRIPCION, g.DESCRIPCION,
               a.COEFICIENTE, r.COEFICIENTE
        FROM "FMA_STOCK"(NULL, NULL, '001,003', 1, 1) s
        JOIN "ARTICULOS" a ON a.CODIGOARTICULO = s.ID_ARTICULO
        LEFT JOIN "RUBROS" r ON r.CODIGORUBRO = a.CODIGORUBRO
        LEFT JOIN "SUPERRUBROS" sr ON sr.CODIGOSUPERRUBRO = r.CODIGOSUPERRUBRO
        LEFT JOIN "GRUPOSUPERRUBROS" g ON g.CODIGOGRUPOSUPERRUBRO = sr.CODIGOGRUPOSUPERRUBRO
        WHERE {where_sql} AND s.STOCKREMANENTE > 0
        ORDER BY a.DESCRIPCION
    """
    c1 = conn()
    cur1 = c1.cursor()
    cur1.execute(sql, params)
    rows = cur1.fetchall()
    c1.close()

    # ── 3. Remanente por depósito (paralelo para mayor velocidad)
    from concurrent.futures import ThreadPoolExecutor

    def _fetch_rem(dep):
        try:
            cx = conn()
            cu = cx.cursor()
            cu.execute(f'SELECT ID_ARTICULO, STOCKREMANENTE FROM "FMA_STOCK"(NULL, NULL, \'{dep}\', 1, 1)')
            result = {row[0]: float(row[1] or 0) for row in cu.fetchall()}
            cx.close()
            return result
        except Exception:
            return {}

    with ThreadPoolExecutor(max_workers=2) as ex:
        f001 = ex.submit(_fetch_rem, '001')
        f003 = ex.submit(_fetch_rem, '003')
        rem_001_map = f001.result()
        rem_003_map = f003.result()

    # ── 4. Armar resultado
    result = []
    for r in rows:
        moneda = (r[8] or '').strip().upper()
        factor = cambio_usd if moneda == 'DOLARES' else 1.0
        precio = _redondear_precio(r[3], factor)
        result.append({
            "codigo":           r[1] or r[0],
            "descripcion":      (r[2] or '').strip(),
            "stock":            float(r[6] or 0),
            "rem_001":          rem_001_map.get(r[0], 0),
            "rem_003":          rem_003_map.get(r[0], 0),
            "rem_total":        float(r[7] or 0),
            "precio":           precio,
            "iva":              round((float(r[13] or 0) if str(r[4] or '0').strip() == '1' else float(r[12] or 0)) * 21, 2),
            "rubro":            (r[9] or '').strip(),
            "superrubro":       (r[10] or '').strip(),
            "gruposuperrubro":  (r[11] or '').strip(),
        })
    return result


@app.get("/stock/exportar-excel")
def exportar_stock_excel(
    buscar: Optional[str] = None,
    gruposuperrubro: Optional[str] = None,
    superrubro: Optional[str] = None,
    rubro: Optional[str] = None,
    marca: Optional[str] = None,
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        rows = _fetch_stock_data(buscar, gruposuperrubro, superrubro, rubro, marca)
        # Aplicar reservas (igual que en /stock)
        for r in rows:
            r['remanente_001'] = r['rem_001']
            r['remanente_003'] = r['rem_003']
            r['remanente'] = r['rem_total']
            r['codigoparticular'] = r['codigo']
        _apply_reservas(rows, _get_reservas_activas(), rem_key='remanente')
        for r in rows:
            r['rem_001'] = r['remanente_001']
            r['rem_003'] = r['remanente_003']
            r['rem_total'] = r['remanente_001'] + r['remanente_003']

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock Remanente"

        hdr_fill = PatternFill("solid", fgColor="1A56DB")
        hdr_font = Font(bold=True, color="FFFFFFFF", size=10)
        alt_fill = PatternFill("solid", fgColor="FFEFF6FF")
        center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
        right_al = Alignment(horizontal="right",  vertical="center")
        left_al  = Alignment(horizontal="left",   vertical="center")

        headers    = ["Gr. Super Rubro", "Super Rubro", "Rubro",
                      "Código", "Descripción",
                      "Rem. VAC-LOG (001)", "Rem. Pacheco (003)", "Rem. Total",
                      "Precio s/IVA", "IVA %"]
        col_widths = [22, 22, 18, 14, 52, 18, 18, 14, 18, 8]

        ws.row_dimensions[1].height = 24
        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center
            ws.column_dimensions[get_column_letter(ci)].width = w

        for ri, row in enumerate(rows, 2):
            vals = [row["gruposuperrubro"], row["superrubro"], row["rubro"],
                    row["codigo"], row["descripcion"],
                    int(row["rem_001"]), int(row["rem_003"]), int(row["rem_total"]),
                    row["precio"], float(row["iva"] or 0)]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.alignment = left_al if ci <= 5 else right_al
                if ci == 10:
                    cell.number_format = '0.00'
                if ri % 2 == 0:
                    cell.fill = alt_fill

        ws.freeze_panes = "A2"
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=stock_remanente.xlsx"})
    except Exception as e:
        import traceback
        raise HTTPException(status_code=500, detail=traceback.format_exc())


@app.get("/stock/exportar-pdf")
def exportar_stock_pdf(
    buscar: Optional[str] = None,
    gruposuperrubro: Optional[str] = None,
    superrubro: Optional[str] = None,
    rubro: Optional[str] = None,
    marca: Optional[str] = None,
):
    from datetime import datetime
    from reportlab.lib.pagesizes import landscape
    rows = _fetch_stock_data(buscar, gruposuperrubro, superrubro, rubro, marca)
    # Aplicar reservas (igual que en /stock)
    for r in rows:
        r['remanente_001'] = r['rem_001']
        r['remanente_003'] = r['rem_003']
        r['remanente'] = r['rem_total']
        r['codigoparticular'] = r['codigo']
    _apply_reservas(rows, _get_reservas_activas(), rem_key='remanente')
    for r in rows:
        r['rem_001'] = r['remanente_001']
        r['rem_003'] = r['remanente_003']
        r['rem_total'] = r['remanente_001'] + r['remanente_003']

    # Datos empresa
    razon_soc = cuit_emp = dir_emp = tel_emp = email_emp = ''
    try:
        cp = conn()
        ccp = cp.cursor()
        ccp.execute('SELECT RAZONSOCIAL, CUIT, DIRECCION, TELEFONO, EMAIL FROM "PARAMETROS" WHERE CODIGOPARAMETRO = 1')
        rp = ccp.fetchone()
        cp.close()
        if rp:
            razon_soc, cuit_emp, dir_emp, tel_emp, email_emp = [(v or '').strip() for v in rp]
    except Exception:
        pass

    buf = BytesIO()
    PAGE_W, PAGE_H = landscape(A4)
    mg = 14 * mm

    s_title = ParagraphStyle('t', fontSize=13, leading=16, fontName='Helvetica-Bold',
                              textColor=colors.HexColor('#1a56db'))
    s_sub   = ParagraphStyle('s', fontSize=8,  leading=10, fontName='Helvetica',
                              textColor=colors.HexColor('#6b7280'))
    s_ft    = ParagraphStyle('f', fontSize=6.5, leading=9, fontName='Helvetica',
                              alignment=TA_CENTER, textColor=colors.HexColor('#6b7280'))
    s_hdr   = ParagraphStyle('h', fontSize=7.5, leading=9, fontName='Helvetica-Bold',
                              alignment=TA_CENTER, textColor=colors.white)
    s_cell  = ParagraphStyle('c', fontSize=7.5, leading=9, fontName='Helvetica')
    s_cell_r= ParagraphStyle('cr', fontSize=7.5, leading=9, fontName='Helvetica', alignment=TA_RIGHT)

    usable_w = PAGE_W - 2 * mg
    footer_txt = f'{razon_soc}  ·  CUIT {cuit_emp}  ·  {dir_emp}  ·  Tel {tel_emp}  ·  {email_emp}'

    def _on_page(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor('#e5e7eb'))
        canvas.setLineWidth(0.5)
        canvas.line(mg, 10*mm, PAGE_W - mg, 10*mm)
        canvas.setFont('Helvetica', 6.5)
        canvas.setFillColor(colors.HexColor('#6b7280'))
        canvas.drawCentredString(PAGE_W / 2, 7*mm, footer_txt)
        canvas.drawRightString(PAGE_W - mg, 7*mm, f"Pág. {doc.page}")
        canvas.restoreState()

    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=mg, rightMargin=mg,
                            topMargin=mg, bottomMargin=18*mm)

    story = []

    # Logo + título
    logo_row = []
    if os.path.exists(LOGO_PATH):
        logo_row.append(Image(LOGO_PATH, width=38*mm, height=12*mm))
    else:
        logo_row.append(Paragraph('', s_sub))

    # Obtener descripcion de marca si aplica
    marca_desc = None
    if marca:
        try:
            cm = conn(); cum = cm.cursor()
            cum.execute('SELECT DESCRIPCION FROM "MARCAS" WHERE CODIGOMARCA = ?', (marca,))
            rm2 = cum.fetchone()
            cm.close()
            marca_desc = (rm2[0] or '').strip() if rm2 else marca
        except Exception:
            marca_desc = marca

    filtro_txt = " · ".join(filter(None, [
        f"Buscar: {buscar}" if buscar else None,
        f"GSR: {gruposuperrubro}" if gruposuperrubro else None,
        f"SR: {superrubro}" if superrubro else None,
        f"Rubro: {rubro}" if rubro else None,
        f"Marca: {marca_desc}" if marca_desc else None,
        "Sin filtro" if not any([buscar, gruposuperrubro, superrubro, rubro, marca]) else None,
    ]))
    titulo_cell = [
        Paragraph("Stock Remanente", s_title),
        Paragraph(f"{filtro_txt}  —  {datetime.now().strftime('%d/%m/%Y %H:%M')}  —  {len(rows)} artículos", s_sub),
    ]
    t_hdr = Table([[logo_row[0], titulo_cell]], colWidths=[42*mm, usable_w - 42*mm])
    t_hdr.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN',  (1,0), (1,0),  'RIGHT'),
    ]))
    story.append(t_hdr)
    story.append(Spacer(1, 4*mm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#1a56db')))
    story.append(Spacer(1, 3*mm))

    # Tabla de datos
    HDR_BG = colors.HexColor('#1a56db')
    ALT_BG = colors.HexColor('#eff6ff')
    COL_001 = colors.HexColor('#2563eb')
    COL_003 = colors.HexColor('#7c3aed')
    COL_TOT = colors.HexColor('#059669')

    # Anchos de columna — 10 cols llenando usable_w (landscape A4 ≈ 269mm)
    cw = [28*mm, 28*mm, 24*mm, 16*mm, 90*mm, 18*mm, 18*mm, 16*mm, 22*mm, 10*mm]
    # suma: 28+28+24+16+90+18+18+16+22+10 = 270mm ≈ usable_w

    data = [[
        Paragraph('Gr. Super\nRubro',   s_hdr),
        Paragraph('Super\nRubro',       s_hdr),
        Paragraph('Rubro',              s_hdr),
        Paragraph('Código',             s_hdr),
        Paragraph('Descripción',        s_hdr),
        Paragraph('Rem.\nVAC-LOG',      s_hdr),
        Paragraph('Rem.\nPacheco',      s_hdr),
        Paragraph('Rem.\nTotal',        s_hdr),
        Paragraph('Precio s/IVA',       s_hdr),
        Paragraph('IVA',                s_hdr),
    ]]

    for row in rows:
        data.append([
            Paragraph(row['gruposuperrubro'] or '—', s_cell),
            Paragraph(row['superrubro'] or '—',      s_cell),
            Paragraph(row['rubro'] or '—',           s_cell),
            Paragraph(str(row['codigo']),            s_cell),
            Paragraph(row['descripcion'],            s_cell),
            Paragraph(f"{int(row['rem_001']):,}".replace(',', '.'),   s_cell_r),
            Paragraph(f"{int(row['rem_003']):,}".replace(',', '.'),   s_cell_r),
            Paragraph(f"{int(row['rem_total']):,}".replace(',', '.'), s_cell_r),
            Paragraph('$'+f"{row['precio']:,.2f}".replace(',','X').replace('.',',').replace('X','.'), s_cell_r),
            Paragraph((f"{row['iva']:.2f}".replace('.', ',')+'%') if row['iva'] else '—', s_cell_r),
        ])

    tbl = Table(data, colWidths=cw, repeatRows=1)
    style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), HDR_BG),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, ALT_BG]),
        ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor('#e5e7eb')),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        # Color columnas remanente (índices 5,6,7 sin Stock Total)
        ('TEXTCOLOR', (5,1), (5,-1), COL_001),
        ('TEXTCOLOR', (6,1), (6,-1), COL_003),
        ('TEXTCOLOR', (7,1), (7,-1), COL_TOT),
        ('FONTNAME',  (5,1), (7,-1), 'Helvetica-Bold'),
    ])
    tbl.setStyle(style)
    story.append(tbl)

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": "inline; filename=stock_remanente.pdf"})


@app.get("/stock/batch")
def get_articulos_batch(codigos: str = Query(..., description="Códigos separados por coma")):
    """Devuelve stock y datos de múltiples artículos en una sola llamada."""
    codes = [c.strip() for c in codigos.split(',') if c.strip()]
    if not codes:
        return []
    c = conn()
    cur = c.cursor()
    # Cotización USD
    try:
        cur.execute('SELECT CAMBIO FROM "MONEDAS" WHERE CODIGOMONEDA = ?', ('DOLARES',))
        rm = cur.fetchone()
        cambio_usd = float(rm[0]) if rm else 1.0
    except Exception:
        cambio_usd = 1.0
    _COLS = ('SELECT a.CODIGOARTICULO, a.DESCRIPCION, a.CODIGOMARCA, '
             'a.PRECIOLISTA1, a.PRECIOLISTA5, a.COEFICIENTESEGUNRUBRO, '
             'a.CODIGOMONEDA, a.CODIGOPARTICULAR, a.DTOMAXIMO1, a.APLICABLEABONIFICACION, '
             'a.PERMITESTOCKNEGATIVO, a.COEFICIENTE, r.COEFICIENTE '
             'FROM "ARTICULOS" a LEFT JOIN "RUBROS" r ON r.CODIGORUBRO = a.CODIGORUBRO WHERE ')
    placeholders = ','.join(['?' for _ in codes])
    # Buscar por CODIGOPARTICULAR
    cur.execute(_COLS + f'a.CODIGOPARTICULAR IN ({placeholders})', codes)
    rows_by_part = {str(r[7] or '').strip(): r for r in cur.fetchall()}
    # Los no encontrados, buscar por CODIGOARTICULO
    missing = [c2 for c2 in codes if c2 not in rows_by_part]
    rows_by_art = {}
    if missing:
        ph2 = ','.join(['?' for _ in missing])
        cur.execute(_COLS + f'a.CODIGOARTICULO IN ({ph2})', missing)
        rows_by_art = {str(r[0] or '').strip(): r for r in cur.fetchall()}
    c.close()
    # FMA stock (cache compartido, una llamada paralela) — todos los depósitos activos
    # (antes 6 fijos, dejaba afuera depósitos nuevos como 017 SARANDI).
    _DEPS = _deps_activos()
    rem_bulk = _fma_stock_parallel(_DEPS)
    # Resolver la clave de FMA_STOCK usando el catálogo compartido (_get_catalog),
    # que es el mismo que usa /stock (get_stock) y SÍ matchea correctamente contra
    # _fma_cache. La query propia de este endpoint (arriba, conn() sin WIN1252)
    # puede traer CODIGOARTICULO con una representación distinta (charset/padding)
    # que no coincide como clave de dict aunque "se vea" igual — por eso este
    # endpoint venía devolviendo 0 en todos los depósitos para ciertos artículos.
    _catalog, _ = _get_catalog()
    _cat_by_part = {}
    _cat_by_id_str = {}
    for _aid, _a in _catalog.items():
        _cat_by_id_str[str(_aid).strip()] = _aid
        if _a.get('codigoparticular'):
            _cat_by_part[_a['codigoparticular']] = _aid
    results = []
    for cod in codes:
        row = rows_by_part.get(cod) or rows_by_art.get(cod)
        if not row:
            continue
        moneda = (row[6] or '').strip().upper()
        factor = cambio_usd if moneda == 'DOLARES' else 1.0
        def conv(v, f=factor): return _redondear_precio(v, f)
        codigoarticulo = row[0]
        cod_key = str(codigoarticulo).strip()
        # Clave "buena" resuelta vía catálogo (prioridad), con fallback a la propia
        fma_key = _cat_by_part.get(cod) or _cat_by_id_str.get(cod) or _cat_by_id_str.get(cod_key) or codigoarticulo
        rem = {}
        _debug_dep = None
        for dep in _DEPS:
            d = rem_bulk[dep]
            if fma_key in d:
                rem[dep] = d[fma_key]
            elif codigoarticulo in d:
                rem[dep] = d[codigoarticulo]
            elif cod_key in d:
                rem[dep] = d[cod_key]
            else:
                rem[dep] = 0.0
                _debug_dep = dep  # se guarda el último depósito donde falló el match, para diagnóstico
        dto_max_raw = row[8]
        aplica_b = str(row[9] or '0').strip() == '1'
        dto_max = float(dto_max_raw) if (dto_max_raw is not None and aplica_b) else None
        permite_neg = str(row[10] or '0').strip() == '1'
        item = {
            "codigo": codigoarticulo,
            "codigoparticular": (row[7] or row[0] or '').strip(),
            "descripcion": row[1],
            "marca": (row[2] or '').strip(),
            "precio1": conv(row[3]), "precio5": conv(row[4]),
            "iva": round((float(row[12] or 0) if str(row[5] or '0').strip() == '1' else float(row[11] or 0)) * 21, 2),
            "moneda": row[6],
            "dto_max": dto_max,
            "permite_stock_negativo": permite_neg,
            "_input_cod": cod,
        }
        for _d in _DEPS:
            item[f"remanente_{_d}"] = rem[_d]
        if _debug_dep:
            # Diagnóstico temporal: el remanente de ESTE artículo no matcheó ninguna
            # clave de rem_bulk (ver /stock/batch) — se adjuntan tipo/valor de la clave
            # buscada y 5 claves de muestra de FMA_STOCK para ese depósito, así se puede
            # ver desde el propio JSON (Network tab) si es un problema de tipo de dato
            # (ej. CODIGOARTICULO con padding CHAR vs ID_ARTICULO ya trimeado).
            _sample = list(rem_bulk[_debug_dep].keys())[:5]
            item["_debug_fma"] = {
                "dep_fallo": _debug_dep,
                "codigoarticulo_valor": repr(codigoarticulo),
                "codigoarticulo_tipo": type(codigoarticulo).__name__,
                "fma_key_usada": repr(fma_key),
                "cod_key": cod_key,
                "muestra_claves_fma": [repr(k) for k in _sample],
                "muestra_tipos_fma": [type(k).__name__ for k in _sample],
            }
        results.append(item)
    return results


@app.get("/debug/stock-check")
def debug_stock_check(codigo: str):
    """DEBUG TEMPORAL: para un CODIGOPARTICULAR dado, muestra en una sola
    respuesta cómo lo resuelve el catálogo compartido (_get_catalog, usado por
    /stock — probado correcto), cómo lo resuelve la query propia de
    /stock/batch (conn() sin WIN1252 — sospechosa), y qué devuelve el caché
    FMA_STOCK para cada clave candidata en los depósitos 001/003. Objetivo:
    confirmar de una vez si el mismatch es de clave (tipo/padding) o si el
    valor en caché es legítimamente distinto al de /stock. Sin auth (mismo
    nivel que /stock/batch) — sacar una vez resuelto el bug real."""
    out = {"codigo_buscado": codigo}

    catalog, _ = _get_catalog()
    cat_hit = None
    for aid, a in catalog.items():
        if a.get('codigoparticular') == codigo or str(aid).strip() == codigo:
            cat_hit = (aid, a)
            break
    out["catalog"] = None
    if cat_hit:
        aid, a = cat_hit
        out["catalog"] = {
            "art_id": repr(aid), "art_id_tipo": type(aid).__name__,
            "codigoparticular": a.get('codigoparticular'),
        }

    c = conn()
    cur = c.cursor()
    cur.execute('SELECT CODIGOARTICULO, CODIGOPARTICULAR FROM "ARTICULOS" WHERE CODIGOPARTICULAR=?', (codigo,))
    row = cur.fetchone()
    c.close()
    out["query_batch"] = None
    if row:
        out["query_batch"] = {
            "codigoarticulo": repr(row[0]), "codigoarticulo_tipo": type(row[0]).__name__,
            "codigoparticular": repr(row[1]),
        }

    rem_bulk = _fma_stock_parallel(['001', '003'])
    candidatos = []
    if cat_hit: candidatos.append(('catalog_art_id', cat_hit[0]))
    if row: candidatos.append(('query_codigoarticulo', row[0]))
    out["fma"] = {}
    for dep in ('001', '003'):
        d = rem_bulk[dep]
        entry = {"total_claves_en_dep": len(d), "muestra_claves": [repr(k) for k in list(d.keys())[:3]]}
        for nombre, key in candidatos:
            entry[nombre] = {"clave": repr(key), "encontrada": key in d, "valor": d.get(key)}
        out["fma"][dep] = entry
    return out


@app.get("/conjunto/{codigo}/partes")
def get_conjunto_partes(codigo: str):
    """Dado CODIGOPARTICULAR de un Conjunto, devuelve sus artículos componentes."""
    c = conn()
    cur = c.cursor()
    cur.execute('SELECT CODIGOARTICULO FROM "ARTICULOS" WHERE CODIGOPARTICULAR = ?', (codigo,))
    row = cur.fetchone()
    if not row:
        cur.execute('SELECT CODIGOARTICULO FROM "ARTICULOS" WHERE CODIGOARTICULO = ?', (codigo,))
        row = cur.fetchone()
    if not row:
        c.close()
        raise HTTPException(404, "Conjunto no encontrado")
    cod_interno = row[0]
    cur.execute(
        'SELECT cj.CODIGOARTICULO, cj.CANTIDAD, a.DESCRIPCION, a.DESCRIPCIONADICIONAL, '
        'a.CODIGOPARTICULAR, a.PRECIOLISTA1, a.PRECIOLISTA2, a.PRECIOLISTA3, a.PRECIOLISTA5, '
        'a.COEFICIENTESEGUNRUBRO, a.CODIGOUNIDADMEDIDA, a.CODIGOMONEDA, a.PERMITESTOCKNEGATIVO, cj.COEFICIENTEPRECIO, '
        'a.COEFICIENTE, r.COEFICIENTE '
        'FROM "CONJUNTOS" cj '
        'JOIN "ARTICULOS" a ON a.CODIGOARTICULO = cj.CODIGOARTICULO '
        'LEFT JOIN "RUBROS" r ON r.CODIGORUBRO = a.CODIGORUBRO '
        'WHERE cj.CODIGOCONJUNTO = ? ORDER BY cj.LINEA',
        (cod_interno,)
    )
    partes_rows = cur.fetchall()
    try:
        cur.execute('SELECT CAMBIO FROM "MONEDAS" WHERE CODIGOMONEDA = ?', ('DOLARES',))
        rm = cur.fetchone()
        cambio_usd = float(rm[0]) if rm else 1.0
    except Exception:
        cambio_usd = 1.0
    c.close()
    partes = []
    for p in partes_rows:
        moneda = (p[11] or '').strip().upper()
        factor = cambio_usd if moneda == 'DOLARES' else 1.0
        def conv(v): return _redondear_precio(v, factor)
        partes.append({
            "codigo": (p[0] or '').strip(),
            "codigoparticular": (p[4] or p[0] or '').strip(),
            "cantidad": float(p[1] or 1),
            "descripcion": p[2],
            "descripcion_adicional": p[3],
            "precio1": conv(p[5]), "precio2": conv(p[6]),
            "precio3": conv(p[7]), "precio5": conv(p[8]),
            "iva": round((float(p[15] or 0) if str(p[9] or '0').strip() == '1' else float(p[14] or 0)) * 21, 2),
            "unidad": p[10], "moneda": p[11],
            "permite_stock_negativo": str(p[12] or '0').strip() == '1',
            "coeficiente_precio": float(p[13] or 0)
        })
    return partes


# ─── Catálogo de imágenes (Drive → servidor) ─────────────────────────────────
CATALOGO_IMG_DIR = "catalogo_imagenes"
_IMG_EXTS = ('.jpg', '.jpeg', '.png', '.webp', '.gif')

def _catalogo_leaf_folders():
    """Recorre CATALOGO_IMG_DIR y devuelve carpetas que tienen imágenes
    directamente adentro — cada una es un candidato = un artículo/variante."""
    resultado = []
    base = CATALOGO_IMG_DIR
    if not os.path.isdir(base):
        return resultado
    for root, dirs, files in os.walk(base):
        imgs = sorted([f for f in files if f.lower().endswith(_IMG_EXTS)])
        if imgs:
            rel = os.path.relpath(root, base).replace('\\', '/')
            resultado.append({"carpeta": rel, "imagenes": imgs})
    return sorted(resultado, key=lambda x: x["carpeta"])

def _sin_tildes_cat(s):
    import unicodedata as _ud
    return ''.join(c for c in _ud.normalize('NFKD', s or '') if not _ud.combining(c))

def _candidatos_articulo_por_carpeta(carpeta, limit=8):
    """Sugiere artículos de Firebird parecidos al nombre de una carpeta del catálogo
    (match exacto de código como fast-path, y por texto/similaridad como respaldo)."""
    import re as _re_cat
    import difflib
    _IGNORAR = {'fotos','foto','banner','banners','video','videos','imagen','imagenes',
                'jpg','png','the','de','del','la','el','los','las','y'}
    tokens = [t for t in _re_cat.split(r'[\s/_\-]+', carpeta) if t and not t.isdigit() and t.lower() not in _IGNORAR]
    if not tokens:
        return []
    c = conn()
    cur = c.cursor()
    candidatos = {}
    try:
        # Fast-path: algún token coincide EXACTO con un CODIGOPARTICULAR (ej. carpeta "T033")
        for tok in tokens:
            cur.execute(
                'SELECT CODIGOPARTICULAR, DESCRIPCION FROM "ARTICULOS" '
                'WHERE UPPER(CODIGOPARTICULAR) = UPPER(?) AND ACTIVO = \'1\'',
                (tok,)
            )
            for row in cur.fetchall():
                candidatos[row[0]] = {"codigo_particular": row[0], "descripcion": row[1], "score": 1.0, "match": "codigo_exacto"}
        # Fuzzy: candidatos por texto (cualquier token presente en DESCRIPCION), re-rankeados por similaridad
        ors = " OR ".join(["UPPER(DESCRIPCION) CONTAINING UPPER(?)"] * len(tokens))
        cur.execute(
            f'SELECT FIRST 200 CODIGOPARTICULAR, DESCRIPCION FROM "ARTICULOS" '
            f'WHERE ACTIVO = \'1\' AND ({ors})',
            tuple(tokens)
        )
        texto_norm = _sin_tildes_cat(' '.join(tokens)).upper()
        for row in cur.fetchall():
            cp = (row[0] or '').strip()
            if not cp or cp in candidatos:
                continue
            desc_norm = _sin_tildes_cat(row[1] or '').upper()
            score = difflib.SequenceMatcher(None, texto_norm, desc_norm).ratio()
            candidatos[cp] = {"codigo_particular": cp, "descripcion": row[1], "score": round(score, 3), "match": "texto"}
    finally:
        c.close()
    ranked = sorted(candidatos.values(), key=lambda x: -x["score"])
    return ranked[:limit]

@app.get("/admin/imagenes/pendientes")
def get_imagenes_pendientes(_u=Depends(get_admin_user)):
    """Carpetas del catálogo sincronizado que todavía no están 100% vinculadas
    a un código de artículo, con candidatos sugeridos por carpeta."""
    c = _admin_db()
    ya_vinculadas = {r[0] for r in c.execute(
        "SELECT DISTINCT carpeta_origen FROM articulo_imagenes WHERE carpeta_origen IS NOT NULL"
    ).fetchall()}
    c.close()
    resultado = []
    for grupo in _catalogo_leaf_folders():
        if grupo["carpeta"] in ya_vinculadas:
            continue
        resultado.append({
            "carpeta": grupo["carpeta"],
            "imagenes": grupo["imagenes"],
            "candidatos": _candidatos_articulo_por_carpeta(grupo["carpeta"])
        })
    return resultado

@app.get("/admin/imagenes/articulo/{codigo}")
def get_imagenes_de_articulo(codigo: str, _u=Depends(get_admin_user)):
    c = _admin_db()
    rows = c.execute(
        "SELECT id, ruta_imagen, carpeta_origen, orden FROM articulo_imagenes "
        "WHERE codigo_particular = ? ORDER BY orden, id",
        (codigo,)
    ).fetchall()
    c.close()
    return [dict(r) for r in rows]

class VincularImagenesBody(BaseModel):
    codigo_particular: str
    carpeta: str

@app.post("/admin/imagenes/vincular")
def post_vincular_imagenes(body: VincularImagenesBody, _u=Depends(get_admin_user)):
    base = os.path.join(CATALOGO_IMG_DIR, body.carpeta)
    if not os.path.isdir(base):
        raise HTTPException(404, f"Carpeta no encontrada: {body.carpeta}")
    imgs = sorted([f for f in os.listdir(base) if f.lower().endswith(_IMG_EXTS)])
    if not imgs:
        raise HTTPException(400, "La carpeta no tiene imágenes")
    c = _admin_db()
    cur = c.cursor()
    # orden continúa desde el máximo ya existente para ese código (por si ya tenía otras imágenes)
    row = cur.execute("SELECT COALESCE(MAX(orden), -1) FROM articulo_imagenes WHERE codigo_particular = ?",
                       (body.codigo_particular,)).fetchone()
    siguiente_orden = (row[0] or -1) + 1
    for i, img in enumerate(imgs):
        ruta = f"{body.carpeta}/{img}"
        cur.execute(
            "INSERT OR IGNORE INTO articulo_imagenes (codigo_particular, ruta_imagen, carpeta_origen, orden) "
            "VALUES (?, ?, ?, ?)",
            (body.codigo_particular, ruta, body.carpeta, siguiente_orden + i)
        )
    c.commit()
    rows = cur.execute(
        "SELECT id, ruta_imagen, carpeta_origen, orden FROM articulo_imagenes "
        "WHERE codigo_particular = ? ORDER BY orden, id",
        (body.codigo_particular,)
    ).fetchall()
    c.close()
    return [dict(r) for r in rows]

@app.delete("/admin/imagenes/{id}")
def delete_imagen_vinculada(id: int, _u=Depends(get_admin_user)):
    c = _admin_db()
    c.execute("DELETE FROM articulo_imagenes WHERE id = ?", (id,))
    c.commit()
    c.close()
    return {"ok": True}

@app.get("/articulo-imagenes/{codigo}")
def get_articulo_imagenes_publico(codigo: str, _u=Depends(get_current_user)):
    """Imágenes de un artículo para el carrusel del vendedor en pedido/presupuesto."""
    c = _admin_db()
    rows = c.execute(
        "SELECT ruta_imagen FROM articulo_imagenes WHERE codigo_particular = ? ORDER BY orden, id",
        (codigo,)
    ).fetchall()
    c.close()
    return [f"/catalogo-imagenes/{urllib.parse.quote(r[0])}" for r in rows]

_codigos_con_imagenes_cache = {"ts": 0, "val": None}
_CODIGOS_CON_IMAGENES_TTL = 60

@app.get("/articulos-con-imagenes")
def get_articulos_con_imagenes(_u=Depends(get_current_user)):
    """Set de CODIGOPARTICULAR que tienen al menos una foto vinculada — para marcar
    en rojo el código en el pedido/presupuesto y saber si conviene mostrar el carrusel."""
    now = time.time()
    if _codigos_con_imagenes_cache["val"] is not None and now - _codigos_con_imagenes_cache["ts"] < _CODIGOS_CON_IMAGENES_TTL:
        return _codigos_con_imagenes_cache["val"]
    c = _admin_db()
    rows = c.execute("SELECT DISTINCT codigo_particular FROM articulo_imagenes").fetchall()
    c.close()
    val = [r[0] for r in rows]
    _codigos_con_imagenes_cache["ts"] = now
    _codigos_con_imagenes_cache["val"] = val
    return val


@app.get("/stock/{codigo}")
def get_articulo(codigo: str):
    c = conn()
    cur = c.cursor()
    # Cotización USD actual
    try:
        cur.execute('SELECT CAMBIO FROM "MONEDAS" WHERE CODIGOMONEDA = ?', ('DOLARES',))
        rm = cur.fetchone()
        cambio_usd = float(rm[0]) if rm else 1.0
    except Exception:
        cambio_usd = 1.0
    _COLS = ('SELECT a.CODIGOARTICULO, a.DESCRIPCION, a.DESCRIPCIONADICIONAL, a.CODIGOMARCA, '
             'a.PRECIOLISTA1, a.PRECIOLISTA2, a.PRECIOLISTA3, a.PRECIOLISTA5, a.COEFICIENTESEGUNRUBRO, a.CODIGOUNIDADMEDIDA, '
             'a.CODIGOMONEDA, a.CODIGOPARTICULAR, a.DTOMAXIMO1, a.APLICABLEABONIFICACION, '
             'a.CODIGORUBRO, r.CODIGOSUPERRUBRO, sr.CODIGOGRUPOSUPERRUBRO, a.PERMITESTOCKNEGATIVO, a.PARTECONJUNTO, '
             'g.DESCRIPCION, a.COEFICIENTE, r.COEFICIENTE '
             'FROM "ARTICULOS" a '
             'LEFT JOIN "RUBROS" r ON r.CODIGORUBRO = a.CODIGORUBRO '
             'LEFT JOIN "SUPERRUBROS" sr ON sr.CODIGOSUPERRUBRO = r.CODIGOSUPERRUBRO '
             'LEFT JOIN "GRUPOSUPERRUBROS" g ON g.CODIGOGRUPOSUPERRUBRO = sr.CODIGOGRUPOSUPERRUBRO '
             'WHERE ')
    # Buscar primero por CODIGOPARTICULAR (prioridad); si no existe, por CODIGOARTICULO.
    # El OR con una sola consulta puede devolver el artículo equivocado cuando un
    # CODIGOARTICULO de otro artículo coincide con el CODIGOPARTICULAR buscado.
    cur.execute(_COLS + 'a.CODIGOPARTICULAR = ?', (codigo,))
    row = cur.fetchone()
    if not row:
        cur.execute(_COLS + 'a.CODIGOARTICULO = ?', (codigo,))
        row = cur.fetchone()
    if not row:
        c.close()
        raise HTTPException(404, "Artículo no encontrado")
    if (row[19] or '').strip().upper() in ('TERCEROS', 'SERVICIOS'):
        c.close()
        raise HTTPException(404, "Artículo no encontrado")
    moneda = (row[10] or '').strip().upper()
    factor = cambio_usd if moneda == 'DOLARES' else 1.0
    def conv(v): return _redondear_precio(v, factor)
    codigoarticulo = row[0]
    c.close()
    # Remanente por depósito — paralelo + caché TTL. Todos los depósitos activos
    # (antes 6 fijos, dejaba afuera depósitos nuevos como 017 SARANDI).
    _DEPS_ART = _deps_activos()
    rem_bulk = _fma_stock_parallel(_DEPS_ART)
    cod_key = str(codigoarticulo).strip()
    rem = {dep: rem_bulk[dep].get(codigoarticulo, rem_bulk[dep].get(cod_key, 0.0))
           for dep in _DEPS_ART}
    dto_max_raw = row[12]
    aplica_b    = str(row[13] or '0').strip() == '1'
    dto_max     = float(dto_max_raw) if (dto_max_raw is not None and aplica_b) else None
    codigoparticular = (row[11] or row[0] or '').strip()
    item = {
        "codigo": codigoarticulo, "codigoparticular": codigoparticular,
        "descripcion": row[1], "descripcion_adicional": row[2],
        "marca": (row[3] or '').strip(),
        "precio1": conv(row[4]), "precio2": conv(row[5]), "precio3": conv(row[6]), "precio5": conv(row[7]),
        "iva": round((float(row[21] or 0) if str(row[8] or '0').strip() == '1' else float(row[20] or 0)) * 21, 2),
        "unidad": row[9], "moneda": row[10],
        "codigo_rubro":           (row[14] or '').strip(),
        "codigo_superrubro":      (row[15] or '').strip(),
        "codigo_gruposuperrubro": (row[16] or '').strip(),
        "dto_max": dto_max,
        "permite_stock_negativo": str(row[17] or '0').strip() == '1',
        "es_conjunto": str(row[18] or '').strip().upper() == 'C'
    }
    for _d in _DEPS_ART:
        item[f"remanente_{_d}"] = rem[_d]
    item.pop('codigo_rubro', None)
    item.pop('codigo_superrubro', None)
    item.pop('codigo_gruposuperrubro', None)
    return item

# ─── Clientes (solo del vendedor) ─────────────────────────────────────────────
@app.get("/debug/cliente-iva/{codigo}")
def debug_cliente_iva(codigo: str, db: str = Query("oficial")):
    """Debug: columnas de CLIENTES + campos IVA de CABEZAPRESUPUESTOS."""
    db_path = DATABASE if db in ('oficial','l1') else (DATABASE_EST if db == 'est' else DATABASE_MLT)
    resultado = {}
    try:
        c = conn('WIN1252', db=db_path)
        cur = c.cursor()

        # 1) Ver columnas reales de CLIENTES (primer registro)
        try:
            cur.execute('SELECT FIRST 1 * FROM "CLIENTES"')
            row0 = cur.fetchone()
            cols0 = [d[0] for d in cur.description]
            resultado['columnas_CLIENTES'] = cols0
            resultado['primer_registro_sample'] = {cols0[i]: str(row0[i] or '').strip() for i in range(len(cols0))} if row0 else {}
        except Exception as e1:
            resultado['error_CLIENTES'] = str(e1)

        # 2) Buscar el cliente por código (puede ser int o char)
        try:
            cur.execute('SELECT * FROM "CLIENTES" WHERE CODIGOCLIENTE = ?', (int(codigo),))
            row = cur.fetchone()
            if row:
                cols = [d[0] for d in cur.description]
                resultado['cliente'] = {cols[i]: str(row[i] or '').strip() for i in range(len(cols))}
            else:
                resultado['cliente'] = f'No encontrado con codigo={codigo}'
        except Exception as e2:
            resultado['error_buscar_cliente'] = str(e2)

        # 3) Ver CABEZAPRESUPUESTOS del cliente — campos IVA
        try:
            cur.execute(
                'SELECT FIRST 5 NROPRESUPUESTO, CODIGOCLIENTE, COEFICIENTEIVA, '
                'DESCUENTOPORCENTAJE, DESCUENTOMONTO, DESCUENTODESCRIPCION, TOTAL '
                'FROM "CABEZAPRESUPUESTOS" WHERE CODIGOCLIENTE = ? '
                'ORDER BY NROPRESUPUESTO DESC', (int(codigo),)
            )
            rows_p = cur.fetchall()
            cols_p = [d[0] for d in cur.description]
            resultado['ultimos_presupuestos'] = [
                {cols_p[i]: str(r[i] or '').strip() for i in range(len(cols_p))} for r in rows_p
            ]
        except Exception as e3:
            resultado['error_presupuestos'] = str(e3)

        c.close()
    except Exception as e:
        resultado['error_conexion'] = str(e)
    return resultado

# Caché corto para /clientes — cada letra que el vendedor escribe en el buscador de
# cliente disparaba una consulta CONTAINING nueva a Firebird (a diferencia de
# /buscar-articulos, que desde 2026-07-31 busca contra un catálogo en memoria — ver
# _get_catalog — los clientes no tienen ese catálogo). TTL chico porque un cliente
# puede darse de alta o editarse en cualquier momento, pero alcanza para no repetir
# la misma búsqueda si el vendedor borra/reescribe o navega para atrás y adelante.
_cli_cache: dict = {}   # (vendedor, buscar, limit, offset) -> (ts, resultado_ya_armado)
_CLI_CACHE_TTL = int(os.getenv('CLI_CACHE_TTL', 25))

@app.get("/clientes")
def get_clientes(
    vendedor: Optional[str] = None,
    buscar: Optional[str] = None,
    limit: int = Query(100, le=300),
    offset: int = 0,
    _user=Depends(get_current_user)
):
    _cache_key = ((vendedor or '').upper(), (buscar or '').upper(), limit, offset)
    _now = time.time()
    _cached = _cli_cache.get(_cache_key)
    if _cached and (_now - _cached[0]) < _CLI_CACHE_TTL:
        return _cached[1]

    params = []
    where_vendedor = ""
    if vendedor:
        where_vendedor = "AND CODIGOVENDEDOR = ?"
        params = [vendedor.upper()]

    where_buscar = ""
    if buscar:
        where_buscar = "AND (UPPER(RAZONSOCIAL) CONTAINING UPPER(?) OR CODIGOCLIENTE CONTAINING ?)"
        params += [buscar, buscar]

    try:
        c1 = conn()
        cur1 = c1.cursor()
        cur1.execute(f"""
            SELECT FIRST {limit} SKIP {offset}
                CODIGOCLIENTE, RAZONSOCIAL, NOMBREFANTASIA, CUIT,
                TELEFONO, TELEFONOCELULAR, EMAIL, DIRECCION, LOCALIDAD,
                CODIGOPARTICULAR, REPARTOPROPIO, CONDICIONIVA,
                CODIGOMULTIPLAZO, MULTIPLAZOFIJO, COMENTARIOS,
                LIMITECREDITO, LIMITECREDITODOC
            FROM "CLIENTES"
            WHERE ACTIVO = '1' {where_vendedor}
                {where_buscar}
            ORDER BY RAZONSOCIAL
        """, params)
        rows = cur1.fetchall()
        c1.close()
        resultado = [{
            "codigo": r[0], "razonsocial": r[1], "fantasia": r[2],
            "cuit":   r[3], "telefono":   r[4], "celular":  r[5],
            "email":  r[6], "direccion":  r[7], "localidad": r[8],
            "codigoparticular": (r[9] or "").strip(),
            "tipoiva": (r[11] or "").strip(),
            "discrimina_iva": (r[11] or "").strip().upper() == 'RI',
            "reparto_propio": str(r[10] or '0').strip() == '1',
            "codigomultiplazo": str(r[12] or '0').strip(),
            "multiplazofijo": int(r[13] or 0),
            "comentarios": (r[14] or '').strip(),
            "limitecredito": float(r[15] or 0),
            "limitecreditodoc": float(r[16] or 0),
        } for r in rows]
        _cli_cache[_cache_key] = (_now, resultado)
        if len(_cli_cache) > 300:
            _oldest = sorted(_cli_cache, key=lambda k: _cli_cache[k][0])[:100]
            for _k in _oldest:
                _cli_cache.pop(_k, None)
        return resultado
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error clientes: {e}")

@app.get("/clientes/{codigo}")
def get_cliente(codigo: str):
    c = conn()
    cur = c.cursor()
    cur.execute(
        'SELECT CODIGOCLIENTE, RAZONSOCIAL, NOMBREFANTASIA, CUIT, '
        'TELEFONO, TELEFONOCELULAR, EMAIL, DIRECCION, LOCALIDAD, COMENTARIOS '
        'FROM "CLIENTES" WHERE CODIGOCLIENTE = ? OR CODIGOPARTICULAR = ?', (codigo, codigo)
    )
    row = cur.fetchone()
    c.close()
    if not row:
        raise HTTPException(404, "Cliente no encontrado")
    return {
        "codigo": row[0], "razonsocial": row[1], "fantasia": row[2],
        "cuit": row[3], "telefono": row[4], "celular": row[5],
        "email": row[6], "direccion": row[7], "localidad": row[8],
        "comentarios": row[9]
    }

@app.get("/debug/cliente/{codigo}")
def debug_cliente(codigo: str):
    """Muestra todas las columnas de CLIENTES para un cliente dado (busca en todas las DBs)."""
    _DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    for db_path in [DATABASE, _DB_PROD]:
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            cur.execute(
                'SELECT FIRST 1 * FROM "CLIENTES" WHERE CODIGOCLIENTE = ? '
                'OR TRIM(CODIGOCLIENTE) = ? OR RAZONSOCIAL CONTAINING ?',
                (codigo, codigo.strip(), codigo)
            )
            row = cur.fetchone()
            if not row:
                c.close(); continue
            cols = [d[0] for d in cur.description]
            data = {}
            for k, v in zip(cols, row):
                if v is not None:
                    data[k] = str(v).strip() if isinstance(v, str) else v
            c.close()
            keywords = ['MULTI','PLAZO','COMENT','LIMITE','CREDITO','FIJO','INHAB','DESHAB']
            filtrado = {k: v for k, v in data.items() if any(kw in k.upper() for kw in keywords)}
            return {"db": db_path, "filtrado": filtrado, "todos_los_campos": list(cols)}
        except Exception as e:
            continue
    return {"error": "cliente no encontrado en ninguna DB"}

@app.get("/debug/transportes")
def debug_transportes():
    """Diagnóstico: muestra qué transportes devuelve cada DB y el resultado del merge."""
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    resultado = {}
    rows_map: dict = {}
    for label, db_path in [("prueba", DATABASE), ("microbell", DB_PROD)]:
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            cur.execute('SELECT CODIGOTRANSPORTE, DESCRIPCION FROM "TRANSPORTES" ORDER BY DESCRIPCION')
            rows = cur.fetchall()
            c.close()
            items = [{"codigo": r[0], "descripcion": (r[1] or '').strip()} for r in rows if r[0]]
            resultado[label] = {"count": len(items), "transportes": items}
            for r in rows:
                cod = r[0]
                if cod is not None and str(cod).strip() not in ('', '0') and cod not in rows_map:
                    rows_map[cod] = (r[1] or '').strip()
        except Exception as e:
            resultado[label] = {"error": str(e)}
    merged = sorted(
        [{"codigo": cod, "descripcion": desc} for cod, desc in rows_map.items()],
        key=lambda x: x["descripcion"]
    )
    resultado["merged"] = {"count": len(merged), "transportes": merged}
    resultado["codigo_450"] = rows_map.get(450) or rows_map.get('450') or "NO ENCONTRADO"
    return resultado

@app.get("/debug/campos_bonif")
def debug_campos_bonif():
    """Muestra campos DTO/BONIF de ARTICULOS y campos de permiso de PERFILES."""
    c = conn('WIN1252')
    cur = c.cursor()
    res = {}
    try:
        cur.execute("SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME='ARTICULOS' AND (RDB$FIELD_NAME CONTAINING 'DTO' OR RDB$FIELD_NAME CONTAINING 'BONIF' OR RDB$FIELD_NAME CONTAINING 'MAXIMO') ORDER BY RDB$FIELD_POSITION")
        res["articulos_dto_bonif"] = [r[0] for r in cur.fetchall()]
        # Buscar en USUARIOS los campos de permisos
        cur.execute("SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME='USUARIOS' AND (RDB$FIELD_NAME CONTAINING 'PORCENTAJE' OR RDB$FIELD_NAME CONTAINING 'MAXIMO' OR RDB$FIELD_NAME CONTAINING 'INCREMENT' OR RDB$FIELD_NAME CONTAINING 'DECREMENT' OR RDB$FIELD_NAME CONTAINING 'BONIF') ORDER BY RDB$FIELD_POSITION")
        res["usuarios_permisos_campos"] = [r[0] for r in cur.fetchall()]
        # Valores de un usuario ADV como ejemplo (primer usuario ADV activo)
        cur.execute('SELECT FIRST 1 CODIGOUSUARIO, CODIGOPERFIL FROM "USUARIOS" WHERE CODIGOPERFIL=? AND ACTIVO=?', ('ADV', '1'))
        row_u = cur.fetchone()
        if row_u:
            uid = row_u[0]
            cur2 = c.cursor()
            cur2.execute("SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME='USUARIOS' ORDER BY RDB$FIELD_POSITION")
            cols_u = [r[0] for r in cur2.fetchall()]
            cur3 = c.cursor()
            cur3.execute('SELECT * FROM "USUARIOS" WHERE CODIGOUSUARIO=?', (uid,))
            row_vals = cur3.fetchone()
            if row_vals:
                res["usuario_ADV_sample"] = {k: str(v) for k, v in zip(cols_u, row_vals) if any(x in k for x in ['PORCENTAJE','MAXIMO','INCREMENT','DECREMENT','BONIF','CODIGO','PERFIL'])}
    except Exception as e:
        res["error"] = str(e)
    finally:
        c.close()
    return res

@app.get("/debug/sucursales/{codigo}")
def debug_sucursales(codigo: str):
    """Diagnóstico: muestra el paso a paso de resolución de sucursales para un cliente."""
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    resultado = {"codigo_recibido": codigo, "pasos": []}

    # Paso 1: DB-Prueba
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        cur.execute('SELECT CODIGOCLIENTE, CODIGOPARTICULAR, DIRECCION, RAZONSOCIAL '
                    'FROM "CLIENTES" WHERE CODIGOCLIENTE = ? OR CODIGOPARTICULAR = ?',
                    (codigo, codigo))
        r = cur.fetchone()
        c.close()
        resultado["pasos"].append({"db": "DB-Prueba", "encontrado": r is not None,
            "fila": [str(x) for x in r] if r else None})
        codigoparticular = (r[1] or '').strip() if r else codigo
    except Exception as e:
        resultado["pasos"].append({"db": "DB-Prueba", "error": str(e)})
        codigoparticular = codigo

    resultado["codigoparticular_resuelto"] = codigoparticular

    # Paso 2: DB-Microbell - resolver CODIGOCLIENTE
    try:
        c = conn('WIN1252', db=DB_PROD)
        cur = c.cursor()
        cur.execute('SELECT CODIGOCLIENTE, CODIGOPARTICULAR, DIRECCION, RAZONSOCIAL '
                    'FROM "CLIENTES" WHERE CODIGOPARTICULAR = ? OR CODIGOCLIENTE = ?',
                    (codigoparticular, codigoparticular))
        r = cur.fetchone()
        resultado["pasos"].append({"db": "DB-Microbell CLIENTES", "encontrado": r is not None,
            "fila": [str(x) for x in r] if r else None})
        cod_mb = str(r[0]).strip() if r else codigoparticular
        resultado["codigocliente_microbell"] = cod_mb

        # Paso 3: transporte a nivel cliente en CLIENTES
        cur.execute('SELECT CODIGOTRANSPORTE, TRANSPORTEFIJO, REPARTOPROPIO '
                    'FROM "CLIENTES" WHERE CODIGOCLIENTE = ?', (cod_mb,))
        r_cli2 = cur.fetchone()
        cols_cli2 = [d[0] for d in cur.description] if cur.description else []
        resultado["clientes_columnas_transporte"] = [c for c in cols_cli2 if 'TRANS' in c.upper() or 'REPARTO' in c.upper()]
        transp_cli = str(r_cli2[0]).strip() if r_cli2 and r_cli2[0] else None
        transp_fijo = str(r_cli2[1] or '0').strip() == '1' if r_cli2 else False
        if r_cli2:
            resultado["clientes_transporte_valores"] = {
                "CODIGOTRANSPORTE": transp_cli, "TRANSPORTEFIJO": transp_fijo
            }

        # Paso 4: SUCURSALESXCLIENTES
        cur.execute('SELECT CODIGOSUCURSAL, NOMBRE, DIRECCION, CODIGOTRANSPORTE '
                    'FROM "SUCURSALESXCLIENTES" WHERE CODIGOCLIENTE = ? ORDER BY CODIGOSUCURSAL',
                    (cod_mb,))
        rows = cur.fetchall()
        resultado["sucursales_count"] = len(rows)
        resultado["sucursales"] = [{"cod": str(r[0]), "nombre": str(r[1]), "dir": str(r[2]),
                                    "transp_suc": str(r[3]) if r[3] else None,
                                    "transp_efectivo": str(r[3]).strip() if r[3] else transp_cli,
                                    "fijo": transp_fijo} for r in rows]
        c.close()
    except Exception as e:
        resultado["pasos"].append({"db": "DB-Microbell", "error": str(e)})

    return resultado


@app.get("/debug/despachos/{tipo}/{numero}")
def debug_despachos(tipo: str, numero: str):
    """Busca despachos para un comprobante en todas las tablas posibles."""
    result = {}
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    for db_label, db_path in [('prueba', DATABASE), ('prod', DB_PROD)]:
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            # 1. Columnas de CUERPOCOMPROBANTES
            try:
                cur.execute("SELECT TRIM(f.RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS f WHERE f.RDB$RELATION_NAME='CUERPOCOMPROBANTES' ORDER BY f.RDB$FIELD_POSITION")
                result[f'{db_label}_cuerpo_cols'] = [r[0] for r in cur.fetchall()]
            except Exception as e:
                result[f'{db_label}_cuerpo_cols_err'] = str(e)
            # 2. Primera fila del comprobante (todos los campos)
            try:
                cur.execute('SELECT FIRST 1 * FROM "CUERPOCOMPROBANTES" WHERE TIPOCOMPROBANTE=? AND NUMEROCOMPROBANTE=?', (tipo, numero))
                row = cur.fetchone()
                if row:
                    cols = [d[0] for d in cur.description]
                    result[f'{db_label}_cuerpo_row'] = dict(zip(cols, [str(v) for v in row]))
            except Exception as e:
                result[f'{db_label}_cuerpo_row_err'] = str(e)
            # 3. Buscar tablas con "DESP" en el nombre
            try:
                cur.execute("SELECT DISTINCT TRIM(r.RDB$RELATION_NAME) FROM RDB$RELATIONS r WHERE r.RDB$RELATION_NAME LIKE '%DESP%' ORDER BY 1")
                result[f'{db_label}_tablas_desp'] = [r[0] for r in cur.fetchall()]
            except Exception as e:
                result[f'{db_label}_tablas_desp_err'] = str(e)
            # 4. Intentar cada tabla DESP encontrada
            for tbl in result.get(f'{db_label}_tablas_desp', []):
                try:
                    cur.execute(f'SELECT FIRST 3 * FROM "{tbl}"')
                    cols2 = [d[0] for d in cur.description]
                    rows2 = cur.fetchall()
                    result[f'{db_label}_{tbl}_cols'] = cols2
                    result[f'{db_label}_{tbl}_sample'] = [dict(zip(cols2,[str(v) for v in r])) for r in rows2]
                    # Buscar si tiene NUMEROCOMPROBANTE o TIPOCOMPROBANTE
                    if any(c in cols2 for c in ['NUMEROCOMPROBANTE','NROCOMPROBANTE']):
                        num_col = 'NUMEROCOMPROBANTE' if 'NUMEROCOMPROBANTE' in cols2 else 'NROCOMPROBANTE'
                        cur.execute(f'SELECT FIRST 5 * FROM "{tbl}" WHERE {num_col}=?', (numero,))
                        rows3 = cur.fetchall()
                        result[f'{db_label}_{tbl}_match'] = [dict(zip(cols2,[str(v) for v in r])) for r in rows3]
                except Exception as e:
                    result[f'{db_label}_{tbl}_err'] = str(e)
            c.close()
        except Exception as e:
            result[f'{db_label}_conn_err'] = str(e)
    return result


@app.get("/debug/cae/{numero}")
def debug_cae(numero: str):
    """Busca el CAE para un comprobante en tablas conocidas de Flexxus."""
    result = {}
    # NUMEROCOMPROBANTE almacenado como float sin formato
    num_float = numero  # el caller pasa el número tal como viene del comp_cols
    try:
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()
        # 1. Buscar en CAMPOSDINAMICOSCOMPROBANTES
        try:
            cur.execute(
                'SELECT CODIGOCAMPODINAMICO, VALOR FROM "CAMPOSDINAMICOSCOMPROBANTES" '
                'WHERE NUMEROCOMPROBANTE = ?', (num_float,)
            )
            result['CAMPOSDINAMICOS'] = [{"cod": r[0], "valor": str(r[1])} for r in cur.fetchall()]
        except Exception as e:
            result['CAMPOSDINAMICOS'] = str(e)
        # 2. Buscar en COMPROBANTESCAE si existe
        try:
            cur.execute(
                'SELECT * FROM "COMPROBANTESCAE" WHERE NUMEROCOMPROBANTE = ?', (num_float,)
            )
            row = cur.fetchone()
            if row:
                cols = [d[0] for d in cur.description]
                result['COMPROBANTESCAE'] = dict(zip(cols, [str(v) for v in row]))
            else:
                result['COMPROBANTESCAE'] = 'no encontrado'
        except Exception as e:
            result['COMPROBANTESCAE'] = str(e)
        # 3. Buscar tablas que contengan "CAE" en campos o nombre
        try:
            cur.execute(
                "SELECT DISTINCT f.RDB$RELATION_NAME "
                "FROM RDB$RELATION_FIELDS f "
                "WHERE f.RDB$FIELD_NAME CONTAINING 'CAE' "
                "AND f.RDB$SYSTEM_FLAG = 0"
            )
            result['tablas_con_CAE'] = [r[0].strip() for r in cur.fetchall()]
        except Exception as e:
            result['tablas_con_CAE'] = str(e)
        # 4. Consultar CAEAFIP
        try:
            cur.execute('SELECT FIRST 1 * FROM "CAEAFIP" WHERE NUMEROCOMPROBANTE = ?', (num_float,))
            row = cur.fetchone()
            if row:
                cols = [d[0] for d in cur.description]
                result['CAEAFIP'] = dict(zip(cols, [str(v) if v is not None else None for v in row]))
            else:
                # Mostrar estructura + últimos registros para ver formato
                cur.execute('SELECT FIRST 3 * FROM "CAEAFIP" ORDER BY 1 DESC')
                rows = cur.fetchall()
                cols = [d[0] for d in cur.description]
                result['CAEAFIP'] = {
                    'sin_dato_para_numero': num_float,
                    'muestra': [dict(zip(cols, [str(v) if v is not None else None for v in r])) for r in rows]
                }
        except Exception as e:
            result['CAEAFIP'] = str(e)
        c.close()
    except Exception as e:
        result['error'] = str(e)
    return result

@app.get("/debug/cuerpo-cols")
def debug_cuerpo_cols(db: str = Query("prod")):
    """Lista TODAS las columnas de CUERPOCOMPROBANTES con su valor en la 1ra FA."""
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    db_path = {'sw': DATABASE_MLT, 'prod': DB_PROD, 'oficial': DATABASE}.get(db, DB_PROD)
    try:
        c = conn('WIN1252', db=db_path)
        cur = c.cursor()
        cur.execute('SELECT FIRST 1 * FROM "CUERPOCOMPROBANTES" WHERE TIPOCOMPROBANTE=?',('FA',))
        row  = cur.fetchone()
        cols = [d[0] for d in cur.description]
        c.close()
        if not row:
            return {"error": "Sin filas FA", "cols": cols}
        data = {cols[i]: (float(row[i]) if isinstance(row[i], (int, float)) else str(row[i]) if row[i] is not None else None)
                for i in range(len(cols))}
        # Destacar las que parecen de costo/precio
        relevantes = {k: v for k, v in data.items() if any(t in k.upper() for t in ['COSTO','PRECIO','COMPRA','REPO'])}
        return {"relevantes": relevantes, "todos": data}
    except Exception as e:
        return {"error": str(e)}




@app.get("/admin/import-comisiones")
def admin_import_comisiones(token: str = Query(...)):
    """Importa comisiones_vendedores desde importar_comisiones.sql. Uso único."""
    if token != "microbell2026":
        raise HTTPException(status_code=403, detail="Token inválido")

    sql_path = os.path.join(os.path.dirname(__file__), 'importar_comisiones.sql')
    if not os.path.exists(sql_path):
        raise HTTPException(status_code=404, detail="importar_comisiones.sql no encontrado en el servidor")

    import sqlite3 as _sq3
    db_path = os.path.join(os.path.dirname(__file__), 'admin.db')
    db = _sq3.connect(db_path)
    db.executescript(open(sql_path, encoding='utf-8').read())
    count = db.execute('SELECT COUNT(*) FROM comisiones_vendedores').fetchone()[0]
    db.close()
    _load_comisiones()  # recarga en memoria
    return {"ok": True, "filas_importadas": count}

@app.get("/debug/comisiones-check")
def debug_comisiones_check():
    """Muestra nombres de USUARIOS vs primeras claves de _COMISIONES para diagnosticar matching."""
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    usuarios = {}
    for _db in [DATABASE, DB_PROD]:
        try:
            c = conn('WIN1252', db=_db)
            cur = c.cursor()
            cur.execute('SELECT CODIGOUSUARIO, RAZONSOCIAL FROM "USUARIOS"')
            for row in cur.fetchall():
                cod = (row[0] or '').strip().upper()
                nom = (row[1] or '').strip().upper()
                if cod and cod not in usuarios:
                    usuarios[cod] = nom
            c.close()
        except Exception:
            pass
    # Primeros 30 keys del dict de comisiones
    sample_com = list(_COMISIONES.keys())[:30]
    # Ver qué vendedores tienen comisiones definidas
    vend_com = sorted({k[0] for k in _COMISIONES.keys()})
    return {
        "usuarios_mapa": usuarios,
        "vendedores_en_comisiones": vend_com[:50],
        "sample_comisiones_keys": [str(k) for k in sample_com],
        "total_comisiones": len(_COMISIONES),
    }

@app.get("/debug/cuerpo-costo/{numero}")
def debug_cuerpo_costo(numero: str, tipo: str = "FA", db: str = Query("oficial")):
    """Muestra todos los campos de CUERPOCOMPROBANTES con COSTO o PRECIO en el nombre,
    para una línea del comprobante dado. Útil para identificar el campo de costo de venta."""
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    db_path = {'sw': DATABASE_MLT, 'prod': DB_PROD}.get(db, DATABASE)
    try:
        c = conn('WIN1252', db=db_path)
        cur = c.cursor()
        cur.execute(
            'SELECT FIRST 3 * FROM "CUERPOCOMPROBANTES" '
            'WHERE TIPOCOMPROBANTE=? AND NUMEROCOMPROBANTE=?',
            (tipo.upper(), numero)
        )
        rows = cur.fetchall()
        if not rows:
            # fallback: últimas 3 líneas FA
            cur.execute(
                "SELECT FIRST 3 * FROM \"CUERPOCOMPROBANTES\" "
                "WHERE TIPOCOMPROBANTE='FA' ORDER BY NUMEROCOMPROBANTE DESC"
            )
            rows = cur.fetchall()
        cols = [d[0] for d in cur.description]
        c.close()
        # Filtrar columnas con COSTO o PRECIO en el nombre (case-insensitive)
        idx_rel = [i for i, c in enumerate(cols) if 'COSTO' in c.upper() or 'PRECIO' in c.upper()]
        result = []
        for row in rows:
            full = {cols[i]: (str(row[i]) if row[i] is not None else None) for i in range(len(cols))}
            rel  = {cols[i]: (str(row[i]) if row[i] is not None else None) for i in idx_rel}
            result.append({"linea": full.get("CODIGOPARTICULAR") or full.get("CODIGOARTICULO"),
                           "campos_precio_costo": rel})
        return {"columnas_totales": len(cols), "todas_las_columnas": cols,
                "lineas": result}
    except Exception as e:
        return {"error": str(e)}


@app.get("/debug/comp_cols/{numero}")
def debug_comp_cols(numero: str, db: str = Query("oficial")):
    """Muestra TODOS los campos de CABEZACOMPROBANTES para un comprobante dado."""
    db_path = DATABASE_MLT if db == 'sw' else DATABASE
    try:
        c = conn('WIN1252', db=db_path)
        cur = c.cursor()
        # Intentar búsqueda exacta primero, luego por LIKE
        cur.execute(
            'SELECT FIRST 1 * FROM "CABEZACOMPROBANTES" '
            'WHERE NUMEROCOMPROBANTE = ? OR NUMEROCOMPROBANTE LIKE ?',
            (numero, f'%{numero}%')
        )
        row = cur.fetchone()
        if not row:
            # Devolver los últimos 5 comprobantes para ver el formato real
            cur.execute(
                'SELECT FIRST 5 TIPOCOMPROBANTE, NUMEROCOMPROBANTE, FECHACOMPROBANTE '
                'FROM "CABEZACOMPROBANTES" ORDER BY FECHACOMPROBANTE DESC'
            )
            sample = cur.fetchall()
            c.close()
            return {"error": f"No encontrado", "formato_muestra": [
                {"tipo": r[0], "numero": r[1], "fecha": str(r[2])} for r in sample
            ]}
        cols = [d[0] for d in cur.description]
        c.close()
        return dict(zip(cols, [str(v) if v is not None else None for v in row]))
    except Exception as e:
        return {"error": str(e)}

# ─── Cuenta Corriente ──────────────────────────────────────────────────────────
def _resolve_codigos_en_db(db_path, codigoparticular):
    """Dado un CODIGOPARTICULAR, devuelve los CODIGOCLIENTEs que lo usan en esa BD via CUERPOCOMPROBANTES."""
    if not codigoparticular or not str(codigoparticular).strip():
        return []
    try:
        c = conn('LATIN1', db=db_path)
        cur = c.cursor()
        cur.execute(
            'SELECT DISTINCT CODIGOCLIENTE FROM "CUERPOCOMPROBANTES" WHERE CODIGOPARTICULAR = ?',
            (str(codigoparticular).strip(),)
        )
        rows = cur.fetchall()
        c.close()
        return [str(r[0]).strip() for r in rows if r[0] and str(r[0]).strip()]
    except Exception:
        return []

def _query_cta_part(db_path, codigos_base, codigoparticular, limit, offset):
    """Como _query_cta pero resuelve el CODIGOCLIENTE correcto para esta BD vía CODIGOPARTICULAR."""
    codigos = set(c for c in codigos_base if c)
    codigos.update(_resolve_codigos_en_db(db_path, codigoparticular))
    if not codigos:
        return []
    return _query_cta(db_path, list(codigos), limit, offset)

def _get_cambios(db_path):
    """Retorna dict {CODIGOMONEDA: CAMBIO} desde tabla MONEDAS."""
    try:
        c = conn('WIN1252', db=db_path)
        cur = c.cursor()
        cur.execute('SELECT CODIGOMONEDA, CAMBIO FROM "MONEDAS"')
        result = {str(r[0]).strip(): float(r[1] or 1) for r in cur.fetchall()}
        c.close()
        return result
    except Exception:
        return {}

def _query_cta(db_path, codigos, limit, offset, vendedor=None):
    """Consulta CABEZACOMPROBANTES sin CAST/JOIN en SQL. Conversión de moneda en Python.
    Usa fetchone() en loop para tolerar registros individuales con valores problemáticos."""
    cambios = _get_cambios(db_path)
    result = []
    try:
        c = conn('WIN1252', db=db_path)
        cur = c.cursor()
        ph = ', '.join(['?'] * len(codigos))
        params = list(codigos)
        _NC_TIPOS = "('NCA','NCB','NCCA','NCCB','NCE','NCCE','SIV','NDA','NDB','NDCA','NDCB')"
        if vendedor:
            # FAs y NCs: ambas filtradas por CODIGOUSUARIO del vendedor logueado
            cur.execute(f"""
                SELECT FIRST {limit} SKIP {offset}
                    TIPOCOMPROBANTE, NUMEROCOMPROBANTE, FECHACOMPROBANTE,
                    TOTAL, IVA1, IVA2, PAGADO, COTIZACION, CODIGOMONEDA,
                    FECHAVENCIMIENTO, CLASECOMPROBANTE
                FROM "CABEZACOMPROBANTES"
                WHERE CODIGOCLIENTE IN ({ph})
                  AND ANULADA = '0'
                  AND TIPOCOMPROBANTE NOT IN ('RE', 'RI', 'INA')
                  AND UPPER(CODIGOUSUARIO) = ?
                  AND (CUENTACORRIENTE = '1' OR TIPOCOMPROBANTE IN {_NC_TIPOS})
                ORDER BY FECHAVENCIMIENTO ASC, FECHACOMPROBANTE ASC
            """, tuple(params + [vendedor.upper()]))
        else:
            cur.execute(f"""
                SELECT FIRST {limit} SKIP {offset}
                    TIPOCOMPROBANTE, NUMEROCOMPROBANTE, FECHACOMPROBANTE,
                    TOTAL, IVA1, IVA2, PAGADO, COTIZACION, CODIGOMONEDA,
                    FECHAVENCIMIENTO, CLASECOMPROBANTE
                FROM "CABEZACOMPROBANTES"
                WHERE CODIGOCLIENTE IN ({ph})
                  AND ANULADA = '0'
                  AND TIPOCOMPROBANTE NOT IN ('RE', 'RI', 'INA')
                  AND (CUENTACORRIENTE = '1' OR TIPOCOMPROBANTE IN {_NC_TIPOS})
                ORDER BY FECHAVENCIMIENTO ASC, FECHACOMPROBANTE ASC
            """, tuple(params))
        while True:
            try:
                r = cur.fetchone()
                if r is None:
                    break
                tipo   = r[0]; num   = r[1]; fecha = r[2]
                total  = float(r[3] or 0)
                iva1   = float(r[4] or 0)
                iva2   = float(r[5] or 0)
                pagado = float(r[6] or 0)
                cotiz  = float(r[7] or 1) or 1.0
                moneda = str(r[8] or '').strip()
                fvto   = r[9]; clase = r[10]
                neto   = total + iva1 + iva2
                debe   = neto - pagado
                cambio = cambios.get(moneda, 1.0) or 1.0
                deuda  = debe * cambio / cotiz
                if abs(deuda) >= 0.01:
                    result.append((tipo, num, fecha, neto, pagado, deuda, fvto, clase))
            except Exception:
                continue
        c.close()
    except Exception:
        pass
    return result

_credito_cache: dict = {}
_CREDITO_CACHE_TTL = int(os.getenv('CREDITO_CACHE_TTL', 25))

@app.get("/clientes/{codigo}/credito")
def get_credito_cliente(codigo: str):
    """Retorna límites de crédito y saldo deudor actual del cliente."""
    _now_cred = time.time()
    _cached_cred = _credito_cache.get(codigo)
    if _cached_cred and (_now_cred - _cached_cred[0]) < _CREDITO_CACHE_TTL:
        return _cached_cred[1]
    DB_PROD     = DATABASE      # DB-Prueba.gdb
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'  # SW producción
    lim_cred = lim_doc = 0.0
    cod_real = codigo
    part_real = None
    _db_fuente = None
    _db_errores = {}
    for db_path in [DATABASE, DB_PROD]:
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            cur.execute(
                'SELECT LIMITECREDITO, LIMITECREDITODOC, CODIGOCLIENTE, '
                'MULTIPLAZOFIJO, CODIGOMULTIPLAZO, CODIGOPARTICULAR '
                'FROM "CLIENTES" WHERE CODIGOCLIENTE = ? OR CODIGOPARTICULAR = ?',
                (codigo, codigo)
            )
            row = cur.fetchone()
            c.close()
            if row:
                lim_cred       = float(row[0] or 0)
                lim_doc        = float(row[1] or 0)
                cod_real       = str(row[2]).strip()
                multiplazofijo = int(row[3] or 0)
                codigomulti    = str(row[4] or '0').strip()
                part_real      = str(row[5] or '').strip() or None
                _db_fuente     = db_path
                break
        except Exception as _e:
            _db_errores[db_path] = str(_e)
    # Saldo deudor: solo CODIGOCLIENTE (nunca mezclar con CODIGOPARTICULAR en CABEZACOMPROBANTES).
    # DB_PROD y DB_MLT_PROD son bases Firebird DISTINTAS — se consultan en paralelo
    # (cada una abre su propia conexión de red) en vez de una tras otra, para no sumar
    # el costo de conexión de ambas de forma secuencial.
    from concurrent.futures import ThreadPoolExecutor as _TPE_credito
    saldo_deudor = 0.0
    seen_sd = set()
    with _TPE_credito(max_workers=2) as _ex_sd:
        _futs_sd = {_ex_sd.submit(_query_cta, db_path, [cod_real], 500, 0): db_path for db_path in [DB_PROD, DB_MLT_PROD]}
        for _fut in _futs_sd:
            try:
                rows = _fut.result()
            except Exception:
                rows = []
            for r in rows:
                key = (r[0], r[1])
                if key not in seen_sd:
                    seen_sd.add(key)
                    saldo_deudor += float(r[5] or 0)
    saldo_deudor = round(saldo_deudor, 2)
    # Pedidos "A preparar": OPERACION='1' (Flexxus setea FECHATERMINADA al confirmar,
    # no al entregar — por eso filtramos por OPERACION, no por FECHATERMINADA IS NULL).
    # TOTAL en Flexxus = monto bruto con IVA incluido.
    pedidos_pendientes = 0.0
    for db_path in [DATABASE, DB_PROD]:
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            # Excluye pedidos que ya tienen remito (CANTIDADREMITIDA > 0 en alguna línea)
            # aunque Flexxus no haya cambiado el OPERACION manualmente.
            cur.execute(
                'SELECT COALESCE(SUM(cb.TOTAL), 0) FROM "CABEZAPEDIDOS" cb '
                'WHERE cb.TIPOCOMPROBANTE = ? AND cb.CODIGOCLIENTE = ? '
                'AND cb.ANULADA = ? AND cb.OPERACION = ? '
                'AND NOT EXISTS ('
                '  SELECT 1 FROM "CUERPOPEDIDOS" cue '
                '  WHERE cue.TIPOCOMPROBANTE = ? '
                '    AND cue.NUMEROCOMPROBANTE = cb.NUMEROCOMPROBANTE '
                '    AND cue.CANTIDADREMITIDA > 0'
                ')',
                ('NP', cod_real, '0', '1', 'NP')
            )
            row = cur.fetchone()
            c.close()
            if row:
                pedidos_pendientes += float(row[0] or 0)
                break  # BD operativa encontrada, no sumar la otra
        except Exception:
            pass
    pedidos_pendientes = round(pedidos_pendientes, 2)
    disponible_total = round(max(0, lim_cred + lim_doc - saldo_deudor - pedidos_pendientes), 2)
    resultado = {
        "limitecredito":      lim_cred,
        "limitecreditodoc":   lim_doc,
        "saldo_deudor":       saldo_deudor,
        "pedidos_pendientes": pedidos_pendientes,
        "disponible_cred":    round(max(0, lim_cred - saldo_deudor), 2),
        "disponible_doc":     round(max(0, lim_doc  - saldo_deudor), 2),
        "disponible_total":   disponible_total,
        "multiplazofijo":     multiplazofijo,
        "codigomultiplazo":   codigomulti,
        "_db_fuente":         _db_fuente,
        "_db_errores":        _db_errores,
    }
    _credito_cache[codigo] = (_now_cred, resultado)
    if len(_credito_cache) > 300:
        _oldest = sorted(_credito_cache, key=lambda k: _credito_cache[k][0])[:100]
        for _k in _oldest:
            _credito_cache.pop(_k, None)
    return resultado


@app.get("/clientes/{codigo}/cuenta_corriente")
def cuenta_corriente(codigo: str, limit: int = Query(200, le=500), offset: int = 0, _user=Depends(get_current_user)):
    # Lookup en DB_PROD (igual que resumen-deudas)
    DB_PROD      = DATABASE                              # DB-Prueba.gdb
    DB_MLT_PROD  = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'  # SW producción
    c_cli = conn('WIN1252', db=DB_PROD)
    cur_cli = c_cli.cursor()
    cur_cli.execute(
        'SELECT CODIGOCLIENTE, CODIGOPARTICULAR FROM "CLIENTES" '
        'WHERE CODIGOCLIENTE = ? OR CODIGOPARTICULAR = ?',
        (codigo, codigo)
    )
    cli = cur_cli.fetchone()
    c_cli.close()
    # Misma lógica que PDF: buscar con AMBOS CODIGOCLIENTE y CODIGOPARTICULAR
    codigos_set = set()
    if cli:
        if cli[0] and cli[0].strip(): codigos_set.add(cli[0].strip())
        if cli[1] and cli[1].strip(): codigos_set.add(cli[1].strip())
    if not codigos_set:
        codigos_set.add(codigo)
    codigos = list(codigos_set)

    # Fetch ALL rows con límite SQL amplio; la paginación se aplica en Python
    # después del filtro de deuda para no cortar registros como NDAs tardías.
    SQL_LIMIT = 2000
    rows_prod     = _query_cta(DB_PROD,     codigos, SQL_LIMIT, 0)
    rows_mlt_prod = _query_cta(DB_MLT_PROD, codigos, SQL_LIMIT, 0)

    # Combinar y deduplicar por (tipo, numero)
    seen = set()
    combined = []
    for r in rows_prod + rows_mlt_prod:
        key = (r[0], r[1])  # tipo + numero
        if key not in seen:
            seen.add(key)
            combined.append(r)

    # Ordenar por fecha vencimiento
    combined.sort(key=lambda r: (r[6] or r[2], r[2]))

    # Paginación Python post-filtro
    combined = combined[offset: offset + limit]

    return [{
        "tipo":      r[0], "numero":    r[1], "fecha":   r[2],
        "total":     float(r[3]) if r[3] else 0,
        "pagado":    float(r[4]) if r[4] else 0,
        "deuda":     float(r[5]) if r[5] else 0,
        "fecha_vto": r[6].isoformat() if r[6] else None, "clase": r[7],
    } for r in combined]


@app.get("/que-vendi/clientes")
def que_vendi_clientes(vendedor: str, buscar: Optional[str] = None, _user=Depends(get_current_user)):
    """Clientes del vendedor para autocomplete en ¿Qué Vendí?"""
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    params = [vendedor.upper()]
    where_buscar = ""
    if buscar:
        where_buscar = "AND (UPPER(RAZONSOCIAL) CONTAINING UPPER(?) OR CODIGOCLIENTE CONTAINING ? OR CODIGOPARTICULAR CONTAINING ?)"
        params += [buscar, buscar, buscar]
    try:
        c = conn('WIN1252', DB_PROD)
        cur = c.cursor()
        cur.execute(f"""
            SELECT FIRST 30 CODIGOCLIENTE, RAZONSOCIAL, CODIGOPARTICULAR
            FROM "CLIENTES"
            WHERE ACTIVO = '1' AND UPPER(CODIGOVENDEDOR) = ?
            {where_buscar}
            ORDER BY RAZONSOCIAL
        """, params)
        rows = cur.fetchall()
        c.close()
        return [{"codigo": (r[0] or '').strip(),
                 "razonsocial": (r[1] or '').strip(),
                 "codigoparticular": (r[2] or '').strip()} for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/que-vendi")
def que_vendi(
    vendedor: str,
    cliente: str,
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    limit: int = Query(500, le=2000),
    offset: int = 0,
    _user=Depends(get_current_user)
):
    """Artículos facturados a un cliente. Facturas en BDs Prueba; lookup de código en todas las BDs."""
    from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeout
    DB_PROD     = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    # Solo producción — las BDs de prueba tienen datos sucios con códigos distintos
    DBS_FACT   = [DB_PROD, DB_MLT_PROD]
    # Lookup solo en las BDs que tienen tabla CLIENTES
    DBS_LOOKUP = [DATABASE, DB_PROD]

    # ── Lookup de códigos del cliente ────────────────────────────────────────
    def _lookup_codigos(db_path):
        try:
            c = conn('WIN1252', db_path)
            cur = c.cursor()
            cur.execute(
                'SELECT CODIGOCLIENTE, CODIGOPARTICULAR FROM "CLIENTES" '
                'WHERE CODIGOCLIENTE = ? OR CODIGOPARTICULAR = ?',
                (cliente, cliente)
            )
            row = cur.fetchone()
            c.close()
            if row:
                # str() porque CODIGOCLIENTE puede ser INTEGER en Firebird
                return [str(v).strip() for v in row if v is not None and str(v).strip()]
        except Exception:
            pass
        return []

    codigos = set([str(cliente).strip()])
    with ThreadPoolExecutor(max_workers=3) as ex:
        futs = {ex.submit(_lookup_codigos, db): db for db in DBS_LOOKUP}
        for f in futs:
            try:
                for cod in f.result(timeout=8):
                    codigos.add(cod)
            except FutureTimeout:
                pass
    codigos = list(codigos)

    # WHERE dinámico — filtra por CODIGOCLIENTE usando todos los códigos resueltos
    ph = ','.join('?' * len(codigos))
    TIPOS_FACTURA = ("('FA','FB','FE','FCA','FCB','FCE',"
                     "'FCCA','FCCB','FCCE',"
                     "'NCA','NCB','NCCA','NCCB')")
    TIPOS_NC = {'NCA', 'NCB', 'NCCA', 'NCCB'}   # valores se negarán
    params_extra = []
    where_parts = [f"cb.CODIGOCLIENTE IN ({ph})",
                   "cb.ANULADA = '0'",
                   f"cb.TIPOCOMPROBANTE IN {TIPOS_FACTURA}"]
    if desde:
        where_parts.append("cb.FECHACOMPROBANTE >= ?")
        params_extra.append(desde)
    if hasta:
        where_parts.append("cb.FECHACOMPROBANTE <= ?")
        params_extra.append(hasta)
    params_base = list(codigos) + params_extra
    where_sql = " AND ".join(where_parts)

    sql = f"""
        SELECT FIRST {limit} SKIP {offset}
            TRIM(cu.CODIGOARTICULO)                        AS COD_ART,
            COALESCE(NULLIF(TRIM(cu.DESCRIPCION),''), '') AS DESCR,
            cb.FECHACOMPROBANTE,
            cb.TIPOCOMPROBANTE,
            cb.NUMEROCOMPROBANTE,
            CAST(cu.CANTIDAD       AS DOUBLE PRECISION)  AS CANT,
            CAST(cu.PRECIOUNITARIO AS DOUBLE PRECISION)  AS PU,
            CAST(cu.PRECIOTOTAL    AS DOUBLE PRECISION)  AS SUBTOTAL,
            COALESCE(CAST(cu.PORCENTAJEIVA AS DOUBLE PRECISION), 21) AS IVA_PCT
        FROM "CUERPOCOMPROBANTES" cu
        JOIN "CABEZACOMPROBANTES" cb
             ON cb.TIPOCOMPROBANTE   = cu.TIPOCOMPROBANTE
            AND cb.NUMEROCOMPROBANTE = cu.NUMEROCOMPROBANTE
        WHERE {where_sql}
        ORDER BY cb.FECHACOMPROBANTE DESC, cb.NUMEROCOMPROBANTE DESC
    """

    # ── Consulta paralela a las 4 BDs, timeout 25s por BD ────────────────────
    _db_errors = {}
    def _query_db(db_path):
        try:
            c = conn('LATIN1', db=db_path)
            cur = c.cursor()
            cur.execute(sql, params_base)
            rows = cur.fetchall()
            c.close()
            _QV_LAST_COUNTS[db_path] = len(rows)
            return rows
        except Exception as _e:
            _db_errors[db_path] = str(_e)
            _QV_LAST_ERRORS[db_path] = str(_e)
            _QV_LAST_COUNTS[db_path] = f"ERROR: {_e}"
            return []

    all_rows = []
    seen = set()
    with ThreadPoolExecutor(max_workers=4) as ex:
        futs = {ex.submit(_query_db, db): db for db in DBS_FACT}
        for f in futs:
            try:
                for r in f.result(timeout=25):
                    key = (str(r[3]).strip(), str(r[4]).strip(), str(r[0]).strip())
                    if key not in seen:
                        seen.add(key)
                        all_rows.append(r)
            except FutureTimeout:
                pass

    def _fmt(v):
        if hasattr(v, 'strftime'):
            return v.strftime('%Y-%m-%d')
        return str(v)[:10] if v else ''

    all_rows.sort(key=lambda r: (_fmt(r[2]), str(r[4] or '')), reverse=True)

    result = []
    for r in all_rows:
        try:
            tipo    = (r[3] or '').strip().upper()
            signo   = -1 if tipo in TIPOS_NC else 1
            cant    = signo * float(r[5] or 0)
            pu      = float(r[6] or 0)          # precio unitario siempre positivo
            subtot  = signo * float(r[7] or 0)
            iva_pct = float(r[8] or 21)
            total   = round(subtot * (1 + iva_pct / 100), 2)
            result.append({
                "cod_articulo":    (r[0] or '').strip(),
                "descripcion":     (r[1] or '').strip(),
                "fecha":           _fmt(r[2]),
                "tipo":            tipo,
                "numero":          str(r[4] or '').strip().replace('.0',''),
                "cantidad":        cant,
                "precio_unitario": pu,
                "importe":         subtot,
                "iva_pct":         iva_pct,
                "total":           total,
            })
        except Exception:
            pass   # fila con datos inválidos: se omite
    return result


@app.get("/que-vendi/pdf")
def que_vendi_pdf(vendedor: str, cliente: str,
                  desde: Optional[str] = None, hasta: Optional[str] = None,
                  razon: Optional[str] = None):
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from io import BytesIO
    from datetime import datetime

    rows = que_vendi(vendedor=vendedor, cliente=cliente, desde=desde, hasta=hasta, limit=2000)

    buf = BytesIO()
    # Landscape A4: usable width = 297 - 24 = 273 mm
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12*mm, rightMargin=12*mm,
                            topMargin=10*mm, bottomMargin=10*mm)

    azul = colors.HexColor('#1e429f')
    AR   = lambda v: f"{float(v or 0):,.2f}".replace(',','X').replace('.',',').replace('X','.')
    sDesc= ParagraphStyle('desc', fontSize=7, fontName='Helvetica',    leading=8)
    sHdr = ParagraphStyle('hdr',  fontSize=7, fontName='Helvetica-Bold',
                          textColor=colors.white, leading=8, alignment=1)  # 1=CENTER
    sSub = ParagraphStyle('sub',  fontSize=8, fontName='Helvetica', leading=10)
    sTit = ParagraphStyle('tit',  fontSize=12, fontName='Helvetica-Bold', leading=14)

    # ── Encabezado: logo izquierda + título derecha ─────────────────────────
    per = f"{desde or ''} a {hasta or ''}"
    emi = datetime.now().strftime('%d/%m/%Y %H:%M')
    logo_cell = ''
    if os.path.exists(LOGO_PATH):
        logo_cell = Image(LOGO_PATH, width=38*mm, height=12*mm,
                          kind='proportional')
    hdr_table = Table(
        [[logo_cell,
          [Paragraph('Microbell S.A. — ¿Qué Vendí?', sTit),
           Paragraph(f'Cliente: {razon or cliente}', sSub),
           Paragraph(f'Período: {per}   |   Emisión: {emi}', sSub)]]],
        colWidths=[42*mm, 231*mm]
    )
    hdr_table.setStyle(TableStyle([
        ('VALIGN',  (0,0),(-1,-1), 'MIDDLE'),
        ('ALIGN',   (1,0),(1,0),   'LEFT'),
        ('LEFTPADDING',  (0,0),(-1,-1), 0),
        ('RIGHTPADDING', (0,0),(-1,-1), 4),
        ('TOPPADDING',   (0,0),(-1,-1), 0),
        ('BOTTOMPADDING',(0,0),(-1,-1), 4),
    ]))

    # ── Columnas: suma = 273 mm ──────────────────────────────────────────────
    # CodArt Descripcion Fecha  Tipo  Nro   Cant  PUnit  Importe IVA  Total
    cw = [16*mm, 100*mm, 18*mm, 12*mm, 24*mm, 13*mm, 22*mm, 22*mm, 10*mm, 22*mm]
    # Encabezados como Paragraph para que también hagan wrap si es necesario
    hdrs = [Paragraph(h, sHdr) for h in
            ['Cód.Art.','Descripción','Fecha','Tipo','Nro. Comp.','Cant.','P.Unit.','Importe','IVA%','Total']]
    data = [hdrs]
    tot_imp = tot_tot = 0.0
    for r in rows:
        nro = str(r['numero']).replace('.0','').zfill(10)
        tot_imp += float(r['importe'] or 0)
        tot_tot += float(r['total']   or 0)
        fecha = r['fecha'][8:10]+'/'+r['fecha'][5:7]+'/'+r['fecha'][:4] if r['fecha'] else ''
        data.append([
            r['cod_articulo'],
            Paragraph(r['descripcion'] or '', sDesc),   # ← wrap automático
            fecha, r['tipo'], nro,
            str(int(float(r['cantidad'] or 0))), f"${AR(r['precio_unitario'])}",
            f"${AR(r['importe'])}", AR(r['iva_pct']), f"${AR(r['total'])}",
        ])
    data.append(['','','','','','',
                 Paragraph('TOTALES', ParagraphStyle('tb', fontSize=7, fontName='Helvetica-Bold')),
                 f"${AR(tot_imp)}", '', f"${AR(tot_tot)}"])

    t = Table(data, colWidths=cw, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',   (0,0), (-1,0),  azul),
        ('ROWBACKGROUNDS',(0,1),(-1,-2), [colors.white, colors.HexColor('#f3f4f6')]),
        ('BACKGROUND',   (0,-1),(-1,-1), colors.HexColor('#e0e7ff')),
        ('FONTNAME',     (0,1), (-1,-1), 'Helvetica'),
        ('FONTNAME',     (0,-1),(-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE',     (0,0), (-1,-1), 7),
        ('ALIGN',        (0,0), (-1,-1),  'LEFT'),
        ('ALIGN',        (0,0), (-1,0),  'CENTER'),   # toda la fila de headers
        ('ALIGN',        (0,1), (0,-1),  'CENTER'),   # Cód.Art. (datos)
        ('ALIGN',        (2,0), (2,-1),  'CENTER'),   # Fecha
        ('ALIGN',        (3,0), (3,-1),  'CENTER'),   # Tipo
        ('ALIGN',        (4,0), (4,-1),  'CENTER'),   # Nro
        ('ALIGN',        (5,0), (5,-1),  'CENTER'),   # Cant: centrado
        ('ALIGN',        (6,0), (-1,-1), 'RIGHT'),    # P.Unit → Total
        ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
        ('GRID',         (0,0), (-1,-1), 0.3, colors.HexColor('#d1d5db')),
        ('TOPPADDING',   (0,0), (-1,-1), 2),
        ('BOTTOMPADDING',(0,0), (-1,-1), 2),
        ('LEFTPADDING',  (0,0), (-1,-1), 3),
        ('RIGHTPADDING', (0,0), (-1,-1), 3),
    ]))

    story = [hdr_table, t]
    doc.build(story)
    buf.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(buf, media_type='application/pdf',
        headers={'Content-Disposition': f'inline; filename="que_vendi_{cliente}.pdf"'})


@app.get("/que-vendi/excel")
def que_vendi_excel(vendedor: str, cliente: str,
                    desde: Optional[str] = None, hasta: Optional[str] = None,
                    razon: Optional[str] = None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    rows = que_vendi(vendedor=vendedor, cliente=cliente, desde=desde, hasta=hasta, limit=2000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Que Vendi'

    azul  = PatternFill('solid', fgColor='1e429f')
    gris  = PatternFill('solid', fgColor='f3f4f6')
    azulc = PatternFill('solid', fgColor='e0e7ff')
    bF    = Font(bold=True, color='FFFFFF', size=9)
    bN    = Font(bold=True, size=9)
    nN    = Font(size=9)
    cen   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    der   = Alignment(horizontal='right',  vertical='center')
    izq   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin  = Side(style='thin', color='d1d5db')
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

    hdrs  = ['Cód.Art.','Descripción','Fecha','Tipo','Nro. Comp.','Cant.','P.Unit.','Importe','IVA%','Total']
    widths= [12,        45,           12,     8,     14,          10,     16,       16,       8,     16]
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = azul; cell.font = bF; cell.alignment = cen; cell.border = brd
        ws.column_dimensions[get_column_letter(ci)].width = w

    tot_imp = tot_tot = 0.0
    for ri, r in enumerate(rows, 2):
        nro = str(r['numero']).replace('.0','').zfill(10)
        fg  = None if ri % 2 == 0 else gris
        fecha = r['fecha'][8:10]+'/'+r['fecha'][5:7]+'/'+r['fecha'][:4] if r['fecha'] else ''
        vals = [r['cod_articulo'], r['descripcion'], fecha, r['tipo'], nro,
                float(r['cantidad'] or 0), float(r['precio_unitario'] or 0),
                float(r['importe'] or 0), float(r['iva_pct'] or 0), float(r['total'] or 0)]
        tot_imp += float(r['importe'] or 0); tot_tot += float(r['total'] or 0)
        aligns = [cen, izq, cen, cen, cen, cen, der, der, der, der]
        for ci, (v, al) in enumerate(zip(vals, aligns), 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = nN; cell.alignment = al; cell.border = brd
            if fg: cell.fill = fg
            if ci == 6:              cell.number_format = '#,##0'           # Cant: entero
            elif ci in (7, 8, 10): cell.number_format = '"$"#,##0.00'    # P.Unit/Importe/Total: $
            elif ci >= 7:          cell.number_format = '#,##0.00'        # IVA%: decimal

    tr = len(rows) + 2
    for ci, v in enumerate(['']*6 + ['TOTALES', tot_imp, '', tot_tot], 1):
        cell = ws.cell(row=tr, column=ci, value=v)
        cell.fill = azulc; cell.font = bN; cell.border = brd
        cell.alignment = der if ci >= 6 else cen
        if ci in (8, 10): cell.number_format = '"$"#,##0.00'

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="que_vendi_{cliente}.xlsx"'})


# ══════════════════════════════════════════════════════════════════════════════
#  ANÁLISIS DE VENTAS (admin) — /ventas/*
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/ventas/vendedores")
def ventas_vendedores(_u=Depends(get_admin_user)):
    """Lista de vendedores activos para el combo del panel de Ventas."""
    try:
        c = conn('WIN1252', DATABASE)
        cur = c.cursor()
        cur.execute(
            "SELECT CODIGOUSUARIO, RAZONSOCIAL FROM \"USUARIOS\" "
            "WHERE ACTIVO='1' AND UPPER(TRIM(ESVENDEDOR))='1' ORDER BY RAZONSOCIAL"
        )
        rows = cur.fetchall()
        c.close()
        return [{"codigo": (r[0] or '').strip(), "nombre": (r[1] or '').strip()} for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/ventas/clientes")
def ventas_clientes(vendedor: Optional[str] = None, buscar: Optional[str] = None,
                    _u=Depends(get_admin_user)):
    """Clientes para autocomplete. Si vendedor dado, filtra por CODIGOVENDEDOR."""
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    params = []
    where_parts = ["ACTIVO = '1'"]
    if vendedor:
        where_parts.append("UPPER(CODIGOVENDEDOR) = ?")
        params.append(vendedor.upper())
    if buscar:
        where_parts.append(
            "(UPPER(RAZONSOCIAL) CONTAINING UPPER(?) OR CODIGOCLIENTE CONTAINING ? OR CODIGOPARTICULAR CONTAINING ?)"
        )
        params += [buscar, buscar, buscar]
    where_sql = " AND ".join(where_parts)
    try:
        c = conn('WIN1252', DB_PROD)
        cur = c.cursor()
        cur.execute(
            f"SELECT FIRST 40 CODIGOCLIENTE, RAZONSOCIAL, CODIGOPARTICULAR "
            f"FROM \"CLIENTES\" WHERE {where_sql} ORDER BY RAZONSOCIAL",
            params
        )
        rows = cur.fetchall()
        c.close()
        return [{"codigo": (r[0] or '').strip(),
                 "razonsocial": (r[1] or '').strip(),
                 "codigoparticular": (r[2] or '').strip()} for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def _ventas_query(vendedor: Optional[str], cliente: Optional[str],
                  desde: Optional[str], hasta: Optional[str], limit: int = 2000,
                  grupo: Optional[str] = None, superrubro: Optional[str] = None,
                  rubro: Optional[str] = None, marca: Optional[str] = None,
                  articulo: Optional[str] = None):
    """Core query reutilizable por /ventas, /ventas/pdf y /ventas/excel."""
    # Cuando hay filtros de categoría el límite se aplica ANTES de filtrar
    # por catálogo, por lo que se necesita traer TODOS los registros del período.
    if grupo or superrubro or rubro or marca or articulo:
        limit = 100_000
    from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeout
    DB_PROD     = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    DBS_FACT    = [DATABASE, DATABASE_MLT, DB_PROD, DB_MLT_PROD]
    DBS_LOOKUP  = [DATABASE, DB_PROD]

    TIPOS_FACTURA = ("('FA','FB','FE','FCA','FCB','FCE',"
                     "'FCCA','FCCB','FCCE',"
                     "'NCA','NCB','NCCA','NCCB')")
    TIPOS_NC = {'NCA', 'NCB', 'NCCA', 'NCCB'}

    # Resolver códigos del cliente si se especificó
    codigos_cliente = []
    if cliente:
        def _lookup(db_path):
            try:
                c = conn('WIN1252', db_path)
                cur = c.cursor()
                cur.execute(
                    'SELECT CODIGOCLIENTE, CODIGOPARTICULAR FROM "CLIENTES" '
                    'WHERE CODIGOCLIENTE = ? OR CODIGOPARTICULAR = ?',
                    (cliente, cliente)
                )
                row = cur.fetchone()
                c.close()
                if row:
                    return [str(v).strip() for v in row if v is not None and str(v).strip()]
            except Exception:
                pass
            return []

        codigos_set = set([str(cliente).strip()])
        with ThreadPoolExecutor(max_workers=2) as ex:
            for r in ex.map(_lookup, DBS_LOOKUP):
                codigos_set.update(r)
        codigos_cliente = list(codigos_set)

    # WHERE dinámico
    where_parts = [f"cb.ANULADA = '0'", f"cb.TIPOCOMPROBANTE IN {TIPOS_FACTURA}"]
    params_extra = []
    if codigos_cliente:
        ph = ','.join('?' * len(codigos_cliente))
        where_parts.append(f"cb.CODIGOCLIENTE IN ({ph})")
        params_extra += codigos_cliente
    if vendedor:
        where_parts.append("UPPER(TRIM(cb.CODIGOUSUARIO)) = ?")
        params_extra.append(vendedor.upper())
    if desde:
        where_parts.append("cb.FECHACOMPROBANTE >= ?")
        params_extra.append(desde)
    if hasta:
        where_parts.append("cb.FECHACOMPROBANTE <= ?")
        params_extra.append(hasta)
    where_sql = " AND ".join(where_parts)

    sql = f"""
        SELECT FIRST {limit} SKIP 0
            TRIM(cu.CODIGOARTICULO)                                                  AS COD_ART,
            COALESCE(NULLIF(TRIM(cu.DESCRIPCION),''), '')                            AS DESCR,
            cb.FECHACOMPROBANTE,
            cb.TIPOCOMPROBANTE,
            cb.NUMEROCOMPROBANTE,
            CAST(cu.CANTIDAD       AS DOUBLE PRECISION) AS CANT,
            CAST(cu.PRECIOUNITARIO AS DOUBLE PRECISION) AS PU,
            CAST(cu.PRECIOTOTAL    AS DOUBLE PRECISION) AS SUBTOTAL,
            COALESCE(CAST(cu.PORCENTAJEIVA AS DOUBLE PRECISION), 21) AS IVA_PCT,
            TRIM(cb.CODIGOCLIENTE)  AS COD_CLI,
            TRIM(cb.RAZONSOCIAL)    AS RAZON,
            TRIM(cb.CODIGOUSUARIO)  AS COD_VEND,
            CAST(cu.COSTOVENTA AS DOUBLE PRECISION) AS COSTO_VENTA,
            CAST(cb.TOTAL  AS DOUBLE PRECISION) AS CB_TOTAL,
            CAST(cb.IVA1   AS DOUBLE PRECISION) AS CB_IVA1,
            CAST(cb.IVA2   AS DOUBLE PRECISION) AS CB_IVA2,
            CAST(cb.PAGADO AS DOUBLE PRECISION) AS CB_PAGADO
        FROM "CUERPOCOMPROBANTES" cu
        JOIN "CABEZACOMPROBANTES" cb
             ON cb.TIPOCOMPROBANTE   = cu.TIPOCOMPROBANTE
            AND cb.NUMEROCOMPROBANTE = cu.NUMEROCOMPROBANTE
        WHERE {where_sql}
        ORDER BY cb.FECHACOMPROBANTE DESC, cb.NUMEROCOMPROBANTE DESC
    """

    def _query_db(db_path):
        try:
            c = conn('LATIN1', db=db_path)
            cur = c.cursor()
            cur.execute(sql, params_extra)
            rows = cur.fetchall()
            c.close()
            return rows
        except Exception:
            return []

    all_rows = []
    seen = set()
    with ThreadPoolExecutor(max_workers=4) as ex:
        futs = {ex.submit(_query_db, db): db for db in DBS_FACT}
        for f in futs:
            try:
                for r in f.result(timeout=25):
                    key = (str(r[3]).strip(), str(r[4]).strip(), str(r[0]).strip(), str(r[9] or '').strip())
                    if key not in seen:
                        seen.add(key)
                        all_rows.append(r)
            except FutureTimeout:
                pass

    def _fmt(v):
        if hasattr(v, 'strftime'):
            return v.strftime('%Y-%m-%d')
        return str(v)[:10] if v else ''

    all_rows.sort(key=lambda r: (_fmt(r[2]), str(r[4] or '')), reverse=True)

    result = []
    for r in all_rows:
        try:
            tipo   = (r[3] or '').strip().upper()
            signo  = -1 if tipo in TIPOS_NC else 1
            cant   = signo * float(r[5] or 0)
            pu     = float(r[6] or 0)
            subtot = signo * float(r[7] or 0)
            iva_pct= float(r[8] or 21)
            total  = round(subtot * (1 + iva_pct / 100), 2)
            # Porcentaje cobrado de la cabeza del comprobante
            cb_total  = float(r[13] or 0) + float(r[14] or 0) + float(r[15] or 0)
            cb_pagado = float(r[16] or 0)
            if tipo in TIPOS_NC:
                pct_cobrado = 1.0   # NC: comisión sobre el total (ya está aplicada a una FA)
            elif cb_total > 0:
                pct_cobrado = min(cb_pagado / cb_total, 1.0)
            else:
                pct_cobrado = 0.0
            result.append({
                "cod_articulo":    (r[0] or '').strip(),
                "descripcion":     (r[1] or '').strip(),
                "fecha":           _fmt(r[2]),
                "tipo":            tipo,
                "numero":          str(r[4] or '').strip().replace('.0', ''),
                "cantidad":        cant,
                "precio_unitario": pu,
                "importe":         subtot,
                "iva_pct":         iva_pct,
                "total":           total,
                "codigocliente":   (r[9]  or '').strip(),
                "razonsocial":     (r[10] or '').strip(),
                "codigovendedor":  (r[11] or '').strip(),
                "costo_venta":      float(r[12] or 0),
                "pct_cobrado":      round(pct_cobrado * 100, 1),
            })
        except Exception:
            pass

    # ── Enriquecimiento: costo, jerarquía y comisión ─────────────────────────
    try:
        DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
        catalog, cambio_usd = _get_catalog()

        # Índice inverso: por codigoparticular Y por codigoarticulo interno
        cat_by_part: dict = {}
        cat_by_art:  dict = {}
        for art_id, art in catalog.items():
            cp = art.get('codigoparticular', '').strip()
            if cp:
                cat_by_part[cp] = art
            cat_by_art[str(art_id).strip()] = art

        # Vendedor asociado al cliente (fallback para comisiones cuando
        # CODIGOUSUARIO del comprobante no matchea tabla de comisiones)
        cliente_vend_code = ''
        if cliente and codigos_cliente:
            for _db_cli in [DB_PROD, DATABASE]:
                try:
                    _c_cli = conn('WIN1252', db=_db_cli)
                    _cur_cli = _c_cli.cursor()
                    _ph_cli = ','.join(['?'] * len(codigos_cliente))
                    _cur_cli.execute(
                        f'SELECT FIRST 1 UPPER(TRIM(CODIGOVENDEDOR)) FROM "CLIENTES" '
                        f'WHERE CODIGOCLIENTE IN ({_ph_cli}) AND ACTIVO=\'1\'',
                        codigos_cliente)
                    _row_cli = _cur_cli.fetchone()
                    _c_cli.close()
                    if _row_cli and _row_cli[0]:
                        cliente_vend_code = (_row_cli[0] or '').strip().upper()
                        break
                except Exception:
                    pass

        # Costos — tomados de CUERPOCOMPROBANTES.COSTOVENTA (ya en cada row)

        # Nombres de vendedores (CODIGOUSUARIO → RAZONSOCIAL uppercase)
        # Busca en ambas bases para mayor cobertura
        vend_nombre_map: dict = {}
        for _db_vend in [DATABASE, DB_PROD]:
            try:
                _cu = conn('WIN1252', _db_vend)
                _curu = _cu.cursor()
                _curu.execute('SELECT CODIGOUSUARIO, RAZONSOCIAL FROM "USUARIOS"')
                for row in _curu.fetchall():
                    cu = (row[0] or '').strip().upper()
                    if cu and cu not in vend_nombre_map:
                        vend_nombre_map[cu] = (row[1] or '').strip().upper()
                _cu.close()
            except Exception:
                pass

        for r in result:
            cod = r['cod_articulo']
            # Buscar en catálogo: cod es CUERPOCOMPROBANTES.CODIGOARTICULO (código
            # interno), por eso cat_by_art tiene que probarse PRIMERO. Si se probaba
            # cat_by_part primero, un cod interno que coincidiera por casualidad con
            # el codigoparticular de OTRO artículo traía la jerarquía (rubro/super
            # rubro/grupo) de ese otro artículo — bug detectado en Reactivación con
            # un artículo de rubro JUGUETES mostrando rubro TERMOS.
            art = cat_by_art.get(cod) or cat_by_part.get(cod) or {}
            moneda = art.get('codigomoneda', '').upper()

            # Costo de venta — COSTOVENTA es el total de la línea (no unitario)
            _cant = abs(r['cantidad']) or 1
            r['costo_total']    = round(r['costo_venta'], 2)
            r['costo_unitario'] = round(r['costo_venta'] / _cant, 2)

            # Jerarquía de categorías (descripciones, para mostrar)
            gsr  = art.get('gruposuperrubro', '').upper()
            sr   = art.get('superrubro',      '').upper()
            rubr = art.get('rubro',           '').upper()
            r['gruposuperrubro'] = gsr
            r['superrubro']      = sr
            r['rubro']           = rubr
            # Códigos (para matchear contra offer_category_filters/filtros, que
            # guardan código — NUNCA la descripción de arriba, que es distinta).
            r['codigo_gruposuperrubro'] = art.get('codigo_gruposuperrubro', '')
            r['codigo_superrubro']      = art.get('codigo_superrubro', '')
            r['codigo_rubro']           = art.get('codigo_rubro', '')
            # No hay descripción de marca separada en el catálogo — el código de
            # marca ya es un nombre legible (ej. "HASBRO"), se usa tal cual.
            r['marca']                  = (art.get('codigomarca', '') or '').upper()
            r['codigomarca']            = art.get('codigomarca', '')

            # Comisión: usar vendedor del comprobante con fallback al vendedor del cliente
            vend_code   = r['codigovendedor'].upper() or cliente_vend_code
            vend_nombre = vend_nombre_map.get(vend_code, '')
            if not vend_nombre and cliente_vend_code and vend_code != cliente_vend_code:
                vend_nombre = vend_nombre_map.get(cliente_vend_code, '')
            pct = _COMISIONES.get((vend_nombre, gsr, sr, rubr), 0)
            pct_cob = r.get('pct_cobrado', 0) / 100.0  # 0.0–1.0
            r['comision_pct']    = pct
            # Comisión solo sobre la parte cobrada (NC siempre al 100%)
            r['comision']        = round(r['importe'] * pct / 100 * pct_cob, 2)
            r['vendedor_nombre'] = vend_nombre
    except Exception:
        # Si el enriquecimiento falla no rompemos el endpoint
        for r in result:
            r.setdefault('costo_unitario', 0)
            r.setdefault('costo_total', 0)
            r.setdefault('gruposuperrubro', '')
            r.setdefault('superrubro', '')
            r.setdefault('rubro', '')
            r.setdefault('codigo_gruposuperrubro', '')
            r.setdefault('codigo_superrubro', '')
            r.setdefault('codigo_rubro', '')
            r.setdefault('marca', '')
            r.setdefault('codigomarca', '')
            r.setdefault('comision_pct', 0)
            r.setdefault('comision', 0)
            r.setdefault('vendedor_nombre', '')

    # ── Filtro por categoría (post-enriquecimiento) ──────────────────────────
    if grupo or superrubro or rubro or marca or articulo:
        try:
            cat_filt, _ = _get_catalog()
            articulo_norm = articulo.upper().strip() if articulo else None
            matching_ids: set = set()
            for art_id, art in cat_filt.items():
                if grupo      and art.get('codigo_gruposuperrubro', '') != grupo:      continue
                if superrubro and art.get('codigo_superrubro',      '') != superrubro: continue
                if rubro      and art.get('codigo_rubro',           '') != rubro:      continue
                if marca      and art.get('codigomarca',            '') != marca:      continue
                if articulo_norm:
                    cp   = art.get('codigoparticular', '').upper()
                    desc = art.get('descripcion',      '').upper()
                    if articulo_norm not in cp and articulo_norm not in desc:
                        continue
                matching_ids.add(str(art_id).strip())
            antes = len(result)
            result = [r for r in result if r['cod_articulo'] in matching_ids]
            print(f"[VENTAS CATFILT] matching_ids={len(matching_ids)} antes={antes} despues={len(result)}")
        except Exception as e:
            print(f"[VENTAS CATFILT ERROR] {e}")

    return result


# ══════════════════════════════════════════════════════════════════════════════
#  REACTIVACIÓN DE CLIENTES (admin) — /admin/reactivacion/*
# ══════════════════════════════════════════════════════════════════════════════

# Cuentas de vendedor que NO son corredores reales (casillas internas/house accounts)
# — nunca deben generar detección ni notificación de reactivación de clientes.
_REACTIVACION_VENDEDORES_EXCLUIDOS = {
    'ECOMMERCE', 'JAVIER DIPPOLITO', 'DIPPOLITO JAVIER', 'FERNANDO LAJE', 'LAJE FERNANDO',
    'MICROBELL', 'TTC', 'ADMINISTRACION DEL SISTEMA', 'ADMINISTRADOR SISTEMA',
    'EDUARDO MORENO', 'GRECO EMILIANO', 'MARISA EMILIANA ACOSTA', 'MAZZUCHELI GERMAN',
    'ROSSI JUAN MANUEL', 'VENDEDOR WEB'
}

# Notas de crédito: no cuentan ni como facturación (Periodo A) ni como actividad
# reciente (Periodo B) para este análisis — a diferencia de /ventas, que las resta
# del neto, acá se ignoran directamente (no son una venta nueva al cliente).
_REACTIVACION_TIPOS_NC = {'NCA', 'NCB', 'NCCA', 'NCCB', 'NCE'}


def _reactivacion_clientes_vendedor_map() -> dict:
    """CODIGOCLIENTE -> {vendedor, vendedor_nombre, vendedor_activo}. Se usa para
    resolver a qué corredor pertenece cada cliente (CLIENTES.CODIGOVENDEDOR),
    independiente de quién haya procesado cada factura histórica. Solo incluye
    clientes cuyo vendedor asignado tiene perfil VENDEDORES y está ACTIVO en
    Flexxus, y excluye además cuentas de _REACTIVACION_VENDEDORES_EXCLUIDOS
    (casillas de sistema/house accounts que puedan colarse con ese perfil)."""
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    out = {}
    try:
        c = conn('WIN1252', DB_PROD)
        cur = c.cursor()
        cur.execute("""
            SELECT cl.CODIGOCLIENTE, TRIM(cl.CODIGOVENDEDOR), u.RAZONSOCIAL, u.ACTIVO, u.CODIGOPERFIL,
                   TRIM(cl.CODIGOPARTICULAR)
            FROM "CLIENTES" cl
            LEFT JOIN "USUARIOS" u ON u.CODIGOUSUARIO = cl.CODIGOVENDEDOR
            WHERE cl.ACTIVO = '1'
        """)
        for r in cur.fetchall():
            cod_cli = str(r[0] or '').strip()
            if not cod_cli:
                continue
            vendedor_nombre = (r[2] or '').strip()
            vendedor_activo = str(r[3] or '0').strip() == '1'
            vendedor_perfil = str(r[4] or '').strip().upper()
            if vendedor_nombre.upper() in _REACTIVACION_VENDEDORES_EXCLUIDOS:
                continue
            if vendedor_perfil != 'VENDEDORES' or not vendedor_activo:
                continue
            out[cod_cli] = {
                'vendedor':        (r[1] or '').strip().upper(),
                'vendedor_nombre': vendedor_nombre,
                'vendedor_activo': vendedor_activo,
                'codigoparticular': (r[5] or '').strip() or cod_cli,
            }
        c.close()
    except Exception as e:
        print(f"[REACTIVACION] error mapa cliente->vendedor: {e}")
    return out


def _reactivacion_calcular(cfg: dict) -> list:
    """Corre UN análisis de reactivación y devuelve hasta 10 clientes POR VENDEDOR
    (no 10 en total — si el análisis cubre "Todos" los vendedores activos, cada uno
    recibe su propio top-10, no compiten entre sí por un cupo global), cada uno con
    importe_total (Periodo A, ya filtrado por vendedor/categoría) y su detalle línea
    a línea (factura+artículo), para clientes que además no tuvieron NINGUNA venta
    (de ningún rubro) en los últimos N días.
    No persiste nada — eso lo hace el llamador (endpoint /ejecutar)."""
    from datetime import date, timedelta

    gsr           = (cfg.get('gruposuperrubro') or '').strip() or None
    sr            = (cfg.get('superrubro') or '').strip() or None
    rubro         = (cfg.get('rubro') or '').strip() or None
    periodo_a_desde = (cfg.get('periodo_a_desde') or '').strip() or None
    periodo_a_hasta = (cfg.get('periodo_a_hasta') or '').strip() or None
    periodo_b_dias  = int(cfg.get('periodo_b_dias') or 60)
    vendedor_filtro = (cfg.get('vendedor_codigo') or '').strip().upper() or None

    fecha_b_desde = (date.today() - timedelta(days=periodo_b_dias)).isoformat()

    # Periodo A: líneas de facturación filtradas por categoría (el vendedor "dueño"
    # del cliente se resuelve aparte, vía CLIENTES.CODIGOVENDEDOR — no por quién
    # procesó cada factura, que es lo que filtraría vendedor= en _ventas_query).
    lineas_a = _ventas_query(vendedor=None, cliente=None,
                              desde=periodo_a_desde, hasta=periodo_a_hasta,
                              grupo=gsr, superrubro=sr, rubro=rubro)

    clientes_vend = _reactivacion_clientes_vendedor_map()

    por_cliente = {}
    for r in lineas_a:
        if (r.get('tipo') or '').strip().upper() in _REACTIVACION_TIPOS_NC:
            continue  # Notas de crédito: no cuentan como facturación para este análisis
        cod_cli = r.get('codigocliente') or ''
        if not cod_cli:
            continue
        info = clientes_vend.get(cod_cli)
        if not info or not info.get('vendedor_activo') or not info.get('vendedor'):
            continue  # cliente sin vendedor activo asignado: fuera de alcance
        if vendedor_filtro and info['vendedor'] != vendedor_filtro:
            continue
        entry = por_cliente.get(cod_cli)
        if entry is None:
            entry = {
                'codigocliente':    cod_cli,
                'codigoparticular': info.get('codigoparticular') or cod_cli,
                'razonsocial':      r.get('razonsocial') or '',
                'vendedor_codigo':  info['vendedor'],
                'vendedor_nombre':  info.get('vendedor_nombre', ''),
                'importe_total':    0.0,
                'detalle':          [],
            }
            por_cliente[cod_cli] = entry
        entry['importe_total'] += r.get('total', 0)
        entry['detalle'].append({
            'tipo_comprobante':      r.get('tipo', ''),
            'numero_comprobante':    r.get('numero', ''),
            'fecha':                 r.get('fecha', ''),
            'codigo_articulo':       r.get('cod_articulo', ''),
            'descripcion_articulo':  r.get('descripcion', ''),
            'rubro':                 r.get('rubro', ''),
            'superrubro':            r.get('superrubro', ''),
            # Códigos (no descripción) — para matchear contra offer_category_filters
            # y contra el catálogo de stock (codigo_rubro/codigomarca), que usan
            # código, no la descripción de arriba.
            'codigo_rubro':          r.get('codigo_rubro', ''),
            'codigo_superrubro':     r.get('codigo_superrubro', ''),
            'codigomarca':           r.get('codigomarca', ''),
            'cantidad':              r.get('cantidad', 0),
            'importe':               r.get('total', 0),
        })

    if not por_cliente:
        return []

    # Periodo B: actividad GLOBAL (sin filtro de categoría) — un cliente que sigue
    # comprando otras cosas no cuenta como "inactivo", aunque haya dejado de
    # comprar específicamente en el rubro analizado. Las notas de crédito tampoco
    # cuentan como "actividad" (no son una compra nueva).
    lineas_b = _ventas_query(vendedor=None, cliente=None, desde=fecha_b_desde, hasta=None)
    clientes_activos_b = set(
        r.get('codigocliente') for r in lineas_b
        if r.get('codigocliente') and (r.get('tipo') or '').strip().upper() not in _REACTIVACION_TIPOS_NC
    )

    candidatos = [e for cod, e in por_cliente.items() if cod not in clientes_activos_b]

    # Top 10 POR VENDEDOR, no un top-10 global — así un análisis "Todos los
    # vendedores" le da a cada corredor sus propios 10 clientes de mayor
    # facturación inactivos, en vez de que unos pocos vendedores con clientes
    # grandes acaparen el único cupo de 10 de todo el análisis.
    por_vendedor = {}
    for e in candidatos:
        por_vendedor.setdefault(e['vendedor_codigo'], []).append(e)

    resultado = []
    for vend_cod, lst in por_vendedor.items():
        lst.sort(key=lambda e: e['importe_total'], reverse=True)
        resultado.extend(lst[:10])

    for e in resultado:
        e['importe_total'] = round(e['importe_total'], 2)
    return resultado


def _reactivacion_ofertas_sugeridas(rubros_vistos, superrubros_vistos, codigos_articulo_vistos=None, marcas_vistas=None) -> list:
    """Ofertas activas (admin.db) que aplican a lo que el cliente compraba antes —
    ya sea porque el ALCANCE DE CATEGORÍA de la oferta (rubro/superrubro/marca)
    coincide, o porque alguno de los artículos que compró está puntualmente cargado
    en la oferta (offer_product_details) — muchas promos (ej. "Bonifacion Jugueteria")
    se cargan por SKU y no por rubro, así que matchear solo por categoría las deja
    afuera. Para cada oferta devuelve también el % de bonificación representativo
    y la condición comercial asociada (con su descripción resuelta en Firebird),
    para que el PDF pueda mostrarlo. No valida stock remanente en vivo.
    IMPORTANTE: rubros_vistos/superrubros_vistos/marcas_vistas van por CÓDIGO
    (codigo_rubro/codigo_superrubro/codigomarca), no por descripción."""
    import json as _json
    from datetime import date
    rubros_vistos       = sorted(set(r for r in (rubros_vistos or []) if r))
    superrubros_vistos  = sorted(set(s for s in (superrubros_vistos or []) if s))
    marcas_vistas        = sorted(set(m for m in (marcas_vistas or []) if m))
    codigos_articulo    = set(a for a in (codigos_articulo_vistos or []) if a)
    if not rubros_vistos and not superrubros_vistos and not marcas_vistas and not codigos_articulo:
        return []
    # offer_product_details.codigo_producto se carga como CODIGOPARTICULAR (así lo
    # guarda el buscador de artículos del modal de Ofertas), mientras que
    # codigos_articulo acá viene del CODIGOARTICULO interno (CUERPOCOMPROBANTES) —
    # son esquemas de código distintos que pueden coincidir por casualidad (mismo
    # bug de colisión que en _ventas_query). Traducimos a codigoparticular antes de
    # matchear para no sugerir ofertas de un artículo distinto al que compró.
    if codigos_articulo:
        try:
            catalog, _ = _get_catalog()
            traducidos = set()
            for cod in codigos_articulo:
                art = catalog.get(cod)
                traducidos.add((art.get('codigoparticular') or cod) if art else cod)
            codigos_articulo = traducidos
        except Exception:
            pass
    hoy = date.today().isoformat()
    c = _admin_db()
    ofertas_activas = c.execute("""
        SELECT id, nombre, descripcion, fecha_hasta, financial_escalones
        FROM offers
        WHERE activo = 1 AND (fecha_hasta = '' OR fecha_hasta IS NULL OR fecha_hasta >= ?)
    """, (hoy,)).fetchall()

    matched = {}  # offer_id -> dict base
    if rubros_vistos or superrubros_vistos or marcas_vistas:
        conds, params = [], []
        if rubros_vistos:
            conds.append(f"(f.nivel='rubro' AND f.valor IN ({','.join('?'*len(rubros_vistos))}))")
            params += rubros_vistos
        if superrubros_vistos:
            conds.append(f"(f.nivel='superrubro' AND f.valor IN ({','.join('?'*len(superrubros_vistos))}))")
            params += superrubros_vistos
        if marcas_vistas:
            conds.append(f"(f.nivel='marca' AND f.valor IN ({','.join('?'*len(marcas_vistas))}))")
            params += marcas_vistas
        for r in c.execute(f"""
            SELECT DISTINCT offer_id FROM offer_category_filters f WHERE {' OR '.join(conds)}
        """, params).fetchall():
            matched[r['offer_id']] = None

    prod_pct_por_oferta = {}  # offer_id -> lista de bonificacion_pct de artículos que el cliente compró
    if codigos_articulo:
        ph = ','.join('?' * len(codigos_articulo))
        for r in c.execute(f"""
            SELECT offer_id, codigo_producto, bonificacion_pct FROM offer_product_details
            WHERE codigo_producto IN ({ph})
        """, list(codigos_articulo)).fetchall():
            matched[r['offer_id']] = None
            prod_pct_por_oferta.setdefault(r['offer_id'], []).append(r['bonificacion_pct'] or 0)

    if not matched:
        c.close()
        return []

    # Condición comercial cargada directo en la oferta (offer_conditions) — fallback
    # si el escalón financiero no trae una propia.
    cond_por_oferta = {}
    ids = list(matched.keys())
    ph_ids = ','.join('?' * len(ids))
    for r in c.execute(f"SELECT offer_id, condicion_comercial FROM offer_conditions WHERE offer_id IN ({ph_ids})", ids).fetchall():
        cond_por_oferta.setdefault(r['offer_id'], r['condicion_comercial'])
    c.close()

    resultado = []
    codigos_condicion = set()
    for o in ofertas_activas:
        oid = o['id']
        if oid not in matched:
            continue
        pct = 0.0
        cond_codigo = cond_por_oferta.get(oid)
        escalones_info = []
        fe_raw = o['financial_escalones']
        if fe_raw:
            try:
                escalones = _json.loads(fe_raw)
                if escalones:
                    # Se informan TODOS los escalones (no solo el más accesible) — cada
                    # uno con su propio monto mínimo y % acumulado, para que el vendedor
                    # sepa qué gana en cada nivel de facturación, no solo en el primero.
                    escalones = sorted(escalones, key=lambda e: e.get('monto_minimo', 0))
                    for esc in escalones:
                        escalones_info.append({
                            'monto_minimo': esc.get('monto_minimo', 0),
                            'porcentajes': [float(p) for p in (esc.get('porcentajes') or [])],
                            'condicion_comercial': esc.get('condicion_comercial'),
                        })
                        if esc.get('condicion_comercial'):
                            codigos_condicion.add(str(esc['condicion_comercial']))
                    if escalones[0].get('condicion_comercial'):
                        cond_codigo = escalones[0]['condicion_comercial']
            except Exception:
                pass
        if not escalones_info and prod_pct_por_oferta.get(oid):
            pct = round(max(prod_pct_por_oferta[oid]), 2)
        if cond_codigo:
            codigos_condicion.add(str(cond_codigo))
        resultado.append({
            'id': oid, 'nombre': o['nombre'], 'descripcion': o['descripcion'],
            'fecha_hasta': o['fecha_hasta'], 'bonificacion_pct': pct, 'escalones': escalones_info,
            'condicion_codigo': cond_codigo, 'condicion_desc': None,
        })

    # Resolver descripción de las condiciones comerciales en Firebird (una sola conexión)
    if codigos_condicion:
        try:
            fb = conn('WIN1252')
            cur = fb.cursor()
            ph = ','.join('?' * len(codigos_condicion))
            desc_map = {}
            cur.execute(f'SELECT CODIGOMULTIPLAZO, DESCRIPCION FROM "MULTIPLAZOS" WHERE CODIGOMULTIPLAZO IN ({ph})', list(codigos_condicion))
            for row in cur.fetchall():
                desc_map[str(row[0]).strip()] = str(row[1] or '').strip()
            fb.close()
            for o in resultado:
                if o['condicion_codigo']:
                    o['condicion_desc'] = desc_map.get(str(o['condicion_codigo']).strip())
                for esc in o.get('escalones') or []:
                    if esc.get('condicion_comercial'):
                        esc['condicion_desc'] = desc_map.get(str(esc['condicion_comercial']).strip())
        except Exception:
            pass

    return resultado


def _reactivacion_generar_pdf(analisis_nombre: str, vendedor_nombre: str, clientes: list, ofertas_sugeridas: list,
                               periodo_a_desde: str = None, periodo_a_hasta: str = None,
                               descuento_pct: float = 0, oferta_vencimiento: str = None,
                               periodo_b_dias: int = None, disponibilidad: dict = None,
                               descuento_monto_minimo: float = 0) -> bytes:
    """PDF para el corredor: código, razón social e importe total neto facturado por
    cliente (orden desc) + por cada cliente con stock disponible, el detalle de esos
    artículos (código/descripción/remanente/bonificación %, si tiene) con aviso de
    urgencia porque el stock no está reservado + sección de ofertas generales
    sugeridas + párrafo de descuento adicional propio de este análisis (solo si se
    cargó uno). El detalle completo de facturas queda en el modal de auditoría."""
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from io import BytesIO

    def _ddmmyyyy(s):
        s = (s or '').strip()
        if not s:
            return None
        try:
            return datetime.strptime(s[:10], '%Y-%m-%d').strftime('%d/%m/%Y')
        except Exception:
            return s  # ya viene en otro formato: se muestra tal cual antes que romper

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=12*mm, bottomMargin=12*mm)
    azul = colors.HexColor('#1e429f')
    AR = lambda v: f"{float(v or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    sTit = ParagraphStyle('tit', fontSize=13, fontName='Helvetica-Bold', leading=16)
    sSub = ParagraphStyle('sub', fontSize=9,  fontName='Helvetica',      leading=12)
    sHdr = ParagraphStyle('hdr', fontSize=8,  fontName='Helvetica-Bold', textColor=colors.white, leading=10)

    story = []
    if os.path.exists(LOGO_PATH):
        story.append(Image(LOGO_PATH, width=38*mm, height=12*mm, kind='proportional'))
    story.append(Spacer(1, 6))
    story.append(Paragraph(f'Reactivación de clientes — {analisis_nombre}', sTit))
    story.append(Paragraph(f'Vendedor: {vendedor_nombre}', sSub))
    story.append(Paragraph(f'Emitido: {datetime.now().strftime("%d/%m/%Y %H:%M")}', sSub))
    story.append(Spacer(1, 10))

    primer_nombre = (vendedor_nombre or '').split(' ')[0].title()
    if periodo_a_desde or periodo_a_hasta:
        periodo_txt = f'el período {_ddmmyyyy(periodo_a_desde) or "…"} a {_ddmmyyyy(periodo_a_hasta) or "…"}'
    else:
        periodo_txt = 'toda la historia'
    inactividad_txt = f'en los últimos {int(periodo_b_dias)} días' if periodo_b_dias else 'en el período de inactividad analizado'
    story.append(Paragraph(
        f'Estimado {primer_nombre}: observamos que en {periodo_txt} se registraron ventas a los '
        f'siguientes clientes tuyos, que no registran operaciones {inactividad_txt}. '
        'Te dejamos el detalle:', sSub))
    story.append(Spacer(1, 10))

    data = [[Paragraph('Código', sHdr), Paragraph('Cliente', sHdr), Paragraph('Facturación total', sHdr)]]
    for cl in clientes:
        data.append([cl.get('codigoparticular') or cl['codigocliente'], cl['razonsocial'], f"${AR(cl['importe_total'])}"])
    t = Table(data, colWidths=[30*mm, 100*mm, 50*mm])
    t.setStyle(TableStyle([
        ('BACKGROUND',      (0,0), (-1,0), azul),
        ('FONTNAME',        (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE',        (0,0), (-1,-1), 9),
        ('ALIGN',           (2,0), (2,-1), 'RIGHT'),
        ('GRID',            (0,0), (-1,-1), 0.3, colors.HexColor('#d1d5db')),
        ('ROWBACKGROUNDS',  (0,1), (-1,-1), [colors.white, colors.HexColor('#f3f4f6')]),
        ('TOPPADDING',      (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',   (0,0), (-1,-1), 4),
    ]))
    story.append(t)
    story.append(Spacer(1, 10))

    # Por cliente, artículos con stock disponible AHORA (remanente real, no
    # reservado) dentro de los mismos rubros que ya le compraba — no se limita a
    # los SKU puntuales facturados antes, para ofrecer variedad real.
    disponibilidad = disponibilidad or {}
    clientes_con_stock = [cl for cl in clientes if disponibilidad.get(cl['codigocliente'])]
    if clientes_con_stock:
        # Lista única de artículos (no se repite por cliente ni se muestra el
        # cliente): solo código, descripción, precio unit. y bonificación activa,
        # sin la cantidad de remanente.
        vistos = {}
        for cl in clientes_con_stock:
            for it in disponibilidad[cl['codigocliente']]:
                if it['codigo'] not in vistos:
                    vistos[it['codigo']] = it
        items_unicos = sorted(vistos.values(), key=lambda x: (-(x.get('bonificacion_pct') or 0), x['descripcion']))

        story.append(Paragraph('Artículos con stock disponible para ofrecerles ya mismo', ParagraphStyle(
            'h2', fontSize=10, fontName='Helvetica-Bold', textColor=azul, leading=13)))
        story.append(Paragraph(
            'Urgente: este remanente no está reservado — se puede agotar por otras ventas antes de que '
            'contactes al cliente. Si el artículo tiene bonificación activa, se pierde la oportunidad si no '
            'se concreta a tiempo.', ParagraphStyle('warn', fontSize=8.5, fontName='Helvetica-Oblique',
                                                     textColor=colors.HexColor('#b91c1c'), leading=11)))
        story.append(Spacer(1, 6))
        for it in items_unicos:
            partes = [f"{it['codigo']} — {it['descripcion']}", f"precio unit. ${AR(it['precio'])}"]
            if it.get('bonificacion_pct'):
                partes.append(f"bonificación activa {it['bonificacion_pct']:g}%")
            story.append(Paragraph('&nbsp;&nbsp;• ' + ' — '.join(partes), sSub))
        story.append(Spacer(1, 6))

    story.append(Paragraph(
        'Te sugerimos volver a comunicarte con ellos. Te informamos que contamos con las siguientes '
        'promociones vigentes para ofrecerles:' if ofertas_sugeridas else
        'Te sugerimos volver a comunicarte con ellos.', sSub))

    if ofertas_sugeridas:
        story.append(Spacer(1, 6))
        for o in ofertas_sugeridas:
            vto = _ddmmyyyy(o.get('fecha_hasta')) or 'sin vencimiento'
            desc = o.get('descripcion') or ''
            escalones = o.get('escalones') or []
            partes = [f'<b>{o["nombre"]}</b>']
            if desc:
                partes.append(desc)
            if not escalones and o.get('bonificacion_pct'):
                partes.append(f'bonificación {o["bonificacion_pct"]:g}%')
            if o.get('condicion_desc'):
                partes.append(f'condición de venta: {o["condicion_desc"]}')
            partes.append(f'vigente hasta {vto}')
            story.append(Paragraph('• ' + ' — '.join(partes), sSub))
            # Escalones por monto mínimo de facturación neta: se informa CADA nivel
            # (no solo el más accesible), con el % acumulado de ese nivel — ej. primer
            # escalón 5%, segundo escalón 5%+5%, etc.
            for i, esc in enumerate(escalones, start=1):
                pcts_txt = '+'.join(f'{p:g}%' for p in (esc.get('porcentajes') or [])) or '0%'
                esc_cond = f", condición: {esc['condicion_desc']}" if esc.get('condicion_desc') else ''
                story.append(Paragraph(
                    f'&nbsp;&nbsp;&nbsp;&nbsp;– Escalón {i}: {pcts_txt} de descuento a partir de '
                    f'${AR(esc.get("monto_minimo"))} de neto facturado{esc_cond}', sSub))

    if descuento_pct and float(descuento_pct) > 0:
        story.append(Spacer(1, 10))
        minimo_txt = (f' y alcanza el mínimo de facturación de ${AR(descuento_monto_minimo)}'
                      if descuento_monto_minimo and float(descuento_monto_minimo) > 0 else '')
        texto_bonif = (
            f'Si alcanza los objetivos previstos en las Listas informadas{minimo_txt}, podés ofrecerle a estos '
            f'clientes una bonificación adicional: hasta el {_ddmmyyyy(oferta_vencimiento) or "—"}, el cliente '
            f'obtiene un {float(descuento_pct):g}% de descuento sobre el total neto facturado.'
        )
        box_style = ParagraphStyle('box', fontSize=9.5, fontName='Helvetica-Bold',
                                    textColor=colors.HexColor('#065f46'), leading=13)
        box = Table([[Paragraph(texto_bonif, box_style)]], colWidths=[180*mm])
        box.setStyle(TableStyle([
            ('BOX',        (0,0), (-1,-1), 1.2, colors.HexColor('#059669')),
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#ecfdf5')),
            ('TOPPADDING',    (0,0), (-1,-1), 8),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
            ('LEFTPADDING',   (0,0), (-1,-1), 10),
            ('RIGHTPADDING',  (0,0), (-1,-1), 10),
        ]))
        story.append(box)

    doc.build(story)
    buf.seek(0)
    return buf.read()


def _reactivacion_enviar_email(destinatario: str, asunto: str, cuerpo: str, pdf_bytes: bytes, pdf_filename: str) -> bool:
    if not SMTP_HOST or not destinatario:
        return False
    msg = MIMEMultipart()
    msg["From"] = SMTP_FROM or SMTP_USER
    msg["To"] = destinatario
    msg["Subject"] = asunto
    msg.attach(MIMEText(cuerpo, "plain", "utf-8"))
    part = MIMEBase("application", "octet-stream")
    part.set_payload(pdf_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", "attachment", filename=pdf_filename)
    msg.attach(part)
    raw = msg.as_bytes()
    remitente = msg["From"]
    try:
        if SMTP_PORT == 465:
            with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=20) as srv:
                srv.login(SMTP_USER, SMTP_PASS)
                srv.sendmail(remitente, [destinatario], raw)
        else:
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20) as srv:
                srv.ehlo(); srv.starttls()
                srv.login(SMTP_USER, SMTP_PASS)
                srv.sendmail(remitente, [destinatario], raw)
        return True
    except Exception as e:
        print(f"[REACTIVACION] error enviando mail a {destinatario}: {e}")
        return False


def _reactivacion_disponibilidad_stock(clientes: list, vendedor_codigo: str) -> dict:
    """Para cada cliente detectado, arma la lista de artículos CON STOCK disponible
    en los depósitos alcanzados por este vendedor (remanente real, no reservado),
    dentro del/los mismo(s) RUBRO(S) y MARCA(S) que el cliente ya compraba (unión de
    ambos criterios) — no se limita a los SKU puntuales que facturó antes, para
    ofrecerle variedad real de todo lo que hay disponible, no siempre lo mismo.
    Incluye precio unitario (Lista 1) y bonificación % activa si el SKU está en
    alguna oferta vigente. Devuelve
    {codigocliente: [{codigo, descripcion, precio, stock, bonificacion_pct}, ...]}
    (máx. 8 artículos por cliente, priorizando bonificación activa y mayor stock)."""
    from datetime import date
    try:
        flags = get_flags_for_vendor(vendedor_codigo)
    except Exception:
        flags = {}
    dep_exclusivo = (flags.get('deposito_exclusivo') or '').strip()
    dep_lista = [d.strip() for d in dep_exclusivo.split(',') if d.strip()] if dep_exclusivo else \
        ['001', '002', '003', '005', '013', '016']  # sin restricción configurada → depósitos habituales
    if not dep_lista:
        return {cl['codigocliente']: [] for cl in clientes}

    try:
        catalog, cambio_usd = _get_catalog()
        fma_data = _fma_stock_parallel(dep_lista)
    except Exception:
        return {cl['codigocliente']: [] for cl in clientes}

    # Bonificación activa por codigoparticular (mismo criterio de vigencia que
    # _reactivacion_ofertas_sugeridas).
    hoy = date.today().isoformat()
    bonif_por_sku = {}
    try:
        cdb = _admin_db()
        for r in cdb.execute("""
            SELECT pd.codigo_producto, pd.bonificacion_pct FROM offer_product_details pd
            JOIN offers o ON o.id = pd.offer_id
            WHERE o.activo=1 AND (o.fecha_hasta='' OR o.fecha_hasta IS NULL OR o.fecha_hasta>=?)
              AND pd.bonificacion_pct > 0
        """, (hoy,)).fetchall():
            bonif_por_sku[r['codigo_producto']] = max(bonif_por_sku.get(r['codigo_producto'], 0), r['bonificacion_pct'])
        cdb.close()
    except Exception:
        pass

    # Una sola pasada del catálogo: indexado por rubro Y por marca (código), todos
    # los artículos con stock>0 en los depósitos de este vendedor — independiente de
    # si el cliente compró ESE SKU puntual antes.
    por_rubro = {}
    por_marca = {}
    for art_id, art in catalog.items():
        stock = sum(fma_data.get(dep, {}).get(art_id, 0) for dep in dep_lista)
        if stock <= 0:
            continue
        factor = cambio_usd if (art.get('codigomoneda') or '').upper() == 'DOLARES' else 1.0
        precio = _redondear_precio(art.get('precio1', 0), factor)
        cod_part = art.get('codigoparticular') or art_id
        item = {
            'codigo': cod_part,
            'descripcion': art.get('descripcion', ''),
            'precio': precio,
            'stock': stock,
            'bonificacion_pct': bonif_por_sku.get(cod_part, 0),
        }
        rubro_cod = art.get('codigo_rubro')
        if rubro_cod:
            por_rubro.setdefault(rubro_cod, []).append(item)
        marca_cod = art.get('codigomarca')
        if marca_cod:
            por_marca.setdefault(marca_cod, []).append(item)

    resultado = {}
    for cl in clientes:
        rubros_cliente = {d.get('codigo_rubro') for d in cl.get('detalle', []) if d.get('codigo_rubro')}
        marcas_cliente = {d.get('codigomarca') for d in cl.get('detalle', []) if d.get('codigomarca')}
        vistos = set()
        items = []
        for rc in rubros_cliente:
            for it in por_rubro.get(rc, []):
                if it['codigo'] in vistos:
                    continue
                vistos.add(it['codigo']); items.append(it)
        for mc in marcas_cliente:
            for it in por_marca.get(mc, []):
                if it['codigo'] in vistos:
                    continue
                vistos.add(it['codigo']); items.append(it)
        items.sort(key=lambda x: (-x['bonificacion_pct'], -x['stock']))
        resultado[cl['codigocliente']] = items[:8]
    return resultado


def _reactivacion_notificar(cfg: dict, candidatos: list) -> dict:
    """Agrupa los clientes detectados por vendedor, genera UN pdf consolidado por
    vendedor, lo manda por mail (adjunto) y avisa por WhatsApp (best-effort, requiere
    plantilla ya aprobada por Meta — ver /admin/wa/crear-plantilla-reactivacion).
    Devuelve {vendedor_codigo: {'mail_ok':bool,'wa_ok':bool}} para marcar estado."""
    analisis_nombre = cfg['nombre']
    por_vend = {}
    for cand in candidatos:
        por_vend.setdefault(cand['vendedor_codigo'], []).append(cand)

    c = _admin_db()
    contactos = {r['codigo']: dict(r) for r in c.execute(
        "SELECT * FROM vendedores_contacto WHERE activo=1").fetchall()}
    c.close()

    resultado = {}
    for vend_cod, clientes in por_vend.items():
        vend_nombre = clientes[0]['vendedor_nombre']
        contacto = contactos.get(vend_cod)
        rubros, superrubros, marcas, codigos_articulo = [], [], [], []
        for cl in clientes:
            for d in cl.get('detalle', []):
                if d.get('codigo_rubro'): rubros.append(d['codigo_rubro'])
                if d.get('codigo_superrubro'): superrubros.append(d['codigo_superrubro'])
                if d.get('codigomarca'): marcas.append(d['codigomarca'])
                if d.get('codigo_articulo'): codigos_articulo.append(d['codigo_articulo'])
        ofertas = _reactivacion_ofertas_sugeridas(rubros, superrubros, codigos_articulo, marcas)
        try:
            disponibilidad = _reactivacion_disponibilidad_stock(clientes, vend_cod)
        except Exception as e:
            print(f"[REACTIVACION] error calculando disponibilidad de stock: {e}")
            disponibilidad = {}
        pdf_bytes = _reactivacion_generar_pdf(
            analisis_nombre, vend_nombre, clientes, ofertas,
            periodo_a_desde=cfg.get('periodo_a_desde'), periodo_a_hasta=cfg.get('periodo_a_hasta'),
            descuento_pct=cfg.get('descuento_pct') or 0, oferta_vencimiento=cfg.get('oferta_vencimiento'),
            periodo_b_dias=cfg.get('periodo_b_dias'), disponibilidad=disponibilidad,
            descuento_monto_minimo=cfg.get('descuento_monto_minimo') or 0)

        mail_ok = False
        if contacto and contacto.get('mail'):
            cuerpo = (f"Hola {vend_nombre},\n\n"
                      f"Adjuntamos el detalle de {len(clientes)} cliente(s) con alta facturación "
                      f"histórica que no registran ventas en el período analizado.\n\n"
                      f"Saludos,\nMicrobell S.A.")
            mail_ok = _reactivacion_enviar_email(
                contacto['mail'], f"Reactivación de clientes — {analisis_nombre}",
                cuerpo, pdf_bytes, f"reactivacion_{vend_cod}.pdf")

        wa_ok = False
        wa_err = ''
        if contacto and contacto.get('celular'):
            wa_ok, wa_err = _send_whatsapp_reactivacion(contacto['celular'], vend_nombre, len(clientes))
        elif contacto:
            wa_err = 'sin celular cargado en Contactos de Vendedores'

        resultado[vend_cod] = {'mail_ok': mail_ok, 'wa_ok': wa_ok, 'wa_err': wa_err}
    return resultado


class ReactivacionAnalisisBody(BaseModel):
    nombre: str
    vendedor_codigo: Optional[str] = ''
    gruposuperrubro: Optional[str] = ''
    superrubro: Optional[str] = ''
    rubro: Optional[str] = ''
    periodo_a_desde: Optional[str] = ''
    periodo_a_hasta: Optional[str] = ''
    periodo_b_dias: int = 60
    descuento_pct: float = 0
    oferta_vencimiento: Optional[str] = ''
    descuento_monto_minimo: float = 0
    condicion_comercial_extra: Optional[str] = ''
    dia_semana: int = 0
    hora: str = '09:00'
    activo: int = 1


@app.get("/admin/reactivacion")
def reactivacion_listar(_u=Depends(get_admin_user)):
    c = _admin_db()
    rows = c.execute("SELECT * FROM reactivacion_analisis ORDER BY id DESC").fetchall()
    c.close()
    return [dict(r) for r in rows]


@app.post("/admin/reactivacion")
def reactivacion_crear(body: ReactivacionAnalisisBody, _u=Depends(get_admin_user)):
    c = _admin_db()
    cur = c.cursor()
    cur.execute("""
        INSERT INTO reactivacion_analisis
            (nombre, vendedor_codigo, gruposuperrubro, superrubro, rubro,
             periodo_a_desde, periodo_a_hasta, periodo_b_dias,
             descuento_pct, oferta_vencimiento, descuento_monto_minimo, condicion_comercial_extra, dia_semana, hora, activo)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (body.nombre, body.vendedor_codigo, body.gruposuperrubro, body.superrubro, body.rubro,
          body.periodo_a_desde, body.periodo_a_hasta, body.periodo_b_dias,
          body.descuento_pct, body.oferta_vencimiento, body.descuento_monto_minimo, body.condicion_comercial_extra, body.dia_semana, body.hora, body.activo))
    new_id = cur.lastrowid
    c.commit(); c.close()
    return {"id": new_id}


@app.put("/admin/reactivacion/{analisis_id}")
def reactivacion_editar(analisis_id: int, body: ReactivacionAnalisisBody, _u=Depends(get_admin_user)):
    c = _admin_db()
    cur = c.cursor()
    cur.execute("""
        UPDATE reactivacion_analisis SET
            nombre=?, vendedor_codigo=?, gruposuperrubro=?, superrubro=?, rubro=?,
            periodo_a_desde=?, periodo_a_hasta=?, periodo_b_dias=?,
            descuento_pct=?, oferta_vencimiento=?, descuento_monto_minimo=?, condicion_comercial_extra=?, dia_semana=?, hora=?, activo=?
        WHERE id=?
    """, (body.nombre, body.vendedor_codigo, body.gruposuperrubro, body.superrubro, body.rubro,
          body.periodo_a_desde, body.periodo_a_hasta, body.periodo_b_dias,
          body.descuento_pct, body.oferta_vencimiento, body.descuento_monto_minimo, body.condicion_comercial_extra, body.dia_semana, body.hora, body.activo,
          analisis_id))
    if cur.rowcount == 0:
        c.close()
        raise HTTPException(404, "Análisis no encontrado")
    c.commit(); c.close()
    return {"ok": True}


@app.delete("/admin/reactivacion/{analisis_id}")
def reactivacion_borrar(analisis_id: int, _u=Depends(get_admin_user)):
    c = _admin_db()
    cur = c.cursor()
    cur.execute("DELETE FROM reactivacion_analisis WHERE id=?", (analisis_id,))
    c.commit(); c.close()
    return {"ok": True}


def _reactivacion_matched_offer_ids(c, rubros_vistos, superrubros_vistos, codigos_articulo, marcas_vistas=None) -> list:
    """Devuelve los ids de oferta activa que matchean por categoría (offer_category_filters,
    incluye rubro/superrubro/marca) o por SKU puntual (offer_product_details) — misma
    lógica de _reactivacion_ofertas_sugeridas, extraída acá para reutilizarla también al
    vincular el cliente de recupero. Traduce codigos_articulo (interno) a codigoparticular
    antes de matchear por SKU. IMPORTANTE: rubros_vistos/superrubros_vistos/marcas_vistas
    tienen que ser CÓDIGOS (codigo_rubro/codigo_superrubro/codigomarca), no la descripción."""
    marcas_vistas = marcas_vistas or []
    matched = set()
    if rubros_vistos or superrubros_vistos or marcas_vistas:
        conds, params = [], []
        if rubros_vistos:
            conds.append(f"(f.nivel='rubro' AND f.valor IN ({','.join('?'*len(rubros_vistos))}))")
            params += list(rubros_vistos)
        if superrubros_vistos:
            conds.append(f"(f.nivel='superrubro' AND f.valor IN ({','.join('?'*len(superrubros_vistos))}))")
            params += list(superrubros_vistos)
        if marcas_vistas:
            conds.append(f"(f.nivel='marca' AND f.valor IN ({','.join('?'*len(marcas_vistas))}))")
            params += list(marcas_vistas)
        for r in c.execute(f"SELECT DISTINCT offer_id FROM offer_category_filters f WHERE {' OR '.join(conds)}", params).fetchall():
            matched.add(r['offer_id'])
    if codigos_articulo:
        try:
            catalog, _ = _get_catalog()
            traducidos = set()
            for cod in codigos_articulo:
                art = catalog.get(cod)
                traducidos.add((art.get('codigoparticular') or cod) if art else cod)
        except Exception:
            traducidos = set(codigos_articulo)
        if traducidos:
            ph = ','.join('?' * len(traducidos))
            for r in c.execute(f"SELECT DISTINCT offer_id FROM offer_product_details WHERE codigo_producto IN ({ph})", list(traducidos)).fetchall():
                matched.add(r['offer_id'])
    if not matched:
        return []
    ph_ids = ','.join('?' * len(matched))
    ids_activas = {r['id'] for r in c.execute(
        f"SELECT id FROM offers WHERE id IN ({ph_ids}) AND activo=1", list(matched)).fetchall()}
    return [i for i in matched if i in ids_activas]


def _reactivacion_vincular_cliente_recupero(cfg: dict, cand: dict) -> list:
    """'Recupero de cartera': en vez de crear una oferta aparte, guarda el % de
    descuento adicional + vencimiento DENTRO de la(s) oferta(s) de categoría que ya
    le aplican a este cliente (Jugueteria/Outdoors/Tecnologia/etc, según lo que
    compró) — como una fila extra en offer_clients de ESA oferta. El % solo se
    inyecta en /ofertas dentro del escalón ya alcanzado (ver get_ofertas_for_vendor) —
    nunca destraba el escalón por sí solo. Devuelve la lista de offer_ids vinculados
    (vacía si no hay descuento cargado o no matcheó ninguna oferta existente)."""
    descuento_pct = float(cfg.get('descuento_pct') or 0)
    condicion_extra = (cfg.get('condicion_comercial_extra') or '').strip()
    # La recompensa puede ser % de descuento, condición comercial distinta, o ambas a
    # la vez — no son excluyentes (así se puede reactivar con una condición de venta
    # distinta cuando el % queda en 0, o sumar las dos cosas si corresponde).
    tipo_cartera = 'ambos' if (descuento_pct > 0 and condicion_extra) else ('condicion' if condicion_extra else 'descuento')
    if descuento_pct <= 0 and not condicion_extra:
        return []
    rubros, superrubros, marcas, codigos_articulo = [], [], [], []
    for d in cand.get('detalle', []):
        if d.get('codigo_rubro'): rubros.append(d['codigo_rubro'])
        if d.get('codigo_superrubro'): superrubros.append(d['codigo_superrubro'])
        if d.get('codigomarca'): marcas.append(d['codigomarca'])
        if d.get('codigo_articulo'): codigos_articulo.append(d['codigo_articulo'])
    c = _admin_db()
    offer_ids = _reactivacion_matched_offer_ids(c, rubros, superrubros, codigos_articulo, marcas)
    if not offer_ids:
        c.close()
        return []
    monto_minimo_extra = float(cfg.get('descuento_monto_minimo') or 0)
    for oid in offer_ids:
        c.execute("""
            INSERT INTO offer_clients (offer_id, codigocliente, razonsocial, descuento_extra_pct, vencimiento_extra, monto_minimo_extra, tipo_cartera, condicion_comercial_extra)
            VALUES (?,?,?,?,?,?,?,?)
            ON CONFLICT(offer_id, codigocliente) DO UPDATE SET
                razonsocial=excluded.razonsocial,
                descuento_extra_pct=excluded.descuento_extra_pct,
                vencimiento_extra=excluded.vencimiento_extra,
                monto_minimo_extra=excluded.monto_minimo_extra,
                tipo_cartera=excluded.tipo_cartera,
                condicion_comercial_extra=excluded.condicion_comercial_extra
        """, (oid, cand['codigocliente'], cand['razonsocial'],
              descuento_pct,
              cfg.get('oferta_vencimiento') or '', monto_minimo_extra,
              tipo_cartera, condicion_extra))
    c.commit()
    c.close()
    return offer_ids


def _reactivacion_ejecutar_analisis(analisis_id: int) -> dict:
    """Lógica compartida por el endpoint manual y el scheduler automático."""
    c = _admin_db()
    row = c.execute("SELECT * FROM reactivacion_analisis WHERE id=?", (analisis_id,)).fetchone()
    if not row:
        c.close()
        raise HTTPException(404, "Análisis no encontrado")
    cfg = dict(row)

    # Oferta especial vencida: se cierra la campaña — se sigue registrando el
    # detectado para auditoría, pero no se vuelve a notificar a nadie.
    oferta_vencida = bool(cfg.get('oferta_vencimiento')) and datetime.now().date().isoformat() > cfg['oferta_vencimiento']
    if oferta_vencida:
        c.execute("""
            UPDATE reactivacion_resultados SET estado='cerrado'
            WHERE analisis_id=? AND estado IN ('pendiente','notificado')
        """, (analisis_id,))

    candidatos = _reactivacion_calcular(cfg)

    errores = []  # se persisten en reactivacion_analisis.ultimo_error para poder
                  # diagnosticar sin necesitar acceso a la consola del servidor.

    fecha_corrida = datetime.now().isoformat(timespec='seconds')
    cur = c.cursor()
    for cand in candidatos:
        cur.execute("""
            INSERT INTO reactivacion_resultados
                (analisis_id, fecha_corrida, codigocliente, codigoparticular, razonsocial,
                 vendedor_codigo, vendedor_nombre, importe_total, estado)
            VALUES (?,?,?,?,?,?,?,?, 'pendiente')
        """, (analisis_id, fecha_corrida, cand['codigocliente'], cand.get('codigoparticular') or cand['codigocliente'],
              cand['razonsocial'], cand['vendedor_codigo'], cand['vendedor_nombre'], cand['importe_total']))
        cand['resultado_id'] = cur.lastrowid
        for d in cand['detalle']:
            cur.execute("""
                INSERT INTO reactivacion_resultado_detalle
                    (resultado_id, tipo_comprobante, numero_comprobante, fecha,
                     codigo_articulo, descripcion_articulo, rubro, superrubro, cantidad, importe)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (cand['resultado_id'], d['tipo_comprobante'], d['numero_comprobante'], d['fecha'],
                  d['codigo_articulo'], d['descripcion_articulo'], d['rubro'], d['superrubro'],
                  d['cantidad'], d['importe']))
    cur.execute("UPDATE reactivacion_analisis SET ultima_corrida=? WHERE id=?", (fecha_corrida, analisis_id))
    c.commit(); c.close()

    # Vinculación del descuento adicional (recupero de cartera) al cliente puntual,
    # DENTRO de la(s) oferta(s) de categoría que ya le aplican (Jugueteria/Outdoors/
    # Tecnologia/etc, según lo que compró) — se hace en una pasada APARTE, ya con la
    # transacción de arriba commiteada y cerrada, para no pelear por el lock de
    # SQLite con esa misma conexión.
    if not oferta_vencida:
        c3 = _admin_db()
        for cand in candidatos:
            try:
                offer_ids = _reactivacion_vincular_cliente_recupero(cfg, cand)
                if offer_ids:
                    c3.execute("UPDATE reactivacion_resultados SET offer_id=? WHERE id=?", (offer_ids[0], cand['resultado_id']))
                    c3.commit()
                elif float(cfg.get('descuento_pct') or 0) > 0 or (cfg.get('condicion_comercial_extra') or '').strip():
                    errores.append(f"cliente {cand['codigocliente']}: descuento adicional cargado pero no matcheó "
                                    f"ninguna oferta existente (Jugueteria/Outdoors/Tecnologia/etc) para vincularlo")
            except Exception as e:
                msg = f"vinculando recupero cliente {cand['codigocliente']}: {e}"
                print(f"[REACTIVACION] error {msg}")
                errores.append(msg)
        c3.close()

    notificados = 0
    if candidatos and not oferta_vencida:
        try:
            resultado_envio = _reactivacion_notificar(cfg, candidatos)
            c2 = _admin_db()
            vend_sin_contacto = set()
            vend_wa_error = {}
            for cand in candidatos:
                info = resultado_envio.get(cand['vendedor_codigo'], {})
                if info.get('mail_ok') or info.get('wa_ok'):
                    notificados += 1
                    c2.execute(
                        "UPDATE reactivacion_resultados SET estado='notificado', fecha_ultima_notificacion=? WHERE id=?",
                        (fecha_corrida, cand['resultado_id']))
                else:
                    vend_sin_contacto.add(cand['vendedor_codigo'])
                if not info.get('wa_ok') and info.get('wa_err'):
                    vend_wa_error[cand['vendedor_codigo']] = info['wa_err']
            for vc in vend_sin_contacto:
                errores.append(f"vendedor {vc}: no se pudo notificar (sin mail/celular cargado en Contactos, o falló el envío)")
            for vc, werr in vend_wa_error.items():
                errores.append(f"vendedor {vc}: WhatsApp de reactivación falló — {werr}")
            c2.commit(); c2.close()
        except Exception as e:
            msg = f"notificando: {e}"
            print(f"[REACTIVACION] error {msg}")
            errores.append(msg)

    c4 = _admin_db()
    c4.execute("UPDATE reactivacion_analisis SET ultimo_error=? WHERE id=?",
               ('; '.join(errores[:20]) if errores else '', analisis_id))
    c4.commit(); c4.close()

    return {"fecha_corrida": fecha_corrida, "detectados": len(candidatos), "notificados": notificados,
            "oferta_vencida": oferta_vencida, "errores": errores}


@app.post("/admin/reactivacion/{analisis_id}/ejecutar")
def reactivacion_ejecutar(analisis_id: int, _u=Depends(get_admin_user)):
    """Corre el análisis AHORA (recalcula todo desde cero), persiste el resultado y
    notifica a cada vendedor (mail con PDF + WhatsApp best-effort). Uso: botón manual
    en el panel — el scheduler automático llama _reactivacion_ejecutar_analisis directo."""
    return _reactivacion_ejecutar_analisis(analisis_id)


# ── Scheduler automático: corre cada análisis activo en su día/hora configurados ──
_REACTIVACION_SCHEDULER_ULTIMA = {}  # analisis_id -> 'YYYY-MM-DD' de la última corrida disparada

def _reactivacion_scheduler_loop():
    import time as _time
    while True:
        try:
            ahora = datetime.now()
            hoy_str = ahora.date().isoformat()
            hora_str = ahora.strftime('%H:%M')
            dia_actual = ahora.weekday()  # Monday=0 ... Sunday=6, mismo criterio que dia_semana
            c = _admin_db()
            activos = c.execute(
                "SELECT id, dia_semana, hora FROM reactivacion_analisis WHERE activo=1"
            ).fetchall()
            c.close()
            for a in activos:
                if a['dia_semana'] != dia_actual or a['hora'] != hora_str:
                    continue
                if _REACTIVACION_SCHEDULER_ULTIMA.get(a['id']) == hoy_str:
                    continue  # ya se disparó hoy, no repetir dentro del mismo minuto/hora
                _REACTIVACION_SCHEDULER_ULTIMA[a['id']] = hoy_str
                try:
                    print(f"[REACTIVACION SCHEDULER] corriendo análisis {a['id']}")
                    _reactivacion_ejecutar_analisis(a['id'])
                except Exception as e:
                    print(f"[REACTIVACION SCHEDULER] error en análisis {a['id']}: {e}")
        except Exception as e:
            print(f"[REACTIVACION SCHEDULER] error de loop: {e}")
        _time.sleep(60)

threading.Thread(target=_reactivacion_scheduler_loop, daemon=True).start()


def _reactivacion_obtener_resultados(analisis_id: int) -> dict:
    """Última corrida de este análisis, agrupada Vendedor -> Cliente -> Importe,
    con el detalle expandible (facturas/artículos). Reutilizado por el endpoint JSON
    del modal de auditoría y por las exportaciones a PDF/Excel."""
    c = _admin_db()
    ultima = c.execute(
        "SELECT MAX(fecha_corrida) FROM reactivacion_resultados WHERE analisis_id=?",
        (analisis_id,)
    ).fetchone()
    fecha_corrida = ultima[0] if ultima else None
    if not fecha_corrida:
        c.close()
        return {"fecha_corrida": None, "vendedores": []}

    resultados = c.execute("""
        SELECT * FROM reactivacion_resultados
        WHERE analisis_id=? AND fecha_corrida=?
        ORDER BY vendedor_nombre, importe_total DESC
    """, (analisis_id, fecha_corrida)).fetchall()

    por_vendedor = {}
    for r in resultados:
        r = dict(r)
        r['detalle'] = [dict(d) for d in c.execute(
            "SELECT * FROM reactivacion_resultado_detalle WHERE resultado_id=? ORDER BY fecha DESC",
            (r['id'],)
        ).fetchall()]
        vend = por_vendedor.setdefault(r['vendedor_codigo'], {
            'vendedor_codigo': r['vendedor_codigo'],
            'vendedor_nombre': r['vendedor_nombre'],
            'clientes': [],
        })
        vend['clientes'].append(r)
    c.close()
    return {"fecha_corrida": fecha_corrida, "vendedores": list(por_vendedor.values())}


@app.get("/admin/reactivacion/{analisis_id}/resultados")
def reactivacion_resultados(analisis_id: int, _u=Depends(get_admin_user)):
    """Última corrida de este análisis, agrupada Vendedor -> Cliente -> Importe,
    con el detalle expandible (facturas/artículos) para el modal de auditoría.
    Nunca se envía a los corredores — es solo para consulta interna."""
    return _reactivacion_obtener_resultados(analisis_id)


@app.get("/admin/reactivacion/{analisis_id}/resultados/exportar-excel")
def reactivacion_resultados_exportar_excel(analisis_id: int, _u=Depends(get_admin_download_auth)):
    """Excel de auditoría interna: hoja Resumen (vendedor/cliente/importe) + hoja
    Detalle (facturas/artículos). Usa CODIGOPARTICULAR como código de cliente."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    data = _reactivacion_obtener_resultados(analisis_id)
    if not data['vendedores']:
        raise HTTPException(404, "Este análisis todavía no se corrió, o no detectó clientes")

    _c_an = _admin_db()
    an = _c_an.execute("SELECT nombre FROM reactivacion_analisis WHERE id=?", (analisis_id,)).fetchone()
    _c_an.close()
    nombre_analisis = an['nombre'] if an else f"analisis_{analisis_id}"

    wb = openpyxl.Workbook()
    hdr_fill  = PatternFill("solid", fgColor="1A56DB")
    hdr_font  = Font(bold=True, color="FFFFFFFF", size=10)
    alt_fill  = PatternFill("solid", fgColor="EFF6FF")
    right_al  = Alignment(horizontal="right", vertical="center")
    left_al   = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws1 = wb.active
    ws1.title = "Resumen"
    headers1 = ["Vendedor", "Código cliente", "Razón social", "Facturación total"]
    for ci, h in enumerate(headers1, 1):
        cell = ws1.cell(1, ci, h); cell.font = hdr_font; cell.fill = hdr_fill
    ri = 2
    for v in data['vendedores']:
        for cl in v['clientes']:
            fill = alt_fill if ri % 2 == 0 else None
            vals = [v['vendedor_nombre'] or v['vendedor_codigo'],
                     cl.get('codigoparticular') or cl['codigocliente'],
                     cl['razonsocial'], round(cl['importe_total'], 2)]
            for ci, val in enumerate(vals, 1):
                cell = ws1.cell(ri, ci, val)
                if fill: cell.fill = fill
                cell.alignment = right_al if ci == 4 else left_al
            ri += 1
    for i, w in enumerate([28, 14, 40, 18], 1):
        ws1.column_dimensions[chr(64 + i)].width = w

    ws2 = wb.create_sheet("Detalle")
    headers2 = ["Vendedor", "Código cliente", "Razón social", "Fecha", "Comprobante",
                "Artículo", "Rubro", "Cantidad", "Importe"]
    for ci, h in enumerate(headers2, 1):
        cell = ws2.cell(1, ci, h); cell.font = hdr_font; cell.fill = hdr_fill
    ri = 2
    for v in data['vendedores']:
        for cl in v['clientes']:
            for d in cl['detalle']:
                fill = alt_fill if ri % 2 == 0 else None
                vals = [v['vendedor_nombre'] or v['vendedor_codigo'],
                         cl.get('codigoparticular') or cl['codigocliente'], cl['razonsocial'],
                         _fecha_ddmmyyyy(d['fecha']), f"{d['tipo_comprobante']} {d['numero_comprobante']}",
                         f"{d['codigo_articulo']} — {d['descripcion_articulo']}", d['rubro'] or '',
                         d['cantidad'], round(d['importe'], 2)]
                for ci, val in enumerate(vals, 1):
                    cell = ws2.cell(ri, ci, val)
                    if fill: cell.fill = fill
                    if ci in (8, 9): cell.alignment = right_al
                    elif ci in (6,): cell.alignment = left_al
                ri += 1
    for i, w in enumerate([28, 14, 30, 12, 16, 40, 16, 10, 14], 1):
        ws2.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    from fastapi.responses import StreamingResponse
    fname = f"reactivacion_{nombre_analisis}_{data['fecha_corrida'][:10]}.xlsx".replace(' ', '_')
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              headers={"Content-Disposition": f"attachment; filename={fname}"})


@app.get("/admin/reactivacion/{analisis_id}/resultados/exportar-pdf")
def reactivacion_resultados_exportar_pdf(analisis_id: int, _u=Depends(get_admin_download_auth)):
    """PDF de auditoría interna: mismo contenido que el modal (Vendedor -> Cliente
    -> detalle de facturas/artículos), con CODIGOPARTICULAR como código de cliente."""
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    data = _reactivacion_obtener_resultados(analisis_id)
    if not data['vendedores']:
        raise HTTPException(404, "Este análisis todavía no se corrió, o no detectó clientes")

    _c_an = _admin_db()
    an = _c_an.execute("SELECT nombre FROM reactivacion_analisis WHERE id=?", (analisis_id,)).fetchone()
    _c_an.close()
    nombre_analisis = an['nombre'] if an else f"análisis {analisis_id}"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=12*mm, rightMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
    azul = colors.HexColor('#1e429f')
    AR = lambda v: f"{float(v or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    sTit = ParagraphStyle('tit', fontSize=13, fontName='Helvetica-Bold', leading=16)
    sSub = ParagraphStyle('sub', fontSize=9,  fontName='Helvetica',      leading=12)
    sVend = ParagraphStyle('vend', fontSize=10, fontName='Helvetica-Bold', textColor=azul, leading=13, spaceBefore=10, spaceAfter=4)
    sHdr = ParagraphStyle('hdr', fontSize=7.5, fontName='Helvetica-Bold', textColor=colors.white, leading=9)
    sCell = ParagraphStyle('cell', fontSize=7.5, fontName='Helvetica', leading=9)

    story = [
        Paragraph(f'Reactivación de clientes — {nombre_analisis}', sTit),
        Paragraph(f'Corrida: {_fecha_ddmmyyyy(data["fecha_corrida"])} {data["fecha_corrida"][11:16]}  ·  Solo para consulta interna, nunca se envía a los corredores', sSub),
        Spacer(1, 8),
    ]
    for v in data['vendedores']:
        story.append(Paragraph(f'👤 {v["vendedor_nombre"] or v["vendedor_codigo"]}', sVend))
        resumen = [[Paragraph('Código', sHdr), Paragraph('Cliente', sHdr), Paragraph('Facturación', sHdr)]]
        for cl in v['clientes']:
            resumen.append([cl.get('codigoparticular') or cl['codigocliente'], cl['razonsocial'], f"${AR(cl['importe_total'])}"])
        t = Table(resumen, colWidths=[25*mm, 100*mm, 35*mm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), azul),
            ('FONTSIZE', (0,1), (-1,-1), 8),
            ('ALIGN', (2,0), (2,-1), 'RIGHT'),
            ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#d1d5db')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f3f4f6')]),
            ('TOPPADDING', (0,0), (-1,-1), 3), ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ]))
        story.append(t)
        for cl in v['clientes']:
            if not cl['detalle']:
                continue
            story.append(Spacer(1, 3))
            story.append(Paragraph(f'Detalle — {cl.get("codigoparticular") or cl["codigocliente"]} · {cl["razonsocial"]}', sSub))
            det = [[Paragraph(h, sHdr) for h in ('Fecha', 'Comprobante', 'Artículo', 'Rubro', 'Cant.', 'Importe')]]
            for d in cl['detalle']:
                det.append([
                    Paragraph(_fecha_ddmmyyyy(d['fecha']), sCell),
                    Paragraph(f"{d['tipo_comprobante']} {d['numero_comprobante']}", sCell),
                    Paragraph(f"{d['codigo_articulo']} — {d['descripcion_articulo']}", sCell),
                    Paragraph(d['rubro'] or '', sCell),
                    Paragraph(str(d['cantidad']), sCell),
                    Paragraph(f"${AR(d['importe'])}", sCell),
                ])
            td = Table(det, colWidths=[18*mm, 25*mm, 65*mm, 22*mm, 15*mm, 20*mm])
            td.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#6b7280')),
                ('ALIGN', (4,0), (5,-1), 'RIGHT'),
                ('GRID', (0,0), (-1,-1), 0.25, colors.HexColor('#e5e7eb')),
                ('TOPPADDING', (0,0), (-1,-1), 2), ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ]))
            story.append(td)
            story.append(Spacer(1, 6))

    doc.build(story)
    buf.seek(0)
    from fastapi.responses import StreamingResponse
    fname = f"reactivacion_{nombre_analisis}_{data['fecha_corrida'][:10]}.pdf".replace(' ', '_')
    return StreamingResponse(buf, media_type="application/pdf",
                              headers={"Content-Disposition": f"attachment; filename={fname}"})


@app.get("/ventas")
def ventas(
    vendedor:    Optional[str] = None,
    cliente:     Optional[str] = None,
    desde:       Optional[str] = None,
    hasta:       Optional[str] = None,
    grupo:       Optional[str] = None,
    superrubro:  Optional[str] = None,
    rubro:       Optional[str] = None,
    marca:       Optional[str] = None,
    articulo:    Optional[str] = None,
    _u=Depends(get_admin_user)
):
    """Análisis de ventas para el panel admin. vendedor y cliente son opcionales."""
    if not vendedor and not cliente and not desde and not hasta \
       and not grupo and not superrubro and not rubro and not marca and not articulo:
        raise HTTPException(status_code=400, detail="Especificá al menos un filtro (vendedor, cliente o período).")
    return _ventas_query(vendedor=vendedor, cliente=cliente, desde=desde, hasta=hasta,
                         grupo=grupo, superrubro=superrubro, rubro=rubro, marca=marca, articulo=articulo)


@app.get("/ventas/pdf")
def ventas_pdf(
    vendedor:   Optional[str] = None,
    cliente:    Optional[str] = None,
    desde:      Optional[str] = None,
    hasta:      Optional[str] = None,
    titulo:     Optional[str] = None,
    grupo:      Optional[str] = None,
    superrubro: Optional[str] = None,
    rubro:      Optional[str] = None,
    marca:      Optional[str] = None,
    articulo:   Optional[str] = None,
    _u=Depends(get_admin_user)
):
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image
    from reportlab.lib.styles import ParagraphStyle
    from io import BytesIO
    from datetime import datetime

    rows = _ventas_query(vendedor=vendedor, cliente=cliente, desde=desde, hasta=hasta,
                         grupo=grupo, superrubro=superrubro, rubro=rubro, marca=marca, articulo=articulo)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12*mm, rightMargin=12*mm,
                            topMargin=10*mm, bottomMargin=10*mm)

    azul = colors.HexColor('#1e429f')
    AR   = lambda v: f"{float(v or 0):,.2f}".replace(',','X').replace('.',',').replace('X','.')
    sDesc= ParagraphStyle('desc', fontSize=7, fontName='Helvetica', leading=8)
    sHdr = ParagraphStyle('hdr',  fontSize=7, fontName='Helvetica-Bold',
                          textColor=colors.white, leading=8, alignment=1)
    sSub = ParagraphStyle('sub',  fontSize=8, fontName='Helvetica', leading=10)
    sTit = ParagraphStyle('tit',  fontSize=12, fontName='Helvetica-Bold', leading=14)

    per = f"{desde or ''} a {hasta or ''}"
    emi = datetime.now().strftime('%d/%m/%Y %H:%M')
    sub_linea = titulo or (f"Vendedor: {vendedor}" if vendedor else '') + (' | ' if vendedor and cliente else '') + (f"Cliente: {cliente}" if cliente else '')
    logo_cell = ''
    if os.path.exists(LOGO_PATH):
        logo_cell = Image(LOGO_PATH, width=38*mm, height=12*mm, kind='proportional')
    hdr_table = Table(
        [[logo_cell,
          [Paragraph('Microbell S.A. — Análisis de Ventas', sTit),
           Paragraph(sub_linea, sSub),
           Paragraph(f'Período: {per}   |   Emisión: {emi}', sSub)]]],
        colWidths=[42*mm, 231*mm]
    )
    hdr_table.setStyle(TableStyle([
        ('VALIGN',  (0,0),(-1,-1), 'MIDDLE'),
        ('ALIGN',   (1,0),(1,0),   'LEFT'),
        ('LEFTPADDING',  (0,0),(-1,-1), 0),
        ('RIGHTPADDING', (0,0),(-1,-1), 4),
        ('TOPPADDING',   (0,0),(-1,-1), 0),
        ('BOTTOMPADDING',(0,0),(-1,-1), 4),
    ]))

    # CodArt Descripcion Cliente Fecha Tipo Nro Cant PUnit Importe IVA Total
    cw  = [14*mm, 70*mm, 50*mm, 16*mm, 10*mm, 22*mm, 11*mm, 18*mm, 18*mm, 8*mm, 18*mm]
    hdrs = [Paragraph(h, sHdr) for h in
            ['Cód.Art.','Descripción','Cliente','Fecha','Tipo','Nro. Comp.','Cant.','P.Unit.','Importe','IVA%','Total']]
    data = [hdrs]
    tot_imp = tot_tot = 0.0
    for r in rows:
        nro   = str(r['numero']).replace('.0','').zfill(10)
        fecha = r['fecha'][8:10]+'/'+r['fecha'][5:7]+'/'+r['fecha'][:4] if r['fecha'] else ''
        tot_imp += float(r['importe'] or 0)
        tot_tot += float(r['total']   or 0)
        data.append([
            r['cod_articulo'],
            Paragraph(r['descripcion'], sDesc),
            Paragraph(r['razonsocial'], sDesc),
            fecha, r['tipo'], nro,
            str(int(float(r['cantidad'] or 0))),
            f"${AR(r['precio_unitario'])}",
            f"${AR(r['importe'])}", AR(r['iva_pct']), f"${AR(r['total'])}",
        ])
    data.append(['','','','','','','',
                 Paragraph('TOTALES', ParagraphStyle('tb', fontSize=7, fontName='Helvetica-Bold')),
                 f"${AR(tot_imp)}", '', f"${AR(tot_tot)}"])

    t = Table(data, colWidths=cw, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),  (-1,0),  azul),
        ('ROWBACKGROUNDS',(0,1),  (-1,-2), [colors.white, colors.HexColor('#f3f4f6')]),
        ('BACKGROUND',    (0,-1), (-1,-1), colors.HexColor('#e0e7ff')),
        ('FONTNAME',      (0,1),  (-1,-1), 'Helvetica'),
        ('FONTNAME',      (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0),  (-1,-1), 7),
        ('ALIGN',         (0,0),  (-1,-1), 'LEFT'),
        ('ALIGN',         (0,0),  (-1,0),  'CENTER'),
        ('ALIGN',         (3,0),  (3,-1),  'CENTER'),
        ('ALIGN',         (4,0),  (4,-1),  'CENTER'),
        ('ALIGN',         (5,0),  (5,-1),  'CENTER'),
        ('ALIGN',         (6,0),  (6,-1),  'CENTER'),
        ('ALIGN',         (7,0),  (-1,-1), 'RIGHT'),
        ('VALIGN',        (0,0),  (-1,-1), 'MIDDLE'),
        ('GRID',          (0,0),  (-1,-1), 0.3, colors.HexColor('#d1d5db')),
        ('TOPPADDING',    (0,0),  (-1,-1), 2),
        ('BOTTOMPADDING', (0,0),  (-1,-1), 2),
        ('LEFTPADDING',   (0,0),  (-1,-1), 3),
        ('RIGHTPADDING',  (0,0),  (-1,-1), 3),
    ]))

    doc.build([hdr_table, t])
    buf.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(buf, media_type='application/pdf',
        headers={'Content-Disposition': 'inline; filename="ventas.pdf"'})


@app.get("/ventas/excel")
def ventas_excel(
    vendedor:   Optional[str] = None,
    cliente:    Optional[str] = None,
    desde:      Optional[str] = None,
    hasta:      Optional[str] = None,
    grupo:      Optional[str] = None,
    superrubro: Optional[str] = None,
    rubro:      Optional[str] = None,
    marca:      Optional[str] = None,
    articulo:   Optional[str] = None,
    _u=Depends(get_admin_user)
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    rows = _ventas_query(vendedor=vendedor, cliente=cliente, desde=desde, hasta=hasta,
                         grupo=grupo, superrubro=superrubro, rubro=rubro, marca=marca, articulo=articulo)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ventas'

    azul  = PatternFill('solid', fgColor='1e429f')
    gris  = PatternFill('solid', fgColor='f3f4f6')
    azulc = PatternFill('solid', fgColor='e0e7ff')
    bF    = Font(bold=True, color='FFFFFF', size=9)
    bN    = Font(bold=True, size=9)
    nN    = Font(size=9)
    cen   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    der   = Alignment(horizontal='right',  vertical='center')
    izq   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin  = Side(style='thin', color='d1d5db')
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

    hdrs   = ['Cód.Art.','Descripción','Cliente','Vendedor','Fecha','Tipo','Nro. Comp.','Cant.','P.Unit.','Importe','IVA%','Total']
    widths = [12,         45,           35,        15,       12,     8,     14,           10,     16,       16,      8,     16]
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = azul; cell.font = bF; cell.alignment = cen; cell.border = brd
        ws.column_dimensions[get_column_letter(ci)].width = w

    tot_imp = tot_tot = 0.0
    for ri, r in enumerate(rows, 2):
        nro   = str(r['numero']).replace('.0','').zfill(10)
        fg    = None if ri % 2 == 0 else gris
        fecha = r['fecha'][8:10]+'/'+r['fecha'][5:7]+'/'+r['fecha'][:4] if r['fecha'] else ''
        vals  = [r['cod_articulo'], r['descripcion'], r['razonsocial'], r['codigovendedor'],
                 fecha, r['tipo'], nro,
                 float(r['cantidad'] or 0), float(r['precio_unitario'] or 0),
                 float(r['importe'] or 0), float(r['iva_pct'] or 0), float(r['total'] or 0)]
        tot_imp += float(r['importe'] or 0); tot_tot += float(r['total'] or 0)
        aligns = [cen, izq, izq, cen, cen, cen, cen, cen, der, der, der, der]
        for ci, (v, al) in enumerate(zip(vals, aligns), 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = nN; cell.alignment = al; cell.border = brd
            if fg: cell.fill = fg
            if ci == 8:             cell.number_format = '#,##0'
            elif ci in (9, 10, 12): cell.number_format = '"$"#,##0.00'
            elif ci == 11:          cell.number_format = '#,##0.00'

    tr = len(rows) + 2
    for ci, v in enumerate(['']*8 + ['TOTALES', tot_imp, '', tot_tot], 1):
        cell = ws.cell(row=tr, column=ci, value=v)
        cell.fill = azulc; cell.font = bN; cell.border = brd
        cell.alignment = der if ci >= 8 else cen
        if ci in (10, 12): cell.number_format = '"$"#,##0.00'

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="ventas.xlsx"'})


@app.get("/resumen-deudas")
def resumen_deudas(vendedor: Optional[str] = None, _user=Depends(get_current_user)):
    """Suma de deuda pendiente por cliente, ordenado por deuda desc."""
    DB_PROD     = DATABASE      # DB-Prueba.gdb
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'  # SW producción

    c_cli = conn('WIN1252', DB_PROD)
    cur_cli = c_cli.cursor()
    if vendedor:
        cur_cli.execute(
            'SELECT CODIGOCLIENTE, RAZONSOCIAL, CODIGOPARTICULAR FROM "CLIENTES" '
            'WHERE ACTIVO=? AND UPPER(CODIGOVENDEDOR)=? ORDER BY RAZONSOCIAL',
            ('1', vendedor.upper())
        )
    else:
        cur_cli.execute(
            'SELECT CODIGOCLIENTE, RAZONSOCIAL, CODIGOPARTICULAR FROM "CLIENTES" '
            "WHERE ACTIVO='1' ORDER BY RAZONSOCIAL"
        )
    clientes_rows = cur_cli.fetchall()
    c_cli.close()

    def _calcular_deuda_cliente(cod, razon, part, vend=None):
        codigos = list({cod, part} - {''}) or [cod]
        total_deuda = 0.0
        tiene_deuda_positiva = False
        seen_cta = set()
        for db_path in [DB_PROD, DB_MLT_PROD]:
            try:
                rows = _query_cta(db_path, codigos, 500, 0, vendedor=vend)
                for r in rows:
                    key = (r[0], r[1])
                    if key not in seen_cta:
                        seen_cta.add(key)
                        valor = float(r[5] or 0)
                        total_deuda += valor
                        if valor > 0:
                            tiene_deuda_positiva = True
            except Exception:
                pass
        return total_deuda, tiene_deuda_positiva

    deudas = {}
    for cod, razon, part in clientes_rows:
        cod = (cod or '').strip(); razon = (razon or '').strip(); part = (part or '').strip()
        if not cod: continue
        total_deuda, tiene_deuda_positiva = _calcular_deuda_cliente(cod, razon, part, vend=vendedor)
        # Mostrar positivos siempre; negativos solo si también tienen comprobantes con deuda
        if total_deuda >= 0.01 or (total_deuda <= -0.01 and tiene_deuda_positiva):
            deudas[cod] = {'codigo': part or cod, 'razonsocial': razon, 'deuda': round(total_deuda, 2)}

    # Agregar clientes con NCA/NCB/etc. emitidas por este vendedor
    # aunque no estén asignados como clientes del vendedor en CLIENTES
    if vendedor:
        _NC_TIPOS = "('NCA','NCB','NCCA','NCCB','NCE','NCCE','SIV')"
        for db_path in [DB_PROD, DB_MLT_PROD]:
            try:
                c_nc = conn('WIN1252', db=db_path)
                cur_nc = c_nc.cursor()
                cur_nc.execute(f"""
                    SELECT DISTINCT CODIGOCLIENTE FROM "CABEZACOMPROBANTES"
                    WHERE TIPOCOMPROBANTE IN {_NC_TIPOS}
                      AND UPPER(CODIGOUSUARIO) = ?
                      AND ANULADA = '0'
                """, (vendedor.upper(),))
                nc_codigos = [r[0].strip() for r in cur_nc.fetchall() if r[0] and r[0].strip()]
                c_nc.close()
                for nc_cod in nc_codigos:
                    if nc_cod in deudas:
                        continue  # ya incluido
                    # Buscar datos del cliente
                    c_li = conn('WIN1252', db=DB_PROD)
                    cur_li = c_li.cursor()
                    cur_li.execute(
                        'SELECT CODIGOCLIENTE, RAZONSOCIAL, CODIGOPARTICULAR FROM "CLIENTES" '
                        'WHERE CODIGOCLIENTE=? OR CODIGOPARTICULAR=?',
                        (nc_cod, nc_cod)
                    )
                    row = cur_li.fetchone()
                    c_li.close()
                    if not row:
                        continue
                    cod2  = (row[0] or '').strip()
                    razon2 = (row[1] or '').strip()
                    part2 = (row[2] or '').strip()
                    if not cod2 or cod2 in deudas:
                        continue
                    total_deuda, tiene_deuda_positiva = _calcular_deuda_cliente(cod2, razon2, part2, vend=vendedor)
                    if total_deuda >= 0.01 or (total_deuda <= -0.01 and tiene_deuda_positiva):
                        deudas[cod2] = {'codigo': part2 or cod2, 'razonsocial': razon2, 'deuda': round(total_deuda, 2)}
            except Exception:
                pass

    result = sorted(deudas.values(), key=lambda x: x['razonsocial'])
    return result

@app.get("/resumen-deudas/pdf")
def resumen_deudas_pdf(vendedor: Optional[str] = None):
    """PDF con resumen de deudas por cliente — logo + detalle de comprobantes."""
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, KeepTogether
    from reportlab.platypus.flowables import Flowable
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
    from reportlab.platypus import Image
    from datetime import date

    DB_PROD     = DATABASE      # DB-Prueba.gdb
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'  # SW producción

    CELESTE = colors.HexColor('#4A90D9')
    GRIS    = colors.HexColor('#f3f4f6')
    ROJO    = colors.HexColor('#dc2626')
    AZUL_CLI= colors.HexColor('#1e3a5f')

    sN  = ParagraphStyle('rdN',  fontSize=7, leading=9,  fontName='Helvetica')
    sNc = ParagraphStyle('rdNc', fontSize=7, leading=9,  fontName='Helvetica',      alignment=TA_CENTER)
    sNr = ParagraphStyle('rdNr', fontSize=7, leading=9,  fontName='Helvetica',      alignment=TA_RIGHT)
    sNb = ParagraphStyle('rdNb', fontSize=7, leading=9,  fontName='Helvetica-Bold', alignment=TA_CENTER)
    sNbr= ParagraphStyle('rdNbr',fontSize=7, leading=9,  fontName='Helvetica-Bold', alignment=TA_CENTER)

    def _fmt_num(v):
        """Convierte 4400014955.0 → '4400014955' (10 dígitos, sin .0)"""
        try:
            return str(int(float(v))).zfill(10)
        except Exception:
            return str(v or '')

    def _fmt_date(v):
        """Convierte datetime o 'yyyy-mm-dd' → 'dd/mm/yyyy'"""
        if not v:
            return '—'
        try:
            from datetime import date as _date, datetime as _dt
            if hasattr(v, 'strftime'):
                return v.strftime('%d/%m/%Y')
            s = str(v)[:10]
            d = _dt.strptime(s, '%Y-%m-%d')
            return d.strftime('%d/%m/%Y')
        except Exception:
            return str(v)[:10]
    sCli= ParagraphStyle('rdCli',fontSize=9, leading=11, fontName='Helvetica-Bold', textColor=AZUL_CLI)
    sT  = ParagraphStyle('rdT',  fontSize=12,leading=15, fontName='Helvetica-Bold')
    sSub= ParagraphStyle('rdSub',fontSize=8, leading=10, fontName='Helvetica', textColor=colors.HexColor('#6b7280'))

    hoy = date.today().strftime('%d/%m/%Y')

    # ── Obtener clientes con deuda y sus comprobantes ─────────────────────────
    # Leer CLIENTES y comprobantes L1 desde DB-Prueba; SW desde DB-MLT-Prueba
    DB_PROD     = DATABASE      # DB-Prueba.gdb
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'  # SW producción
    c_cli = conn('WIN1252', DB_PROD)
    cur_cli = c_cli.cursor()
    if vendedor:
        cur_cli.execute(
            'SELECT CODIGOCLIENTE, RAZONSOCIAL, CODIGOPARTICULAR, CUIT FROM "CLIENTES" '
            "WHERE ACTIVO='1' AND UPPER(CODIGOVENDEDOR)=? ORDER BY RAZONSOCIAL",
            (vendedor.upper(),)
        )
    else:
        cur_cli.execute(
            'SELECT CODIGOCLIENTE, RAZONSOCIAL, CODIGOPARTICULAR, CUIT FROM "CLIENTES" '
            "WHERE ACTIVO='1' ORDER BY RAZONSOCIAL"
        )
    clientes_rows = cur_cli.fetchall()
    c_cli.close()

    clientes_data = []
    gran_total_deuda = 0.0
    for cod, razon, part, cuit in clientes_rows:
        cod   = (cod   or '').strip()
        razon = (razon or '').strip()
        part  = (part  or '').strip()
        cuit  = (cuit  or '').strip()
        if not cod: continue
        comprobantes = []
        seen = set()

        for db_path in [DB_PROD, DB_MLT_PROD]:
            try:
                rows = _query_cta(db_path, [cod], 500, 0)
                for r in rows:
                    key = (r[0], r[1])
                    if key not in seen:
                        seen.add(key)
                        comprobantes.append(r)
            except Exception:
                pass

        comprobantes.sort(key=lambda r: (r[6] or r[2], r[2]))
        total_deuda = sum(float(r[5] or 0) for r in comprobantes)
        if total_deuda > 0:
            gran_total_deuda += total_deuda
            clientes_data.append({
                'codigo': part or cod, 'razonsocial': razon, 'cuit': cuit,
                'comprobantes': comprobantes, 'total_deuda': total_deuda
            })

    clientes_data.sort(key=lambda x: x['razonsocial'])

    # ── Armar PDF ─────────────────────────────────────────────────────────────
    buf = BytesIO()
    PAGE_W, PAGE_H = A4
    mg = 14 * mm
    usable_w = PAGE_W - 2 * mg

    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=mg, rightMargin=mg,
                            topMargin=12*mm, bottomMargin=12*mm)

    def _on_page(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 6.5)
        canvas.setFillColor(colors.HexColor('#9ca3af'))
        canvas.drawCentredString(PAGE_W/2, 8*mm,
            f'microbell S.A.  ·  CUIT 30-70839018-2  ·  Resumen de Deudas al {hoy}  ·  Pág. {doc.page}')
        canvas.restoreState()

    story = []

    # ── Header con logo ───────────────────────────────────────────────────────
    logo_cell = Image(LOGO_PATH, width=38*mm, height=13*mm) if os.path.exists(LOGO_PATH) \
                else Paragraph('<b>microbell S.A.</b>', sT)
    titulo_txt = f'Resumen de Deudas Pendientes'
    subtitulo  = f'Fecha: {hoy}'
    if vendedor:
        subtitulo += f'   |   Vendedor: {vendedor}'
    hdr_tbl = Table([[
        logo_cell,
        [Paragraph(titulo_txt, sT), Paragraph(subtitulo, sSub)]
    ]], colWidths=[45*mm, usable_w - 45*mm])
    hdr_tbl.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('ALIGN',(1,0),(1,0),'RIGHT'),
    ]))
    story.append(hdr_tbl)
    story.append(HRFlowable(width=usable_w, thickness=1.5, color=CELESTE, spaceAfter=4))

    # ── Resumen global ────────────────────────────────────────────────────────
    story.append(Paragraph(
        f'<b>Total general deuda: {_pesos(gran_total_deuda)}</b>  &nbsp;·&nbsp;  '
        f'{len(clientes_data)} clientes con saldo pendiente',
        ParagraphStyle('rdResumen', fontSize=9, fontName='Helvetica-Bold',
                       textColor=ROJO, leading=12)
    ))
    story.append(Spacer(1, 5*mm))

    # ── Detalle por cliente ───────────────────────────────────────────────────
    cw_tipo = 14*mm; cw_num = 28*mm; cw_fcomp = 20*mm; cw_fvto = 20*mm
    cw_total = 30*mm; cw_pago = 30*mm; cw_deuda = 30*mm
    cw_cli = usable_w  # encabezado cliente ocupa todo el ancho

    for cli in clientes_data:
        bloque = []

        # Encabezado cliente
        cli_hdr = Table([[
            Paragraph(f'{cli["razonsocial"]}', sCli),
            Paragraph(f'Cód: {cli["codigo"]}  |  CUIT: {cli["cuit"] or "—"}', sN),
            Paragraph(f'Deuda: <b>{_pesos(cli["total_deuda"])}</b>',
                      ParagraphStyle('rdDeu', fontSize=8, fontName='Helvetica-Bold',
                                     textColor=ROJO, alignment=TA_RIGHT)),
        ]], colWidths=[usable_w*0.45, usable_w*0.3, usable_w*0.25])
        cli_hdr.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,-1), colors.HexColor('#e8f0fe')),
            ('TOPPADDING',(0,0),(-1,-1),3), ('BOTTOMPADDING',(0,0),(-1,-1),3),
            ('LEFTPADDING',(0,0),(-1,-1),5), ('RIGHTPADDING',(0,0),(-1,-1),5),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ]))
        bloque.append(cli_hdr)

        # Tabla de comprobantes
        comp_rows = [[
            Paragraph('<b>Tipo</b>',    sNb),
            Paragraph('<b>Número</b>',  sNb),
            Paragraph('<b>F.Comp.</b>', sNb),
            Paragraph('<b>F.Vto.</b>',  sNb),
            Paragraph('<b>Total</b>',   sNbr),
            Paragraph('<b>Pagado</b>',  sNbr),
            Paragraph('<b>Saldo</b>',   sNbr),
        ]]
        for r in cli['comprobantes']:
            deuda = float(r[5] or 0)
            if deuda <= 0:
                continue
            comp_rows.append([
                Paragraph(str(r[0] or ''),          sNc),
                Paragraph(_fmt_num(r[1]),            sNc),
                Paragraph(_fmt_date(r[2]),           sNc),
                Paragraph(_fmt_date(r[6]),           sNc),
                Paragraph(_pesos(float(r[3] or 0)), sNr),
                Paragraph(_pesos(float(r[4] or 0)), sNr),
                Paragraph(_pesos(deuda),             sNr),
            ])

        comp_tbl = Table(comp_rows,
                         colWidths=[cw_tipo, cw_num, cw_fcomp, cw_fvto, cw_total, cw_pago, cw_deuda],
                         repeatRows=1)
        comp_tbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0), CELESTE), ('TEXTCOLOR',(0,0),(-1,0), colors.white),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, GRIS]),
            ('BOX',(0,0),(-1,-1),0.4,colors.grey),
            ('INNERGRID',(0,0),(-1,-1),0.2,colors.lightgrey),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),2), ('BOTTOMPADDING',(0,0),(-1,-1),2),
            ('LEFTPADDING',(0,0),(-1,-1),3),
        ]))
        bloque.append(comp_tbl)
        bloque.append(Spacer(1, 4*mm))

        story.append(KeepTogether(bloque))

    # ── Pie resumen ───────────────────────────────────────────────────────────
    story.append(HRFlowable(width=usable_w, thickness=1, color=CELESTE, spaceAfter=3))
    story.append(Paragraph(
        f'<b>TOTAL GENERAL DEUDA: {_pesos(gran_total_deuda)}</b>',
        ParagraphStyle('rdTot', fontSize=11, fontName='Helvetica-Bold',
                       textColor=ROJO, alignment=TA_RIGHT)
    ))

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    buf.seek(0)
    return StreamingResponse(buf, media_type='application/pdf',
        headers={'Content-Disposition': f'inline; filename="resumen_deudas_{hoy}.pdf"'})

@app.get("/clientes/{codigo}/cuenta_corriente/pdf")
def cuenta_corriente_pdf(codigo: str, limit: int = Query(500, le=2000), offset: int = 0, _user=Depends(get_current_user)):
    from datetime import datetime

    # ── 1. Datos cliente (lookup en DB_PROD igual que resumen-deudas)
    DB_PROD     = DATABASE      # DB-Prueba.gdb
    c_cli = conn('WIN1252', db=DB_PROD)
    cur_cli = c_cli.cursor()
    cur_cli.execute(
        'SELECT CODIGOCLIENTE, CODIGOPARTICULAR, RAZONSOCIAL, CUIT, TELEFONO, TELEFONOCELULAR, DIRECCION, LOCALIDAD '
        'FROM "CLIENTES" WHERE CODIGOCLIENTE = ? OR CODIGOPARTICULAR = ?',
        (codigo, codigo)
    )
    cli = cur_cli.fetchone()
    c_cli.close()
    if not cli:
        raise HTTPException(404, "Cliente no encontrado")
    cli_cod_visible = (cli[1] or cli[0] or '').strip()
    cli_razon = (cli[2] or '').strip()
    cli_cuit  = (cli[3] or '').strip()
    cli_tel   = (cli[4] or cli[5] or '').strip()
    cli_dir   = (cli[6] or '').strip()
    cli_loc   = (cli[7] or '').strip()

    # ── 2. Movimientos (misma logica que /cuenta_corriente)
    codigos = set()
    if cli[0] and cli[0].strip(): codigos.add(cli[0].strip())
    if cli[1] and cli[1].strip(): codigos.add(cli[1].strip())
    if not codigos: codigos.add(codigo)
    codigos = list(codigos)

    DB_PROD     = DATABASE      # DB-Prueba.gdb
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'  # SW producción
    rows_prod = _query_cta(DB_PROD,      codigos, limit, offset)
    rows_mp   = _query_cta(DB_MLT_PROD,  codigos, limit, offset)

    seen, combined = set(), []
    for r in rows_prod + rows_mp:
        key = (r[0], r[1])
        if key not in seen:
            seen.add(key)
            combined.append(r)
    combined.sort(key=lambda r: (r[6] or r[2], r[2]))

    # ── 3. Datos empresa
    razon_soc = cuit_emp = dir_emp = tel_emp = email_emp = ''
    try:
        cp = conn(); ccp = cp.cursor()
        ccp.execute('SELECT RAZONSOCIAL, CUIT, DIRECCION, TELEFONO, EMAIL FROM "PARAMETROS" WHERE CODIGOPARAMETRO = 1')
        rp = ccp.fetchone(); cp.close()
        if rp:
            razon_soc, cuit_emp, dir_emp, tel_emp, email_emp = [(v or '').strip() for v in rp]
    except Exception:
        pass

    # ── 4. PDF
    buf = BytesIO()
    PAGE_W, PAGE_H = A4
    mg = 14 * mm

    s_title = ParagraphStyle('t', fontSize=13, leading=16, fontName='Helvetica-Bold',
                              textColor=colors.HexColor('#1a56db'))
    s_sub   = ParagraphStyle('s', fontSize=8,  leading=10, fontName='Helvetica',
                              textColor=colors.HexColor('#6b7280'))
    s_label = ParagraphStyle('l', fontSize=8,  leading=10, fontName='Helvetica-Bold',
                              textColor=colors.HexColor('#374151'))
    s_val   = ParagraphStyle('v', fontSize=8,  leading=10, fontName='Helvetica')
    s_hdr   = ParagraphStyle('h', fontSize=7.5, leading=9, fontName='Helvetica-Bold',
                              alignment=TA_CENTER, textColor=colors.white)
    s_cell  = ParagraphStyle('c', fontSize=7.5, leading=9, fontName='Helvetica')
    s_cell_r= ParagraphStyle('cr', fontSize=7.5, leading=9, fontName='Helvetica', alignment=TA_RIGHT)
    s_total = ParagraphStyle('tt', fontSize=10, leading=12, fontName='Helvetica-Bold',
                              alignment=TA_RIGHT, textColor=colors.HexColor('#dc2626'))

    usable_w = PAGE_W - 2 * mg
    footer_txt = f'{razon_soc}  ·  CUIT {cuit_emp}  ·  {dir_emp}  ·  Tel {tel_emp}  ·  {email_emp}'

    def _on_page(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor('#e5e7eb'))
        canvas.setLineWidth(0.5)
        canvas.line(mg, 10*mm, PAGE_W - mg, 10*mm)
        canvas.setFont('Helvetica', 6.5)
        canvas.setFillColor(colors.HexColor('#6b7280'))
        canvas.drawCentredString(PAGE_W / 2, 7*mm, footer_txt)
        canvas.drawRightString(PAGE_W - mg, 7*mm, f"Pág. {doc.page}")
        canvas.restoreState()

    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=mg, rightMargin=mg,
                            topMargin=mg, bottomMargin=18*mm)
    story = []

    # Header logo + titulo
    logo_cell = Image(LOGO_PATH, width=38*mm, height=12*mm) if os.path.exists(LOGO_PATH) else Paragraph('', s_sub)
    titulo_cell = [
        Paragraph("Resumen de Cuenta Corriente", s_title),
        Paragraph(datetime.now().strftime('%d/%m/%Y %H:%M'), s_sub),
    ]
    t_hdr = Table([[logo_cell, titulo_cell]], colWidths=[42*mm, usable_w - 42*mm])
    t_hdr.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN',  (1,0), (1,0),  'RIGHT'),
    ]))
    story.append(t_hdr)
    story.append(Spacer(1, 4*mm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#1a56db')))
    story.append(Spacer(1, 3*mm))

    # Datos cliente
    cli_data = [
        [Paragraph('Cliente:',   s_label), Paragraph(cli_razon, s_val),
         Paragraph('Código:',    s_label), Paragraph(cli_cod_visible, s_val)],
        [Paragraph('CUIT:',      s_label), Paragraph(cli_cuit or '—', s_val),
         Paragraph('Teléfono:',  s_label), Paragraph(cli_tel or '—', s_val)],
        [Paragraph('Dirección:', s_label), Paragraph(cli_dir or '—', s_val),
         Paragraph('Localidad:', s_label), Paragraph(cli_loc or '—', s_val)],
    ]
    t_cli = Table(cli_data, colWidths=[22*mm, (usable_w/2 - 22*mm), 22*mm, (usable_w/2 - 22*mm)])
    t_cli.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOX',    (0,0), (-1,-1), 0.5, colors.HexColor('#e5e7eb')),
        ('INNERGRID', (0,0), (-1,-1), 0.3, colors.HexColor('#f3f4f6')),
        ('TOPPADDING',    (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING',   (0,0), (-1,-1), 6),
        ('RIGHTPADDING',  (0,0), (-1,-1), 6),
    ]))
    story.append(t_cli)
    story.append(Spacer(1, 4*mm))

    # Tabla movimientos
    HDR_BG = colors.HexColor('#1a56db')
    ALT_BG = colors.HexColor('#eff6ff')
    cw = [16*mm, 32*mm, 24*mm, 24*mm, 28*mm, 28*mm, 30*mm]

    data = [[
        Paragraph('Tipo',     s_hdr),
        Paragraph('Número',   s_hdr),
        Paragraph('F. Comp.', s_hdr),
        Paragraph('F. Vto.',  s_hdr),
        Paragraph('Total',    s_hdr),
        Paragraph('Pagado',   s_hdr),
        Paragraph('Deuda',    s_hdr),
    ]]

    def _fmt_money(n):
        return '$' + f"{float(n or 0):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    def _fmt_fecha(f):
        if not f: return '—'
        s = str(f)
        return s[:10] if len(s) >= 10 else s

    total_total = total_pagado = total_deuda = 0.0
    hoy = datetime.now().date()

    for r in combined:
        tot = float(r[3] or 0); pag = float(r[4] or 0); deu = float(r[5] or 0)
        total_total += tot; total_pagado += pag; total_deuda += deu
        vto_str = _fmt_fecha(r[6])
        # Marcar vencido en rojo
        try:
            vto_d = r[6].date() if hasattr(r[6], 'date') else None
            vencido = bool(vto_d and vto_d < hoy and deu > 0.01)
        except Exception:
            vencido = False
        s_vto = ParagraphStyle('vt', parent=s_cell, textColor=colors.HexColor('#dc2626')) if vencido else s_cell

        data.append([
            Paragraph(str(r[0] or ''),    s_cell),
            Paragraph(str(r[1] or ''),    s_cell),
            Paragraph(_fmt_fecha(r[2]),   s_cell),
            Paragraph(vto_str + (' ⚠' if vencido else ''), s_vto),
            Paragraph(_fmt_money(tot),    s_cell_r),
            Paragraph(_fmt_money(pag),    s_cell_r),
            Paragraph(_fmt_money(deu),
                      ParagraphStyle('dr', parent=s_cell_r,
                                     textColor=colors.HexColor('#dc2626'),
                                     fontName='Helvetica-Bold')),
        ])

    tbl = Table(data, colWidths=cw, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND',     (0,0), (-1,0), HDR_BG),
        ('TEXTCOLOR',      (0,0), (-1,0), colors.white),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, ALT_BG]),
        ('GRID',           (0,0), (-1,-1), 0.4, colors.HexColor('#e5e7eb')),
        ('VALIGN',         (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',     (0,0), (-1,-1), 3),
        ('BOTTOMPADDING',  (0,0), (-1,-1), 3),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 4*mm))

    # Totales
    tot_data = [[
        Paragraph(f'{len(combined)} comprobantes', s_sub),
        Paragraph(f'Total: {_fmt_money(total_total)}    Pagado: {_fmt_money(total_pagado)}', s_sub),
        Paragraph(f'Deuda: {_fmt_money(total_deuda)}', s_total),
    ]]
    t_tot = Table(tot_data, colWidths=[40*mm, usable_w - 40*mm - 60*mm, 60*mm])
    t_tot.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN',  (2,0), (2,0),  'RIGHT'),
    ]))
    story.append(t_tot)

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    buf.seek(0)
    fname = f"CtaCte_{cli_cod_visible}.pdf"
    return StreamingResponse(buf, media_type='application/pdf',
        headers={"Content-Disposition": f"inline; filename={fname}"})


# ─── Transportes / Sucursales ────────────────────────────────────────────────
_transportes_cache: dict = {"ts": 0, "val": None}
_TRANSPORTES_CACHE_TTL = int(os.getenv('TRANSPORTES_CACHE_TTL', 300))  # dato casi estático

@app.get("/transportes")
def get_transportes():
    _now = time.time()
    if _transportes_cache["val"] is not None and (_now - _transportes_cache["ts"]) < _TRANSPORTES_CACHE_TTL:
        return _transportes_cache["val"]
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    rows_map: dict = {}
    from concurrent.futures import ThreadPoolExecutor as _TPE_transp

    def _fetch_transp(db_path):
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            cur.execute('SELECT CODIGOTRANSPORTE, DESCRIPCION FROM "TRANSPORTES" ORDER BY DESCRIPCION')
            data = cur.fetchall()
            c.close()
            return data
        except Exception:
            return []

    with _TPE_transp(max_workers=2) as _ex_transp:
        for rows in _ex_transp.map(_fetch_transp, [DATABASE, DB_PROD]):
            for r in rows:
                cod = r[0]
                if cod is not None and str(cod).strip() not in ('', '0') and cod not in rows_map:
                    rows_map[cod] = (r[1] or '').strip()
    resultado = sorted(
        [{"codigo": cod, "descripcion": desc} for cod, desc in rows_map.items()],
        key=lambda x: x["descripcion"]
    )
    _transportes_cache["ts"] = _now
    _transportes_cache["val"] = resultado
    return resultado

@app.get("/clientes/{codigo}/sucursales")
def get_sucursales_cliente(codigo: str):
    """
    Devuelve domicilios de entrega (SUCURSALESXCLIENTES) desde DB-Microbell.gdb.
    Resolución de código:
      1. Busca CODIGOPARTICULAR en DB-Prueba (fuente del codigo recibido)
      2. Usa CODIGOPARTICULAR para encontrar CODIGOCLIENTE en DB-Microbell
      3. Consulta SUCURSALESXCLIENTES con ese CODIGOCLIENTE
    """
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    sucursales = []
    direccion_principal = ''
    transp_cli  = None
    transp_fijo = False
    reparto_propio_cli = False

    # ── Paso 1: resolver CODIGOPARTICULAR desde DB-Prueba ───────────────────
    codigoparticular = codigo   # fallback: usar el código tal cual
    try:
        c_pru = conn('WIN1252')   # DB-Prueba.gdb
        cur_pru = c_pru.cursor()
        cur_pru.execute(
            'SELECT CODIGOPARTICULAR, DIRECCION, LOCALIDAD FROM "CLIENTES" '
            'WHERE CODIGOCLIENTE = ? OR CODIGOPARTICULAR = ?',
            (codigo, codigo)
        )
        row_pru = cur_pru.fetchone()
        c_pru.close()
        if row_pru:
            cp = (row_pru[0] or '').strip()
            if cp:
                codigoparticular = cp
            # Dirección de fallback desde DB-Prueba
            partes_dir = [p for p in [(row_pru[1] or '').strip(),
                                      (row_pru[2] or '').strip()] if p]
            if partes_dir:
                # Separador simple (no em dash): este texto se puede reenviar tal
                # cual como domicilio_entrega al guardar pedido/presupuesto, y esos
                # INSERT usan conexión Firebird LATIN1 — el em dash (—, U+2014) no
                # es codificable en latin-1 y rompía el guardado.
                direccion_principal = ' - '.join(partes_dir)
    except Exception:
        pass

    # ── Paso 2 y 3: buscar CODIGOCLIENTE en Microbell y traer sucursales ────
    c_prod = None
    try:
        c_prod = conn('WIN1252', db=DB_PROD)
        cur_prod = c_prod.cursor()

        # Resolver CODIGOCLIENTE interno en DB-Microbell via CODIGOPARTICULAR
        cur_prod.execute(
            'SELECT CODIGOCLIENTE, DIRECCION, LOCALIDAD FROM "CLIENTES" '
            'WHERE CODIGOPARTICULAR = ? OR CODIGOCLIENTE = ?',
            (codigoparticular, codigoparticular)
        )
        row_mb = cur_prod.fetchone()
        if row_mb:
            cod_mb = str(row_mb[0] or '').strip()
            partes_mb = [p for p in [str(row_mb[1] or '').strip(),
                                     str(row_mb[2] or '').strip()] if p]
            if partes_mb:
                direccion_principal = ' - '.join(partes_mb)
        else:
            cod_mb = codigoparticular

        # Transporte a nivel cliente (TRANSPORTEFIJO vive en CLIENTES, no en SUCURSALESXCLIENTES)
        cur_prod.execute(
            'SELECT CODIGOTRANSPORTE, TRANSPORTEFIJO, REPARTOPROPIO '
            'FROM "CLIENTES" WHERE CODIGOCLIENTE = ?', (cod_mb,)
        )
        row_transp = cur_prod.fetchone()
        transp_cli   = str(row_transp[0]).strip() if row_transp and row_transp[0] else None
        transp_fijo  = str(row_transp[1] or '0').strip() == '1' if row_transp else False
        reparto_propio_cli = str(row_transp[2] or '0').strip() == '1' if row_transp else False

        # Domicilios de entrega — intentar con REPARTOPROPIO, fallback sin él
        _suc_rows = []
        _suc_has_reparto = False
        try:
            cur_prod.execute(
                'SELECT s.CODIGOSUCURSAL, s.NOMBRE, s.DIRECCION, s.CODIGOTRANSPORTE, '
                'p.NOMBRE, l.NOMBRE, s.TELEFONO, s.OBSERVACIONES, s.REPARTOPROPIO '
                'FROM "SUCURSALESXCLIENTES" s '
                'LEFT JOIN "PROVINCIAS" p ON p.CODIGOPROVINCIA = s.CODIGOPROVINCIA '
                'LEFT JOIN "LOCALIDADES" l '
                '       ON l.CODIGOPROVINCIA = s.CODIGOPROVINCIA '
                '      AND l.CODIGOLOCALIDAD = s.CODIGOLOCALIDAD '
                'WHERE s.CODIGOCLIENTE = ? '
                'ORDER BY s.CODIGOSUCURSAL',
                (cod_mb,)
            )
            _suc_rows = cur_prod.fetchall()
            _suc_has_reparto = True
        except Exception:
            cur_prod.execute(
                'SELECT s.CODIGOSUCURSAL, s.NOMBRE, s.DIRECCION, s.CODIGOTRANSPORTE, '
                'p.NOMBRE, l.NOMBRE, s.TELEFONO, s.OBSERVACIONES '
                'FROM "SUCURSALESXCLIENTES" s '
                'LEFT JOIN "PROVINCIAS" p ON p.CODIGOPROVINCIA = s.CODIGOPROVINCIA '
                'LEFT JOIN "LOCALIDADES" l '
                '       ON l.CODIGOPROVINCIA = s.CODIGOPROVINCIA '
                '      AND l.CODIGOLOCALIDAD = s.CODIGOLOCALIDAD '
                'WHERE s.CODIGOCLIENTE = ? '
                'ORDER BY s.CODIGOSUCURSAL',
                (cod_mb,)
            )
            _suc_rows = cur_prod.fetchall()
        for r in _suc_rows:
            transp_suc = str(r[3]).strip() if r[3] is not None and r[3] != '' else None
            reparto_suc = (str(r[8] or '0').strip() == '1') if _suc_has_reparto else reparto_propio_cli
            sucursales.append({
                "codigo":        str(r[0] or "").strip(),
                "nombre":        str(r[1] or "").strip(),
                "direccion":     str(r[2] or "").strip(),
                "transporte":    transp_suc or transp_cli,
                "provincia":     str(r[4] or "").strip(),
                "localidad":     str(r[5] or "").strip(),
                "telefono":      str(r[6] or "").strip(),
                "observaciones": str(r[7] or "").strip(),
                "transporteFijo": transp_fijo,
                "repartoPropio": reparto_suc,
            })
    except Exception as _e_suc:
        sucursales = []
        _suc_error = str(_e_suc)
    else:
        _suc_error = None
    finally:
        if c_prod:
            try: c_prod.close()
            except Exception: pass

    # Fallback: si Microbell no devolvió sucursales, intentar en DATABASE (DB-Prueba)
    if not sucursales:
        try:
            c_fb = conn('WIN1252', db=DATABASE)
            cur_fb = c_fb.cursor()
            # Transporte del cliente en DATABASE
            if not transp_cli:
                cur_fb.execute(
                    'SELECT CODIGOTRANSPORTE, TRANSPORTEFIJO, REPARTOPROPIO FROM "CLIENTES" WHERE CODIGOCLIENTE = ?',
                    (codigo,)
                )
                row_fb = cur_fb.fetchone()
                if row_fb:
                    transp_cli = str(row_fb[0] or '').strip() or None
                    transp_fijo = str(row_fb[1] or '0').strip() == '1'
                    reparto_propio_cli = str(row_fb[2] or '0').strip() == '1'
            cur_fb.execute(
                'SELECT s.CODIGOSUCURSAL, s.NOMBRE, s.DIRECCION, s.CODIGOTRANSPORTE, '
                'p.NOMBRE, l.NOMBRE, s.TELEFONO, s.OBSERVACIONES '
                'FROM "SUCURSALESXCLIENTES" s '
                'LEFT JOIN "PROVINCIAS" p ON p.CODIGOPROVINCIA = s.CODIGOPROVINCIA '
                'LEFT JOIN "LOCALIDADES" l ON l.CODIGOPROVINCIA = s.CODIGOPROVINCIA '
                '   AND l.CODIGOLOCALIDAD = s.CODIGOLOCALIDAD '
                'WHERE s.CODIGOCLIENTE = ? ORDER BY s.CODIGOSUCURSAL',
                (codigo,)
            )
            for r in cur_fb.fetchall():
                sucursales.append({
                    "codigo":        str(r[0] or "").strip(),
                    "nombre":        str(r[1] or "").strip(),
                    "direccion":     str(r[2] or "").strip(),
                    "transporte":    str(r[3] or "").strip() or transp_cli,
                    "provincia":     str(r[4] or "").strip(),
                    "localidad":     str(r[5] or "").strip(),
                    "telefono":      str(r[6] or "").strip(),
                    "observaciones": str(r[7] or "").strip(),
                    "transporteFijo": transp_fijo,
                    "repartoPropio": reparto_propio_cli,
                })
            c_fb.close()
            if not direccion_principal:
                cur_fb2 = conn('WIN1252', db=DATABASE).cursor()
                cur_fb2.execute('SELECT DIRECCION, LOCALIDAD FROM "CLIENTES" WHERE CODIGOCLIENTE = ?', (codigo,))
                row_d = cur_fb2.fetchone()
                if row_d:
                    partes = [p for p in [(row_d[0] or '').strip(), (row_d[1] or '').strip()] if p]
                    direccion_principal = ' - '.join(partes)
        except Exception:
            pass

    resp = {
        "sucursales": sucursales,
        "direccion_principal": direccion_principal,
        "transporte_codigo": transp_cli,
        "transporte_fijo": transp_fijo,
        "reparto_propio": reparto_propio_cli,
    }
    if _suc_error:
        resp["_error"] = _suc_error   # solo para diagnóstico; se puede quitar luego
    return resp

# ─── Informar Pago ────────────────────────────────────────────────────────────
@app.post("/clientes/{id}/informar-pago")
async def informar_pago(
    id: str,
    nombre: str = Form(...),
    vendedor: str = Form(...),
    comentario: str = Form(""),
    comprobante: Optional[UploadFile] = File(None),
):
    """Envía notificación de pago por email con adjunto opcional."""
    if not SMTP_HOST or not SMTP_TO_PAGOS:
        raise HTTPException(status_code=503, detail="Servicio de email no configurado")

    msg = MIMEMultipart()
    msg["From"]    = SMTP_FROM or SMTP_USER
    msg["To"]      = SMTP_TO_PAGOS
    msg["Subject"] = f"Informar Pago - Cliente {nombre} (Vendedor: {vendedor})"

    cuerpo = f"""Se ha recibido un aviso de pago:

Cliente:  {nombre} (código: {id})
Vendedor: {vendedor}
Comentario: {comentario or '(sin comentario)'}
"""
    msg.attach(MIMEText(cuerpo, "plain", "utf-8"))

    if comprobante and comprobante.filename:
        datos = await comprobante.read()
        part = MIMEBase("application", "octet-stream")
        part.set_payload(datos)
        encoders.encode_base64(part)
        safe_fn = comprobante.filename.replace("\r", "").replace("\n", "")
        part.add_header("Content-Disposition", "attachment", filename=safe_fn)
        msg.attach(part)

    raw = msg.as_bytes()
    remitente = msg["From"]

    def _send():
        # Puerto 465 → SMTP_SSL directo (sin starttls)
        # Puerto 587 → SMTP + starttls
        if SMTP_PORT == 465:
            with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=20) as srv:
                srv.login(SMTP_USER, SMTP_PASS)
                srv.sendmail(remitente, [SMTP_TO_PAGOS], raw)
        else:
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20) as srv:
                srv.ehlo()
                srv.starttls()
                srv.login(SMTP_USER, SMTP_PASS)
                srv.sendmail(remitente, [SMTP_TO_PAGOS], raw)

    try:
        await asyncio.to_thread(_send)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al enviar email: {e}")

    return {"ok": True, "mensaje": "Pago informado correctamente"}

# ─── Pedidos (solo del vendedor) ───────────────────────────────────────────────
@app.get("/pedidos")
def get_pedidos(
    vendedor: str,
    cliente: Optional[str] = None,
    limit: int = Query(50, le=200),
    offset: int = 0,
    db: str = Query("oficial"),
    _user=Depends(get_current_user)
):
    if db == 'sw':
        # SW → DATABASE_MLT → CABEZACOMPROBANTES
        c = conn('WIN1252', db=DATABASE_MLT)
        cur = c.cursor()
        params = [vendedor.upper()]
        where_cli = ""
        if cliente:
            where_cli = "AND CODIGOCLIENTE = ?"
            params.append(cliente)
        cur.execute(f"""
            SELECT FIRST {limit} SKIP {offset}
                NUMEROCOMPROBANTE, CODIGOCLIENTE, RAZONSOCIAL,
                FECHACOMPROBANTE, TOTAL, ANULADA, COMENTARIOS, FECHAVENCIMIENTO,
                NULL, NULL, CODIGOUSUARIO2
            FROM "CABEZACOMPROBANTES"
            WHERE TIPOCOMPROBANTE = 'NP' AND CODIGOUSUARIO = ? AND ANULADA = 0
                {where_cli}
            ORDER BY FECHACOMPROBANTE DESC, NUMEROCOMPROBANTE DESC
        """, params)
        rows = cur.fetchall()
        c.close()
    else:
        # L1 → DATABASE → CABEZAPEDIDOS
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()
        params = [vendedor.upper()]
        where_cli = ""
        if cliente:
            where_cli = "AND CODIGOCLIENTE = ?"
            params.append(cliente)
        cur.execute(f"""
            SELECT FIRST {limit} SKIP {offset}
                NUMEROCOMPROBANTE, CODIGOCLIENTE, RAZONSOCIAL,
                FECHACOMPROBANTE, TOTAL, ANULADA, COMENTARIOS, FECHAENTREGA,
                OPERACION, FECHATERMINADA, CODIGOUSUARIO2
            FROM "CABEZAPEDIDOS"
            WHERE TIPOCOMPROBANTE = 'NP' AND CODIGOUSUARIO = ? AND ANULADA = '0'
                {where_cli} AND (PRIORIDAD IS NULL OR PRIORIDAD = '1')
            ORDER BY FECHACOMPROBANTE DESC, NUMEROCOMPROBANTE DESC
        """, params)
        rows = cur.fetchall()
        c.close()
    return [{
        "numero": r[0], "cod_cliente": r[1], "razonsocial": r[2],
        "fecha":  r[3], "total": r[4], "anulada": r[5],
        "comentarios": r[6], "fecha_entrega": r[7],
        "operacion": str(r[8]).strip() if r[8] is not None else "1",
        "fecha_terminada": str(r[9]) if r[9] else None,
        "responsable": str(r[10]).strip() if r[10] else None
    } for r in rows]

@app.post("/pedidos/{numero}/terminar")
def terminar_pedido(numero: str, codigousuario: str = Query(...)):
    """Marca el pedido como terminado: FECHATERMINADA = ahora, CODIGOUSUARIO2 = responsable."""
    from datetime import datetime
    ahora = datetime.now()
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        cur.execute("""
            UPDATE "CABEZAPEDIDOS"
            SET FECHATERMINADA = ?, CODIGOUSUARIO2 = ?, FECHAMODIFICACION = ?
            WHERE TIPOCOMPROBANTE = 'NP' AND NUMEROCOMPROBANTE = ?
        """, (ahora, codigousuario.upper(), ahora, numero))
        c.commit()
        c.close()
        return {"ok": True, "numero": numero, "fechaterminada": str(ahora), "responsable": codigousuario.upper()}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/pedidos/{numero}/detalle")
def get_pedido_detalle(numero: str, db: str = Query("oficial")):
    if db == 'sw':
        # SW → DATABASE_MLT → CUERPOCOMPROBANTES (sin JOIN a ARTICULOS que no existe)
        c = conn('WIN1252', db=DATABASE_MLT)
        cur = c.cursor()
        cur.execute(
            "SELECT cc.LINEA, "
            "COALESCE(NULLIF(TRIM(cc.CODIGOPARTICULAR),''), TRIM(cc.CODIGOARTICULO)), "
            "cc.DESCRIPCION, cc.CANTIDAD, "
            "cc.DESCUENTO, cc.PRECIOUNITARIO, cc.PRECIOTOTAL, cc.PORCENTAJEIVA "
            "FROM \"CUERPOCOMPROBANTES\" cc "
            "WHERE cc.TIPOCOMPROBANTE = 'NP' AND cc.NUMEROCOMPROBANTE = ? ORDER BY cc.LINEA",
            (numero,)
        )
    else:
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()
        cur.execute(
            "SELECT cp.LINEA, "
            "COALESCE(NULLIF(TRIM(cp.CODIGOPARTICULAR),''), NULLIF(TRIM(a.CODIGOPARTICULAR),''), TRIM(cp.CODIGOARTICULO)), "
            "cp.DESCRIPCION, cp.CANTIDAD, "
            "cp.DESCUENTO, cp.PRECIOUNITARIO, cp.PRECIOTOTAL, cp.PORCENTAJEIVA "
            "FROM \"CUERPOPEDIDOS\" cp "
            "LEFT JOIN \"ARTICULOS\" a ON a.CODIGOARTICULO = cp.CODIGOARTICULO "
            "WHERE cp.TIPOCOMPROBANTE = 'NP' AND cp.NUMEROCOMPROBANTE = ? ORDER BY cp.LINEA",
            (numero,)
        )
    rows = cur.fetchall()
    c.close()
    return [{
        "linea": r[0], "codigo": r[1], "descripcion": r[2],
        "cantidad": r[3], "descuento": r[4],
        "precio_unitario": r[5], "precio_total": r[6], "iva": r[7]
    } for r in rows]

@app.get("/pedidos/{numero}/copia-datos")
def get_pedido_copia(numero: str, db: str = Query("oficial"), _u=Depends(get_current_user)):
    """Devuelve cabecera + items de un pedido para pre-cargar en nuevo NV."""
    db_path = DATABASE_MLT if db == 'sw' else DATABASE
    c = conn(db=db_path)
    cur = c.cursor()
    cur.execute(
        'SELECT CODIGOCLIENTE, RAZONSOCIAL, CODIGOMULTIPLAZO, CODIGOTRANSPORTE, '
        'DIRECCION, CODIGOUSUARIO '
        'FROM "CABEZAPEDIDOS" WHERE NUMEROCOMPROBANTE = ? AND TIPOCOMPROBANTE = \'NP\'',
        (numero,)
    )
    cab = cur.fetchone()
    if not cab:
        c.close()
        raise HTTPException(404, f"Pedido {numero} no encontrado")
    codigocliente, razonsocial, codigomultiplazo, codigotransporte, direccion, codigousuario = cab
    # Deposito del primer item
    cur.execute(
        'SELECT FIRST 1 CODIGODEPOSITO FROM "CUERPOPEDIDOS" WHERE NUMEROCOMPROBANTE = ? AND TIPOCOMPROBANTE = \'NP\'',
        (numero,)
    )
    dep_row = cur.fetchone()
    codigodeposito = str(dep_row[0] or '001').strip() if dep_row else '001'
    # Items (misma conexión, mismo DB)
    cur.execute(
        'SELECT COALESCE(NULLIF(TRIM(it.CODIGOPARTICULAR),\'\'), NULLIF(TRIM(a.CODIGOPARTICULAR),\'\'), TRIM(it.CODIGOARTICULO)), '
        'it.DESCRIPCION, it.CANTIDAD, it.PRECIOUNITARIO, it.DESCUENTO, it.PORCENTAJEIVA '
        'FROM "CUERPOPEDIDOS" it '
        'LEFT JOIN "ARTICULOS" a ON a.CODIGOARTICULO = it.CODIGOARTICULO '
        'WHERE it.NUMEROCOMPROBANTE = ? AND it.TIPOCOMPROBANTE = \'NP\' ORDER BY it.LINEA',
        (numero,)
    )
    items = [{"codigo": r[0], "descripcion": r[1], "cantidad": float(r[2] or 1),
              "precio_unitario": float(r[3] or 0), "descuento": float(r[4] or 0),
              "iva": float(r[5] or 21)} for r in cur.fetchall()]
    c.close()
    return {
        "codigocliente": str(codigocliente or '').strip(),
        "razonsocial": str(razonsocial or '').strip(),
        "codigomultiplazo": str(codigomultiplazo or '').strip(),
        "codigotransporte": str(codigotransporte or '0').strip(),
        "domicilio_entrega": str(direccion or '').strip(),
        "codigodeposito": codigodeposito,
        "codigousuario": str(codigousuario or '').strip(),
        "items": items
    }

@app.get("/presupuestos/{numero}/debug-raw")
def get_presupuesto_debug(numero: str):
    """DEBUG TEMPORAL: devuelve todos los campos de CABEZAPRESUPUESTOS para comparar."""
    try:
        c = conn(db=DATABASE)
        cur = c.cursor()
        cur.execute(
            'SELECT TIPOCOMPROBANTE, NUMEROCOMPROBANTE, CODIGOCLIENTE, RAZONSOCIAL, '
            'FECHACOMPROBANTE, FECHAVENCIMIENTO, TOTAL, ANULADA, CODIGOUSUARIO, '
            'CODIGOMULTIPLAZO, CODIGOTRANSPORTE, CODIGOOPERACION, CLASECOMPROBANTE, '
            'CODIGORESPONSABLE, CODIGOUSUARIO2, DESCUENTOPORCENTAJE, FECHAAPROBADO, '
            'CODIGOUSUARIOAPROBACION, TIPOIVA '
            'FROM "CABEZAPRESUPUESTOS" WHERE NUMEROCOMPROBANTE = ?', (numero,)
        )
        row = cur.fetchone()
        c.close()
        if not row:
            raise HTTPException(404, f"Presupuesto {numero} no encontrado")
        cols = ['tipocomprobante','numerocomprobante','codigocliente','razonsocial',
                'fechacomprobante','fechavencimiento','total','anulada','codigousuario',
                'codigomultiplazo','codigotransporte','codigooperacion','clasecomprobante',
                'codigoresponsable','codigousuario2','descuentoporcentaje','fechaaprobado',
                'codigousuarioaprobacion','tipoiva']
        return {c: str(v) if v is not None else None for c, v in zip(cols, row)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/presupuestos/{numero}/copia-datos")
def get_presupuesto_copia(numero: str, _u=Depends(get_current_user)):
    """Devuelve cabecera + items de un presupuesto para pre-cargar en nuevo NV."""
    try:
        c = conn(db=DATABASE)
        cur = c.cursor()
        cur.execute(
            'SELECT CODIGOCLIENTE, RAZONSOCIAL, CODIGOMULTIPLAZO, CODIGOTRANSPORTE, '
            'DIRECCION, CODIGOUSUARIO '
            'FROM "CABEZAPRESUPUESTOS" WHERE NUMEROCOMPROBANTE = ?',
            (numero,)
        )
        cab = cur.fetchone()
        if not cab:
            c.close()
            raise HTTPException(404, f"Presupuesto {numero} no encontrado")
        codigocliente, razonsocial, codigomultiplazo, codigotransporte, direccion, codigousuario = cab
        codigodeposito = '001'  # CABEZAPRESUPUESTOS y CUERPOPRESUPUESTOS no tienen CODIGODEPOSITO
        cur.execute(
            'SELECT COALESCE(NULLIF(TRIM(cp.CODIGOPARTICULAR),\'\'), NULLIF(TRIM(a.CODIGOPARTICULAR),\'\'), TRIM(cp.CODIGOARTICULO)), '
            'cp.DESCRIPCION, cp.CANTIDAD, cp.PRECIOUNITARIO, cp.BONIFICACION, cp.PORCENTAJEIVA '
            'FROM "CUERPOPRESUPUESTOS" cp '
            'LEFT JOIN "ARTICULOS" a ON a.CODIGOARTICULO = cp.CODIGOARTICULO '
            'WHERE cp.NUMEROCOMPROBANTE = ? ORDER BY cp.LINEA',
            (numero,)
        )
        items = [{"codigo": r[0], "descripcion": r[1], "cantidad": float(r[2] or 1),
                  "precio_unitario": float(r[3] or 0), "descuento": float(r[4] or 0),
                  "iva": float(r[5] or 21)} for r in cur.fetchall()]
        c.close()
        return {
            "codigocliente": str(codigocliente or '').strip(),
            "razonsocial": str(razonsocial or '').strip(),
            "codigomultiplazo": str(codigomultiplazo or '').strip(),
            "codigotransporte": str(codigotransporte or '0').strip(),
            "domicilio_entrega": str(direccion or '').strip(),
            "codigodeposito": str(codigodeposito or '001').strip() or '001',
            "codigousuario": str(codigousuario or '').strip(),
            "items": items
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error interno al obtener presupuesto {numero}: {type(e).__name__}: {e}")

def _domicilio_con_localidad(db_path_dom, cod_cliente, direccion_texto):
    """Agrega ' - Loc. X - Prov. Y' a una dirección de entrega de pedido/presupuesto,
    igual que lo hace Flexxus ERP en sus propios comprobantes. CABEZAPEDIDOS/
    CABEZAPRESUPUESTOS solo guardan la dirección como texto plano (sin localidad ni
    provincia), así que se busca la sucursal de SUCURSALESXCLIENTES cuya DIRECCION
    coincida con ese texto (mismo criterio que /clientes/{codigo}/sucursales), y si
    no hay match se cae a la dirección principal del cliente en CLIENTES."""
    direccion_texto = (direccion_texto or '').strip()
    if not direccion_texto or not cod_cliente:
        return direccion_texto
    def _norm(s):
        return ' '.join(str(s or '').split()).upper()
    localidad = provincia = ''
    try:
        c = conn('WIN1252', db=db_path_dom)
        cur = c.cursor()
        cur.execute(
            'SELECT s.DIRECCION, l.NOMBRE, p.NOMBRE '
            'FROM "SUCURSALESXCLIENTES" s '
            'LEFT JOIN "PROVINCIAS" p ON p.CODIGOPROVINCIA = s.CODIGOPROVINCIA '
            'LEFT JOIN "LOCALIDADES" l ON l.CODIGOPROVINCIA = s.CODIGOPROVINCIA '
            '   AND l.CODIGOLOCALIDAD = s.CODIGOLOCALIDAD '
            'WHERE s.CODIGOCLIENTE = ?', (cod_cliente,)
        )
        for r in cur.fetchall():
            if _norm(r[0]) == _norm(direccion_texto):
                localidad = (r[1] or '').strip()
                provincia = (r[2] or '').strip()
                break
        if not localidad and not provincia:
            cur.execute(
                'SELECT DIRECCION, LOCALIDAD, CODIGOPROVINCIA FROM "CLIENTES" WHERE CODIGOCLIENTE = ?',
                (cod_cliente,)
            )
            r_cli = cur.fetchone()
            if r_cli and _norm(r_cli[0]) == _norm(direccion_texto):
                localidad = (r_cli[1] or '').strip()
                cod_prov = (r_cli[2] or '').strip()
                if cod_prov:
                    cur.execute('SELECT NOMBRE FROM "PROVINCIAS" WHERE CODIGOPROVINCIA = ?', (cod_prov,))
                    r_p = cur.fetchone()
                    if r_p: provincia = (r_p[0] or '').strip()
        c.close()
    except Exception:
        pass
    extra = ''
    if localidad: extra += f' - Loc. {localidad}'
    if provincia: extra += f' - Prov. {provincia}'
    return direccion_texto + extra

# ─── PDF Nota de Pedido ────────────────────────────────────────────────────────
@app.get("/pedidos/{numero}/pdf")
def pedido_pdf(numero: str, db: str = Query("oficial"),
               descuento_promo_pct: float = Query(0.0),
               descuento_promo_nombre: str = Query("")):
    db_path = DATABASE_MLT if db == 'sw' else DATABASE

    # ── 1. Datos empresa ───────────────────────────────────────────────────────
    razon_soc = 'MICROBELL S.A.'
    dir_emp   = 'PATAGONES 2675 PISO 3 - CABA - C1437JEA'
    tel_emp   = '+54 11 3988-0024'
    email_emp = 'info@microbellsa.com.ar'
    web_emp   = 'www.microbellsa.com'
    cuit_emp  = '30-70839018-2'
    try:
        c_e = conn('WIN1252', db=db_path)
        cur_e = c_e.cursor()
        cur_e.execute(
            'SELECT RAZONSOCIAL, DIRECCION, TELEFONO, EMAIL, DIRECCIONWEB, CUIT '
            'FROM "SUCURSALES" WHERE CODIGOSUCURSAL = ?', ('PRINCIPAL',)
        )
        emp = cur_e.fetchone()
        c_e.close()
        if emp:
            razon_soc = (emp[0] or razon_soc).strip()
            dir_emp   = (emp[1] or dir_emp).strip()
            tel_emp   = (emp[2] or tel_emp).strip()
            email_emp = (emp[3] or email_emp).strip()
            web_emp   = (emp[4] or web_emp).strip()
            cuit_emp  = (emp[5] or cuit_emp).strip()
    except Exception:
        pass

    # ── 2. Cabeza pedido ── L1 y SW ambos en CABEZAPEDIDOS de DATABASE ──────────
    extra_where = " AND TIPOCOMPROBANTE = 'NP'"
    try:
        c_h = conn('WIN1252', db=DATABASE)
        cur_h = c_h.cursor()
        cur_h.execute(
            'SELECT CODIGOCLIENTE, RAZONSOCIAL, FECHACOMPROBANTE, FECHAENTREGA, '
            'TOTAL, IVA1, COMENTARIOS, CODIGOUSUARIO, CODIGOMULTIPLAZO, '
            'CODIGOTRANSPORTE, DIRECCION, TIPOIVA, TELEFONO '
            f'FROM "CABEZAPEDIDOS" WHERE NUMEROCOMPROBANTE = ?{extra_where}',
            (numero,)
        )
        cab = cur_h.fetchone()
        c_h.close()
    except Exception as ex:
        raise HTTPException(500, f"Cabeza: {ex}")
    if not cab:
        raise HTTPException(404, f"Pedido {numero} no encontrado")

    cod_cli, rs_cli, fec_comp, fec_entrega, total_cab, iva1_cab, \
    comentarios, cod_usu, cod_multi, cod_transp, dir_cli, tipo_iva, tel_cli = cab
    dir_cli = _domicilio_con_localidad(DATABASE, cod_cli, dir_cli)

    subtotal_cab = float(total_cab or 0)
    iva1_val     = float(iva1_cab or 0)
    total_final  = subtotal_cab + iva1_val

    # ── 3. Items ── L1 y SW ambos en CUERPOPEDIDOS de DATABASE ──────────────────
    try:
        c_it = conn('WIN1252', db=DATABASE)
        cur_it = c_it.cursor()
        cur_it.execute(
            'SELECT COALESCE(NULLIF(TRIM(it.CODIGOPARTICULAR),\'\'), NULLIF(TRIM(a.CODIGOPARTICULAR),\'\'), TRIM(it.CODIGOARTICULO)), '
            'it.DESCRIPCION, it.CANTIDAD, it.PRECIOUNITARIO, '
            'it.DESCUENTO, it.PRECIOTOTAL, it.PORCENTAJEIVA '
            'FROM "CUERPOPEDIDOS" it '
            'LEFT JOIN "ARTICULOS" a ON a.CODIGOARTICULO = it.CODIGOARTICULO '
            f'WHERE it.NUMEROCOMPROBANTE = ? AND it.TIPOCOMPROBANTE = \'NP\' ORDER BY it.LINEA',
            (numero,)
        )
        items = cur_it.fetchall()
        c_it.close()
    except Exception as ex:
        raise HTTPException(500, f"Items: {ex}")

    # ── 4. Datos cliente ───────────────────────────────────────────────────────
    # CLIENTES y USUARIOS siempre en DATABASE (tabla maestra)
    cuit_cli = ''
    tel_pdf  = str(tel_cli or '').strip().lstrip('-').strip().split()[0] if str(tel_cli or '').strip().lstrip('-').strip() else ''
    vendedor_nombre = cod_usu or ''
    try:
        c_cl = conn('WIN1252', db=DATABASE)
        cur_cl = c_cl.cursor()
        cur_cl.execute(
            'SELECT CUIT, TELEFONO, TELEFONOCELULAR, CODIGOPARTICULAR FROM "CLIENTES" WHERE CODIGOCLIENTE = ?',
            (cod_cli,)
        )
        r_cl = cur_cl.fetchone()
        if r_cl:
            cuit_cli = (r_cl[0] or '').strip()
            _cod_particular = (r_cl[3] or '').strip()
            if _cod_particular:
                cod_cli = _cod_particular
            if not tel_pdf:
                def _tp(v):
                    s = str(v or '').strip().lstrip('-').strip()
                    try: s = str(int(float(s))) if s and ('E' in s.upper() or '.' in s) else s
                    except: pass
                    return s
                tel_pdf = _tp(r_cl[1]) or _tp(r_cl[2])
        cur_cl.execute(
            'SELECT NOMBRE, APELLIDO FROM "USUARIOS" WHERE CODIGOUSUARIO = ?',
            (cod_usu,)
        )
        r_usu = cur_cl.fetchone()
        if r_usu:
            vendedor_nombre = f"{(r_usu[0] or '').strip()} {(r_usu[1] or '').strip()}".strip() or cod_usu
        c_cl.close()
    except Exception:
        pass

    # ── 5. Transporte descripción ──────────────────────────────────────────────
    transporte_desc = 'A CONVENIR'
    try:
        if cod_transp and str(cod_transp).strip() not in ('', '0'):
            # Buscar transporte en DATABASE principal (MLT no tiene la tabla completa)
            for _tr_db in ([db_path, DATABASE] if db == 'sw' else [db_path]):
                try:
                    c_tr = conn('WIN1252', db=_tr_db)
                    cur_tr = c_tr.cursor()
                    cur_tr.execute('SELECT DESCRIPCION FROM "TRANSPORTES" WHERE CODIGOTRANSPORTE = ?', (cod_transp,))
                    r_tr = cur_tr.fetchone()
                    c_tr.close()
                    if r_tr and (r_tr[0] or '').strip():
                        transporte_desc = r_tr[0].strip()
                        break
                except Exception:
                    pass
    except Exception:
        pass

    # ── 6. Condición de venta ──────────────────────────────────────────────────
    cond_venta = ''
    try:
        if cod_multi:
            c_mp = conn('WIN1252', db=DATABASE)
            cur_mp = c_mp.cursor()
            cur_mp.execute('SELECT DESCRIPCION FROM "MULTIPLAZOS" WHERE CODIGOMULTIPLAZO = ?', (cod_multi,))
            r_mp = cur_mp.fetchone()
            if r_mp: cond_venta = (r_mp[0] or '').strip()
            c_mp.close()
    except Exception:
        pass

    # ── 7. Depósito ────────────────────────────────────────────────────────────
    deposito_desc = ''
    try:
        c_dep = conn('WIN1252', db=DATABASE)
        cur_dep = c_dep.cursor()
        cur_dep.execute(
            f'SELECT FIRST 1 CODIGODEPOSITO FROM "CUERPOPEDIDOS" WHERE NUMEROCOMPROBANTE = ?{extra_where}',
            (numero,)
        )
        r_dep = cur_dep.fetchone()
        if r_dep:
            cod_dep = str(r_dep[0] or '').strip()
            cur_dep.execute('SELECT DESCRIPCION FROM "DEPOSITOS" WHERE CODIGODEPOSITO = ?', (cod_dep,))
            r_dn = cur_dep.fetchone()
            deposito_desc = (r_dn[0] or cod_dep).strip() if r_dn else cod_dep
        c_dep.close()
    except Exception:
        pass

    # ── 8. Construcción PDF ────────────────────────────────────────────────────
    buf  = BytesIO()
    PAGE_W, PAGE_H = A4
    mg   = 14 * mm
    BOTTOM_BLOCK = 48 * mm   # espacio reservado al pie para totales + footer

    s_norm  = ParagraphStyle('pn',  fontSize=8,  leading=11, fontName='Helvetica')
    s_sm    = ParagraphStyle('psm', fontSize=7,  leading=9,  fontName='Helvetica')
    s_bold  = ParagraphStyle('pb',  fontSize=8,  leading=11, fontName='Helvetica-Bold')
    s_h2    = ParagraphStyle('ph2', fontSize=10, leading=13, fontName='Helvetica-Bold')
    s_c     = ParagraphStyle('pc',  fontSize=7.5,leading=10, fontName='Helvetica',      alignment=TA_CENTER)
    s_c_b   = ParagraphStyle('pcb', fontSize=7.5,leading=10, fontName='Helvetica-Bold', alignment=TA_CENTER)
    s_r     = ParagraphStyle('pr',  fontSize=8,  leading=11, fontName='Helvetica',      alignment=TA_RIGHT)
    s_r_b   = ParagraphStyle('prb', fontSize=9,  leading=12, fontName='Helvetica-Bold', alignment=TA_RIGHT)
    s_label = ParagraphStyle('pl',  fontSize=8,  leading=11, fontName='Helvetica-Bold', textColor=colors.HexColor('#374151'))
    s_val   = ParagraphStyle('pv',  fontSize=8,  leading=11, fontName='Helvetica',      textColor=colors.HexColor('#111827'))

    usable_w = PAGE_W - 2 * mg
    num_fmt  = f"0001-{int(numero):08d}"
    footer_txt = f'{razon_soc}  ·  CUIT {cuit_emp}  ·  {dir_emp}  ·  Tel {tel_emp}  ·  {email_emp}'
    s_ft = ParagraphStyle('ft', fontSize=6.5, leading=9, fontName='Helvetica',
                           alignment=TA_CENTER, textColor=colors.HexColor('#6b7280'))

    # ── Canvas callbacks: totales + firma al pie ───────────────────────────────
    def _draw_bottom(canvas, doc):
        canvas.saveState()
        x0    = mg
        right = PAGE_W - mg
        tot_y = BOTTOM_BLOCK - 8*mm        # Y desde la base de la hoja

        # Línea separadora
        canvas.setStrokeColor(colors.HexColor('#e5e7eb'))
        canvas.setLineWidth(0.5)
        canvas.line(x0, tot_y + 28*mm, right, tot_y + 28*mm)

        # Observaciones (si existen)
        obs_y = tot_y + 30*mm
        obs = str(comentarios or '').strip()
        if obs:
            canvas.setFont('Helvetica-Bold', 8)
            canvas.setFillColor(colors.HexColor('#374151'))
            canvas.drawString(x0, obs_y, 'Observaciones:')
            canvas.setFont('Helvetica', 8)
            canvas.drawString(x0 + 28*mm, obs_y, obs[:120])

        # Bloque totales (derecha)
        tw = 50 * mm
        tx = right - tw
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#374151'))
        _dto_p = max(0.0, min(float(descuento_promo_pct or 0), 100.0))
        if _dto_p > 0:
            _sub_bruto = subtotal_cab / (1 - _dto_p / 100) if _dto_p < 100 else subtotal_cab
            _dto_monto = _sub_bruto - subtotal_cab
            _lbl_promo = descuento_promo_nombre.strip() or f'Desc. promo combo ({_dto_p:g}%)'
            canvas.drawRightString(tx - 2, tot_y + 26*mm, 'Subtotal sin descuento:')
            canvas.drawRightString(right,  tot_y + 26*mm, _fmt(_sub_bruto))
            canvas.setFillColor(colors.HexColor('#b45309'))
            canvas.drawRightString(tx - 2, tot_y + 20*mm, f'{_lbl_promo}:')
            canvas.drawRightString(right,  tot_y + 20*mm, f'- {_fmt(_dto_monto)}')
            canvas.setFillColor(colors.HexColor('#374151'))
            canvas.drawRightString(tx - 2, tot_y + 14*mm, 'Subtotal c/descuento:')
            canvas.drawRightString(right,  tot_y + 14*mm, _fmt(subtotal_cab))
            canvas.drawRightString(tx - 2, tot_y + 8*mm, 'IVA 21%:')
            canvas.drawRightString(right,  tot_y + 8*mm, _fmt(iva1_val))
            canvas.setLineWidth(0.8); canvas.setStrokeColor(colors.black)
            canvas.line(tx - 5*mm, tot_y + 6*mm, right, tot_y + 6*mm)
            canvas.setFont('Helvetica-Bold', 10); canvas.setFillColor(colors.black)
            canvas.drawRightString(tx - 2, tot_y + 1*mm, 'TOTAL:')
            canvas.drawRightString(right,  tot_y + 1*mm, _fmt(total_final))
        else:
            canvas.drawRightString(tx - 2, tot_y + 20*mm, 'Subtotal:')
            canvas.drawRightString(right,  tot_y + 20*mm, _fmt(subtotal_cab))
            canvas.drawRightString(tx - 2, tot_y + 12*mm, 'IVA 21%:')
            canvas.drawRightString(right,  tot_y + 12*mm, _fmt(iva1_val))
            # Línea sobre total
            canvas.setLineWidth(0.8)
            canvas.setStrokeColor(colors.black)
            canvas.line(tx - 5*mm, tot_y + 10*mm, right, tot_y + 10*mm)
            canvas.setFont('Helvetica-Bold', 10)
            canvas.setFillColor(colors.black)
            canvas.drawRightString(tx - 2, tot_y + 3*mm, 'TOTAL:')
            canvas.drawRightString(right,  tot_y + 3*mm, _fmt(total_final))

        # Líneas de firma (izquierda)
        sig_y = tot_y + 5*mm
        canvas.setLineWidth(0.4)
        canvas.setStrokeColor(colors.HexColor('#9ca3af'))
        canvas.line(x0, sig_y, x0 + 60*mm, sig_y)
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(colors.HexColor('#6b7280'))
        canvas.drawString(x0, sig_y - 4*mm, 'Firma y aclaración')

        # Footer empresa
        canvas.setFont('Helvetica', 6.5)
        canvas.setFillColor(colors.HexColor('#9ca3af'))
        canvas.drawCentredString(PAGE_W / 2, 8*mm, footer_txt)
        canvas.restoreState()

    doc_obj = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=mg, rightMargin=mg,
        topMargin=5*mm, bottomMargin=BOTTOM_BLOCK
    )

    story = []

    # ── Header ────────────────────────────────────────────────────────────────
    logo_w = 38 * mm
    if os.path.exists(LOGO_PATH):
        logo_cell = [Image(LOGO_PATH, width=logo_w, height=logo_w * 0.32)]
    else:
        logo_cell = [Paragraph(f'<b>{razon_soc}</b>', s_h2)]

    doc_w  = 52 * mm
    emp_fixed = 72 * mm                          # ancho fijo del bloque empresa
    spacer_w  = usable_w - logo_w - doc_w - emp_fixed  # empuja emp_cell al margen derecho

    doc_cell = [
        Paragraph('<b>NOTA DE PEDIDO</b>',
                  ParagraphStyle('dt', fontSize=13, fontName='Helvetica-Bold', alignment=TA_CENTER)),
        Paragraph(f'<b>N°  {num_fmt}</b>',
                  ParagraphStyle('dn', fontSize=11, fontName='Helvetica-Bold', alignment=TA_CENTER)),
        Spacer(1, 3*mm),
        Paragraph(f'Fecha: <b>{_d(fec_comp)}</b>',
                  ParagraphStyle('df', fontSize=8.5, fontName='Helvetica', alignment=TA_CENTER)),
        Paragraph(f'Fecha Entrega: <b>{_d(fec_entrega)}</b>',
                  ParagraphStyle('de', fontSize=8.5, fontName='Helvetica', alignment=TA_CENTER)),
    ]
    s_emp_l  = ParagraphStyle('pel',  fontSize=8, leading=11, fontName='Helvetica')
    s_emp_lb = ParagraphStyle('pelb', fontSize=8, leading=11, fontName='Helvetica-Bold')
    emp_cell = [
        Paragraph(f'<b>{razon_soc}</b>', s_emp_lb),
        Paragraph(dir_emp,              s_emp_l),
        Paragraph(f'Tel: {tel_emp}',    s_emp_l),
        Paragraph(f'CUIT: {cuit_emp}',  s_emp_l),
        Paragraph(email_emp,            s_emp_l),
        Paragraph(web_emp,              s_emp_l),
    ]

    hdr_tbl = Table([[logo_cell, doc_cell, Paragraph('', s_emp_l), emp_cell]],
                    colWidths=[logo_w, doc_w, spacer_w, emp_fixed])
    hdr_tbl.setStyle(TableStyle([
        ('VALIGN',       (0,0),(-1,-1), 'MIDDLE'),
        ('ALIGN',        (1,0),(1,0),   'CENTER'),
        ('ALIGN',        (3,0),(3,0),   'LEFT'),
        ('LEFTPADDING',  (0,0),(-1,-1), 4),
        ('RIGHTPADDING', (0,0),(-1,-1), 0),
        ('TOPPADDING',   (0,0),(-1,-1), 0),
        ('BOTTOMPADDING',(0,0),(-1,-1), 0),
    ]))
    story.append(hdr_tbl)
    story.append(HRFlowable(width='100%', thickness=1.5, color=colors.HexColor('#1a56db'),
                            spaceBefore=4*mm, spaceAfter=4*mm))

    # ── Datos cliente ──────────────────────────────────────────────────────────
    pad = TableStyle([
        ('TOPPADDING',   (0,0),(-1,-1), 3),
        ('BOTTOMPADDING',(0,0),(-1,-1), 3),
        ('LEFTPADDING',  (0,0),(-1,-1), 4),
        ('RIGHTPADDING', (0,0),(-1,-1), 4),
    ])
    lbl_w = 24 * mm
    half  = (usable_w - 5*mm) / 2
    cli_nombre = f"{cod_cli} - {rs_cli}".strip(' -') if cod_cli and rs_cli else (rs_cli or cod_cli or '')
    cli_l = Table([
        [Paragraph('Cliente:',   s_label), Paragraph(cli_nombre,    s_val)],
        [Paragraph('CUIT:',      s_label), Paragraph(cuit_cli or '', s_val)],
        [Paragraph('Dirección:', s_label), Paragraph(dir_cli or '',  s_val)],
        [Paragraph('Teléfono:',  s_label), Paragraph(tel_pdf or '',  s_val)],
    ], colWidths=[lbl_w, half - lbl_w])
    cli_r = Table([
        [Paragraph('Cond. Venta:', s_label), Paragraph(cond_venta or '',      s_val)],
        [Paragraph('Transporte:',  s_label), Paragraph(transporte_desc or '', s_val)],
        [Paragraph('Depósito:',    s_label), Paragraph(deposito_desc or '',   s_val)],
        [Paragraph('Vendedor:',    s_label), Paragraph(vendedor_nombre or '', s_val)],
    ], colWidths=[26*mm, half - 26*mm])
    for t in (cli_l, cli_r):
        t.setStyle(pad)

    cli_outer = Table([[cli_l, Spacer(5*mm,1), cli_r]], colWidths=[half, 5*mm, half])
    cli_outer.setStyle(TableStyle([
        ('VALIGN', (0,0),(-1,-1), 'TOP'),
        ('BOX',    (0,0),(0,0),   0.5, colors.HexColor('#d1d5db')),
        ('BOX',    (2,0),(2,0),   0.5, colors.HexColor('#d1d5db')),
        ('BACKGROUND',(0,0),(0,0), colors.HexColor('#f9fafb')),
        ('BACKGROUND',(2,0),(2,0), colors.HexColor('#f9fafb')),
    ]))
    story.append(cli_outer)
    story.append(Spacer(1, 5*mm))

    # ── Tabla artículos ────────────────────────────────────────────────────────
    # col_w suma = usable_w exacto
    col_w = [18*mm, 78*mm, 16*mm, 28*mm, 14*mm, 28*mm]   # = 182mm
    hdr_row = [
        Paragraph('Código',      s_c_b),
        Paragraph('Descripción', s_c_b),
        Paragraph('Cantidad',    s_c_b),
        Paragraph('P. Unitario', s_c_b),
        Paragraph('Dto %',       s_c_b),
        Paragraph('P. Total',    s_c_b),
    ]
    s_ri = ParagraphStyle('ri', fontSize=8, leading=11, fontName='Helvetica', alignment=TA_RIGHT)
    items_data = [hdr_row]
    for it in items:
        cod_art, desc_art, cant, pu, bonif, ptotal, piva = it
        cod_str = str(cod_art or '').strip()
        if cod_str.endswith('.0'):
            cod_str = cod_str[:-2]
        try:
            cant_str = str(int(float(cant))) if float(cant) == int(float(cant)) else str(cant)
        except Exception:
            cant_str = str(cant)
        items_data.append([
            Paragraph(cod_str,                    s_c),
            Paragraph(str(desc_art or '').strip(), s_norm),
            Paragraph(cant_str,                   s_c),
            Paragraph(_fmt(pu),                   s_ri),
            Paragraph(f"{float(bonif or 0):.2f}%", s_c),
            Paragraph(_fmt(ptotal),               s_ri),
        ])

    items_tbl = Table(items_data, colWidths=col_w, repeatRows=1)
    items_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),(-1,0),  colors.HexColor('#1a56db')),
        ('TEXTCOLOR',     (0,0),(-1,0),  colors.white),
        ('FONTNAME',      (0,0),(-1,0),  'Helvetica-Bold'),
        ('ROWBACKGROUNDS',(0,1),(-1,-1), [colors.white, colors.HexColor('#eff6ff')]),
        ('GRID',          (0,0),(-1,-1), 0.3, colors.HexColor('#e5e7eb')),
        ('TOPPADDING',    (0,0),(-1,-1), 4),
        ('BOTTOMPADDING', (0,0),(-1,-1), 4),
        ('LEFTPADDING',   (0,0),(-1,-1), 4),
        ('RIGHTPADDING',  (0,0),(-1,-1), 4),
        ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
    ]))
    story.append(items_tbl)

    doc_obj.build(story, onFirstPage=_draw_bottom, onLaterPages=_draw_bottom)
    buf.seek(0)
    fname = f"Nota_Pedido_{numero}.pdf"
    return StreamingResponse(buf, media_type='application/pdf',
                             headers={'Content-Disposition': f'inline; filename="{fname}"'})


class ItemDoc(BaseModel):
    codigoarticulo: str
    codigoparticular: str = ""
    descripcion: str
    cantidad: float
    preciounitario: float
    descuento: float = 0.0
    porcentajeiva: float = 21.0
    linea_presupuesto: Optional[int] = None   # línea en CUERPOPRESUPUESTOS si viene de presupuesto

class NuevoPedido(BaseModel):
    codigocliente: str
    razonsocial: str
    codigousuario: str
    comentarios: str = ""
    codigomultiplazo: Optional[str] = None
    codigotransporte: str = "0"
    codigodeposito: str = "001"
    domicilio_entrega: Optional[str] = None
    numero_presupuesto: Optional[str] = None  # si el pedido absorbe un presupuesto aprobado
    descuento_general: float = 0.0
    descuento_promo_pct: float = 0.0      # descuento por escalón de combos (tipo producto)
    descuento_promo_nombre: str = ""       # nombre del escalón aplicado
    alerta_cobranzas: bool = False   # cliente tenía observación en el ABM y el corredor confirmó cargar igual
    motivo_alerta: str = ""          # texto original de la observación del cliente (CLIENTES.COMENTARIOS)
    nombre_vendedor: str = ""        # nombre del corredor logueado, para el aviso a Cobranzas
    items: list[ItemDoc]

@app.post("/pedidos")
def crear_pedido(body: NuevoPedido, db: str = Query("oficial")):
    db_path = DATABASE_MLT if db == 'sw' else DATABASE
    # LOG para debug de routing
    try:
        import datetime
        with open('C:/api_vendedores/pedido_debug.log', 'a', encoding='utf-8') as _lf:
            _lf.write(f"{datetime.datetime.now()} | db_param={db!r} | db_path={db_path!r} | usuario={body.codigousuario!r}\n")
    except Exception:
        pass
    # Obtener datos del cliente — busca en oficial si SW no tiene CLIENTES
    direccion, tipoiva, telefono, atencion_p, cuit_cli = '-', 'CF', '-', '', ''
    cli_transp_p, cli_reparto_p = '', ''
    def _tel_p(v):
        if v is None: return ''
        s = str(v).strip().lstrip('-').strip()
        try:
            s = str(int(float(s))) if s and ('E' in s.upper() or '.' in s) else s
        except Exception:
            pass
        return s
    for cli_db in ([db_path, DATABASE] if db == 'sw' else [db_path]):
        try:
            c_cli = conn('WIN1252', db=cli_db)
            cur_cli = c_cli.cursor()
            cur_cli.execute(
                'SELECT DIRECCION, TELEFONO, TELEFONOCELULAR, '
                'NOMBRE, APELLIDO, CODIGOTRANSPORTE, REPARTOPROPIO, CUIT, CONDICIONIVA '
                'FROM "CLIENTES" WHERE CODIGOCLIENTE = ?',
                (body.codigocliente,)
            )
            cli = cur_cli.fetchone()
            c_cli.close()
            if cli:
                direccion      = (cli[0] or '').strip() or '-'
                _VALID_IVA = {'CF','EX','EXPO','MOA','NR','RI','RM'}
                _cond  = (cli[8] or '').strip().upper()
                if _cond in _VALID_IVA:
                    tipoiva = _cond
                elif 'INSCRIPTO' in _cond:
                    tipoiva = 'RI'
                elif 'EXENTO' in _cond:
                    tipoiva = 'EX'
                elif 'MONOTRIB' in _cond:
                    tipoiva = 'MOA'
                elif 'NO RESPONSABL' in _cond:
                    tipoiva = 'NR'
                elif 'EXPORTAC' in _cond:
                    tipoiva = 'EXPO'
                else:
                    tipoiva = 'CF'
                tel_r          = _tel_p(cli[1])
                tel_c          = _tel_p(cli[2])
                telefono       = tel_r or tel_c or '-'
                nombre_c       = (cli[3] or '').strip()
                apellido_c     = (cli[4] or '').strip()
                atencion_p     = f"{nombre_c} {apellido_c}".strip()
                cli_transp_p   = str(cli[5] or '').strip()
                cli_reparto_p  = str(cli[6] or '').strip()
                cuit_cli       = str(cli[7] or '').strip()
                break
        except Exception:
            continue

    if body.domicilio_entrega:
        direccion = body.domicilio_entrega

    # Si cliente tiene reparto propio y no se eligió transporte, buscar código de REPARTO PROPIO
    if cli_reparto_p == '1' and (not body.codigotransporte or body.codigotransporte == '0'):
        try:
            c_tr = conn('WIN1252', db=DATABASE)
            cur_tr = c_tr.cursor()
            cur_tr.execute(
                "SELECT FIRST 1 CODIGOTRANSPORTE FROM \"TRANSPORTES\" "
                "WHERE UPPER(DESCRIPCION) CONTAINING 'REPARTO PROPIO' AND ACTIVO <> '0'"
            )
            row_tr = cur_tr.fetchone()
            c_tr.close()
            if row_tr:
                cli_transp_p = str(row_tr[0]).strip()
        except Exception:
            pass

    from datetime import datetime, timedelta
    now = datetime.now()
    fecha = now.strftime('%Y-%m-%d %H:%M:%S')
    fecha_entrega = (now + timedelta(days=2)).strftime('%Y-%m-%d %H:%M:%S')
    subtotal = sum(it.cantidad * it.preciounitario * (1 - it.descuento / 100) for it in body.items)
    iva1 = sum(it.cantidad * it.preciounitario * (1 - it.descuento / 100) * (it.porcentajeiva / 100) for it in body.items)
    dto_gral = max(0.0, min(float(body.descuento_general or 0), 100.0))
    dto_promo = max(0.0, min(float(body.descuento_promo_pct or 0), 100.0))
    total = subtotal * (1 - dto_gral / 100) * (1 - dto_promo / 100)
    # DESCUENTOPORCENTAJE/DESCUENTOMONTO/DESCUENTODESCRIPCION (CABEZAPEDIDOS): mismo
    # criterio que en crear_presupuesto — dto_gral (admin) y dto_promo (combo/escalón,
    # incluye recupero de cartera) son mutuamente excluyentes, se graba el que esté activo.
    _dto_pct_final = dto_gral if dto_gral > 0 else dto_promo
    _dto_monto_final = round(subtotal * dto_gral / 100, 2) if dto_gral > 0 else round(subtotal * dto_promo / 100, 2)
    _dto_desc_final = (f'{dto_gral:.6f} %' if dto_gral > 0
                        else (f'{dto_promo:.6f} %' if dto_promo > 0 else '0,000000 %'))

    # Resolver CODIGOCLIENTE correcto según BD destino
    # body.codigocliente viene siempre en código L1; para SW hay que traducirlo via CODIGOPARTICULAR
    codigocliente_destino = body.codigocliente
    if db == 'sw':
        try:
            _c1 = conn('WIN1252', db=DATABASE)
            _cur1 = _c1.cursor()
            _cur1.execute(
                'SELECT CODIGOPARTICULAR FROM "CLIENTES" WHERE CODIGOCLIENTE = ?',
                (body.codigocliente,)
            )
            _r1 = _cur1.fetchone()
            _c1.close()
            cod_particular = (_r1[0] or '').strip() if _r1 else ''
            if cod_particular:
                _c2 = conn('WIN1252', db=DATABASE_MLT)
                _cur2 = _c2.cursor()
                _cur2.execute(
                    'SELECT CODIGOCLIENTE FROM "CLIENTES" WHERE CODIGOPARTICULAR = ? OR CODIGOCLIENTE = ?',
                    (cod_particular, cod_particular)
                )
                _r2 = _cur2.fetchone()
                _c2.close()
                if _r2:
                    codigocliente_destino = str(_r2[0]).strip()
        except Exception:
            pass  # si falla la traducción, usa el código original

    # Pedidos siempre en DATABASE (CABEZAPEDIDOS existe solo en L1)
    c = conn('WIN1252', db=DATABASE)
    cur = c.cursor()
    try:
        # Transporte: usar el del cliente como fallback
        transp_final = (
            body.codigotransporte if body.codigotransporte and body.codigotransporte != '0'
            else (cli_transp_p if cli_transp_p and cli_transp_p != '0' else '0')
        )
        deposito = body.codigodeposito or '001'

        # L1 y SW → DATABASE → CABEZAPEDIDOS / CUERPOPEDIDOS
        # SW usa ENTREGAR='1', INTERES=0.0, CODIGOFINANCIACION='0'; L1 los deja en NULL
        cur.execute(
            "SELECT VALOR FROM \"PARAMETROS\" WHERE TRIM(TIPODOCUMENTO) = 'NP' WITH LOCK"
        )
        row_param_np = cur.fetchone()
        nuevo_num_int_o = int(float(row_param_np[0])) if row_param_np else 0
        cur.execute(
            'SELECT MAX(CAST(NUMEROCOMPROBANTE AS INTEGER)) FROM "CABEZAPEDIDOS"'
            ' WHERE TIPOCOMPROBANTE = ?', ('NP',)
        )
        max_ped = int(cur.fetchone()[0] or 0)
        if nuevo_num_int_o <= max_ped:
            nuevo_num_int_o = max_ped + 1
        nuevo_num = str(nuevo_num_int_o)
        cur.execute(
            "UPDATE \"PARAMETROS\" SET VALOR = ? WHERE TRIM(TIPODOCUMENTO) = 'NP'",
            (nuevo_num_int_o + 1,)
        )
        cur.execute("""
            INSERT INTO "CABEZAPEDIDOS"
            (TIPOCOMPROBANTE, NUMEROCOMPROBANTE, CODIGOCLIENTE, RAZONSOCIAL,
             FECHACOMPROBANTE, TOTAL, ANULADA, CODIGOUSUARIO, COMENTARIOS,
             PORCIVA1, IVA1, PORCIVA2, IVA2, PAGADO, OPERACION, ENTREGAR,
             DIRECCION, TIPOIVA, TELEFONO, FECHAENTREGA, PRIORIDAD,
             VENTECNICOS, COMENTARIOSST, FORMAPAGO, CODIGOTRANSPORTE,
             CODIGOMULTIPLAZO,
             COEFICIENTEIVA, CODIGOMONEDA, COTIZACION,
             NUMEROTRANSACCION, COTIZACIONFIJA, PORCENTAJEFLETE, MONTOFLETE,
             LISTAPRECIO, CLASECOMPROBANTE, CODIGOACOPIO,
             VALIDACTACTE, NUMEROAUTORIZACIONENTREGA, CUIT,
             DESCUENTOPORCENTAJE, DESCUENTOMONTO, DESCUENTODESCRIPCION,
             CODIGOUSUARIO2, FECHATERMINADA, FECHAMODIFICACION,
             INTERES, CODIGOFINANCIACION)
            VALUES ('NP', ?, ?, ?, ?, ?, '0', ?, ?,
                    0.0, ?, 0.0, 0.0, 0.0, '1', ?,
                    ?, ?, ?, ?, '1', '0', ' ', '', ?,
                    ?,
                    1.0, 'PESOS', 1.0, '0', '0', 0.0, 0.0,
                    '1', '0', '0',
                    1, '0', ?,
                    ?, ?, ?,
                    ?, NULL, ?,
                    ?, ?)
        """, (nuevo_num, body.codigocliente, body.razonsocial, fecha,
              round(total, 2), body.codigousuario.upper(), body.comentarios,
              iva1, '1' if db == 'sw' else '0',
              direccion, tipoiva, telefono, fecha_entrega,
              transp_final,
              int(body.codigomultiplazo) if body.codigomultiplazo else 0,
              cuit_cli,
              -round(_dto_pct_final, 6), -_dto_monto_final, _dto_desc_final,
              body.codigousuario.upper(), fecha,
              0.0 if db == 'sw' else None,
              '0' if db == 'sw' else None))
        # Resolver codigoparticular → CODIGOARTICULO real en Firebird (evita código incorrecto en ERP)
        _cod_parts = list({(it.codigoparticular or it.codigoarticulo).strip() for it in body.items if it.codigoarticulo})
        _cod_map = {}
        if _cod_parts:
            try:
                _ph = ','.join('?' * len(_cod_parts))
                _rows = c.cursor().execute(
                    f'SELECT TRIM(CODIGOARTICULO), TRIM(CODIGOPARTICULAR) FROM "ARTICULOS" WHERE TRIM(CODIGOPARTICULAR) IN ({_ph})',
                    _cod_parts).fetchall()
                for _r in _rows:
                    if _r[1]: _cod_map[str(_r[1]).strip()] = str(_r[0]).strip()
            except Exception: pass

        for i, it in enumerate(body.items, 1):
            _real_cod = _cod_map.get((it.codigoparticular or it.codigoarticulo).strip(), it.codigoarticulo)
            subtotal_item = it.cantidad * it.preciounitario * (1 - it.descuento / 100)
            cur.execute("""
                INSERT INTO "CUERPOPEDIDOS"
                (TIPOCOMPROBANTE, NUMEROCOMPROBANTE, LINEA, CODIGOARTICULO,
                 DESCRIPCION, CANTIDAD, DESCUENTO, PRECIOUNITARIO, PRECIOTOTAL,
                 PORCENTAJEIVA, CANTIDADREMITIDA, ESCONJUNTO,
                 GARANTIA, LOTE, FECHAMODIFICACION, NUMEROTRANSACCION,
                 CODIGODEPOSITO, CANTIDADPREPARADA, CANTIDADCOMPRA,
                 CANTIDADPRODUCCION, CANTIDADENVIADA, CANTIDADCANCELADA,
                 COEFICIENTECONVERSION, ESEXENTO, ESPRECIOPACTADO,
                 CODIGOPARTICULAR)
                VALUES ('NP', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '0',
                        0, '000', ?, '0', ?, ?, 0,
                        0, 0, 0,
                        0.0, '0', '0',
                        ?)
            """, (nuevo_num, i, _real_cod, it.descripcion,
                  it.cantidad, it.descuento, it.preciounitario, subtotal_item,
                  it.porcentajeiva, 0, fecha, deposito, 0,
                  it.codigoparticular or it.codigoarticulo))

        c.commit()

        # Si el pedido absorbe un presupuesto aprobado, actualizar CANTIDADREMITIDA
        if body.numero_presupuesto:
            try:
                c_pr = conn('WIN1252', db=db_path)
                cur_pr = c_pr.cursor()
                for it in body.items:
                    if it.linea_presupuesto is not None:
                        cur_pr.execute(
                            'UPDATE "CUERPOPRESUPUESTOS" '
                            'SET CANTIDADREMITIDA = COALESCE(CANTIDADREMITIDA, 0) + ? '
                            'WHERE NUMEROCOMPROBANTE = ? AND LINEA = ?',
                            (it.cantidad, body.numero_presupuesto, it.linea_presupuesto)
                        )
                c_pr.commit()
                c_pr.close()
            except Exception:
                pass  # No rollback del pedido por fallo en presupuesto

        # Invalidar caché del depósito afectado para que otros vendedores vean stock actualizado
        threading.Thread(target=_fma_cache_invalidate, args=([body.codigodeposito],), daemon=True).start()
        # Avisar a Cobranzas si el cliente tenía observación en el ABM y el corredor confirmó cargar igual
        if body.alerta_cobranzas:
            threading.Thread(target=_send_whatsapp_cobranzas, args=(
                'pedido', nuevo_num, body.nombre_vendedor, body.codigocliente,
                body.razonsocial, body.motivo_alerta
            ), daemon=True).start()
        return {"ok": True, "numero": nuevo_num, "total": round(total + iva1, 2), "_db_usado": db, "_db_path": db_path}
    except Exception as e:
        c.rollback()
        raise HTTPException(500, str(e))
    finally:
        c.close()


class CreditoInternoBody(BaseModel):
    codigocliente: str
    codigousuario: str
    cliente_razonsocial: str = ""
    monto: float
    oferta_id: Optional[int] = None
    oferta_nombre: str = ""
    escalon_monto_minimo: Optional[float] = None

@app.post("/pedidos/{numero}/credito_interno")
def generar_credito_interno_pedido(numero: str, body: CreditoInternoBody, db: str = Query("oficial"), _u=Depends(get_current_user)):
    """Genera el comprobante CI (crédito interno) real en Firebird para un Pedido YA
    CONFIRMADO, cuando el corredor eligió 'crédito' en vez de 'bonificación %' para un
    escalón de oferta que tenía ambos beneficios cargados (admin.html →
    financial_escalones[].monto_credito). Se llama SIEMPRE después de que /pedidos ya
    confirmó el pedido en Firebird, nunca antes — y si esto falla, el pedido NO se
    revierte (ya es una venta real; ver decisión de Eduardo 2026-07-31): se le informa
    al corredor con el número de pedido para que se resuelva aparte."""
    db_path = DATABASE_MLT if db == 'sw' else DATABASE
    partes_desc = [f"CREDITO POR PROMOCION - {body.oferta_nombre or 'Oferta'}"]
    if body.escalon_monto_minimo:
        partes_desc.append(f"(Escalon ${body.escalon_monto_minimo:,.0f})".replace(',', '.'))
    partes_desc.append(f"- Pedido Nro {numero}")
    descripcion = " ".join(partes_desc)

    admin_c = _admin_db()
    try:
        resultado = _generar_credito_interno_ci(
            codigo_cliente=body.codigocliente,
            codigo_usuario=body.codigousuario,
            monto=body.monto,
            descripcion=descripcion,
            db_path=db_path,
        )
    except Exception as e:
        try:
            admin_c.execute(
                "INSERT INTO creditos_internos_log "
                "(codigousuario, codigocliente, cliente_razonsocial, oferta_id, oferta_nombre, "
                "escalon_monto_minimo, monto, pedido_numero, pedido_db, estado, error_detalle) "
                "VALUES (?,?,?,?,?,?,?,?,?,'error',?)",
                (body.codigousuario, body.codigocliente, body.cliente_razonsocial, body.oferta_id,
                 body.oferta_nombre, body.escalon_monto_minimo, body.monto, numero, db, str(e))
            )
            admin_c.commit()
        except Exception:
            pass
        finally:
            admin_c.close()
        raise HTTPException(
            500,
            f"El pedido {numero} se confirmó correctamente, pero el crédito interno de "
            f"${body.monto:,.0f} NO se pudo generar en Firebird ({e}). El pedido queda "
            f"en pie — avisá para que el crédito se cargue manualmente."
        )

    try:
        admin_c.execute(
            "INSERT INTO creditos_internos_log "
            "(codigousuario, codigocliente, cliente_razonsocial, oferta_id, oferta_nombre, "
            "escalon_monto_minimo, monto, pedido_numero, pedido_db, numero_ci, codigo_asiento, estado) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,'ok')",
            (body.codigousuario, body.codigocliente, body.cliente_razonsocial, body.oferta_id,
             body.oferta_nombre, body.escalon_monto_minimo, body.monto, numero, db,
             resultado["numero_ci"], resultado["codigo_asiento"])
        )
        admin_c.commit()
    except Exception:
        pass  # el CI ya se generó en Firebird; que falle el log de auditoría no debe romper la respuesta
    finally:
        admin_c.close()

    from datetime import datetime
    return {
        "ok": True,
        "numero_ci": resultado["numero_ci"],
        "codigo_asiento": resultado["codigo_asiento"],
        "monto": body.monto,
        "fecha": datetime.now().strftime('%d/%m/%Y'),
        "oferta_nombre": body.oferta_nombre,
        "cliente_razonsocial": body.cliente_razonsocial,
        "codigocliente": body.codigocliente,
        "pedido_numero": numero,
        "descripcion": descripcion,
    }

@app.get("/debug/cab_full/{numero}")
def debug_cab_full(numero: str):
    """Muestra TODOS los campos de CABEZAPEDIDOS para un NP dado."""
    try:
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()
        cur.execute(
            'SELECT * FROM "CABEZAPEDIDOS" WHERE TIPOCOMPROBANTE = ? AND NUMEROCOMPROBANTE = ?',
            ('NP', numero)
        )
        row = cur.fetchone()
        if not row:
            c.close()
            return {"error": f"NP {numero} no encontrado"}
        cols = [d[0] for d in cur.description]
        c.close()
        return dict(zip(cols, [str(v) if v is not None else None for v in row]))
    except Exception as e:
        return {"error": str(e)}

_PROXIMO_CACHE_TTL = int(os.getenv('PROXIMO_CACHE_TTL', 5))  # solo texto de previsualización
_proximo_np_cache: dict = {"ts": 0, "val": None}
_proximo_pr_cache: dict = {}

@app.get("/pedidos/proximo")
def get_proximo_pedido(db: str = Query("oficial")):
    # El contador NP siempre está en DATABASE (BD principal), sin importar si es SW u oficial
    _now = time.time()
    if _proximo_np_cache["val"] is not None and (_now - _proximo_np_cache["ts"]) < _PROXIMO_CACHE_TTL:
        return _proximo_np_cache["val"]
    try:
        c = conn('LATIN1', db=DATABASE)
        cur = c.cursor()
        cur.execute("SELECT VALOR FROM \"PARAMETROS\" WHERE TRIM(TIPODOCUMENTO) = 'NP'")
        row = cur.fetchone()
        c.close()
        if not row:
            raise HTTPException(500, "No se encontró parámetro de numeración NP")
        resultado = {"proximo": int(float(row[0]))}
        _proximo_np_cache["ts"] = _now
        _proximo_np_cache["val"] = resultado
        return resultado
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/presupuestos/proximo")
def get_proximo_presupuesto(db: str = Query("oficial")):
    db_path = DATABASE_MLT if db == 'sw' else DATABASE
    _now = time.time()
    _cached = _proximo_pr_cache.get(db_path)
    if _cached and (_now - _cached[0]) < _PROXIMO_CACHE_TTL:
        return _cached[1]
    try:
        c = conn('LATIN1', db=db_path)
        cur = c.cursor()
        cur.execute("SELECT VALOR FROM \"PARAMETROS\" WHERE TIPODOCUMENTO = 'PR'")
        row = cur.fetchone()
        c.close()
        if not row:
            raise HTTPException(500, "No se encontró parámetro de numeración PR")
        resultado = {"proximo": int(float(row[0]))}
        _proximo_pr_cache[db_path] = (_now, resultado)
        return resultado
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))

# ─── Presupuestos (solo del vendedor) ─────────────────────────────────────────
@app.get("/presupuestos")
def get_presupuestos(
    vendedor: str,
    cliente: Optional[str] = None,
    limit: int = Query(50, le=200),
    offset: int = 0,
    _user=Depends(get_current_user)
):
    c = conn()
    cur = c.cursor()
    params = [vendedor.upper()]
    where_cli = ""
    if cliente:
        where_cli = "AND CODIGOCLIENTE = ?"
        params.append(cliente)

    cur.execute(f"""
        SELECT FIRST {limit} SKIP {offset}
            NUMEROCOMPROBANTE, CODIGOCLIENTE, RAZONSOCIAL,
            FECHACOMPROBANTE, FECHAVENCIMIENTO, TOTAL, ANULADA, COMENTARIOS
        FROM "CABEZAPRESUPUESTOS"
        WHERE CODIGOUSUARIO = ? AND ANULADA = '0'
            {where_cli}
        ORDER BY FECHACOMPROBANTE DESC, NUMEROCOMPROBANTE DESC
    """, params)
    rows = cur.fetchall()
    c.close()
    return [{
        "numero": r[0], "cod_cliente": r[1], "razonsocial": r[2],
        "fecha": r[3], "fecha_vto": r[4], "total": r[5],
        "anulada": r[6], "comentarios": r[7]
    } for r in rows]

@app.get("/presupuestos/pendientes/{codigocliente}")
def get_presupuestos_pendientes(codigocliente: str, db: str = Query("oficial")):
    """Presupuestos aprobados con items aún no completados para el cliente."""
    # CABEZAPRESUPUESTOS siempre en BD principal
    c = conn('LATIN1', db=DATABASE)
    cur = c.cursor()
    # Cabezas: aprobadas (FECHAAPROBADO IS NOT NULL), no anuladas, con items pendientes
    cur.execute("""
        SELECT DISTINCT cp.NUMEROCOMPROBANTE, cp.FECHACOMPROBANTE, cp.TOTAL, cp.COMENTARIOS,
               cp.CODIGOMULTIPLAZO, cp.CODIGOTRANSPORTE, cp.DIRECCION
        FROM "CABEZAPRESUPUESTOS" cp
        JOIN "CUERPOPRESUPUESTOS" cu ON cu.NUMEROCOMPROBANTE = cp.NUMEROCOMPROBANTE
        WHERE cp.CODIGOCLIENTE = ?
          AND cp.ANULADA = '0'
          AND cp.FECHAAPROBADO IS NOT NULL
          AND cp.FECHAAPROBADO > CAST('1900-01-02 00:00:00' AS TIMESTAMP)
          AND (cu.CANTIDAD - COALESCE(cu.CANTIDADREMITIDA, 0)) > 0
        ORDER BY cp.NUMEROCOMPROBANTE DESC
    """, (codigocliente,))
    cabezas = cur.fetchall()
    # Descripción de la condición de venta de cada presupuesto — incluye MULTIPLAZOS
    # inactivos/dados de baja desde entonces, porque el presupuesto puede ser viejo y
    # la condición ya no estar vigente hoy (igual se necesita mostrarla/copiarla).
    multi_desc = {}
    codigos_mp = list({str(cab[4]).strip() for cab in cabezas if cab[4]})
    if codigos_mp:
        try:
            ph = ','.join('?' * len(codigos_mp))
            cur_mp = c.cursor()
            cur_mp.execute(f'SELECT CODIGOMULTIPLAZO, DESCRIPCION FROM "MULTIPLAZOS" WHERE CODIGOMULTIPLAZO IN ({ph})', codigos_mp)
            multi_desc = {str(r[0]).strip(): (r[1] or '').strip() for r in cur_mp.fetchall()}
        except Exception:
            pass
    result = []
    for cab in cabezas:
        numero = cab[0]
        cur2 = c.cursor()
        cur2.execute("""
            SELECT LINEA, CODIGOARTICULO, DESCRIPCION,
                   CANTIDAD, COALESCE(CANTIDADREMITIDA, 0),
                   BONIFICACION, PRECIOUNITARIO, PORCENTAJEIVA,
                   CODIGOPARTICULAR
            FROM "CUERPOPRESUPUESTOS"
            WHERE NUMEROCOMPROBANTE = ?
              AND (CANTIDAD - COALESCE(CANTIDADREMITIDA, 0)) > 0
            ORDER BY LINEA
        """, (numero,))
        items = []
        for r in cur2.fetchall():
            pendiente = float(r[3]) - float(r[4])
            cod_particular = (r[8] or '').strip() or (r[1] or '').strip()
            items.append({
                # "codigo" = CODIGOPARTICULAR (código de negocio), igual que en el resto
                # de la app — antes era r[1] (CODIGOARTICULO interno de Firebird), lo que
                # hacía que incorporarAlPedido() pidiera /stock/batch con el código interno
                # y a veces trajera el stock de OTRO artículo (coincidencia numérica entre
                # CODIGOARTICULO de uno y CODIGOPARTICULAR de otro).
                "linea": r[0], "codigo": cod_particular, "codigoparticular": cod_particular,
                "codigo_interno": r[1],
                "descripcion": r[2],
                "cantidad_total": float(r[3]),
                "cantidad_remitida": float(r[4]),
                "cantidad_pendiente": pendiente,
                "descuento": float(r[5]),
                "precio_unitario": float(r[6]),
                "iva": float(r[7])
            })
        result.append({
            "numero": numero,
            "fecha": str(cab[1]),
            "total": float(cab[2]) if cab[2] else 0,
            "comentarios": cab[3] or "",
            "codigomultiplazo": str(cab[4] or '0'),
            "multiplazo_descripcion": multi_desc.get(str(cab[4]).strip(), '') if cab[4] else '',
            "codigotransporte": str(cab[5] or '0'),
            "direccion": cab[6] or "",
            "codigodeposito": "001",
            "items": items
        })
    c.close()
    return result

@app.get("/debug/presupuesto/{numero}/multiplazo")
def debug_presupuesto_multiplazo(numero: str):
    """DIAGNÓSTICO TEMPORAL — no usado por el frontend. Muestra el valor y tipo crudo de
    CODIGOMULTIPLAZO en CABEZAPRESUPUESTOS para este presupuesto, lo compara contra
    MULTIPLAZOS con distintos casteos, y muestra cómo está guardado el código 51 para
    poder comparar tipos/formato exacto. Borrar una vez resuelto el problema de matching."""
    c = conn('LATIN1', db=DATABASE)
    cur = c.cursor()
    cur.execute('SELECT CODIGOMULTIPLAZO FROM "CABEZAPRESUPUESTOS" WHERE NUMEROCOMPROBANTE = ?', (numero,))
    row = cur.fetchone()
    raw = row[0] if row else None
    info = {
        "numero": numero,
        "raw_value": raw,
        "raw_repr": repr(raw),
        "raw_type": type(raw).__name__,
    }
    candidatos = []
    if raw is not None:
        candidatos.append(("str_stripped", str(raw).strip()))
        try:
            candidatos.append(("int", int(raw)))
        except Exception:
            pass
    intentos = {}
    for etiqueta, valor in candidatos:
        try:
            cur2 = c.cursor()
            cur2.execute('SELECT CODIGOMULTIPLAZO, DESCRIPCION FROM "MULTIPLAZOS" WHERE CODIGOMULTIPLAZO = ?', (valor,))
            r = cur2.fetchone()
            intentos[etiqueta] = {"parametro_usado": repr(valor), "match": {"codigo": r[0], "descripcion": r[1]} if r else None}
        except Exception as e:
            intentos[etiqueta] = {"parametro_usado": repr(valor), "error": str(e)}
    # Cómo está guardado el código 51 en MULTIPLAZOS, para comparar tipo/formato
    try:
        cur3 = c.cursor()
        cur3.execute('SELECT CODIGOMULTIPLAZO, DESCRIPCION FROM "MULTIPLAZOS" WHERE CODIGOMULTIPLAZO = ?', ('51',))
        r3 = cur3.fetchone()
        info["multiplazo_51_via_str"] = {"codigo": r3[0], "codigo_repr": repr(r3[0]), "codigo_type": type(r3[0]).__name__, "descripcion": r3[1]} if r3 else None
    except Exception as e:
        info["multiplazo_51_via_str_error"] = str(e)
    try:
        cur4 = c.cursor()
        cur4.execute('SELECT CODIGOMULTIPLAZO, DESCRIPCION FROM "MULTIPLAZOS" WHERE CODIGOMULTIPLAZO = ?', (51,))
        r4 = cur4.fetchone()
        info["multiplazo_51_via_int"] = {"codigo": r4[0], "codigo_repr": repr(r4[0]), "codigo_type": type(r4[0]).__name__, "descripcion": r4[1]} if r4 else None
    except Exception as e:
        info["multiplazo_51_via_int_error"] = str(e)
    info["intentos_match"] = intentos
    c.close()
    return info

@app.get("/presupuestos/{numero}/cabecera")
def get_presupuesto_cabecera(numero: str):
    """Condición de venta (multiplazo) de un presupuesto puntual — para mostrar en el
    modal de 'Últimos presupuestos' sin cambiar la forma de /detalle (lista de ítems)."""
    c = conn('LATIN1', db=DATABASE)
    cur = c.cursor()
    cur.execute('SELECT CODIGOMULTIPLAZO FROM "CABEZAPRESUPUESTOS" WHERE NUMEROCOMPROBANTE = ?', (numero,))
    row = cur.fetchone()
    codigomultiplazo = str(row[0]).strip() if row and row[0] else ''
    descripcion = ''
    if codigomultiplazo:
        try:
            cur2 = c.cursor()
            cur2.execute('SELECT DESCRIPCION FROM "MULTIPLAZOS" WHERE CODIGOMULTIPLAZO = ?', (codigomultiplazo,))
            r2 = cur2.fetchone()
            descripcion = (r2[0] or '').strip() if r2 else ''
        except Exception:
            pass
    c.close()
    return {"codigomultiplazo": codigomultiplazo, "multiplazo_descripcion": descripcion}

@app.get("/presupuestos/{numero}/detalle")
def get_presupuesto_detalle(numero: str):
    c = conn()
    cur = c.cursor()
    cur.execute(
        "SELECT cp.LINEA, "
        "COALESCE(NULLIF(TRIM(cp.CODIGOPARTICULAR),''), NULLIF(TRIM(a.CODIGOPARTICULAR),''), TRIM(cp.CODIGOARTICULO)), "
        "cp.DESCRIPCION, cp.CANTIDAD, "
        "cp.BONIFICACION, cp.PRECIOUNITARIO, cp.PRECIOTOTAL, cp.PORCENTAJEIVA "
        "FROM \"CUERPOPRESUPUESTOS\" cp "
        "LEFT JOIN \"ARTICULOS\" a ON a.CODIGOARTICULO = cp.CODIGOARTICULO "
        "WHERE cp.NUMEROCOMPROBANTE = ? ORDER BY cp.LINEA",
        (numero,)
    )
    rows = cur.fetchall()
    c.close()
    return [{
        "linea": r[0], "codigo": r[1], "descripcion": r[2],
        "cantidad": r[3], "descuento": r[4],
        "precio_unitario": r[5], "precio_total": r[6], "iva": r[7]
    } for r in rows]

class NuevoPresupuesto(BaseModel):
    codigocliente: str
    razonsocial: str
    codigousuario: str
    comentarios: str = ""
    codigomultiplazo: Optional[str] = None
    codigotransporte: str = "0"
    domicilio_entrega: Optional[str] = None
    descuento_general: float = 0.0
    descuento_promo_pct: float = 0.0
    descuento_promo_nombre: str = ""
    alerta_cobranzas: bool = False
    motivo_alerta: str = ""
    nombre_vendedor: str = ""
    items: list[ItemDoc]

@app.post("/presupuestos")
def crear_presupuesto(body: NuevoPresupuesto, db: str = Query("oficial")):
    db_path = DATABASE_MLT if db == 'sw' else DATABASE
    # Obtener datos del cliente (WIN1252)
    try:
        c_cli = conn('WIN1252', db=db_path)
        cur_cli = c_cli.cursor()
        cur_cli.execute(
            'SELECT DIRECCION, TELEFONO, TELEFONOCELULAR, NOMBRE, APELLIDO, '
            'CODIGOTRANSPORTE, REPARTOPROPIO, CONDICIONIVA '
            'FROM "CLIENTES" WHERE CODIGOCLIENTE = ?',
            (body.codigocliente,)
        )
        cli = cur_cli.fetchone()
        c_cli.close()
        if cli:
            direccion     = (cli[0] or '').strip() or ''
            _VALID_IVA_P = {'CF','EX','EXPO','MOA','NR','RI','RM'}
            _cond_pr = (cli[7] or '').strip().upper()
            if _cond_pr in _VALID_IVA_P:
                tipoiva = _cond_pr
            elif 'INSCRIPTO' in _cond_pr:
                tipoiva = 'RI'
            elif 'EXENTO' in _cond_pr:
                tipoiva = 'EX'
            elif 'MONOTRIB' in _cond_pr:
                tipoiva = 'MOA'
            elif 'NO RESPONSABL' in _cond_pr:
                tipoiva = 'NR'
            elif 'EXPORTAC' in _cond_pr:
                tipoiva = 'EXPO'
            else:
                tipoiva = 'CF'
            def _tel(v):
                if v is None: return ''
                s = str(v).strip().lstrip('-').strip()
                # Si Firebird devuelve numérico (ej: 4.51E+14), convertir a entero
                try:
                    s = str(int(float(s))) if s and ('E' in s.upper() or '.' in s) else s
                except Exception:
                    pass
                return s
            tel_raw       = _tel(cli[1])
            tel_cel       = _tel(cli[2])
            telefono      = tel_raw or tel_cel
            nombre_c      = (cli[3] or '').strip()
            apellido_c    = (cli[4] or '').strip()
            atencion      = f"{nombre_c} {apellido_c}".strip()
            cli_transp    = str(cli[5] or '').strip()
            cli_reparto   = str(cli[6] or '').strip()
        else:
            direccion, tipoiva, telefono, atencion = '', 'CF', '', ''
            cli_transp, cli_reparto = '', ''
    except Exception:
        direccion, tipoiva, telefono, atencion = '', 'CF', '', ''
        cli_transp, cli_reparto = '', ''

    if body.domicilio_entrega:
        direccion = body.domicilio_entrega

    from datetime import datetime, timedelta
    now = datetime.now()
    fecha = now.strftime('%Y-%m-%d %H:%M:%S')
    fecha_vto = (now + timedelta(days=7)).strftime('%Y-%m-%d %H:%M:%S')
    subtotal = sum(it.cantidad * it.preciounitario * (1 - it.descuento / 100) for it in body.items)
    dto_gral = max(0.0, min(float(body.descuento_general or 0), 100.0))
    dto_promo = max(0.0, min(float(body.descuento_promo_pct or 0), 100.0))
    _factor_dto = (1 - dto_gral / 100) * (1 - dto_promo / 100)
    iva1 = sum(it.cantidad * it.preciounitario * (1 - it.descuento / 100) * (it.porcentajeiva / 100) for it in body.items) * _factor_dto
    total = subtotal * _factor_dto
    # DESCUENTOPORCENTAJE/DESCUENTOMONTO/DESCUENTODESCRIPCION (CABEZAPRESUPUESTOS): antes
    # solo se grababa dto_gral (descuento manual de admin) — dto_promo (combo, bonif. por
    # monto, o descuento por escalón/recupero de cartera) afectaba el TOTAL pero nunca
    # quedaba registrado en estos campos, así que Flexxus nunca lo mostraba en su propio
    # PDF/pantalla. Son mutuamente excluyentes en el frontend (nunca se envían los dos a
    # la vez), así que se graba el que esté activo.
    _dto_pct_final = dto_gral if dto_gral > 0 else dto_promo
    _dto_monto_final = round(subtotal * dto_gral / 100, 2) if dto_gral > 0 else round(subtotal * dto_promo / 100, 2)
    _dto_desc_final = (f'{dto_gral:.6f} %' if dto_gral > 0
                        else (f'{dto_promo:.6f} %' if dto_promo > 0 else '0,000000 %'))

    c = conn('LATIN1', db=db_path)
    cur = c.cursor()
    try:
        cur.execute("SELECT VALOR FROM \"PARAMETROS\" WHERE TIPODOCUMENTO = 'PR'")
        row_param = cur.fetchone()
        if not row_param:
            raise Exception("No se encontró parámetro de numeración PR en PARAMETROS")
        nuevo_num = str(int(float(row_param[0])))

        cur.execute("""
                INSERT INTO "CABEZAPRESUPUESTOS"
                (TIPOCOMPROBANTE, NUMEROCOMPROBANTE, CODIGOCLIENTE, RAZONSOCIAL,
                 FECHACOMPROBANTE, FECHAVENCIMIENTO, TOTAL, ANULADA,
                 CODIGOUSUARIO, COMENTARIOS,
                 PORCIVA1, IVA1, PORCIVA2, IVA2,
                 DIRECCION, TIPOIVA, TELEFONO,
                 CODIGOUSUARIO2, COEFICIENTEIVA, FECHAMODIFICACION,
                 CODIGORESPONSABLE, CODIGOMONEDA, COTIZACION,
                 ATENCION, NUMEROTRANSACCION, FORMAPAGO, CODIGOOPERACION,
                 LISTAPRECIO, COTIZACIONFIJA, PORCENTAJEFLETE, MONTOFLETE,
                 CLASECOMPROBANTE, CODIGOACOPIO,
                 CODIGOMULTIPLAZO, CODIGOTRANSPORTE,
                 DESCUENTOPORCENTAJE, DESCUENTOMONTO, DESCUENTODESCRIPCION,
                 FECHAAPROBADO, CODIGOUSUARIOAPROBACION)
                VALUES ('PR', ?, ?, ?, ?, ?, ?, '0', ?, ?,
                        21.0, ?, 0.0, 0.0,
                        ?, ?, ?,
                        ?, 1.0, ?,
                        ?, 'PESOS', 1.0,
                        ?, '0', '', '1',
                        '1', '0', 0.0, 0.0,
                        '0', '0',
                        ?, ?,
                        ?, ?, ?,
                        '1900-01-01 00:00:00', '')
            """, (nuevo_num, body.codigocliente, body.razonsocial, fecha,
                  fecha_vto, round(total, 2), body.codigousuario.upper(), body.comentarios,
                  iva1,
                  direccion, tipoiva, telefono,
                  body.codigousuario.upper(), fecha,
                  body.codigousuario.upper(), atencion,
                  int(body.codigomultiplazo) if body.codigomultiplazo else 0,
                  body.codigotransporte if body.codigotransporte and body.codigotransporte != '0'
                  else (cli_transp if cli_transp and cli_transp != '0' else '0'),
                  -round(_dto_pct_final, 6), -_dto_monto_final, _dto_desc_final))

        # Resolver codigoparticular → CODIGOARTICULO real en Firebird
        _cod_parts_p = list({(it.codigoparticular or it.codigoarticulo).strip() for it in body.items if it.codigoarticulo})
        _cod_map_p = {}
        if _cod_parts_p:
            try:
                _ph_p = ','.join('?' * len(_cod_parts_p))
                _rows_p = c.cursor().execute(
                    f'SELECT TRIM(CODIGOARTICULO), TRIM(CODIGOPARTICULAR) FROM "ARTICULOS" WHERE TRIM(CODIGOPARTICULAR) IN ({_ph_p})',
                    _cod_parts_p).fetchall()
                for _r in _rows_p:
                    if _r[1]: _cod_map_p[str(_r[1]).strip()] = str(_r[0]).strip()
            except Exception: pass

        for i, it in enumerate(body.items, 1):
            _real_cod_p = _cod_map_p.get((it.codigoparticular or it.codigoarticulo).strip(), it.codigoarticulo)
            subtotal_item = it.cantidad * it.preciounitario * (1 - it.descuento / 100)
            cur.execute("""
                INSERT INTO "CUERPOPRESUPUESTOS"
                (TIPOCOMPROBANTE, NUMEROCOMPROBANTE, LINEA, CODIGOARTICULO,
                 DESCRIPCION, CANTIDAD, BONIFICACION, PRECIOUNITARIO, PRECIOTOTAL,
                 PORCENTAJEIVA, CANTIDADREMITIDA, ESCONJUNTO,
                 GARANTIA, FECHAMODIFICACION, NUMEROTRANSACCION,
                 CODIGOPARTICULAR,
                 LOTE, INTERES, COEFICIENTECONVERSION,
                 ITEMGANADO, ESEXENTO, ESALTERNATIVO, ESPRECIOPACTADO)
                VALUES ('PR', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '0',
                        0, ?, '0',
                        ?,
                        '000', 0.0, 0.0,
                        '1', '0', '0', '0')
            """, (nuevo_num, i, _real_cod_p, it.descripcion,
                  it.cantidad, it.descuento, it.preciounitario, subtotal_item,
                  it.porcentajeiva, 0, fecha,
                  it.codigoparticular or it.codigoarticulo))

        cur.execute(
            "UPDATE \"PARAMETROS\" SET VALOR = ? WHERE TIPODOCUMENTO = 'PR'",
            (int(float(row_param[0])) + 1,)
        )
        c.commit()
        c.close()
        # Invalidar caché (presupuesto no tiene deposito en modelo, invalidar todos)
        threading.Thread(target=_fma_cache_invalidate, daemon=True).start()
        # Avisar a Cobranzas si el cliente tenía observación en el ABM y el corredor confirmó cargar igual
        if body.alerta_cobranzas:
            threading.Thread(target=_send_whatsapp_cobranzas, args=(
                'presupuesto', nuevo_num, body.nombre_vendedor, body.codigocliente,
                body.razonsocial, body.motivo_alerta
            ), daemon=True).start()
        return {"ok": True, "numero": nuevo_num, "total": round(total + iva1, 2)}

    except Exception as e:
        c.rollback()
        c.close()
        raise HTTPException(500, str(e))

# ─── PDF Simulación de Presupuesto (Modo Simulador) ────────────────────────────
# A diferencia de /presupuestos/{numero}/pdf (que lee CABEZAPRESUPUESTOS/CUERPOPRESUPUESTOS
# de Firebird por número de comprobante), este endpoint NO toca Firebird para nada —
# ni lectura ni escritura. Recibe los datos ya resueltos por el frontend (nombre de
# cliente, condición de venta, ítems) y arma un PDF autocontenido, rotulado como
# SIMULACIÓN, para que el corredor pueda probar escenarios de bonificación sin que
# quede ningún rastro en el ERP.
class SimuladorItem(BaseModel):
    codigo: str = ''
    descripcion: str = ''
    cantidad: float = 0
    precio: float = 0
    descuento: float = 0
    iva: float = 21

class SimuladorPresupuestoBody(BaseModel):
    cliente_codigo: str = ''
    cliente_nombre: str = ''
    vendedor_nombre: str = ''
    condicion_venta_texto: str = ''
    comentarios: str = ''
    discrimina_iva: bool = True
    items: List[SimuladorItem] = []

@app.post("/presupuestos/simular/pdf")
def presupuesto_simular_pdf(body: SimuladorPresupuestoBody):
    from datetime import datetime
    if not body.items:
        raise HTTPException(400, "Sin ítems para simular")

    razon_soc = 'MICROBELL S.A.'
    dir_emp   = 'PATAGONES 2675 PISO 3 - CABA - C1437JEA'
    tel_emp   = '+54 11 3988-0024'
    email_emp = 'info@microbellsa.com.ar'

    discIva = body.discrimina_iva
    subtotal_neto = 0.0
    iva_monto = 0.0
    rows = [['Cód.', 'Descripción', 'Cant.', 'Precio $', 'Bon%', 'IVA%', 'Subtotal']]
    for it in body.items:
        sub = it.cantidad * it.precio * (1 - (it.descuento or 0) / 100)
        subtotal_neto += sub
        iva_item = sub * (it.iva or 0) / 100 if discIva else 0
        iva_monto += iva_item
        rows.append([
            it.codigo, it.descripcion,
            f"{it.cantidad:g}", f"${it.precio:,.2f}".replace(',', '.'),
            f"{it.descuento:g}%", f"{it.iva:g}%" if discIva else '—',
            f"${sub:,.2f}".replace(',', '.'),
        ])
    total_final = subtotal_neto + iva_monto

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=12*mm, rightMargin=12*mm, topMargin=10*mm, bottomMargin=12*mm)
    styles = getSampleStyleSheet()
    s_title  = ParagraphStyle('title', fontSize=14, fontName='Helvetica-Bold', alignment=TA_CENTER)
    s_watermark = ParagraphStyle('wm', fontSize=11, fontName='Helvetica-Bold', alignment=TA_CENTER, textColor=colors.HexColor('#dc2626'))
    s_norm   = ParagraphStyle('norm', fontSize=8, leading=10, fontName='Helvetica')
    s_bold   = ParagraphStyle('bold', fontSize=8, leading=10, fontName='Helvetica-Bold')
    s_small  = ParagraphStyle('small', fontSize=7, leading=9, fontName='Helvetica', textColor=colors.HexColor('#6b7280'))
    s_r      = ParagraphStyle('r', fontSize=9, leading=11, fontName='Helvetica', alignment=TA_RIGHT)
    s_rb     = ParagraphStyle('rb', fontSize=10, leading=13, fontName='Helvetica-Bold', alignment=TA_RIGHT)

    story = []
    if os.path.exists(LOGO_PATH):
        try:
            _logo = Image(LOGO_PATH, width=36*mm, height=36*0.32*mm)
            _logo.hAlign = 'CENTER'
            story.append(_logo)
        except Exception:
            pass
    story.append(Paragraph(razon_soc, s_title))
    story.append(Spacer(1, 4))
    story.append(Paragraph('SIMULACIÓN DE PRESUPUESTO — NO VÁLIDO COMO PRESUPUESTO NI FACTURA', s_watermark))
    story.append(Paragraph('Documento generado solo a modo de prueba de escenario. No fue registrado en Flexxus ERP.', s_small))
    story.append(Spacer(1, 8))
    story.append(HRFlowable(width='100%', color=colors.HexColor('#dc2626'), thickness=1))
    story.append(Spacer(1, 8))

    info_lines = (
        f"<b>Cliente:</b> {body.cliente_nombre or '—'} ({body.cliente_codigo or 's/cod.'})<br/>"
        f"<b>Vendedor:</b> {body.vendedor_nombre or '—'}<br/>"
        f"<b>Condición de venta:</b> {body.condicion_venta_texto or '—'}<br/>"
        f"<b>Fecha simulación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    story.append(Paragraph(info_lines, s_norm))
    if body.comentarios:
        story.append(Spacer(1, 4))
        story.append(Paragraph(f"<b>Comentarios:</b> {body.comentarios}", s_norm))
    story.append(Spacer(1, 10))

    col_widths = [22*mm, 62*mm, 14*mm, 22*mm, 16*mm, 14*mm, 24*mm]
    tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F3864')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 7.5),
        ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#d1d5db')),
        ('ALIGN',      (2,1), (-1,-1), 'RIGHT'),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 12))

    tot_rows = [['Subtotal:', f"${subtotal_neto:,.2f}".replace(',', '.')]]
    if discIva:
        tot_rows.append(['IVA:', f"${iva_monto:,.2f}".replace(',', '.')])
    tot_rows.append(['TOTAL c/IVA:', f"${total_final:,.2f}".replace(',', '.')])
    tot_tbl = Table(tot_rows, colWidths=[40*mm, 32*mm])
    tot_tbl.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-2), 'Helvetica'),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN',    (0,0), (-1,-1), 'RIGHT'),
        ('LINEABOVE',(0,-1),(-1,-1), 0.75, colors.black),
        ('TOPPADDING', (0,-1),(-1,-1), 4),
    ]))
    tot_tbl.hAlign = 'RIGHT'
    story.append(tot_tbl)
    story.append(Spacer(1, 16))
    story.append(HRFlowable(width='100%', color=colors.HexColor('#dc2626'), thickness=1))
    story.append(Spacer(1, 4))
    story.append(Paragraph('⚠ Esta simulación no genera ni compromete ningún documento en Flexxus ERP ni en Microbell S.A. '
                            'Las bonificaciones aquí probadas no son válidas hasta convertir el documento en un presupuesto real.',
                            s_small))

    doc.build(story)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                              headers={"Content-Disposition": 'inline; filename="simulacion_presupuesto.pdf"'})

# ─── PDF Presupuesto ───────────────────────────────────────────────────────────
@app.get("/presupuestos/{numero}/pdf")
def presupuesto_pdf(numero: str, db: str = Query("oficial"),
                    descuento_promo_pct: float = Query(0.0),
                    descuento_promo_nombre: str = Query("")):
    from datetime import datetime
    db_path = DATABASE_MLT if db == 'sw' else DATABASE

    # ── 1. Datos empresa ───────────────────────────────────────────────────────
    razon_soc = 'MICROBELL S.A.'
    dir_emp   = 'PATAGONES 2675 PISO 3 - CABA - C1437JEA'
    tel_emp   = '+54 11 3988-0024'
    email_emp = 'info@microbellsa.com.ar'
    web_emp   = 'www.microbellsa.com'
    cuit_emp  = '30-70839018-2'
    iibb_emp  = 'CM 901-068199-0'
    fi_emp    = '2021-05-01'
    try:
        c_e = conn('WIN1252', db=db_path)
        cur_e = c_e.cursor()
        cur_e.execute(
            'SELECT RAZONSOCIAL, DIRECCION, TELEFONO, EMAIL, DIRECCIONWEB, '
            'CUIT, INGRESOSBRUTOS, FE_FECHAINC '
            'FROM "SUCURSALES" WHERE CODIGOSUCURSAL = ?', ('PRINCIPAL',)
        )
        emp = cur_e.fetchone()
        c_e.close()
        if emp:
            razon_soc = (emp[0] or razon_soc).strip()
            dir_emp   = (emp[1] or dir_emp).strip()
            tel_emp   = (emp[2] or tel_emp).strip()
            email_emp = (emp[3] or email_emp).strip()
            web_emp   = (emp[4] or web_emp).strip()
            cuit_emp  = (emp[5] or cuit_emp).strip()
            iibb_emp  = (emp[6] or iibb_emp).strip()
            fi_emp    = emp[7] if emp[7] else fi_emp
    except Exception:
        pass  # usa los valores hardcodeados

    # ── 2. Cabeza presupuesto ──────────────────────────────────────────────────
    try:
        c_h = conn('WIN1252', db=db_path)
        cur_h = c_h.cursor()
        cur_h.execute(
            'SELECT CODIGOCLIENTE, RAZONSOCIAL, FECHACOMPROBANTE, FECHAVENCIMIENTO, '
            'TOTAL, IVA1, IVA2, COMENTARIOS, CODIGOUSUARIO, CODIGOMULTIPLAZO, '
            'CODIGOTRANSPORTE, DIRECCION, TIPOIVA, TELEFONO, '
            'COALESCE(DESCUENTOPORCENTAJE, 0) '
            'FROM "CABEZAPRESUPUESTOS" WHERE NUMEROCOMPROBANTE = ?', (numero,)
        )
        cab = cur_h.fetchone()
        c_h.close()
    except Exception as ex:
        raise HTTPException(500, f"Cabeza: {ex}")

    if not cab:
        raise HTTPException(404, f"Presupuesto {numero} no encontrado")

    cod_cli, rs_cli, fec_comp, fec_vto, total_cab, iva1_cab, iva2_cab, \
    comentarios, cod_usu, cod_multi, cod_transp, dir_cli, tipo_iva, tel_cli, \
    dto_pct_cab = cab
    dir_cli = _domicilio_con_localidad(db_path, cod_cli, dir_cli)

    subtotal_cab  = float(total_cab or 0)  # neto después del descuento, sin IVA
    iva1_val      = float(iva1_cab or 0)
    iva2_val      = float(iva2_cab or 0)
    dto_gral_pdf  = float(dto_pct_cab or 0)
    total_final   = subtotal_cab + iva1_val + iva2_val

    # ── 3. Items ───────────────────────────────────────────────────────────────
    try:
        c_it = conn('WIN1252', db=db_path)
        cur_it = c_it.cursor()
        cur_it.execute(
            'SELECT COALESCE(NULLIF(TRIM(it.CODIGOPARTICULAR),\'\'), NULLIF(TRIM(a.CODIGOPARTICULAR),\'\'), TRIM(it.CODIGOARTICULO)), '
            'it.DESCRIPCION, it.CANTIDAD, it.PRECIOUNITARIO, '
            'it.BONIFICACION, it.PRECIOTOTAL, it.PORCENTAJEIVA '
            'FROM "CUERPOPRESUPUESTOS" it '
            'LEFT JOIN "ARTICULOS" a ON a.CODIGOARTICULO = it.CODIGOARTICULO '
            'WHERE it.NUMEROCOMPROBANTE = ? ORDER BY it.LINEA',
            (numero,)
        )
        items = cur_it.fetchall()
        c_it.close()
    except Exception as ex:
        raise HTTPException(500, f"Items: {ex}")

    # Recalcular subtotal desde ítems (TOTAL en CABEZA es IVA-inclusive para Flexxus)
    _dto_abs_pdf = abs(dto_gral_pdf)
    _dto_promo_pdf = max(0.0, min(float(descuento_promo_pct), 100.0))
    subtotal_cab = sum(float(it[5] or 0) for it in items) * (1 - _dto_abs_pdf / 100) * (1 - _dto_promo_pdf / 100)
    total_final  = subtotal_cab + iva1_val + iva2_val

    # ── 4. Datos cliente (CUIT, teléfono, vendedor) ───────────────────────────
    cuit_cli = ''
    vendedor_nombre = cod_usu or ''
    # tel_cli viene de CABEZAPRESUPUESTOS; si está vacío lo buscamos en CLIENTES
    tel_pdf = str(tel_cli or '').strip().lstrip('-').strip()
    try:
        c_cl = conn('WIN1252', db=db_path)
        cur_cl = c_cl.cursor()
        cur_cl.execute(
            'SELECT CUIT, TELEFONO, TELEFONOCELULAR FROM "CLIENTES" WHERE CODIGOCLIENTE = ?',
            (cod_cli,)
        )
        r_cl = cur_cl.fetchone()
        if r_cl:
            cuit_cli = (r_cl[0] or '').strip()
            if not tel_pdf:
                tel_pdf = (r_cl[1] or r_cl[2] or '').strip()
        cur_cl.execute(
            'SELECT RAZONSOCIAL FROM "USUARIOS" WHERE CODIGOUSUARIO = ?', (cod_usu,)
        )
        r_vend = cur_cl.fetchone()
        if r_vend:
            vendedor_nombre = (r_vend[0] or '').strip()
        c_cl.close()
    except Exception:
        pass

    # ── 5. Transporte ──────────────────────────────────────────────────────────
    transporte_desc = str(cod_transp or '').strip()
    if transporte_desc and transporte_desc != '0':
        for _tr_db in ([db_path, DATABASE] if db == 'sw' else [db_path]):
            try:
                c_tr = conn('WIN1252', db=_tr_db)
                cur_tr = c_tr.cursor()
                cur_tr.execute('SELECT DESCRIPCION FROM "TRANSPORTES" WHERE CODIGOTRANSPORTE = ?', (cod_transp,))
                r_tr = cur_tr.fetchone()
                c_tr.close()
                if r_tr and (r_tr[0] or '').strip():
                    transporte_desc = r_tr[0].strip()
                    break
            except Exception:
                pass
        else:
            transporte_desc = 'A CONVENIR'
    else:
        transporte_desc = 'A CONVENIR'

    # ── 6. Condición de pago ───────────────────────────────────────────────────
    cond_pago = str(cod_multi or '').strip()
    if cond_pago and cond_pago not in ('0', ''):
        try:
            c_mp = conn('WIN1252', db=db_path)
            cur_mp = c_mp.cursor()
            cur_mp.execute('SELECT DESCRIPCION FROM "MULTIPLAZOS" WHERE CODIGOMULTIPLAZO = ?', (cod_multi,))
            r_mp = cur_mp.fetchone()
            if r_mp:
                cond_pago = (r_mp[0] or '').strip()
            c_mp.close()
        except Exception:
            pass
    else:
        cond_pago = ''

    # ── 7. Construir PDF ───────────────────────────────────────────────────────
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=10*mm, rightMargin=10*mm,
        topMargin=4*mm, bottomMargin=12*mm
    )

    styles = getSampleStyleSheet()
    W = A4[0] - 20*mm  # ancho útil

    s_norm  = ParagraphStyle('norm',  fontSize=7,  leading=9,  fontName='Helvetica')
    s_bold  = ParagraphStyle('bold',  fontSize=7,  leading=9,  fontName='Helvetica-Bold')
    s_title = ParagraphStyle('title', fontSize=11, leading=13, fontName='Helvetica-Bold', alignment=TA_CENTER)
    s_small = ParagraphStyle('small', fontSize=6,  leading=7,  fontName='Helvetica')
    s_c     = ParagraphStyle('c',     fontSize=7,  leading=9,  fontName='Helvetica', alignment=TA_CENTER)
    s_r     = ParagraphStyle('r',     fontSize=7,  leading=9,  fontName='Helvetica', alignment=TA_RIGHT)
    s_rb    = ParagraphStyle('rb',    fontSize=7,  leading=9,  fontName='Helvetica-Bold', alignment=TA_RIGHT)

    GRIS   = colors.HexColor('#D9D9D9')
    OSCURO = colors.HexColor('#1F3864')
    AZUL   = colors.HexColor('#2E74B5')

    story = []

    # ── ENCABEZADO ─────────────────────────────────────────────────────────────
    # Logo
    logo_cell = ''
    if os.path.exists(LOGO_PATH):
        try:
            _logo_img = Image(LOGO_PATH, width=36*mm, height=36*0.32*mm)
            _logo_img.hAlign = 'CENTER'
            logo_cell = _logo_img
        except Exception:
            logo_cell = Paragraph(razon_soc, s_bold)
    else:
        logo_cell = Paragraph(razon_soc, s_bold)

    emp_info = (
        f"<b>{razon_soc}</b><br/>"
        f"{dir_emp}<br/>"
        f"Tel: {tel_emp}<br/>"
        f"{email_emp}<br/>"
        f"{web_emp}"
    )

    # Caja tipo comprobante "P" (Presupuesto)
    tipo_box = Table(
        [[Paragraph('<b>P</b>', ParagraphStyle('pb', fontSize=22, fontName='Helvetica-Bold', alignment=TA_CENTER, leading=26))]],
        colWidths=[15*mm], rowHeights=[15*mm]
    )
    tipo_box.setStyle(TableStyle([
        ('BOX',            (0,0),(0,0), 1.5, colors.black),
        ('VALIGN',         (0,0),(0,0), 'MIDDLE'),
        ('ALIGN',          (0,0),(0,0), 'CENTER'),
        ('TOPPADDING',     (0,0),(0,0), 6),
        ('BOTTOMPADDING',  (0,0),(0,0), 0),
        ('LEFTPADDING',    (0,0),(0,0), 0),
        ('RIGHTPADDING',   (0,0),(0,0), 0),
    ]))
    cod_info = Paragraph('COD.908', ParagraphStyle('cod', fontSize=8, fontName='Helvetica', alignment=TA_CENTER))

    nro_fmt = numero.zfill(8)
    doc_info = (
        f"<b>PRESUPUESTO</b><br/>"
        f"Nº 0001-{nro_fmt}<br/>"
        f"Fecha: {_d(fec_comp)}<br/>"
        f"CUIT: {cuit_emp}<br/>"
        f"IIBB: {iibb_emp}<br/>"
        f"F.Inicio: {_d(fi_emp)}<br/>"
        f"<font size='5'><b>DOCUMENTO NO VÁLIDO COMO FACTURA</b></font>"
    )

    header_data = [[
        logo_cell,
        Paragraph(emp_info, s_norm),
        Table([[tipo_box],[cod_info]], colWidths=[20*mm]),
        Paragraph(doc_info, s_norm),
    ]]
    header_table = Table(header_data, colWidths=[45*mm, 39*mm, 22*mm, None])
    header_table.setStyle(TableStyle([
        ('VALIGN',    (0,0),(-1,-1), 'TOP'),
        ('ALIGN',     (0,0),(0,0),   'CENTER'),
        ('LEFTPADDING',  (0,0),(-1,-1), 2),
        ('RIGHTPADDING', (0,0),(-1,-1), 2),
        ('TOPPADDING',   (0,0),(-1,-1), 0),
        ('BOTTOMPADDING',(0,0),(-1,-1), 0),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 3*mm))

    # ── DATOS CLIENTE ──────────────────────────────────────────────────────────
    cli_left = (
        f"<b>Cliente:</b> {cod_cli} - {(rs_cli or '').strip()}<br/>"
        f"<b>Dirección:</b> {(dir_cli or '').strip()}<br/>"
        f"<b>Tel:</b> {tel_pdf}<br/>"
        f"<b>CUIT:</b> {cuit_cli}"
    )
    cli_right = (
        f"<b>Cond. IVA:</b> {(tipo_iva or '').strip()}<br/>"
        f"<b>Transporte:</b> {transporte_desc}<br/>"
        f"<b>Asistente:</b> {cod_usu or ''}<br/>"
        f"<b>Vendedor:</b> {vendedor_nombre}"
    )
    cli_table = Table(
        [[Paragraph(cli_left, s_norm), Paragraph(cli_right, s_norm)]],
        colWidths=[W*0.55, W*0.45]
    )
    cli_table.setStyle(TableStyle([
        ('BACKGROUND',   (0,0),(-1,-1), GRIS),
        ('BOX',          (0,0),(-1,-1), 0.5, colors.grey),
        ('LEFTPADDING',  (0,0),(-1,-1), 4),
        ('RIGHTPADDING', (0,0),(-1,-1), 4),
        ('TOPPADDING',   (0,0),(-1,-1), 3),
        ('BOTTOMPADDING',(0,0),(-1,-1), 3),
        ('VALIGN',       (0,0),(-1,-1), 'TOP'),
    ]))
    story.append(cli_table)
    story.append(Spacer(1, 2*mm))

    # ── TABLA ITEMS ────────────────────────────────────────────────────────────
    col_w = [22*mm, None, 18*mm, 28*mm, 26*mm, 28*mm]
    # cabecera
    hdr = ['Código', 'Descripción', 'Cantidad', 'Precio Unit.', 'Bonif.%', 'Subtotal']
    items_data = [[Paragraph(f'<b>{h}</b>', s_c) for h in hdr]]

    for it in items:
        cod_art, desc_art, cant, pu, bonif, ptotal, piva = it
        # Limpiar código: quitar espacios y el .0 si viene como float
        cod_str = str(cod_art or '').strip()
        if cod_str.endswith('.0'):
            cod_str = cod_str[:-2]
        try:
            cant_str = str(int(float(cant))) if float(cant) == int(float(cant)) else str(cant)
        except Exception:
            cant_str = str(cant or '')
        items_data.append([
            Paragraph(cod_str, s_c),
            Paragraph(str(desc_art or ''), s_norm),
            Paragraph(cant_str, s_c),
            Paragraph(_fmt(pu), s_r),
            Paragraph(f"{float(bonif or 0):.2f}%", s_c),
            Paragraph(_fmt(ptotal), s_r),
        ])

    # fila totales
    _dto_p_pr = max(0.0, min(float(descuento_promo_pct or 0), 100.0))
    s_dto = ParagraphStyle('sdto', fontSize=7, leading=9, fontName='Helvetica-Bold', alignment=TA_RIGHT, textColor=colors.HexColor('#b45309'))
    if _dto_p_pr > 0:
        # Descuento por promo combo
        _sub_bruto_pr = subtotal_cab / (1 - _dto_p_pr / 100) if _dto_p_pr < 100 else subtotal_cab
        _dto_m_pr = _sub_bruto_pr - subtotal_cab
        _lbl_pr = (descuento_promo_nombre.strip() or f'Desc. promo combo ({_dto_p_pr:g}%)')
        items_data.append(['', '', '', '', Paragraph('<b>SUBTOTAL S/DESC.</b>', s_rb), Paragraph(_fmt(_sub_bruto_pr), s_rb)])
        items_data.append(['', '', '', '', Paragraph(f'<b>{_lbl_pr}</b>', s_dto), Paragraph(f'<b>- {_fmt(_dto_m_pr)}</b>', s_dto)])
        items_data.append(['', '', '', '', Paragraph('<b>SUBTOTAL</b>', s_rb), Paragraph(_fmt(subtotal_cab), s_rb)])
    elif dto_gral_pdf != 0:
        # Descuento financiero por monto (dto_gral_pdf es negativo en DB → abs para cálculos)
        _dto_abs = abs(dto_gral_pdf)
        _sub_bruto_gral = subtotal_cab / (1 - _dto_abs / 100) if _dto_abs < 100 else subtotal_cab
        _dto_m_gral = _sub_bruto_gral - subtotal_cab
        _lbl_gral = f'Descuento {_dto_abs:g}%'
        items_data.append(['', '', '', '', Paragraph('<b>SUBTOTAL</b>', s_rb), Paragraph(_fmt(_sub_bruto_gral), s_rb)])
        items_data.append(['', '', '', '', Paragraph(f'<b>{_lbl_gral}</b>', s_dto), Paragraph(f'<b>- {_fmt(_dto_m_gral)}</b>', s_dto)])
        items_data.append(['', '', '', '', Paragraph('<b>SUBTOTAL C/DESC.</b>', s_rb), Paragraph(_fmt(subtotal_cab), s_rb)])
    else:
        items_data.append(['', '', '', '', Paragraph('<b>SUBTOTAL</b>', s_rb), Paragraph(_fmt(subtotal_cab), s_rb)])
    if iva1_val:
        items_data.append(['', '', '', '', Paragraph('<b>IVA 21%</b>', s_rb), Paragraph(_fmt(iva1_val), s_rb)])
    if iva2_val:
        items_data.append(['', '', '', '', Paragraph('<b>IVA 10.5%</b>', s_rb), Paragraph(_fmt(iva2_val), s_rb)])
    items_data.append(['', '', '', '', Paragraph('<b>TOTAL</b>', s_rb), Paragraph(_fmt(total_final), s_rb)])

    items_table = Table(items_data, colWidths=col_w, repeatRows=1)
    n_items = len(items) + 1  # +1 header
    ts = TableStyle([
        ('BACKGROUND',    (0,0),(-1,0),          AZUL),
        ('TEXTCOLOR',     (0,0),(-1,0),          colors.white),
        ('FONTNAME',      (0,0),(-1,0),          'Helvetica-Bold'),
        ('FONTSIZE',      (0,0),(-1,-1),         7),
        ('ROWBACKGROUNDS',(0,1),(-1,n_items-1),  [colors.white, colors.HexColor('#EEF3F8')]),
        ('GRID',          (0,0),(-1,n_items-1),  0.3, colors.grey),
        ('LINEABOVE',     (0,n_items),(-1,n_items), 0.8, colors.black),
        ('TOPPADDING',    (0,0),(-1,-1),         2),
        ('BOTTOMPADDING', (0,0),(-1,-1),         2),
        ('LEFTPADDING',   (0,0),(-1,-1),         3),
        ('RIGHTPADDING',  (0,0),(-1,-1),         3),
        ('VALIGN',        (0,0),(-1,-1),         'MIDDLE'),
    ])
    # línea separadora antes del TOTAL
    ts.add('LINEABOVE', (4,-1), (-1,-1), 1.2, colors.black)
    items_table.setStyle(ts)
    story.append(items_table)
    story.append(Spacer(1, 3*mm))

    # ── CONDICIONES DE VENTA ───────────────────────────────────────────────────
    cond_data = [
        [Paragraph('<b>Condiciones de Venta</b>', s_bold), ''],
        [Paragraph(f'<b>Cond. Pago:</b> {cond_pago}', s_norm),
         Paragraph(f'<b>Tipo Cambio:</b> $ 1,00', s_norm)],
        [Paragraph(f'<b>Domicilio Entrega:</b> {(dir_cli or "").strip()}', s_norm),
         Paragraph(f'<b>Fecha Venc.:</b> {_d(fec_vto)}', s_norm)],
        [Paragraph(f'<b>Observaciones:</b> {(comentarios or "").strip()}', s_norm), ''],
    ]
    cond_table = Table(cond_data, colWidths=[W*0.6, W*0.4])
    cond_table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),(-1,0), GRIS),
        ('SPAN',          (0,0),(-1,0)),
        ('SPAN',          (0,3),(-1,3)),
        ('BOX',           (0,0),(-1,-1), 0.5, colors.grey),
        ('INNERGRID',     (0,0),(-1,-1), 0.2, colors.lightgrey),
        ('LEFTPADDING',   (0,0),(-1,-1), 4),
        ('RIGHTPADDING',  (0,0),(-1,-1), 4),
        ('TOPPADDING',    (0,0),(-1,-1), 2),
        ('BOTTOMPADDING', (0,0),(-1,-1), 2),
        ('FONTSIZE',      (0,0),(-1,-1), 7),
    ]))
    story.append(cond_table)
    story.append(Spacer(1, 3*mm))

    # ── CUENTAS BANCARIAS (en flujo, centrada, angosta) ───────────────────────
    s_bh  = ParagraphStyle('bh',  fontSize=6, fontName='Helvetica-Bold', textColor=colors.white, alignment=TA_CENTER, leading=8)
    s_bno = ParagraphStyle('bno', fontSize=6, fontName='Helvetica', leading=8)
    s_bc  = ParagraphStyle('bc',  fontSize=6, fontName='Helvetica', leading=8, alignment=TA_CENTER)
    s_bct = ParagraphStyle('bct', fontSize=7, fontName='Helvetica-Bold', textColor=colors.white, alignment=TA_CENTER, leading=9)

    # Fila título "CUENTAS BANCARIAS" que abarca todas las columnas
    # Fila cabecera columnas
    bank_data = [
        [Paragraph('<b>CUENTAS BANCARIAS</b>', s_bct), '', '', ''],
        [Paragraph('<b>BANCO</b>',   s_bh),
         Paragraph('<b>CBU</b>',     s_bh),
         Paragraph('<b>SUCURSAL</b>',s_bh),
         Paragraph('<b>CTA. CTE.</b>',s_bh)],
    ]
    for nombre, cbu, suc, cta in _BANCOS:
        bank_data.append([
            Paragraph(nombre, s_bno), Paragraph(cbu, s_bno),
            Paragraph(suc, s_bc),     Paragraph(cta, s_bno),
        ])
    bank_data.append([
        Paragraph('<b>Mercado Pago</b>', ParagraphStyle('bmp', fontSize=6, fontName='Helvetica-Bold', leading=8)),
        Paragraph(f'{_MP_EMAIL}', s_bno),
        Paragraph('CVU', s_bc),
        Paragraph(_MP_CVU, s_bno),
    ])

    # Ancho total ~118mm, columnas: banco 26 + CBU 54 + suc 15 + cta 23
    BW = [26*mm, 54*mm, 15*mm, 23*mm]
    bank_table = Table(bank_data, colWidths=BW)
    bank_table.setStyle(TableStyle([
        # Título
        ('BACKGROUND',    (0,0),(-1,0),  OSCURO),
        ('SPAN',          (0,0),(-1,0)),
        ('ALIGN',         (0,0),(-1,0),  'CENTER'),
        ('TOPPADDING',    (0,0),(-1,0),  3),
        ('BOTTOMPADDING', (0,0),(-1,0),  3),
        # Cabecera columnas
        ('BACKGROUND',    (0,1),(-1,1),  OSCURO),
        # Filas datos
        ('ROWBACKGROUNDS',(0,2),(-1,-1), [colors.white, colors.HexColor('#EEF3F8')]),
        ('BOX',           (0,0),(-1,-1), 0.5, colors.grey),
        ('INNERGRID',     (0,1),(-1,-1), 0.2, colors.lightgrey),
        ('LEFTPADDING',   (0,0),(-1,-1), 3),
        ('RIGHTPADDING',  (0,0),(-1,-1), 3),
        ('TOPPADDING',    (0,1),(-1,-1), 2),
        ('BOTTOMPADDING', (0,1),(-1,-1), 2),
        ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
    ]))
    bank_table.hAlign = 'CENTER'

    story.append(Spacer(1, 4*mm))
    story.append(bank_table)

    doc.build(story)
    buf.seek(0)
    fname = f"Presupuesto_{nro_fmt}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{fname}"'}
    )

# ─── Detalle de comprobante ────────────────────────────────────────────────────
@app.get("/comprobantes/{tipo}/{numero}/detalle")
def detalle_comprobante(tipo: str, numero: str):
    sql = """
        SELECT CODIGOARTICULO, DESCRIPCION, CANTIDAD,
               PRECIOUNITARIO, PORCENTAJEIVA, DESCUENTO, PRECIOTOTAL
        FROM "CUERPOCOMPROBANTES"
        WHERE TIPOCOMPROBANTE = ? AND NUMEROCOMPROBANTE = ?
        ORDER BY LINEA
    """
    DB_PROD     = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    for db_path in [DATABASE, DATABASE_MLT, DB_PROD, DB_MLT_PROD]:
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            cur.execute(sql, (tipo, numero))
            rows = cur.fetchall()
            c.close()
            if rows:
                return [{
                    "codigo":      r[0], "descripcion": r[1],
                    "cantidad":    float(r[2]) if r[2] else 0,
                    "precio_unit": float(r[3]) if r[3] else 0,
                    "iva":         float(r[4]) if r[4] else 0,
                    "descuento":   float(r[5]) if r[5] else 0,
                    "total":       float(r[6]) if r[6] else 0,
                } for r in rows]
        except Exception:
            continue
    return []

# ─── PDF Comprobante (FA/FB/FCA/FE/NCA/NCB/NDA/NDB y Proforma SW) ────────────

_TIPO_NOMBRE = {
    'FA':'FACTURA A','FB':'FACTURA B','FCA':'FACTURA DE CREDITO A',
    'FCB':'FACTURA DE CREDITO B','FE':'FACTURA E',
    'NCA':'NOTA DE CREDITO A','NCB':'NOTA DE CREDITO B',
    'NCCA':'NOTA DE CREDITO DE CREDITO A','NCE':'NOTA DE CREDITO E',
    'NDA':'NOTA DE DEBITO A','NDB':'NOTA DE DEBITO B','NDE':'NOTA DE DEBITO E',
}
_TIPO_LETRA = {
    'FA':'A','FCA':'A','NCA':'A','NDA':'A','NCCA':'A',
    'FB':'B','FCB':'B','NCB':'B','NDB':'B',
    'FE':'E','NCE':'E','NDE':'E',
}
_BANCOS = [
    ('Santander Rio','0720131420000001149872','131','11498/7'),
    ('Prov. Bs. As.','0140004501400404115211','4004','041152/1'),
    ('HSBC','1500607500060732055732','607','607-3-205573'),
    ('Galicia','0070154520000005006724','154','5006-7-154/2'),
]
_CVU = 'E-mail: marketing@microbellsa.com.ar\nCVU: 0000003100004756934965'

def _fmt_num(numero: str):
    """4400014918.0 → ('0044','00014918','0044-00014918')"""
    try:
        n = int(float(numero))
        pv  = n // 100_000_000
        seq = n %  100_000_000
        return f'{pv:04d}', f'{seq:08d}', f'{pv:04d}-{seq:08d}'
    except Exception:
        return '0001', numero, numero

def _pesos(v): return f'$ {float(v or 0):,.2f}'.replace(',','X').replace('.',',').replace('X','.')

@app.get("/comprobantes/{tipo}/{numero}/pdf")
def comprobante_pdf_route(tipo: str, numero: str):
    from reportlab.platypus import KeepTogether
    from reportlab.platypus.flowables import Flowable

    class BottomSpacer(Flowable):
        """Empuja el contenido siguiente al pie de página."""
        def __init__(self, footer_h):
            Flowable.__init__(self)
            self._fh = footer_h
        def wrap(self, availWidth, availHeight):
            return availWidth, max(0, availHeight - self._fh)
        def draw(self):
            pass
    from reportlab.lib.utils import ImageReader
    DB_PROD     = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'

    # ── 1. Buscar cabeza ───────────────────────────────────────────────────────
    cab = None; items = []; found_db = None; is_mlt = False
    for db_path, mlt in [(DATABASE,False),(DB_PROD,False),(DATABASE_MLT,True),(DB_MLT_PROD,True)]:
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            cur.execute('SELECT * FROM "CABEZACOMPROBANTES" WHERE TIPOCOMPROBANTE=? AND NUMEROCOMPROBANTE=?',(tipo,numero))
            row = cur.fetchone()
            if row:
                cols = [d[0] for d in cur.description]
                cab = {k: (v.strip() if isinstance(v,str) else v) for k,v in zip(cols,row)}
                cur.execute(
                    'SELECT LINEA,CODIGOARTICULO,DESCRIPCION,CANTIDAD,'
                    'PRECIOUNITARIO,DESCUENTO,PORCENTAJEIVA,PRECIOTOTAL '
                    'FROM "CUERPOCOMPROBANTES" '
                    'WHERE TIPOCOMPROBANTE=? AND NUMEROCOMPROBANTE=? ORDER BY LINEA',
                    (tipo, numero)
                )
                items = cur.fetchall()
                found_db = db_path; is_mlt = mlt
                c.close(); break
            c.close()
        except Exception:
            pass
    if not cab:
        raise HTTPException(404, f"{tipo} {numero} no encontrado")

    # ── 2. CAE (L1 y SW — ambos sistemas emiten facturas electrónicas con CAE) ──
    cae = None; vto_cae = None
    _dbs_cae = [found_db, DATABASE, DB_PROD] if not is_mlt else [found_db, DATABASE_MLT, DB_MLT_PROD]
    for db_path in _dbs_cae:
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            cur.execute('SELECT CAE,VENCIMIENTOCAE FROM "CAEAFIP" WHERE TIPOCOMPROBANTE=? AND NUMEROCOMPROBANTE=?',(tipo,numero))
            r = cur.fetchone(); c.close()
            if r and r[0]:
                cae = str(r[0]).strip()
                vto_cae = r[1]; break
        except Exception:
            pass

    # ── 2b. Despachos (solo Línea 1) ──────────────────────────────────────────
    # Busca despachos por los códigos de artículo que componen la factura
    # usando STOCKXDESPACHO (CODIGOARTICULO → DESPACHO → ADUANA)
    despachos_lista = []  # lista de "nro - aduana" únicos
    if not is_mlt and items:
        codigos_articulos = list({str(it[1] or '').strip() for it in items if it[1]})
        if codigos_articulos:
            for db_path in [found_db, DATABASE, DB_PROD]:
                try:
                    c_d = conn('WIN1252', db=db_path)
                    cur_d = c_d.cursor()
                    placeholders = ','.join(['?' for _ in codigos_articulos])
                    # MAX(DESPACHO) por artículo = despacho más reciente de importación
                    cur_d.execute(
                        f'SELECT MAX(s.DESPACHO), MAX(s.ADUANA) '
                        f'FROM "STOCKXDESPACHO" s '
                        f'WHERE s.CODIGOARTICULO IN ({placeholders}) '
                        f"AND TRIM(s.DESPACHO) <> '' "
                        f'GROUP BY s.CODIGOARTICULO',
                        codigos_articulos
                    )
                    for r in cur_d.fetchall():
                        nrd    = str(r[0] or '').strip()
                        aduana = str(r[1] or '').strip()
                        # descartar despachos vacíos o con solo ceros
                        if not nrd or not nrd.replace('0','').strip():
                            continue
                        if not aduana or aduana == '-':
                            aduana = 'Aduana de Buenos Aires'
                        entry = f'{nrd} - {aduana}'
                        if entry not in despachos_lista:
                            despachos_lista.append(entry)
                    despachos_lista.sort()
                    c_d.close()
                    if despachos_lista:
                        break
                except Exception:
                    pass

    # ── 3. Datos cliente extra ─────────────────────────────────────────────────
    cli_ingbrutos = ''; cli_localidad = ''; cli_provincia = ''; cli_cp = ''
    try:
        c_cl = conn('WIN1252', db=DATABASE)
        cur_cl = c_cl.cursor()
        cur_cl.execute(
            'SELECT INGRESOSBRUTOS, LOCALIDAD, PROVINCIA, CP '
            'FROM "CLIENTES" WHERE CODIGOCLIENTE=?', (cab.get('CODIGOCLIENTE',''),)
        )
        r_cl = cur_cl.fetchone(); c_cl.close()
        if r_cl:
            cli_ingbrutos = str(r_cl[0] or '').strip()
            cli_localidad = str(r_cl[1] or '').strip()
            cli_provincia = str(r_cl[2] or '').strip()
            cli_cp        = str(r_cl[3] or '').strip()
    except Exception:
        pass

    # ── 4. Vendedor / asistente ────────────────────────────────────────────────
    cod_vend = str(cab.get('CODIGOUSUARIO','') or '').strip()
    cod_asist= str(cab.get('CODIGOUSUARIO2','') or '').strip()

    # ── 5. Condición de venta ──────────────────────────────────────────────────
    cond_venta = ''
    try:
        c_mp = conn('WIN1252', db=DATABASE)
        cur_mp = c_mp.cursor()
        cur_mp.execute('SELECT DESCRIPCION FROM "MULTIPLAZOS" WHERE CODIGOMULTIPLAZO=?',(cab.get('CODIGOMULTIPLAZO',''),))
        r_mp = cur_mp.fetchone(); c_mp.close()
        if r_mp: cond_venta = str(r_mp[0] or '').strip()
    except Exception:
        pass

    pv_str, seq_str, num_display = _fmt_num(numero)

    # ── 6. Generar PDF ─────────────────────────────────────────────────────────
    buf = BytesIO()
    W, H = A4
    m = 14*mm
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=m, rightMargin=m,
                            topMargin=10*mm, bottomMargin=10*mm)

    AZUL    = colors.HexColor('#1a3a5c')
    CELESTE = colors.HexColor('#4A90D9')   # azul claro para tablas bancarias
    GRIS    = colors.HexColor('#f4f6f9')
    NEGRO   = colors.black
    ROJO    = colors.HexColor('#cc0000')

    base   = getSampleStyleSheet()
    sN     = ParagraphStyle('sN',  fontSize=7,  leading=9,  fontName='Helvetica')
    sNb    = ParagraphStyle('sNb', fontSize=7,  leading=9,  fontName='Helvetica-Bold')
    sNc    = ParagraphStyle('sNc', fontSize=7,  leading=9,  fontName='Helvetica', alignment=TA_CENTER)
    sNr    = ParagraphStyle('sNr', fontSize=7,  leading=9,  fontName='Helvetica', alignment=TA_RIGHT)
    sT     = ParagraphStyle('sT',  fontSize=9,  leading=11, fontName='Helvetica-Bold')
    sTc    = ParagraphStyle('sTc', fontSize=9,  leading=11, fontName='Helvetica-Bold', alignment=TA_CENTER)
    sCAE   = ParagraphStyle('sCAE',fontSize=8,  leading=10, fontName='Helvetica-Bold')
    sCAEv  = ParagraphStyle('sCAEv',fontSize=8, leading=10, fontName='Helvetica')

    cw = W - 2*m  # ancho útil

    story = []

    # ── Helper QR compartido (L1 y SW) ───────────────────────────────────────
    def _make_qr_img(url, w=32*mm, h=32*mm):
        """Genera imagen QR para ReportLab. Auto-instala qrcode si falta."""
        try:
            import qrcode as _qrc
        except ImportError:
            import subprocess as _sp, sys as _sys
            _sp.check_call(
                [_sys.executable, '-m', 'pip', 'install', 'qrcode[pil]', '--break-system-packages', '-q'],
                stdout=_sp.DEVNULL, stderr=_sp.DEVNULL)
            import qrcode as _qrc
        _qr = _qrc.QRCode(error_correction=_qrc.constants.ERROR_CORRECT_L, box_size=6, border=2)
        _qr.add_data(url); _qr.make(fit=True)
        # Intento 1: PIL/Pillow
        try:
            _img = _qr.make_image(fill_color='black', back_color='white')
            _buf = BytesIO(); _img.save(_buf, format='PNG'); _buf.seek(0)
            return Image(_buf, width=w, height=h)
        except Exception:
            pass
        # Intento 2: PyPNGImage (sin Pillow)
        try:
            from qrcode.image.pure import PyPNGImage
            _img = _qr.make_image(image_factory=PyPNGImage)
            _buf = BytesIO(); _img.save(_buf); _buf.seek(0)
            return Image(_buf, width=w, height=h)
        except Exception:
            pass
        # Intento 3: PNG puro via struct+zlib (sin dependencias externas)
        import struct as _st, zlib as _zl
        _mat = _qr.modules; _sz = len(_mat); _sc = 8
        _rows = []
        for _row in _mat:
            _rb = b'\x00'
            for _c in _row:
                _rb += (b'\x00\x00\x00' if _c else b'\xff\xff\xff') * _sc
            _rows.extend([_rb] * _sc)
        _raw = b''.join(_rows); _iw = _ih = _sz * _sc
        def _ck(_t, _d):
            _cd = _t + _d
            return _st.pack('>I', len(_d)) + _cd + _st.pack('>I', _zl.crc32(_cd) & 0xffffffff)
        _png = (b'\x89PNG\r\n\x1a\n'
                + _ck(b'IHDR', _st.pack('>IIBBBBB', _iw, _ih, 8, 2, 0, 0, 0))
                + _ck(b'IDAT', _zl.compress(_raw))
                + _ck(b'IEND', b''))
        return Image(BytesIO(_png), width=w, height=h)

    if is_mlt:
        # ══════════════════════════════════════════════════════════════════════
        # PROFORMA REMITO (SW / DATABASE_MLT)
        # ══════════════════════════════════════════════════════════════════════
        fecha_str = ''
        if cab.get('FECHACOMPROBANTE'):
            try: fecha_str = cab['FECHACOMPROBANTE'].strftime('%d/%m/%Y')
            except Exception: fecha_str = str(cab['FECHACOMPROBANTE'])[:10]

        # Cabeza: logo | badge A | "Proforma Remito / RN-N° / FECHA"
        logo_img = Image(LOGO_PATH, width=40*mm, height=14*mm) if os.path.exists(LOGO_PATH) else Paragraph('microbell S.A.', sT)

        badge_tbl = Table([[Paragraph('<b>A</b>',ParagraphStyle('ba',fontSize=28,fontName='Helvetica-Bold',alignment=TA_CENTER))]],
                          colWidths=[18*mm], rowHeights=[20*mm])
        badge_tbl.setStyle(TableStyle([('BOX',(0,0),(-1,-1),1.5,NEGRO),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(0,0),(-1,-1),'CENTER')]))

        sRsw  = ParagraphStyle('sRsw',  fontSize=7, leading=9, fontName='Helvetica',      alignment=TA_RIGHT)
        sRbsw = ParagraphStyle('sRbsw', fontSize=7, leading=9, fontName='Helvetica-Bold', alignment=TA_RIGHT)
        right_hdr = [
            Paragraph('<b>Proforma Remito</b>',             ParagraphStyle('ph', fontSize=11,fontName='Helvetica-Bold',alignment=TA_RIGHT)),
            Paragraph('<b>DOCUMENTO NO VALIDO COMO FACTURA</b>', ParagraphStyle('phd',fontSize=7,fontName='Helvetica-Bold',textColor=ROJO,alignment=TA_RIGHT)),
            Paragraph(f'RN - N°: {num_display}',            ParagraphStyle('phn',fontSize=10,fontName='Helvetica-Bold',alignment=TA_RIGHT)),
            Paragraph(f'FECHA: {fecha_str}',                sRbsw),
            Paragraph(f'CUIT: 30-70839018-2',               sRsw),
            Paragraph(f'INGRESOS BRUTOS: CM 901-068199-0',  sRsw),
            Paragraph(f'Inicio de Actividades: 05/09/2005', sRsw),
        ]

        emp_left_sw = [
            logo_img,
            Paragraph('Dirección: PATAGONES 2675 PISO 3', sN),
            Paragraph('C.A.B.A. CAPITAL FEDERAL C1437JEA', sN),
            Paragraph('Télefono: +54 11 3988-0024', sN),
            Paragraph('Email: info@microbellsa.com.ar  www.microbellsa.com', sN),
        ]

        hdr_tbl = Table([[emp_left_sw, badge_tbl, right_hdr]],
                        colWidths=[cw*0.35, 22*mm, cw*0.50])
        hdr_tbl.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP'),('LEFTPADDING',(0,0),(0,-1),4),('RIGHTPADDING',(0,0),(-1,-1),0)]))
        story.append(hdr_tbl)
        story.append(HRFlowable(width=cw, thickness=1.5, color=AZUL, spaceAfter=4))

        # Bloque cliente
        cod_cli  = str(cab.get('CODIGOCLIENTE','') or '').strip()
        rs_cli   = str(cab.get('RAZONSOCIAL','') or '').strip()
        dir_cli  = str(cab.get('DIRECCION','') or '').strip()
        cuit_cli = str(cab.get('CUIT','') or '').strip()
        tel_cli  = str(cab.get('TELEFONO','') or '').strip()

        cli_rows = [
            [Paragraph(f'<b>Cliente:</b> {cod_cli}  {rs_cli}', sNb), Paragraph(f'<b>CUIT:</b> {cuit_cli}', sN)],
            [Paragraph(f'<b>Dirección:</b> {dir_cli}', sN), Paragraph(f'<b>Cod.Vend.:</b> {cod_vend}', sN)],
            [Paragraph(f'<b>Localidad:</b> {cli_localidad}   <b>Provincia:</b> {cli_provincia}', sN), Paragraph(f'<b>Cod.Asist.:</b> {cod_asist}', sN)],
            [Paragraph(f'<b>Cond. de Venta:</b> {cond_venta}', sN), Spacer(1,1)],
        ]
        cli_tbl = Table(cli_rows, colWidths=[cw*0.62, cw*0.38])
        cli_tbl.setStyle(TableStyle([
            ('BOX',(0,0),(-1,-1),0.5,colors.grey),
            ('INNERGRID',(0,0),(-1,-1),0.3,colors.lightgrey),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
            ('LEFTPADDING',(0,0),(-1,-1),6),
        ]))
        story.append(cli_tbl)
        story.append(Spacer(1,4))

        # Items SW
        hdrs_sw = ['CANT','CODIGO','ARTÍCULO','DTO','PRECIO UNITARIO','TOTAL']
        cws_sw  = [14*mm, 18*mm, cw-14*mm-18*mm-18*mm-32*mm-28*mm, 18*mm, 32*mm, 28*mm]
        it_rows = [[Paragraph(h, sNb) for h in hdrs_sw]]
        total_sw = 0.0
        for it in items:
            cant  = float(it[3] or 0)
            pu    = float(it[4] or 0)
            dto   = float(it[5] or 0)
            total_item = float(it[7] or 0)
            total_sw += total_item
            obs_it = ''
            it_rows.append([
                Paragraph(str(int(cant)) if cant == int(cant) else str(cant), sNc),
                Paragraph(str(it[1] or ''), sN),
                Paragraph(str(it[2] or ''), sN),
                Paragraph(f'{dto:.2f} %', sNr),
                Paragraph(_pesos(pu), sNr),
                Paragraph(_pesos(total_item), sNr),
            ])
        it_tbl = Table(it_rows, colWidths=cws_sw, repeatRows=1)
        it_tbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),CELESTE),('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,GRIS]),
            ('BOX',(0,0),(-1,-1),0.5,colors.grey),
            ('INNERGRID',(0,0),(-1,-1),0.3,colors.lightgrey),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
            ('LEFTPADDING',(0,0),(-1,-1),4),
        ]))
        story.append(it_tbl)
        story.append(Spacer(1,6))

        # Observaciones + cuentas bancarias + total
        obs_sw = str(cab.get('COMENTARIOS','') or '').strip()
        if obs_sw:
            story.append(Paragraph(f'<b>Observaciones:</b> {obs_sw}', sN))
            story.append(Spacer(1,4))

        story.append(BottomSpacer(105*mm))  # empuja footer al pie (banco + QR ~105mm)

        # Tabla bancaria compacta (125mm) centrada
        bw = [26*mm, 46*mm, 18*mm, 26*mm]   # total ~116mm
        bk_data_sw = [
            [Paragraph('<b>CUENTAS BANCARIAS</b>', ParagraphStyle('bct',fontSize=7,fontName='Helvetica-Bold',textColor=colors.white,alignment=TA_CENTER,leading=9)),'','',''],
            [Paragraph('<b>BANCO</b>',sNb),Paragraph('<b>CBU</b>',sNb),Paragraph('<b>SUCURSAL</b>',sNb),Paragraph('<b>CTA. CTE.</b>',sNb)],
        ]
        for b,cbu,suc,cta in _BANCOS:
            bk_data_sw.append([Paragraph(b,sN),Paragraph(cbu,sN),Paragraph(suc,sNc),Paragraph(cta,sN)])
        bk_data_sw.append([Paragraph(_CVU,sN),'','',''])
        bk_tbl_sw = Table(bk_data_sw, colWidths=bw)
        bk_tbl_sw.setStyle(TableStyle([
            ('SPAN',(0,0),(-1,0)),
            ('BACKGROUND',(0,0),(-1,0),CELESTE),('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('BACKGROUND',(0,1),(-1,1),colors.HexColor('#d0e4f7')),
            ('ROWBACKGROUNDS',(0,2),(-1,-2),[colors.white,GRIS]),
            ('SPAN',(0,-1),(-1,-1)),
            ('BOX',(0,0),(-1,-1),0.5,colors.grey),('INNERGRID',(0,0),(-1,-1),0.3,colors.lightgrey),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2),
            ('LEFTPADDING',(0,0),(-1,-1),4),
        ]))

        bk_total_sw = sum(bw)
        pad_sw = (cw - bk_total_sw) / 2   # centrado

        cotiz = float(cab.get('COTIZACION',1) or 1)
        tot_str = _pesos(total_sw)
        sTotal = ParagraphStyle('sTot', fontSize=12, fontName='Helvetica-Bold', alignment=TA_RIGHT)

        footer_sw = Table([
            [Spacer(1,1), bk_tbl_sw, Spacer(1,1)],
            ['', Paragraph(f'Tipo de Cambio: $ {cotiz:.0f}', sN), ''],
            ['', Paragraph(f'<b>TOTAL:  {tot_str}</b>', sTotal), ''],
        ], colWidths=[pad_sw, bk_total_sw, pad_sw])
        footer_sw.setStyle(TableStyle([
            ('VALIGN',(0,0),(-1,-1),'TOP'),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ]))
        story.append(KeepTogether([footer_sw]))

        # QR ARCA/CAE + QR sitio web (igual que Línea 1)
        story.append(HRFlowable(width=cw, thickness=0.5, color=colors.grey, spaceAfter=6))
        if cae:
            vto_cae_str_sw = ''
            if vto_cae:
                try: vto_cae_str_sw = vto_cae.strftime('%d/%m/%Y')
                except Exception: vto_cae_str_sw = str(vto_cae)[:10]
            import json as _json_sw, base64 as _b64_sw, urllib.parse as _uparse_sw
            _TIPO_AFIP_MAP_SW = {
                'FA':1,'NDA':2,'NCA':3,'FB':6,'NDB':7,'NCB':8,
                'FCA':201,'FCB':206,'FE':19,'NCE':13,'NDE':12,
            }
            pv_sw, seq_sw, _ = _fmt_num(numero)
            try: pv_int_sw = int(pv_sw)
            except Exception: pv_int_sw = 1
            try: seq_int_sw = int(seq_sw)
            except Exception: seq_int_sw = 0
            fecha_afip_sw = ''
            if cab.get('FECHACOMPROBANTE'):
                try: fecha_afip_sw = cab['FECHACOMPROBANTE'].strftime('%Y-%m-%d')
                except Exception: fecha_afip_sw = str(cab['FECHACOMPROBANTE'])[:10]
            gran_total_sw_cae = float(cab.get('TOTAL',0) or 0) + float(cab.get('IVA1',0) or 0) + float(cab.get('IVA2',0) or 0)
            cuit_sw_raw = str(cab.get('CUIT','0') or '0')
            cuit_sw_num = int(''.join(c for c in cuit_sw_raw if c.isdigit()) or '0')
            qr_data_sw = {
                "ver":1, "fecha":fecha_afip_sw,
                "cuit":30708390182, "ptoVta":pv_int_sw,
                "tipoCmp":_TIPO_AFIP_MAP_SW.get(tipo,6), "nroCmp":seq_int_sw,
                "importe":round(gran_total_sw_cae,2),
                "moneda":"PES", "ctz":1,
                "tipoDocRec":80, "nroDocRec":cuit_sw_num,
                "tipoCodAut":"E", "codAut":int(cae) if cae.isdigit() else 0
            }
            _jb_sw = _json_sw.dumps(qr_data_sw, separators=(',',':')).encode()
            qr_b64_sw = _b64_sw.b64encode(_jb_sw).decode().rstrip('=')
            qr_afip_url_sw = 'https://servicioscf.afip.gob.ar/publico/comprobantes/cae.aspx?p=' + _uparse_sw.quote(qr_b64_sw, safe='')
            qr_web_url_sw  = 'https://www.microbellsa.com'
            # SW: solo QR sitio web Microbell (derecha) — sin QR AFIP
            qr_ir = None
            try:
                qr_ir = _make_qr_img(qr_web_url_sw)
            except Exception as _qe_sw:
                qr_ir = Paragraph(f'[{type(_qe_sw).__name__}]', sN)
            sCAEc_sw = ParagraphStyle('sCAEc_sw', fontSize=8, leading=11, fontName='Helvetica')
            sCAEb_sw = ParagraphStyle('sCAEb_sw', fontSize=8, leading=11, fontName='Helvetica-Bold')
            cae_center_sw = [
                Paragraph('Factura Electrónica / CAE:', sCAEc_sw),
                Paragraph(f'<b>{cae}</b>', sCAEb_sw),
                Spacer(1,4),
                Paragraph('Fecha Vencimiento CAE:', sCAEc_sw),
                Paragraph(f'<b>{vto_cae_str_sw}</b>', sCAEb_sw),
            ]
            qr_w_sw = 34*mm
            mid_w_sw = cw - qr_w_sw  # solo columna CAE + QR derecha
            cae_tbl_sw = Table([[cae_center_sw, qr_ir]], colWidths=[mid_w_sw, qr_w_sw])
            cae_tbl_sw.setStyle(TableStyle([
                ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                ('ALIGN',(0,0),(0,-1),'LEFT'),
                ('ALIGN',(1,0),(1,-1),'RIGHT'),
                ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2),
            ]))
            story.append(cae_tbl_sw)
        else:
            story.append(Paragraph('Factura Electrónica / CAE: (pendiente)', sN))

    else:
        # ══════════════════════════════════════════════════════════════════════
        # COMPROBANTE LÍNEA 1 (FA, NCA, NDA, FCA, FE, etc.)
        # ══════════════════════════════════════════════════════════════════════
        tipo_nombre = _TIPO_NOMBRE.get(tipo, tipo)
        letra       = _TIPO_LETRA.get(tipo, 'A')

        fecha_str = ''
        if cab.get('FECHACOMPROBANTE'):
            try: fecha_str = cab['FECHACOMPROBANTE'].strftime('%d/%m/%Y')
            except Exception: fecha_str = str(cab['FECHACOMPROBANTE'])[:10]

        # Logo
        logo_img = Image(LOGO_PATH, width=38*mm, height=13*mm) if os.path.exists(LOGO_PATH) else Paragraph('<b>microbell S.A.</b>', sT)

        emp_left = [
            logo_img,
            Paragraph('Dirección: PATAGONES 2675 PISO 3', sN),
            Paragraph('C.A.B.A. CAPITAL FEDERAL C1437JEA', sN),
            Paragraph('Télefono: +54 11 3988-0024', sN),
            Paragraph('Email: info@microbellsa.com.ar  www.microbellsa.com', sN),
        ]

        badge_cell = Table([[Paragraph(f'<b>{letra}</b>', ParagraphStyle('ltr',fontSize=30,fontName='Helvetica-Bold',alignment=TA_CENTER))]],
                            colWidths=[20*mm], rowHeights=[22*mm])
        badge_cell.setStyle(TableStyle([('BOX',(0,0),(-1,-1),2,NEGRO),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(0,0),(-1,-1),'CENTER')]))

        cod_pv_label = Paragraph(f'COD. N° {pv_str}', sNc)

        badge_full = Table([[badge_cell],[cod_pv_label]], colWidths=[22*mm])
        badge_full.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(0,0),(-1,-1),'CENTER')]))

        sR  = ParagraphStyle('sR',  fontSize=7, leading=9, fontName='Helvetica',      alignment=TA_RIGHT)
        sRb = ParagraphStyle('sRb', fontSize=7, leading=9, fontName='Helvetica-Bold', alignment=TA_RIGHT)
        emp_right = [
            Paragraph(f'<b>{tipo_nombre}</b>', ParagraphStyle('tn',fontSize=9,fontName='Helvetica-Bold',alignment=TA_RIGHT)),
            Paragraph(f'N° {num_display}',     ParagraphStyle('nn',fontSize=9,fontName='Helvetica-Bold',alignment=TA_RIGHT)),
            Paragraph(f'Fecha emisión: {fecha_str}',        sR),
            Paragraph(f'CUIT: 30-70839018-2',               sR),
            Paragraph(f'Ing. Brutos: CM 901-068199-0',       sR),
            Paragraph(f'Inic. Activ.: 05/09/2005',           sR),
            Paragraph(f'RESPONSABLE INSCRIPTO',              sRb),
        ]

        hdr_tbl = Table([[emp_left, badge_full, emp_right]],
                        colWidths=[cw*0.38, 26*mm, cw*0.50])
        hdr_tbl.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP'),('LEFTPADDING',(0,0),(0,-1),4),('RIGHTPADDING',(0,0),(-1,-1),0)]))
        story.append(hdr_tbl)
        story.append(HRFlowable(width=cw, thickness=2, color=AZUL, spaceAfter=4))

        # Bloque cliente
        cod_cli  = str(cab.get('CODIGOCLIENTE','') or '').strip()
        rs_cli   = str(cab.get('RAZONSOCIAL','') or '').strip()
        dir_cli  = str(cab.get('DIRECCION','') or '').strip()
        cuit_cli = str(cab.get('CUIT','') or '').strip()
        tipoiva  = str(cab.get('TIPOIVA','') or '').strip()
        tel_cli  = str(cab.get('TELEFONO','') or '').strip()

        localidad_full = cli_localidad
        if cli_provincia: localidad_full += f'  Provincia: {cli_provincia}'
        if cli_cp:        localidad_full += f'  C.P.: {cli_cp}'

        cli_data = [
            [Paragraph(f'<b>Cliente:</b>   {cod_cli}  {rs_cli}', sN),
             Paragraph(f'<b>C.U.I.T.:</b>  {cuit_cli}', sN)],
            [Paragraph(f'<b>Dirección:</b>  {dir_cli}', sN),
             Paragraph(f'<b>Ing. Brutos:</b>  {cli_ingbrutos}', sN)],
            [Paragraph(f'<b>Localidad:</b>  {localidad_full}', sN),
             Paragraph(f'<b>Cond. Vta.:</b>  {cond_venta}', sN)],
            [Paragraph(f'<b>I.V.A.:</b>  {tipoiva}   <b>C. P.:</b>  {cli_cp}', sN),
             Paragraph(f'<b>Cod. Vend.:</b>  {cod_vend}   <b>Cod. Asist.:</b>  {cod_asist}', sN)],
        ]
        cli_tbl = Table(cli_data, colWidths=[cw*0.55, cw*0.45])
        cli_tbl.setStyle(TableStyle([
            ('BOX',(0,0),(-1,-1),0.5,colors.grey),
            ('INNERGRID',(0,0),(-1,-1),0.3,colors.lightgrey),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
            ('LEFTPADDING',(0,0),(-1,-1),6),
        ]))
        story.append(cli_tbl)
        story.append(Spacer(1,5))

        # Items
        hdrs_l1 = ['CODIGO','DESCRIPCIÓN','CANTIDAD','P. UNITARIO','DESCUENTO','IVA','TOTAL']
        cws_l1  = [18*mm, cw-18*mm-22*mm-26*mm-22*mm-14*mm-26*mm, 22*mm, 26*mm, 22*mm, 14*mm, 26*mm]
        it_rows = [[Paragraph(h, sNb) for h in hdrs_l1]]
        subtotal = 0.0; iva21_tot = 0.0; iva105_tot = 0.0
        total_neto = 0.0
        for it in items:
            cant  = float(it[3] or 0)
            pu    = float(it[4] or 0)
            dto   = float(it[5] or 0)
            piva  = float(it[6] or 0)
            ptot  = float(it[7] or 0)
            neto_item = pu * cant * (1 - dto/100)
            total_neto += neto_item
            iva_item = neto_item * piva / 100
            if piva >= 20: iva21_tot  += iva_item
            elif piva > 0: iva105_tot += iva_item
            subtotal += neto_item
            it_rows.append([
                Paragraph(str(it[1] or ''), sNc),
                Paragraph(str(it[2] or ''), sN),
                Paragraph(f'{cant:g}', sNc),
                Paragraph(_pesos(pu), sNr),
                Paragraph(f'{dto:.2f}%', sNc),
                Paragraph(f'{piva:.2f}%', sNr),
                Paragraph(_pesos(ptot), sNr),
            ])

        it_tbl = Table(it_rows, colWidths=cws_l1, repeatRows=1)
        it_tbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),CELESTE),('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,GRIS]),
            ('BOX',(0,0),(-1,-1),0.5,colors.grey),
            ('INNERGRID',(0,0),(-1,-1),0.3,colors.lightgrey),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
            ('LEFTPADDING',(0,0),(-1,-1),4),
        ]))
        story.append(it_tbl)

        # Sección Despachos (antes del footer) — en línea separados por |
        if despachos_lista:
            story.append(Spacer(1,4))
            desp_txt = ' &nbsp;|&nbsp; '.join(despachos_lista)
            story.append(Paragraph(f'<b>Despachos:</b> {desp_txt}',
                                   ParagraphStyle('sDesp', fontSize=7, leading=9,
                                                  fontName='Helvetica', wordWrap='CJK')))

        # Footer: cuentas bancarias + totales en flujo normal (antes del BottomSpacer)
        obs   = str(cab.get('COMENTARIOS','') or '').strip()
        cotiz = float(cab.get('COTIZACION',1) or 1)
        total_cab = float(cab.get('TOTAL',0) or 0)
        iva1_cab  = float(cab.get('IVA1',0) or 0)
        iva2_cab  = float(cab.get('IVA2',0) or 0)
        gran_total = total_cab + iva1_cab + iva2_cab
        dto_total  = float(cab.get('DESCUENTOMONTO',0) or 0)

        bw_l1 = [26*mm, 46*mm, 18*mm, 26*mm]   # 116mm total
        bk_data = [[Paragraph('<b>CUENTAS BANCARIAS</b>', sTc), '', '', '']]
        bk_data.append([Paragraph('<b>BANCO</b>',sNb),Paragraph('<b>CBU</b>',sNb),Paragraph('<b>SUC.</b>',sNb),Paragraph('<b>CTA. CTE.</b>',sNb)])
        for b,cbu,suc,cta in _BANCOS:
            bk_data.append([Paragraph(b,sN),Paragraph(cbu,sN),Paragraph(suc,sNc),Paragraph(cta,sN)])
        bk_data.append([Paragraph(_CVU, sN), '', '', ''])
        bk_tbl = Table(bk_data, colWidths=bw_l1)
        bk_tbl.setStyle(TableStyle([
            ('SPAN',(0,0),(-1,0)), ('BACKGROUND',(0,0),(-1,0),CELESTE), ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('BACKGROUND',(0,1),(-1,1),colors.HexColor('#d0e4f7')),
            ('BOX',(0,0),(-1,-1),0.5,colors.grey), ('INNERGRID',(0,0),(-1,-1),0.3,colors.lightgrey),
            ('SPAN',(0,-1),(-1,-1)),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'), ('TOPPADDING',(0,0),(-1,-1),2),
            ('BOTTOMPADDING',(0,0),(-1,-1),2), ('LEFTPADDING',(0,0),(-1,-1),3),
        ]))

        sTotL1 = ParagraphStyle('sTotL1', fontSize=12, fontName='Helvetica-Bold', alignment=TA_RIGHT)
        tot_rows = [
            [Paragraph('Neto gravado:', sN), Paragraph(_pesos(total_neto), sNr)],
            [Paragraph('Descuento %:', sN),  Paragraph(_pesos(dto_total), sNr)],
            [Paragraph('Subtotal:', sN),      Paragraph(_pesos(total_cab), sNr)],
            [Paragraph('Perc:', sN),          Paragraph('$ 0,00', sNr)],
            [Paragraph('I.V.A. 21%:', sN),    Paragraph(_pesos(iva1_cab), sNr)],
            [Paragraph('I.V.A. 10,5%:', sN),  Paragraph(_pesos(iva2_cab), sNr)],
            [Paragraph('<b>TOTAL:</b>', sTotL1), Paragraph(f'<b>{_pesos(gran_total)}</b>', sTotL1)],
        ]
        tot_tbl = Table(tot_rows, colWidths=[32*mm, 40*mm])
        tot_tbl.setStyle(TableStyle([
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'), ('TOPPADDING',(0,0),(-1,-1),2), ('BOTTOMPADDING',(0,0),(-1,-1),2),
            ('LINEABOVE',(0,-1),(-1,-1),1.5,CELESTE),
        ]))

        bk_total_l1 = sum(bw_l1)
        tot_total_l1 = 72*mm
        pad_l1 = (cw - bk_total_l1 - tot_total_l1) / 2

        footer_tbl = Table(
            [[Spacer(1,1), bk_tbl, Spacer(1,4), tot_tbl]],
            colWidths=[pad_l1, bk_total_l1, 4*mm, tot_total_l1]
        )
        footer_tbl.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'BOTTOM')]))

        # Todo el footer al pie de página (banco + totales + obs + QR)
        story.append(BottomSpacer(95*mm))

        footer_items = [footer_tbl]
        if obs:
            footer_items.append(Spacer(1,3))
            footer_items.append(Paragraph(f'Observaciones: {obs}', sN))
        footer_items.append(Paragraph(f'Tipo de Cambio: $ {cotiz:.2f}', sN))
        footer_items.append(Spacer(1,8))
        footer_items.append(HRFlowable(width=cw, thickness=0.5, color=colors.grey, spaceAfter=6))

        if cae:
            vto_cae_str = ''
            if vto_cae:
                try: vto_cae_str = vto_cae.strftime('%d/%m/%Y')
                except Exception: vto_cae_str = str(vto_cae)[:10]

            import json as _json, base64 as _b64
            _TIPO_AFIP_MAP = {
                'FA':1,'NDA':2,'NCA':3,'FB':6,'NDB':7,'NCB':8,
                'FCA':201,'FCB':206,'FE':19,'NCCA':203,'NDE':12,'NCE':13,
                'FE_AFIP':19,
            }
            gran_total_cae = float(cab.get('TOTAL',0) or 0) + float(cab.get('IVA1',0) or 0) + float(cab.get('IVA2',0) or 0)
            cuit_rec_raw   = str(cab.get('CUIT','0') or '0')
            cuit_rec_num   = int(''.join(c for c in cuit_rec_raw if c.isdigit()) or '0')
            pv_int         = int(pv_str)
            seq_int        = int(seq_str)
            fecha_afip     = ''
            if cab.get('FECHACOMPROBANTE'):
                try: fecha_afip = cab['FECHACOMPROBANTE'].strftime('%Y-%m-%d')
                except Exception: fecha_afip = str(cab['FECHACOMPROBANTE'])[:10]
            qr_data = {
                "ver":1, "fecha":fecha_afip,
                "cuit":30708390182, "ptoVta":pv_int,
                "tipoCmp":_TIPO_AFIP_MAP.get(tipo,1), "nroCmp":seq_int,
                "importe":round(gran_total_cae,2), "moneda":"PES", "ctz":1.0,
                "tipoDocRec":80, "nroDocRec":cuit_rec_num,
                "tipoCodAut":"E", "codAut":int(cae)
            }
            import urllib.parse as _uparse
            _json_bytes = _json.dumps(qr_data, separators=(',',':')).encode()
            qr_b64 = _b64.b64encode(_json_bytes).decode().rstrip('=')
            qr_afip_url = ('https://servicioscf.afip.gob.ar/publico/comprobantes/cae.aspx?p='
                           + _uparse.quote(qr_b64, safe=''))
            qr_web_url = 'https://www.microbellsa.com'

            qr_img_l = qr_img_r = None
            try:
                qr_img_l = _make_qr_img(qr_afip_url)
                qr_img_r = _make_qr_img(qr_web_url)
            except Exception as _qe:
                _qe_txt = f'[{type(_qe).__name__}]'
                qr_img_l = Paragraph(_qe_txt, sN)
                qr_img_r = Paragraph(_qe_txt, sN)

            sCAEc = ParagraphStyle('sCAEc', fontSize=8, leading=11, fontName='Helvetica')
            sCAEb = ParagraphStyle('sCAEb', fontSize=8, leading=11, fontName='Helvetica-Bold')
            cae_center = [
                Paragraph(f'Factura Electrónica / CAE:', sCAEc),
                Paragraph(f'<b>{cae}</b>', sCAEb),
                Spacer(1,4),
                Paragraph(f'Fecha Vencimiento CAE:', sCAEc),
                Paragraph(f'<b>{vto_cae_str}</b>', sCAEb),
            ]
            qr_w = 34*mm
            # QR Mercado Pago — solo para Facturas tipo A
            import os as _os_mp
            _QR_MP_PATH = _os_mp.path.join(_os_mp.path.dirname(__file__), 'qr_mercadopago.jpeg')
            _qr_mp_cell = None
            if tipo.upper() in ('FA','FCA','FCE','FCCA','FCCE','FE') and _os_mp.path.exists(_QR_MP_PATH):
                from reportlab.platypus import Image as _RLImgMP
                _sMP = ParagraphStyle('sMP', fontSize=6, fontName='Helvetica-Bold',
                                      alignment=TA_CENTER, leading=7, textColor=colors.HexColor('#009ee3'))
                _qr_mp_cell = [_RLImgMP(_QR_MP_PATH, width=qr_w, height=qr_w),
                               Paragraph('Mercado Pago', _sMP)]

            if _qr_mp_cell:
                mid_w = cw - 3*qr_w
                cae_tbl = Table([[qr_img_l, cae_center, _qr_mp_cell, qr_img_r]],
                                colWidths=[qr_w, mid_w, qr_w, qr_w])
                cae_tbl.setStyle(TableStyle([
                    ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                    ('ALIGN',(0,0),(0,-1),'LEFT'),
                    ('ALIGN',(2,0),(2,-1),'CENTER'),
                    ('ALIGN',(3,0),(3,-1),'RIGHT'),
                    ('ALIGN',(1,0),(1,-1),'LEFT'),
                    ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2),
                ]))
            else:
                mid_w = cw - 2*qr_w
                cae_tbl = Table([[qr_img_l, cae_center, qr_img_r]],
                                colWidths=[qr_w, mid_w, qr_w])
                cae_tbl.setStyle(TableStyle([
                    ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                    ('ALIGN',(0,0),(0,-1),'LEFT'),
                    ('ALIGN',(2,0),(2,-1),'RIGHT'),
                    ('ALIGN',(1,0),(1,-1),'LEFT'),
                    ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2),
                ]))
            footer_items.append(cae_tbl)
        else:
            footer_items.append(Paragraph('Factura Electrónica / CAE: (pendiente)', sN))
        story.append(KeepTogether(footer_items))

    doc.build(story)
    buf.seek(0)
    pv_s, seq_s, _ = _fmt_num(numero)
    fname = f'{tipo}_{pv_s}-{seq_s}.pdf'
    return StreamingResponse(buf, media_type='application/pdf',
                             headers={'Content-Disposition': f'inline; filename="{fname}"'})

# ─── Debug despachos comprobante ──────────────────────────────────────────────
@app.get("/debug/despachos2/{tipo}/{numero}")
def debug_despachos2(tipo: str, numero: str):
    """Muestra contenido de DETALLEDESPACHOVENTAS para el comprobante dado,
       y los artículos del comprobante cruzados con STOCKXDESPACHO."""
    result = {}
    _DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    for db_path in [DATABASE, _DB_PROD]:
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()

            # 1. Artículos de la factura
            cur.execute(
                'SELECT LINEA,CODIGOARTICULO FROM "CUERPOCOMPROBANTES" '
                'WHERE TIPOCOMPROBANTE=? AND NUMEROCOMPROBANTE=? ORDER BY LINEA',
                (tipo, numero)
            )
            arts = [{'linea': r[0], 'articulo': str(r[1] or '').strip()} for r in cur.fetchall()]
            result['articulos'] = arts

            # 2. DETALLEDESPACHOVENTAS — raw con exactamente ese tipo/numero
            cur.execute(
                'SELECT FIRST 50 TIPOCOMPROBANTE, NUMEROCOMPROBANTE, LINEA, DESPACHO, CANTIDAD '
                'FROM "DETALLEDESPACHOVENTAS" '
                'WHERE TIPOCOMPROBANTE=? AND NUMEROCOMPROBANTE=? ORDER BY LINEA, DESPACHO',
                (tipo, numero)
            )
            rows_ddv = [{'tipo': str(r[0]).strip(), 'numero': str(r[1]).strip(),
                         'linea': r[2], 'despacho': str(r[3] or '').strip(), 'cantidad': r[4]}
                        for r in cur.fetchall()]
            result['DETALLEDESPACHOVENTAS_exacto'] = rows_ddv

            # 3. DETALLEDESPACHOVENTAS — busca con LIKE para ver si numero está en otro formato
            cur.execute(
                'SELECT FIRST 20 TIPOCOMPROBANTE, NUMEROCOMPROBANTE, LINEA, DESPACHO '
                'FROM "DETALLEDESPACHOVENTAS" '
                "WHERE TIPOCOMPROBANTE=? AND TRIM(DESPACHO)<>'' "
                "AND DESPACHO<>'000000000000000' ORDER BY NUMEROCOMPROBANTE DESC",
                (tipo,)
            )
            rows_sample = [{'tipo': str(r[0]).strip(), 'numero': str(r[1]).strip(),
                            'linea': r[2], 'despacho': str(r[3] or '').strip()}
                           for r in cur.fetchall()]
            result['DETALLEDESPACHOVENTAS_sample_mismo_tipo'] = rows_sample

            # 4. STOCKXDESPACHO por artículos de la factura
            codigos = [a['articulo'] for a in arts if a['articulo']]
            if codigos:
                ph = ','.join(['?' for _ in codigos])
                cur.execute(
                    f'SELECT CODIGOARTICULO, DESPACHO, ADUANA '
                    f'FROM "STOCKXDESPACHO" WHERE CODIGOARTICULO IN ({ph}) '
                    f"AND TRIM(DESPACHO)<>'' ORDER BY CODIGOARTICULO, DESPACHO",
                    codigos
                )
                result['STOCKXDESPACHO_por_articulo'] = [
                    {'articulo': str(r[0]).strip(), 'despacho': str(r[1] or '').strip(),
                     'aduana': str(r[2] or '').strip()}
                    for r in cur.fetchall()
                ]
            c.close()
            result['db_usado'] = db_path
            break
        except Exception as e:
            result[f'error_{db_path}'] = str(e)
    return result

# ─── Rubros / jerarquía ────────────────────────────────────────────────────────
@app.get("/debug/tablas-mlt")
def debug_tablas_mlt():
    """Lista todas las tablas de DB-MLT-Prueba.gdb"""
    try:
        c = conn('WIN1252', db=DATABASE_MLT)
        cur = c.cursor()
        cur.execute("SELECT TRIM(RDB$RELATION_NAME) FROM RDB$RELATIONS WHERE RDB$SYSTEM_FLAG=0 ORDER BY RDB$RELATION_NAME")
        tablas = [r[0] for r in cur.fetchall()]
        c.close()
        return {"tablas": tablas, "db": DATABASE_MLT}
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/tablas_rubro")
def debug_tablas_rubro():
    c = conn()
    cur = c.cursor()
    # Columnas de cada tabla de jerarquía
    result = {}
    for tabla in ['GRUPOSUPERRUBROS','SUPERRUBROS','RUBROS']:
        cur.execute(f"SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME='{tabla}' ORDER BY RDB$FIELD_POSITION")
        result[f'cols_{tabla}'] = [r[0] for r in cur.fetchall()]
        try:
            cur.execute(f'SELECT FIRST 2 * FROM "{tabla}"')
            result[f'muestra_{tabla}'] = [list(r) for r in cur.fetchall()]
        except Exception as e:
            result[f'muestra_{tabla}_err'] = str(e)
    c.close()
    return result

@app.get("/gruposuperrubros")
def get_gruposuperrubros():
    return _get_pub_filtros()['gsr']

@app.get("/superrubros")
def get_superrubros(grupo: Optional[str] = None):
    if not grupo:
        return _get_pub_filtros()['sr']
    # Filtrado en cascada (usuario ya seleccionó un GSR): consulta puntual en vivo
    c = conn()
    cur = c.cursor()
    cur.execute("""
        SELECT DISTINCT sr.CODIGOSUPERRUBRO, sr.DESCRIPCION
        FROM "SUPERRUBROS" sr
        WHERE sr.CODIGOGRUPOSUPERRUBRO = ?
          AND EXISTS (
            SELECT 1 FROM "ARTICULOS" a
            JOIN "RUBROS" r ON r.CODIGORUBRO = a.CODIGORUBRO
            WHERE r.CODIGOSUPERRUBRO = sr.CODIGOSUPERRUBRO AND a.ACTIVO = '1'
          )
        ORDER BY sr.DESCRIPCION
    """, (grupo,))
    rows = cur.fetchall()
    c.close()
    return [{"codigo": r[0], "descripcion": r[1], "grupo": r[2] if len(r)>2 else None} for r in rows]

@app.get("/rubros")
def get_rubros(superrubro: Optional[str] = None, grupo: Optional[str] = None):
    if not superrubro and not grupo:
        return _get_pub_filtros()['rubro']
    # Filtrado en cascada (usuario ya seleccionó un SR o GSR): consulta puntual en vivo
    c = conn()
    cur = c.cursor()
    params = []
    filtro_sr = ""
    if superrubro:
        filtro_sr = "AND r.CODIGOSUPERRUBRO = ?"
        params.append(superrubro)
    elif grupo:
        filtro_sr = "AND sr.CODIGOGRUPOSUPERRUBRO = ?"
        params.append(grupo)
    cur.execute(f"""
        SELECT DISTINCT r.CODIGORUBRO, r.DESCRIPCION, r.CODIGOSUPERRUBRO
        FROM "RUBROS" r
        JOIN "SUPERRUBROS" sr ON sr.CODIGOSUPERRUBRO = r.CODIGOSUPERRUBRO
        WHERE EXISTS (
            SELECT 1 FROM "ARTICULOS" a
            WHERE a.CODIGORUBRO = r.CODIGORUBRO AND a.ACTIVO = '1'
        )
        {filtro_sr}
        ORDER BY r.DESCRIPCION
    """, params)
    rows = cur.fetchall()
    c.close()
    return [{"codigo": r[0], "descripcion": r[1], "superrubro": r[2]} for r in rows]

# ─── DEBUG: verificar matching de reservas para un artículo ──────────────────
@app.get("/debug/reservas/{codigoparticular}")
def debug_reservas_articulo(codigoparticular: str, token: Optional[str] = None, request: Request = None):
    """Muestra reservas activas y cómo matchean con el artículo dado (solo admin)."""
    reservas = _get_reservas_activas()
    # Obtener datos del artículo desde Firebird
    c = conn()
    cur = c.cursor()
    cur.execute(
        'SELECT a.CODIGOARTICULO, a.CODIGOPARTICULAR, a.DESCRIPCION, a.CODIGOMARCA,'
        ' a.CODIGORUBRO, r.CODIGOSUPERRUBRO, sr.CODIGOGRUPOSUPERRUBRO'
        ' FROM "ARTICULOS" a'
        ' LEFT JOIN "RUBROS" r ON r.CODIGORUBRO = a.CODIGORUBRO'
        ' LEFT JOIN "SUPERRUBROS" sr ON sr.CODIGOSUPERRUBRO = r.CODIGOSUPERRUBRO'
        ' WHERE a.CODIGOPARTICULAR = ?',
        (codigoparticular,)
    )
    row = cur.fetchone()
    c.close()
    if not row:
        return {"error": f"Artículo {codigoparticular!r} no encontrado"}
    item = {
        "codigo":           str(row[0] or '').strip(),
        "codigoparticular": str(row[1] or '').strip(),
        "descripcion":      row[2],
        "marca":            str(row[3] or '').strip(),
        "codigo_rubro":     str(row[4] or '').strip(),
        "codigo_superrubro": str(row[5] or '').strip(),
        "codigo_gruposuperrubro": str(row[6] or '').strip(),
    }
    resultado = []
    for rv in reservas:
        rv_art  = str(rv.get('codigo_articulo')  or '').strip()
        rv_part = str(rv.get('codigo_particular') or '').strip()
        it_cod  = item["codigo"]
        it_part = item["codigoparticular"]
        matches = {
            "rv_art==it_cod":  rv_art  == it_cod  if rv_art  else None,
            "rv_part==it_part": rv_part == it_part if rv_part else None,
            "rv_art==it_part": rv_art  == it_part if rv_art  else None,
            "rv_part==it_cod": rv_part == it_cod  if rv_part else None,
        }
        applies = any(v for v in matches.values() if v is not None)
        resultado.append({
            "reserva_id":       rv.get('id'),
            "tipo":             rv.get('tipo'),
            "deposito":         rv.get('deposito'),
            "cantidad":         rv.get('cantidad'),
            "cantidad_utilizada": rv.get('cantidad_utilizada'),
            "motivo":           rv.get('motivo'),
            "rv_codigo_articulo":  rv_art or None,
            "rv_codigo_particular": rv_part or None,
            "matches":          matches,
            "aplica":           applies,
        })
    # Replicar EXACTAMENTE lo que hace /stock para este artículo:
    # FMA_STOCK(NULL, NULL, dep, 1, 1) — misma query bulk de producción
    codigo_interno = item['codigo']  # "03421"
    rem_prod = {}
    for dep in ['001', '003']:
        try:
            c2 = conn()
            cur2 = c2.cursor()
            cur2.execute(f'SELECT ID_ARTICULO, STOCKREMANENTE FROM "FMA_STOCK"(NULL, NULL, \'{dep}\', 1, 1)')
            rem_map = {str(r[0]).strip(): float(r[1] or 0) for r in cur2.fetchall()}
            c2.close()
            # buscar por codigo_interno con y sin strip
            val = rem_map.get(codigo_interno, rem_map.get(codigo_interno.strip(), None))
            # fallback: buscar por codigoparticular
            if val is None:
                val = rem_map.get(item['codigoparticular'], 0.0)
            rem_prod[dep] = val if val is not None else 0.0
        except Exception as e:
            rem_prod[dep] = f"ERROR: {e}"
    # Construir item igual que /stock y aplicar reservas
    test_item = {
        "codigo":           codigo_interno,
        "codigoparticular": item['codigoparticular'],
        "remanente":        (rem_prod.get('001', 0) or 0) + (rem_prod.get('003', 0) or 0),
        "remanente_001":    rem_prod.get('001', 0),
        "remanente_003":    rem_prod.get('003', 0),
        "marca":            item.get('marca', ''),
        "codigo_rubro":     item.get('codigo_rubro', ''),
        "codigo_superrubro": item.get('codigo_superrubro', ''),
        "codigo_gruposuperrubro": item.get('codigo_gruposuperrubro', ''),
    }
    _apply_reservas([test_item], _get_reservas_activas(), rem_key='remanente')
    return {
        "articulo": item,
        "reservas_activas": resultado,
        "simulacion_exacta_produccion": {
            "remanente_001_firebird_raw":  rem_prod.get('001'),
            "remanente_003_firebird_raw":  rem_prod.get('003'),
            "reservado_deposito_001":      (test_item.get("reservado_por_deposito") or {}).get("001", 0),
            "remanente_001_post_reserva":  test_item.get("remanente_001"),
            "remanente_003_post_reserva":  test_item.get("remanente_003"),
            "lo_que_ve_el_vendedor_001":   test_item.get("remanente_001"),
        }
    }

@app.get("/marcas")
def get_marcas():
    return _get_pub_filtros()['marca']

# ─── Debug ────────────────────────────────────────────────────────────────────
@app.get("/debug/stock")
def debug_stock():
    try:
        c = conn()
        cur = c.cursor()
        resultado = {}

        # Columnas de STOCK
        cur.execute("SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME='STOCK' ORDER BY RDB$FIELD_POSITION")
        resultado["columnas_stock"] = [r[0] for r in cur.fetchall()]

        # Primeras 5 filas de STOCK (ver qué campos tiene)
        cur.execute('SELECT FIRST 5 * FROM "STOCK"')
        cols = [d[0] for d in cur.description]
        resultado["stock_muestra_cols"] = cols
        resultado["stock_muestra_rows"] = [list(r) for r in cur.fetchall()]

        # Valores distintos del campo que identifica depósito (CODIGODEPOSITO o CODIGOSUCURSAL)
        for campo in ["CODIGODEPOSITO", "CODIGOSUCURSAL", "DEPOSITO"]:
            try:
                cur.execute(f'SELECT DISTINCT "{campo}" FROM "STOCK"')
                resultado[f"distintos_{campo}"] = [r[0] for r in cur.fetchall()]
            except Exception:
                pass

        c.close()
        return resultado
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/debug/presupuesto/{numero}")
def debug_presupuesto(numero: str):
    """Muestra los campos de aprobación y remisión tal como están en la DB."""
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        cur.execute(
            'SELECT NUMEROCOMPROBANTE, FECHAAPROBADO, CODIGOUSUARIOAPROBACION, '
            'CODIGOOPERACION, ANULADA, CLASECOMPROBANTE, CODIGOMULTIPLAZO, '
            'COTIZACIONFIJA, LISTAPRECIO, NUMEROTRANSACCION, CODIGORESPONSABLE, '
            'FECHAVENCIMIENTO, CODIGOUSUARIO, CODIGOUSUARIO2 '
            'FROM "CABEZAPRESUPUESTOS" WHERE NUMEROCOMPROBANTE = ?', (numero,)
        )
        cab = cur.fetchone()
        cur.execute(
            'SELECT LINEA, CANTIDAD, CANTIDADREMITIDA '
            'FROM "CUERPOPRESUPUESTOS" WHERE NUMEROCOMPROBANTE = ? ORDER BY LINEA',
            (numero,)
        )
        items = cur.fetchall()
        # Revisar triggers sobre CABEZAPRESUPUESTOS (nombre + tipo + source)
        cur.execute(
            "SELECT TRIM(RDB$TRIGGER_NAME), RDB$TRIGGER_TYPE, RDB$TRIGGER_SOURCE "
            "FROM RDB$TRIGGERS "
            "WHERE RDB$RELATION_NAME = 'CABEZAPRESUPUESTOS' AND RDB$SYSTEM_FLAG = 0"
        )
        triggers = [{"name": r[0], "type": r[1], "source": str(r[2])[:800] if r[2] else None}
                    for r in cur.fetchall()]
        # Default del campo CODIGOUSUARIOAPROBACION
        cur.execute("""
            SELECT rf.RDB$DEFAULT_SOURCE
            FROM RDB$RELATION_FIELDS rf
            WHERE rf.RDB$RELATION_NAME = 'CABEZAPRESUPUESTOS'
              AND TRIM(rf.RDB$FIELD_NAME) = 'CODIGOUSUARIOAPROBACION'
        """)
        col_def = cur.fetchone()
        c.close()
        return {
            "cabeza": {
                "numero":                   cab[0],
                "fechaaprobado":            str(cab[1]),
                "codigousuarioaprobacion":  repr(cab[2]),
                "codigooperacion":          cab[3],
                "anulada":                  cab[4],
                "clasecomprobante":         cab[5],
                "codigomultiplazo":         cab[6],
                "cotizacionfija":           cab[7],
                "listaprecio":              cab[8],
                "numerotransaccion":        cab[9],
                "codigoresponsable":        repr(cab[10]),
                "fechavencimiento":         str(cab[11]),
                "codigousuario":            cab[12],
                "codigousuario2":           cab[13],
            } if cab else None,
            "items": [{"linea": r[0], "cantidad": r[1], "cantidadremitida": r[2]} for r in items],
            "triggers": triggers,
            "col_default_aprobacion": str(col_def[0]) if col_def and col_def[0] else None,
        }
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/debug/depositos")
def debug_depositos():
    try:
        c = conn()
        cur = c.cursor()
        resultado = {}

        for tabla in ["DEPOSITOS", "SUBDEPOSITOS", "WEB_STOCK", "STOCKPORUSUARIO", "STOCKPORUSUARIODETALLE"]:
            try:
                cur.execute(f"SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME='{tabla}' ORDER BY RDB$FIELD_POSITION")
                resultado[f"cols_{tabla}"] = [r[0] for r in cur.fetchall()]
            except Exception as ex:
                resultado[f"cols_{tabla}_error"] = str(ex)

        # Muestra los depósitos existentes
        try:
            cur.execute('SELECT FIRST 20 * FROM "DEPOSITOS"')
            cols = [d[0] for d in cur.description]
            resultado["depositos_cols"] = cols
            resultado["depositos_rows"] = [list(r) for r in cur.fetchall()]
        except Exception as ex:
            resultado["depositos_error"] = str(ex)

        # Muestra sample de WEB_STOCK
        try:
            cur.execute('SELECT FIRST 3 * FROM "WEB_STOCK"')
            cols = [d[0] for d in cur.description]
            resultado["web_stock_cols"] = cols
            resultado["web_stock_rows"] = [list(r) for r in cur.fetchall()]
        except Exception as ex:
            resultado["web_stock_error"] = str(ex)

        c.close()
        return resultado
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/debug/articulos_monedas")
def debug_articulos_monedas():
    try:
        c = conn()
        cur = c.cursor()
        resultado = {}
        # Campos de ARTICULOS que tengan CODIGO o PARTICULAR
        cur.execute("SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME='ARTICULOS' AND (RDB$FIELD_NAME CONTAINING 'PARTICULAR' OR RDB$FIELD_NAME CONTAINING 'PRECIO') ORDER BY RDB$FIELD_POSITION")
        resultado["articulos_precio_particular"] = [r[0] for r in cur.fetchall()]
        # Muestra artículo 01315 - valores raw de precio
        cur.execute(
            "SELECT CODIGOARTICULO, CODIGOPARTICULAR, CODIGOMONEDA, "
            "PRECIOLISTA1, PRECIOLISTA2, PRECIOLISTA3 "
            "FROM \"ARTICULOS\" WHERE CODIGOARTICULO = '01315' OR CODIGOPARTICULAR = '01315'"
        )
        row = cur.fetchone()
        if row:
            resultado["art_01315"] = {
                "CODIGOARTICULO": row[0], "CODIGOPARTICULAR": row[1],
                "CODIGOMONEDA": row[2],
                "PRECIOLISTA1_raw": str(row[3]), "PRECIOLISTA1_float": float(row[3]) if row[3] else None,
                "PRECIOLISTA2_raw": str(row[4]), "PRECIOLISTA3_raw": str(row[5]),
            }
        # CAMBIO raw de MONEDAS para DOLARES
        cur.execute('SELECT CODIGOMONEDA, CAMBIO, CAST(CAMBIO AS VARCHAR(30)) FROM "MONEDAS" WHERE CODIGOMONEDA = \'DOLARES\'')
        rm = cur.fetchone()
        if rm:
            resultado["dolares_cambio"] = {"raw": str(rm[1]), "str": rm[2], "float": float(rm[1]) if rm[1] else None}
            if row and row[3] and rm[1]:
                resultado["calculo"] = {
                    "preciolista1_float": float(row[3]),
                    "cambio_float": float(rm[1]),
                    "resultado_sin_round": float(row[3]) * float(rm[1]),
                    "resultado_round2": round(float(row[3]) * float(rm[1]), 2),
                    "resultado_round4": round(float(row[3]) * float(rm[1]), 4),
                }
        # Estructura y datos de MONEDAS
        cur.execute("SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME='MONEDAS' ORDER BY RDB$FIELD_POSITION")
        resultado["monedas_cols"] = [r[0] for r in cur.fetchall()]
        # Columnas de CLIENTES (para identificar campo IVA)
        cur.execute("SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS WHERE RDB$RELATION_NAME='CLIENTES' ORDER BY RDB$FIELD_POSITION")
        all_cols = [r[0] for r in cur.fetchall()]
        resultado["clientes_cols"] = all_cols
        # Busca columnas relacionadas a IVA/condición fiscal
        resultado["clientes_cols_iva"] = [c for c in all_cols if any(k in c for k in ['IVA','CONDIC','CATEG','FISCAL','RESPON'])]
        # Buscar generadores/secuencias de Firebird (para numeración PR)
        cur.execute("SELECT TRIM(RDB$GENERATOR_NAME), RDB$GENERATOR_ID FROM RDB$GENERATORS WHERE RDB$SYSTEM_FLAG=0 ORDER BY RDB$GENERATOR_NAME")
        resultado["generators"] = [{"name": r[0], "id": r[1]} for r in cur.fetchall()]
        # Tablas que podrían tener el contador (buscar tablas con columnas NUMERO+TIPO)
        cur.execute("""
            SELECT DISTINCT TRIM(a.RDB$RELATION_NAME)
            FROM RDB$RELATION_FIELDS a
            JOIN RDB$RELATION_FIELDS b ON a.RDB$RELATION_NAME=b.RDB$RELATION_NAME
            WHERE (a.RDB$FIELD_NAME CONTAINING 'NUMERO' OR a.RDB$FIELD_NAME = 'NUMERO')
              AND (b.RDB$FIELD_NAME CONTAINING 'TIPO' OR b.RDB$FIELD_NAME CONTAINING 'COMPROBANTE' OR b.RDB$FIELD_NAME CONTAINING 'DOCUMENTO')
              AND a.RDB$RELATION_NAME NOT STARTING WITH 'RDB$'
              AND a.RDB$RELATION_NAME NOT CONTAINING 'CABEZA'
              AND a.RDB$RELATION_NAME NOT CONTAINING 'CUERPO'
            ORDER BY 1
        """)
        resultado["tablas_contador_candidatas"] = [r[0] for r in cur.fetchall()]

        # Buscar triggers sobre CABEZAPRESUPUESTOS
        try:
            cur.execute("""
                SELECT TRIM(t.RDB$TRIGGER_NAME), t.RDB$TRIGGER_TYPE,
                       CAST(t.RDB$TRIGGER_SOURCE AS VARCHAR(500))
                FROM RDB$TRIGGERS t
                WHERE t.RDB$RELATION_NAME = 'CABEZAPRESUPUESTOS'
                  AND t.RDB$SYSTEM_FLAG = 0
            """)
            resultado["triggers_cabezapresupuestos"] = [
                {"name": r[0], "type": r[1], "src_snippet": str(r[2])[:300] if r[2] else ""}
                for r in cur.fetchall()
            ]
        except Exception as ex:
            resultado["triggers_err"] = str(ex)

        # Valor actual del generador AUXILIAR (podría ser el contador PR)
        try:
            cur.execute("SELECT GEN_ID(AUXILIAR, 0) FROM RDB$DATABASE")
            resultado["gen_auxiliar_value"] = cur.fetchone()[0]
        except Exception as ex:
            resultado["gen_auxiliar_err"] = str(ex)

        c.close()
        return resultado
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/setup/fechaaprobado_nullable")
def setup_fechaaprobado_nullable():
    """Ejecutar una sola vez: quita NOT NULL de FECHAAPROBADO en CABEZAPRESUPUESTOS."""
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        # Quitar NOT NULL y DEFAULT de FECHAAPROBADO
        cur.execute("""
            UPDATE RDB$RELATION_FIELDS
            SET RDB$NULL_FLAG = NULL,
                RDB$DEFAULT_VALUE = NULL,
                RDB$DEFAULT_SOURCE = NULL
            WHERE RDB$RELATION_NAME = 'CABEZAPRESUPUESTOS'
              AND RDB$FIELD_NAME = 'FECHAAPROBADO'
        """)
        c.commit()
        # Verificar
        cur.execute("""
            SELECT RDB$NULL_FLAG, RDB$DEFAULT_SOURCE
            FROM RDB$RELATION_FIELDS
            WHERE RDB$RELATION_NAME = 'CABEZAPRESUPUESTOS'
              AND RDB$FIELD_NAME = 'FECHAAPROBADO'
        """)
        row = cur.fetchone()
        c.close()
        return {"ok": True,
                "RDB$NULL_FLAG": row[0] if row else '?',
                "RDB$DEFAULT_SOURCE": row[1] if row else '?',
                "msg": "Ambos deben ser None/null para que FECHAAPROBADO quede en NULL al insertar"}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/debug/procedimientos")
def debug_procedimientos():
    try:
        c = conn()
        cur = c.cursor()
        resultado = {}
        # Procedimientos almacenados
        cur.execute("SELECT TRIM(RDB$PROCEDURE_NAME) FROM RDB$PROCEDURES WHERE RDB$SYSTEM_FLAG=0 ORDER BY RDB$PROCEDURE_NAME")
        procs = [r[0] for r in cur.fetchall()]
        resultado["procedimientos"] = [p for p in procs if any(x in p for x in ['STOCK','DEPO','REMANEN','SALDO'])]
        resultado["total_procedimientos"] = len(procs)
        # Probar si STOCK tiene datos recientes por depósito via CUERPOCOMPROBANTES
        # Artículo 02785 depósito 003
        cur.execute("""
            SELECT cc.CODIGODEPOSITO,
                   SUM(CASE WHEN cab.TIPOCOMPROBANTE IN ('FA','FE','FCA','FCE','RE','NC','NCA','SIV','FDI')
                             THEN cc.CANTIDAD ELSE 0 END) AS CANT_COMP,
                   COUNT(*) AS FILAS
            FROM "CUERPOCOMPROBANTES" cc
            JOIN "CABEZACOMPROBANTES" cab ON cab.TIPOCOMPROBANTE=cc.TIPOCOMPROBANTE AND cab.NUMEROCOMPROBANTE=cc.NUMEROCOMPROBANTE
            WHERE cc.CODIGOARTICULO='02785' AND cc.CODIGODEPOSITO IN ('001','003')
            GROUP BY cc.CODIGODEPOSITO
        """)
        cols = [d[0] for d in cur.description]
        resultado["movimientos_02785"] = [dict(zip(cols,list(r))) for r in cur.fetchall()]
        # Ver tipos de comprobante distintos en CUERPOCOMPROBANTES
        cur.execute("SELECT DISTINCT TIPOCOMPROBANTE FROM \"CUERPOCOMPROBANTES\" ORDER BY TIPOCOMPROBANTE")
        resultado["tipos_comprobante_cuerpo"] = [r[0] for r in cur.fetchall()]
        c.close()
        return resultado
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/debug/procs_stock")
def debug_procs_stock():
    try:
        c = conn()
        cur = c.cursor()
        resultado = {}
        procs_interesantes = [
            'FMA_CALCULA_STOCKREMANENTE','FMA_CALCULASTOCKREMANENTE',
            'FMA_CALCULA_STOCKREAL','FMA_CALCULASTOCKREAL',
            'FMA_DETALLESTOCK','FMA_STOCK','FMA_DEPOSITOS'
        ]
        for proc in procs_interesantes:
            try:
                cur.execute("""
                    SELECT TRIM(p.RDB$PARAMETER_NAME), p.RDB$PARAMETER_TYPE,
                           TRIM(f.RDB$FIELD_TYPE), p.RDB$PARAMETER_NUMBER
                    FROM RDB$PROCEDURE_PARAMETERS p
                    JOIN RDB$FIELDS f ON f.RDB$FIELD_NAME = p.RDB$FIELD_SOURCE
                    WHERE p.RDB$PROCEDURE_NAME = ?
                    ORDER BY p.RDB$PARAMETER_TYPE, p.RDB$PARAMETER_NUMBER
                """, (proc,))
                params = [{"nombre": r[0], "tipo": "INPUT" if r[1]==0 else "OUTPUT", "campo": r[2]} for r in cur.fetchall()]
                if params:
                    resultado[proc] = params
            except Exception as ex:
                resultado[proc + "_error"] = str(ex)
        # Intentar llamar FMA_DETALLESTOCK con artículo conocido
        for call in [
            ("FMA_DETALLESTOCK", "EXECUTE PROCEDURE \"FMA_DETALLESTOCK\" '02785'"),
            ("FMA_STOCK", "EXECUTE PROCEDURE \"FMA_STOCK\" '02785'"),
        ]:
            try:
                cur.execute(call[1])
                cols = [d[0] for d in cur.description]
                rows = cur.fetchall()
                resultado[call[0]+"_result"] = [dict(zip(cols,list(r))) for r in rows[:5]]
            except Exception as ex:
                resultado[call[0]+"_call_error"] = str(ex)
        c.close()
        return resultado
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/debug/stock_condiciones")
def debug_stock_condiciones():
    """Verifica OPERACIONES.COMPROMETESTOCK y CASILLEROS para entender por qué no descuenta stock"""
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        res = {}
        # 1. Ver COMPROMETESTOCK para cada OPERACION usada en NPs
        cur.execute('SELECT CODIGOOPERACION, DESCRIPCION, COMPROMETESTOCK FROM "OPERACIONES" ORDER BY CODIGOOPERACION')
        res['operaciones'] = [{'codigo': str(r[0]), 'descripcion': str(r[1]).strip(), 'comprometestock': r[2]} for r in cur.fetchall()]
        # 2. Ver si existen registros en CASILLEROS para artículo 00590 (el de la antena)
        cur.execute('SELECT CODIGOARTICULO, LOTE, CODIGODEPOSITO FROM "CASILLEROS" WHERE CODIGOARTICULO = ? ORDER BY LOTE', ('00590',))
        rows = cur.fetchall()
        res['casilleros_00590'] = [{'art': r[0], 'lote': str(r[1]), 'deposito': str(r[2])} for r in rows]
        # 3. Ver qué LOTE usamos en CUERPOPEDIDOS para ese artículo
        cur.execute("SELECT FIRST 3 LOTE, CODIGODEPOSITO, CANTIDAD, CANTIDADREMITIDA FROM \"CUERPOPEDIDOS\" WHERE CODIGOARTICULO = '00590' ORDER BY NUMEROCOMPROBANTE DESC")
        res['cuerpopedidos_00590'] = [{'lote': str(r[0]), 'deposito': str(r[1]), 'cantidad': r[2], 'remitida': r[3]} for r in cur.fetchall()]
        # 4. Ver cuántos casilleros existen en total (para saber si la tabla está poblada)
        cur.execute('SELECT COUNT(*) FROM "CASILLEROS"')
        res['total_casilleros'] = cur.fetchone()[0]
        c.close()
        return res
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/debug/stock_profundo")
def debug_stock_profundo():
    """Fuente de FMA_CALCULA_STOCKPEDIDO + artículo real en CUERPOPEDIDOS + CASILLEROS sample"""
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        res = {}
        # 1. Fuente de FMA_CALCULA_STOCKPEDIDO (sin el filtro EXISTS)
        for sp in ['FMA_CALCULA_STOCKPEDIDO', 'FMA_STOCK']:
            cur.execute('SELECT RDB$PROCEDURE_SOURCE FROM RDB$PROCEDURES WHERE TRIM(RDB$PROCEDURE_NAME)=?', (sp,))
            row = cur.fetchone()
            res[f'src_{sp}'] = str(row[0])[:2000] if row and row[0] else 'NO ENCONTRADO'
        # 2. Buscar el artículo 00590 en ARTICULOS para ver su CODIGOARTICULO interno
        cur.execute("SELECT CODIGOARTICULO, CODIGOPARTICULAR, DESCRIPCION FROM \"ARTICULOS\" WHERE CODIGOPARTICULAR='00590' OR CODIGOARTICULO='00590'")
        rows = cur.fetchall()
        res['articulos_00590'] = [{'interno': r[0], 'particular': r[1], 'desc': str(r[2])[:40]} for r in rows]
        # 3. CUERPOPEDIDOS de los últimos 5 pedidos NP (cualquier artículo)
        cur.execute("SELECT FIRST 5 NUMEROCOMPROBANTE, CODIGOARTICULO, LOTE, CANTIDAD, CODIGODEPOSITO FROM \"CUERPOPEDIDOS\" WHERE TIPOCOMPROBANTE='NP' ORDER BY NUMEROCOMPROBANTE DESC")
        res['cuerpopedidos_ultimos'] = [{'num': str(r[0]), 'art': r[1], 'lote': str(r[2]), 'cant': r[3], 'dep': str(r[4])} for r in cur.fetchall()]
        # 4. Sample de CASILLEROS para ver qué lote usan
        cur.execute("SELECT FIRST 5 CODIGOARTICULO, LOTE, CODIGODEPOSITO FROM \"CASILLEROS\"")
        res['casilleros_sample'] = [{'art': r[0], 'lote': str(r[1]), 'dep': str(r[2])} for r in cur.fetchall()]
        c.close()
        return res
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/debug/sp_remanente")
def debug_sp_remanente():
    """Lee el código fuente de FMA_CALCULASTOCKREMANENTE para ver qué tablas/condiciones usa"""
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        res = {}
        for sp in ['FMA_CALCULASTOCKREMANENTE', 'FMA_CALCULA_STOCKREMANENTE']:
            try:
                cur.execute("""
                    SELECT TRIM(RDB$PROCEDURE_NAME), RDB$PROCEDURE_SOURCE
                    FROM RDB$PROCEDURES
                    WHERE TRIM(RDB$PROCEDURE_NAME) = ?
                """, (sp,))
                row = cur.fetchone()
                res[sp] = str(row[1]) if row and row[1] else 'NO ENCONTRADO'
            except Exception as e2:
                res[sp] = str(e2)
        c.close()
        return res
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/debug/fma_stock")
def debug_fma_stock():
    try:
        c = conn()
        cur = c.cursor()
        resultado = {}
        # Test FMA_STOCK con filtro de depósitos 001 y 003
        for depositos in ["001,003", "001|003", "001;003", "001 003"]:
            try:
                cur.execute(f"""
                    SELECT FIRST 3 ID_ARTICULO, CODIGO_PRODUCTO, STOCKREAL, STOCKREMANENTE
                    FROM "FMA_STOCK"(NULL, NULL, '{depositos}', 1, 1)
                    WHERE STOCKREAL > 0
                """)
                cols = [d[0] for d in cur.description]
                rows = [dict(zip(cols, list(r))) for r in cur.fetchall()]
                resultado[f"FMA_STOCK_depositos_{depositos}"] = rows
                break
            except Exception as ex:
                resultado[f"FMA_STOCK_depositos_{depositos}_error"] = str(ex)
        # Test FMA_CALCULASTOCKREMANENTE para artículo 02785
        for depo in ['001', '003']:
            try:
                cur.execute('EXECUTE PROCEDURE "FMA_CALCULASTOCKREMANENTE" ?, ?, ?', ('02785', '000', depo))
                row = cur.fetchone()
                resultado[f"remanente_02785_dep{depo}"] = float(row[0]) if row else None
            except Exception as ex:
                resultado[f"remanente_02785_dep{depo}_error"] = str(ex)
        # Test FMA_CALCULA_STOCKREAL para artículo 02785
        for depo in ['001', '003']:
            try:
                cur.execute('EXECUTE PROCEDURE "FMA_CALCULA_STOCKREAL" ?, ?, ?', ('02785', '000', depo))
                row = cur.fetchone()
                resultado[f"stockreal_02785_dep{depo}"] = float(row[0]) if row else None
            except Exception as ex:
                resultado[f"stockreal_02785_dep{depo}_error"] = str(ex)
        c.close()
        return resultado
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/debug/cta/{codigo}")
def debug_cta(codigo: str):
    resultado = {}
    # 1. Lookup en CLIENTES
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        cur.execute('SELECT CODIGOCLIENTE, CODIGOPARTICULAR, RAZONSOCIAL FROM "CLIENTES" WHERE CODIGOCLIENTE = ?', (codigo,))
        row = cur.fetchone()
        c.close()
        resultado['clientes'] = {"codigocliente": row[0], "codigoparticular": row[1], "razonsocial": row[2]} if row else None
    except Exception as e:
        resultado['clientes_error'] = str(e)

    DATABASE_MLT = 'c:/flexxus/db/DB-MLT-Microbell.gdb'
    # 2. BD principal - todos los comprobantes del cliente (sin filtro de saldo)
    for db_key, db_path in [('db_main', DATABASE), ('db_est', DATABASE_EST), ('db_mlt', DATABASE_MLT)]:
        try:
            c = conn('LATIN1', db=db_path)
            cur = c.cursor()
            # Buscar por codigo y por codigoparticular
            cp = resultado.get('clientes', {}) or {}
            codigos = list({codigo, cp.get('codigoparticular','')})
            codigos = [x for x in codigos if x]
            placeholders = ','.join(['?' for _ in codigos])
            cur.execute(f"""
                SELECT TIPOCOMPROBANTE, NUMEROCOMPROBANTE, FECHACOMPROBANTE,
                       CODIGOCLIENTE, TOTAL, IVA1, IVA2, PAGADO, CUENTACORRIENTE, ANULADA
                FROM "CABEZACOMPROBANTES"
                WHERE CODIGOCLIENTE IN ({placeholders})
                  AND TIPOCOMPROBANTE IN ('FA','FB','FE','FCA','FCB','DI','SIV')
                ORDER BY FECHACOMPROBANTE DESC
            """, tuple(codigos))
            rows = cur.fetchall()
            c.close()
            resultado[db_key] = [{"tipo":r[0],"numero":r[1],"fecha":str(r[2]),"cod_cli":r[3],
                                   "total":float(r[4] or 0),"iva1":float(r[5] or 0),"iva2":float(r[6] or 0),
                                   "pagado":float(r[7] or 0),"ctacte":r[8],"anulada":r[9]} for r in rows]
        except Exception as e:
            # Si falla, listar tablas disponibles en esa BD
            resultado[db_key+'_error'] = str(e)
            try:
                c2 = conn('LATIN1', db=db_path)
                cur2 = c2.cursor()
                cur2.execute("SELECT TRIM(RDB$RELATION_NAME) FROM RDB$RELATIONS WHERE RDB$SYSTEM_FLAG = 0 ORDER BY RDB$RELATION_NAME")
                tablas = [r[0] for r in cur2.fetchall()]
                c2.close()
                resultado[db_key+'_tablas'] = tablas
            except Exception as e2:
                resultado[db_key+'_tablas_error'] = str(e2)
    return resultado

@app.get("/debug/cta_detalle/{codigo}")
def debug_cta_detalle(codigo: str, vendedor: Optional[str] = None):
    """Diagnostica por qué un cliente no muestra movimientos en el detalle.
    Muestra: lookup CLIENTES, codigos resueltos, registros en CABEZACOMPROBANTES por tipo y cuentacorriente."""
    DB_PROD_D     = DATABASE
    DB_MLT_PROD_D = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    resultado = {"codigo_ingresado": codigo}

    # 1. Lookup en CLIENTES (igual que endpoint detalle)
    try:
        c = conn('WIN1252', db=DB_PROD_D)
        cur = c.cursor()
        cur.execute(
            'SELECT CODIGOCLIENTE, CODIGOPARTICULAR, RAZONSOCIAL, CODIGOVENDEDOR '
            'FROM "CLIENTES" WHERE CODIGOCLIENTE=? OR CODIGOPARTICULAR=?',
            (codigo, codigo)
        )
        rows_cli = cur.fetchall()
        c.close()
        resultado['clientes_rows'] = [
            {"codigocliente": r[0], "codigoparticular": r[1], "razonsocial": r[2], "codigovendedor": r[3]}
            for r in rows_cli
        ]
        # Armar codigos igual que el endpoint
        codigos_set = set()
        if rows_cli:
            cli = rows_cli[0]
            if cli[0] and str(cli[0]).strip(): codigos_set.add(str(cli[0]).strip())
            if cli[1] and str(cli[1]).strip(): codigos_set.add(str(cli[1]).strip())
        if not codigos_set:
            codigos_set.add(codigo)
        codigos = list(codigos_set)
        resultado['codigos_usados'] = codigos
    except Exception as e:
        resultado['clientes_error'] = str(e)
        codigos = [codigo]

    # 2. Para cada DB: contar por TIPOCOMPROBANTE + CUENTACORRIENTE + ANULADA
    _NC_TIPOS_D = ('NCA','NCB','NCCA','NCCB','NCE','NCCE','SIV','NDA','NDB','NDCA','NDCB')
    for db_key, db_path in [('DB_PROD', DB_PROD_D), ('DB_MLT_PROD', DB_MLT_PROD_D)]:
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            ph = ','.join(['?']*len(codigos))
            # a) Todo lo que existe para estos codigos
            cur.execute(f"""
                SELECT TIPOCOMPROBANTE, CUENTACORRIENTE, ANULADA,
                       CODIGOCLIENTE, CODIGOUSUARIO,
                       COUNT(*), SUM(TOTAL+IVA1+IVA2-PAGADO)
                FROM "CABEZACOMPROBANTES"
                WHERE CODIGOCLIENTE IN ({ph})
                GROUP BY TIPOCOMPROBANTE, CUENTACORRIENTE, ANULADA, CODIGOCLIENTE, CODIGOUSUARIO
                ORDER BY TIPOCOMPROBANTE
            """, tuple(codigos))
            rows_all = cur.fetchall()
            resultado[db_key+'_todos'] = [
                {"tipo": r[0], "ctacte": r[1], "anulada": r[2],
                 "codigocliente": r[3], "codigousuario": r[4],
                 "cant": r[5], "saldo_total": float(r[6] or 0)}
                for r in rows_all
            ]
            # b) Lo que pasaría con el filtro del detalle
            cur.execute(f"""
                SELECT TIPOCOMPROBANTE, NUMEROCOMPROBANTE, FECHACOMPROBANTE,
                       CODIGOCLIENTE, CODIGOUSUARIO,
                       TOTAL, IVA1, IVA2, PAGADO, CUENTACORRIENTE, COTIZACION, CODIGOMONEDA
                FROM "CABEZACOMPROBANTES"
                WHERE CODIGOCLIENTE IN ({ph})
                  AND ANULADA = '0'
                  AND TIPOCOMPROBANTE NOT IN ('RE','RI','INA')
                ORDER BY FECHACOMPROBANTE DESC
            """, tuple(codigos))
            rows_ctacte = cur.fetchall()
            resultado[db_key+'_sin_filtro_ctacte'] = [
                {"tipo": r[0], "num": r[1], "fecha": str(r[2]),
                 "cod_cli": r[3], "usuario": r[4],
                 "total": float(r[5] or 0), "iva1": float(r[6] or 0), "iva2": float(r[7] or 0),
                 "pagado": float(r[8] or 0), "ctacte": r[9],
                 "cotiz": float(r[10] or 1), "moneda": r[11]}
                for r in rows_ctacte
            ]
            c.close()
        except Exception as e:
            resultado[db_key+'_error'] = str(e)

    return resultado

@app.get("/debug/query_cta/{codigo}")
def debug_query_cta(codigo: str):
    """Corre _query_cta EXACTAMENTE como lo hace el endpoint de detalle y muestra resultado + errores."""
    DB_PROD_D     = DATABASE
    DB_MLT_PROD_D = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    resultado = {}

    # Lookup igual que el endpoint
    try:
        c = conn('WIN1252', db=DB_PROD_D)
        cur = c.cursor()
        cur.execute(
            'SELECT CODIGOCLIENTE, CODIGOPARTICULAR FROM "CLIENTES" WHERE CODIGOCLIENTE=? OR CODIGOPARTICULAR=?',
            (codigo, codigo)
        )
        cli = cur.fetchone()
        c.close()
        codigos_set = set()
        if cli:
            if cli[0] and str(cli[0]).strip(): codigos_set.add(str(cli[0]).strip())
            if cli[1] and str(cli[1]).strip(): codigos_set.add(str(cli[1]).strip())
        if not codigos_set:
            codigos_set.add(codigo)
        codigos = list(codigos_set)
        resultado['codigos'] = codigos
    except Exception as e:
        resultado['lookup_error'] = str(e)
        codigos = [codigo]

    # Correr _query_cta con captura de errores por fila
    for db_key, db_path in [('DB_PROD', DB_PROD_D), ('DB_MLT_PROD', DB_MLT_PROD_D)]:
        try:
            cambios = _get_cambios(db_path)
            resultado[db_key+'_cambios'] = cambios
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            ph = ', '.join(['?'] * len(codigos))
            params = list(codigos)
            _NC_T = "('NCA','NCB','NCCA','NCCB','NCE','NCCE','SIV','NDA','NDB','NDCA','NDCB')"
            sql = f"""
                SELECT FIRST 200 SKIP 0
                    TIPOCOMPROBANTE, NUMEROCOMPROBANTE, FECHACOMPROBANTE,
                    TOTAL, IVA1, IVA2, PAGADO, COTIZACION, CODIGOMONEDA,
                    FECHAVENCIMIENTO, CLASECOMPROBANTE
                FROM "CABEZACOMPROBANTES"
                WHERE CODIGOCLIENTE IN ({ph})
                  AND ANULADA = '0'
                  AND TIPOCOMPROBANTE NOT IN ('RE', 'RI', 'INA')
                  AND (CUENTACORRIENTE = '1' OR TIPOCOMPROBANTE IN {_NC_T})
                ORDER BY FECHAVENCIMIENTO ASC, FECHACOMPROBANTE ASC
            """
            cur.execute(sql, tuple(params))
            rows_raw = []
            row_errors = []
            while True:
                try:
                    r = cur.fetchone()
                    if r is None:
                        break
                    tipo=r[0]; num=r[1]; fecha=r[2]
                    total=float(r[3] or 0); iva1=float(r[4] or 0); iva2=float(r[5] or 0)
                    pagado=float(r[6] or 0); cotiz=float(r[7] or 1) or 1.0
                    moneda=str(r[8] or '').strip(); fvto=r[9]; clase=r[10]
                    neto=total+iva1+iva2; debe=neto-pagado
                    cambio=cambios.get(moneda,1.0) or 1.0
                    deuda=debe*cambio/cotiz
                    rows_raw.append({
                        "tipo":tipo,"num":str(num),"fecha":str(fecha),
                        "neto":round(neto,2),"pagado":round(pagado,2),
                        "deuda":round(deuda,2),"moneda":moneda,
                        "cotiz":cotiz,"cambio":cambio,"incluido":abs(deuda)>=0.01
                    })
                except Exception as row_e:
                    row_errors.append(str(row_e))
            c.close()
            resultado[db_key+'_filas'] = rows_raw
            resultado[db_key+'_filas_incluidas'] = sum(1 for r in rows_raw if r['incluido'])
            if row_errors:
                resultado[db_key+'_errores_fila'] = row_errors
        except Exception as e:
            resultado[db_key+'_error'] = str(e)

    return resultado

@app.get("/debug/esquema_docs")
def debug_esquema_docs():
    """Inspecciona columnas, NOT NULL, defaults y generators de pedidos/presupuestos."""
    try:
        c = conn('LATIN1')
        cur = c.cursor()
        resultado = {}

        # Columnas + NOT NULL + DEFAULT de cada tabla
        for tabla in ['CABEZAPEDIDOS', 'CUERPOPEDIDOS', 'CABEZAPRESUPUESTOS', 'CUERPOPRESUPUESTOS']:
            try:
                cur.execute(f"""
                    SELECT TRIM(rf.RDB$FIELD_NAME),
                           rf.RDB$NULL_FLAG,
                           TRIM(rf.RDB$DEFAULT_SOURCE)
                    FROM RDB$RELATION_FIELDS rf
                    WHERE rf.RDB$RELATION_NAME = '{tabla}'
                    ORDER BY rf.RDB$FIELD_POSITION
                """)
                resultado[f'cols_{tabla}'] = [
                    {"campo": r[0], "not_null": r[1]==1, "default": r[2]}
                    for r in cur.fetchall()
                ]
            except Exception as ex:
                resultado[f'cols_{tabla}_error'] = str(ex)

        # Muestra fila real de CABEZAPEDIDOS (para ver qué trae Flexxus)
        for tabla in ['CABEZAPEDIDOS', 'CABEZAPRESUPUESTOS']:
            try:
                cur.execute(f'SELECT FIRST 1 * FROM "{tabla}" ORDER BY NUMEROCOMPROBANTE DESC')
                cols = [d[0] for d in cur.description]
                row = cur.fetchone()
                resultado[f'muestra_{tabla}'] = dict(zip(cols, [str(v) if v is not None else None for v in row])) if row else None
            except Exception as ex:
                resultado[f'muestra_{tabla}_error'] = str(ex)

        # Generators relacionados con pedidos/presupuestos
        try:
            cur.execute("""
                SELECT TRIM(RDB$GENERATOR_NAME), RDB$GENERATOR_ID
                FROM RDB$GENERATORS
                WHERE RDB$SYSTEM_FLAG = 0
                ORDER BY RDB$GENERATOR_NAME
            """)
            todos = [{"nombre": r[0], "id": r[1]} for r in cur.fetchall()]
            resultado['generators'] = [g for g in todos if any(x in g['nombre'].upper() for x in ['PED','PRE','COMP','DOC','NUMERO','NP'])]
            resultado['generators_todos'] = todos
        except Exception as ex:
            resultado['generators_error'] = str(ex)

        # Triggers en CABEZAPEDIDOS (puede haber lógica de numeración)
        try:
            cur.execute("""
                SELECT TRIM(RDB$TRIGGER_NAME), RDB$TRIGGER_TYPE, TRIM(RDB$TRIGGER_SOURCE)
                FROM RDB$TRIGGERS
                WHERE RDB$RELATION_NAME IN ('CABEZAPEDIDOS','CABEZAPRESUPUESTOS')
                  AND RDB$SYSTEM_FLAG = 0
                ORDER BY RDB$TRIGGER_NAME
            """)
            resultado['triggers'] = [{"nombre": r[0], "tipo": r[1], "fuente": r[2]} for r in cur.fetchall()]
        except Exception as ex:
            resultado['triggers_error'] = str(ex)

        c.close()
        return resultado
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/debug/esquema_pedidos")
def debug_esquema_pedidos():
    """Columnas CABEZAPEDIDOS + valores OPERACION + muestra filas NP recientes"""
    res = {}
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        # Columnas de CABEZAPEDIDOS
        cur.execute("""
            SELECT TRIM(rf.RDB$FIELD_NAME), rf.RDB$FIELD_POSITION,
                   f.RDB$NULL_FLAG, f.RDB$DEFAULT_VALUE
            FROM RDB$RELATION_FIELDS rf
            JOIN RDB$FIELDS f ON rf.RDB$FIELD_SOURCE = f.RDB$FIELD_NAME
            WHERE rf.RDB$RELATION_NAME = 'CABEZAPEDIDOS'
            ORDER BY rf.RDB$FIELD_POSITION
        """)
        res['cols_cabezapedidos'] = [r[0] for r in cur.fetchall()]
        # Valores distintos de OPERACION en NPs
        cur.execute("SELECT DISTINCT OPERACION, COUNT(*) FROM \"CABEZAPEDIDOS\" WHERE TIPOCOMPROBANTE='NP' GROUP BY OPERACION ORDER BY OPERACION")
        res['operacion_values'] = [{'operacion': str(r[0]), 'count': r[1]} for r in cur.fetchall()]
        # Últimos 3 NPs con campos clave
        try:
            cur.execute("""
                SELECT FIRST 3 NUMEROCOMPROBANTE, OPERACION, FECHATERMINADA,
                       CODIGOUSUARIO, CODIGOUSUARIO2, CODIGOTECNICO, CLASECOMPROBANTE
                FROM "CABEZAPEDIDOS" WHERE TIPOCOMPROBANTE='NP' AND OPERACION IN ('2','4')
                ORDER BY NUMEROCOMPROBANTE DESC
            """)
            res['sample_np_terminados'] = [
                {'num': str(r[0]), 'operacion': str(r[1]), 'fechaterminada': str(r[2]),
                 'usuario': str(r[3]), 'usuario2': str(r[4]), 'tecnico': str(r[5]), 'clase': str(r[6])}
                for r in cur.fetchall()
            ]
        except Exception as e2:
            res['sample_np_extended_error'] = str(e2)
        c.close()
    except Exception as e:
        res['error'] = str(e)
    return res

@app.get("/debug/comparar_np/{num_a}/{num_b}")
def debug_comparar_np(num_a: str, num_b: str):
    """Compara campo a campo dos NPs en CABEZAPEDIDOS."""
    res = {}
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        cur.execute("SELECT * FROM \"CABEZAPEDIDOS\" WHERE NUMEROCOMPROBANTE IN (?,?) AND TIPOCOMPROBANTE='NP'", (num_a, num_b))
        cols = [d[0] for d in cur.description]
        rows = {str(r[cols.index('NUMEROCOMPROBANTE')]): dict(zip(cols, [str(v) if v is not None else None for v in r])) for r in cur.fetchall()}
        res['cabeza'] = rows

        # Diferencias entre los dos
        if num_a in rows and num_b in rows:
            diffs = {}
            for col in cols:
                va = rows[num_a].get(col)
                vb = rows[num_b].get(col)
                if va != vb:
                    diffs[col] = {num_a: va, num_b: vb}
            res['diferencias'] = diffs

        # Triggers sobre CABEZAPEDIDOS
        cur.execute("""
            SELECT TRIM(t.RDB$TRIGGER_NAME), t.RDB$TRIGGER_TYPE, t.RDB$TRIGGER_SOURCE
            FROM RDB$TRIGGERS t
            WHERE t.RDB$RELATION_NAME = 'CABEZAPEDIDOS' AND t.RDB$SYSTEM_FLAG = 0
            ORDER BY t.RDB$TRIGGER_SEQUENCE
        """)
        res['triggers'] = [{'nombre': r[0], 'tipo': r[1], 'source': (r[2] or '')[:300]} for r in cur.fetchall()]
        c.close()
    except Exception as e:
        res['error'] = str(e)
    return res

@app.get("/debug/operacion_np")
def debug_operacion_np():
    """Muestra OPERACION de los últimos 10 NPs y todos los valores distintos existentes."""
    res = {}
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        # Últimos 10 NPs: número, operacion, fechacomprobante
        cur.execute("""
            SELECT FIRST 10 NUMEROCOMPROBANTE, OPERACION, FECHACOMPROBANTE, CODIGOUSUARIO
            FROM "CABEZAPEDIDOS" WHERE TIPOCOMPROBANTE='NP'
            ORDER BY CAST(NUMEROCOMPROBANTE AS INTEGER) DESC
        """)
        res['ultimos_10'] = [{'num': str(r[0]), 'operacion': str(r[1]).strip(),
                               'fecha': str(r[2])[:10], 'usuario': str(r[3]).strip()}
                              for r in cur.fetchall()]
        # Todos los valores de OPERACION con conteo
        cur.execute("SELECT DISTINCT OPERACION, COUNT(*) FROM \"CABEZAPEDIDOS\" WHERE TIPOCOMPROBANTE='NP' GROUP BY OPERACION ORDER BY OPERACION")
        res['operacion_dist'] = [{'valor': repr(r[0]), 'count': r[1]} for r in cur.fetchall()]
        c.close()
    except Exception as e:
        res['error'] = str(e)
    return res

@app.get("/debug/comparar_pedidos")
def debug_comparar_pedidos():
    """Compara todos los campos de dos NPs: 100023473 (Flexxus) vs 100023547 (APP)"""
    res = {}
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        for num in ['100023473', '100023547']:
            cur.execute('SELECT * FROM "CABEZAPEDIDOS" WHERE TIPOCOMPROBANTE=\'NP\' AND NUMEROCOMPROBANTE=?', (num,))
            row = cur.fetchone()
            if row:
                cols = [d[0] for d in cur.description]
                res[num] = {cols[i]: str(row[i]) if row[i] is not None else None for i in range(len(cols))}
            else:
                res[num] = 'NO ENCONTRADO'
        c.close()
    except Exception as e:
        res['error'] = str(e)
    return res

@app.get("/debug/esquema_mlt")
def debug_esquema_mlt():
    """Columnas de CABEZACOMPROBANTES y CUERPOCOMPROBANTES en DB-MLT"""
    try:
        c = conn('LATIN1', db=DATABASE_MLT)
        cur = c.cursor()
        resultado = {}
        for tabla in ['CABEZACOMPROBANTES', 'CUERPOCOMPROBANTES']:
            cur.execute("""
                SELECT TRIM(f.RDB$FIELD_NAME),
                       TRIM(tp.RDB$TYPE_NAME),
                       f.RDB$NULL_FLAG,
                       f.RDB$DEFAULT_SOURCE
                FROM RDB$RELATION_FIELDS f
                JOIN RDB$FIELDS ff ON ff.RDB$FIELD_NAME = f.RDB$FIELD_SOURCE
                LEFT JOIN RDB$TYPES tp ON tp.RDB$TYPE = ff.RDB$FIELD_TYPE AND tp.RDB$FIELD_NAME = 'RDB$FIELD_TYPE'
                WHERE f.RDB$RELATION_NAME = ?
                ORDER BY f.RDB$FIELD_POSITION
            """, (tabla,))
            resultado[tabla] = [{"col": r[0], "tipo": r[1], "notnull": r[2], "default": r[3]} for r in cur.fetchall()]
        # Muestra de 1 fila para ver valores reales
        for tabla in ['CABEZACOMPROBANTES', 'CUERPOCOMPROBANTES']:
            try:
                cur.execute(f'SELECT FIRST 1 * FROM "{tabla}"')
                cols = [d[0] for d in cur.description]
                row = cur.fetchone()
                resultado[tabla+'_muestra'] = dict(zip(cols, [str(v) for v in row])) if row else {}
            except Exception as ex:
                resultado[tabla+'_muestra_err'] = str(ex)
        c.close()
        return resultado
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/mlt_cab/{numero}")
def debug_mlt_cab(numero: str):
    """Ver CABEZACOMPROBANTES de DATABASE_MLT para un número de comprobante."""
    try:
        c = conn('LATIN1', DATABASE_MLT)
        cur = c.cursor()
        cur.execute("SELECT FIRST 1 * FROM \"CABEZACOMPROBANTES\" "
                    "WHERE NUMEROCOMPROBANTE = ?", (numero,))
        row = cur.fetchone()
        cols = [d[0] for d in cur.description]
        c.close()
        if row:
            return dict(zip(cols, [str(v) for v in row]))
        return {"error": "no encontrado", "numero": numero}
    except Exception as e:
        return {"error": str(e)}


@app.get("/debug/mlt_tablas")
def debug_mlt_tablas():
    """Lista todas las tablas de DATABASE_MLT."""
    try:
        c = conn('LATIN1', DATABASE_MLT)
        cur = c.cursor()
        cur.execute("SELECT TRIM(RDB$RELATION_NAME) FROM RDB$RELATIONS "
                    "WHERE RDB$SYSTEM_FLAG = 0 ORDER BY RDB$RELATION_NAME")
        tablas = [r[0] for r in cur.fetchall()]
        c.close()
        return {"tablas": tablas}
    except Exception as e:
        return {"error": str(e)}


@app.get("/debug/quevendi_mlt/{cliente}")
def debug_quevendi_mlt(cliente: str):
    """Diagnostica por qué que_vendi no trae resultados de DATABASE_MLT para un cliente."""
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    result = {"cliente": cliente, "pasos": []}

    # 1. Lookup en DB-MLT-Microbell.gdb (prod MLT tiene CLIENTES, DATABASE_MLT no)
    cod_mlt = cliente
    try:
        c = conn('WIN1252', DB_MLT_PROD)
        cur = c.cursor()
        cur.execute('SELECT CODIGOCLIENTE, CODIGOPARTICULAR FROM "CLIENTES" '
                    'WHERE CODIGOCLIENTE = ? OR CODIGOPARTICULAR = ?', (cliente, cliente))
        row = cur.fetchone()
        c.close()
        result["pasos"].append({"paso": "lookup_clientes_mlt_prod", "encontrado": row is not None,
                                "fila": [str(x) for x in row] if row else None})
        if row:
            cod_mlt = str(row[0]).strip()
    except Exception as e:
        result["pasos"].append({"paso": "lookup_clientes_mlt_prod", "error": str(e)})

    result["codigocliente_mlt"] = cod_mlt

    # 2. Verificar columnas de CUERPOCOMPROBANTES en DATABASE_MLT
    try:
        c = conn('LATIN1', DATABASE_MLT)
        cur = c.cursor()
        cur.execute("SELECT FIRST 1 * FROM \"CUERPOCOMPROBANTES\"")
        cols = [d[0] for d in cur.description]
        result["cuerpocomprobantes_cols"] = cols
        result["tiene_codigoparticular"] = "CODIGOPARTICULAR" in cols

        # 3. Buscar comprobantes del cliente
        cur.execute(
            'SELECT FIRST 5 cb.TIPOCOMPROBANTE, cb.NUMEROCOMPROBANTE, cb.FECHACOMPROBANTE '
            'FROM "CABEZACOMPROBANTES" cb '
            "WHERE cb.CODIGOCLIENTE = ? AND cb.ANULADA = '0' "
            "AND cb.TIPOCOMPROBANTE IN ('FA','FB','FE','FCA','FCB','FCE','FCCA','FCCB','FCCE','NCA','NCB','NCCA','NCCB') "
            'ORDER BY cb.FECHACOMPROBANTE DESC',
            (cod_mlt,)
        )
        rows_cab = cur.fetchall()
        result["cabeza_count"] = len(rows_cab)
        result["cabeza_sample"] = [[str(x) for x in r] for r in rows_cab]

        # 4. Si tiene CODIGOPARTICULAR: probar el SQL completo
        if "CODIGOPARTICULAR" in cols and rows_cab:
            try:
                cur.execute(
                    'SELECT FIRST 3 '
                    'COALESCE(NULLIF(TRIM(cu.CODIGOPARTICULAR),\'\'), TRIM(cu.CODIGOARTICULO)) AS COD_ART, '
                    'cu.DESCRIPCION, cb.TIPOCOMPROBANTE, cb.NUMEROCOMPROBANTE, '
                    'CAST(cu.CANTIDAD AS DOUBLE PRECISION) '
                    'FROM "CUERPOCOMPROBANTES" cu '
                    'JOIN "CABEZACOMPROBANTES" cb '
                    '  ON cb.TIPOCOMPROBANTE = cu.TIPOCOMPROBANTE '
                    ' AND cb.NUMEROCOMPROBANTE = cu.NUMEROCOMPROBANTE '
                    "WHERE cb.CODIGOCLIENTE = ? AND cb.ANULADA = '0' "
                    "AND cb.TIPOCOMPROBANTE IN ('FA','FB','FE','FCA','FCB','FCE','FCCA','FCCB','FCCE','NCA','NCB','NCCA','NCCB')",
                    (cod_mlt,)
                )
                rows_cuerpo = cur.fetchall()
                result["cuerpo_count"] = len(rows_cuerpo)
                result["cuerpo_sample"] = [[str(x) for x in r] for r in rows_cuerpo]
            except Exception as e2:
                result["cuerpo_error"] = str(e2)
        elif "CODIGOPARTICULAR" not in cols:
            result["nota"] = "CODIGOPARTICULAR no existe en CUERPOCOMPROBANTES de DATABASE_MLT"
        c.close()
    except Exception as e:
        result["pasos"].append({"paso": "query_mlt", "error": str(e)})

    return result


@app.get("/debug/quevendi_prod/{cliente}")
def debug_quevendi_prod(cliente: str):
    """Busca FA 100001668 y 100001684 en las 4 BDs y diagnostica que_vendi para un cliente."""
    DB_PROD     = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    result = {"cliente": cliente}

    # 1. Resolver codigos
    codigos = set([cliente])
    for db_label, db_path in [("DATABASE", DATABASE), ("DB_PROD", DB_PROD)]:
        try:
            c = conn('WIN1252', db_path)
            cur = c.cursor()
            cur.execute('SELECT CODIGOCLIENTE, CODIGOPARTICULAR FROM "CLIENTES" '
                        'WHERE CODIGOCLIENTE = ? OR CODIGOPARTICULAR = ?', (cliente, cliente))
            row = cur.fetchone()
            c.close()
            if row:
                for v in row:
                    if v is not None and str(v).strip():
                        codigos.add(str(v).strip())
            result[f"lookup_{db_label}"] = [str(x) for x in row] if row else None
        except Exception as e:
            result[f"lookup_{db_label}_error"] = str(e)

    codigos = list(codigos)
    result["codigos"] = codigos
    ph = ','.join('?' * len(codigos))

    # 2. Buscar FA 100001668 y 100001684 en las 4 BDs
    for db_label, db_path in [("DATABASE", DATABASE), ("DATABASE_MLT", DATABASE_MLT),
                               ("DB_PROD", DB_PROD), ("DB_MLT_PROD", DB_MLT_PROD)]:
        try:
            c = conn('LATIN1', db=db_path)
            cur = c.cursor()
            # Buscar las facturas objetivo
            cur.execute(
                'SELECT TIPOCOMPROBANTE, NUMEROCOMPROBANTE, CODIGOCLIENTE, FECHACOMPROBANTE '
                'FROM "CABEZACOMPROBANTES" WHERE NUMEROCOMPROBANTE IN (100001668, 100001684)'
            )
            rows_t = cur.fetchall()
            result[f"facturas_target_{db_label}"] = [[str(x) for x in r] for r in rows_t]

            # Contar comprobantes del cliente
            cur.execute(
                f'SELECT COUNT(*) FROM "CABEZACOMPROBANTES" '
                f"WHERE CODIGOCLIENTE IN ({ph}) AND ANULADA = '0' "
                f"AND TIPOCOMPROBANTE IN ('FA','FB','FE','FCA','FCB','FCE','FCCA','FCCB','FCCE','NCA','NCB','NCCA','NCCB')",
                codigos
            )
            result[f"count_{db_label}"] = cur.fetchone()[0]

            # Contar en CUERPOCOMPROBANTES (para verificar el JOIN)
            cur.execute(
                f'SELECT COUNT(*) FROM "CUERPOCOMPROBANTES" cu '
                f'JOIN "CABEZACOMPROBANTES" cb '
                f'  ON cb.TIPOCOMPROBANTE = cu.TIPOCOMPROBANTE '
                f' AND cb.NUMEROCOMPROBANTE = cu.NUMEROCOMPROBANTE '
                f"WHERE cb.CODIGOCLIENTE IN ({ph}) AND cb.ANULADA = '0' "
                f"AND cb.TIPOCOMPROBANTE IN ('FA','FB','FE','FCA','FCB','FCE','FCCA','FCCB','FCCE','NCA','NCB','NCCA','NCCB')",
                codigos
            )
            result[f"count_con_join_{db_label}"] = cur.fetchone()[0]
            c.close()
        except Exception as e:
            result[f"error_{db_label}"] = str(e)

    return result


@app.get("/debug/qv_errors")
def debug_qv_errors():
    """Muestra errores y conteo de filas por BD de la última llamada a que_vendi."""
    return {"errors": _QV_LAST_ERRORS, "counts": _QV_LAST_COUNTS}


@app.get("/debug/flexxus_deudas_schema")
def debug_flexxus_deudas_schema():
    """Busca tablas relevantes (saldo/deuda/vendedor/cuenta) y columnas de CABEZACOMPROBANTES en prod."""
    DB_PROD     = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    result = {}
    keywords = ['DEUDA', 'SALDO', 'VENDEDOR', 'CUENTA', 'COBRANZA', 'VENC']
    for label, db_path in [('DB_PROD', DB_PROD), ('DB_MLT_PROD', DB_MLT_PROD)]:
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            # Todas las tablas
            cur.execute("SELECT TRIM(RDB$RELATION_NAME) FROM RDB$RELATIONS "
                        "WHERE RDB$SYSTEM_FLAG = 0 ORDER BY RDB$RELATION_NAME")
            todas = [r[0] for r in cur.fetchall()]
            relevantes = [t for t in todas if any(k in t.upper() for k in keywords)]
            result[f'{label}_tablas_relevantes'] = relevantes
            # Columnas de CABEZACOMPROBANTES
            cur.execute("SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS "
                        "WHERE TRIM(RDB$RELATION_NAME) = 'CABEZACOMPROBANTES' "
                        "ORDER BY RDB$FIELD_POSITION")
            cols_cab = [r[0] for r in cur.fetchall()]
            result[f'{label}_cabeza_cols'] = cols_cab
            # Columnas de CUERPOCOMPROBANTES (solo las 20 primeras)
            cur.execute("SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS "
                        "WHERE TRIM(RDB$RELATION_NAME) = 'CUERPOCOMPROBANTES' "
                        "ORDER BY RDB$FIELD_POSITION")
            cols_cue = [r[0] for r in cur.fetchall()]
            result[f'{label}_cuerpo_cols'] = cols_cue
            c.close()
        except Exception as e:
            result[f'{label}_error'] = str(e)
    return result


@app.get("/debug/cliente_vendedor/{nombre}")
def debug_cliente_vendedor(nombre: str):
    """Muestra CODIGOVENDEDOR de un cliente en DB_PROD y sus comprobantes pendientes."""
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    result = {}
    try:
        c = conn('WIN1252', DB_PROD)
        cur = c.cursor()
        cur.execute(
            "SELECT CODIGOCLIENTE, CODIGOPARTICULAR, RAZONSOCIAL, CODIGOVENDEDOR, ACTIVO "
            'FROM "CLIENTES" WHERE UPPER(RAZONSOCIAL) CONTAINING UPPER(?)',
            (nombre,)
        )
        rows = cur.fetchall()
        result['clientes'] = [{'cod': r[0], 'part': r[1], 'razon': r[2], 'vendedor': r[3], 'activo': r[4]} for r in rows]
        # Para cada cliente encontrado, contar sus comprobantes pendientes
        for cli in result['clientes']:
            cod = (cli['cod'] or '').strip()
            if not cod: continue
            cur.execute(
                "SELECT COUNT(*) FROM \"CABEZACOMPROBANTES\" "
                "WHERE CODIGOCLIENTE=? AND CUENTACORRIENTE='1' AND ANULADA='0' "
                "AND TIPOCOMPROBANTE NOT IN ('RE','RI','INA')",
                (cod,)
            )
            cli['comprobantes_cta'] = cur.fetchone()[0]
        c.close()
    except Exception as e:
        result['error'] = str(e)
    return result


@app.get("/debug/cliente_en_mlt/{nombre}")
def debug_cliente_en_mlt(nombre: str):
    """Diagnostica la presencia de un cliente en DB_MLT_PROD buscando por nombre parcial en CABEZACOMPROBANTES."""
    DB_PROD     = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    result = {}

    # 1. Buscar en DB_PROD CLIENTES
    try:
        c = conn('WIN1252', DB_PROD)
        cur = c.cursor()
        cur.execute(
            "SELECT CODIGOCLIENTE, CODIGOPARTICULAR, RAZONSOCIAL, CODIGOVENDEDOR "
            'FROM "CLIENTES" WHERE UPPER(RAZONSOCIAL) CONTAINING UPPER(?) AND ACTIVO=\'1\'',
            (nombre,)
        )
        rows = cur.fetchall()
        c.close()
        result['DB_PROD_clientes'] = [
            {'cod': r[0], 'part': r[1], 'razon': r[2], 'vendedor': r[3]} for r in rows
        ]
    except Exception as e:
        result['DB_PROD_error'] = str(e)

    # 2. Buscar en DB_PROD CABEZACOMPROBANTES por RAZONSOCIAL (para verificar si la FA está ahí)
    try:
        c = conn('WIN1252', DB_PROD)
        cur = c.cursor()
        cur.execute(
            "SELECT FIRST 10 CODIGOCLIENTE, TIPOCOMPROBANTE, NUMEROCOMPROBANTE, RAZONSOCIAL, FECHACOMPROBANTE, CUENTACORRIENTE, ANULADA "
            'FROM "CABEZACOMPROBANTES" WHERE UPPER(RAZONSOCIAL) CONTAINING UPPER(?) '
            "AND CUENTACORRIENTE='1' AND ANULADA='0' ORDER BY FECHACOMPROBANTE DESC",
            (nombre,)
        )
        rows = cur.fetchall()
        c.close()
        result['DB_PROD_cabeza_por_razon'] = [
            {'cod': r[0], 'tipo': r[1], 'num': r[2], 'razon': r[3], 'fecha': str(r[4]), 'cta': r[5], 'anul': r[6]} for r in rows
        ]
    except Exception as e:
        result['DB_PROD_cabeza_error'] = str(e)

    # 3. Buscar en DB_MLT_PROD CABEZACOMPROBANTES por RAZONSOCIAL
    try:
        c = conn('WIN1252', DB_MLT_PROD)
        cur = c.cursor()
        cur.execute(
            "SELECT FIRST 5 CODIGOCLIENTE, TIPOCOMPROBANTE, NUMEROCOMPROBANTE, RAZONSOCIAL, FECHACOMPROBANTE "
            'FROM "CABEZACOMPROBANTES" WHERE UPPER(RAZONSOCIAL) CONTAINING UPPER(?)',
            (nombre,)
        )
        rows = cur.fetchall()
        c.close()
        result['DB_MLT_PROD_cabeza_por_razon'] = [
            {'cod': r[0], 'tipo': r[1], 'num': r[2], 'razon': r[3], 'fecha': str(r[4])} for r in rows
        ]
    except Exception as e:
        result['DB_MLT_PROD_cabeza_error'] = str(e)

    # 3. Si encontramos clientes en DB_PROD, buscar su CODIGOPARTICULAR en CUERPOCOMPROBANTES de DB_MLT_PROD
    for cli in result.get('DB_PROD_clientes', []):
        part = cli.get('part', '')
        cod  = cli.get('cod', '')
        key  = f"lookup_part_{part or cod}"
        if part:
            try:
                c = conn('LATIN1', DB_MLT_PROD)
                cur = c.cursor()
                cur.execute(
                    'SELECT FIRST 5 DISTINCT CODIGOCLIENTE FROM "CUERPOCOMPROBANTES" WHERE CODIGOPARTICULAR = ?',
                    (part,)
                )
                rows = cur.fetchall()
                c.close()
                result[key] = {'codigoscliente_en_mlt': [str(r[0]) for r in rows]}
            except Exception as e:
                result[key] = {'error': str(e)}

    return result


@app.get("/debug/vista_deuda_directo/{codigocliente}")
def debug_vista_deuda_directo(codigocliente: str):
    """Consulta directa de VISTADEUDACLIENTES + CABEZACOMPROBANTES para un CODIGOCLIENTE."""
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    result = {}
    # 1. VISTADEUDACLIENTES directo
    for charset in ['LATIN1', 'WIN1252']:
        try:
            c = conn(charset, DB_PROD)
            cur = c.cursor()
            cur.execute(
                'SELECT TIPOCOMPROBANTE, NUMEROCOMPROBANTE, PAGADO, NETO, DEBE, TOTALACTUALIZADO, CODIGOMONEDA, COTIZACION '
                'FROM "VISTADEUDACLIENTES" WHERE CODIGOCLIENTE = ?',
                (codigocliente,)
            )
            rows = cur.fetchall()
            c.close()
            result[f'vista_{charset}'] = [
                {'tipo': r[0], 'num': r[1], 'pagado': r[2], 'neto': r[3],
                 'debe': r[4], 'total_act': r[5], 'moneda': r[6], 'cotiz': r[7]}
                for r in rows
            ]
            result[f'vista_{charset}_suma'] = sum(float(r[5] or 0) for r in rows)
        except Exception as e:
            result[f'vista_{charset}_error'] = str(e)
    # 2. CABEZACOMPROBANTES directo (sin MONEDAS) para ver registros raw
    try:
        c = conn('WIN1252', DB_PROD)
        cur = c.cursor()
        cur.execute(
            "SELECT TIPOCOMPROBANTE, NUMEROCOMPROBANTE, TOTAL, IVA1, IVA2, PAGADO, CODIGOMONEDA, COTIZACION, ANULADA, CUENTACORRIENTE "
            'FROM "CABEZACOMPROBANTES" WHERE CODIGOCLIENTE = ? '
            "AND CUENTACORRIENTE='1' AND ANULADA='0' "
            "AND TIPOCOMPROBANTE NOT IN ('RE','RI','INA') "
            "AND ABS(CAST(TOTAL AS DOUBLE PRECISION)+CAST(IVA1 AS DOUBLE PRECISION)+CAST(IVA2 AS DOUBLE PRECISION)-CAST(PAGADO AS DOUBLE PRECISION)) >= 0.01 "
            "ORDER BY FECHACOMPROBANTE DESC",
            (codigocliente,)
        )
        rows = cur.fetchall()
        c.close()
        result['cabeza_raw'] = [
            {'tipo': r[0], 'num': r[1], 'total': r[2], 'iva1': r[3], 'iva2': r[4],
             'pagado': r[5], 'moneda': r[6], 'cotiz': r[7], 'anul': r[8], 'cta': r[9]}
            for r in rows
        ]
        result['cabeza_raw_suma_simple'] = sum(
            float((r[2] or 0)) + float((r[3] or 0)) + float((r[4] or 0)) - float((r[5] or 0))
            for r in rows
        )
    except Exception as e:
        result['cabeza_raw_error'] = str(e)
    return result


@app.get("/debug/vista_deuda_clientes")
def debug_vista_deuda_clientes():
    """Inspecciona VISTADEUDACLIENTES: columnas, definición SQL y muestra de datos."""
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    result = {}
    try:
        c = conn('WIN1252', db=DB_PROD)
        cur = c.cursor()
        # Columnas de la vista
        cur.execute("SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS "
                    "WHERE TRIM(RDB$RELATION_NAME) = 'VISTADEUDACLIENTES' "
                    "ORDER BY RDB$FIELD_POSITION")
        cols = [r[0] for r in cur.fetchall()]
        result['columnas'] = cols
        # Definición SQL de la vista
        cur.execute("SELECT RDB$VIEW_SOURCE FROM RDB$RELATIONS "
                    "WHERE TRIM(RDB$RELATION_NAME) = 'VISTADEUDACLIENTES'")
        row = cur.fetchone()
        result['sql_vista'] = str(row[0]) if row else None
        # Muestra de 3 filas
        try:
            cur.execute('SELECT FIRST 3 * FROM "VISTADEUDACLIENTES"')
            sample_cols = [d[0] for d in cur.description]
            sample_rows = cur.fetchall()
            result['muestra'] = [dict(zip(sample_cols, [str(v) if v is not None else None for v in r])) for r in sample_rows]
        except Exception as e2:
            result['muestra_error'] = str(e2)
        # Total deuda en la vista (para comparar con Flexxus)
        try:
            deuda_col = next((col for col in cols if 'DEUDA' in col.upper() or 'SALDO' in col.upper()), None)
            if deuda_col:
                cur.execute(f'SELECT COUNT(*), SUM(CAST("{deuda_col}" AS DOUBLE PRECISION)) FROM "VISTADEUDACLIENTES"')
                r = cur.fetchone()
                result[f'total_filas'] = r[0]
                result[f'suma_{deuda_col}'] = float(r[1] or 0)
        except Exception as e3:
            result['suma_error'] = str(e3)
        c.close()
    except Exception as e:
        result['error'] = str(e)
    return result


@app.get("/debug/cta_vendedor/{vendedor}")
def debug_cta_vendedor(vendedor: str):
    """Diagnóstico: suma deuda directo desde CABEZACOMPROBANTES con CODIGOVENDEDOR si existe."""
    DB_PROD     = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    result = {}
    for label, db_path in [('DB_PROD', DB_PROD), ('DB_MLT_PROD', DB_MLT_PROD)]:
        try:
            c = conn('LATIN1', db=db_path)
            cur = c.cursor()
            # Verificar si CABEZACOMPROBANTES tiene CODIGOVENDEDOR
            cur.execute("SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS "
                        "WHERE TRIM(RDB$RELATION_NAME) = 'CABEZACOMPROBANTES' "
                        "AND TRIM(RDB$FIELD_NAME) CONTAINING 'VENDEDOR'")
            cols_vend = [r[0] for r in cur.fetchall()]
            result[f'{label}_cols_vendedor_en_cabeza'] = cols_vend
            if cols_vend:
                col = cols_vend[0]
                cur.execute(f"""
                    SELECT COUNT(*),
                           SUM(CAST(TOTAL AS DOUBLE PRECISION) + CAST(IVA1 AS DOUBLE PRECISION) + CAST(IVA2 AS DOUBLE PRECISION)),
                           SUM(CAST(TOTAL AS DOUBLE PRECISION) + CAST(IVA1 AS DOUBLE PRECISION) + CAST(IVA2 AS DOUBLE PRECISION) - CAST(PAGADO AS DOUBLE PRECISION))
                    FROM "CABEZACOMPROBANTES"
                    WHERE UPPER({col}) = UPPER(?)
                      AND CUENTACORRIENTE = '1'
                      AND ANULADA = '0'
                      AND TIPOCOMPROBANTE IN ('FA','FB','FE','FCA','FCB','DI','SIV','NCA','NCB','NDA','NDB','NCAE','NDAE')
                      AND ABS(CAST(TOTAL AS DOUBLE PRECISION) + CAST(IVA1 AS DOUBLE PRECISION) + CAST(IVA2 AS DOUBLE PRECISION) - CAST(PAGADO AS DOUBLE PRECISION)) > 0.01
                """, (vendedor.upper(),))
                row = cur.fetchone()
                result[f'{label}_directo_count'] = row[0]
                result[f'{label}_directo_total_bruto'] = float(row[1] or 0)
                result[f'{label}_directo_deuda'] = float(row[2] or 0)
            c.close()
        except Exception as e:
            result[f'{label}_error'] = str(e)
    return result


@app.get("/debug/camping_query/{cod}")
def debug_camping_query(cod: str):
    """Debug exacto de _query_cta para un cliente específico (ej: 348 = CAMPING LA PLATA)."""
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    out = {"cod": cod, "db_prod": {}, "db_mlt": {}}
    for db_key, db_path in [("db_prod", DB_PROD), ("db_mlt", DB_MLT_PROD)]:
        res = {"charset_test": {}, "fetchone_loop": [], "errors": []}
        # Test 1: conectar y contar sin filtros extra
        for charset in ['WIN1252', 'LATIN1']:
            try:
                c = conn(charset, db=db_path)
                cur = c.cursor()
                cur.execute('SELECT COUNT(*) FROM "CABEZACOMPROBANTES" WHERE CODIGOCLIENTE = ?', (cod,))
                row = cur.fetchone()
                c.close()
                res["charset_test"][charset] = {"count_total": row[0] if row else None}
            except Exception as e:
                res["charset_test"][charset] = {"error": str(e)}

        # Test 2: SIN FILTROS — ver todos los registros para detectar debe>0 excluidos
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            cur.execute("""
                SELECT FIRST 500 SKIP 0
                    TIPOCOMPROBANTE, NUMEROCOMPROBANTE, FECHACOMPROBANTE,
                    TOTAL, IVA1, IVA2, PAGADO, COTIZACION, CODIGOMONEDA,
                    FECHAVENCIMIENTO, CLASECOMPROBANTE, CUENTACORRIENTE, ANULADA
                FROM "CABEZACOMPROBANTES"
                WHERE CODIGOCLIENTE = ?
                ORDER BY FECHACOMPROBANTE ASC
            """, (cod,))
            all_rows = []
            while True:
                try:
                    r = cur.fetchone()
                    if r is None: break
                    total=float(r[3] or 0); iva1=float(r[4] or 0); iva2=float(r[5] or 0)
                    pagado=float(r[6] or 0); neto=total+iva1+iva2; debe=neto-pagado
                    if abs(debe) >= 0.01:
                        all_rows.append({"tipo":str(r[0]),"num":str(r[1]),"fecha":str(r[2]),
                            "neto":round(neto,2),"pagado":round(pagado,2),"debe":round(debe,2),
                            "cotiz":float(r[7] or 1),"moneda":str(r[8] or '').strip(),
                            "ctacte":str(r[11]),"anulada":str(r[12])})
                except Exception as e:
                    all_rows.append({"row_error": str(e)})
                    break
            res["sin_filtros_con_saldo"] = all_rows
            c.close()
        except Exception as exec_e:
            res["sin_filtros_error"] = str(exec_e)

        # Test 3: fetchone loop con WIN1252 — misma query que _query_cta
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            cur.execute("""
                SELECT FIRST 500 SKIP 0
                    TIPOCOMPROBANTE, NUMEROCOMPROBANTE, FECHACOMPROBANTE,
                    TOTAL, IVA1, IVA2, PAGADO, COTIZACION, CODIGOMONEDA,
                    FECHAVENCIMIENTO, CLASECOMPROBANTE
                FROM "CABEZACOMPROBANTES"
                WHERE CODIGOCLIENTE = ?
                  AND CUENTACORRIENTE = '1'
                  AND ANULADA = '0'
                  AND TIPOCOMPROBANTE NOT IN ('RE', 'RI', 'INA')
                ORDER BY FECHAVENCIMIENTO ASC, FECHACOMPROBANTE ASC
            """, (cod,))
            n = 0
            while True:
                try:
                    r = cur.fetchone()
                    if r is None:
                        break
                    n += 1
                    # intento conversión
                    try:
                        total = float(r[3] or 0); iva1 = float(r[4] or 0); iva2 = float(r[5] or 0)
                        pagado = float(r[6] or 0); cotiz = float(r[7] or 1) or 1.0
                        neto = total + iva1 + iva2; debe = neto - pagado
                        res["fetchone_loop"].append({
                            "n": n, "tipo": str(r[0]), "num": str(r[1]),
                            "fecha": str(r[2]), "neto": neto, "debe": debe,
                            "cotiz": cotiz, "moneda": str(r[8] or '').strip()
                        })
                    except Exception as conv_e:
                        res["fetchone_loop"].append({"n": n, "conv_error": str(conv_e), "raw": [str(x) for x in r]})
                except Exception as fetch_e:
                    res["errors"].append({"at_row": n, "fetch_error": str(fetch_e)})
                    break
            res["total_fetched"] = n
            c.close()
        except Exception as exec_e:
            res["execute_error"] = str(exec_e)

        out[db_key] = res
    return out

@app.get("/debug/deuda_por_bd/{vendedor}")
def debug_deuda_por_bd(vendedor: str):
    """Muestra por cada cliente del vendedor: resultado de DB_PROD vs DB_MLT_PROD por separado."""
    DB_PROD     = 'c:/flexxus/DB/DB-Microbell.gdb'
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    try:
        c = conn('WIN1252', db=DB_PROD)
        cur = c.cursor()
        cur.execute(
            'SELECT CODIGOCLIENTE, RAZONSOCIAL FROM "CLIENTES" '
            'WHERE ACTIVO=? AND UPPER(CODIGOVENDEDOR)=? ORDER BY RAZONSOCIAL',
            ('1', vendedor.upper())
        )
        clientes = cur.fetchall()
        c.close()
    except Exception as e:
        return {"error_clientes": str(e)}

    resultado = []
    for cod_raw, razon in clientes:
        cod = (cod_raw or '').strip()
        if not cod: continue
        d_prod = d_mlt = 0.0
        err_prod = err_mlt = None
        try:
            rows = _query_cta(DB_PROD, [cod], 500, 0)
            d_prod = sum(float(r[5] or 0) for r in rows)
        except Exception as e:
            err_prod = str(e)
        try:
            rows = _query_cta(DB_MLT_PROD, [cod], 500, 0)
            d_mlt = sum(float(r[5] or 0) for r in rows)
        except Exception as e:
            err_mlt = str(e)
        total = d_prod + d_mlt
        if abs(total) >= 0.01 or err_prod or err_mlt:
            entry = {
                "cod": cod, "razon": razon,
                "db_prod": round(d_prod, 2), "db_mlt": round(d_mlt, 2),
                "total": round(total, 2)
            }
            if err_prod: entry["err_prod"] = err_prod
            if err_mlt:  entry["err_mlt"]  = err_mlt
            resultado.append(entry)

    total_general = sum(r["total"] for r in resultado)
    return {"total_general": round(total_general, 2), "clientes": len(resultado), "detalle": resultado}

@app.get("/debug/listar_dbs")
def listar_dbs():
    import glob
    archivos = glob.glob('c:/flexxus/**/*.gdb', recursive=True) + \
               glob.glob('c:/flexxus/**/*.FDB', recursive=True)
    return {"archivos": sorted(archivos)}

@app.get("/debug/generators")
def debug_generators():
    """Lista generators y su valor actual en ambas BDs"""
    resultado = {}
    for nombre, db_path in [('oficial', DATABASE), ('sw', DATABASE_MLT)]:
        try:
            c = conn('LATIN1', db=db_path)
            cur = c.cursor()
            cur.execute(
                "SELECT TRIM(RDB$GENERATOR_NAME), RDB$GENERATOR_ID "
                "FROM RDB$GENERATORS WHERE RDB$SYSTEM_FLAG = 0 "
                "ORDER BY RDB$GENERATOR_NAME"
            )
            gens = cur.fetchall()
            vals = {}
            for g in gens:
                try:
                    cur.execute(f'SELECT GEN_ID("{g[0]}", 0) FROM RDB$DATABASE')
                    vals[g[0]] = cur.fetchone()[0]
                except Exception:
                    vals[g[0]] = None
            resultado[nombre] = vals
            c.close()
        except Exception as e:
            resultado[nombre] = {"error": str(e)}
    return resultado

@app.get("/debug/tablas_mlt")
def debug_tablas_mlt():
    """Lista todas las tablas de usuario en DB-MLT-Microbell.gdb"""
    try:
        c = conn('LATIN1', db=DATABASE_MLT)
        cur = c.cursor()
        cur.execute(
            "SELECT TRIM(RDB$RELATION_NAME) FROM RDB$RELATIONS "
            "WHERE RDB$SYSTEM_FLAG = 0 AND RDB$VIEW_BLR IS NULL "
            "ORDER BY RDB$RELATION_NAME"
        )
        tablas = [r[0] for r in cur.fetchall()]
        c.close()
        return {"tablas": tablas}
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/esquema_clientes")
def debug_esquema_clientes():
    """Muestra todas las columnas de la tabla CLIENTES en DB-Prueba.gdb"""
    try:
        c = conn()
        cur = c.cursor()
        cur.execute(
            "SELECT TRIM(RDB$FIELD_NAME), RDB$FIELD_POSITION "
            "FROM RDB$RELATION_FIELDS "
            "WHERE TRIM(RDB$RELATION_NAME) = 'CLIENTES' "
            "ORDER BY RDB$FIELD_POSITION"
        )
        cols = [{"pos": r[1], "nombre": r[0]} for r in cur.fetchall()]
        c.close()
        return {"columnas": cols}
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/muestra_clientes")
def debug_muestra_clientes(vendedor: str = "KRAFFT"):
    """Muestra 3 clientes del vendedor para ver valores de columnas IVA"""
    try:
        c = conn()
        cur = c.cursor()
        cur.execute(
            'SELECT FIRST 3 CODIGOCLIENTE, RAZONSOCIAL, CODIGOPARTICULAR '
            'FROM "CLIENTES" WHERE ACTIVO = ? AND CODIGOVENDEDOR = ? ORDER BY RAZONSOCIAL',
            ('1', vendedor.upper())
        )
        rows = cur.fetchall()
        c.close()
        return [{"codigo": r[0], "razonsocial": r[1], "codigoparticular": r[2]} for r in rows]
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/tablas_of")
def debug_tablas_of():
    """Lista todas las tablas de DB-Prueba.gdb"""
    try:
        c = conn('LATIN1', db=DATABASE)
        cur = c.cursor()
        cur.execute("SELECT RDB$RELATION_NAME FROM RDB$RELATIONS WHERE RDB$SYSTEM_FLAG=0 ORDER BY RDB$RELATION_NAME")
        rows = cur.fetchall()
        c.close()
        return [r[0].strip() for r in rows]
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/contadores_of")
def debug_contadores_of():
    """Muestra todos los contadores de la BD oficial (DB-Prueba.gdb)"""
    try:
        c = conn('LATIN1', db=DATABASE)
        cur = c.cursor()
        cur.execute('SELECT CODIGOCONTADOR, DESCRIPCION, VALOR FROM "CONTADORES" ORDER BY CODIGOCONTADOR')
        rows = cur.fetchall()
        c.close()
        return [{"codigo": r[0], "descripcion": (r[1] or "").strip(), "valor": r[2]} for r in rows]
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/contadores_sw")
def debug_contadores_sw():
    """Muestra todos los contadores de la BD SW (DB-MLT-Prueba.gdb)"""
    try:
        c = conn('LATIN1', db=DATABASE_MLT)
        cur = c.cursor()
        cur.execute('SELECT CODIGOCONTADOR, DESCRIPCION, VALOR FROM "CONTADORES" ORDER BY CODIGOCONTADOR')
        rows = cur.fetchall()
        c.close()
        return [{"codigo": r[0], "descripcion": (r[1] or "").strip(), "valor": r[2]} for r in rows]
    except Exception as e:
        return {"error": str(e)}


# ============================================================================
# CRÉDITO INTERNO POR ESCALÓN — función real (2026-07-31)
#
# ESTADO: implementada para revisión. NO está conectada a ningún endpoint que la
# dispare desde producción todavía — hay que llamarla explícitamente desde Python
# (por ejemplo, desde el debug endpoint de prueba controlada que armemos después,
# una vez que Eduardo dé el visto bueno para probar contra un cliente real/de prueba).
#
# Qué hace, en una sola transacción Firebird (todo o nada — si algo falla, rollback
# completo, no queda ni el CI ni el asiento a medias):
#   1) Llama al procedimiento OFICIAL de Flexxus FMA_GENERARCREDITOINTERNO (el mismo
#      que usa la app de escritorio) para crear el comprobante CI — así reusamos su
#      numeración correcta (CONTADORES), sus validaciones y todos los datos del
#      cliente, sin reinventar el INSERT.
#   2) Corrige el texto que el procedimiento oficial hardcodea ('CREDITO POR PAGO')
#      por una descripción real del origen promocional (oferta/escalón).
#   3) Genera la cabecera del asiento contable vía FMA_CABEZAASIENTOS (procedimiento
#      oficial, numeración propia vía PARAMETROS tipo 'AS').
#   4) Inserta las 2 líneas del asiento: Debe la cuenta de "Descuentos Otorgados"
#      (51205013 — confirmado con Eduardo, PARAMETROSCUENTAS orden 6001), Haber la
#      cuenta "Deudores por Ventas" — resuelta dinámicamente igual que lo hace
#      Flexxus internamente (PARAMETROSCUENTAS PANTALLA='VENTAS' ORDEN=4001), nunca
#      hardcodeada.
#   5) Vincula el asiento al comprobante CI vía FMA_COMPROBANTESASIENTOS (procedimiento
#      oficial, tabla COMPROBANTESASIENTOS).
#
# Fuente de cada pieza: /debug/ci_investigacion, /debug/ci_triggers_source,
# /debug/ci_asientos_investigacion, /debug/ci_procedimiento_generar y
# /debug/ci_plan_cuentas (investigación de solo lectura hecha antes de escribir esto).
# ============================================================================
CUENTA_DEBE_DESCUENTOS_OTORGADOS = 51205013  # PLANDECUENTAS — confirmado con Eduardo 2026-07-31

def _generar_credito_interno_ci(codigo_cliente, codigo_usuario, monto, descripcion, cuenta_debe=None, db_path=None):
    """Genera un comprobante CI (crédito interno) en Firebird/Flexxus para un cliente,
    con su asiento contable asociado, en una sola transacción atómica.

    Parámetros:
      codigo_cliente: CODIGOCLIENTE del cliente que recibe el crédito (str).
      codigo_usuario: CODIGOUSUARIO del vendedor/usuario que genera el crédito (str).
      monto: monto POSITIVO del crédito en pesos (float). La función se encarga de
        pasarlo en negativo al procedimiento oficial de Flexxus (igual que hace la
        app de escritorio) para que reste saldo deudor del cliente.
      descripcion: texto real del origen del crédito (ej. "CREDITO POR PROMOCION -
        Oferta Día del Niño - Escalón 2"), reemplaza el 'CREDITO POR PAGO' que el
        procedimiento oficial deja hardcodeado.
      cuenta_debe: código PLANDECUENTAS a debitar. Default: Descuentos Otorgados
        (51205013).
      db_path: ruta de la base Firebird a usar (DATABASE u DATABASE_MLT según el
        pedido/presupuesto haya sido confirmado en 'oficial' o 'sw'). Default: DATABASE
        — IMPORTANTE pasar el mismo db que usó el pedido, si no el código de cliente
        puede no existir o referirse a otro cliente en esa base.

    Devuelve: {"numero_ci": <numero de comprobante CI>, "codigo_asiento": <código del
    asiento generado>}.

    Lanza Exception si algo falla — con rollback automático, no queda nada a medias
    (ni el CI se genera sin su asiento, ni el asiento sin el CI)."""
    if not codigo_cliente or not str(codigo_cliente).strip():
        raise ValueError("codigo_cliente es obligatorio")
    if not codigo_usuario or not str(codigo_usuario).strip():
        raise ValueError("codigo_usuario es obligatorio")
    monto = float(monto or 0)
    if monto <= 0:
        raise ValueError("monto debe ser mayor a 0")
    cuenta_debe = float(cuenta_debe if cuenta_debe is not None else CUENTA_DEBE_DESCUENTOS_OTORGADOS)

    c = conn('WIN1252', db=(db_path or DATABASE))
    try:
        cur = c.cursor()

        # 1) Comprobante CI vía el procedimiento oficial de Flexxus
        cur.execute(
            'EXECUTE PROCEDURE "FMA_GENERARCREDITOINTERNO" ?, ?, ?',
            (codigo_cliente, codigo_usuario, -abs(monto))
        )
        row = cur.fetchone()
        if not row or row[0] is None:
            raise Exception("FMA_GENERARCREDITOINTERNO no devolvió NUMEROCI")
        numero_ci = row[0]

        # 2) Corregir el texto hardcodeado ('CREDITO POR PAGO') por el motivo real
        desc = (descripcion or '').strip()[:100] or 'CREDITO POR PROMOCION'
        cur.execute(
            'UPDATE "CABEZACOMPROBANTES" SET COMENTARIOS=? '
            "WHERE TIPOCOMPROBANTE='CI' AND NUMEROCOMPROBANTE=?",
            (desc, numero_ci)
        )
        cur.execute(
            'UPDATE "CUERPOCOMPROBANTES" SET DESCRIPCION=? '
            "WHERE TIPOCOMPROBANTE='CI' AND NUMEROCOMPROBANTE=? AND LINEA=1",
            (desc, numero_ci)
        )

        # 3) Cabecera del asiento contable (procedimiento oficial, numeración propia)
        cur.execute('EXECUTE PROCEDURE "FMA_CABEZAASIENTOS"')
        row = cur.fetchone()
        if not row or row[0] is None:
            raise Exception("FMA_CABEZAASIENTOS no devolvió CODIGOASIENTOOUT")
        codigo_asiento = row[0]

        # 4) Cuenta "Deudores por Ventas" — resuelta dinámicamente, igual que lo hace
        # Flexxus internamente en FMA_CUERPOASIENTOS (nunca hardcodeada acá)
        cur.execute(
            'SELECT FIRST 1 PAC.CUENTA FROM "PARAMETROSCUENTAS" PAC '
            "WHERE PAC.PANTALLA='VENTAS' AND PAC.ORDEN=4001 ORDER BY PAC.ORDEN"
        )
        row = cur.fetchone()
        if not row or row[0] is None:
            raise Exception("No se encontró la cuenta 'Deudores por Ventas' parametrizada (PARAMETROSCUENTAS VENTAS/4001)")
        cuenta_haber = row[0]

        # 5) Líneas del asiento: Debe cuenta de descuento/gasto, Haber Deudores por Ventas
        cur.execute(
            'INSERT INTO "CUERPOASIENTOS" '
            '(CODIGOASIENTO, LINEA, CUENTA, MONTO, ESDEBE, FECHAMODIFICACION, NUMEROTRANSACCION) '
            'VALUES (?, 1, ?, ?, 1, CURRENT_TIMESTAMP, 0)',
            (codigo_asiento, cuenta_debe, monto)
        )
        cur.execute(
            'INSERT INTO "CUERPOASIENTOS" '
            '(CODIGOASIENTO, LINEA, CUENTA, MONTO, ESDEBE, FECHAMODIFICACION, NUMEROTRANSACCION) '
            'VALUES (?, 2, ?, ?, 0, CURRENT_TIMESTAMP, 0)',
            (codigo_asiento, cuenta_haber, monto)
        )

        # 6) Vincular el asiento al comprobante CI (procedimiento oficial)
        cur.execute(
            'EXECUTE PROCEDURE "FMA_COMPROBANTESASIENTOS" ?, ?, ?',
            (codigo_asiento, 'CI', numero_ci)
        )
        row = cur.fetchone()
        if not row or row[0] != 1:
            raise Exception("FMA_COMPROBANTESASIENTOS no devolvió RESULTADO=1")

        c.commit()
        return {"numero_ci": numero_ci, "codigo_asiento": codigo_asiento}
    except Exception:
        try:
            c.rollback()
        except Exception:
            pass
        raise
    finally:
        try:
            c.close()
        except Exception:
            pass


def _fb_datos_cliente_para_ci_di(cur, codigo_cliente, codigo_usuario):
    """Repite EXACTAMENTE las mismas consultas que hace el procedimiento oficial
    FMA_GENERARCREDITOINTERNO (según su código fuente real, ver /debug/ci_procedimiento_generar)
    para resolver punto de venta, depósito, caja, sucursal, datos del cliente, IVA,
    moneda y vendedor. La usa _generar_debito_interno_di para poder armar el INSERT a
    mano, ya que no existe un procedimiento oficial equivalente a FMA_GENERARCREDITOINTERNO
    para DI (confirmado en /debug/ci_reversion_investigacion, 2026-07-31)."""
    cur.execute(
        'SELECT U.CODIGOPUNTOVENTAAUX, D.CODIGODEPOSITO, U.CODIGOCAJA, P.CODIGOSUCURSAL '
        'FROM "USUARIOS" U '
        'LEFT JOIN "DEPOSITOS" D ON D.CODIGODEPOSITO = U.CODIGODEPOSITO '
        'LEFT JOIN "PUNTOSDEVENTA" P ON P.CODIGOPUNTOVENTA = U.CODIGOPUNTOVENTA '
        'LEFT JOIN "CAJAS" CJ ON CJ.CODIGOCAJA = U.CODIGOCAJA '
        'WHERE UPPER(U.CODIGOUSUARIO) = UPPER(?)',
        (codigo_usuario,)
    )
    row = cur.fetchone()
    if not row:
        raise Exception(f"Usuario '{codigo_usuario}' no encontrado en USUARIOS")
    codigo_ptovta, codigo_deposito, codigo_caja, codigo_sucursal = row[0], row[1], row[2], row[3]

    cur.execute(
        'SELECT C.RAZONSOCIAL, C.DIRECCION, C.CONDICIONIVA, '
        'SUBSTRING(C.TELEFONO FROM 1 FOR 18), C.DOCUMENTO, C.CUIT, C.CODIGOMULTIPLAZO '
        'FROM "CLIENTES" C WHERE C.CODIGOCLIENTE = ?',
        (codigo_cliente,)
    )
    row = cur.fetchone()
    if not row:
        raise Exception(f"Cliente '{codigo_cliente}' no encontrado en CLIENTES")
    (razonsocial, direccioncliente, condicioniva, telefonocliente,
     documentocliente, cuitcliente, codigomultiplazo) = row

    alicuotaiva = 0
    if condicioniva:
        cur.execute('SELECT T.IVA1 FROM "TIPOIVA" T WHERE T.CODIGOTIPO = ?', (condicioniva,))
        r = cur.fetchone()
        alicuotaiva = r[0] if r and r[0] is not None else 0

    cur.execute(
        'SELECT M.CODIGOMONEDA FROM "MONEDAS" M '
        'INNER JOIN "MONEDASCAJAS" MC ON M.CODIGOMONEDA = MC.CODIGOMONEDA '
        'WHERE M.MONEDABASE = 1 AND MC.CAMBIO = 1 AND MC.CODIGOCAJA = ?',
        (codigo_caja,)
    )
    r = cur.fetchone()
    codigomoneda = r[0] if (r and r[0]) else 'PESOS'

    codigovendedor = None
    cur.execute(
        'SELECT C.CODIGOVENDEDOR FROM "CLIENTES" C WHERE C.CODIGOCLIENTE = ? AND C.VENDEDORFIJO = 1',
        (codigo_cliente,)
    )
    r = cur.fetchone()
    if r and r[0]:
        codigovendedor = r[0]
    if not codigovendedor:
        codigovendedor = codigo_usuario

    return {
        "codigo_ptovta": codigo_ptovta, "codigo_deposito": codigo_deposito,
        "codigo_caja": codigo_caja, "codigo_sucursal": codigo_sucursal,
        "razonsocial": razonsocial, "direccioncliente": direccioncliente,
        "condicioniva": condicioniva, "telefonocliente": telefonocliente,
        "documentocliente": documentocliente, "cuitcliente": cuitcliente,
        "codigomultiplazo": codigomultiplazo, "alicuotaiva": alicuotaiva,
        "codigomoneda": codigomoneda, "codigovendedor": codigovendedor,
    }


# ============================================================================
# REVERSIÓN DE CRÉDITO INTERNO (DI) — función real (2026-07-31)
#
# ESTADO: implementada para revisión. NO está conectada a ningún endpoint todavía.
# Por decisión explícita de Eduardo (2026-07-31): esta función NUNCA se dispara sola
# ni automáticamente al detectar que un pedido ya no alcanza un escalón — se llama
# solo después de que un humano (Eduardo o MPEREZ) revisa y aprueba la reversión
# desde el panel de control. El flujo de "detectar el desajuste y dejarlo pendiente
# de aprobación" todavía no está diseñado — es el próximo paso, aparte de esto.
#
# A diferencia de _generar_credito_interno_ci, NO existe un procedimiento oficial de
# Flexxus equivalente para generar DI (confirmado en /debug/ci_reversion_investigacion:
# la búsqueda de FMA_GENERARDEBITOINTERNO y similares no encontró nada). El ejemplo
# real de DI investigado antes (cheque rebotado, "DEVUELTOS") aparenta cargarse a mano
# desde la pantalla genérica de comprobantes de Flexxus, no vía lógica automatizada.
#
# Por eso acá replicamos, campo por campo, EXACTAMENTE el mismo patrón de INSERT que
# sí está oficialmente validado en el código fuente real de FMA_GENERARCREDITOINTERNO
# (ver /debug/ci_procedimiento_generar) — mismas columnas, mismos valores por defecto
# — cambiando solo: TIPOCOMPROBANTE='DI', TOTAL positivo (un DI le vuelve a sumar
# deuda al cliente, cancelando el crédito previo), numeración propia vía CONTADOR
# 00015 (en vez de 00014 que usa CI), y el asiento contable invertido (Debe/Haber
# cambiados respecto al asiento del CI original).
# ============================================================================

def _generar_debito_interno_di(codigo_cliente, codigo_usuario, monto, descripcion,
                                 ci_original_numero=None, cuenta_debe=None, db_path=None):
    """Genera un comprobante DI (Débito Interno) que REVIERTE un crédito interno (CI)
    previamente otorgado por el sistema de ofertas, cuando el pedido/escalón que lo
    originó ya no corresponde (pedido modificado por debajo del escalón, o cancelado).

    Parámetros:
      codigo_cliente, codigo_usuario: igual que en _generar_credito_interno_ci.
      monto: monto POSITIVO a revertir (float) — debe ser el mismo monto del CI
        original que se está anulando.
      descripcion: texto real del motivo (ej. "Reversión pedido 000123 modificado,
        ya no alcanza Escalón 2 de Oferta Día del Niño").
      ci_original_numero: número del comprobante CI que se está revirtiendo, para
        trazabilidad (se agrega a la descripción si no viene ya incluido).
      cuenta_debe: código PLANDECUENTAS que se había debitado en el CI original
        (default: Descuentos Otorgados, igual que en _generar_credito_interno_ci) —
        en la reversión esta cuenta va del lado del HABER (se invierte).
      db_path: MISMA base Firebird (DATABASE u DATABASE_MLT) que se usó para generar
        el CI original — obligatorio pasarla igual, o el DI puede terminar en la base
        equivocada. Default: DATABASE.

    Devuelve: {"numero_di": ..., "codigo_asiento": ...}
    Lanza Exception si algo falla, con rollback automático — no queda nada a medias."""
    if not codigo_cliente or not str(codigo_cliente).strip():
        raise ValueError("codigo_cliente es obligatorio")
    if not codigo_usuario or not str(codigo_usuario).strip():
        raise ValueError("codigo_usuario es obligatorio")
    monto = float(monto or 0)
    if monto <= 0:
        raise ValueError("monto debe ser mayor a 0")
    cuenta_credito_original = float(cuenta_debe if cuenta_debe is not None else CUENTA_DEBE_DESCUENTOS_OTORGADOS)

    desc = (descripcion or '').strip()[:100] or 'REVERSION CREDITO INTERNO'
    if ci_original_numero and str(ci_original_numero) not in desc:
        desc = (desc + f" (CI {ci_original_numero})")[:100]

    c = conn('WIN1252', db=(db_path or DATABASE))
    try:
        cur = c.cursor()
        datos = _fb_datos_cliente_para_ci_di(cur, codigo_cliente, codigo_usuario)

        # Numeración propia del DI (CONTADOR 00015), mismo patrón que usa Flexxus para CI
        cur.execute(
            'UPDATE "CONTADORES" SET VALOR = VALOR + 1 WHERE CODIGOCONTADOR = ('
            'SELECT C.CODIGOCONTADOR FROM "PARAMETROSPUNTOSDEVENTA" P '
            'INNER JOIN "CONTADORES" C ON P.CODIGOCONTADOR = C.CODIGOCONTADOR '
            "WHERE P.TIPODOCUMENTO = 'DI' AND P.CODIGOPUNTOVENTA = ?)",
            (datos["codigo_ptovta"],)
        )
        cur.execute(
            'SELECT C.VALOR - 1 FROM "PARAMETROSPUNTOSDEVENTA" P '
            'INNER JOIN "CONTADORES" C ON P.CODIGOCONTADOR = C.CODIGOCONTADOR '
            "WHERE P.TIPODOCUMENTO = 'DI' AND P.CODIGOPUNTOVENTA = ?",
            (datos["codigo_ptovta"],)
        )
        row = cur.fetchone()
        if not row or row[0] is None:
            raise Exception("No se pudo obtener numeración DI (PARAMETROSPUNTOSDEVENTA/CONTADORES)")
        numero_di = row[0]

        fecha_sql = ("CAST(EXTRACT(MONTH FROM CURRENT_TIMESTAMP) || '/' || "
                     "EXTRACT(DAY FROM CURRENT_TIMESTAMP) || '/' || "
                     "EXTRACT(YEAR FROM CURRENT_TIMESTAMP) AS DATE)")

        # CABEZACOMPROBANTES — mismo patrón exacto que FMA_GENERARCREDITOINTERNO,
        # con TIPOCOMPROBANTE='DI', TOTAL positivo y COMENTARIOS con el motivo real.
        sql_cabeza = (
            'INSERT INTO "CABEZACOMPROBANTES" ('
            'TIPOCOMPROBANTE, NUMEROCOMPROBANTE, CODIGOCLIENTE, FECHACOMPROBANTE, '
            'RAZONSOCIAL, DIRECCION, PORCIVA1, PORCIVA2, IVA1, IVA2, TOTAL, PAGADO, '
            'CUENTACORRIENTE, HORA, CODIGOUSUARIO, TIPOIVA, REMITOFACTURADO, COMENTARIOS, '
            'TELEFONO, FECHAVENCIMIENTO, IMPRIME, ANULADA, CUIT, COMPRA, CODIGOTRANSPORTE, '
            'MONTOTRANSPORTE, CODIGOMULTIPLAZO, EXENTO, CLASECOMPROBANTE, CODIGOUSUARIO2, '
            'COEFICIENTEIVA, FECHAMODIFICACION, DESCCOMPROBANTE, CODIGOMONEDA, COTIZACION, '
            'NUMEROTRANSACCION, CANTIDADBULTOS, NROPUNTODEVENTA, CODIGOPROYECTO, DESCUENTOPORCENTAJE, '
            'DESCUENTOMONTO, DESCUENTODESCRIPCION, CANTIDADPAGINAS, LISTAPRECIO, VALIDACTACTE, '
            'MONTOTOTALII, FECHAVENCIMIENTO2, RECARGOVENCIMIENTO2, FECHAVENCIMIENTO3, RECARGOVENCIMIENTO3'
            ') VALUES (?, ?, ?, ' + fecha_sql + ', ?, ?, ?, 0, 0, 0, ?, 0, 1, CURRENT_TIMESTAMP, ?, ?, 0, ?, '
            '?, ' + fecha_sql + ', 0, 0, ?, 0, 2, 0, ?, 0, 0, ?, 0, ' + fecha_sql + ", '', ?, 1, 0, 0, 4, '0', 0, 0, '0 %', "
            '1, 1, 1, 0, ' + fecha_sql + ', 0, ' + fecha_sql + ', 0)'
        )
        cur.execute(sql_cabeza, (
            'DI', numero_di, codigo_cliente,
            datos["razonsocial"], datos["direccioncliente"], datos["alicuotaiva"],
            monto,
            datos["codigovendedor"], datos["condicioniva"], desc,
            datos["telefonocliente"],
            datos["cuitcliente"],
            datos["codigomultiplazo"],
            codigo_usuario,
            datos["codigomoneda"],
        ))

        # CUERPOCOMPROBANTES — ídem, mismo patrón exacto con línea única CODIGOARTICULO='*'
        sql_cuerpo = (
            'INSERT INTO "CUERPOCOMPROBANTES" ('
            'TIPOCOMPROBANTE, NUMEROCOMPROBANTE, LINEA, CODIGOARTICULO, DESCRIPCION, CANTIDAD, '
            'DESCUENTO, PRECIOUNITARIO, PRECIOTOTAL, GARANTIA, INTERES, CANTIDADREMITIDA, LOTE, '
            'ESCONJUNTO, FECHAMODIFICACION, CODIGODEPOSITO, COSTOVENTA, NUMEROTRANSACCION, '
            'CODIGOPARTICULAR, PORCENTAJEIVA, DESCDESCUENTO, TIPOPRECIO, PORCENTAJEDESCUENTOS, '
            'MONTOII, COEFICIENTECONVERSION, CODIGOEMPAQUE, DESCRIPCIONEMPAQUE, OBSERVACIONES'
            ") VALUES (?, ?, 1, '*', ?, 1, 0, ?, ?, 0, 0, 0, '000', 0, " + fecha_sql + ', ?, 0, 0, '
            "'*', ?, NULL, NULL, 0, 0, 0, '', '', NULL)"
        )
        cur.execute(sql_cuerpo, (
            'DI', numero_di, desc, monto, monto, datos["codigo_deposito"], datos["alicuotaiva"]
        ))

        # Asiento contable INVERSO al del CI original: Debe Deudores por Ventas,
        # Haber la cuenta que se había debitado (ej. Descuentos Otorgados).
        cur.execute('EXECUTE PROCEDURE "FMA_CABEZAASIENTOS"')
        row = cur.fetchone()
        if not row or row[0] is None:
            raise Exception("FMA_CABEZAASIENTOS no devolvió CODIGOASIENTOOUT")
        codigo_asiento = row[0]

        cur.execute(
            'SELECT FIRST 1 PAC.CUENTA FROM "PARAMETROSCUENTAS" PAC '
            "WHERE PAC.PANTALLA='VENTAS' AND PAC.ORDEN=4001 ORDER BY PAC.ORDEN"
        )
        row = cur.fetchone()
        if not row or row[0] is None:
            raise Exception("No se encontró la cuenta 'Deudores por Ventas' parametrizada (PARAMETROSCUENTAS VENTAS/4001)")
        cuenta_deudores = row[0]

        cur.execute(
            'INSERT INTO "CUERPOASIENTOS" '
            '(CODIGOASIENTO, LINEA, CUENTA, MONTO, ESDEBE, FECHAMODIFICACION, NUMEROTRANSACCION) '
            'VALUES (?, 1, ?, ?, 1, CURRENT_TIMESTAMP, 0)',
            (codigo_asiento, cuenta_deudores, monto)
        )
        cur.execute(
            'INSERT INTO "CUERPOASIENTOS" '
            '(CODIGOASIENTO, LINEA, CUENTA, MONTO, ESDEBE, FECHAMODIFICACION, NUMEROTRANSACCION) '
            'VALUES (?, 2, ?, ?, 0, CURRENT_TIMESTAMP, 0)',
            (codigo_asiento, cuenta_credito_original, monto)
        )

        # Vincular el asiento al comprobante DI (procedimiento oficial)
        cur.execute(
            'EXECUTE PROCEDURE "FMA_COMPROBANTESASIENTOS" ?, ?, ?',
            (codigo_asiento, 'DI', numero_di)
        )
        row = cur.fetchone()
        if not row or row[0] != 1:
            raise Exception("FMA_COMPROBANTESASIENTOS no devolvió RESULTADO=1")

        c.commit()
        return {"numero_di": numero_di, "codigo_asiento": codigo_asiento}
    except Exception:
        try:
            c.rollback()
        except Exception:
            pass
        raise
    finally:
        try:
            c.close()
        except Exception:
            pass


@app.get("/debug/ci_investigacion")
def debug_ci_investigacion(_u=Depends(get_admin_user)):
    """DEBUG TEMPORAL (2026-07-31): investigación de solo lectura (sin ningún INSERT/
    UPDATE) para la feature de crédito interno por escalón (bono acreditado en cuenta
    corriente vía comprobante tipo CI, visto en la pantalla Archivos > Contadores de
    Flexxus: código 00014 "CREDITO INTERNO", y su contraparte 00015 "DEBITO INTERNO").
    Junta en una sola llamada todo lo necesario para diseñar el INSERT real:
      1) Columnas de CABEZACOMPROBANTES y CUERPOCOMPROBANTES.
      2) Fila(s) de CONTADORES que correspondan a CI/DI (numeración propia).
      3) Si ya existen comprobantes TIPOCOMPROBANTE='CI' o 'DI' cargados (aunque sea
         de prueba manual desde Flexxus) — cabecera y cuerpo del más reciente de cada
         uno, para ver la forma REAL que toman sus campos. Si no hay ninguno, lo
         informa así sabemos que hay que armar el INSERT sin un ejemplo de referencia.
      4) Triggers sobre CABEZACOMPROBANTES (por si hay lógica automática — numeración,
         impacto en cuenta corriente, validaciones — que corre sola al insertar)."""
    out = {}
    try:
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()

        # 1) Columnas de ambas tablas
        for tabla in ('CABEZACOMPROBANTES', 'CUERPOCOMPROBANTES'):
            try:
                cur.execute(
                    "SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS "
                    "WHERE RDB$RELATION_NAME=? ORDER BY RDB$FIELD_POSITION", (tabla,)
                )
                out[f'columnas_{tabla}'] = [r[0] for r in cur.fetchall()]
            except Exception as e:
                out[f'columnas_{tabla}_error'] = str(e)

        # 2) Contador(es) CI/DI — búsqueda dirigida + fallback con todos por si el
        # texto de DESCRIPCION no matchea exacto.
        try:
            cur.execute(
                "SELECT CODIGOCONTADOR, DESCRIPCION, VALOR FROM \"CONTADORES\" "
                "WHERE UPPER(DESCRIPCION) CONTAINING UPPER(?) "
                "OR UPPER(DESCRIPCION) CONTAINING UPPER(?) "
                "ORDER BY CODIGOCONTADOR", ('CREDITO INTERNO', 'DEBITO INTERNO')
            )
            out['contadores_ci_di'] = [
                {"codigo": r[0], "descripcion": (r[1] or '').strip(), "valor": r[2]}
                for r in cur.fetchall()
            ]
        except Exception as e:
            out['contadores_ci_di_error'] = str(e)

        # 3) Comprobantes CI/DI ya cargados, si existen — cabecera + cuerpo del más
        # reciente de cada tipo.
        for tipo in ('CI', 'DI'):
            try:
                cur.execute(
                    'SELECT FIRST 3 * FROM "CABEZACOMPROBANTES" WHERE TIPOCOMPROBANTE=? '
                    'ORDER BY NUMEROCOMPROBANTE DESC', (tipo,)
                )
                rows = cur.fetchall()
                cols = [d[0] for d in cur.description]
                out[f'muestra_cabeza_{tipo}'] = [
                    {cols[i]: (str(row[i]) if row[i] is not None else None) for i in range(len(cols))}
                    for row in rows
                ]
                if rows and 'NUMEROCOMPROBANTE' in cols:
                    numero0 = rows[0][cols.index('NUMEROCOMPROBANTE')]
                    cur.execute(
                        'SELECT FIRST 5 * FROM "CUERPOCOMPROBANTES" WHERE TIPOCOMPROBANTE=? AND NUMEROCOMPROBANTE=?',
                        (tipo, numero0)
                    )
                    rows_c = cur.fetchall()
                    cols_c = [d[0] for d in cur.description]
                    out[f'muestra_cuerpo_{tipo}'] = [
                        {cols_c[i]: (str(r[i]) if r[i] is not None else None) for i in range(len(cols_c))}
                        for r in rows_c
                    ]
                elif not rows:
                    out[f'muestra_cabeza_{tipo}_nota'] = f'No hay comprobantes TIPOCOMPROBANTE={tipo} cargados todavía.'
            except Exception as e:
                out[f'muestra_{tipo}_error'] = str(e)

        # 4) Triggers sobre CABEZACOMPROBANTES
        try:
            cur.execute(
                "SELECT TRIM(RDB$TRIGGER_NAME) FROM RDB$TRIGGERS "
                "WHERE RDB$RELATION_NAME='CABEZACOMPROBANTES' AND RDB$SYSTEM_FLAG=0"
            )
            out['triggers_cabezacomprobantes'] = [r[0] for r in cur.fetchall()]
        except Exception as e:
            out['triggers_error'] = str(e)

        c.close()
    except Exception as e:
        out['error_general'] = str(e)
    return out

@app.get("/debug/ci_triggers_source")
def debug_ci_triggers_source(_u=Depends(get_admin_user)):
    """DEBUG TEMPORAL (2026-07-31): trae el código fuente PSQL de los 4 triggers reales
    detectados sobre CABEZACOMPROBANTES (ver 'triggers_cabezacomprobantes' en
    /debug/ci_investigacion): CABEZACOMPROBANTES_NT, TRIG_PAGADO_VENTAS,
    CABEZACOMPROBANTES_BI0, MTO_TGCABEZACOMPROBANTES. Solo lectura (RDB$TRIGGERS es
    tabla de sistema, no toca datos). El objetivo es confirmar si la generación
    automática del asiento contable (visto en Flexxus vía 'Ver Asiento' para un
    comprobante CI) ocurre a nivel trigger de base (entonces un INSERT crudo desde
    Python la replicaría solo) o si corre únicamente desde la app de escritorio Delphi
    (entonces habría que reproducir esa lógica a mano)."""
    triggers = [
        'CABEZACOMPROBANTES_NT',
        'TRIG_PAGADO_VENTAS',
        'CABEZACOMPROBANTES_BI0',
        'MTO_TGCABEZACOMPROBANTES',
    ]
    out = {}
    try:
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()
        for t in triggers:
            try:
                cur.execute(
                    "SELECT RDB$TRIGGER_SOURCE FROM RDB$TRIGGERS WHERE RDB$TRIGGER_NAME=?", (t,)
                )
                row = cur.fetchone()
                if row and row[0] is not None:
                    src = row[0]
                    if isinstance(src, bytes):
                        src = src.decode('latin1', errors='replace')
                    out[t] = src
                else:
                    out[t] = None
            except Exception as e:
                out[t + '_error'] = str(e)
        c.close()
    except Exception as e:
        out['error_general'] = str(e)
    return out

@app.get("/debug/ci_asientos_investigacion")
def debug_ci_asientos_investigacion(_u=Depends(get_admin_user)):
    """DEBUG TEMPORAL (2026-07-31): sigue la investigación de crédito interno por
    escalón. El comprobante CI de ejemplo (6399) que vimos resultó ser un crédito
    generado por un PAGO (cheque de 3ros que superó lo adeudado) — toca caja/valores,
    por eso el asiento debita 'Cheques de 3° Directos'. Lo que necesitamos nosotros es
    un crédito por bonificación/promoción comercial, SIN movimiento de caja, que
    debería debitar una cuenta de gasto (ej. GASTOS DE COMERCIALIZACION) contra el
    haber de Deudores por Ventas del cliente — un circuito contable distinto.
    Este endpoint busca:
      1) Triggers sobre CUERPOCOMPROBANTES (no revisado aún).
      2) Tablas cuyo nombre sugiera manejo de asientos/contabilidad.
      3) Procedimientos almacenados (RDB$PROCEDURES) que sugieran generación de
         asientos o estén vinculados a pagos/cobros (posible lugar donde Delphi
         genera el asiento contable, si no es un trigger)."""
    out = {}
    try:
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()

        # 1) Triggers sobre CUERPOCOMPROBANTES
        try:
            cur.execute(
                "SELECT TRIM(RDB$TRIGGER_NAME) FROM RDB$TRIGGERS "
                "WHERE RDB$RELATION_NAME='CUERPOCOMPROBANTES' AND RDB$SYSTEM_FLAG=0"
            )
            out['triggers_cuerpocomprobantes'] = [r[0] for r in cur.fetchall()]
        except Exception as e:
            out['triggers_cuerpocomprobantes_error'] = str(e)

        # 2) Tablas relacionadas a asientos/contabilidad
        try:
            out['tablas_contables'] = []
            for kw in ('ASIENTO', 'CONTAB', 'DIARIO', 'PLANCTA', 'MAYOR', 'MOVCTA', 'CUENTA'):
                cur.execute(
                    "SELECT TRIM(RDB$RELATION_NAME) FROM RDB$RELATIONS "
                    "WHERE RDB$SYSTEM_FLAG=0 AND UPPER(RDB$RELATION_NAME) CONTAINING UPPER(?)",
                    (kw,)
                )
                for r in cur.fetchall():
                    if r[0] not in out['tablas_contables']:
                        out['tablas_contables'].append(r[0])
        except Exception as e:
            out['tablas_contables_error'] = str(e)

        # 3) Procedimientos almacenados relacionados a asientos / pagos / cobros
        try:
            out['procedimientos_relacionados'] = []
            for kw in ('ASIENTO', 'CONTAB', 'PAGO', 'COBRO', 'CREDITOINTERNO', 'CTACTE'):
                cur.execute(
                    "SELECT TRIM(RDB$PROCEDURE_NAME) FROM RDB$PROCEDURES "
                    "WHERE UPPER(RDB$PROCEDURE_NAME) CONTAINING UPPER(?)",
                    (kw,)
                )
                for r in cur.fetchall():
                    if r[0] not in out['procedimientos_relacionados']:
                        out['procedimientos_relacionados'].append(r[0])
        except Exception as e:
            out['procedimientos_relacionados_error'] = str(e)

        # 4) Triggers en TODAS las tablas cuyo nombre sugiera pagos/cobros (posible
        # lugar real donde se genera el asiento, ya que el ejemplo CI que vimos vino
        # de un pago, no de la carga directa del comprobante)
        try:
            out['triggers_tablas_pagos_cobros'] = []
            for kw in ('PAGO', 'COBRO', 'RECIBO', 'CHEQUE'):
                cur.execute(
                    "SELECT TRIM(RDB$RELATION_NAME) FROM RDB$RELATIONS "
                    "WHERE RDB$SYSTEM_FLAG=0 AND UPPER(RDB$RELATION_NAME) CONTAINING UPPER(?)",
                    (kw,)
                )
                tablas = [r[0] for r in cur.fetchall()]
                for tabla in tablas:
                    cur.execute(
                        "SELECT TRIM(RDB$TRIGGER_NAME) FROM RDB$TRIGGERS "
                        "WHERE RDB$RELATION_NAME=? AND RDB$SYSTEM_FLAG=0", (tabla,)
                    )
                    trigs = [r[0] for r in cur.fetchall()]
                    if trigs:
                        out['triggers_tablas_pagos_cobros'].append({"tabla": tabla, "triggers": trigs})
        except Exception as e:
            out['triggers_tablas_pagos_cobros_error'] = str(e)

        c.close()
    except Exception as e:
        out['error_general'] = str(e)
    return out

@app.get("/debug/ci_procedimiento_generar")
def debug_ci_procedimiento_generar(_u=Depends(get_admin_user)):
    """DEBUG TEMPORAL (2026-07-31): trae parámetros + código fuente PSQL completo de
    FMA_GENERARCREDITOINTERNO (hallado en /debug/ci_asientos_investigacion — procedimiento
    almacenado que genera el crédito interno, posiblemente incluyendo numeración,
    INSERT en CABEZACOMPROBANTES/CUERPOCOMPROBANTES y el asiento contable en un solo
    paso). También trae FMA_CABEZAASIENTOS, FMA_CUERPOASIENTOS y
    FMA_COMPROBANTESASIENTOS por si son invocados desde adentro. Si logramos llamar
    a este procedimiento vía EXECUTE PROCEDURE desde Python en vez de armar un INSERT
    crudo, replicaríamos exactamente la lógica real de Flexxus (numeración correcta,
    asiento contable correcto) sin tener que reconstruirla nosotros."""
    procs = [
        'FMA_GENERARCREDITOINTERNO',
        'FMA_CABEZAASIENTOS',
        'FMA_CUERPOASIENTOS',
        'FMA_COMPROBANTESASIENTOS',
    ]
    out = {}
    try:
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()
        for p in procs:
            try:
                cur.execute(
                    "SELECT TRIM(RDB$PARAMETER_NAME), RDB$PARAMETER_NUMBER, RDB$PARAMETER_TYPE, "
                    "TRIM(RDB$FIELD_SOURCE) FROM RDB$PROCEDURE_PARAMETERS "
                    "WHERE RDB$PROCEDURE_NAME=? ORDER BY RDB$PARAMETER_TYPE, RDB$PARAMETER_NUMBER",
                    (p,)
                )
                out[f'{p}_parametros'] = [
                    {
                        "nombre": r[0],
                        "posicion": r[1],
                        "tipo": "INPUT" if r[2] == 0 else "OUTPUT",
                        "dominio": r[3],
                    }
                    for r in cur.fetchall()
                ]
            except Exception as e:
                out[f'{p}_parametros_error'] = str(e)
            try:
                cur.execute(
                    "SELECT RDB$PROCEDURE_SOURCE FROM RDB$PROCEDURES WHERE RDB$PROCEDURE_NAME=?", (p,)
                )
                row = cur.fetchone()
                if row and row[0] is not None:
                    src = row[0]
                    if isinstance(src, bytes):
                        src = src.decode('latin1', errors='replace')
                    out[f'{p}_source'] = src
                else:
                    out[f'{p}_source'] = None
            except Exception as e:
                out[f'{p}_source_error'] = str(e)
        c.close()
    except Exception as e:
        out['error_general'] = str(e)
    return out

@app.get("/debug/ci_plan_cuentas")
def debug_ci_plan_cuentas(_u=Depends(get_admin_user)):
    """DEBUG TEMPORAL (2026-07-31): última pieza para diseñar el asiento propio del
    crédito interno por escalón. Necesitamos:
      1) Columnas de PLANDECUENTAS (para saber cómo se relaciona el CODIGOCUENTA
         numérico usado en CUERPOASIENTOS con el número de cuenta tipo '5.1.2.05.012'
         que se ve en la UI de Flexxus).
      2) La fila de PLANDECUENTAS para la cuenta de gastos de comercialización
         (buscada por descripción 'COMERCIALIZ' y por número '5.1.2.05%').
      3) Confirmar cómo se resuelve la cuenta 'Deudores por Ventas' que ya usa
         Flexxus internamente (vía PARAMETROSCUENTAS PANTALLA='VENTAS' ORDEN=4001,
         visto en FMA_CUERPOASIENTOS) — reusamos esa misma consulta para el haber."""
    out = {}
    try:
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()

        try:
            cur.execute(
                "SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS "
                "WHERE RDB$RELATION_NAME='PLANDECUENTAS' ORDER BY RDB$FIELD_POSITION"
            )
            out['columnas_plandecuentas'] = [r[0] for r in cur.fetchall()]
        except Exception as e:
            out['columnas_plandecuentas_error'] = str(e)

        try:
            cur.execute('SELECT * FROM "PLANDECUENTAS" WHERE UPPER(DESCRIPCION) CONTAINING UPPER(?)', ('COMERCIALIZ',))
            rows = cur.fetchall()
            cols = [d[0] for d in cur.description]
            out['plandecuentas_comercializacion'] = [
                {cols[i]: (str(row[i]) if row[i] is not None else None) for i in range(len(cols))}
                for row in rows
            ]
        except Exception as e:
            out['plandecuentas_comercializacion_error'] = str(e)

        try:
            cur.execute(
                "SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS "
                "WHERE RDB$RELATION_NAME='PARAMETROSCUENTAS' ORDER BY RDB$FIELD_POSITION"
            )
            out['columnas_parametroscuentas'] = [r[0] for r in cur.fetchall()]
        except Exception as e:
            out['columnas_parametroscuentas_error'] = str(e)

        try:
            cur.execute('SELECT * FROM "PARAMETROSCUENTAS" WHERE PANTALLA=?', ('VENTAS',))
            rows = cur.fetchall()
            cols = [d[0] for d in cur.description]
            out['parametroscuentas_ventas'] = [
                {cols[i]: (str(row[i]) if row[i] is not None else None) for i in range(len(cols))}
                for row in rows
            ]
        except Exception as e:
            out['parametroscuentas_ventas_error'] = str(e)

        try:
            cur.execute(
                'SELECT FIRST 1 PAC.CUENTA FROM "PARAMETROSCUENTAS" PAC '
                "WHERE PAC.PANTALLA='VENTAS' AND PAC.ORDEN=4001 ORDER BY PAC.ORDEN"
            )
            row = cur.fetchone()
            cuentaventas = row[0] if row else None
            out['cuentaventas_codigo'] = cuentaventas
            if cuentaventas is not None:
                cur.execute('SELECT * FROM "PLANDECUENTAS" WHERE CODIGOCUENTA=?', (cuentaventas,))
                rows = cur.fetchall()
                cols = [d[0] for d in cur.description]
                out['cuentaventas_detalle'] = [
                    {cols[i]: (str(row[i]) if row[i] is not None else None) for i in range(len(cols))}
                    for row in rows
                ]
        except Exception as e:
            out['cuentaventas_error'] = str(e)

        c.close()
    except Exception as e:
        out['error_general'] = str(e)
    return out

@app.get("/debug/ci_reversion_investigacion")
def debug_ci_reversion_investigacion(_u=Depends(get_admin_user)):
    """DEBUG TEMPORAL (2026-07-31): investiga cómo revertir un crédito interno (CI) ya
    generado, para el caso en que el pedido que lo originó se modifique/cancele después
    y deje de alcanzar el escalón. Hipótesis a confirmar: Flexxus nunca borra/edita un
    comprobante ya emitido — el mecanismo oficial probablemente sea generar un DI
    (Débito Interno, la contraparte de CI que ya vimos en CONTADORES código 00015) por
    el mismo monto, que compensa el efecto en cuenta corriente sin tocar el CI original.
    Busca:
      1) Procedimientos almacenados relacionados a generar DI / anular comprobantes /
         contra-asientos.
      2) Si existe FMA_GENERARDEBITOINTERNO (o similar), sus parámetros y código fuente.
      3) Configuración de numeración para TIPODOCUMENTO='DI' en PARAMETROSPUNTOSDEVENTA
         (para confirmar que está tan lista para usarse como CI)."""
    out = {}
    try:
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()

        try:
            out['procedimientos_reversion'] = []
            for kw in ('DEBITOINTERNO', 'GENERARDEBITO', 'ANULA', 'REVERTIR', 'CONTRAASIENTO', 'CANCELA'):
                cur.execute(
                    "SELECT TRIM(RDB$PROCEDURE_NAME) FROM RDB$PROCEDURES "
                    "WHERE UPPER(RDB$PROCEDURE_NAME) CONTAINING UPPER(?)",
                    (kw,)
                )
                for r in cur.fetchall():
                    if r[0] not in out['procedimientos_reversion']:
                        out['procedimientos_reversion'].append(r[0])
        except Exception as e:
            out['procedimientos_reversion_error'] = str(e)

        for p in out.get('procedimientos_reversion', []):
            try:
                cur.execute(
                    "SELECT TRIM(RDB$PARAMETER_NAME), RDB$PARAMETER_NUMBER, RDB$PARAMETER_TYPE, "
                    "TRIM(RDB$FIELD_SOURCE) FROM RDB$PROCEDURE_PARAMETERS "
                    "WHERE RDB$PROCEDURE_NAME=? ORDER BY RDB$PARAMETER_TYPE, RDB$PARAMETER_NUMBER",
                    (p,)
                )
                out[f'{p}_parametros'] = [
                    {"nombre": r[0], "posicion": r[1], "tipo": "INPUT" if r[2] == 0 else "OUTPUT", "dominio": r[3]}
                    for r in cur.fetchall()
                ]
            except Exception as e:
                out[f'{p}_parametros_error'] = str(e)
            try:
                cur.execute("SELECT RDB$PROCEDURE_SOURCE FROM RDB$PROCEDURES WHERE RDB$PROCEDURE_NAME=?", (p,))
                row = cur.fetchone()
                if row and row[0] is not None:
                    src = row[0]
                    if isinstance(src, bytes):
                        src = src.decode('latin1', errors='replace')
                    out[f'{p}_source'] = src
                else:
                    out[f'{p}_source'] = None
            except Exception as e:
                out[f'{p}_source_error'] = str(e)

        try:
            cur.execute(
                'SELECT P.CODIGOPUNTOVENTA, P.TIPODOCUMENTO, P.CODIGOCONTADOR, C.DESCRIPCION, C.VALOR '
                'FROM "PARAMETROSPUNTOSDEVENTA" P INNER JOIN "CONTADORES" C ON P.CODIGOCONTADOR = C.CODIGOCONTADOR '
                "WHERE P.TIPODOCUMENTO IN ('CI', 'DI')"
            )
            out['numeracion_ci_di_por_puntoventa'] = [
                {"punto_venta": r[0], "tipo": r[1], "codigo_contador": r[2], "descripcion_contador": r[3], "valor_actual": r[4]}
                for r in cur.fetchall()
            ]
        except Exception as e:
            out['numeracion_ci_di_por_puntoventa_error'] = str(e)

        c.close()
    except Exception as e:
        out['error_general'] = str(e)
    return out

@app.get("/debug/info_bd")
def debug_info_bd():
    """Muestra las rutas de BD configuradas y cuenta de registros clave"""
    result = {
        "DATABASE": DATABASE,
        "DATABASE_MLT": DATABASE_MLT,
    }
    try:
        c = conn()
        cur = c.cursor()
        cur.execute('SELECT COUNT(*) FROM "CABEZAPRESUPUESTOS"')
        result["presupuestos_total"] = cur.fetchone()[0]
        cur.execute('SELECT MAX(CAST(NUMEROCOMPROBANTE AS INTEGER)) FROM "CABEZAPRESUPUESTOS"')
        result["presupuesto_max"] = cur.fetchone()[0]
        cur.execute('SELECT COUNT(*) FROM "CABEZAPEDIDOS"')
        result["pedidos_total"] = cur.fetchone()[0]
        cur.execute('SELECT MAX(CAST(NUMEROCOMPROBANTE AS INTEGER)) FROM "CABEZAPEDIDOS"')
        result["pedido_max"] = cur.fetchone()[0]
        c.close()
    except Exception as e:
        result["error"] = str(e)
    return result

@app.get("/debug/config_prueba")
def debug_config_prueba():
    """Lee configuracion.ini del perfil Prueba"""
    import os
    result = {}
    paths = [
        'c:/flexxus/FlexxusERP/Prueba/bin/configuracion.ini',
        'c:/flexxus/FlexxusERP/BIN/FlexxusServer.ini',
    ]
    for p in paths:
        try:
            with open(p, 'r', encoding='latin1', errors='replace') as f:
                result[p] = f.read()
        except Exception as e:
            result[p] = f"Error: {e}"
    return result

@app.get("/debug/config_flexxus")
def debug_config_flexxus():
    """Busca archivos de configuración de Flexxus que definen la BD por empresa"""
    import glob, os
    result = {"archivos": [], "contenido": {}}
    patterns = [
        'c:/flexxus/**/*.ini', 'c:/flexxus/**/*.cfg',
        'c:/flexxus/**/*.config', 'c:/flexxus/**/Empresas*.xml',
        'c:/flexxus/**/empresas*.ini', 'c:/flexxus/**/conexion*.ini',
        'c:/flexxus/FlexxusERP/*.ini', 'c:/flexxus/FlexxusERP/*.cfg',
        'c:/flexxus/FlexxusERP/*.xml',
    ]
    for p in patterns:
        for f in glob.glob(p, recursive=True):
            result["archivos"].append(f)
            try:
                with open(f, 'r', encoding='latin1', errors='replace') as fp:
                    content = fp.read(3000)
                if 'DB-' in content or 'Prueba' in content or 'Microbell' in content or 'gdb' in content.lower():
                    result["contenido"][f] = content
            except Exception as e:
                result["contenido"][f] = f"Error: {e}"
    return result

@app.get("/debug/comparar_bases")
def debug_comparar_bases():
    """Compara presupuesto MAX en DB-Prueba vs DB-Microbell"""
    result = {}
    for nombre, path in [
        ("DB-Prueba", "c:/flexxus/DB/DB-Prueba.gdb"),
        ("DB-Microbell", "c:/flexxus/DB/DB-Microbell.gdb"),
    ]:
        try:
            c = firebirdsql.connect(host=HOST, port=PORT, database=path,
                                    user=DB_USER, password=DB_PASS, charset='LATIN1')
            cur = c.cursor()
            cur.execute('SELECT MAX(CAST(NUMEROCOMPROBANTE AS INTEGER)), COUNT(*) FROM "CABEZAPRESUPUESTOS"')
            row = cur.fetchone()
            cur.execute('SELECT MAX(CAST(NUMEROCOMPROBANTE AS INTEGER)), COUNT(*) FROM "CABEZAPEDIDOS"')
            row2 = cur.fetchone()
            c.close()
            result[nombre] = {
                "presupuesto_max": row[0], "presupuestos_total": row[1],
                "pedido_max": row2[0], "pedidos_total": row2[1]
            }
        except Exception as e:
            result[nombre] = {"error": str(e)}
    return result

@app.get("/debug/listar_gdbs")
def debug_listar_gdbs():
    """Lista todos los archivos .gdb y .fdb en c:/flexxus/"""
    import glob, os
    result = []
    for pattern in ['c:/flexxus/**/*.gdb', 'c:/flexxus/**/*.fdb',
                    'c:/Flexxus/**/*.gdb', 'c:/Flexxus/**/*.fdb']:
        for f in glob.glob(pattern, recursive=True):
            try:
                size = os.path.getsize(f)
                mtime = os.path.getmtime(f)
                from datetime import datetime
                result.append({
                    "path": f,
                    "size_mb": round(size / 1024 / 1024, 1),
                    "modificado": datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M')
                })
            except Exception:
                result.append({"path": f})
    return sorted(result, key=lambda x: x.get("modificado",""), reverse=True)

@app.get("/debug/esquema_cuerpos")
def debug_esquema_cuerpos():
    """Columnas reales de CUERPOPRESUPUESTOS y CUERPOPEDIDOS en DB-Prueba.gdb"""
    result = {}
    c = conn()
    cur = c.cursor()
    for tabla in ['CUERPOPRESUPUESTOS', 'CUERPOPEDIDOS']:
        cur.execute(
            "SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS "
            "WHERE TRIM(RDB$RELATION_NAME) = ? ORDER BY RDB$FIELD_POSITION",
            (tabla,))
        result[tabla] = [r[0] for r in cur.fetchall()]
    c.close()
    return result

@app.get("/debug/esquema_cabezas")
def debug_esquema_cabezas():
    """Columnas de CABEZAPRESUPUESTOS y CABEZAPEDIDOS"""
    result = {}
    c = conn()
    cur = c.cursor()
    for tabla in ['CABEZAPRESUPUESTOS', 'CABEZAPEDIDOS']:
        cur.execute(
            "SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS "
            "WHERE TRIM(RDB$RELATION_NAME) = ? ORDER BY RDB$FIELD_POSITION",
            (tabla,))
        result[tabla] = [r[0] for r in cur.fetchall()]
    c.close()
    return result

@app.get("/debug/muestra_presupuesto/{numero}")
def debug_muestra_presupuesto(numero: str):
    """Muestra cabeza + cuerpo de un presupuesto específico"""
    try:
        c = conn()
        cur = c.cursor()
        cur.execute('SELECT * FROM "CABEZAPRESUPUESTOS" WHERE NUMEROCOMPROBANTE = ?', (numero,))
        cols_cab = [d[0] for d in cur.description]
        row = cur.fetchone()
        cabeza = dict(zip(cols_cab, [str(v) if v is not None else None for v in row])) if row else None
        cur.execute('SELECT * FROM "CUERPOPRESUPUESTOS" WHERE NUMEROCOMPROBANTE = ? ORDER BY LINEA', (numero,))
        cols_cue = [d[0] for d in cur.description]
        cuerpos = [dict(zip(cols_cue, [str(v) if v is not None else None for v in r])) for r in cur.fetchall()]
        c.close()
        return {"cabeza": cabeza, "cuerpos": cuerpos}
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/codigos_presupuesto/{numero}")
def debug_codigos_presupuesto(numero: str):
    """Compara CODIGOARTICULO vs CODIGOPARTICULAR en CUERPOPRESUPUESTOS."""
    try:
        c = conn()
        cur = c.cursor()
        cur.execute(
            "SELECT LINEA, CODIGOARTICULO, CODIGOPARTICULAR, DESCRIPCION "
            "FROM \"CUERPOPRESUPUESTOS\" WHERE NUMEROCOMPROBANTE = ? ORDER BY LINEA",
            (numero,)
        )
        rows = cur.fetchall()
        c.close()
        return [{"linea": r[0], "codigoarticulo": r[1], "codigoparticular": r[2], "descripcion": r[3]} for r in rows]
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/ultimos_presupuestos")
def debug_ultimos_presupuestos():
    """Últimos 5 presupuestos en DB-Prueba.gdb"""
    try:
        c = conn()
        cur = c.cursor()
        cur.execute(
            'SELECT FIRST 5 TIPOCOMPROBANTE, NUMEROCOMPROBANTE, CODIGOCLIENTE, '
            'RAZONSOCIAL, FECHACOMPROBANTE, TOTAL, CODIGOUSUARIO '
            'FROM "CABEZAPRESUPUESTOS" ORDER BY FECHAMODIFICACION DESC')
        rows = cur.fetchall()
        c.close()
        return [{"tipo": r[0], "numero": r[1], "cliente": r[2],
                 "razon": r[3], "fecha": str(r[4]), "total": r[5], "usuario": r[6]}
                for r in rows]
    except Exception as e:
        return {"error": str(e)}


@app.get("/debug/comparar/{num_bueno}/{num_malo}")
def debug_comparar(num_bueno: str, num_malo: str):
    """
    Compara TODAS las columnas de dos presupuestos.
    Uso: /debug/comparar/7849/7857
    Resalta las columnas que difieren.
    """
    try:
        c = conn()
        cur = c.cursor()

        def leer(numero):
            cur.execute('SELECT * FROM "CABEZAPRESUPUESTOS" WHERE NUMEROCOMPROBANTE = ?', (numero,))
            cols = [d[0] for d in cur.description]
            row = cur.fetchone()
            if not row:
                return None, cols
            return dict(zip(cols, [str(v) if v is not None else '__NULL__' for v in row])), cols

        datos_bueno, cols = leer(num_bueno)
        datos_malo, _    = leer(num_malo)

        # Defaults del dominio para cada columna
        cur.execute("""
            SELECT TRIM(rf.RDB$FIELD_NAME), rf.RDB$NULL_FLAG, rf.RDB$DEFAULT_SOURCE
            FROM RDB$RELATION_FIELDS rf
            WHERE rf.RDB$RELATION_NAME = 'CABEZAPRESUPUESTOS'
            ORDER BY rf.RDB$FIELD_POSITION
        """)
        defaults = {r[0].strip(): {"not_null": r[1]==1, "default": str(r[2]).strip() if r[2] else None}
                    for r in cur.fetchall()}

        # Todos los triggers (INSERT + UPDATE)
        cur.execute("""
            SELECT TRIM(RDB$TRIGGER_NAME), RDB$TRIGGER_TYPE, CAST(RDB$TRIGGER_SOURCE AS VARCHAR(4000))
            FROM RDB$TRIGGERS
            WHERE RDB$RELATION_NAME = 'CABEZAPRESUPUESTOS' AND RDB$SYSTEM_FLAG = 0
        """)
        triggers = [{"name": r[0], "type": r[1],
                     "type_desc": {1:"BEFORE INSERT",2:"AFTER INSERT",3:"BEFORE UPDATE",4:"AFTER UPDATE",
                                   5:"BEFORE DELETE",6:"AFTER DELETE"}.get(r[1],"?"),
                     "source": r[2]} for r in cur.fetchall()]

        c.close()

        if not datos_bueno:
            return {"error": f"No encontrado: {num_bueno}"}
        if not datos_malo:
            return {"error": f"No encontrado: {num_malo}"}

        diferencias = {}
        iguales = {}
        for col in cols:
            v1 = datos_bueno.get(col)
            v2 = datos_malo.get(col)
            if v1 != v2:
                diferencias[col] = {
                    f"bueno_{num_bueno}": v1,
                    f"malo_{num_malo}": v2,
                    "default_info": defaults.get(col.strip())
                }
            else:
                iguales[col] = v1

        return {
            "diferencias": diferencias,
            "total_diferencias": len(diferencias),
            "iguales": iguales,
            "triggers": triggers,
            "defaults_columnas": defaults,
        }
    except Exception as e:
        return {"error": str(e)}


@app.post("/setup/restaurar_defaults_aprobacion")
def restaurar_defaults_aprobacion():
    """
    Restaura el DEFAULT original de FECHAAPROBADO ('1900-01-01') y CODIGOUSUARIOAPROBACION ('')
    que fue eliminado por setup/fechaaprobado_nullable.
    Ejecutar UNA SOLA VEZ.
    """
    try:
        import struct
        c = conn('WIN1252')
        cur = c.cursor()

        # Restaurar DEFAULT '1900-01-01 00:00:00' en FECHAAPROBADO
        # En Firebird, el DEFAULT_SOURCE es texto SQL, y DEFAULT_VALUE es BLR binario.
        # Actualizamos solo DEFAULT_SOURCE (texto); Firebird lo recompila al arrancar.
        cur.execute("""
            UPDATE RDB$RELATION_FIELDS
            SET RDB$DEFAULT_SOURCE = 'DEFAULT ''1900-01-01 00:00:00'''
            WHERE RDB$RELATION_NAME = 'CABEZAPRESUPUESTOS'
              AND TRIM(RDB$FIELD_NAME) = 'FECHAAPROBADO'
        """)
        # Restaurar DEFAULT '' en CODIGOUSUARIOAPROBACION
        cur.execute("""
            UPDATE RDB$RELATION_FIELDS
            SET RDB$DEFAULT_SOURCE = 'DEFAULT '''''
            WHERE RDB$RELATION_NAME = 'CABEZAPRESUPUESTOS'
              AND TRIM(RDB$FIELD_NAME) = 'CODIGOUSUARIOAPROBACION'
        """)
        c.commit()

        # Verificar
        cur.execute("""
            SELECT TRIM(RDB$FIELD_NAME), RDB$NULL_FLAG, RDB$DEFAULT_SOURCE
            FROM RDB$RELATION_FIELDS
            WHERE RDB$RELATION_NAME = 'CABEZAPRESUPUESTOS'
              AND TRIM(RDB$FIELD_NAME) IN ('FECHAAPROBADO', 'CODIGOUSUARIOAPROBACION')
        """)
        rows = cur.fetchall()
        c.close()
        return {
            "ok": True,
            "columnas": [{"campo": r[0], "not_null": r[1], "default": str(r[2]) if r[2] else None} for r in rows],
            "instruccion": "Reiniciar Firebird para que tome los nuevos defaults"
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/debug/parametros_np")
def debug_parametros_np():
    """Verifica el valor de NP en PARAMETROS y hace prueba de TRIM."""
    res = {}
    try:
        c = conn('LATIN1')
        cur = c.cursor()
        # Buscar con TRIM
        cur.execute("SELECT TIPODOCUMENTO, VALOR FROM \"PARAMETROS\" WHERE TRIM(TIPODOCUMENTO) = 'NP'")
        row = cur.fetchone()
        res['con_trim'] = {'tipodocumento': repr(row[0]), 'valor': row[1]} if row else None
        # Buscar sin TRIM para ver el valor crudo
        cur.execute("SELECT TIPODOCUMENTO, VALOR FROM \"PARAMETROS\" WHERE TIPODOCUMENTO LIKE '%NP%'")
        rows = cur.fetchall()
        res['like_NP'] = [{'tipodocumento': repr(r[0]), 'valor': r[1]} for r in rows]
        # Ver todas las filas con comprobante tipo 2 chars
        cur.execute("SELECT TIPODOCUMENTO, VALOR FROM \"PARAMETROS\" ORDER BY TIPODOCUMENTO")
        all_rows = cur.fetchall()
        res['primeros_20'] = [{'tipodocumento': repr(r[0]), 'valor': r[1]} for r in all_rows[:20]]
        c.close()
    except Exception as e:
        res['error'] = str(e)
    return res

@app.get("/debug/mlt_tablas_pedidos")
def debug_mlt_tablas_pedidos():
    """Verifica si DB-MLT tiene CABEZAPEDIDOS y cuántos registros NP hay en cada tabla."""
    res = {"db": DATABASE_MLT}
    try:
        c = conn('WIN1252', db=DATABASE_MLT)
        cur = c.cursor()
        # Tablas que existen
        cur.execute("SELECT TRIM(RDB$RELATION_NAME) FROM RDB$RELATIONS WHERE RDB$SYSTEM_FLAG=0 AND RDB$RELATION_NAME CONTAINING 'CABEZA'")
        res["tablas_cabeza"] = [r[0] for r in cur.fetchall()]
        # Contar NP en CABEZACOMPROBANTES
        try:
            cur.execute("SELECT COUNT(*) FROM \"CABEZACOMPROBANTES\" WHERE TIPOCOMPROBANTE='NP'")
            res["cabezacomprobantes_NP"] = cur.fetchone()[0]
        except Exception as e:
            res["cabezacomprobantes_NP_error"] = str(e)
        # Contar NP en CABEZAPEDIDOS si existe
        try:
            res["cabezapedidos_NP"] = cur.fetchone()[0]
        except Exception as e:
            res["cabezapedidos_NP_error"] = str(e)
        c.close()
    except Exception as e:
        res["error"] = str(e)
    return res

@app.get("/debug/fix_cuentacorriente_sw")
def debug_fix_cuentacorriente_sw(numero: str = Query(...)):
    """Corrige CUENTACORRIENTE=1 en CABEZACOMPROBANTES de DB-MLT para un pedido SW."""
    try:
        c = conn('WIN1252', db=DATABASE_MLT)
        cur = c.cursor()
        cur.execute(
            'UPDATE "CABEZACOMPROBANTES" SET CUENTACORRIENTE=1 '
            'WHERE NUMEROCOMPROBANTE=? AND TIPOCOMPROBANTE=?',
            (numero, 'NP')
        )
        affected = cur.rowcount
        c.commit()
        c.close()
        return {"ok": True, "numero": numero, "filas_actualizadas": affected}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/debug/inspect_sw_pedido")
def debug_inspect_sw_pedido(numero: str = Query(...)):
    """Devuelve todos los campos de CABEZACOMPROBANTES en DB-MLT para un número dado,
    y también el último NP nativo (creado por Flexxus) para comparar."""
    try:
        c = conn('WIN1252', db=DATABASE_MLT)
        cur = c.cursor()
        # Registro del pedido pedido
        cur.execute('SELECT * FROM "CABEZACOMPROBANTES" WHERE NUMEROCOMPROBANTE=? AND TIPOCOMPROBANTE=?', (numero, 'NP'))
        row = cur.fetchone()
        cols = [d[0] for d in cur.description]
        target = dict(zip(cols, [str(v) for v in row])) if row else None
        # Último NP que NO sea el nuestro (para comparar)
        cur.execute(
            'SELECT FIRST 1 * FROM "CABEZACOMPROBANTES" WHERE TIPOCOMPROBANTE=? AND NUMEROCOMPROBANTE<>? ORDER BY FECHACOMPROBANTE DESC',
            ('NP', numero)
        )
        row2 = cur.fetchone()
        nativo = dict(zip(cols, [str(v) for v in row2])) if row2 else None
        c.close()
        # Diferencias
        diffs = {}
        if target and nativo:
            for k in cols:
                if target.get(k) != nativo.get(k):
                    diffs[k] = {"nuestro": target.get(k), "nativo": nativo.get(k)}
        return {"pedido": numero, "encontrado": target is not None, "diferencias_con_nativo": diffs, "nuestro": target, "nativo": nativo}
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/check_dbs")
def debug_check_dbs():
    """Muestra qué bases de datos está usando la API."""
    import os
    return {
        "DATABASE": DATABASE,
        "DATABASE_MLT": DATABASE_MLT,
        "DB_L1_env": os.getenv('DB_L1', '(no seteado)'),
        "DB_MLT_env": os.getenv('DB_MLT', '(no seteado)'),
    }

@app.get("/debug/comparar_sw_produccion")
def debug_comparar_sw_produccion(n_prod: str = Query(...), n_prueba: str = Query(...)):
    """Compara un pedido nativo de DB-MLT-Microbell.gdb (producción SW)
    contra uno creado por la API en DB-MLT-Prueba.gdb."""
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    def _get(db, numero):
        try:
            c = conn('WIN1252', db=db)
            cur = c.cursor()
            cur.execute('SELECT * FROM "CABEZACOMPROBANTES" WHERE NUMEROCOMPROBANTE=? AND TIPOCOMPROBANTE=?', (numero, 'NP'))
            row = cur.fetchone()
            cols = [d[0] for d in cur.description]
            c.close()
            return dict(zip(cols, [str(v) for v in row])) if row else None
        except Exception as e:
            return {"error": str(e)}
    prod  = _get(DB_MLT_PROD, n_prod)
    prueba = _get(DATABASE_MLT, n_prueba)
    diffs = {}
    if prod and prueba and "error" not in prod and "error" not in prueba:
        for k in prod:
            if k in prueba and prod[k] != prueba[k]:
                diffs[k] = {"produccion_nativo": prod[k], "prueba_api": prueba[k]}
    return {"n_prod": n_prod, "n_prueba": n_prueba,
            "produccion_encontrado": prod is not None and "error" not in (prod or {}),
            "prueba_encontrado": prueba is not None and "error" not in (prueba or {}),
            "diferencias": diffs}

@app.get("/debug/tablas_mlt_prod")
def debug_tablas_mlt_prod():
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    try:
        c = conn('WIN1252', db=DB_MLT_PROD)
        cur = c.cursor()
        cur.execute("SELECT TRIM(RDB$RELATION_NAME) FROM RDB$RELATIONS WHERE RDB$SYSTEM_FLAG=0 AND RDB$RELATION_NAME CONTAINING 'CABEZA' ORDER BY 1")
        tablas = [r[0] for r in cur.fetchall()]
        # Buscar el pedido en todas las tablas CABEZA*
        encontrado = {}
        for t in tablas:
            try:
                cur.execute(f'SELECT COUNT(*) FROM "{t}" WHERE CAST(NUMEROCOMPROBANTE AS VARCHAR(20))=? AND TIPOCOMPROBANTE=?', ('100023558', 'NP'))
                n = cur.fetchone()[0]
                if n > 0:
                    encontrado[t] = n
            except Exception:
                pass
        c.close()
        return {"tablas_cabeza": tablas, "pedido_100023558_en": encontrado}
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/mlt_prod_np")
def debug_mlt_prod_np():
    """Muestra los últimos 5 NP en CABEZACOMPROBANTES de DB-MLT-Microbell.gdb y total de registros."""
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    try:
        c = conn('WIN1252', db=DB_MLT_PROD)
        cur = c.cursor()
        cur.execute("SELECT COUNT(*) FROM \"CABEZACOMPROBANTES\"")
        total = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM \"CABEZACOMPROBANTES\" WHERE TIPOCOMPROBANTE='NP'")
        total_np = cur.fetchone()[0]
        cur.execute("SELECT FIRST 5 TIPOCOMPROBANTE, NUMEROCOMPROBANTE, CODIGOCLIENTE, RAZONSOCIAL, FECHACOMPROBANTE FROM \"CABEZACOMPROBANTES\" WHERE TIPOCOMPROBANTE='NP' ORDER BY FECHACOMPROBANTE DESC")
        rows = cur.fetchall()
        c.close()
        return {
            "db": DB_MLT_PROD,
            "total_comprobantes": total,
            "total_NP": total_np,
            "ultimos_5_NP": [{"tipo": r[0], "numero": str(r[1]), "cliente": r[2], "razon": r[3], "fecha": str(r[4])} for r in rows]
        }
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/mlt_prod_parametros")
def debug_mlt_prod_parametros():
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    try:
        c = conn('WIN1252', db=DB_MLT_PROD)
        cur = c.cursor()
        cur.execute("SELECT TIPODOCUMENTO, VALOR FROM \"PARAMETROS\" WHERE TRIM(TIPODOCUMENTO) IN ('NP','PP')")
        params = {str(r[0]).strip(): str(r[1]) for r in cur.fetchall()}
        # Max número en CABEZACOMPROBANTES
        cur.execute("SELECT MAX(NUMEROCOMPROBANTE) FROM \"CABEZACOMPROBANTES\" WHERE TIPOCOMPROBANTE='NP'")
        max_np = str(cur.fetchone()[0])
        c.close()
        return {"db": DB_MLT_PROD, "parametros_NP_PP": params, "max_NP_en_tabla": max_np}
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/comparar_l1_vs_sw_nativo")
def debug_comparar_l1_vs_sw_nativo(n_sw: str = Query(...), n_l1: str = Query(...)):
    """Compara un pedido SW nativo vs uno L1 nativo, ambos en CABEZAPEDIDOS de DB-Prueba."""
    try:
        c = conn('WIN1252', db=DATABASE)
        cur = c.cursor()
        def _get(numero):
            cur.execute('SELECT * FROM "CABEZAPEDIDOS" WHERE NUMEROCOMPROBANTE=? AND TIPOCOMPROBANTE=?', (numero, 'NP'))
            row = cur.fetchone()
            cols = [d[0] for d in cur.description]
            return dict(zip(cols, [str(v) for v in row])) if row else None
        sw  = _get(n_sw)
        l1  = _get(n_l1)
        c.close()
        diffs = {}
        if sw and l1:
            for k in sw:
                if sw[k] != l1[k]:
                    diffs[k] = {"sw_nativo": sw[k], "l1_nativo": l1[k]}
        return {"diferencias": diffs, "sw": sw, "l1": l1}
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/buscar_pedido")
def debug_buscar_pedido(numero: str = Query(...)):
    """Busca un pedido en L1 (CABEZAPEDIDOS) y SW (CABEZACOMPROBANTES) y reporta en cuál está."""
    resultado = {"numero": numero, "l1": None, "sw": None}
    # L1
    try:
        c1 = conn('WIN1252', db=DATABASE)
        cur1 = c1.cursor()
        cur1.execute(
            'SELECT NUMEROCOMPROBANTE, CODIGOCLIENTE, RAZONSOCIAL, TOTAL, FECHACOMPROBANTE '
            'FROM "CABEZAPEDIDOS" WHERE NUMEROCOMPROBANTE = ? AND TIPOCOMPROBANTE = ?',
            (numero, 'NP'))
        r = cur1.fetchone()
        c1.close()
        resultado["l1"] = {"encontrado": r is not None, "db": DATABASE,
                           "fila": [str(x) for x in r] if r else None}
    except Exception as e:
        resultado["l1"] = {"error": str(e)}
    # SW
    try:
        c2 = conn('WIN1252', db=DATABASE_MLT)
        cur2 = c2.cursor()
        cur2.execute(
            'SELECT NUMEROCOMPROBANTE, CODIGOCLIENTE, RAZONSOCIAL, TOTAL, FECHACOMPROBANTE '
            'FROM "CABEZACOMPROBANTES" WHERE NUMEROCOMPROBANTE = ? AND TIPOCOMPROBANTE = ?',
            (numero, 'NP'))
        r2 = cur2.fetchone()
        c2.close()
        resultado["sw"] = {"encontrado": r2 is not None, "db": DATABASE_MLT,
                           "fila": [str(x) for x in r2] if r2 else None}
    except Exception as e:
        resultado["sw"] = {"error": str(e)}
    return resultado

@app.get("/debug/ultimos_pedidos")
def debug_ultimos_pedidos():
    """Últimos 5 pedidos en DB-Prueba.gdb"""
    try:
        c = conn()
        cur = c.cursor()
        cur.execute(
            'SELECT FIRST 5 TIPOCOMPROBANTE, NUMEROCOMPROBANTE, CODIGOCLIENTE, '
            'RAZONSOCIAL, FECHACOMPROBANTE, TOTAL, CODIGOUSUARIO '
            'FROM "CABEZAPEDIDOS" ORDER BY FECHAMODIFICACION DESC')
        rows = cur.fetchall()
        c.close()
        return [{"tipo": r[0], "numero": r[1], "cliente": r[2],
                 "razon": r[3], "fecha": str(r[4]), "total": r[5], "usuario": r[6]}
                for r in rows]
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/test_mlt")
def debug_test_mlt():
    """Verifica conectividad, tablas y datos de DB-MLT-Microbell.gdb."""
    DB_MLT_PROD = 'c:/flexxus/DB/DB-MLT-Microbell.gdb'
    result = {"db_path": DB_MLT_PROD}

    def _fresh(sql, params=None):
        c2 = firebirdsql.connect(host=HOST, port=PORT, database=DB_MLT_PROD,
                                  user=DB_USER, password=DB_PASS, charset='WIN1252')
        cur2 = c2.cursor()
        if params:
            cur2.execute(sql, params)
        else:
            cur2.execute(sql)
        rows = cur2.fetchall()
        c2.close()
        return rows

    try:
        result["tablas_count"] = len(_fresh(
            "SELECT RDB$RELATION_NAME FROM RDB$RELATIONS WHERE RDB$SYSTEM_FLAG=0"))
        result["conexion"] = "OK"
    except Exception as e:
        result["conexion_error"] = str(e)
        return result

    for label, sql, params in [
        ("total",            'SELECT COUNT(*) FROM "CABEZACOMPROBANTES"', None),
        ("cta_1",            'SELECT COUNT(*) FROM "CABEZACOMPROBANTES" WHERE CUENTACORRIENTE=?', ('1',)),
        ("cta_0",            'SELECT COUNT(*) FROM "CABEZACOMPROBANTES" WHERE CUENTACORRIENTE=?', ('0',)),
        ("anulada_0",        'SELECT COUNT(*) FROM "CABEZACOMPROBANTES" WHERE ANULADA=?', ('0',)),
        ("anulada_1",        'SELECT COUNT(*) FROM "CABEZACOMPROBANTES" WHERE ANULADA=?', ('1',)),
        ("filtros_cta",      'SELECT COUNT(*) FROM "CABEZACOMPROBANTES" '
                             "WHERE CUENTACORRIENTE='1' AND ANULADA='0' "
                             "AND TIPOCOMPROBANTE NOT IN ('RE','RI','INA')", None),
    ]:
        try:
            result[f"count_{label}"] = _fresh(sql, params)[0][0]
        except Exception as e:
            result[f"count_{label}_error"] = str(e)

    # Tipos de comprobante
    try:
        rows = _fresh('SELECT TIPOCOMPROBANTE, COUNT(*) FROM "CABEZACOMPROBANTES" '
                      'GROUP BY TIPOCOMPROBANTE ORDER BY 2 DESC')
        result["dist_tipo"] = {str(r[0]).strip(): r[1] for r in rows}
    except Exception as e:
        result["dist_tipo_error"] = str(e)

    # Suma bruta de debe con filtros _query_cta
    try:
        rows = _fresh(
            'SELECT COUNT(*), SUM(TOTAL+IVA1+IVA2-PAGADO) FROM "CABEZACOMPROBANTES" '
            "WHERE CUENTACORRIENTE='1' AND ANULADA='0' "
            "AND TIPOCOMPROBANTE NOT IN ('RE','RI','INA') "
            "AND (TOTAL+IVA1+IVA2-PAGADO) > 0"
        )
        result["registros_con_debe"] = rows[0][0]
        result["suma_debe_bruta"] = float(rows[0][1] or 0)
    except Exception as e:
        result["suma_debe_error"] = str(e)

    # Muestra últimas 5 filas
    try:
        rows = _fresh("SELECT FIRST 5 CODIGOCLIENTE, TIPOCOMPROBANTE, TOTAL, PAGADO, FECHACOMPROBANTE "
                      'FROM "CABEZACOMPROBANTES" ORDER BY FECHACOMPROBANTE DESC')
        result["muestra"] = [
            {"cod": str(r[0]).strip(), "tipo": str(r[1]).strip(),
             "total": float(r[2] or 0), "pagado": float(r[3] or 0), "fecha": str(r[4])}
            for r in rows
        ]
    except Exception as e:
        result["muestra_error"] = str(e)

    return result

@app.get("/debug/fa_en_todas_las_bds/{numero}")
def debug_fa_en_todas_las_bds(numero: str):
    """
    Busca un comprobante por NUMEROCOMPROBANTE en TODAS las BDs disponibles.
    Muestra TOTAL, IVA1, IVA2, PAGADO, CODIGOCLIENTE para comparar qué BD usa Flexxus.
    Ejemplo: /debug/fa_en_todas_las_bds/4400014878
    """
    bds = [
        ('DB-Prueba',    'c:/flexxus/DB/DB-Prueba.gdb'),
        ('DB-Microbell', 'c:/flexxus/DB/DB-Microbell.gdb'),
        ('DB-MLT-Prueba',    'c:/flexxus/DB/DB-MLT-Prueba.gdb'),
        ('DB-MLT-Microbell', 'c:/flexxus/DB/DB-MLT-Microbell.gdb'),
    ]
    result = {}
    for nombre, path in bds:
        for charset in ['WIN1252', 'LATIN1']:
            try:
                c2 = firebirdsql.connect(host=HOST, port=PORT, database=path,
                                          user=DB_USER, password=DB_PASS, charset=charset)
                cur2 = c2.cursor()
                cur2.execute(
                    'SELECT TIPOCOMPROBANTE, NUMEROCOMPROBANTE, CODIGOCLIENTE, '
                    'TOTAL, IVA1, IVA2, PAGADO, COTIZACION, CODIGOMONEDA, CUENTACORRIENTE, ANULADA '
                    'FROM "CABEZACOMPROBANTES" WHERE NUMEROCOMPROBANTE = ?',
                    (numero,)
                )
                rows = cur2.fetchall()
                c2.close()
                if rows:
                    result[nombre] = [{
                        "tipo": str(r[0]).strip(), "num": str(r[1]).strip(),
                        "cliente": str(r[2]).strip(),
                        "total": float(r[3] or 0), "iva1": float(r[4] or 0),
                        "iva2": float(r[5] or 0), "pagado": float(r[6] or 0),
                        "neto": float(r[3] or 0)+float(r[4] or 0)+float(r[5] or 0),
                        "debe": float(r[3] or 0)+float(r[4] or 0)+float(r[5] or 0)-float(r[6] or 0),
                        "cotiz": float(r[7] or 1), "moneda": str(r[8] or '').strip(),
                        "cta_corriente": str(r[9] or '').strip(), "anulada": r[10],
                    } for r in rows]
                else:
                    result[nombre] = "no_encontrado"
                break
            except Exception as e:
                result[f"{nombre}_{charset}"] = str(e)
    return result

@app.get("/debug/cliente_en_bds/{codigo}")
def debug_cliente_en_bds(codigo: str):
    """Muestra registro CLIENTES y todos sus comprobantes en DB-Prueba y DB-MLT-Prueba."""
    result = {}
    for nombre, db_path in [('L1_Prueba', DATABASE), ('SW_MLT', DATABASE_MLT)]:
        try:
            c = conn('WIN1252', db=db_path)
            cur = c.cursor()
            # Registro en CLIENTES
            cur.execute(
                'SELECT CODIGOCLIENTE, CODIGOPARTICULAR, RAZONSOCIAL, CODIGOVENDEDOR, ACTIVO '
                'FROM "CLIENTES" WHERE CODIGOCLIENTE=? OR CODIGOPARTICULAR=?',
                (codigo, codigo)
            )
            cli = cur.fetchone()
            result[nombre] = {'cliente': dict(zip(
                ['cod', 'part', 'razon', 'vendedor', 'activo'], cli
            )) if cli else None}
            # Comprobantes bajo ese código
            codigos = []
            if cli:
                if cli[0] and str(cli[0]).strip(): codigos.append(str(cli[0]).strip())
                if cli[1] and str(cli[1]).strip(): codigos.append(str(cli[1]).strip())
            if not codigos: codigos = [codigo]
            ph = ','.join(['?']*len(codigos))
            cur.execute(
                f'SELECT TIPOCOMPROBANTE, NUMEROCOMPROBANTE, CODIGOCLIENTE, TOTAL, PAGADO, '
                f'CUENTACORRIENTE, ANULADA FROM "CABEZACOMPROBANTES" '
                f'WHERE CODIGOCLIENTE IN ({ph}) ORDER BY NUMEROCOMPROBANTE',
                tuple(codigos)
            )
            rows = cur.fetchall()
            c.close()
            result[nombre]['comprobantes_total'] = len(rows)
            result[nombre]['comprobantes'] = [{
                'tipo': str(r[0]).strip(), 'num': str(r[1]).strip(),
                'cliente': str(r[2]).strip(),
                'total': float(r[3] or 0), 'pagado': float(r[4] or 0),
                'debe': float(r[3] or 0) - float(r[4] or 0),
                'cta': str(r[5] or '').strip(), 'anulada': r[6]
            } for r in rows]
        except Exception as e:
            result[nombre] = {'error': str(e)}
    return result

@app.get("/debug/tablas_usuarios")
def debug_tablas_usuarios():
    """Lista tablas que podrían contener usuarios/operadores y muestra sus columnas."""
    candidatas = ['USUARIOS', 'OPERADORES', 'VENDEDORES', 'USERS', 'EMPLEADOS',
                  'OPERADOR', 'USUARIO', 'AGENTES', 'PERSONAL']
    result = {}
    try:
        c = conn('WIN1252')
        cur = c.cursor()
        # Todas las tablas del sistema
        cur.execute("""
            SELECT RDB$RELATION_NAME FROM RDB$RELATIONS
            WHERE RDB$SYSTEM_FLAG = 0 AND RDB$VIEW_BLR IS NULL
            ORDER BY RDB$RELATION_NAME
        """)
        todas = [r[0].strip() for r in cur.fetchall()]
        result['todas_las_tablas'] = todas
        # Buscar candidatas
        for tabla in candidatas:
            if tabla in todas:
                try:
                    cur.execute(f'SELECT FIRST 3 * FROM "{tabla}"')
                    cols = [d[0] for d in cur.description]
                    rows = cur.fetchall()
                    result[tabla] = {
                        'columnas': cols,
                        'muestra': [dict(zip(cols, [str(v) for v in r])) for r in rows]
                    }
                except Exception as e:
                    result[tabla] = {'error': str(e)}
        c.close()
    except Exception as e:
        result['error'] = str(e)
    return result

@app.get("/debug/cotizacion_fa_akrafft")
def debug_cotizacion_fa_akrafft():
    """
    Muestra distribución de COTIZACION en FA con deuda para AKRAFFT.
    Si COTIZACION != 1 en algunos FA de pesos, _query_cta da distinto al SQL crudo.
    """
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'

    def _fresh(sql, params=None):
        c2 = firebirdsql.connect(host=HOST, port=PORT, database=DB_PROD,
                                  user=DB_USER, password=DB_PASS, charset='WIN1252')
        cur2 = c2.cursor()
        if params: cur2.execute(sql, params)
        else: cur2.execute(sql)
        rows = cur2.fetchall()
        c2.close()
        return rows

    try:
        rows = _fresh('SELECT CODIGOCLIENTE FROM "CLIENTES" WHERE ACTIVO=? AND UPPER(CODIGOVENDEDOR)=?',
                      ('1', 'AKRAFFT'))
        codigos = [str(r[0]).strip() for r in rows if (r[0] or '').strip()]
    except Exception as e:
        return {"error": str(e)}

    ph = ', '.join(['?'] * len(codigos))

    # Distribución de COTIZACION en registros con debe>0
    try:
        rows = _fresh(
            f'SELECT COTIZACION, CODIGOMONEDA, COUNT(*), SUM(TOTAL+IVA1+IVA2-PAGADO) '
            f'FROM "CABEZACOMPROBANTES" WHERE CODIGOCLIENTE IN ({ph}) '
            f"AND CUENTACORRIENTE='1' AND ANULADA='0' "
            f"AND TIPOCOMPROBANTE NOT IN ('RE','RI','INA') "
            f"AND (TOTAL+IVA1+IVA2-PAGADO) > 0 "
            f'GROUP BY COTIZACION, CODIGOMONEDA ORDER BY 4 DESC',
            tuple(codigos)
        )
        dist = [{"cotiz": float(r[0] or 1), "moneda": str(r[1]).strip(),
                 "count": r[2], "suma_bruta": float(r[3] or 0)} for r in rows]
    except Exception as e:
        return {"error_dist": str(e)}

    # Calcular lo que daría _query_cta para cada grupo (cambio PESOS=1, DOLARES=1475)
    cambios = {"PESOS": 1.0, "DOLARES": 1475.0}
    total_app = 0.0
    for d in dist:
        cambio = cambios.get(d["moneda"], 1.0)
        cotiz  = d["cotiz"] or 1.0
        # Cada FA tiene un COTIZACION diferente, pero aproximamos con la suma del grupo
        deuda_convertida = d["suma_bruta"] * cambio / cotiz
        d["deuda_convertida_aprox"] = round(deuda_convertida, 2)
        total_app += deuda_convertida

    return {
        "total_sql_bruto": round(sum(d["suma_bruta"] for d in dist), 2),
        "total_app_aprox": round(total_app, 2),
        "diferencia": round(sum(d["suma_bruta"] for d in dist) - total_app, 2),
        "dist_cotizacion": dist
    }

@app.get("/debug/gap2_akrafft")
def debug_gap2_akrafft():
    """
    Investigación de gap Parte 2:
    - Compara suma con/sin filtro CUENTACORRIENTE
    - Muestra valores reales de MONEDAS
    - Muestra distribución CUENTACORRIENTE y ANULADA para registros AKRAFFT
    """
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'

    def _fresh(sql, params=None):
        c2 = firebirdsql.connect(host=HOST, port=PORT, database=DB_PROD,
                                  user=DB_USER, password=DB_PASS, charset='WIN1252')
        cur2 = c2.cursor()
        if params:
            cur2.execute(sql, params)
        else:
            cur2.execute(sql)
        rows = cur2.fetchall()
        c2.close()
        return rows

    result = {}

    # Códigos AKRAFFT
    try:
        rows = _fresh('SELECT CODIGOCLIENTE FROM "CLIENTES" WHERE ACTIVO=? AND UPPER(CODIGOVENDEDOR)=?',
                      ('1', 'AKRAFFT'))
        codigos = [str(r[0]).strip() for r in rows if (r[0] or '').strip()]
    except Exception as e:
        return {"error_clientes": str(e)}

    ph = ', '.join(['?'] * len(codigos))

    # 1. Distribución CUENTACORRIENTE para registros AKRAFFT
    try:
        rows = _fresh(
            f'SELECT CUENTACORRIENTE, COUNT(*), SUM(TOTAL+IVA1+IVA2-PAGADO) '
            f'FROM "CABEZACOMPROBANTES" WHERE CODIGOCLIENTE IN ({ph}) '
            f"AND TIPOCOMPROBANTE NOT IN ('RE','RI','INA') AND ANULADA='0' "
            f'GROUP BY CUENTACORRIENTE',
            tuple(codigos)
        )
        result["dist_cuentacorriente"] = [
            {"cta": str(r[0]).strip() if r[0] is not None else "NULL",
             "count": r[1], "suma_debe": float(r[2] or 0)}
            for r in rows
        ]
    except Exception as e:
        result["error_dist_cta"] = str(e)

    # 2. Suma SIN filtro de CUENTACORRIENTE (todo tipo no RE/RI/INA, no anulado, debe>0)
    try:
        rows = _fresh(
            f'SELECT COUNT(*), SUM(TOTAL+IVA1+IVA2-PAGADO) FROM "CABEZACOMPROBANTES" '
            f"WHERE CODIGOCLIENTE IN ({ph}) AND ANULADA='0' "
            f"AND TIPOCOMPROBANTE NOT IN ('RE','RI','INA') "
            f"AND (TOTAL+IVA1+IVA2-PAGADO) > 0",
            tuple(codigos)
        )
        result["sin_filtro_cta_count"] = rows[0][0]
        result["sin_filtro_cta_suma"] = float(rows[0][1] or 0)
    except Exception as e:
        result["error_sin_cta"] = str(e)


# ─── Debug: estructura de pagos en Flexxus para comisiones ───────────────────
@app.get("/debug/pagos-estructura")
def debug_pagos_estructura(vendedor: str = 'RBOCHOR', _u=Depends(get_admin_user)):
    """
    Inspecciona cómo Flexxus registra los pagos de facturas.
    Muestra:
    - Tablas del sistema que contienen 'RECIBO','COBRO','IMPUTAC','PAGO'
    - Columnas de CABEZACOMPROBANTES relacionadas con pago
    - Muestra FA/NCA de un vendedor con TOTAL, IVA1, IVA2, PAGADO y % cobrado
    """
    DB_PROD = 'c:/flexxus/DB/DB-Microbell.gdb'
    result = {}

    def _q(sql, params=()):
        c = conn('WIN1252', DB_PROD)
        cur = c.cursor()
        cur.execute(sql, params)
        rows = cur.fetchall()
        cols = [d[0] for d in cur.description] if cur.description else []
        c.close()
        return cols, rows

    # 1. Tablas del sistema relacionadas con pagos/cobros
    try:
        cols, rows = _q(
            "SELECT TRIM(RDB$RELATION_NAME) FROM RDB$RELATIONS "
            "WHERE RDB$SYSTEM_FLAG=0 AND RDB$VIEW_BLR IS NULL "
            "ORDER BY RDB$RELATION_NAME"
        )
        all_tables = [r[0] for r in rows if r[0]]
        pago_tables = [t for t in all_tables if any(
            x in t.upper() for x in ['RECIBO','COBRO','IMPUTAC','PAGO','APLICA','CAJA','BANCO']
        )]
        result["tablas_pago_relacionadas"] = pago_tables
        result["total_tablas"] = len(all_tables)
    except Exception as e:
        result["error_tablas"] = str(e)

    # 2. Columnas de CABEZACOMPROBANTES (campo PAGADO y relacionados)
    try:
        cols, rows = _q(
            "SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS "
            "WHERE RDB$RELATION_NAME='CABEZACOMPROBANTES' "
            "ORDER BY RDB$FIELD_POSITION"
        )
        all_cols = [r[0] for r in rows if r[0]]
        pago_cols = [c for c in all_cols if any(
            x in c.upper() for x in ['PAG','COBR','SALDO','CANCEL','IMPUT']
        )]
        result["cols_pago_en_cabeza"] = pago_cols
    except Exception as e:
        result["error_cols"] = str(e)

    # 3. Muestra de FA/NCA del vendedor con estado de pago
    try:
        cols, rows = _q(
            "SELECT FIRST 20 TIPOCOMPROBANTE, NUMEROCOMPROBANTE, FECHACOMPROBANTE, "
            "CAST(TOTAL AS DOUBLE PRECISION), CAST(IVA1 AS DOUBLE PRECISION), "
            "CAST(IVA2 AS DOUBLE PRECISION), CAST(PAGADO AS DOUBLE PRECISION), "
            "CODIGOCLIENTE "
            "FROM \"CABEZACOMPROBANTES\" "
            "WHERE UPPER(TRIM(CODIGOUSUARIO))=? AND ANULADA='0' "
            "AND TIPOCOMPROBANTE IN ('FA','NCA','FCA','NCB') "
            "ORDER BY FECHACOMPROBANTE DESC",
            (vendedor.upper(),)
        )
        muestra = []
        for r in rows:
            tipo, num, fecha = str(r[0]).strip(), str(r[1]).strip(), r[2]
            total = float(r[3] or 0); iva1 = float(r[4] or 0); iva2 = float(r[5] or 0)
            pagado = float(r[6] or 0); cli = str(r[7] or '').strip()
            gran_total = total + iva1 + iva2
            pct = round(pagado / gran_total * 100, 1) if gran_total > 0 else 0
            muestra.append({
                "tipo": tipo, "numero": num,
                "fecha": fecha.strftime('%Y-%m-%d') if hasattr(fecha,'strftime') else str(fecha)[:10],
                "total": round(gran_total, 2), "pagado": round(pagado, 2),
                "deuda": round(gran_total - pagado, 2), "pct_cobrado": pct,
                "cliente": cli
            })
        result["muestra_fa_nca"] = muestra
    except Exception as e:
        result["error_muestra"] = str(e)

    # 4. Si existe tabla RECIBOS, mostrar sus columnas
    for tabla in (result.get("tablas_pago_relacionadas") or []):
        try:
            _, rows2 = _q(
                "SELECT FIRST 3 * FROM \"" + tabla + "\""
            )
            _, col_rows = _q(
                "SELECT TRIM(RDB$FIELD_NAME) FROM RDB$RELATION_FIELDS "
                f"WHERE RDB$RELATION_NAME='{tabla}' ORDER BY RDB$FIELD_POSITION"
            )
            result[f"cols_{tabla}"] = [r[0] for r in col_rows if r[0]]
            result[f"sample_{tabla}_count"] = len(rows2)
        except Exception as e:
            result[f"error_{tabla}"] = str(e)

    return result
